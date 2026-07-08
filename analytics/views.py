"""
Views for the analytics application.
Contains all analytical and bot-management views previously hosted in
panel/views.py. Extracted to reduce panel/views.py complexity and give
analytics its own manageable module.
---
Vistas para la aplicacion analytics.
Contiene todas las vistas analiticas y de gestion del bot previamente
alojadas en panel/views.py. Extraidas para reducir la complejidad de
panel/views.py y dar a analytics su propio modulo manejable.
"""
import difflib
import io
import json as _json
import logging
import re as _re
from collections import defaultdict
from datetime import date, datetime as _datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages as django_messages
from django.db.models import Q
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    JsonResponse,
)
from django.shortcuts import redirect, render
from django.utils.timezone import localdate, now
from django.views.generic import View

from budgets.models import Budget
from chat.models import BreakdownTicket
from fleet.models import MachineAsset
from ivr_config.models import (
    CompanyUser,
    Contact,
    PhoneNumber,
    PresenceStatus,
    Section,
    SectionContact,
    WorkshopFamilyMapping,
)
from panel.mixins import (
    AdminRoleRequiredMixin,
    CompanyUserRequiredMixin,
    SupervisorAccessMixin,
)
from panel.models import AnalyticsProfile
from whatsapp.models import WhatsAppSession, WhatsAppTemplate
from whatsapp.services import WhatsAppChatService
from work_order_processor.models import (
    OperatorMonthlyCost,
    WorkOrder,
    WorkOrderEntry,
    WorkOrderEntryLine,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _get_own_presence(company_user):
    """
    Returns the current active PresenceStatus for the given CompanyUser,
    or None if no active presence record exists.
    ---
    Devuelve el PresenceStatus activo actual para el CompanyUser dado,
    o None si no existe registro de presencia activo.
    """
    return (
        PresenceStatus.objects.filter(
            company_user=company_user,
            starts_at__lte=now(),
        )
        .filter(
            Q(ends_at__isnull=True) | Q(ends_at__gt=now())
        )
        .order_by("-starts_at")
        .first()
    )


def _excel_serial_to_date(serial):
    """
    Converts an Excel date serial number (days since 1900-01-01) into
    a Python date, correcting for Excel's fictitious 1900 leap-year
    bug (serials greater than 60 are shifted back by 2 days).
    ---
    Convierte un numero de serie de fecha de Excel (dias desde
    1900-01-01) en un date de Python, corrigiendo el bug del año
    bisiesto ficticio de 1900 (los seriales mayores que 60 se
    desplazan 2 dias hacia atras).
    """
    serial = int(serial)
    if serial > 60:
        serial -= 2
    return date(1900, 1, 1) + timedelta(days=serial - 1)


def _parse_cost_date_cell(value):
    """
    Parses the FECHA cell of an OperatorMonthlyCost import row, which
    may arrive as a native date/datetime (Excel date-formatted cell),
    an Excel serial number, or a 'YYYY-MM-DD' string. Returns a
    (year, month) tuple.
    ---
    Parsea la celda FECHA de una fila de importacion de
    OperatorMonthlyCost, que puede llegar como date/datetime nativo
    (celda con formato de fecha en Excel), numero de serie de Excel,
    o cadena 'YYYY-MM-DD'. Devuelve una tupla (year, month).
    """
    if isinstance(value, _datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    if isinstance(value, (int, float)):
        parsed = _excel_serial_to_date(value)
        return parsed.year, parsed.month
    parsed = _datetime.strptime(str(value).strip(), "%Y-%m-%d")
    return parsed.year, parsed.month


def _fuzzy_match_worker(raw_name, candidates):
    """
    Matches a raw operator name against the worker_name values known
    for a given company/month using difflib. Returns a 3-tuple of
    (status, suggested_worker_name, scored_candidates). status is
    'auto' when exactly one candidate scores >= 0.8, 'ambiguous' when
    several candidates are plausible (or the single best one scores
    below 0.8), and 'no_match' when nothing scores >= 0.5.
    ---
    Empareja un nombre de operario en bruto contra los valores
    worker_name conocidos para una empresa/mes usando difflib.
    Devuelve una tupla de 3 (status, suggested_worker_name,
    scored_candidates). status es 'auto' cuando exactamente un
    candidato puntua >= 0.8, 'ambiguous' cuando varios candidatos son
    plausibles (o el mejor puntua por debajo de 0.8), y 'no_match'
    cuando nada puntua >= 0.5.
    """
    scored = sorted(
        (
            (
                candidate,
                difflib.SequenceMatcher(
                    None, raw_name.upper(), candidate.upper(),
                ).ratio(),
            )
            for candidate in candidates
        ),
        key=lambda pair: pair[1],
        reverse=True,
    )
    top = [
        {"worker_name": name, "score": round(score, 3)}
        for name, score in scored[:3]
        if score >= 0.5
    ]
    high = [item for item in top if item["score"] >= 0.8]
    if len(high) == 1:
        return "auto", high[0]["worker_name"], top
    if top:
        return "ambiguous", None, top
    return "no_match", None, []


# ---------------------------------------------------------------------------
# AnalyticsView
# ---------------------------------------------------------------------------

class AnalyticsView(SupervisorAccessMixin, View):
    """
    Renders the analytics dashboard shell for the authenticated user's
    company. All chart data is fetched client-side via AnalyticsDataView.
    ---
    Renderiza el shell del panel de analitica para la empresa del usuario
    autenticado. Todos los datos se obtienen en el cliente via
    AnalyticsDataView.
    """

    template_name = "panel/analytics.html"

    def get(self, request):
        """
        Renders the analytics page. No chart data is computed server-side.
        ---
        Renderiza la pagina de analitica. No se calculan datos en servidor.
        """
        company_user = request.user.company_user
        company = company_user.company
        return render(request, self.template_name, {
            "company": company,
            "company_user": company_user,
            "own_presence": _get_own_presence(company_user),
            "active_nav": "analytics",
        })


# ---------------------------------------------------------------------------
# AnalyticsDataView
# ---------------------------------------------------------------------------

class AnalyticsDataView(SupervisorAccessMixin, View):
    """
    JSON endpoint returning all WorkOrderEntryLine records for the
    authenticated company, enriched with machine asset metadata, work
    date and WorkOrder reference. The client-side chart builder consumes
    this payload to render Plotly charts without further server round-trips.
    ---
    Endpoint JSON que devuelve todos los registros WorkOrderEntryLine de
    la empresa autenticada, enriquecidos con metadatos del activo, fecha
    de trabajo y referencia al WorkOrder. El constructor de graficos
    client-side los consume para renderizar graficos Plotly.
    """

    def get(self, request):
        """
        Queries and serialises all WorkOrderEntryLine records for the
        company. Returns HTTP 403 if the user has no CompanyUser profile.
        ---
        Consulta y serializa todos los registros WorkOrderEntryLine de la
        empresa. Devuelve HTTP 403 si el usuario no tiene perfil CompanyUser.
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        qs = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                machine_asset__isnull=False,
                entry__work_date__isnull=False,
            )
            .select_related(
                "machine_asset",
                "entry",
                "entry__work_order",
            )
            .order_by("entry__work_date", "machine_asset__code")
        )

        lines = []
        for line in qs:
            work_date = line.entry.work_date
            delta = (
                float(line.delta_hours)
                if line.delta_hours is not None
                else None
            )
            pdf_name = (
                line.entry.work_order.source_pdf.name.split("/")[-1]
            )
            pdf_label = _re.sub(
                r'_[A-Za-z0-9]{7}(\.[^.]+)$', r'', pdf_name,
            )
            lines.append({
                "id": line.pk,
                "work_date": (
                    work_date.isoformat() if work_date else None
                ),
                "work_order": line.entry.work_order_id,
                "pdf_name": pdf_label,
                "code": line.machine_asset.code,
                "brand_model": line.machine_asset.brand_model,
                "delta_hours": delta,
                "weekday": (
                    work_date.weekday() if work_date else None
                ),
            })

        wo_qs = (
            WorkOrder.objects
            .filter(company=company, status=WorkOrder.Status.DONE)
            .order_by("id")
        )
        work_orders = []
        for wo in wo_qs:
            raw = wo.source_pdf.name.split("/")[-1]
            label = _re.sub(
                r'_[A-Za-z0-9]{7}(\.[^.]+)$', r'', raw,
            )
            work_orders.append({"id": wo.pk, "label": label})

        seen_assets: dict = {}
        for line in lines:
            c = line["code"]
            if c not in seen_assets:
                seen_assets[c] = line["brand_model"]
        assets = [
            {"code": c, "brand_model": m}
            for c, m in sorted(seen_assets.items())
        ]

        return JsonResponse({
            "lines": lines,
            "work_orders": work_orders,
            "assets": assets,
        })


# ---------------------------------------------------------------------------
# AnalyticsLabView
# ---------------------------------------------------------------------------

class AnalyticsLabView(AdminRoleRequiredMixin, View):
    """
    Renders the Unified Analytics Laboratory shell. Passes selector data
    (operators, machines, fault categories, default date range) to the
    template. Only dimensions with available data are included in context.
    ---
    Renderiza el shell del Laboratorio de Analisis Unificado. Pasa al
    template los datos de los selectores. Solo se incluyen en contexto
    las dimensiones con datos disponibles.
    """

    template_name = "panel/analytics_lab.html"

    _FAULT_CAT_MAP = {
        "TYRES_RUNNING_GEAR": "Neumaticos y rodadura",
        "BRAKES_STEERING_SUSPENSION": "Frenos, direccion y suspension",
        "HYDRAULIC": "Hidraulica",
        "BODYWORK_CHASSIS": "Carroceria y chasis",
        "ENGINE_TRANSMISSION": "Motor y transmision",
        "LIFTING_STRUCTURE": "Estructura de elevacion",
        "ELECTRICAL_ELECTRONIC": "Electrico y electronico",
        "OTHER": "Otros",
    }

    def get(self, request):
        """
        Builds context and renders the lab template. Date defaults:
        date_from = first day of current month, date_to = today.
        ---
        Construye el contexto y renderiza el template del laboratorio.
        Fechas por defecto: date_from = primer dia del mes, date_to = hoy.
        """
        company_user = request.user.company_user
        company = company_user.company
        today = localdate()
        date_from_def = today.replace(day=1)
        date_to_def = today

        has_d1 = (
            WorkOrderEntry.objects
            .filter(
                work_order__company=company,
                worker_name__gt="",
            )
            .values("worker_name")
            .distinct()
            .count()
        ) > 0

        has_d2 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                machine_asset__isnull=False,
            )
            .values("machine_asset")
            .distinct()
            .count()
        ) > 0

        has_d3 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                fault_category__gt="",
            )
            .values("fault_category")
            .distinct()
            .count()
        ) > 0

        has_d4 = WorkOrderEntry.objects.filter(
            work_order__company=company,
            work_date__isnull=False,
        ).exists()

        # D6 — fault_subcategory
        has_d6 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                fault_subcategory__gt="",
            )
            .values("fault_subcategory")
            .distinct()
            .count()
        ) > 0

        # D7 — machine fleet family (MachineAsset.family)
        has_d7 = (
            MachineAsset.objects
            .filter(company=company, family__gt="")
            .values("family")
            .distinct()
            .count()
        ) > 0

        # D8 — machine type name (MachineAsset.type_name)
        has_d8 = (
            MachineAsset.objects
            .filter(company=company, type_name__gt="")
            .values("type_name")
            .distinct()
            .count()
        ) > 0

        # D9 — work order source (PDF_UPLOAD / DIGITAL / GENERATED)
        has_d9 = WorkOrder.objects.filter(company=company).exists()

        # D10 — O.R. (orden de reparacion)
        has_d10 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                or_val__gt="",
            )
            .exists()
        )

        # D11 — has_diet (jornada con dieta)
        has_d11 = (
            WorkOrderEntry.objects
            .filter(
                work_order__company=company,
                has_diet=True,
            )
            .exists()
        )

        # D12 — is_on_site (trabajo in situ / desplazamiento)
        has_d12 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                is_on_site=True,
            )
            .exists()
        )

        # D13 — con ticket de averia vinculado
        has_d13 = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                breakdown_ticket__isnull=False,
            )
            .exists()
        )

        # D14 — parte revisado por supervisor
        has_d14 = WorkOrder.objects.filter(company=company).exists()

        # D15 — sin pausa de comida
        has_d15 = WorkOrderEntry.objects.filter(
            work_order__company=company,
        ).exists()

        # D16 — empresa origen de la maquina (MachineAsset.company_code)
        has_d16 = (
            MachineAsset.objects
            .filter(company=company, company_code__gt="")
            .values("company_code")
            .distinct()
            .count()
        ) > 0

        # --- Build selector data ---
        operators = []
        if has_d1:
            operators = list(
                WorkOrderEntry.objects
                .filter(
                    work_order__company=company,
                    worker_name__gt="",
                )
                .values_list("worker_name", flat=True)
                .distinct()
                .order_by("worker_name")
            )

        machines = []
        if has_d2:
            machines = list(
                MachineAsset.objects
                .filter(company=company)
                .values("pk", "code", "brand_model")
                .order_by("code")
            )

        fault_categories = []
        if has_d3:
            raw_keys = list(
                WorkOrderEntryLine.objects
                .filter(
                    entry__work_order__company=company,
                    fault_category__gt="",
                )
                .values_list("fault_category", flat=True)
                .distinct()
                .order_by("fault_category")
            )
            fault_categories = [
                {"key": k, "label": self._FAULT_CAT_MAP.get(k, k)}
                for k in raw_keys
            ]

        fault_subcategories = []
        if has_d6:
            from work_order_processor.models import FaultSubcategory as _FSC
            raw_keys6 = list(
                WorkOrderEntryLine.objects
                .filter(
                    entry__work_order__company=company,
                    fault_subcategory__gt="",
                )
                .values_list("fault_subcategory", flat=True)
                .distinct()
                .order_by("fault_subcategory")
            )
            fsc_map = dict(_FSC.choices)
            fault_subcategories = [
                {"key": k, "label": fsc_map.get(k, k)}
                for k in raw_keys6
            ]

        machine_families = []
        if has_d7:
            machine_families = list(
                MachineAsset.objects
                .filter(company=company, family__gt="")
                .values_list("family", flat=True)
                .distinct()
                .order_by("family")
            )

        machine_types = []
        if has_d8:
            machine_types = list(
                MachineAsset.objects
                .filter(company=company, type_name__gt="")
                .values_list("type_name", flat=True)
                .distinct()
                .order_by("type_name")
            )

        wo_sources = []
        if has_d9:
            wo_sources = [
                {"key": k, "label": v}
                for k, v in WorkOrder.Source.choices
            ]

        or_values = []
        if has_d10:
            or_values = list(
                WorkOrderEntryLine.objects
                .filter(
                    entry__work_order__company=company,
                    or_val__gt="",
                )
                .values_list("or_val", flat=True)
                .distinct()
                .order_by("or_val")
            )

        machine_companies = []
        if has_d16:
            machine_companies = list(
                MachineAsset.objects
                .filter(company=company, company_code__gt="")
                .values("company_code", "company_name")
                .distinct()
                .order_by("company_code")
            )

        return render(request, self.template_name, {
            "company": company,
            "company_user": company_user,
            "own_presence": _get_own_presence(company_user),
            "active_nav": "analytics_lab",
            "date_from_default": date_from_def.isoformat(),
            "date_to_default": date_to_def.isoformat(),
            "has_d1": has_d1,
            "has_d2": has_d2,
            "has_d3": has_d3,
            "has_d4": has_d4,
            "has_d6": has_d6,
            "has_d7": has_d7,
            "has_d8": has_d8,
            "has_d9": has_d9,
            "has_d10": has_d10,
            "has_d11": has_d11,
            "has_d12": has_d12,
            "has_d13": has_d13,
            "has_d14": has_d14,
            "has_d15": has_d15,
            "has_d16": has_d16,
            "operators": operators,
            "machines": machines,
            "fault_categories": fault_categories,
            "fault_subcategories": fault_subcategories,
            "machine_families": machine_families,
            "machine_types": machine_types,
            "wo_sources": wo_sources,
            "or_values": or_values,
            "machine_companies": machine_companies,
        })


# ---------------------------------------------------------------------------
# AnalyticsLabDataView
# ---------------------------------------------------------------------------

class AnalyticsLabDataView(AdminRoleRequiredMixin, View):
    """
    JSON endpoint for the Unified Analytics Laboratory.
    GET /panel/analytics/lab/data/
        Parameters:
          dimension   (str) -- d1 | d2 | d3 | d4 | d5
          entity_pk   (str) -- pk of operator/machine/fault; omit for d4/d5
          date_from   (str) -- YYYY-MM-DD
          date_to     (str) -- YYYY-MM-DD
          granularity (str) -- day | week | month  (default: month)
          chart_type  (str) -- bar | line | scatter | pie | heatmap |
                               treemap  (default depends on dimension)
    ---
    Endpoint JSON para el Laboratorio de Analisis Unificado.
    GET /panel/analytics/lab/data/
        Parametros:
          dimension   (str) -- d1 | d2 | d3 | d4 | d5
          entity_pk   (str) -- pk de operario/maquina/familia; omitir d4/d5
          date_from   (str) -- YYYY-MM-DD
          date_to     (str) -- YYYY-MM-DD
          granularity (str) -- day | week | month  (por defecto: month)
          chart_type  (str) -- bar | line | scatter | pie | heatmap |
                               treemap  (por defecto segun dimension)
    """

    _DEFAULT_CHART = {
        "d1":  "bar",
        "d2":  "bar",
        "d3":  "pie",
        "d4":  "bar",
        "d6":  "pie",
        "d7":  "pie",
        "d8":  "bar",
        "d9":  "pie",
        "d10": "bar",
        "d11": "pie",
        "d12": "pie",
        "d13": "pie",
        "d14": "pie",
        "d15": "pie",
        "d16": "pie",
        "d18": "bar",  # ordinary_hours
        "d19": "bar",
    }

    _FAULT_CAT_LABELS = {
        "TYRES_RUNNING_GEAR": "Neumaticos y rodadura",
        "BRAKES_STEERING_SUSPENSION": "Frenos, direccion y suspension",
        "HYDRAULIC": "Hidraulica",
        "BODYWORK_CHASSIS": "Carroceria y chasis",
        "ENGINE_TRANSMISSION": "Motor y transmision",
        "LIFTING_STRUCTURE": "Estructura de elevacion",
        "ELECTRICAL_ELECTRONIC": "Electrico y electronico",
        "OTHER": "Otros",
    }

    _FAULT_SUBCAT_LABELS = {
        "ET_ENGINE":          "Motor",
        "ET_TRANSMISSION":    "Transmision",
        "ET_PTO":             "Toma de fuerza (PTO)",
        "ET_COOLING":         "Sistema de refrigeracion",
        "ET_FUEL":            "Sistema de combustible",
        "HY_PUMP":            "Bomba hidraulica",
        "HY_CYLINDERS":       "Cilindros hidraulicos",
        "HY_VALVES":          "Valvulas hidraulicas",
        "HY_OIL":             "Aceite y circuito hidraulico",
        "HY_CENTRAL":         "Central hidraulica",
        "EE_WIRING":          "Cableado y conectores",
        "EE_SENSORS":         "Sensores y sondas",
        "EE_CONTROLS":        "Mandos y controles",
        "EE_LIGHTS":          "Iluminacion",
        "EE_BATTERY":         "Bateria y sistema de carga",
        "BSS_BRAKES":         "Frenos",
        "BSS_STEERING":       "Direccion",
        "BSS_SUSPENSION":     "Suspension",
        "TRG_TYRES":          "Neumaticos",
        "TRG_AXLES":          "Ejes y transmision de rueda",
        "TRG_TRACKS":         "Cadenas y rodadura oruga",
        "LS_BOOM":            "Pluma y brazo",
        "LS_HOOK_PULLEYS":    "Gancho y poleas",
        "LS_CABLE":           "Cable de elevacion",
        "LS_ROTATION":        "Sistema de rotacion",
        "LS_STABILIZERS":     "Estabilizadores y apoyos",
        "LS_MAST":            "Mastil y horquillas",
        "LS_PLATFORM":        "Plataforma elevadora",
        "LS_FIFTH_WHEEL":     "Quinta rueda",
        "LS_CHASSIS_TRAILER": "Chasis de semirremolque",
        "BC_BODYWORK":        "Carroceria",
        "BC_CHASSIS":         "Chasis estructural",
        "OT_OTHER":           "Otra averia no clasificada",
    }

    def _translate_fault_cat(self, key):
        """
        Returns the Spanish display label for a FaultCategory internal key.
        Falls back to the raw key if no translation is found.
        ---
        Devuelve la etiqueta en castellano para una clave interna de
        FaultCategory. Retrocede a la clave cruda si no hay traduccion.
        """
        return self._FAULT_CAT_LABELS.get(key, key)

    @staticmethod
    def _bucket(work_date, granularity):
        """
        Returns a string label for the time bucket of work_date given the
        requested granularity (day / week / month).
        ---
        Devuelve una etiqueta para el bucket temporal de work_date segun
        la granularidad solicitada (day / week / month).
        """
        if granularity == "day":
            return work_date.strftime("%Y-%m-%d")
        if granularity == "week":
            return work_date.strftime("%G-W%V")
        return work_date.strftime("%Y-%m")

    @staticmethod
    def _parse_date(value, fallback):
        """
        Parses a YYYY-MM-DD string into a date object.
        Returns fallback on any parsing error.
        ---
        Parsea una cadena YYYY-MM-DD en un objeto date.
        Devuelve fallback ante cualquier error de parseo.
        """
        try:
            return _datetime.strptime(
                value.strip(), "%Y-%m-%d"
            ).date()
        except (ValueError, AttributeError):
            return fallback

    def _handle_d1(
        self, company, entity_pk, date_from, date_to,
        granularity, chart_type,
    ):
        """
        D1 -- Operator analysis.
        Metrics: worked hours per bucket, number of entries, dominant
        fault categories, time evolution, hours/part ratio.
        ---
        D1 -- Analisis de operario.
        Metricas: horas trabajadas por bucket, numero de entradas,
        familias de averia dominantes, evolucion temporal.
        """
        worker_name = str(entity_pk) if entity_pk else None

        qs = WorkOrderEntry.objects.filter(
            work_order__company=company,
            work_date__gte=date_from,
            work_date__lte=date_to,
            work_date__isnull=False,
        )
        if worker_name:
            qs = qs.filter(worker_name=worker_name)

        entry_pks = list(qs.values_list("pk", flat=True))
        lines_qs = WorkOrderEntryLine.objects.filter(
            entry__pk__in=entry_pks,
            delta_hours__isnull=False,
        )

        bucket_hours = defaultdict(float)
        bucket_parts = defaultdict(int)

        for line in lines_qs.select_related("entry"):
            b = self._bucket(line.entry.work_date, granularity)
            bucket_hours[b] += float(line.delta_hours or 0)

        for entry in qs:
            b = self._bucket(entry.work_date, granularity)
            bucket_parts[b] += 1

        all_buckets = sorted(
            set(bucket_hours.keys()) | set(bucket_parts.keys())
        )
        hours_series = [
            round(bucket_hours.get(b, 0.0), 2) for b in all_buckets
        ]
        parts_series = [
            bucket_parts.get(b, 0) for b in all_buckets
        ]

        if chart_type == "heatmap":
            hm_entry_filter = {
                "work_order__company": company,
                "work_date__gte": date_from,
                "work_date__lte": date_to,
                "worker_name__gt": "",
            }
            hm_line_filter = {
                "entry__work_order__company": company,
                "entry__work_date__gte": date_from,
                "entry__work_date__lte": date_to,
                "fault_category__gt": "",
            }
            if worker_name:
                hm_entry_filter["worker_name"] = worker_name
                hm_line_filter["entry__worker_name"] = worker_name
            operators_in_range = list(
                WorkOrderEntry.objects
                .filter(**hm_entry_filter)
                .values_list("worker_name", flat=True)
                .distinct()
                .order_by("worker_name")
            )
            faults_in_range = [
                self._translate_fault_cat(k)
                for k in sorted(
                    WorkOrderEntryLine.objects
                    .filter(**hm_line_filter)
                    .values_list("fault_category", flat=True)
                    .distinct()
                )
            ]
            heat_data = defaultdict(int)
            for line in (
                WorkOrderEntryLine.objects
                .filter(**hm_line_filter)
                .select_related("entry")
            ):
                op = line.entry.worker_name
                fc = self._translate_fault_cat(line.fault_category)
                if op and fc:
                    heat_data[(op, fc)] += 1

            hm_points = [
                [xi, yi, heat_data.get((op, fc), 0)]
                for xi, op in enumerate(operators_in_range)
                for yi, fc in enumerate(faults_in_range)
            ]
            total_hours = sum(
                float(ln.delta_hours or 0)
                for ln in WorkOrderEntryLine.objects.filter(
                    entry__pk__in=entry_pks,
                    delta_hours__isnull=False,
                )
            )
            total_parts = qs.count()
            table_rows = [
                [op, fc, heat_data[(op, fc)]]
                for (op, fc) in sorted(heat_data.keys())
                if heat_data[(op, fc)] > 0
            ]
            return {
                "ok": True,
                "chart": {
                    "type": "heatmap",
                    "title": (
                        "Heatmap Operario x Familia"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": operators_in_range,
                    "yAxis": faults_in_range,
                    "series": [
                        {
                            "name": "Intervenciones",
                            "data": hm_points,
                        },
                    ],
                    "visualMap": {
                        "min": 0,
                        "max": max(
                            (p[2] for p in hm_points), default=1,
                        ),
                    },
                },
                "table": {
                    "columns": [
                        "Operario",
                        "Familia de averia",
                        "Intervenciones",
                    ],
                    "rows": table_rows,
                },
                "summary": {
                    "total_hours": round(total_hours, 2),
                    "total_parts": total_parts,
                    "avg_hours_per_part": (
                        round(total_hours / total_parts, 2)
                        if total_parts else 0.0
                    ),
                },
            }

        title = (
            f"Operario: {worker_name or 'Todos'} -- "
            f"Horas trabajadas ({date_from} / {date_to})"
        )
        total_hours = sum(hours_series)
        total_parts = sum(parts_series)
        # When a specific operator is selected, show aggregated
        # series. When "all" is selected, break down by operator.
        # Con operario especifico: serie agregada. Con "Todos":
        # desglosar por operario.
        if worker_name:
            chart_series = [
                {
                    "name": "Horas trabajadas",
                    "data": hours_series,
                },
                {"name": "N entradas", "data": parts_series},
            ]
            table_cols = [
                "Periodo", "Horas trabajadas", "N entradas",
            ]
            table_rows = [
                [b, round(bucket_hours.get(b, 0.0), 2),
                 bucket_parts.get(b, 0)]
                for b in all_buckets
            ]
        else:
            # Breakdown by operator -- Desglose por operario
            all_operators = sorted(
                WorkOrderEntry.objects
                .filter(
                    work_order__company=company,
                    work_date__gte=date_from,
                    work_date__lte=date_to,
                    worker_name__gt="",
                )
                .values_list("worker_name", flat=True)
                .distinct()
            )
            op_bucket_hours = defaultdict(lambda: defaultdict(float))
            for line in lines_qs.select_related("entry"):
                op = line.entry.worker_name
                b = self._bucket(line.entry.work_date, granularity)
                op_bucket_hours[op][b] += float(
                    line.delta_hours or 0
                )
            chart_series = [
                {
                    "name": op,
                    "data": [
                        round(op_bucket_hours[op].get(b, 0.0), 2)
                        for b in all_buckets
                    ],
                }
                for op in all_operators
            ]
            table_cols = ["Operario", "Periodo", "Horas trabajadas"]
            table_rows = [
                [op, b, round(op_bucket_hours[op].get(b, 0.0), 2)]
                for op in all_operators
                for b in all_buckets
                if op_bucket_hours[op].get(b, 0.0) > 0
            ]
        return {
            "ok": True,
            "chart": {
                "type": (
                    chart_type
                    if chart_type in ("bar", "line", "pie")
                    else "bar"
                ),
                "title": title,
                "xAxis": all_buckets,
                "series": chart_series,
            },
            "table": {
                "columns": table_cols,
                "rows": table_rows,
            },
            "summary": {
                "total_hours": round(total_hours, 2),
                "total_parts": total_parts,
                "avg_hours_per_part": (
                    round(total_hours / total_parts, 2)
                    if total_parts else 0.0
                ),
            },
        }

    def _handle_d2(
        self, company, entity_pk, date_from, date_to,
        granularity, chart_type,
    ):
        """
        D2 -- Machine / Cost Centre analysis.
        Metrics: accumulated labour hours, number of interventions,
        dominant fault categories, approximate MTBF, time evolution.
        ---
        D2 -- Analisis de Maquina / Centro de Gasto.
        Metricas: horas de mano de obra acumuladas, numero de
        intervenciones, familias de averia dominantes, MTBF aproximado,
        evolucion temporal.
        """
        machine = None
        if entity_pk:
            try:
                machine = MachineAsset.objects.get(
                    pk=int(entity_pk), company=company,
                )
            except (ValueError, TypeError, MachineAsset.DoesNotExist):
                pass

        lines_qs = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                entry__work_date__gte=date_from,
                entry__work_date__lte=date_to,
                entry__work_date__isnull=False,
                machine_asset__isnull=False,
            )
            .select_related("entry", "machine_asset")
        )
        if machine:
            lines_qs = lines_qs.filter(machine_asset=machine)

        if chart_type == "scatter":
            scatter_data = defaultdict(
                lambda: {"hours": 0.0, "count": 0, "code": ""}
            )
            for line in lines_qs:
                pk_m = line.machine_asset.pk
                scatter_data[pk_m]["hours"] += float(
                    line.delta_hours or 0
                )
                scatter_data[pk_m]["count"] += 1
                scatter_data[pk_m]["code"] = line.machine_asset.code

            scatter_points = [
                [round(v["hours"], 2), v["count"], v["code"]]
                for v in scatter_data.values()
            ]
            total_hours = sum(p[0] for p in scatter_points)
            total_parts = sum(p[1] for p in scatter_points)
            return {
                "ok": True,
                "chart": {
                    "type": "scatter",
                    "title": (
                        "Maquinas -- Horas vs Intervenciones"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": [],
                    "series": [
                        {
                            "name": "Horas vs Intervenciones",
                            "data": scatter_points,
                        },
                    ],
                },
                "table": {
                    "columns": [
                        "Codigo maquina",
                        "Horas acumuladas",
                        "Intervenciones",
                    ],
                    "rows": [
                        [v["code"], round(v["hours"], 2), v["count"]]
                        for v in sorted(
                            scatter_data.values(),
                            key=lambda x: x["hours"],
                            reverse=True,
                        )
                    ],
                },
                "summary": {
                    "total_hours": round(total_hours, 2),
                    "total_parts": total_parts,
                    "avg_hours_per_part": (
                        round(total_hours / total_parts, 2)
                        if total_parts else 0.0
                    ),
                },
            }

        if chart_type == "pie":
            fault_counts = defaultdict(int)
            for line in lines_qs.filter(fault_category__gt=""):
                fault_counts[
                    self._translate_fault_cat(line.fault_category)
                ] += 1
            pie_data = [
                {"name": fc, "value": cnt}
                for fc, cnt in sorted(
                    fault_counts.items(),
                    key=lambda x: x[1],
                    reverse=True,
                )
            ]
            total_parts = sum(d["value"] for d in pie_data)
            m_label = (
                machine.code if machine else "Todas las maquinas"
            )
            return {
                "ok": True,
                "chart": {
                    "type": "pie",
                    "title": (
                        f"Familias de averia -- {m_label}"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": [],
                    "series": [
                        {
                            "name": "Intervenciones",
                            "data": pie_data,
                        },
                    ],
                },
                "table": {
                    "columns": [
                        "Familia de averia", "Intervenciones",
                    ],
                    "rows": [
                        [d["name"], d["value"]] for d in pie_data
                    ],
                },
                "summary": {
                    "total_hours": 0.0,
                    "total_parts": total_parts,
                    "avg_hours_per_part": 0.0,
                },
            }

        bucket_hours = defaultdict(float)
        bucket_counts = defaultdict(int)
        for line in lines_qs:
            b = self._bucket(line.entry.work_date, granularity)
            bucket_hours[b] += float(line.delta_hours or 0)
            bucket_counts[b] += 1

        all_buckets = sorted(
            set(bucket_hours.keys()) | set(bucket_counts.keys())
        )
        hours_series = [
            round(bucket_hours.get(b, 0.0), 2) for b in all_buckets
        ]
        count_series = [
            bucket_counts.get(b, 0) for b in all_buckets
        ]
        label = machine.code if machine else "Todas las maquinas"
        title = (
            f"Maquina: {label} -- "
            f"Horas de mano de obra ({date_from} / {date_to})"
        )
        total_hours = sum(hours_series)
        total_parts = sum(count_series)
        # When a specific machine is selected, show aggregated
        # series. When "all" is selected, break down by machine.
        # Con maquina especifica: serie agregada. Con "Todas":
        # desglosar por maquina.
        if machine:
            chart_series = [
                {
                    "name": "Horas de mano de obra",
                    "data": hours_series,
                },
                {"name": "Intervenciones", "data": count_series},
            ]
            table_cols = [
                "Periodo", "Horas de mano de obra",
                "Intervenciones",
            ]
            table_rows_d2 = [
                [b, round(bucket_hours.get(b, 0.0), 2),
                 bucket_counts.get(b, 0)]
                for b in all_buckets
            ]
        else:
            # Breakdown by machine -- Desglose por maquina
            all_machines = list(
                MachineAsset.objects
                .filter(
                    company=company,
                    workorderentryline__entry__work_date__gte=(
                        date_from
                    ),
                    workorderentryline__entry__work_date__lte=(
                        date_to
                    ),
                )
                .distinct()
                .order_by("code")
            )
            m_bucket_hours = defaultdict(
                lambda: defaultdict(float)
            )
            for line in lines_qs.select_related(
                "entry", "machine_asset"
            ):
                m_pk = line.machine_asset.pk
                b = self._bucket(
                    line.entry.work_date, granularity
                )
                m_bucket_hours[m_pk][b] += float(
                    line.delta_hours or 0
                )
            chart_series = [
                {
                    "name": m.code,
                    "data": [
                        round(
                            m_bucket_hours[m.pk].get(b, 0.0), 2
                        )
                        for b in all_buckets
                    ],
                }
                for m in all_machines
            ]
            table_cols = [
                "Maquina", "Periodo", "Horas de mano de obra",
            ]
            table_rows_d2 = [
                [
                    m.code, b,
                    round(m_bucket_hours[m.pk].get(b, 0.0), 2),
                ]
                for m in all_machines
                for b in all_buckets
                if m_bucket_hours[m.pk].get(b, 0.0) > 0
            ]
        return {
            "ok": True,
            "chart": {
                "type": (
                    chart_type
                    if chart_type in ("bar", "line", "pie")
                    else "bar"
                ),
                "title": title,
                "xAxis": all_buckets,
                "series": [
                    {"name": "Horas M.O.", "data": hours_series},
                    {
                        "name": "Intervenciones",
                        "data": count_series,
                    },
                ],
            },
            "table": {
                "columns": [
                    "Periodo", "Horas M.O.", "Intervenciones",
                ],
                "rows": [
                    [b, round(bucket_hours.get(b, 0.0), 2),
                     bucket_counts.get(b, 0)]
                    for b in all_buckets
                ],
            },
            "summary": {
                "total_hours": round(total_hours, 2),
                "total_parts": total_parts,
                "avg_hours_per_part": (
                    round(total_hours / total_parts, 2)
                    if total_parts else 0.0
                ),
            },
        }

    def _handle_d3(
        self, company, entity_pk, date_from, date_to,
        granularity, chart_type,
    ):
        """
        D3 -- Fault Category analysis.
        Metrics: frequency per bucket, most affected machines, operators
        that handle it most, time evolution, distribution by machine.
        ---
        D3 -- Analisis de Familia de Averia.
        Metricas: frecuencia por bucket, maquinas mas afectadas, operarios
        que mas la atienden, evolucion temporal, distribucion por maquina.
        """
        fault_cat = str(entity_pk) if entity_pk else None

        lines_qs = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                entry__work_date__gte=date_from,
                entry__work_date__lte=date_to,
                entry__work_date__isnull=False,
                fault_category__gt="",
            )
            .select_related(
                "entry", "entry__work_order", "machine_asset",
            )
        )
        if fault_cat:
            lines_qs = lines_qs.filter(fault_category=fault_cat)

        if chart_type in ("bar", "stacked"):
            cat_bucket = defaultdict(lambda: defaultdict(int))
            all_buckets_set = set()
            for line in lines_qs:
                b = self._bucket(line.entry.work_date, granularity)
                fc = self._translate_fault_cat(line.fault_category)
                cat_bucket[fc][b] += 1
                all_buckets_set.add(b)
            all_buckets = sorted(all_buckets_set)
            all_cats = sorted(cat_bucket.keys())
            series_list = [
                {
                    "name": fc,
                    "data": [
                        cat_bucket[fc].get(b, 0)
                        for b in all_buckets
                    ],
                    "stack": "averia",
                }
                for fc in all_cats
            ]
            total_parts = sum(
                sum(d["data"]) for d in series_list
            )
            table_rows = [
                [fc] + [
                    cat_bucket[fc].get(b, 0) for b in all_buckets
                ]
                for fc in all_cats
            ]
            table_cols = ["Familia"] + all_buckets
            return {
                "ok": True,
                "chart": {
                    "type": "bar",
                    "title": (
                        "Familias de averia -- Distribucion temporal"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": all_buckets,
                    "series": series_list,
                },
                "table": {
                    "columns": table_cols,
                    "rows": table_rows,
                },
                "summary": {
                    "total_hours": 0.0,
                    "total_parts": total_parts,
                    "avg_hours_per_part": 0.0,
                },
            }

        if chart_type == "treemap":
            machine_fault = defaultdict(lambda: defaultdict(int))
            for line in lines_qs:
                m_code = (
                    line.machine_asset.code
                    if line.machine_asset
                    else "Sin maquina"
                )
                fc = self._translate_fault_cat(line.fault_category)
                machine_fault[m_code][fc] += 1

            treemap_data = []
            for m_code, faults in sorted(machine_fault.items()):
                children = [
                    {"name": fc, "value": cnt}
                    for fc, cnt in sorted(
                        faults.items(),
                        key=lambda x: x[1],
                        reverse=True,
                    )
                ]
                treemap_data.append({
                    "name": m_code,
                    "value": sum(c["value"] for c in children),
                    "children": children,
                })
            total_parts = sum(d["value"] for d in treemap_data)
            table_rows = [
                [m_code, fc, machine_fault[m_code][fc]]
                for m_code in sorted(machine_fault.keys())
                for fc in sorted(machine_fault[m_code].keys())
                if machine_fault[m_code][fc] > 0
            ]
            fc_label = (
                self._translate_fault_cat(fault_cat)
                if fault_cat
                else "Todas las familias"
            )
            return {
                "ok": True,
                "chart": {
                    "type": "treemap",
                    "title": (
                        f"Distribucion por maquina -- {fc_label}"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": [],
                    "series": [
                        {
                            "name": "Intervenciones",
                            "data": treemap_data,
                        },
                    ],
                },
                "table": {
                    "columns": [
                        "Maquina",
                        "Familia de averia",
                        "Intervenciones",
                    ],
                    "rows": table_rows,
                },
                "summary": {
                    "total_hours": 0.0,
                    "total_parts": total_parts,
                    "avg_hours_per_part": 0.0,
                },
            }

        bucket_counts = defaultdict(int)
        for line in lines_qs:
            b = self._bucket(line.entry.work_date, granularity)
            bucket_counts[b] += 1
        all_buckets = sorted(bucket_counts.keys())
        count_series = [
            bucket_counts.get(b, 0) for b in all_buckets
        ]
        total_parts = sum(count_series)
        fc_label = (
            self._translate_fault_cat(fault_cat)
            if fault_cat
            else "Todas las familias"
        )
        return {
            "ok": True,
            "chart": {
                "type": "line",
                "title": (
                    f"Evolucion temporal -- {fc_label}"
                    f" ({date_from} / {date_to})"
                ),
                "xAxis": all_buckets,
                "series": [
                    {"name": fc_label, "data": count_series},
                ],
            },
            "table": {
                "columns": ["Periodo", "Intervenciones"],
                "rows": [
                    [b, bucket_counts.get(b, 0)]
                    for b in all_buckets
                ],
            },
            "summary": {
                "total_hours": 0.0,
                "total_parts": total_parts,
                "avg_hours_per_part": 0.0,
            },
        }

    def _handle_d4(
        self, company, entity_pk, date_from, date_to,
        granularity, chart_type,
    ):
        """
        D4 -- Time Period cross-analysis.
        Metrics: total hours, processed entries, distribution by operator,
        distribution by fault family, top intervened machines.
        ---
        D4 -- Analisis de Periodo Temporal cruzado.
        Metricas: horas totales, entradas procesadas, distribucion por
        operario, distribucion por familia, top maquinas intervenidas.
        """
        entries_qs = WorkOrderEntry.objects.filter(
            work_order__company=company,
            work_date__gte=date_from,
            work_date__lte=date_to,
            work_date__isnull=False,
        )
        lines_qs = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                entry__work_date__gte=date_from,
                entry__work_date__lte=date_to,
                entry__work_date__isnull=False,
            )
            .select_related("entry", "machine_asset")
        )

        op_bucket_hours = defaultdict(lambda: defaultdict(float))
        all_buckets_set = set()
        all_ops_set = set()

        for line in lines_qs.filter(delta_hours__isnull=False):
            b = self._bucket(line.entry.work_date, granularity)
            op = line.entry.worker_name or "Sin operario"
            op_bucket_hours[op][b] += float(line.delta_hours or 0)
            all_buckets_set.add(b)
            all_ops_set.add(op)

        all_buckets = sorted(all_buckets_set)
        all_ops = sorted(all_ops_set)
        series_list = [
            {
                "name": op,
                "data": [
                    round(op_bucket_hours[op].get(b, 0.0), 2)
                    for b in all_buckets
                ],
            }
            for op in all_ops
        ]
        total_hours = sum(
            v
            for op_data in op_bucket_hours.values()
            for v in op_data.values()
        )
        total_parts = entries_qs.count()
        table_rows = [
            [op] + [
                round(op_bucket_hours[op].get(b, 0.0), 2)
                for b in all_buckets
            ]
            for op in all_ops
        ]
        table_cols = ["Operario"] + all_buckets
        return {
            "ok": True,
            "chart": {
                "type": (
                    chart_type
                    if chart_type in ("bar", "line", "pie")
                    else "bar"
                ),
                "title": (
                    f"Resumen de periodo: {date_from} / {date_to}"
                ),
                "xAxis": all_buckets,
                "series": series_list,
            },
            "table": {
                "columns": table_cols,
                "rows": table_rows,
            },
            "summary": {
                "total_hours": round(total_hours, 2),
                "total_parts": total_parts,
                "avg_hours_per_part": (
                    round(total_hours / total_parts, 2)
                    if total_parts else 0.0
                ),
            },
        }

    def _handle_d5(
        self, company, entity_pk, date_from, date_to,
        granularity, chart_type,
    ):
        """
        D5 -- Budget / Assistance analysis.
        Metrics: budgets per insurer, average amounts, service
        distribution, temporal evolution.
        ---
        D5 -- Analisis de Presupuestos / Asistencia.
        Metricas: presupuestos por aseguradora, importes medios,
        distribucion de servicios, evolucion temporal.
        """
        budgets_qs = (
            Budget.objects
            .filter(
                company=company,
                service_date__gte=date_from,
                service_date__lte=date_to,
                service_date__isnull=False,
            )
            .select_related("insurer")
        )

        if chart_type in ("bar",):
            insurer_data = defaultdict(
                lambda: {"amount": 0.0, "count": 0}
            )
            for budget in budgets_qs:
                lbl = (
                    budget.insurer.name
                    if budget.insurer
                    else "Sin aseguradora"
                )
                insurer_data[lbl]["amount"] += float(
                    budget.total_amount or 0
                )
                insurer_data[lbl]["count"] += 1

            sorted_insurers = sorted(
                insurer_data.items(),
                key=lambda x: x[1]["amount"],
                reverse=True,
            )
            x_labels = [k for k, _ in sorted_insurers]
            amount_series = [
                round(v["amount"], 2) for _, v in sorted_insurers
            ]
            count_series = [
                v["count"] for _, v in sorted_insurers
            ]
            total_amount = sum(amount_series)
            total_parts = sum(count_series)
            table_rows = [
                [
                    lbl,
                    round(v["amount"], 2),
                    v["count"],
                    round(v["amount"] / v["count"], 2)
                    if v["count"] else 0.0,
                ]
                for lbl, v in sorted_insurers
            ]
            return {
                "ok": True,
                "chart": {
                    "type": "bar",
                    "title": (
                        "Presupuestos por aseguradora"
                        f" ({date_from} / {date_to})"
                    ),
                    "xAxis": x_labels,
                    "series": [
                        {
                            "name": "Importe total (EUR)",
                            "data": amount_series,
                        },
                        {
                            "name": "N presupuestos",
                            "data": count_series,
                        },
                    ],
                },
                "table": {
                    "columns": [
                        "Aseguradora",
                        "Importe total (EUR)",
                        "N presupuestos",
                        "Importe medio (EUR)",
                    ],
                    "rows": table_rows,
                },
                "summary": {
                    "total_hours": 0.0,
                    "total_parts": total_parts,
                    "avg_hours_per_part": (
                        round(total_amount / total_parts, 2)
                        if total_parts else 0.0
                    ),
                },
            }

        bucket_amounts = defaultdict(float)
        bucket_counts = defaultdict(int)
        for budget in budgets_qs:
            b = self._bucket(budget.service_date, granularity)
            bucket_amounts[b] += float(budget.total_amount or 0)
            bucket_counts[b] += 1

        all_buckets = sorted(
            set(bucket_amounts.keys()) | set(bucket_counts.keys())
        )
        total_parts = sum(bucket_counts.values())
        total_amount = sum(bucket_amounts.values())
        return {
            "ok": True,
            "chart": {
                "type": "line",
                "title": (
                    "Evolucion temporal de presupuestos"
                    f" ({date_from} / {date_to})"
                ),
                "xAxis": all_buckets,
                "series": [
                    {
                        "name": "Importe (EUR)",
                        "data": [
                            round(bucket_amounts.get(b, 0.0), 2)
                            for b in all_buckets
                        ],
                    },
                    {
                        "name": "N presupuestos",
                        "data": [
                            bucket_counts.get(b, 0)
                            for b in all_buckets
                        ],
                    },
                ],
            },
            "table": {
                "columns": [
                    "Periodo", "Importe (EUR)", "N presupuestos",
                ],
                "rows": [
                    [b, round(bucket_amounts.get(b, 0.0), 2),
                     bucket_counts.get(b, 0)]
                    for b in all_buckets
                ],
            },
            "summary": {
                "total_hours": 0.0,
                "total_parts": total_parts,
                "avg_hours_per_part": (
                    round(total_amount / total_parts, 2)
                    if total_parts else 0.0
                ),
            },
        }

    # ----------------------------------------------------------
    # Field type -> legacy dimension code mapping
    # Mapeo tipo de campo -> codigo de dimension legacy
    # ----------------------------------------------------------
    _FIELD_TYPE_TO_DIM = {
        "worker":            "d1",
        "machine":           "d2",
        "fault_category":    "d3",
        "period":            "d4",
        "fault_subcategory": "d6",
        "machine_family":    "d7",
        "machine_type":      "d8",
        "work_order_source": "d9",
        "or_val":            "d10",
        "has_diet":          "d11",
        "is_on_site":        "d12",
        "has_ticket":        "d13",
        "reviewed":          "d14",
        "no_lunch_break":    "d15",
        "machine_company":   "d16",
        "ordinary_hours":    "d18",
        "extra_hours":       "d19",
    }

    def get(self, request):
        """
        Validates required parameters, dispatches to the appropriate
        handler and returns the structured JSON payload.
        Accepts either:
          - legacy: dimension=d1..d5 + entity_pk
          - multidimensional: fields=JSON array of {type, value}
            up to 5 active fields.
        Returns HTTP 400 on invalid input.
        Returns HTTP 403 on missing CompanyUser profile.
        Returns HTTP 500 on unhandled handler exception.
        ---
        Valida parametros, despacha al handler adecuado y devuelve
        el payload JSON estructurado.
        Acepta modo legacy (dimension + entity_pk) o multidimensional
        (fields = array JSON de {type, value}, max 5 campos activos).
        HTTP 400 ante entrada invalida.
        HTTP 403 ante perfil CompanyUser ausente.
        HTTP 500 ante excepcion no manejada en el handler.
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"ok": False, "error": "Sin perfil de empresa asociado."},
                status=403,
            )

        granularity = (
            request.GET.get("granularity", "month").strip().lower()
        )
        if granularity not in ("day", "week", "month"):
            granularity = "month"

        chart_type = (
            request.GET.get("chart_type", "").strip().lower()
        )

        today = localdate()
        fallback_from = today.replace(day=1)
        fallback_to = today
        date_from = self._parse_date(
            request.GET.get("date_from", "").strip(), fallback_from,
        )
        date_to = self._parse_date(
            request.GET.get("date_to", "").strip(), fallback_to,
        )
        if date_from > date_to:
            date_from, date_to = date_to, date_from

        # -------------------------------------------------------
        # Parse active fields -- Parsear campos activos
        # -------------------------------------------------------
        fields_raw = request.GET.get("fields", "").strip()
        legacy_dim = request.GET.get("dimension", "").strip().lower()

        if fields_raw:
            # Multidimensional mode -- Modo multidimensional
            try:
                fields = _json.loads(fields_raw)
                if not isinstance(fields, list):
                    raise ValueError("fields must be a list")
            except (ValueError, _json.JSONDecodeError) as exc:
                return JsonResponse(
                    {"ok": False, "error": f"fields JSON invalido: {exc}"},
                    status=400,
                )
            # Filter out empty types -- Filtrar tipos vacios
            fields = [
                f for f in fields
                if isinstance(f, dict) and f.get("type")
            ]
            if not fields:
                return JsonResponse(
                    {"ok": False, "error": "Sin campos activos."},
                    status=400,
                )
            if len(fields) > 5:
                fields = fields[:5]
        elif legacy_dim:
            # Legacy mode -- Modo legacy
            if legacy_dim not in ("d1", "d2", "d3", "d4", "d5"):
                return JsonResponse(
                    {
                        "ok": False,
                        "error": (
                            f"Dimension no valida: '{legacy_dim}'. "
                            "Valores permitidos: d1, d2, d3, d4, d5."
                        ),
                    },
                    status=400,
                )
            dim_to_type = {v: k for k, v in self._FIELD_TYPE_TO_DIM.items()}
            entity_pk = (
                request.GET.get("entity_pk", "").strip() or None
            )
            fields = [{
                "type": dim_to_type.get(legacy_dim, "period"),
                "value": entity_pk or "*",
            }]
        else:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Parametro fields o dimension requerido.",
                },
                status=400,
            )

        if not chart_type:
            # Derive default chart type from first field type
            # Derivar tipo de grafico por defecto del primer campo
            first_dim = self._FIELD_TYPE_TO_DIM.get(
                fields[0]["type"], "d4"
            )
            chart_type = self._DEFAULT_CHART.get(first_dim, "bar")

        try:
            result = self._dispatch(
                company, fields, date_from, date_to,
                granularity, chart_type,
            )
        except Exception as exc:
            logger.exception(
                "# [ANALYTICS LAB DATA] Error fields=%s: %s",
                fields_raw or legacy_dim, exc,
            )
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Error interno al procesar la solicitud.",
                },
                status=500,
            )

        return JsonResponse(result)

    def _dispatch(
        self, company, fields, date_from, date_to,
        granularity, chart_type,
    ):
        """
        Routes to the appropriate handler based on the active field
        type combination. Single-field requests use the legacy handlers
        (d1..d5). Multi-field requests use _handle_cross.
        ---
        Enruta al handler adecuado segun la combinacion de tipos de
        campo activos. Solicitudes de un campo usan los handlers legacy
        (d1..d5). Solicitudes multiples usan _handle_cross.
        """
        types = [f["type"] for f in fields]
        type_set = frozenset(types)

        if len(fields) == 1:
            # Single field -- Campo unico
            ftype = types[0]
            dim = self._FIELD_TYPE_TO_DIM.get(ftype, "d4")
            entity_pk = fields[0].get("value") or None
            if entity_pk == "*":
                entity_pk = None
            legacy_handlers = {
                "d1": self._handle_d1,
                "d2": self._handle_d2,
                "d3": self._handle_d3,
                "d4": self._handle_d4,
            }
            if dim in legacy_handlers:
                return legacy_handlers[dim](
                    company, entity_pk, date_from, date_to,
                    granularity, chart_type,
                )
            # New single-field types go through _handle_cross
            # Los nuevos tipos de campo unico pasan por _handle_cross
            return self._handle_cross(
                company, fields, type_set, date_from, date_to,
                granularity, chart_type,
            )

        # Multi-field -- Multiples campos: handler de cruce
        return self._handle_cross(
            company, fields, type_set, date_from, date_to,
            granularity, chart_type,
        )

    def _handle_cross(
        self, company, fields, type_set, date_from, date_to,
        granularity, chart_type,
    ):
        """
        Cross-dimensional handler. Builds a single queryset joining
        WorkOrderEntryLine + WorkOrderEntry + MachineAsset filtered by
        all active field values, then aggregates according to the active
        dimension combination.
        Supported combinations (any order, up to 5 fields):
          worker, machine, fault_category, period, spare_part.
        Output: one series per group-by key, table with all dimensions.
        ---
        Handler multidimensional. Construye un queryset unico uniendo
        WorkOrderEntryLine + WorkOrderEntry + MachineAsset filtrado por
        todos los valores de campos activos, luego agrega segun la
        combinacion de dimensiones activas.
        Combinaciones soportadas (cualquier orden, hasta 5 campos):
          worker, machine, fault_category, period, spare_part.
        Salida: una serie por clave de agrupacion, tabla con todas las
        dimensiones.
        """
        from collections import defaultdict

        # Build base filter -- Construir filtro base
        qs_filter = {
            "entry__work_order__company": company,
            "entry__work_date__gte": date_from,
            "entry__work_date__lte": date_to,
            "entry__work_date__isnull": False,
            "delta_hours__isnull": False,
        }

        # Apply per-field value filters -- Aplicar filtros de valor
        for field in fields:
            ftype = field.get("type")
            fval = field.get("value") or "*"
            if fval == "*":
                continue
            if ftype == "worker":
                qs_filter["entry__worker_name"] = fval
            elif ftype == "machine":
                try:
                    qs_filter["machine_asset__pk"] = int(fval)
                except (ValueError, TypeError):
                    qs_filter["machine_asset__code__iexact"] = fval
            elif ftype == "fault_category":
                qs_filter["fault_category__iexact"] = fval
            elif ftype == "fault_subcategory":
                qs_filter["fault_subcategory__iexact"] = fval
            elif ftype == "machine_family":
                qs_filter["machine_asset__family__iexact"] = fval
            elif ftype == "machine_type":
                qs_filter["machine_asset__type_name__iexact"] = fval
            elif ftype == "work_order_source":
                qs_filter["entry__work_order__source__iexact"] = fval
            elif ftype == "or_val":
                qs_filter["or_val__iexact"] = fval
            elif ftype == "has_diet":
                qs_filter["entry__has_diet"] = (fval.lower() == "true")
            elif ftype == "is_on_site":
                qs_filter["is_on_site"] = (fval.lower() == "true")
            elif ftype == "has_ticket":
                if fval.lower() == "true":
                    qs_filter["breakdown_ticket__isnull"] = False
                elif fval.lower() == "false":
                    qs_filter["breakdown_ticket__isnull"] = True
            elif ftype == "reviewed":
                qs_filter["entry__work_order__reviewed"] = (
                    fval.lower() == "true"
                )
            elif ftype == "no_lunch_break":
                qs_filter["entry__no_lunch_break"] = (
                    fval.lower() == "true"
                )
            elif ftype == "machine_company":
                qs_filter["machine_asset__company_code__iexact"] = fval
            # ordinary_hours y extra_hours son metricas, no dimensiones de agrupacion
            # ordinary_hours and extra_hours are metrics, not grouping dimensions

        # Determine group-by dimensions -- Determinar dimensiones de agrupacion
        group_by_worker         = "worker"            in type_set
        group_by_machine        = "machine"           in type_set
        group_by_fault          = "fault_category"    in type_set
        group_by_fault_sub      = "fault_subcategory" in type_set
        group_by_machine_family = "machine_family"    in type_set
        group_by_machine_type   = "machine_type"      in type_set
        group_by_source         = "work_order_source" in type_set
        group_by_or_val         = "or_val"            in type_set
        group_by_has_diet       = "has_diet"          in type_set
        group_by_is_on_site     = "is_on_site"        in type_set
        group_by_has_ticket     = "has_ticket"        in type_set
        group_by_reviewed       = "reviewed"          in type_set
        group_by_no_lunch       = "no_lunch_break"    in type_set
        group_by_mach_company   = "machine_company"   in type_set
        show_ordinary_hours     = "ordinary_hours"    in type_set  # metric col only
        show_extra_hours        = "extra_hours"       in type_set  # metric col only
        group_by_period         = "period"            in type_set

        lines_qs = (
            WorkOrderEntryLine.objects
            .filter(**qs_filter)
            .select_related("entry", "entry__work_order", "machine_asset")
        )

        # Pre-calculate daily hours per (worker, date) — always needed for
        # extra hours metric. Siempre necesario para la metrica de horas extra.
        daily_hours: dict = defaultdict(float)
        for line in lines_qs:
            daily_hours[(
                line.entry.worker_name or "",
                line.entry.work_date,
            )] += float(line.delta_hours or 0)

        # Aggregate -- Agregar
        agg = defaultdict(lambda: {"hours": 0.0, "count": 0, "extra": 0.0, "ordinary": 0.0})

        # Helper: extract grouping value for a field type from a line
        # Helper: extraer valor de agrupacion para un tipo de campo desde una linea
        _SRC_LABELS = {
            "PDF_UPLOAD": "PDF",
            "DIGITAL":    "Digital",
            "GENERATED":  "Generado",
        }

        def _field_value(ftype, line):
            if ftype == "worker":
                return line.entry.worker_name or "Desconocido"
            if ftype == "machine":
                return line.machine_asset.code if line.machine_asset else "Sin maquina"
            if ftype == "fault_category":
                return (
                    self._translate_fault_cat(line.fault_category)
                    if line.fault_category else "Sin familia"
                )
            if ftype == "fault_subcategory":
                return (
                    self._FAULT_SUBCAT_LABELS.get(
                        line.fault_subcategory, line.fault_subcategory
                    ) if line.fault_subcategory else "Sin subcategoria"
                )
            if ftype == "machine_family":
                return (
                    line.machine_asset.family
                    if line.machine_asset and line.machine_asset.family
                    else "Sin familia flota"
                )
            if ftype == "machine_type":
                return (
                    line.machine_asset.type_name
                    if line.machine_asset and line.machine_asset.type_name
                    else "Sin tipo"
                )
            if ftype == "work_order_source":
                src = line.entry.work_order.source if line.entry.work_order else ""
                return _SRC_LABELS.get(src, src or "Desconocido")
            if ftype == "or_val":
                return line.or_val or "Sin O.R."
            if ftype == "has_diet":
                return "Con dieta" if line.entry.has_diet else "Sin dieta"
            if ftype == "is_on_site":
                return "In situ" if line.is_on_site else "En taller"
            if ftype == "has_ticket":
                return "Con ticket" if line.breakdown_ticket_id else "Sin ticket"
            if ftype == "reviewed":
                return "Revisado" if line.entry.work_order.reviewed else "Sin revisar"
            if ftype == "no_lunch_break":
                return (
                    "Sin pausa comida"
                    if line.entry.no_lunch_break else "Con pausa comida"
                )
            if ftype == "machine_company":
                return (
                    line.machine_asset.company_code
                    if line.machine_asset and line.machine_asset.company_code
                    else "Desconocida"
                )
            if ftype == "period":
                return self._bucket(line.entry.work_date, granularity)
            return None  # metric fields (ordinary_hours, extra_hours) — no key part

        # Ordered field types that produce grouping keys (metrics excluded)
        # Tipos de campo ordenados que producen claves de agrupacion (metricas excluidas)
        _METRIC_TYPES = frozenset(("ordinary_hours", "extra_hours"))
        ordered_dim_types = [
            f["type"] for f in fields
            if f["type"] not in _METRIC_TYPES
        ]

        for line in lines_qs:
            key_parts = [_field_value(ft, line) for ft in ordered_dim_types]
            key = tuple(key_parts) if key_parts else ("Total",)
            line_hours = float(line.delta_hours or 0)
            agg[key]["hours"] += line_hours
            agg[key]["count"] += 1
            # Calculate ordinary and extra hours proportionally
            # Calcular horas ordinarias y extra proporcionalmente
            dk = (line.entry.worker_name or "", line.entry.work_date)
            total_day = daily_hours.get(dk, 0.0)
            if total_day > 0:
                if total_day > 8.0:
                    prop = line_hours / total_day
                    agg[key]["extra"]    += prop * (total_day - 8.0)
                    agg[key]["ordinary"] += prop * 8.0
                else:
                    agg[key]["ordinary"] += line_hours

        # Build column headers in field order -- Cabeceras en orden de campos
        _DIM_LABELS = {
            "worker":            "Operario",
            "machine":           "Maquina / CdG",
            "fault_category":    "Familia averia",
            "fault_subcategory": "Subcategoria averia",
            "machine_family":    "Familia flota",
            "machine_type":      "Tipo maquina",
            "work_order_source": "Origen parte",
            "or_val":            "O.R.",
            "has_diet":          "Dieta",
            "is_on_site":        "In situ",
            "has_ticket":        "Ticket averia",
            "reviewed":          "Revisado",
            "no_lunch_break":    "Pausa comida",
            "machine_company":   "Empresa maquina",
            "period":            "Periodo",
        }
        col_labels = [
            _DIM_LABELS[ft] for ft in ordered_dim_types if ft in _DIM_LABELS
        ]
        col_labels += ["Horas trabajadas", "Intervenciones"]
        if show_ordinary_hours: col_labels.append("Horas ordinarias")
        if show_extra_hours:    col_labels.append("Horas extra")

        # Sort keys -- Ordenar claves
        sorted_keys = sorted(agg.keys())

        # Build table rows -- Construir filas de tabla
        table_rows = []
        for key in sorted_keys:
            row = list(key) + [
                round(agg[key]["hours"], 2),
                agg[key]["count"],
            ]
            if show_ordinary_hours:
                row.append(round(agg[key]["ordinary"], 2))
            if show_extra_hours:
                row.append(round(agg[key]["extra"], 2))
            table_rows.append(row)

        total_hours = sum(v["hours"] for v in agg.values())
        total_parts = sum(v["count"] for v in agg.values())

        # Build chart series -- Construir series de grafico
        # Group by first dimension key, x-axis = second dimension
        # or bucket if period is active.
        # Agrupar por primera clave de dimension, eje X = segunda
        # dimension o bucket si periodo esta activo.
        if group_by_period:
            # X-axis: time buckets -- Eje X: periodos temporales
            x_axis = sorted(
                set(
                    key[-1] for key in sorted_keys
                    if key
                )
            )
            # One series per non-period combination
            # Una serie por combinacion no-periodo
            series_keys = sorted(
                set(
                    key[:-1] for key in sorted_keys
                )
            )
            chart_series = [
                {
                    "name": " / ".join(str(s) for s in sk) or "Total",
                    "data": [
                        round(
                            agg.get(sk + (b,), {}).get("hours", 0.0),
                            2,
                        )
                        for b in x_axis
                    ],
                }
                for sk in series_keys
            ]
        else:
            # X-axis: all group keys as labels
            # Eje X: todas las claves de grupo como etiquetas
            x_axis = [
                " / ".join(str(k) for k in key)
                for key in sorted_keys
            ]
            chart_series = [
                {
                    "name": "Horas trabajadas",
                    "data": [
                        round(agg[key]["hours"], 2)
                        for key in sorted_keys
                    ],
                },
                {
                    "name": "Intervenciones",
                    "data": [
                        agg[key]["count"]
                        for key in sorted_keys
                    ],
                },
            ]

        # Build cross title in field order -- Titulo en orden de campos
        _TITLE_LABELS = {
            "worker":            "Operario",
            "machine":           "Maquina",
            "fault_category":    "Familia averia",
            "fault_subcategory": "Subcategoria averia",
            "machine_family":    "Familia flota",
            "machine_type":      "Tipo maquina",
            "work_order_source": "Origen parte",
            "or_val":            "O.R.",
            "has_diet":          "Dieta",
            "is_on_site":        "In situ",
            "has_ticket":        "Ticket averia",
            "reviewed":          "Revisado",
            "no_lunch_break":    "Pausa comida",
            "machine_company":   "Empresa maquina",
            "period":            "Periodo",
        }
        dim_names = [
            _TITLE_LABELS[ft]
            for ft in ordered_dim_types
            if ft in _TITLE_LABELS
        ]
        if show_ordinary_hours: dim_names.append("Horas ordinarias")
        if show_extra_hours:    dim_names.append("Horas extra")
        title = " x ".join(dim_names) + f" ({date_from} / {date_to})"

        return {
            "ok": True,
            "chart": {
                "type": (
                    chart_type
                    if chart_type in (
                        "bar", "line", "scatter",
                        "pie", "heatmap", "treemap",
                    )
                    else "bar"
                ),
                "title": title,
                "xAxis": x_axis,
                "series": chart_series,
            },
            "table": {
                "columns": col_labels,
                "rows": table_rows,
            },
            "summary": {
                "total_hours": round(total_hours, 2),
                "total_parts": total_parts,
                "avg_hours_per_part": (
                    round(total_hours / total_parts, 2)
                    if total_parts else 0.0
                ),
            },
        }


# ---------------------------------------------------------------------------
# AnalyticsLabExportView
# ---------------------------------------------------------------------------

class AnalyticsLabExportView(AdminRoleRequiredMixin, View):
    """
    Generates and streams an Excel file from the Analytics Laboratory
    table data. POST body must contain JSON-encoded 'columns' and 'rows'.
    POST /panel/analytics/lab/export/
    ---
    Genera y devuelve en streaming un Excel desde los datos de tabla del
    Laboratorio de Analisis. El cuerpo POST debe contener 'columns' y
    'rows' codificados en JSON.
    POST /panel/analytics/lab/export/
    """

    def post(self, request, *args, **kwargs):
        """
        Parses columns and rows from POST body, builds an openpyxl
        workbook and streams it as an xlsx attachment.
        ---
        Parsea columns y rows del cuerpo POST, construye un libro
        openpyxl y lo devuelve en streaming como adjunto xlsx.
        """
        try:
            columns = _json.loads(
                request.POST.get("columns", "[]")
            )
            rows = _json.loads(request.POST.get("rows", "[]"))
            dimension = request.POST.get("dimension", "lab").strip()
            date_from = request.POST.get("date_from", "").strip()
            date_to = request.POST.get("date_to", "").strip()
        except (_json.JSONDecodeError, TypeError) as exc:
            logger.warning(
                "# [ANALYTICS LAB EXPORT] Payload invalido: %s",
                exc,
            )
            return HttpResponseBadRequest(
                "# [ANALYTICS LAB EXPORT] Payload JSON invalido."
            )

        if not columns or not isinstance(rows, list):
            return HttpResponseBadRequest(
                "# [ANALYTICS LAB EXPORT] "
                "Se requieren 'columns' y 'rows' en el cuerpo."
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laboratorio"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            fill_type="solid", fgColor="2C3E50",
        )
        header_align = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        for col_idx, col_header in enumerate(columns, start=1):
            cell = ws.cell(
                row=1, column=col_idx, value=str(col_header),
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                if hasattr(cell_value, "__float__"):
                    cell_value = float(cell_value)
                ws.cell(
                    row=row_idx, column=col_idx, value=cell_value,
                )

        for col_cells in ws.columns:
            max_length = max(
                (
                    len(str(c.value)) if c.value is not None else 0
                    for c in col_cells
                ),
                default=8,
            )
            ws.column_dimensions[
                col_cells[0].column_letter
            ].width = min(max_length + 4, 60)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = (
            f"lab_{dimension}_{date_from}_{date_to}.xlsx"
        )
        response = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        logger.info(
            "# [ANALYTICS LAB EXPORT] dimension=%s (%d filas) "
            "por %s.",
            dimension, len(rows), request.user.username,
        )
        return response


# ---------------------------------------------------------------------------
# AnalyticsProfileListCreateView
# ---------------------------------------------------------------------------

class AnalyticsProfileListCreateView(SupervisorAccessMixin, View):
    """
    JSON endpoint for listing and creating/updating AnalyticsProfile
    records belonging to the authenticated CompanyUser.
    GET  /panel/analytics/profiles/
    POST /panel/analytics/profiles/
    ---
    Endpoint JSON para listar y crear/actualizar registros
    AnalyticsProfile del CompanyUser autenticado.
    GET  /panel/analytics/profiles/
    POST /panel/analytics/profiles/
    """

    def get(self, request):
        """
        Returns the list of AnalyticsProfile records for the current
        CompanyUser ordered by name.
        ---
        Devuelve la lista de registros AnalyticsProfile del CompanyUser
        actual ordenados por nombre.
        """
        try:
            company_user = request.user.company_user
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        profiles = (
            AnalyticsProfile.objects
            .filter(company_user=company_user)
            .order_by("nombre")
            .values("id", "nombre", "config")
        )
        return JsonResponse({"profiles": list(profiles)})

    def post(self, request):
        """
        Creates or updates an AnalyticsProfile for the current
        CompanyUser using update_or_create (upsert semantics).
        ---
        Crea o actualiza un AnalyticsProfile para el CompanyUser actual
        usando update_or_create (semantica upsert).
        """
        try:
            company_user = request.user.company_user
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            payload = _json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse(
                {"error": "Cuerpo JSON invalido."},
                status=400,
            )

        nombre = payload.get("nombre", "").strip()
        config = payload.get("config")

        if not nombre:
            return JsonResponse(
                {"error": "El campo 'nombre' es obligatorio."},
                status=400,
            )
        if not isinstance(config, dict):
            return JsonResponse(
                {
                    "error": (
                        "El campo 'config' debe ser un objeto JSON."
                    ),
                },
                status=400,
            )

        profile, _ = AnalyticsProfile.objects.update_or_create(
            company_user=company_user,
            nombre=nombre,
            defaults={"config": config},
        )
        return JsonResponse({
            "id": profile.pk,
            "nombre": profile.nombre,
            "config": profile.config,
        })


# ---------------------------------------------------------------------------
# AnalyticsProfileDeleteView
# ---------------------------------------------------------------------------

class AnalyticsProfileDeleteView(SupervisorAccessMixin, View):
    """
    Deletes a single AnalyticsProfile owned by the current CompanyUser.
    DELETE /panel/analytics/profiles/<pk>/
    ---
    Elimina un AnalyticsProfile del CompanyUser actual.
    DELETE /panel/analytics/profiles/<pk>/
    """

    def delete(self, request, pk):
        """
        Deletes the AnalyticsProfile with the given pk if it belongs to
        the current CompanyUser. Returns HTTP 404 if not found.
        ---
        Elimina el AnalyticsProfile con el pk dado si pertenece al
        CompanyUser actual. Devuelve HTTP 404 si no se encuentra.
        """
        try:
            company_user = request.user.company_user
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            profile = AnalyticsProfile.objects.get(
                pk=pk, company_user=company_user,
            )
        except AnalyticsProfile.DoesNotExist:
            return JsonResponse(
                {"error": "Perfil no encontrado."},
                status=404,
            )

        profile.delete()
        return JsonResponse({"deleted": pk})


# ---------------------------------------------------------------------------
# AnalyticsProfileUpdateView
# ---------------------------------------------------------------------------

class AnalyticsProfileUpdateView(SupervisorAccessMixin, View):
    """
    Updates nombre and/or config of a single AnalyticsProfile.
    PATCH /panel/analytics/profiles/<pk>/update/
    ---
    Actualiza nombre y/o config de un AnalyticsProfile.
    PATCH /panel/analytics/profiles/<pk>/update/
    """

    def patch(self, request, pk):
        """
        Partial update: accepts nombre and/or config fields.
        Nombre change enforces uniqueness at (company_user, nombre) level.
        ---
        Actualizacion parcial: acepta campos nombre y/o config.
        El cambio de nombre respeta la unicidad (company_user, nombre).
        """
        try:
            company_user = request.user.company_user
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            profile = AnalyticsProfile.objects.get(
                pk=pk, company_user=company_user,
            )
        except AnalyticsProfile.DoesNotExist:
            return JsonResponse(
                {"error": "Perfil no encontrado."},
                status=404,
            )

        try:
            payload = _json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse(
                {"error": "Cuerpo JSON invalido."},
                status=400,
            )

        nuevo_nombre = payload.get("nombre", "").strip()
        nuevo_config = payload.get("config")

        if nuevo_nombre and nuevo_nombre != profile.nombre:
            if AnalyticsProfile.objects.filter(
                company_user=company_user,
                nombre=nuevo_nombre,
            ).exclude(pk=pk).exists():
                return JsonResponse(
                    {"error": f"Ya existe una plantilla con el nombre '{nuevo_nombre}'."},
                    status=409,
                )
            profile.nombre = nuevo_nombre

        if isinstance(nuevo_config, dict):
            profile.config = nuevo_config

        profile.save()
        return JsonResponse({
            "id": profile.pk,
            "nombre": profile.nombre,
            "config": profile.config,
        })


# ---------------------------------------------------------------------------
# AnalyticsProfileCloneView
# ---------------------------------------------------------------------------

class AnalyticsProfileCloneView(SupervisorAccessMixin, View):
    """
    Clones an existing AnalyticsProfile under a new name.
    POST /panel/analytics/profiles/<pk>/clone/
    Body: {"nombre": "Nuevo nombre"}
    ---
    Clona un AnalyticsProfile existente con un nuevo nombre.
    POST /panel/analytics/profiles/<pk>/clone/
    Cuerpo: {"nombre": "Nuevo nombre"}
    """

    def post(self, request, pk):
        """
        Creates a new AnalyticsProfile with the same config as the
        source profile but under the given nombre.
        ---
        Crea un nuevo AnalyticsProfile con la misma config que el perfil
        origen pero con el nombre indicado.
        """
        try:
            company_user = request.user.company_user
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            source = AnalyticsProfile.objects.get(
                pk=pk, company_user=company_user,
            )
        except AnalyticsProfile.DoesNotExist:
            return JsonResponse(
                {"error": "Perfil origen no encontrado."},
                status=404,
            )

        try:
            payload = _json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse(
                {"error": "Cuerpo JSON invalido."},
                status=400,
            )

        nuevo_nombre = payload.get("nombre", "").strip()
        if not nuevo_nombre:
            return JsonResponse(
                {"error": "El campo 'nombre' es obligatorio."},
                status=400,
            )

        if AnalyticsProfile.objects.filter(
            company_user=company_user,
            nombre=nuevo_nombre,
        ).exists():
            return JsonResponse(
                {"error": f"Ya existe una plantilla con el nombre '{nuevo_nombre}'."},
                status=409,
            )

        clone = AnalyticsProfile.objects.create(
            company_user=company_user,
            nombre=nuevo_nombre,
            config=source.config,
        )
        return JsonResponse({
            "id": clone.pk,
            "nombre": clone.nombre,
            "config": clone.config,
        })


# ---------------------------------------------------------------------------
# OperatorMonthlyCostListView
# ---------------------------------------------------------------------------

class OperatorMonthlyCostListView(SupervisorAccessMixin, View):
    """
    Lists OperatorMonthlyCost records for the current company.
    GET /panel/analytics/costs/
    ---
    Lista los registros OperatorMonthlyCost de la empresa actual.
    GET /panel/analytics/costs/
    """

    def get(self, request):
        """
        Returns every cost record owned by the company, serialised
        with monthly_cost as a string (Decimal is not JSON-native).
        ---
        Devuelve todos los registros de coste de la empresa,
        serializados con monthly_cost como cadena (Decimal no es
        nativo de JSON).
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        records = [
            {
                "id": record.pk,
                "worker_name": record.worker_name,
                "year": record.year,
                "month": record.month,
                "monthly_cost": str(record.monthly_cost),
            }
            for record in OperatorMonthlyCost.objects.filter(
                company=company,
            )
        ]
        return JsonResponse({"records": records})


# ---------------------------------------------------------------------------
# OperatorMonthlyCostCreateView
# ---------------------------------------------------------------------------

class OperatorMonthlyCostCreateView(SupervisorAccessMixin, View):
    """
    Creates or updates a single OperatorMonthlyCost record via manual
    entry (upsert on the (company, worker_name, year, month) unique
    constraint).
    POST /panel/analytics/costs/create/
    Body: {"worker_name": str, "year": int, "month": int,
           "monthly_cost": number}
    ---
    Crea o actualiza un registro OperatorMonthlyCost mediante entrada
    manual (upsert sobre la restriccion unica (company, worker_name,
    year, month)).
    POST /panel/analytics/costs/create/
    Cuerpo: {"worker_name": str, "year": int, "month": int,
             "monthly_cost": numero}
    """

    def post(self, request):
        """
        Validates the payload and performs an update_or_create against
        OperatorMonthlyCost scoped to the current company.
        ---
        Valida el cuerpo de la peticion y ejecuta un update_or_create
        sobre OperatorMonthlyCost acotado a la empresa actual.
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            payload = _json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse(
                {"error": "Cuerpo JSON invalido."},
                status=400,
            )

        worker_name = str(payload.get("worker_name", "")).strip()
        if not worker_name:
            return JsonResponse(
                {"error": "El campo 'worker_name' es obligatorio."},
                status=400,
            )

        try:
            year = int(payload.get("year"))
            month = int(payload.get("month"))
            if not 1 <= month <= 12:
                raise ValueError
        except (TypeError, ValueError):
            return JsonResponse(
                {
                    "error": (
                        "'year' y 'month' deben ser enteros validos "
                        "('month' entre 1 y 12)."
                    ),
                },
                status=400,
            )

        try:
            monthly_cost = Decimal(
                str(payload.get("monthly_cost")).replace(",", "."),
            )
            if monthly_cost < 0:
                raise InvalidOperation
        except (InvalidOperation, TypeError, ValueError):
            return JsonResponse(
                {
                    "error": (
                        "'monthly_cost' debe ser un numero valido "
                        "y no negativo."
                    ),
                },
                status=400,
            )

        record, created = OperatorMonthlyCost.objects.update_or_create(
            company=company,
            worker_name=worker_name,
            year=year,
            month=month,
            defaults={"monthly_cost": monthly_cost},
        )
        return JsonResponse({
            "id": record.pk,
            "worker_name": record.worker_name,
            "year": record.year,
            "month": record.month,
            "monthly_cost": str(record.monthly_cost),
            "created": created,
        })


# ---------------------------------------------------------------------------
# OperatorMonthlyCostDeleteView
# ---------------------------------------------------------------------------

class OperatorMonthlyCostDeleteView(SupervisorAccessMixin, View):
    """
    Deletes a single OperatorMonthlyCost record owned by the current
    company.
    DELETE /panel/analytics/costs/<pk>/
    ---
    Elimina un registro OperatorMonthlyCost de la empresa actual.
    DELETE /panel/analytics/costs/<pk>/
    """

    def delete(self, request, pk):
        """
        Deletes the record with the given pk if it belongs to the
        current company. Returns HTTP 404 if not found.
        ---
        Elimina el registro con el pk dado si pertenece a la empresa
        actual. Devuelve HTTP 404 si no se encuentra.
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        try:
            record = OperatorMonthlyCost.objects.get(
                pk=pk, company=company,
            )
        except OperatorMonthlyCost.DoesNotExist:
            return JsonResponse(
                {"error": "Registro no encontrado."},
                status=404,
            )

        record.delete()
        return JsonResponse({"deleted": pk})


# ---------------------------------------------------------------------------
# OperatorMonthlyCostImportView
# ---------------------------------------------------------------------------

class OperatorMonthlyCostImportView(SupervisorAccessMixin, View):
    """
    Two-phase Excel importer for OperatorMonthlyCost. Expects columns
    MECANICO, FECHA and COSTE.

    Phase 1 -- PREVIEW: multipart POST with a 'file' field. Parses the
    workbook, resolves each row's (year, month) and attempts a fuzzy
    match of MECANICO against the worker_name values recorded in
    WorkOrderEntry for that company/month. Returns a preview payload;
    nothing is persisted.

    Phase 2 -- CONFIRM: JSON POST with a 'rows' array (each item
    carrying the confirmed worker_name, year, month and monthly_cost).
    Persists every row via update_or_create.

    POST /panel/analytics/costs/import/
    ---
    Importador de Excel en dos fases para OperatorMonthlyCost. Espera
    las columnas MECANICO, FECHA y COSTE.

    Fase 1 -- PREVIEW: POST multipart con campo 'file'. Parsea el
    libro, resuelve (year, month) de cada fila e intenta un
    emparejamiento difuso de MECANICO contra los worker_name
    registrados en WorkOrderEntry para esa empresa/mes. Devuelve un
    payload de previsualizacion; no persiste nada.

    Fase 2 -- CONFIRM: POST JSON con un array 'rows' (cada elemento
    lleva worker_name, year, month y monthly_cost confirmados).
    Persiste cada fila via update_or_create.

    POST /panel/analytics/costs/import/
    """

    def post(self, request):
        """
        Dispatches to preview or confirm mode depending on whether the
        request carries an uploaded file or a JSON 'rows' payload.
        ---
        Despacha a modo preview o confirm segun si la peticion trae un
        fichero subido o un payload JSON 'rows'.
        """
        try:
            company = request.user.company_user.company
        except AttributeError:
            return JsonResponse(
                {"error": "Sin perfil de empresa asociado."},
                status=403,
            )

        uploaded = request.FILES.get("file")
        if uploaded:
            return self._preview(uploaded, company)
        return self._confirm(request, company)

    def _preview(self, uploaded, company):
        """
        Parses the uploaded workbook and returns a per-row preview
        with the fuzzy-match status, without persisting anything.
        ---
        Parsea el libro subido y devuelve una previsualizacion por
        fila con el estado del emparejamiento difuso, sin persistir
        nada.
        """
        try:
            workbook = openpyxl.load_workbook(
                uploaded, data_only=True, read_only=True,
            )
            sheet = workbook.active
        except Exception:
            return JsonResponse(
                {"error": "No se pudo leer el fichero Excel."},
                status=400,
            )

        rows_out = []
        header = None
        for row_index, row in enumerate(
            sheet.iter_rows(values_only=True), start=1,
        ):
            if header is None:
                header = [
                    str(cell).strip().upper() if cell else ""
                    for cell in row
                ]
                continue
            if all(cell in (None, "") for cell in row):
                continue

            mapped = dict(zip(header, row))
            mecanico_raw = mapped.get("MECANICO")
            fecha_raw = mapped.get("FECHA")
            coste_raw = mapped.get("COSTE")

            entry = {
                "row_index": row_index,
                "mecanico_raw": mecanico_raw,
                "fecha_raw": (
                    str(fecha_raw) if fecha_raw is not None else None
                ),
                "coste_raw": coste_raw,
            }

            if (
                not mecanico_raw
                or fecha_raw is None
                or coste_raw is None
            ):
                entry["status"] = "error"
                entry["error"] = (
                    "Faltan valores en MECANICO/FECHA/COSTE."
                )
                rows_out.append(entry)
                continue

            try:
                year, month = _parse_cost_date_cell(fecha_raw)
            except (ValueError, TypeError):
                entry["status"] = "error"
                entry["error"] = "FECHA con formato no reconocido."
                rows_out.append(entry)
                continue

            try:
                monthly_cost = Decimal(
                    str(coste_raw).replace(",", "."),
                )
            except (InvalidOperation, TypeError, ValueError):
                entry["status"] = "error"
                entry["error"] = "COSTE no es un numero valido."
                rows_out.append(entry)
                continue

            candidates = list(
                WorkOrderEntry.objects
                .filter(
                    work_order__company=company,
                    worker_name__gt="",
                    work_date__year=year,
                    work_date__month=month,
                )
                .values_list("worker_name", flat=True)
                .distinct()
            )
            status, suggested, scored = _fuzzy_match_worker(
                str(mecanico_raw).strip(), candidates,
            )

            entry.update({
                "year": year,
                "month": month,
                "monthly_cost": str(monthly_cost),
                "status": status,
                "suggested_worker_name": suggested,
                "candidates": scored,
            })
            rows_out.append(entry)

        return JsonResponse({"rows": rows_out})

    def _confirm(self, request, company):
        """
        Persists the confirmed rows via update_or_create, scoped to
        the current company.
        ---
        Persiste las filas confirmadas via update_or_create, acotadas
        a la empresa actual.
        """
        try:
            payload = _json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse(
                {"error": "Cuerpo JSON invalido."},
                status=400,
            )

        rows = payload.get("rows")
        if not isinstance(rows, list) or not rows:
            return JsonResponse(
                {
                    "error": (
                        "El campo 'rows' debe ser una lista no vacia."
                    ),
                },
                status=400,
            )

        created_count = 0
        updated_count = 0
        errors = []

        for index, row in enumerate(rows):
            worker_name = str(row.get("worker_name", "")).strip()

            if not worker_name:
                errors.append(
                    {"row": index, "error": "worker_name vacio."},
                )
                continue
            try:
                year = int(row.get("year"))
                month = int(row.get("month"))
                if not 1 <= month <= 12:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(
                    {"row": index, "error": "year/month invalidos."},
                )
                continue
            try:
                monthly_cost = Decimal(
                    str(row.get("monthly_cost")).replace(",", "."),
                )
                if monthly_cost < 0:
                    raise InvalidOperation
            except (InvalidOperation, TypeError, ValueError):
                errors.append(
                    {"row": index, "error": "monthly_cost invalido."},
                )
                continue

            _, created = OperatorMonthlyCost.objects.update_or_create(
                company=company,
                worker_name=worker_name,
                year=year,
                month=month,
                defaults={"monthly_cost": monthly_cost},
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        return JsonResponse({
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
        })


# ---------------------------------------------------------------------------
# BotManagementView
# ---------------------------------------------------------------------------

class BotManagementView(CompanyUserRequiredMixin, View):
    """
    Central management dashboard for the WhatsApp bot. Provides four
    functional blocks: section onboarding, group broadcast, 1-to-1
    broadcast and breakdown ticket viewer.
    Access: ADMIN (all blocks), SUPERVISOR/WORKSHOPBOSS/WORKSHOP (viewer).
    ---
    Panel central de gestion del bot de WhatsApp. Proporciona cuatro
    bloques funcionales: onboarding por seccion, circular a grupos,
    circular 1:1 y visor de averias.
    Acceso: ADMIN (todos), SUPERVISOR/WORKSHOPBOSS/WORKSHOP (solo visor).
    """

    template_name = "panel/bot/dashboard.html"

    ALLOWED_ROLES = {
        CompanyUser.ROLE_ADMIN,
        CompanyUser.ROLE_SUPERVISOR,
        CompanyUser.ROLE_WORKSHOPBOSS,
        CompanyUser.ROLE_WORKSHOP,
    }

    def _get_workshop_family_for_user(self, company_user):
        """
        Resolves the workshop family visible to the authenticated user.
        ADMIN/SUPERVISOR: None. WORKSHOPBOSS: own family. WORKSHOP: family
        of the WORKSHOPBOSS assigned to their section.
        ---
        Resuelve la familia de taller visible para el usuario autenticado.
        ADMIN/SUPERVISOR: None. WORKSHOPBOSS: su familia. WORKSHOP: familia
        del WORKSHOPBOSS asignado a su seccion.
        """
        role = company_user.role
        if role in (
            CompanyUser.ROLE_ADMIN, CompanyUser.ROLE_SUPERVISOR,
        ):
            return None
        if role == CompanyUser.ROLE_WORKSHOPBOSS:
            return company_user.workshop_family
        if role == CompanyUser.ROLE_WORKSHOP:
            section_contact = (
                SectionContact.objects
                .filter(
                    contact__company_user=company_user,
                    section__company=company_user.company,
                )
                .select_related("section")
                .first()
            )
            if not section_contact:
                return None
            boss = (
                CompanyUser.objects
                .filter(
                    company=company_user.company,
                    role=CompanyUser.ROLE_WORKSHOPBOSS,
                    workshop_family__isnull=False,
                )
                .first()
            )
            return boss.workshop_family if boss else None
        return None

    def _get_breakdown_tickets(self, company_user, family_filter=None):
        """
        Returns a queryset of active BreakdownTickets for the company,
        filtered by workshop family when family_filter is provided.
        ---
        Devuelve un queryset de BreakdownTickets activos para la empresa,
        filtrado por familia de taller cuando se proporciona family_filter.
        """
        qs = (
            BreakdownTicket.objects
            .filter(company=company_user.company)
            .exclude(status=BreakdownTicket.STATUS_CLOSED)
            .select_related("machine", "contact", "assigned_to")
            .order_by("-created_at")
        )
        if family_filter:
            mapped_families = list(
                WorkshopFamilyMapping.objects
                .filter(
                    company=company_user.company,
                    workshop_family=family_filter,
                )
                .values_list("catalogue_family", flat=True)
            )
            if mapped_families:
                qs = qs.filter(
                    machine__family__in=mapped_families,
                )
            else:
                qs = qs.none()
        return qs

    def get(self, request, *args, **kwargs):
        """
        Renders the bot management dashboard with context data for all
        four functional blocks. Applies role-based visibility rules.
        ---
        Renderiza el panel de gestion del bot con datos de contexto para
        los cuatro bloques funcionales. Aplica las reglas de visibilidad
        por rol.
        """
        company_user = request.user.company_user
        if company_user.role not in self.ALLOWED_ROLES:
            return redirect("panel:dashboard")

        is_admin = company_user.role == CompanyUser.ROLE_ADMIN
        is_supervisor = (
            company_user.role == CompanyUser.ROLE_SUPERVISOR
        )
        can_manage = is_admin

        resolved_family = self._get_workshop_family_for_user(
            company_user,
        )
        family_filter = resolved_family
        if is_admin or is_supervisor:
            family_filter = (
                request.GET.get("family", None) or None
            )

        breakdown_tickets = self._get_breakdown_tickets(
            company_user, family_filter,
        )

        sections = (
            Section.objects
            .filter(company=company_user.company)
            .order_by("name")
            if can_manage else []
        )
        broadcast_sections = (
            Section.objects
            .filter(
                company=company_user.company,
                is_broadcast_enabled=True,
            )
            .order_by("name")
            if can_manage else []
        )
        family_choices = (
            CompanyUser.WORKSHOP_FAMILY_CHOICES
            if (is_admin or is_supervisor) else []
        )

        company_users = (
            CompanyUser.objects
            .filter(company=company_user.company, is_active=True)
            .select_related("user")
            .order_by("user__last_name", "user__first_name")
            if can_manage else []
        )

        return render(request, self.template_name, {
            "active_nav": "bot_management",
            "can_manage": can_manage,
            "is_admin": is_admin,
            "is_supervisor": is_supervisor,
            "sections": sections,
            "broadcast_sections": broadcast_sections,
            "breakdown_tickets": breakdown_tickets,
            "family_choices": family_choices,
            "family_filter": family_filter,
            "company_user": company_user,
            "company_users": company_users,
        })

    def post(self, request, *args, **kwargs):
        """
        Handles the three bot management actions dispatched via the hidden
        'action' field: onboarding, group_broadcast, direct_broadcast.
        Only ADMIN role may reach this handler.
        ---
        Gestiona las tres acciones de administracion del bot despachadas
        mediante el campo oculto 'action': onboarding, group_broadcast,
        direct_broadcast. Solo el rol ADMIN puede alcanzar este handler.
        """
        company_user = request.user.company_user
        is_admin = company_user.role == CompanyUser.ROLE_ADMIN
        if not is_admin:
            return redirect("analytics:bot_management")

        company = company_user.company
        action = request.POST.get("action", "")

        bot_number = (
            PhoneNumber.objects
            .filter(
                company=company,
                capabilities__in=[
                    PhoneNumber.CAPABILITY_WHATSAPP,
                    PhoneNumber.CAPABILITY_BOTH,
                ],
                is_active=True,
            )
            .values_list("number", flat=True)
            .first()
        )
        if not bot_number:
            django_messages.error(
                request,
                "No se encontro ningun numero WhatsApp activo "
                "para esta empresa.",
            )
            return redirect("analytics:bot_management")

        if action == "onboarding":
            section_id = request.POST.get("section_id")
            if not section_id:
                django_messages.error(
                    request,
                    "Debes seleccionar una seccion.",
                )
                return redirect("analytics:bot_management")
            try:
                section = Section.objects.get(
                    pk=section_id, company=company,
                )
            except Section.DoesNotExist:
                django_messages.error(
                    request, "Seccion no valida.",
                )
                return redirect("analytics:bot_management")

            try:
                onboarding_template = WhatsAppTemplate.objects.get(
                    name="chat_onboarding", company=company,
                )
                template_sid = onboarding_template.content_sid
            except WhatsAppTemplate.DoesNotExist:
                django_messages.error(
                    request,
                    "No se encontro el template chat_onboarding.",
                )
                return redirect("analytics:bot_management")

            pending_contacts = (
                Contact.objects
                .filter(
                    company=company,
                    section_assignments__section=section,
                )
                .exclude(
                    alias_onboarding_step=Contact.ALIAS_STEP_NONE,
                )
                .exclude(
                    company_user__isnull=False,
                    company_user__alias__isnull=False,
                )
                .exclude(phone_number="")
                .distinct()
            )

            sent_count = 0
            for contact in pending_contacts:
                if not contact.phone_number:
                    continue
                try:
                    WhatsAppChatService.send_quick_reply(
                        from_number=bot_number,
                        to_number=contact.phone_number,
                        content_sid=template_sid,
                        content_variables={
                            "1": contact.name,
                            "2": company.name,
                        },
                    )
                    sent_count += 1
                except Exception as exc:
                    logger.error(
                        "# [BOT MGMT] Error enviando onboarding "
                        "a %s: %s",
                        contact.phone_number, exc,
                    )

            django_messages.success(
                request,
                f"Onboarding lanzado a {sent_count} contacto(s) "
                f"de la seccion '{section.name}'.",
            )
            return redirect("analytics:bot_management")

        if action == "group_broadcast":
            selected_section_pks = request.POST.getlist(
                "section_pks",
            )
            message_body = (
                request.POST.get("message", "").strip()
            )
            if not selected_section_pks or not message_body:
                django_messages.error(
                    request,
                    "Debes seleccionar al menos una seccion "
                    "y escribir un mensaje.",
                )
                return redirect("analytics:bot_management")

            valid_sections = Section.objects.filter(
                pk__in=selected_section_pks,
                company=company,
                is_broadcast_enabled=True,
            )
            if not valid_sections.exists():
                django_messages.error(
                    request,
                    "Ninguna seccion seleccionada es valida "
                    "para circulares.",
                )
                return redirect("analytics:bot_management")

            try:
                renewal_template = WhatsAppTemplate.objects.get(
                    name="chat_session_renewal", company=company,
                )
            except WhatsAppTemplate.DoesNotExist:
                renewal_template = None
                logger.warning(
                    "# [BOT MGMT] Template chat_session_renewal "
                    "no encontrado para empresa pk=%r.",
                    company.pk,
                )

            from django.utils.timezone import now as _now
            from datetime import timedelta as _td
            window_threshold = _now() - _td(hours=24)
            created_at_iso = _now().isoformat()
            total_sent = 0
            total_pending = 0

            for section in valid_sections:
                room_contacts = (
                    Contact.objects
                    .filter(
                        company=company,
                        sections=section,
                        company_user__isnull=False,
                        company_user__is_active=True,
                    )
                    .exclude(phone_number="")
                    .distinct()
                )

                room_sent = 0
                for contact in room_contacts:
                    active_session = (
                        WhatsAppSession.objects
                        .filter(
                            company=company,
                            phone_number=contact.phone_number,
                            is_active=True,
                            last_message_at__gte=window_threshold,
                        )
                        .order_by("-last_message_at")
                        .first()
                    )

                    if active_session is not None:
                        try:
                            WhatsAppChatService.send_reply(
                                from_number=bot_number,
                                to_number=contact.phone_number,
                                reply_text=message_body,
                            )
                            room_sent += 1
                            total_sent += 1
                        except Exception as exc:
                            logger.error(
                                "# [BOT MGMT] Error group_broadcast "
                                "directo a %s: %s",
                                contact.phone_number, exc,
                            )
                    else:
                        out_session = (
                            WhatsAppSession.objects
                            .filter(
                                company=company,
                                phone_number=contact.phone_number,
                            )
                            .order_by("-session_start")
                            .first()
                        )
                        if out_session is None:
                            logger.warning(
                                "# [BOT MGMT] No hay sesion "
                                "WhatsApp para %s -- omitido.",
                                contact.phone_number,
                            )
                            continue

                        pending = list(
                            out_session.pending_broadcast_messages
                            or []
                        )
                        pending.append({
                            "body": message_body,
                            "created_at": created_at_iso,
                        })
                        out_session.pending_broadcast_messages = (
                            pending
                        )
                        out_session.save(
                            update_fields=[
                                "pending_broadcast_messages",
                            ],
                        )

                        if renewal_template:
                            try:
                                WhatsAppChatService.send_quick_reply(
                                    from_number=bot_number,
                                    to_number=contact.phone_number,
                                    content_sid=(
                                        renewal_template.content_sid
                                    ),
                                    content_variables={
                                        "1": (
                                            contact.name
                                            or contact.phone_number
                                        ),
                                        "2": company.name,
                                        "3": "/panel/",
                                    },
                                )
                                total_pending += 1
                            except Exception as exc:
                                logger.error(
                                    "# [BOT MGMT] Error renewal "
                                    "a %s: %s",
                                    contact.phone_number, exc,
                                )

            django_messages.success(
                request,
                f"Circular enviada: {total_sent} entregado(s) "
                f"directamente, {total_pending} renewal(s) "
                f"enviado(s) con mensaje en cola.",
            )
            return redirect("analytics:bot_management")

        if action == "direct_broadcast":
            message_body = (
                request.POST.get("message", "").strip()
            )
            company_user_pk = (
                request.POST.get("company_user_pk", "").strip()
            )
            if not message_body:
                django_messages.error(
                    request, "Debes escribir un mensaje.",
                )
                return redirect("analytics:bot_management")

            contacts_qs = Contact.objects.filter(
                company=company,
                opt_out_broadcast=False,
            ).exclude(phone_number="")

            if company_user_pk:
                contacts_qs = contacts_qs.filter(
                    company_user__pk=company_user_pk,
                )

            sent_count = 0
            for contact in contacts_qs.distinct():
                try:
                    WhatsAppChatService.send_reply(
                        from_number=bot_number,
                        to_number=contact.phone_number,
                        reply_text=message_body,
                    )
                    sent_count += 1
                except Exception as exc:
                    logger.error(
                        "# [BOT MGMT] Error direct_broadcast "
                        "a %s: %s",
                        contact.phone_number, exc,
                    )

            django_messages.success(
                request,
                f"Circular 1:1 enviada a {sent_count} contacto(s).",
            )
            return redirect("analytics:bot_management")

        django_messages.warning(request, "Accion no reconocida.")
        return redirect("analytics:bot_management")



