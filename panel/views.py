


from django .contrib .auth import update_session_auth_hash 
from django .contrib .auth .views import LoginView ,LogoutView 
from django .contrib import messages as django_messages 
from django .views .generic import TemplateView ,View ,ListView ,UpdateView ,CreateView ,DeleteView 
from django .shortcuts import redirect ,render 
from django .db import models as django_models 
from django .db .models import Q ,Prefetch 
from django .utils .timezone import now 
from django .forms import modelformset_factory 
from django .views .decorators .csrf import csrf_exempt 
from django .utils .decorators import method_decorator 

from panel .mixins import CompanyUserRequiredMixin ,AdminRoleRequiredMixin ,WorkshopRequiredMixin ,SupervisorAccessMixin 
from panel .models import AnalyticsProfile 
from panel .forms import (
PanelAuthenticationForm ,
PresenceStatusForm ,
SectionForm ,
SectionScheduleForm ,
ContactForm ,
CallFlowForm ,
CorporateVoiceProfileForm ,
BlockedCallerForm ,
CompanyUserCreateForm ,
PanelPasswordChangeForm ,
PanelSetPasswordForm ,
DataCaptureSetForm ,
)
from ivr_config .models import (
Section ,
SectionSchedule ,
Contact ,
PresenceStatus ,
CompanyUser ,
CallFlow ,
PhoneNumber ,
CorporateVoiceProfile ,
BlockedCaller ,
DataCaptureSet ,
)
from whatsapp .models import WhatsAppTemplate ,WhatsAppSession 
from work_order_processor .models import WorkOrder ,WorkOrderEntry ,WorkOrderEntryLine 
from work_order_processor .services import find_cached_classification 
from work_order_processor .tasks import classify_fault_line ,generate_period_excel ,process_work_order_pdf 
from fleet .models import MachineAsset 
from chat .models import ChatRoom 
from chat .views import ChatRoomView ,ChatMessagesPollingView ,ChatRoomListView 
import logging 
import plotly .graph_objects as go 
import plotly .io as pio 

logger =logging .getLogger (__name__ )


class OperatorDashboardView (WorkshopRequiredMixin ,TemplateView ):
    """
    Landing view for CompanyUsers with the WORKSHOP role.
    Displays a selector with the two remaining work-order entry paths:
    Form (structured web form) and Upload (photo or PDF with Gemini Vision
    extraction). Via B (STT) was removed in Hito 7 / S023.
    Accessible to WORKSHOP and ADMIN roles (WorkshopRequiredMixin).
    ---
    Vista de aterrizaje para CompanyUsers con rol WORKSHOP.
    Muestra un selector con las dos vías de entrada de partes disponibles:
    Form (formulario web estructurado) y Upload (foto o PDF con extracción
    Gemini Vision). La Vía B (STT) fue eliminada en Hito 7 / S023.
    Accesible para los roles WORKSHOP y ADMIN (WorkshopRequiredMixin).
    """

    template_name ="panel/operator/dashboard.html"

    def get_context_data (self ,**kwargs ):
        """
        Build context with company, company_user and own_presence for the operator dashboard.
        Clears any residual error messages left in the session by access-control mixins
        when the authenticated user has the WORKSHOP role — these messages are artefacts
        of prior failed access attempts and are not meaningful to the operator.
        ---
        Construye el contexto con company, company_user y own_presence para el dashboard
        del operario de taller.
        Limpia los mensajes de error residuales dejados en la sesión por los mixins de
        control de acceso cuando el usuario autenticado tiene rol WORKSHOP — estos mensajes
        son artefactos de intentos de acceso previos fallidos y no son significativos
        para el operario.
        """
        context =super ().get_context_data (**kwargs )



        company_user =self .request .user .company_user 










        if company_user .role =="WORKSHOP":
            from django .contrib .messages import get_messages as _get_messages 
            storage =_get_messages (self .request )
            storage .used =True 
        company =company_user .company 



        own_presence =PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

        context ["company"]=company 
        context ["company_user"]=company_user 
        context ["own_presence"]=own_presence 
        context ["active_nav"]="operator_dashboard"

        return context 


class WorkerSignupView (View ):
    """
    Worker self-registration view — DEACTIVATED (H13 redesign).
    Registration is now exclusively managed by Supervisors from the panel.
    Both GET and POST redirect to login with an informative message.
    The URL is preserved to avoid breaking existing links.
    ---
    Vista de auto-registro de operarios — DESACTIVADA (rediseño H13).
    El registro es ahora gestionado exclusivamente por los Supervisores
    desde el panel. GET y POST redirigen al login con un mensaje informativo.
    La URL se preserva para no romper enlaces existentes.
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Redirects to login with an informative message.
        ---
        Redirige al login con un mensaje informativo.
        """
        django_messages .info (
        request ,
        "El registro de trabajadores lo gestiona el supervisor desde el panel.",
        )
        return redirect ("/panel/login/")

    def post (self ,request ,*args ,**kwargs ):
        """
        Redirects to login with an informative message.
        ---
        Redirige al login con un mensaje informativo.
        """
        django_messages .info (
        request ,
        "El registro de trabajadores lo gestiona el supervisor desde el panel.",
        )
        return redirect ("/panel/login/")



class CompanyUserCreateView (SupervisorAccessMixin ,View ):
    """
    Allows a SUPERVISOR or ADMIN to create a new CompanyUser for their company.
    Creates the underlying auth.User with password '1234' and sets
    must_change_password=True so the new user must change it on first login.
    Optionally links the new user to a Section and creates or retrieves the
    associated Contact record if a phone number is provided.
    Updated H13: mixin changed to SupervisorAccessMixin; section, phone_number
    and is_ivr_active fields added; Contact auto-creation/linking logic added.
    ---
    Permite a un SUPERVISOR o ADMIN crear un nuevo CompanyUser para su empresa.
    Crea el auth.User subyacente con contraseña '1234' y establece
    must_change_password=True para forzar el cambio en el primer acceso.
    Opcionalmente vincula el nuevo usuario a una Section y crea o recupera
    el Contact asociado si se indica un número de teléfono.
    Actualización H13: mixin cambiado a SupervisorAccessMixin; campos section,
    phone_number e is_ivr_active añadidos; lógica de auto-creación/vinculación
    de Contact incorporada.
    """

    template_name ="panel/users/create.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _get_context (self ,request ,form =None ):
        """
        Builds base template context with company, company_user, own_presence
        and sections queryset for the section selector JS pre-fill.
        ---
        Construye el contexto base con company, company_user, own_presence
        y el queryset de secciones para el pre-relleno JS del selector de sección.
        """
        cu =request .user .company_user 
        company =cu .company 
        if form is None :
            form =CompanyUserCreateForm ()
            form .fields ["section"].queryset =Section .objects .filter (
            company =company 
            ).order_by ("name")
        return {
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"users",
        "form":form ,
        "sections_data":list (
        Section .objects .filter (company =company ).values (
        "id","default_role"
        )
        ),
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the user creation form with section queryset scoped to company.
        ---
        Renderiza el formulario de creación de usuario con el queryset de secciones
        acotado a la empresa.
        """
        return render (request ,self .template_name ,self ._get_context (request ))

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the form, creates auth.User and CompanyUser, optionally creates
        or retrieves the Contact record and links it to the new CompanyUser,
        then optionally assigns the new user to the selected Section.
        Redirects to user list on success.
        ---
        Valida el formulario, crea auth.User y CompanyUser, opcionalmente crea
        o recupera el Contact y lo vincula al nuevo CompanyUser, y opcionalmente
        asigna el nuevo usuario a la Section seleccionada.
        Redirige a la lista de usuarios en caso de éxito.
        """
        from django .contrib .auth .models import User as AuthUser 

        cu =request .user .company_user 
        company =cu .company 

        form =CompanyUserCreateForm (request .POST )
        form .fields ["section"].queryset =Section .objects .filter (
        company =company 
        ).order_by ("name")

        if not form .is_valid ():
            context =self ._get_context (request ,form )
            return render (request ,self .template_name ,context )



        auth_user =AuthUser .objects .create_user (
        username =form .cleaned_data ["username"],
        first_name =form .cleaned_data .get ("first_name",""),
        last_name =form .cleaned_data .get ("last_name",""),
        password =form .get_initial_password (),
        is_staff =False ,
        is_superuser =False ,
        )



        new_cu =CompanyUser .objects .create (
        user =auth_user ,
        company =company ,
        role =form .cleaned_data ["role"],
        is_active =True ,
        must_change_password =True ,
        )
        logger .info (
        "# [USER CREATE] CompanyUser pk=%s (username='%s') creado por %s.",
        new_cu .pk ,
        auth_user .username ,
        request .user .username ,
        )







        _schedule_pk =request .POST .get ("workday_schedule_pk","").strip ()
        if _schedule_pk :
            try :
                from ivr_config .models import WorkdaySchedule as _WorkdaySchedule 
                _schedule =_WorkdaySchedule .objects .get (pk =int (_schedule_pk ),company =company )
                new_cu .workday_schedule =_schedule 
                new_cu .save (update_fields =["workday_schedule"])
                logger .info (
                "# [USER CREATE] WorkdaySchedule pk=%s asignado a CompanyUser pk=%s.",
                _schedule .pk ,
                new_cu .pk ,
                )
            except (ValueError ,TypeError ):
                logger .warning (
                "# [USER CREATE] workday_schedule_pk='%s' no es un entero valido — ignorado.",
                _schedule_pk ,
                )
            except _WorkdaySchedule .DoesNotExist :
                logger .warning (
                "# [USER CREATE] WorkdaySchedule pk=%s no pertenece a la empresa pk=%s — ignorado.",
                _schedule_pk ,
                company .pk ,
                )










        phone_number =form .cleaned_data .get ("phone_number","")
        is_ivr_active =form .cleaned_data .get ("is_ivr_active",True )
        section =form .cleaned_data .get ("section")



        _display_name =(
        f"{form.cleaned_data.get('first_name', '')} "
        f"{form.cleaned_data.get('last_name', '')}".strip ()
        or auth_user .username 
        )

        contact =None 

        if phone_number :


            contact ,created =Contact .objects .get_or_create (
            company =company ,
            phone_number =phone_number ,
            defaults ={
            "name":_display_name ,
            "is_internal":is_ivr_active ,
            "company_user":new_cu ,
            },
            )
            if not created :


                contact .company_user =new_cu 
                contact .is_internal =is_ivr_active 
                contact .save (update_fields =["company_user","is_internal"])
                logger .info (
                "# [USER CREATE] Contact existente pk=%s vinculado a CompanyUser pk=%s.",
                contact .pk ,
                new_cu .pk ,
                )
            else :
                logger .info (
                "# [USER CREATE] Contact nuevo pk=%s creado y vinculado a CompanyUser pk=%s.",
                contact .pk ,
                new_cu .pk ,
                )
        elif section is not None :




            contact =Contact .objects .create (
            company =company ,
            phone_number ="",
            name =_display_name ,
            is_internal =is_ivr_active ,
            company_user =new_cu ,
            )
            logger .info (
            "# [USER CREATE] Contact sin telefono pk=%s creado para CompanyUser pk=%s.",
            contact .pk ,
            new_cu .pk ,
            )



        if section is not None and contact is not None :
            section .contacts .add (contact )
            logger .info (
            "# [USER CREATE] Contact pk=%s añadido a Section pk=%s.",
            contact .pk ,
            section .pk ,
            )

        django_messages .success (
        request ,
        f"Usuario '{auth_user.username}' creado correctamente. "
        f"Deberá cambiar su contraseña en el primer acceso."
        )
        return redirect ("/panel/users/")



class CompanyUserListView (SupervisorAccessMixin ,ListView ):
    """
    Lists all CompanyUser accounts belonging to the authenticated user's company.
    Accessible only to users with the ADMIN role.
    Supports optional filtering by section via GET parameter ?section=<pk>.
    ---
    Lista todas las cuentas CompanyUser pertenecientes a la empresa del usuario autenticado.
    Solo accesible para usuarios con rol ADMIN.
    Soporta filtrado opcional por sección mediante el parámetro GET ?section=<pk>.
    """

    model =CompanyUser 
    template_name ="panel/users/list.html"
    context_object_name ="company_users"

    def get_queryset (self ):
        """
        Returns CompanyUser records scoped to the authenticated user's company.
        If the GET parameter 'section' is present and valid, filters by the
        contacts assigned to that section (company-scoped, injection-safe).
        ---
        Retorna los registros CompanyUser acotados a la empresa del usuario autenticado.
        Si el parámetro GET 'section' está presente y es válido, filtra por los
        contactos asignados a esa sección (acotado a empresa, seguro contra inyección).
        """
        company =self .request .user .company_user .company 
        qs =CompanyUser .objects .filter (
        company =company 
        ).select_related ("user").order_by ("user__username")


        section_pk =self .request .GET .get ("section","").strip ()
        if section_pk :
            try :
                from ivr_config .models import Section as _Section 
                section_obj =_Section .objects .get (pk =int (section_pk ),company =company )
                qs =qs .filter (
                contact_profile__sections =section_obj ,
                )
            except (ValueError ,TypeError ,_Section .DoesNotExist ):
                pass 

        return qs 

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and sections list to template context.
        Passes the currently selected section pk for filter persistence.
        ---
        Añade company, company_user y lista de secciones al contexto de la plantilla.
        Pasa el pk de sección seleccionado actualmente para persistencia del filtro.
        """
        context =super ().get_context_data (**kwargs )
        company =self .request .user .company_user .company 
        context ["company"]=company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="users"
        context ["sections"]=Section .objects .filter (
        company =company 
        ).order_by ("name")
        context ["selected_section"]=self .request .GET .get ("section","")
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class CompanyUserUpdateView (SupervisorAccessMixin ,UpdateView ):
    """
    Allows an ADMIN to update the role, active status and workday schedule
    of a CompanyUser belonging to the same company.
    Prevents editing users from other companies.
    The workday_schedule queryset is restricted to the authenticated
    company's own schedules via get_form() — no schedule from another
    company can be selected.
    ---
    Permite a un ADMIN actualizar el rol, el estado activo y el horario de
    jornada de un CompanyUser de la misma empresa.
    Impide editar usuarios de otras empresas.
    El queryset de workday_schedule se restringe a los horarios de la empresa
    autenticada mediante get_form() — ningún horario de otra empresa puede
    ser seleccionado.
    """

    model =CompanyUser 
    template_name ="panel/users/form.html"
    fields =["role","is_active","workday_schedule"]

    def get_queryset (self ):
        """
        Restricts the queryset to CompanyUser records of the authenticated user's company.
        ---
        Restringe el queryset a los registros CompanyUser de la empresa del usuario autenticado.
        """
        return CompanyUser .objects .filter (
        company =self .request .user .company_user .company 
        )

    def get_form (self ,form_class =None ):
        """
        Restricts the workday_schedule queryset to the authenticated user's
        company so that no foreign schedule can be selected from the dropdown.
        Marks workday_schedule as optional (blank allowed).
        ---
        Restringe el queryset de workday_schedule a la empresa del usuario
        autenticado para que ningún horario externo pueda seleccionarse.
        Marca workday_schedule como opcional (blank permitido).
        """
        from ivr_config .models import WorkdaySchedule 
        form =super ().get_form (form_class )
        company =self .request .user .company_user .company 
        form .fields ["workday_schedule"].queryset =WorkdaySchedule .objects .filter (
        company =company 
        ).order_by ("label")
        form .fields ["workday_schedule"].required =False 
        form .fields ["workday_schedule"].widget .attrs .update ({"class":"form-select"})
        return form 

    def post (self ,request ,*args ,**kwargs ):
        """
        Handles standard update and the force-reset action.
        Persists the phone_number field by updating or creating the linked
        Contact record. Respects the 'next' POST parameter for redirect.
        If POST contains 'force_reset', sets must_change_password=True and redirects.
        ---
        Gestiona la actualización estándar y la acción de forzar reset.
        Persiste el campo phone_number actualizando o creando el registro
        Contact vinculado. Respeta el parámetro POST 'next' para la redirección.
        Si el POST contiene 'force_reset', establece must_change_password=True y redirige.
        """
        self .object =self .get_object ()
        next_url =request .POST .get ("next","").strip ()or "/panel/users/"






        if self .object .role in ("WORKSHOP","WORKSHOPBOSS","DRIVER"):
            self .object .is_active =True 
            self .object .save (update_fields =["is_active"])
        if "force_reset"in request .POST :
            self .object .must_change_password =True 
            self .object .save (update_fields =["must_change_password"])
            django_messages .success (
            request ,
            f"Se ha forzado el cambio de contraseña para "
            f"'{self.object.user.username}'."
            )
            return redirect (next_url )




        request .session ["_cu_update_next"]=next_url 


        phone_number =request .POST .get ("phone_number","").strip ()
        if phone_number :
            from ivr_config .models import Contact as _Contact 
            company =self .object .company 
            _contact =_Contact .objects .filter (
            company =company ,company_user =self .object 
            ).first ()
            if _contact is not None :


                if _contact .phone_number !=phone_number :
                    _contact .phone_number =phone_number 
                    _contact .save (update_fields =["phone_number"])
                    logger .info (
                    "# [USER UPDATE] Contact pk=%s phone_number actualizado a '%s'.",
                    _contact .pk ,phone_number ,
                    )
            else :


                _display =(
                self .object .user .get_full_name ()or self .object .user .username 
                )
                _Contact .objects .create (
                company =company ,
                phone_number =phone_number ,
                name =_display ,
                is_internal =True ,
                company_user =self .object ,
                )
                logger .info (
                "# [USER UPDATE] Contact nuevo creado para CompanyUser pk=%s con phone '%s'.",
                self .object .pk ,phone_number ,
                )








        _section_pk_raw =request .POST .get ("section_pk","").strip ()
        if _section_pk_raw is not None :
            from ivr_config .models import Contact as _Contact2 
            from ivr_config .models import Section as _Section 
            _cu_contact =_Contact2 .objects .filter (
            company =self .object .company ,
            company_user =self .object ,
            ).first ()
            if _cu_contact is not None :


                _current_sections =_Section .objects .filter (
                company =self .object .company ,
                contacts =_cu_contact ,
                )
                for _sec in _current_sections :
                    _sec .contacts .remove (_cu_contact )
                    logger .info (
                    "# [USER UPDATE] Contact pk=%s desvinculado de Section pk=%s.",
                    _cu_contact .pk ,_sec .pk ,
                    )


                if _section_pk_raw :
                    try :
                        _new_section =_Section .objects .get (
                        pk =int (_section_pk_raw ),
                        company =self .object .company ,
                        )
                        _new_section .contacts .add (_cu_contact )
                        logger .info (
                        "# [USER UPDATE] Contact pk=%s vinculado a Section pk=%s.",
                        _cu_contact .pk ,_new_section .pk ,
                        )
                    except (ValueError ,TypeError ,_Section .DoesNotExist ):
                        logger .warning (
                        "# [USER UPDATE] section_pk='%s' inválido o no pertenece a la empresa — ignorado.",
                        _section_pk_raw ,
                        )
        return super ().post (request ,*args ,**kwargs )

    def get_success_url (self ):
        """
        Redirects to the 'next' URL stored in session, or falls back to user list.
        ---
        Redirige a la URL 'next' almacenada en sesión, o a la lista de usuarios
        como fallback.
        """
        django_messages .success (
        self .request ,
        f"Usuario '{self.object.user.username}' actualizado correctamente."
        )
        next_url =self .request .session .pop ("_cu_update_next","/panel/users/")
        return next_url 

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence, contact_phone and next_url
        to context.
        next_url resolution order: POST['next'] → GET['next'] → HTTP_REFERER → /panel/users/.
        Using POST['next'] first ensures the origin is preserved across
        validation re-renders (when the form is re-POSTed after an error,
        HTTP_REFERER would already point to the edit page itself).
        ---
        Añade company, company_user, own_presence, contact_phone y next_url
        al contexto.
        Orden de resolución de next_url: POST['next'] → GET['next'] →
        HTTP_REFERER → /panel/users/.
        Priorizar POST['next'] garantiza que el origen se preserva en
        re-renders de validación (cuando el formulario se re-envía tras error,
        HTTP_REFERER ya apunta a la propia página de edición).
        """
        if not hasattr (self ,'object')or self .object is None :
            self .object =self .get_object ()
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="users"
        from ivr_config .models import Contact as _Contact 
        _contact =_Contact .objects .filter (
        company =self .object .company ,
        company_user =self .object ,
        ).first ()
        context ["contact_phone"]=_contact .phone_number if _contact else ""


        context ["sections"]=Section .objects .filter (
        company =self .object .company 
        ).order_by ("name")


        context ["current_section_pk"]=(
        _contact .sections .values_list ("pk",flat =True ).first ()
        if _contact else None 
        )


        context ["next_url"]=(
        self .request .POST .get ("next","").strip ()
        or self .request .GET .get ("next","").strip ()
        or self .request .META .get ("HTTP_REFERER","")
        or "/panel/users/"
        )
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()



class CompanyUserBulkDeleteView (SupervisorAccessMixin ,View ):
    """
    Allows a SUPERVISOR or ADMIN to delete one or more CompanyUser accounts
    belonging to their company in a single operation.

    Flow:
      1. First POST (confirmed not set): evaluate selected pks.
         - Verify all pks belong to the authenticated user's company.
         - Detect IVR-risk users: those whose linked Contact is a member
           of at least one active Section (is_active=True).
         - If no IVR-risk users: delete immediately and redirect.
         - If IVR-risk users exist: render confirmation page with a warning
           listing the at-risk users and their affected sections.
      2. Second POST (confirmed=1): delete all selected users including
         IVR-risk ones. Contact.company_user is SET_NULL by the DB cascade;
         auth.User is deleted explicitly to ensure full cleanup.

    POST /panel/users/bulk-delete/
         Body: selected_users (list of pks)
               confirmed       (optional, '1' to bypass risk warning)
    ---
    Permite a un SUPERVISOR o ADMIN eliminar una o varias cuentas CompanyUser
    de su empresa en una única operación.

    Flujo:
      1. Primer POST (sin confirmed): evaluar pks seleccionados.
         - Verificar que todos los pks pertenecen a la empresa autenticada.
         - Detectar usuarios en riesgo IVR: aquellos cuyo Contact vinculado
           es miembro de al menos una Section activa (is_active=True).
         - Sin usuarios en riesgo: eliminar directamente y redirigir.
         - Con usuarios en riesgo: renderizar página de confirmación con
           aviso listando los usuarios en riesgo y sus secciones afectadas.
      2. Segundo POST (confirmed=1): eliminar todos los seleccionados
         incluidos los en riesgo. Contact.company_user queda a NULL por
         el cascade de la BD; auth.User se elimina explícitamente.

    POST /panel/users/bulk-delete/
         Body: selected_users (lista de pks)
               confirmed       (opcional, '1' para saltarse el aviso de riesgo)
    """

    template_name ="panel/users/bulk_delete_confirm.html"

    def _get_own_presence (self ,request ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _resolve_users (self ,pks_raw ,company ):
        """
        Returns a queryset of CompanyUser records for the given pk list,
        scoped to the provided company. Silently drops invalid or
        cross-company pks.
        ---
        Retorna un queryset de CompanyUser para la lista de pks dada,
        acotado a la empresa proporcionada. Descarta silenciosamente los
        pks inválidos o de otras empresas.
        """
        valid_pks =[]
        for raw in pks_raw :
            try :
                valid_pks .append (int (raw ))
            except (ValueError ,TypeError ):
                pass 
        return CompanyUser .objects .filter (
        pk__in =valid_pks ,
        company =company ,
        ).select_related ("user")

    def _classify_risk (self ,users_qs ):
        """
        Splits the given queryset into two lists:
          - safe_users: CompanyUser records with no active IVR contact.
          - risk_users: CompanyUser records whose linked Contact belongs
            to at least one active Section (is_active=True). Each entry is
            a dict with keys 'cu' and 'sections' (list of Section names).
        A CompanyUser is IVR-risk if:
            contact = Contact.objects.filter(company_user=cu).first()
            contact is not None AND contact.sections.filter(is_active=True).exists()
        ---
        Divide el queryset en dos listas:
          - safe_users: CompanyUser sin contacto IVR activo.
          - risk_users: CompanyUser cuyo Contact vinculado pertenece
            a al menos una Section activa (is_active=True). Cada entrada
            es un dict con claves 'cu' y 'sections' (lista de nombres).
        Un CompanyUser es riesgo IVR si:
            contact = Contact.objects.filter(company_user=cu).first()
            contact is not None AND contact.sections.filter(is_active=True).exists()
        """
        safe_users =[]
        risk_users =[]
        for cu in users_qs :
            contact =Contact .objects .filter (company_user =cu ).first ()
            if contact is not None and contact .sections .filter (is_active =True ).exists ():
                active_sections =list (
                contact .sections .filter (is_active =True ).values_list ("name",flat =True )
                )
                risk_users .append ({"cu":cu ,"sections":active_sections })
            else :
                safe_users .append (cu )
        return safe_users ,risk_users 

    def _delete_users (self ,cu_list ):
        """
        Deletes all CompanyUser records in the given list together with
        their underlying auth.User. The DB cascade sets Contact.company_user
        to NULL automatically (SET_NULL). auth.User is deleted explicitly
        to guarantee full cleanup regardless of cascade configuration.
        Returns the count of deleted CompanyUser records.
        ---
        Elimina todos los registros CompanyUser de la lista junto con su
        auth.User subyacente. El cascade de BD pone Contact.company_user
        a NULL automáticamente (SET_NULL). auth.User se elimina de forma
        explícita para garantizar la limpieza total independientemente
        de la configuración del cascade.
        Retorna el contador de CompanyUser eliminados.
        """
        from django .contrib .auth .models import User as AuthUser 
        count =0 
        for cu in cu_list :
            auth_user_pk =cu .user .pk 
            username =cu .user .username 
            cu .delete ()
            AuthUser .objects .filter (pk =auth_user_pk ).delete ()
            logger .info (
            "# [USER BULK DELETE] CompanyUser '%s' eliminado.",
            username ,
            )
            count +=1 
        return count 

    def post (self ,request ,*args ,**kwargs ):
        """
        Handles both the initial evaluation POST and the confirmed deletion POST.
        On the first POST without 'confirmed': resolves users, classifies IVR risk
        and either deletes immediately (no risk) or renders the confirmation page.
        On the second POST with confirmed='1': deletes all selected users including
        those previously flagged as IVR-risk.
        ---
        Gestiona tanto el POST de evaluación inicial como el POST de confirmación.
        En el primer POST sin 'confirmed': resuelve usuarios, clasifica riesgo IVR
        y elimina directamente (sin riesgo) o renderiza la página de confirmación.
        En el segundo POST con confirmed='1': elimina todos los seleccionados
        incluidos los marcados como riesgo IVR.
        """
        company =request .user .company_user .company 
        company_user =request .user .company_user 
        pks_raw =request .POST .getlist ("selected_users")
        confirmed =request .POST .get ("confirmed","")=="1"



        if not pks_raw :
            django_messages .warning (request ,"No se seleccionó ningún usuario.")
            return redirect ("/panel/users/")

        users_qs =self ._resolve_users (pks_raw ,company )



        users_qs =users_qs .exclude (pk =company_user .pk )

        if not users_qs .exists ():
            django_messages .warning (request ,"No se seleccionó ningún usuario válido.")
            return redirect ("/panel/users/")

        if confirmed :


            count =self ._delete_users (list (users_qs ))
            django_messages .success (
            request ,
            f"{count} usuario{'s' if count != 1 else ''} eliminado"
            f"{'s' if count != 1 else ''} correctamente.",
            )
            return redirect ("/panel/users/")



        safe_users ,risk_users =self ._classify_risk (users_qs )

        if not risk_users :


            count =self ._delete_users (safe_users )
            django_messages .success (
            request ,
            f"{count} usuario{'s' if count != 1 else ''} eliminado"
            f"{'s' if count != 1 else ''} correctamente.",
            )
            return redirect ("/panel/users/")



        all_users =[entry ["cu"]for entry in risk_users ]+safe_users 
        context ={
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (request ),
        "active_nav":"users",
        "risk_users":risk_users ,
        "safe_users":safe_users ,
        "selected_pks":[cu .pk for cu in all_users ],
        }
        return render (request ,self .template_name ,context )


class CompanyUserSectionUnlinkView (SupervisorAccessMixin ,View ):
    """
    Unlinks a CompanyUser from a Section by removing their associated Contact
    from the section's contacts M2M relation.
    Redirects back to the section edit form on success.
    Only affects the M2M membership — the CompanyUser account is preserved.
    Scoped to the authenticated user's company; rejects cross-company attempts.

    POST /panel/users/<pk>/unlink-section/
         Body: section_pk (required) — pk of the section to unlink from.
               next (optional)       — URL to redirect to after unlinking.
    ---
    Desvincula un CompanyUser de una Section eliminando su Contact asociado
    de la relación M2M de contactos de la sección.
    Redirige de vuelta al formulario de edición de sección en caso de éxito.
    Solo afecta a la membresía M2M — la cuenta CompanyUser se conserva.
    Acotado a la empresa del usuario autenticado; rechaza intentos entre empresas.

    POST /panel/users/<pk>/unlink-section/
         Body: section_pk (requerido) — pk de la sección de la que desvincular.
               next (opcional)        — URL a la que redirigir tras desvincular.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Removes the Contact associated with the CompanyUser from the given section.
        ---
        Elimina el Contact asociado al CompanyUser de la sección indicada.
        """
        company =request .user .company_user .company 
        section_pk =request .POST .get ("section_pk","").strip ()
        next_url =request .POST .get ("next","/panel/sections/")



        try :
            cu =CompanyUser .objects .get (pk =pk ,company =company )
        except CompanyUser .DoesNotExist :
            django_messages .error (request ,"Usuario no encontrado.")
            return redirect (next_url )



        try :
            section =Section .objects .get (pk =int (section_pk ),company =company )
        except (ValueError ,TypeError ,Section .DoesNotExist ):
            django_messages .error (request ,"Sección no encontrada.")
            return redirect (next_url )



        contact =getattr (cu ,"contact",None )
        if contact is None :
            try :
                from ivr_config .models import Contact as _Contact 
                contact =_Contact .objects .filter (
                company =company ,
                company_user =cu ,
                ).first ()
            except Exception :
                contact =None 

        if contact and section .contacts .filter (pk =contact .pk ).exists ():
            section .contacts .remove (contact )
            django_messages .success (
            request ,
            f"Trabajador '{cu.user.get_full_name() or cu.user.username}' "
            f"desvinculado de la sección '{section.name}'."
            )
        else :
            django_messages .warning (
            request ,
            f"El trabajador no estaba vinculado a la sección '{section.name}'."
            )

        return redirect (next_url )



class WorkerScheduleUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates the workday_schedule FK of a single CompanyUser belonging to
    the authenticated user's company. Designed for AJAX calls from the
    section edit form worker table.
    Returns JSON {"ok": true} on success or {"ok": false, "error": "..."}
    on failure.

    POST /panel/users/<pk>/schedule/
         Body: schedule_pk — pk of the WorkdaySchedule to assign,
                             or empty string to clear the assignment.
    ---
    Actualiza la FK workday_schedule de un CompanyUser perteneciente a la
    empresa del usuario autenticado. Diseñado para llamadas AJAX desde la
    tabla de trabajadores del formulario de edición de sección.
    Devuelve JSON {"ok": true} en éxito o {"ok": false, "error": "..."}
    en caso de fallo.

    POST /panel/users/<pk>/schedule/
         Body: schedule_pk — pk del WorkdaySchedule a asignar,
                             o cadena vacía para limpiar la asignación.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Resolves the CompanyUser and the WorkdaySchedule (if provided),
        validates company ownership for both, and updates the FK.
        ---
        Resuelve el CompanyUser y el WorkdaySchedule (si se proporciona),
        valida la propiedad de empresa en ambos y actualiza la FK.
        """
        import json as _json 
        from django .http import JsonResponse 
        from ivr_config .models import WorkdaySchedule as _WorkdaySchedule 

        company =request .user .company_user .company 



        try :
            cu =CompanyUser .objects .get (pk =pk ,company =company )
        except CompanyUser .DoesNotExist :
            return JsonResponse ({"ok":False ,"error":"Usuario no encontrado."},status =404 )

        schedule_pk =request .POST .get ("schedule_pk","").strip ()

        if schedule_pk :


            try :
                schedule =_WorkdaySchedule .objects .get (pk =int (schedule_pk ),company =company )
            except (ValueError ,TypeError ,_WorkdaySchedule .DoesNotExist ):
                return JsonResponse ({"ok":False ,"error":"Horario no encontrado."},status =404 )
            cu .workday_schedule =schedule 
        else :


            cu .workday_schedule =None 

        cu .save (update_fields =["workday_schedule"])
        logger .info (
        "# [WORKER SCHEDULE] CompanyUser pk=%s — workday_schedule -> %s (por %s).",
        cu .pk ,
        cu .workday_schedule_id ,
        request .user .username ,
        )
        return JsonResponse ({"ok":True })


class SectionListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all Section records belonging to the authenticated user's company.
    Annotates each Section with two filtered counts:
      - ivr_contact_count: contacts whose linked CompanyUser is None or has a
        role other than WORKSHOP/DRIVER (true IVR contacts).
      - worker_count: contacts whose linked CompanyUser has role WORKSHOP or DRIVER.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros Section pertenecientes a la empresa del usuario autenticado.
    Anota cada Section con dos contadores filtrados:
      - ivr_contact_count: contactos cuyo CompanyUser vinculado es None o tiene un rol
        distinto de WORKSHOP/DRIVER (contactos IVR reales).
      - worker_count: contactos cuyo CompanyUser vinculado tiene rol WORKSHOP o DRIVER.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =Section 
    template_name ="panel/sections/list.html"
    context_object_name ="sections"

    def get_queryset (self ):
        """
        Returns Section records scoped to the authenticated user's company,
        annotated with ivr_contact_count and worker_count.
        ivr_contact_count excludes contacts linked to WORKSHOP or DRIVER users.
        worker_count counts only contacts linked to WORKSHOP or DRIVER users.
        ---
        Retorna los registros Section acotados a la empresa del usuario autenticado,
        anotados con ivr_contact_count y worker_count.
        ivr_contact_count excluye contactos vinculados a usuarios WORKSHOP o DRIVER.
        worker_count cuenta solo los contactos vinculados a usuarios WORKSHOP o DRIVER.
        """
        from django .db .models import Count ,Q as _Q 
        return Section .objects .filter (
        company =self .request .user .company_user .company 
        ).annotate (


        ivr_contact_count =Count (
        "contacts",
        filter =_Q (
        _Q (contacts__company_user__isnull =True )
        |~_Q (contacts__company_user__role__in =["WORKSHOP","DRIVER"])
        ),
        distinct =True ,
        ),


        worker_count =Count (
        "contacts",
        filter =_Q (contacts__company_user__role__in =["WORKSHOP","DRIVER"]),
        distinct =True ,
        ),
        ).order_by ("name")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="sections"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class SectionCreateView (AdminRoleRequiredMixin ,CreateView ):
    """
    Allows an ADMIN to create a new Section for their company.
    Automatically assigns the company from the authenticated user's CompanyUser.
    Manages an inline SectionSchedule formset for time slot configuration.
    Updated 2026-04-16 (Step 37.C — Estrategia B): call_flow queryset restricted
    to the company's own active CallFlows.
    ---
    Permite a un ADMIN crear una nueva Section para su empresa.
    Asigna automáticamente la empresa desde el CompanyUser del usuario autenticado.
    Gestiona un formset inline de SectionSchedule para la configuración de horarios.
    Actualización 2026-04-16 (Paso 37.C — Estrategia B): queryset de call_flow
    restringido a los CallFlows activos de la empresa.
    """

    model =Section 
    form_class =SectionForm 
    template_name ="panel/sections/form.html"

    def _get_schedule_formset_class (self ):
        """
        Returns the SectionSchedule inline formset class with 5 empty extra forms.
        ---
        Retorna la clase de formset inline de SectionSchedule con 5 formularios extra vacíos.
        """
        return modelformset_factory (
        SectionSchedule ,
        form =SectionScheduleForm ,
        extra =5 ,
        can_delete =True ,
        )

    def get_form (self ,form_class =None ):
        """
        Restricts the contacts queryset to the authenticated user's company.
        Restricts the call_flow queryset to the company's own active CallFlows
        (Estrategia B — Step 37.C). call_flow is optional.
        Restricts the data_capture_set queryset to the company's own DataCaptureSets.
        data_capture_set is optional — can also be created inline from the section form.
        ---
        Restringe el queryset de contactos a la empresa del usuario autenticado.
        Restringe el queryset de call_flow a los CallFlows activos de la empresa
        (Estrategia B — Paso 37.C). call_flow es opcional.
        Restringe el queryset de data_capture_set a los DataCaptureSets de la empresa.
        data_capture_set es opcional — también puede crearse inline desde el formulario de sección.
        """
        form =super ().get_form (form_class )
        company =self .request .user .company_user .company 




        form .fields ["contacts"].queryset =Contact .objects .filter (
        company =company 
        ).exclude (
        company_user__role__in =["WORKSHOP","DRIVER"]
        )
        form .fields ["call_flow"].queryset =CallFlow .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("name")
        form .fields ["call_flow"].required =False 
        form .fields ["data_capture_set"].queryset =DataCaptureSet .objects .filter (
        company =company ,
        ).order_by ("name")
        form .fields ["data_capture_set"].required =False 


        from ivr_config .models import WorkdaySchedule as _WorkdaySchedule 
        form .fields ["workday_schedule"].queryset =_WorkdaySchedule .objects .filter (
        company =company ,
        ).order_by ("label")
        form .fields ["workday_schedule"].required =False 
        return form 

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the section form with an empty schedule formset.
        ---
        Renderiza el formulario de sección con un formset de horarios vacío.
        """
        self .object =None 
        form =self .get_form ()
        ScheduleFormSet =self ._get_schedule_formset_class ()
        schedule_formset =ScheduleFormSet (
        queryset =SectionSchedule .objects .none (),
        prefix ="schedules",
        )
        return self .render_to_response (
        self .get_context_data (form =form ,schedule_formset =schedule_formset )
        )

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates and saves both the section form and the schedule formset.
        ---
        Valida y guarda tanto el formulario de sección como el formset de horarios.
        """
        self .object =None 
        form =self .get_form ()
        ScheduleFormSet =self ._get_schedule_formset_class ()
        schedule_formset =ScheduleFormSet (
        request .POST ,
        queryset =SectionSchedule .objects .none (),
        prefix ="schedules",
        )
        if form .is_valid ()and schedule_formset .is_valid ():
            return self ._form_valid (form ,schedule_formset )
        return self .render_to_response (
        self .get_context_data (form =form ,schedule_formset =schedule_formset )
        )

    def _form_valid (self ,form ,schedule_formset ):
        """
        Saves the Section and all valid SectionSchedule entries.
        If 'capture_fields_json' is present in POST and non-empty, creates a new
        DataCaptureSet with the submitted fields and links it to the section,
        overriding any selector choice. If the selector has a value but no inline
        fields were submitted, the selected DataCaptureSet is linked as-is.
        ---
        Guarda la Section y todas las entradas SectionSchedule válidas.
        Si 'capture_fields_json' está presente en el POST y no está vacío, crea un
        nuevo DataCaptureSet con los campos enviados y lo vincula a la sección,
        sobreescribiendo cualquier selección del selector. Si el selector tiene valor
        pero no se enviaron campos inline, el DataCaptureSet seleccionado se vincula tal cual.
        """
        import json 
        from django .contrib import messages as django_messages 
        company =self .request .user .company_user .company 
        form .instance .company =company 



        raw_json =self .request .POST .get ("capture_fields_json","").strip ()
        capture_name =self .request .POST .get ("capture_set_name","").strip ()
        if raw_json and raw_json !="[]":


            try :
                parsed_fields =json .loads (raw_json )
            except (ValueError ,TypeError ):
                parsed_fields =[]
            dcs_name =capture_name or f"Captura — {form.cleaned_data.get('name', 'Sección')}"
            new_dcs =DataCaptureSet .objects .create (
            company =company ,
            name =dcs_name ,
            fields =parsed_fields ,
            )
            form .instance .data_capture_set =new_dcs 

        self .object =form .save ()
        schedules =schedule_formset .save (commit =False )
        for schedule in schedules :
            schedule .section =self .object 
            schedule .save ()
        for deleted in schedule_formset .deleted_objects :
            deleted .delete ()
        django_messages .success (
        self .request ,
        f"Sección '{self.object.name}' creada correctamente."
        )
        return redirect ("/panel/sections/")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence, action flag and schedule_formset.
        ---
        Añade company, company_user, own_presence, flag de acción y schedule_formset.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="sections"
        context ["action"]="Crear"
        context ["section_workers"]=[]
        if "schedule_formset"not in context :
            ScheduleFormSet =self ._get_schedule_formset_class ()
            context ["schedule_formset"]=ScheduleFormSet (
            queryset =SectionSchedule .objects .none (),
            prefix ="schedules",
            )
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class SectionUpdateView (AdminRoleRequiredMixin ,UpdateView ):
    """
    Allows an ADMIN to update an existing Section belonging to their company.
    Prevents editing sections from other companies.
    Manages an inline SectionSchedule formset pre-populated with existing time slots.
    ---
    Permite a un ADMIN actualizar una Section existente de su empresa.
    Impide editar secciones de otras empresas.
    Gestiona un formset inline de SectionSchedule prerellenado con las franjas existentes.
    """

    model =Section 
    form_class =SectionForm 
    template_name ="panel/sections/form.html"

    def _get_schedule_formset_class (self ):
        """
        Returns the SectionSchedule inline formset class with 3 extra empty forms.
        ---
        Retorna la clase de formset inline de SectionSchedule con 3 formularios extra vacíos.
        """
        return modelformset_factory (
        SectionSchedule ,
        form =SectionScheduleForm ,
        extra =3 ,
        can_delete =True ,
        )

    def get_queryset (self ):
        """
        Restricts the queryset to Section records of the authenticated user's company.
        ---
        Restringe el queryset a los registros Section de la empresa del usuario autenticado.
        """
        return Section .objects .filter (
        company =self .request .user .company_user .company 
        )

    def get_form (self ,form_class =None ):
        """
        Restricts the contacts queryset to the authenticated user's company.
        Restricts the call_flow queryset to the company's own active CallFlows
        (Estrategia B — Step 37.C). call_flow is optional.
        Restricts the data_capture_set queryset to the company's own DataCaptureSets.
        data_capture_set is optional — can also be created inline from the section form.
        ---
        Restringe el queryset de contactos a la empresa del usuario autenticado.
        Restringe el queryset de call_flow a los CallFlows activos de la empresa
        (Estrategia B — Paso 37.C). call_flow es opcional.
        Restringe el queryset de data_capture_set a los DataCaptureSets de la empresa.
        data_capture_set es opcional — también puede crearse inline desde el formulario de sección.
        """
        form =super ().get_form (form_class )
        company =self .request .user .company_user .company 




        form .fields ["contacts"].queryset =Contact .objects .filter (
        company =company 
        ).exclude (
        company_user__role__in =["WORKSHOP","DRIVER"]
        )
        form .fields ["call_flow"].queryset =CallFlow .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("name")
        form .fields ["call_flow"].required =False 
        form .fields ["data_capture_set"].queryset =DataCaptureSet .objects .filter (
        company =company ,
        ).order_by ("name")
        form .fields ["data_capture_set"].required =False 


        from ivr_config .models import WorkdaySchedule as _WorkdaySchedule 
        form .fields ["workday_schedule"].queryset =_WorkdaySchedule .objects .filter (
        company =company ,
        ).order_by ("label")
        form .fields ["workday_schedule"].required =False 
        return form 

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the section form pre-populated with existing schedules.
        ---
        Renderiza el formulario de sección prerellenado con los horarios existentes.
        """
        self .object =self .get_object ()
        form =self .get_form ()
        ScheduleFormSet =self ._get_schedule_formset_class ()
        schedule_formset =ScheduleFormSet (
        queryset =SectionSchedule .objects .filter (section =self .object ).order_by ("weekday","time_open"),
        prefix ="schedules",
        )
        return self .render_to_response (
        self .get_context_data (form =form ,schedule_formset =schedule_formset )
        )

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates and saves both the section form and the schedule formset.
        ---
        Valida y guarda tanto el formulario de sección como el formset de horarios.
        """
        self .object =self .get_object ()
        form =self .get_form ()
        ScheduleFormSet =self ._get_schedule_formset_class ()
        schedule_formset =ScheduleFormSet (
        request .POST ,
        queryset =SectionSchedule .objects .filter (section =self .object ).order_by ("weekday","time_open"),
        prefix ="schedules",
        )
        if form .is_valid ()and schedule_formset .is_valid ():
            return self ._form_valid (form ,schedule_formset )
        return self .render_to_response (
        self .get_context_data (form =form ,schedule_formset =schedule_formset )
        )

    def _form_valid (self ,form ,schedule_formset ):
        """
        Saves the Section and all valid SectionSchedule entries.
        Handles deletion of removed schedule entries.
        If 'capture_fields_json' is present in POST and non-empty, updates the
        existing linked DataCaptureSet fields in-place, or creates a new one if
        none is linked. If the selector has a value but no inline fields were
        submitted, the selected DataCaptureSet is linked as-is.
        ---
        Guarda la Section y todas las entradas SectionSchedule válidas.
        Gestiona la eliminación de las franjas horarias eliminadas.
        Si 'capture_fields_json' está presente en el POST y no está vacío, actualiza
        los campos del DataCaptureSet vinculado existente en su lugar, o crea uno nuevo
        si no hay ninguno vinculado. Si el selector tiene valor pero no se enviaron
        campos inline, el DataCaptureSet seleccionado se vincula tal cual.
        """
        import json 
        from django .contrib import messages as django_messages 
        company =self .request .user .company_user .company 



        raw_json =self .request .POST .get ("capture_fields_json","").strip ()
        capture_name =self .request .POST .get ("capture_set_name","").strip ()
        if raw_json and raw_json !="[]":


            try :
                parsed_fields =json .loads (raw_json )
            except (ValueError ,TypeError ):
                parsed_fields =[]
            existing_dcs =self .object .data_capture_set if self .object .data_capture_set_id else None 
            if existing_dcs is not None :


                if capture_name :
                    existing_dcs .name =capture_name 
                existing_dcs .fields =parsed_fields 
                existing_dcs .save (update_fields =["name","fields","updated_at"])
            else :


                dcs_name =capture_name or f"Captura — {form.cleaned_data.get('name', 'Sección')}"
                new_dcs =DataCaptureSet .objects .create (
                company =company ,
                name =dcs_name ,
                fields =parsed_fields ,
                )
                form .instance .data_capture_set =new_dcs 











        _worker_contacts =list (
        self .object .contacts .filter (
        company_user__role__in =["WORKSHOP","DRIVER"]
        ).values_list ("pk",flat =True )
        )

        self .object =form .save ()



        if _worker_contacts :
            self .object .contacts .add (*_worker_contacts )
            logger .info (
            "# [SECTION UPDATE] %d contacto(s) WORKSHOP/DRIVER preservados en M2M de Section pk=%s.",
            len (_worker_contacts ),
            self .object .pk ,
            )

        schedules =schedule_formset .save (commit =False )
        for schedule in schedules :
            schedule .section =self .object 
            schedule .save ()
        for deleted in schedule_formset .deleted_objects :
            deleted .delete ()
        django_messages .success (
        self .request ,
        f"Sección '{self.object.name}' actualizada correctamente."
        )
        return redirect ("/panel/sections/")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence, action flag and schedule_formset.
        ---
        Añade company, company_user, own_presence, flag de acción y schedule_formset.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="sections"
        context ["action"]="Guardar"


        _section_contacts =(
        self .object .contacts 
        .filter (company_user__isnull =False )
        .select_related ("company_user__user")
        .order_by ("company_user__user__username")
        )
        _workers =[]
        for _contact in _section_contacts :
            _cu =_contact .company_user 
            _cu .contact_phone =_contact .phone_number or ""
            _cu .is_ivr_active =_contact .is_internal 
            _workers .append (_cu )
        context ["section_workers"]=_workers 


        from ivr_config .models import WorkdaySchedule as _WorkdaySchedule 
        context ["workday_schedules"]=list (
        _WorkdaySchedule .objects .filter (company =self .request .user .company_user .company )
        .order_by ("label")
        )
        if "schedule_formset"not in context :
            ScheduleFormSet =self ._get_schedule_formset_class ()
            context ["schedule_formset"]=ScheduleFormSet (
            queryset =SectionSchedule .objects .filter (section =self .object ).order_by ("weekday","time_open"),
            prefix ="schedules",
            )
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class ContactListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all Contact records belonging to the authenticated user's company.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros Contact pertenecientes a la empresa del usuario autenticado.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =Contact 
    template_name ="panel/contacts/list.html"
    context_object_name ="contacts"

    def get_queryset (self ):
        """
        Returns Contact records scoped to the authenticated user's company.
        ---
        Retorna los registros Contact acotados a la empresa del usuario autenticado.
        """
        return Contact .objects .filter (
        company =self .request .user .company_user .company 
        ).select_related ("company_user__user").order_by ("name")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="contacts"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class ContactCreateView (AdminRoleRequiredMixin ,CreateView ):
    """
    Allows an ADMIN to create a new Contact for their company.
    Automatically assigns the company from the authenticated user's CompanyUser.
    Restricts the company_user field to users belonging to the same company.
    ---
    Permite a un ADMIN crear un nuevo Contact para su empresa.
    Asigna automáticamente la empresa desde el CompanyUser del usuario autenticado.
    Restringe el campo company_user a usuarios de la misma empresa.
    """

    model =Contact 
    form_class =ContactForm 
    template_name ="panel/contacts/form.html"

    def get_form (self ,form_class =None ):
        """
        Restricts the company_user queryset in the form to the authenticated user's company.
        ---
        Restringe el queryset de company_user del formulario a la empresa del usuario autenticado.
        """
        form =super ().get_form (form_class )
        form .fields ["company_user"].queryset =CompanyUser .objects .filter (
        company =self .request .user .company_user .company 
        ).select_related ("user")
        form .fields ["company_user"].required =False 
        return form 

    def form_valid (self ,form ):
        """
        Assigns the company before saving the new Contact.
        ---
        Asigna la empresa antes de guardar el nuevo Contact.
        """
        form .instance .company =self .request .user .company_user .company 
        return super ().form_valid (form )

    def get_success_url (self ):
        """
        Redirects to the contact list after a successful creation.
        ---
        Redirige a la lista de contactos tras una creación correcta.
        """
        from django .contrib import messages as django_messages 
        django_messages .success (
        self .request ,
        f"Contacto '{self.object.name}' creado correctamente."
        )
        return "/panel/contacts/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="contacts"
        context ["action"]="Crear"
        context ["section_workers"]=[]
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class ContactUpdateView (AdminRoleRequiredMixin ,UpdateView ):
    """
    Allows an ADMIN to update an existing Contact belonging to their company.
    Prevents editing contacts from other companies.
    Restricts the company_user field to users belonging to the same company.
    ---
    Permite a un ADMIN actualizar un Contact existente de su empresa.
    Impide editar contactos de otras empresas.
    Restringe el campo company_user a usuarios de la misma empresa.
    """

    model =Contact 
    form_class =ContactForm 
    template_name ="panel/contacts/form.html"

    def get_queryset (self ):
        """
        Restricts the queryset to Contact records of the authenticated user's company.
        ---
        Restringe el queryset a los registros Contact de la empresa del usuario autenticado.
        """
        return Contact .objects .filter (
        company =self .request .user .company_user .company 
        )

    def get_form (self ,form_class =None ):
        """
        Restricts the company_user queryset in the form to the authenticated user's company.
        ---
        Restringe el queryset de company_user del formulario a la empresa del usuario autenticado.
        """
        form =super ().get_form (form_class )
        form .fields ["company_user"].queryset =CompanyUser .objects .filter (
        company =self .request .user .company_user .company 
        ).select_related ("user")
        form .fields ["company_user"].required =False 
        return form 

    def get_success_url (self ):
        """
        Redirects to the contact list after a successful update.
        ---
        Redirige a la lista de contactos tras una actualización correcta.
        """
        from django .contrib import messages as django_messages 
        django_messages .success (
        self .request ,
        f"Contacto '{self.object.name}' actualizado correctamente."
        )
        return "/panel/contacts/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="contacts"
        context ["action"]="Editar"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()



class ContactDeleteView (AdminRoleRequiredMixin ,View ):
    """
    Allows an ADMIN to delete a Contact belonging to their company.
    Scoped to the authenticated user's company — cross-company deletions
    are rejected silently via get_object_or_404.
    Renders a confirmation page on GET; deletes on POST.
    ---
    Permite a un ADMIN eliminar un Contact de su empresa.
    Acotado a la empresa del usuario autenticado — los intentos entre
    empresas son rechazados silenciosamente via get_object_or_404.
    Renderiza una página de confirmación en GET; elimina en POST.
    """

    template_name ="panel/contacts/confirm_delete.html"

    def _get_contact (self ,request ,pk ):
        """
        Returns the Contact scoped to the authenticated user's company
        or raises Http404.
        ---
        Retorna el Contact acotado a la empresa del usuario autenticado
        o lanza Http404.
        """
        from django .shortcuts import get_object_or_404 
        return get_object_or_404 (
        Contact ,
        pk =pk ,
        company =request .user .company_user .company ,
        )

    def _get_own_presence (self ,request ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,pk ,*args ,**kwargs ):
        """
        Renders the confirmation page for the given Contact.
        ---
        Renderiza la página de confirmación para el Contact indicado.
        """
        contact =self ._get_contact (request ,pk )
        context ={
        "company":request .user .company_user .company ,
        "company_user":request .user .company_user ,
        "own_presence":self ._get_own_presence (request ),
        "active_nav":"contacts",
        "contact":contact ,
        }
        return render (request ,self .template_name ,context )

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Deletes the Contact and redirects to the contact list.
        ---
        Elimina el Contact y redirige a la lista de contactos.
        """
        contact =self ._get_contact (request ,pk )
        name =contact .name 
        contact .delete ()
        logger .info (
        "# [CONTACT DELETE] Contact '%s' eliminado por %s.",
        name ,
        request .user .username ,
        )
        django_messages .success (
        request ,
        f"Contacto '{name}' eliminado correctamente.",
        )
        return redirect ("/panel/contacts/")


class CallFlowListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all CallFlow records belonging to the authenticated user's company.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros CallFlow pertenecientes a la empresa del usuario autenticado.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =CallFlow 
    template_name ="panel/callflows/list.html"
    context_object_name ="call_flows"

    def get_queryset (self ):
        """
        Returns CallFlow records scoped to the authenticated user's company.
        ---
        Retorna los registros CallFlow acotados a la empresa del usuario autenticado.
        """
        return CallFlow .objects .filter (
        company =self .request .user .company_user .company 
        ).order_by ("name")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="callflows"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class CallFlowCreateView (AdminRoleRequiredMixin ,CreateView ):
    """
    Allows an ADMIN to create a new CallFlow for their company.
    Automatically assigns the company from the authenticated user's CompanyUser.
    Restricts the notification_contact queryset to the company's own contacts.
    Updated 2026-04-16 (Step 37.C — Estrategia B): fallback_section queryset
    restricted to the company's own active Sections.
    ---
    Permite a un ADMIN crear un nuevo CallFlow para su empresa.
    Asigna automáticamente la empresa desde el CompanyUser del usuario autenticado.
    Restringe el queryset de notification_contact a los contactos de la empresa.
    Actualización 2026-04-16 (Paso 37.C — Estrategia B): queryset de fallback_section
    restringido a las Sections activas de la empresa.
    """

    model =CallFlow 
    form_class =CallFlowForm 
    template_name ="panel/callflows/form.html"

    def get_form (self ,form_class =None ):
        """
        Restricts notification_contact queryset to the authenticated user's company.
        Restricts fallback_section queryset to the company's own active Sections
        (Estrategia B — Step 37.C). Both fields are optional.
        ---
        Restringe el queryset de notification_contact a la empresa del usuario autenticado.
        Restringe el queryset de fallback_section a las Sections activas de la empresa
        (Estrategia B — Paso 37.C). Ambos campos son opcionales.
        """
        form =super ().get_form (form_class )
        company =self .request .user .company_user .company 


        form .fields ["notification_contact"].queryset =Contact .objects .filter (
        company =company 
        ).exclude (
        company_user__role__in =["WORKSHOP","DRIVER"]
        ).order_by ("name")
        form .fields ["notification_contact"].required =False 
        form .fields ["fallback_section"].queryset =Section .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("name")
        form .fields ["fallback_section"].required =False 
        return form 

    def form_valid (self ,form ):
        """
        Assigns the company before saving the new CallFlow.
        ---
        Asigna la empresa antes de guardar el nuevo CallFlow.
        """
        form .instance .company =self .request .user .company_user .company 
        return super ().form_valid (form )

    def get_success_url (self ):
        """
        Redirects to the callflow list after a successful creation.
        ---
        Redirige a la lista de flujos IVR tras una creación correcta.
        """
        from django .contrib import messages as django_messages 
        django_messages .success (
        self .request ,
        f"Flujo IVR '{self.object.name}' creado correctamente."
        )
        return "/panel/callflows/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="callflows"
        context ["action"]="Crear"
        context ["section_workers"]=[]
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class CallFlowUpdateView (AdminRoleRequiredMixin ,UpdateView ):
    """
    Allows an ADMIN to update an existing CallFlow belonging to their company.
    Prevents editing call flows from other companies.
    Updated 2026-04-16 (Step 37.C — Estrategia B): fallback_section queryset
    restricted to the company's own active Sections.
    ---
    Permite a un ADMIN actualizar un CallFlow existente de su empresa.
    Impide editar flujos IVR de otras empresas.
    Actualización 2026-04-16 (Paso 37.C — Estrategia B): queryset de fallback_section
    restringido a las Sections activas de la empresa.
    """

    model =CallFlow 
    form_class =CallFlowForm 
    template_name ="panel/callflows/form.html"

    def get_queryset (self ):
        """
        Restricts the queryset to CallFlow records of the authenticated user's company.
        ---
        Restringe el queryset a los registros CallFlow de la empresa del usuario autenticado.
        """
        return CallFlow .objects .filter (
        company =self .request .user .company_user .company 
        )

    def get_form (self ,form_class =None ):
        """
        Restricts notification_contact queryset to the authenticated user's company.
        Restricts fallback_section queryset to the company's own active Sections
        (Estrategia B — Step 37.C). Both fields are optional.
        ---
        Restringe el queryset de notification_contact a la empresa del usuario autenticado.
        Restringe el queryset de fallback_section a las Sections activas de la empresa
        (Estrategia B — Paso 37.C). Ambos campos son opcionales.
        """
        form =super ().get_form (form_class )
        company =self .request .user .company_user .company 


        form .fields ["notification_contact"].queryset =Contact .objects .filter (
        company =company 
        ).exclude (
        company_user__role__in =["WORKSHOP","DRIVER"]
        ).order_by ("name")
        form .fields ["notification_contact"].required =False 
        form .fields ["fallback_section"].queryset =Section .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("name")
        form .fields ["fallback_section"].required =False 
        return form 

    def form_valid (self ,form ):
        """
        Saves a backup snapshot of the current values before applying the new ones.
        Backup fields store the state BEFORE this save so the ADMIN can restore.

        Uses a direct DB values() query to retrieve the pre-save state, bypassing
        Django's UpdateView object cache which already holds the POST values at this
        point in the request cycle and would cause the backup to store the new value
        instead of the old one.

        ---

        Guarda un snapshot de backup de los valores actuales antes de aplicar los nuevos.
        Los campos backup almacenan el estado ANTERIOR a este guardado para que el ADMIN
        pueda restaurar.

        Usa una query values() directa a BD para obtener el estado pre-guardado, evitando
        la caché de objeto de UpdateView que en este punto del ciclo de petición ya contiene
        los valores del POST y causaría que el backup almacene el valor nuevo en lugar del antiguo.
        """


        pre_save =CallFlow .objects .filter (pk =form .instance .pk ).values (
        "name",
        "system_instruction",
        "initial_greeting",
        "notification_contact_id",
        ).first ()

        if pre_save :
            form .instance .backup_name =pre_save ["name"]
            form .instance .backup_system_instruction =pre_save ["system_instruction"]
            form .instance .backup_initial_greeting =pre_save ["initial_greeting"]
            form .instance .backup_notification_contact_id =pre_save ["notification_contact_id"]

        return super ().form_valid (form )

    def get_success_url (self ):
        """
        Redirects to the callflow list after a successful update.
        ---
        Redirige a la lista de flujos IVR tras una actualización correcta.
        """
        from django .contrib import messages as django_messages 
        django_messages .success (
        self .request ,
        f"Flujo IVR '{self.object.name}' actualizado correctamente. "
        "Puedes restaurar la versión anterior desde el formulario de edición."
        )
        return "/panel/callflows/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence, action flag and has_backup to context.
        ---
        Añade company, company_user, own_presence, flag de acción y has_backup al contexto.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="callflows"
        context ["action"]="Editar"


        obj =self .get_object ()
        context ["has_backup"]=bool (
        obj .backup_name 
        or obj .backup_system_instruction 
        or obj .backup_initial_greeting 
        )
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class PhoneNumberListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all PhoneNumber records belonging to the authenticated user's company.
    Read-only view: Twilio number assignment is managed by the superuser.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros PhoneNumber pertenecientes a la empresa del usuario autenticado.
    Vista de solo lectura: la asignación de números Twilio la gestiona el superusuario.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =PhoneNumber 
    template_name ="panel/phonenumbers/list.html"
    context_object_name ="phone_numbers"

    def get_queryset (self ):
        """
        Returns PhoneNumber records scoped to the authenticated user's company.
        ---
        Retorna los registros PhoneNumber acotados a la empresa del usuario autenticado.
        """
        return PhoneNumber .objects .filter (
        company =self .request .user .company_user .company 
        ).select_related ("call_flow").order_by ("number")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="phonenumbers"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class CorporateVoiceProfileUpdateView (AdminRoleRequiredMixin ,View ):
    """
    Allows an ADMIN to view and update the CorporateVoiceProfile of their company.
    If no profile exists yet, it is created on first POST submission.
    Uses View instead of UpdateView to handle the get-or-create pattern cleanly.
    ---
    Permite a un ADMIN ver y actualizar el CorporateVoiceProfile de su empresa.
    Si no existe perfil todavía, se crea en el primer envío POST.
    Usa View en lugar de UpdateView para gestionar el patrón get-or-create limpiamente.
    """

    template_name ="panel/voiceprofile/detail.html"

    def _get_profile_and_context (self ,request ):
        """
        Retrieves or initialises the CorporateVoiceProfile for the company.
        Returns a dict with company, company_user, own_presence and profile.
        ---
        Obtiene o inicializa el CorporateVoiceProfile de la empresa.
        Retorna un dict con company, company_user, own_presence y profile.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 

        company_user =request .user .company_user 
        company =company_user .company 

        profile ,_ =CorporateVoiceProfile .objects .get_or_create (
        company =company ,
        defaults ={
        "tone_guidelines":"",
        "sample_responses":[],
        "forbidden_phrases":[],
        "is_active":True ,
        }
        )

        own_presence =PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

        return {
        "company":company ,
        "company_user":company_user ,
        "own_presence":own_presence ,
        "active_nav":"voiceprofile",
        "profile":profile ,
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the voice profile form pre-populated with the current profile data.
        ---
        Renderiza el formulario del perfil de voz prerellenado con los datos actuales.
        """
        from django .shortcuts import render 
        ctx =self ._get_profile_and_context (request )
        ctx ["form"]=CorporateVoiceProfileForm (instance =ctx ["profile"])
        return render (request ,self .template_name ,ctx )

    def post (self ,request ,*args ,**kwargs ):
        """
        Updates the CorporateVoiceProfile with the submitted data.
        ---
        Actualiza el CorporateVoiceProfile con los datos enviados.
        """
        from django .shortcuts import render ,redirect 
        from django .contrib import messages as django_messages 

        ctx =self ._get_profile_and_context (request )
        form =CorporateVoiceProfileForm (request .POST ,instance =ctx ["profile"])

        if form .is_valid ():
            profile =ctx ["profile"]


            instance =form .save (commit =False )
            instance .backup_voice_name =profile .voice_name 
            instance .backup_tone_guidelines =profile .tone_guidelines 
            instance .backup_sample_responses =profile .sample_responses 
            instance .backup_forbidden_phrases =profile .forbidden_phrases 
            instance .save ()
            django_messages .success (
            request ,
            "Perfil de voz corporativa actualizado correctamente. "
            "Puedes restaurar la versión anterior desde este mismo formulario."
            )
            return redirect ("panel:voiceprofile_detail")

        ctx ["form"]=form 
        return render (request ,self .template_name ,ctx )


class CallFlowRestoreView (AdminRoleRequiredMixin ,View ):
    """
    Restores a CallFlow to its previous backup snapshot with a single POST.
    Swaps active fields ↔ backup fields so the backup becomes the new backup
    (enabling a second restore to redo the change if needed).
    Restricted to CallFlow records belonging to the authenticated user's company.
    ---
    Restaura un CallFlow a su snapshot de backup anterior con un solo POST.
    Intercambia los campos activos ↔ backup para que el backup sea el nuevo backup
    (permitiendo una segunda restauración para rehacer el cambio si es necesario).
    Restringido a registros CallFlow de la empresa del usuario autenticado.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Performs the active ↔ backup field swap and redirects to the edit form.
        ---
        Realiza el intercambio activo ↔ backup y redirige al formulario de edición.
        """
        try :
            flow =CallFlow .objects .get (
            pk =pk ,
            company =request .user .company_user .company 
            )
        except CallFlow .DoesNotExist :
            django_messages .error (request ,"Flujo IVR no encontrado.")
            return redirect ("/panel/callflows/")

        if (
        not flow .backup_name 
        and not flow .backup_system_instruction 
        and not flow .backup_initial_greeting 
        ):
            django_messages .warning (
            request ,
            "No existe versión anterior para restaurar en este flujo IVR."
            )
            return redirect (f"/panel/callflows/{pk}/edit/")







        active_name =flow .name 
        active_system =flow .system_instruction 
        active_greeting =flow .initial_greeting 
        active_notification =flow .notification_contact_id 
        backup_name =flow .backup_name 
        backup_system =flow .backup_system_instruction 
        backup_greeting =flow .backup_initial_greeting 
        backup_notification =flow .backup_notification_contact_id 

        CallFlow .objects .filter (pk =flow .pk ).update (
        name =backup_name or active_name ,
        backup_name =active_name ,
        system_instruction =backup_system ,
        backup_system_instruction =active_system ,
        initial_greeting =backup_greeting ,
        backup_initial_greeting =active_greeting ,
        notification_contact_id =backup_notification ,
        backup_notification_contact_id =active_notification ,
        )
        django_messages .success (
        request ,
        f"Flujo IVR '{flow.name}' restaurado a la versión anterior correctamente."
        )
        return redirect (f"/panel/callflows/{pk}/edit/")


class VoiceProfileRestoreView (AdminRoleRequiredMixin ,View ):
    """
    Restores the CorporateVoiceProfile to its previous backup snapshot with a single POST.
    Swaps active fields ↔ backup fields to allow bidirectional restore.
    Restricted to the profile of the authenticated user's company.
    ---
    Restaura el CorporateVoiceProfile a su snapshot de backup anterior con un solo POST.
    Intercambia los campos activos ↔ backup para permitir restauración bidireccional.
    Restringido al perfil de la empresa del usuario autenticado.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Performs the active ↔ backup field swap and redirects to the voice profile form.
        ---
        Realiza el intercambio activo ↔ backup y redirige al formulario del perfil de voz.
        """
        try :
            profile =CorporateVoiceProfile .objects .get (
            company =request .user .company_user .company 
            )
        except CorporateVoiceProfile .DoesNotExist :
            django_messages .error (request ,"Perfil de voz no encontrado.")
            return redirect ("panel:voiceprofile_detail")

        if not profile .backup_tone_guidelines and not profile .backup_voice_name :
            django_messages .warning (
            request ,
            "No existe versión anterior para restaurar en el perfil de voz."
            )
            return redirect ("panel:voiceprofile_detail")



        (
        profile .voice_name ,profile .backup_voice_name ,
        )=(
        profile .backup_voice_name ,profile .voice_name ,
        )
        (
        profile .tone_guidelines ,profile .backup_tone_guidelines ,
        )=(
        profile .backup_tone_guidelines ,profile .tone_guidelines ,
        )
        (
        profile .sample_responses ,profile .backup_sample_responses ,
        )=(
        profile .backup_sample_responses ,profile .sample_responses ,
        )
        (
        profile .forbidden_phrases ,profile .backup_forbidden_phrases ,
        )=(
        profile .backup_forbidden_phrases ,profile .forbidden_phrases ,
        )
        profile .save (update_fields =[
        "voice_name",
        "backup_voice_name",
        "tone_guidelines",
        "backup_tone_guidelines",
        "sample_responses",
        "backup_sample_responses",
        "forbidden_phrases",
        "backup_forbidden_phrases",
        ])
        django_messages .success (
        request ,
        "Perfil de voz restaurado a la versión anterior correctamente."
        )
        return redirect ("panel:voiceprofile_detail")


class BlockedCallerListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all BlockedCaller records for the authenticated user's company.
    Shows active blocks (blocked_until > now) and expired history separately.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros BlockedCaller de la empresa del usuario autenticado.
    Muestra los bloqueos activos (blocked_until > now) e historial expirado por separado.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =BlockedCaller 
    template_name ="panel/blockedcallers/list.html"
    context_object_name ="blocked_callers"

    def get_queryset (self ):
        """
        Returns all BlockedCaller records for the company, ordered by most recent first.
        ---
        Retorna todos los registros BlockedCaller de la empresa, ordenados por más reciente primero.
        """
        return BlockedCaller .objects .filter (
        company =self .request .user .company_user .company 
        ).select_related ("blocked_by").order_by ("-blocked_at")

    def get_context_data (self ,**kwargs ):
        """
        Adds active/expired partition, company, company_user and own_presence to context.
        ---
        Añade la partición activos/expirados, company, company_user y own_presence al contexto.
        """
        context =super ().get_context_data (**kwargs )
        company_user =self .request .user .company_user 
        all_records =context ["blocked_callers"]
        context ["active_blocks"]=[b for b in all_records if b .is_active ]
        context ["expired_blocks"]=[b for b in all_records if not b .is_active ]
        context ["company"]=company_user .company 
        context ["company_user"]=company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="blockedcallers"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class BlockedCallerCreateView (AdminRoleRequiredMixin ,CreateView ):
    """
    Allows an ADMIN to manually block a phone number for their company.
    Automatically assigns the company and blocked_by from the authenticated user.
    ---
    Permite a un ADMIN bloquear manualmente un número de teléfono para su empresa.
    Asigna automáticamente la empresa y blocked_by desde el usuario autenticado.
    """

    model =BlockedCaller 
    form_class =BlockedCallerForm 
    template_name ="panel/blockedcallers/form.html"

    def form_valid (self ,form ):
        """
        Assigns company and blocked_by before saving the new BlockedCaller.
        ---
        Asigna company y blocked_by antes de guardar el nuevo BlockedCaller.
        """
        from django .contrib import messages as django_messages 
        form .instance .company =self .request .user .company_user .company 
        form .instance .blocked_by =self .request .user 
        response =super ().form_valid (form )
        django_messages .success (
        self .request ,
        f"Número {self.object.phone_number} bloqueado hasta {self.object.blocked_until:%d/%m/%Y %H:%M}."
        )
        return response 

    def get_success_url (self ):
        """
        Redirects to the blocked callers list after a successful creation.
        ---
        Redirige a la lista de bloqueados tras una creación correcta.
        """
        return "/panel/blockedcallers/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="blockedcallers"
        context ["action"]="Bloquear número"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class BlockedCallerDeleteView (AdminRoleRequiredMixin ,DeleteView ):
    """
    Allows an ADMIN to manually unblock a phone number before its expiry.
    Restricted to BlockedCaller records belonging to the authenticated user's company.
    Uses DELETE method via a confirmation form in the template.
    ---
    Permite a un ADMIN desbloquear manualmente un número antes de su vencimiento.
    Restringido a registros BlockedCaller de la empresa del usuario autenticado.
    Usa el método DELETE mediante un formulario de confirmación en la plantilla.
    """

    model =BlockedCaller 
    template_name ="panel/blockedcallers/confirm_delete.html"

    def get_queryset (self ):
        """
        Restricts deletion to BlockedCaller records of the authenticated user's company.
        ---
        Restringe la eliminación a registros BlockedCaller de la empresa del usuario autenticado.
        """
        return BlockedCaller .objects .filter (
        company =self .request .user .company_user .company 
        )

    def get_success_url (self ):
        """
        Redirects to the blocked callers list after successful deletion.
        ---
        Redirige a la lista de bloqueados tras la eliminación correcta.
        """
        from django .contrib import messages as django_messages 
        django_messages .success (
        self .request ,
        f"Número {self.object.phone_number} desbloqueado correctamente."
        )
        return "/panel/blockedcallers/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="blockedcallers"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class PanelLoginView (LoginView ):
    """
    Login view for the panel application.
    Uses the custom PanelAuthenticationForm.
    Authenticated users with an active CompanyUser are redirected to the dashboard.
    Authenticated users without a CompanyUser (e.g. superusers) see the login form normally.
    ---
    Vista de login para la aplicación panel.
    Usa el PanelAuthenticationForm personalizado.
    Los usuarios autenticados con CompanyUser activo son redirigidos al dashboard.
    Los usuarios autenticados sin CompanyUser (p.ej. superusuarios) ven el formulario normalmente.
    """

    template_name ="panel/login.html"
    authentication_form =PanelAuthenticationForm 
    next_page ="/panel/"

    def dispatch (self ,request ,*args ,**kwargs ):
        """
        Redirect authenticated CompanyUser accounts directly to the dashboard.
        If the user carries a valid trusted-device cookie, authenticate them
        directly without showing the login form.
        Superusers and users without CompanyUser proceed to the login form normally.
        ---
        Redirige las cuentas CompanyUser autenticadas directamente al dashboard.
        Si el usuario porta una cookie de dispositivo de confianza válida, lo
        autentica directamente sin mostrar el formulario de login.
        Los superusuarios y usuarios sin CompanyUser acceden al formulario normalmente.
        """


        if request .user .is_authenticated :
            company_user =getattr (request .user ,"company_user",None )
            if company_user is not None and company_user .is_active :
                from django .shortcuts import redirect 
                return redirect (self .next_page )
        return super ().dispatch (request ,*args ,**kwargs )

    def get_context_data (self ,**kwargs ):
        """
        Injects trusted_device_user into the template context when a valid
        trusted-device cookie is present, so the template can render the
        quick-access mode.
        ---
        Inyecta trusted_device_user en el contexto del template cuando hay
        una cookie de dispositivo de confianza válida, para que el template
        pueda renderizar el modo de acceso rápido.
        """
        from django .core import signing 
        from django .contrib .auth import get_user_model 
        context =super ().get_context_data (**kwargs )
        _raw =self .request .COOKIES .get ("eb_trusted_device","")
        if _raw and not self .request .GET .get ("force_form"):
            try :
                _payload =signing .loads (_raw ,max_age =365 *24 *3600 )
                _user_id =_payload .get ("uid")
                _device_token =_payload .get ("tok")
                User =get_user_model ()
                _user =User .objects .select_related ("company_user").get (pk =_user_id )
                _cu =getattr (_user ,"company_user",None )
                if (
                _cu is not None 
                and _cu .is_active 
                and _cu .trusted_device_token is not None 
                and str (_cu .trusted_device_token )==_device_token 
                ):
                    context ["trusted_device_user"]=_user 
            except Exception as _exc :
                logger .debug ("# [TRUSTED DEVICE] Cookie inválida en login: %s",_exc )
        return context 

    def form_valid (self ,form ):
        """
        After a successful login, redirect to trust_device if the device
        is not yet trusted. Otherwise redirect to the dashboard.
        ---
        Tras un login exitoso, redirige a trust_device si el dispositivo
        no es de confianza todavía. En caso contrario redirige al dashboard.
        """
        response =super ().form_valid (form )
        if not self .request .COOKIES .get ("eb_trusted_device",""):
            from django .shortcuts import redirect 
            return redirect ("panel:trust_device")
        return response 


@method_decorator (csrf_exempt ,name ="dispatch")
class TrustDeviceQuickLoginView (View ):
    """
    Handles the quick-login POST from the login page when a trusted-device
    cookie is present. Authenticates the cookie owner without a password
    and redirects to the dashboard.
    Security is provided by the signed cookie (django.core.signing), not CSRF.
    ---
    Gestiona el POST de acceso rápido desde la página de login cuando hay
    cookie de dispositivo de confianza. Autentica al propietario de la cookie
    sin contraseña y redirige al dashboard.
    La seguridad la proporciona la cookie firmada (django.core.signing), no el CSRF.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the trusted-device cookie and authenticates the user.
        ---
        Valida la cookie de dispositivo de confianza y autentica al usuario.
        """
        from django .core import signing 
        from django .contrib .auth import get_user_model ,login as auth_login 

        _raw =request .COOKIES .get ("eb_trusted_device","")
        if not _raw :
            return redirect ("panel:login")
        try :
            _payload =signing .loads (_raw ,max_age =365 *24 *3600 )
            _user_id =_payload .get ("uid")
            _device_token =_payload .get ("tok")
            User =get_user_model ()
            _user =User .objects .select_related ("company_user").get (pk =_user_id )
            _cu =getattr (_user ,"company_user",None )
            if (
            _cu is not None 
            and _cu .is_active 
            and not _cu .must_change_password 
            and _cu .trusted_device_token is not None 
            and str (_cu .trusted_device_token )==_device_token 
            ):
                auth_login (
                request ,
                _user ,
                backend ="django.contrib.auth.backends.ModelBackend",
                )
                logger .info (
                "# [TRUSTED DEVICE] Acceso rápido concedido a usuario pk=%s.",
                _user .pk ,
                )
                return redirect ("/panel/")
        except Exception as _exc :
            logger .warning ("# [TRUSTED DEVICE] Cookie inválida en quick-login: %s",_exc )
        return redirect ("panel:login")


class TrustDeviceView (View ):
    """
    Shown after a successful login on an untrusted device.
    GET:  Renders the trust-device question page.
    POST: If 'trust=yes', generates a UUID4 token, persists it in
          CompanyUser.trusted_device_token and emits the signed HttpOnly
          cookie 'eb_trusted_device' (max_age=365 days).
          If 'trust=no', redirects to dashboard without emitting the cookie.
    ---
    Se muestra tras un login exitoso en un dispositivo no conocido.
    GET:  Renderiza la página de pregunta de confianza de dispositivo.
    POST: Si 'trust=yes', genera un UUID4, lo persiste en
          CompanyUser.trusted_device_token y emite la cookie HttpOnly firmada
          'eb_trusted_device' (max_age=365 días).
          Si 'trust=no', redirige al dashboard sin emitir la cookie.
    """

    template_name ="panel/trust_device.html"

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the trust-device question page.
        ---
        Renderiza la página de pregunta de confianza de dispositivo.
        """
        if not request .user .is_authenticated :
            return redirect ("panel:login")
        cu =request .user .company_user 
        return render (request ,self .template_name ,{
        "company":cu .company ,
        "company_user":cu ,
        "active_nav":"",
        })

    def post (self ,request ,*args ,**kwargs ):
        """
        Processes the trust-device decision.
        'trust=yes' -> emit cookie and redirect to dashboard.
        'trust=no'  -> redirect to dashboard without cookie.
        ---
        Procesa la decisión de confianza de dispositivo.
        'trust=yes' -> emite cookie y redirige al dashboard.
        'trust=no'  -> redirige al dashboard sin cookie.
        """
        import uuid as _uuid 
        from django .core import signing as _signing 

        response =redirect ("/panel/")
        if request .POST .get ("trust")=="yes":
            cu =request .user .company_user 
            _token =_uuid .uuid4 ()
            cu .trusted_device_token =_token 
            cu .save (update_fields =["trusted_device_token"])
            _payload ={"uid":request .user .pk ,"tok":str (_token )}
            _signed =_signing .dumps (_payload )
            response .set_cookie (
            key ="eb_trusted_device",
            value =_signed ,
            max_age =365 *24 *3600 ,
            httponly =True ,
            secure =True ,
            samesite ="Lax",
            )
            logger .info (
            "# [TRUSTED DEVICE] Cookie emitida para usuario pk=%s.",
            request .user .pk ,
            )
        return response 


class TrustDeviceToggleView (CompanyUserRequiredMixin ,View ):
    """
    Toggles the trusted-device status of the current device from the profile page.
    POST 'action=trust'   -> generates token + emits cookie.
    POST 'action=revoke'  -> clears token + deletes cookie.
    ---
    Alterna el estado de dispositivo de confianza del dispositivo actual desde el perfil.
    POST 'action=trust'   -> genera token + emite cookie.
    POST 'action=revoke'  -> limpia token + borra cookie.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Processes trust/revoke action.
        ---
        Procesa la acción trust/revoke.
        """
        import uuid as _uuid 
        from django .core import signing as _signing 

        cu =request .user .company_user 
        action =request .POST .get ("action","")
        response =redirect ("panel:own_profile")

        if action =="trust":
            _token =_uuid .uuid4 ()
            cu .trusted_device_token =_token 
            cu .save (update_fields =["trusted_device_token"])
            _payload ={"uid":request .user .pk ,"tok":str (_token )}
            _signed =_signing .dumps (_payload )
            response .set_cookie (
            key ="eb_trusted_device",
            value =_signed ,
            max_age =365 *24 *3600 ,
            httponly =True ,
            secure =True ,
            samesite ="Lax",
            )
            django_messages .success (request ,"Este dispositivo ha sido marcado como de confianza.")
            logger .info (
            "# [TRUSTED DEVICE] Cookie emitida desde perfil para usuario pk=%s.",
            request .user .pk ,
            )

        elif action =="revoke":
            cu .trusted_device_token =None 
            cu .save (update_fields =["trusted_device_token"])
            response .delete_cookie ("eb_trusted_device")
            django_messages .success (request ,"La confianza de este dispositivo ha sido revocada.")
            logger .info (
            "# [TRUSTED DEVICE] Cookie revocada desde perfil para usuario pk=%s.",
            request .user .pk ,
            )

        return response 


class PresenceStatusUpdateView (CompanyUserRequiredMixin ,View ):
    """
    View for displaying and updating the authenticated user's own presence status.
    GET:  Renders the presence status form pre-populated with the current active state.
    POST: Closes the current active PresenceStatus and creates a new one.
    ---
    Vista para mostrar y actualizar el estado de presencia del usuario autenticado.
    GET:  Renderiza el formulario de presencia prerellenado con el estado activo actual.
    POST: Cierra el PresenceStatus activo actual y crea uno nuevo.
    """

    template_name ="panel/presence/status.html"

    def _get_active_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the given CompanyUser, or None.
        ---
        Retorna el PresenceStatus activo actual para el CompanyUser dado, o None.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the presence status page with the current active status and the form.
        ---
        Renderiza la página de estado de presencia con el estado activo actual y el formulario.
        """
        from django .shortcuts import render 
        company_user =request .user .company_user 
        active_presence =self ._get_active_presence (company_user )
        form =PresenceStatusForm (instance =active_presence )

        return render (request ,self .template_name ,{
        "company":company_user .company ,
        "company_user":company_user ,
        "own_presence":active_presence ,
        "active_nav":"presence",
        "form":form ,
        })

    def post (self ,request ,*args ,**kwargs ):
        """
        Closes the current active PresenceStatus and creates a new one with
        the submitted status and optional ends_at.
        ---
        Cierra el PresenceStatus activo actual y crea uno nuevo con el estado
        enviado y el ends_at opcional.
        """
        from django .shortcuts import render 
        from django .contrib import messages as django_messages 

        company_user =request .user .company_user 
        form =PresenceStatusForm (request .POST )

        if form .is_valid ():


            PresenceStatus .objects .filter (
            company_user =company_user ,
            starts_at__lte =now (),
            ).filter (
            Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
            ).update (ends_at =now ())



            new_status =form .save (commit =False )
            new_status .company_user =company_user 
            new_status .save ()

            django_messages .success (
            request ,
            f"Estado de presencia actualizado a: {new_status.get_status_display()}"
            )
            return redirect ("panel:presence_status")



        active_presence =self ._get_active_presence (company_user )
        return render (request ,self .template_name ,{
        "company":company_user .company ,
        "company_user":company_user ,
        "own_presence":active_presence ,
        "active_nav":"presence",
        "form":form ,
        })


class PanelLogoutView (LogoutView ):
    """
    Logout view for the panel application.
    Redirects to the panel login page after session termination.

    Django 5.x restricts LogoutView to POST-only by default (CSRF protection).
    The panel sidebar uses a plain <a> link (GET request) to trigger logout.
    The get() override handles this case by delegating to post(), which
    executes Django's standard session termination and redirects cleanly.
    ---
    Vista de logout para la aplicación panel.
    Redirige a la página de login del panel tras la terminación de sesión.

    Django 5.x restringe LogoutView a POST exclusivamente por defecto (protección CSRF).
    El sidebar del panel usa un enlace <a> simple (petición GET) para disparar el logout.
    El override de get() gestiona este caso delegando en post(), que ejecuta la
    terminación de sesión estándar de Django y redirige correctamente.
    """

    http_method_names =['get','post','head','options']
    next_page ="/panel/login/"

    def get (self ,request ,*args ,**kwargs ):
        """
        Handles GET logout requests from the panel sidebar link.
        Delegates to post() to execute standard Django session termination.
        ---
        Gestiona las peticiones GET de logout desde el enlace del sidebar del panel.
        Delega en post() para ejecutar la terminación de sesión estándar de Django.
        """
        return self .post (request ,*args ,**kwargs )


class PanelDashboardView (CompanyUserRequiredMixin ,TemplateView ):
    """
    Main dashboard view for authenticated CompanyUser accounts.
    Provides a summary of the company's active sections, total contacts,
    and the current presence status of the authenticated user.
    ---
    Vista principal del dashboard para cuentas CompanyUser autenticadas.
    Proporciona un resumen de las secciones activas de la empresa, el total
    de contactos y el estado de presencia actual del usuario autenticado.
    """

    template_name ="panel/dashboard.html"

    def dispatch (self ,request ,*args ,**kwargs ):
        """
        Redirect WORKSHOP users to the operator dashboard immediately.
        ADMIN and OPERATOR users proceed to the standard dashboard.
        ---
        Redirige a los usuarios WORKSHOP al dashboard de operario inmediatamente.
        Los usuarios ADMIN y OPERATOR continúan al dashboard estándar.
        """


        response =super ().dispatch (request ,*args ,**kwargs )
        if not request .user .is_authenticated :
            return response 

        company_user =getattr (request .user ,"company_user",None )
        if company_user and company_user .role ==CompanyUser .ROLE_WORKSHOP :
            return redirect ("/panel/operator/")

        return response 

    def get_context_data (self ,**kwargs ):
        """
        Build dashboard context with company summary and own presence status.
        ---
        Construye el contexto del dashboard con el resumen de empresa y el
        estado de presencia propio.
        """
        context =super ().get_context_data (**kwargs )



        company_user =self .request .user .company_user 
        company =company_user .company 



        active_sections =Section .objects .filter (
        company =company ,
        is_active =True ,
        )



        total_contacts =Contact .objects .filter (company =company ).count ()



        from django .utils .timezone import now 
        from django .db .models import Q 

        own_presence =PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

        context ["company"]=company 
        context ["company_user"]=company_user 
        context ["active_sections"]=active_sections 
        context ["active_sections_count"]=active_sections .count ()
        context ["total_contacts"]=total_contacts 
        context ["own_presence"]=own_presence 
        context ["active_nav"]="dashboard"

        return context 


class PanelPasswordChangeView (CompanyUserRequiredMixin ,View ):
    """
    Allows any authenticated CompanyUser to change their own password.
    When must_change_password=True this view is mandatory — the mixin
    blocks all other panel URLs until the password is updated.
    On success, must_change_password is cleared and the session auth hash
    is refreshed so the user stays logged in.
    ---
    Permite a cualquier CompanyUser autenticado cambiar su propia contraseña.
    Cuando must_change_password=True esta vista es obligatoria — el mixin
    bloquea todas las demás URLs del panel hasta que se actualice la contraseña.
    Al guardar, must_change_password se limpia y el hash de sesión se refresca
    para que el usuario permanezca autenticado.
    """

    template_name ="panel/password/change.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _build_form (self ,request ,data =None ):
        """
        Returns the appropriate password form depending on whether the change
        is forced (must_change_password=True) or voluntary.

        Forced change  → PanelSetPasswordForm: does not require old_password.
                         The newly created user does not know the system-assigned
                         initial password ('1234') and must not be asked for it.
        Voluntary change → PanelPasswordChangeForm: requires old_password as
                           an additional security gate.
        ---
        Retorna el formulario de contraseña adecuado según si el cambio es
        forzado (must_change_password=True) o voluntario.

        Cambio forzado   → PanelSetPasswordForm: no requiere old_password.
                           El usuario recién creado desconoce la contraseña
                           inicial asignada por el sistema ('1234') y no debe
                           ser preguntado por ella.
        Cambio voluntario → PanelPasswordChangeForm: requiere old_password como
                            barrera de seguridad adicional.
        """
        cu =request .user .company_user 
        if cu .must_change_password :


            return PanelSetPasswordForm (user =request .user ,data =data )


        return PanelPasswordChangeForm (user =request .user ,data =data )

    def _get_context (self ,request ,form =None ):
        """
        Builds template context including is_forced flag for UI messaging.
        When is_forced=True the template hides the old_password block, since
        PanelSetPasswordForm does not expose that field.
        ---
        Construye el contexto de plantilla incluyendo el flag is_forced para la UI.
        Cuando is_forced=True el template oculta el bloque old_password, ya que
        PanelSetPasswordForm no expone ese campo.
        """
        cu =request .user .company_user 
        return {
        "company":cu .company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"",
        "form":form or self ._build_form (request ),
        "is_forced":cu .must_change_password ,
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the password change form (forced or voluntary depending on context).
        ---
        Renderiza el formulario de cambio de contraseña (forzado o voluntario
        según el contexto).
        """
        return render (request ,self .template_name ,self ._get_context (request ))

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates and saves the new password using the appropriate form.
        On success clears must_change_password (if forced), updates the session
        auth hash so the user stays logged in, and redirects to the dashboard.
        ---
        Valida y guarda la nueva contraseña usando el formulario adecuado.
        En caso de éxito limpia must_change_password (si es forzado), actualiza
        el hash de sesión para que el usuario permanezca autenticado y redirige
        al dashboard.
        """
        form =self ._build_form (request ,data =request .POST )
        if form .is_valid ():
            form .save ()
            update_session_auth_hash (request ,form .user )
            cu =request .user .company_user 
            _was_forced =cu .must_change_password 
            if cu .must_change_password :
                cu .must_change_password =False 
                cu .save (update_fields =["must_change_password"])
            django_messages .success (request ,"Contraseña actualizada correctamente.")









            response =redirect ("/panel/")
            if _was_forced :
                import uuid as _uuid 
                from django .core import signing as _signing 
                _token =_uuid .uuid4 ()
                cu .trusted_device_token =_token 
                cu .save (update_fields =["trusted_device_token"])
                _payload ={"uid":request .user .pk ,"tok":str (_token )}
                _signed =_signing .dumps (_payload )
                _max_age =365 *24 *3600 
                response .set_cookie (
                key ="eb_trusted_device",
                value =_signed ,
                max_age =_max_age ,
                httponly =True ,
                secure =True ,
                samesite ="Lax",
                )
                logger .info (
                "# [TRUSTED DEVICE] Cookie emitida para usuario pk=%s — "
                "token=%s.",
                request .user .pk ,
                _token ,
                )
            return response 
        return render (request ,self .template_name ,self ._get_context (request ,form ))








class WhatsAppTemplateListView (AdminRoleRequiredMixin ,ListView ):
    """
    Displays a read-only list of active WhatsAppTemplate records scoped to
    the authenticated user's company. Accessible only to users with the
    ADMIN role. No create, update or delete actions are exposed — templates
    are managed exclusively via the Twilio Content Template Builder and
    seeded through the seed_whatsapp_templates management command.
    ---
    Muestra un listado de solo lectura de los registros WhatsAppTemplate
    activos acotados a la empresa del usuario autenticado. Solo accesible
    para usuarios con rol ADMIN. No se exponen acciones de creación, edición
    ni borrado — las plantillas se gestionan exclusivamente a través del
    Content Template Builder de Twilio y se registran mediante el comando
    de gestión seed_whatsapp_templates.
    """

    model =WhatsAppTemplate 
    template_name ="panel/whatsapp/template_list.html"
    context_object_name ="templates"

    def get_queryset (self ):
        """
        Returns active WhatsAppTemplate records scoped to the authenticated
        user's company, ordered alphabetically by name.
        ---
        Retorna los registros WhatsAppTemplate activos acotados a la empresa
        del usuario autenticado, ordenados alfabéticamente por nombre.
        """
        return WhatsAppTemplate .objects .filter (
        company =self .request .user .company_user .company ,
        is_active =True ,
        ).order_by ("name")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to the template context,
        following the same pattern as all other panel ListViews.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla,
        siguiendo el mismo patrón que el resto de ListViews del panel.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="whatsapp_templates"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .utils .timezone import now 
        from django .db .models import Q 
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()







class WhatsAppActiveSessionListView (AdminRoleRequiredMixin ,ListView ):
    """
    Displays a list of active WhatsAppSession records scoped to the
    authenticated user's company, ordered by most recently updated.
    A session is considered active when its status field equals 'active'.
    Accessible only to users with the ADMIN role.
    Each row shows: origin number, session start, last inbound message
    excerpt, and a link to the full session history.
    ---
    Muestra un listado de los registros WhatsAppSession activos acotados
    a la empresa del usuario autenticado, ordenados por actualizacion
    mas reciente. Una sesion se considera activa cuando su campo status
    es igual a 'active'. Solo accesible para usuarios con rol ADMIN.
    Cada fila muestra: numero de origen, inicio de sesion, extracto del
    ultimo mensaje entrante y enlace al historial completo.
    """

    model =WhatsAppSession 
    template_name ="panel/whatsapp/active_session_list.html"
    context_object_name ="active_sessions"

    def get_queryset (self ):
        """
        Returns active WhatsAppSession records scoped to the authenticated
        user's company, with last inbound message prefetched for excerpt display.
        ---
        Retorna los registros WhatsAppSession activos acotados a la empresa
        del usuario autenticado, con el ultimo mensaje entrante precargado
        para mostrar el extracto.
        """
        from whatsapp .models import WhatsAppMessage 
        return (
        WhatsAppSession .objects .filter (
        company =self .request .user .company_user .company ,
        is_active =True ,
        )
        .prefetch_related (
        Prefetch (
        "messages",
        queryset =WhatsAppMessage .objects .filter (
        direction ="IN"
        ).order_by ("-timestamp")[:1 ],
        to_attr ="last_inbound",
        )
        )
        .order_by ("-last_message_at")
        )

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to the template context.
        ---
        Anade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="whatsapp_sessions"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()







class DataCaptureSetListView (AdminRoleRequiredMixin ,ListView ):
    """
    Lists all DataCaptureSet records belonging to the authenticated user's company.
    Accessible only to users with the ADMIN role.
    ---
    Lista todos los registros DataCaptureSet pertenecientes a la empresa del usuario autenticado.
    Solo accesible para usuarios con rol ADMIN.
    """

    model =DataCaptureSet 
    template_name ="panel/datacapturesets/list.html"
    context_object_name ="capture_sets"

    def get_queryset (self ):
        """
        Returns DataCaptureSet records scoped to the authenticated user's company,
        ordered alphabetically by name.
        ---
        Retorna los registros DataCaptureSet acotados a la empresa del usuario autenticado,
        ordenados alfabéticamente por nombre.
        """
        return DataCaptureSet .objects .filter (
        company =self .request .user .company_user .company 
        ).order_by ("name")

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user and own_presence to template context.
        ---
        Añade company, company_user y own_presence al contexto de la plantilla.
        """
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="datacapturesets"
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class DataCaptureSetCreateView (AdminRoleRequiredMixin ,CreateView ):
    """
    Allows an ADMIN to create a new DataCaptureSet for their company.
    Automatically assigns the company from the authenticated user's CompanyUser.
    The 'fields' JSONField is populated by the JS dynamic row builder in the
    template, which serialises the field definitions to JSON before submit.
    ---
    Permite a un ADMIN crear un nuevo DataCaptureSet para su empresa.
    Asigna automáticamente la empresa desde el CompanyUser del usuario autenticado.
    El JSONField 'fields' es rellenado por el constructor de filas JS dinámico del
    template, que serializa las definiciones de campos a JSON antes del submit.
    """

    model =DataCaptureSet 
    form_class =DataCaptureSetForm 
    template_name ="panel/datacapturesets/form.html"

    def form_valid (self ,form ):
        """
        Assigns the company before saving the new DataCaptureSet.
        Deserialises the 'fields_json' hidden input into the model's JSONField.
        ---
        Asigna la empresa antes de guardar el nuevo DataCaptureSet.
        Deserializa el campo oculto 'fields_json' en el JSONField del modelo.
        """
        import json 
        form .instance .company =self .request .user .company_user .company 
        raw_json =self .request .POST .get ("fields_json","[]")
        try :
            form .instance .fields =json .loads (raw_json )
        except (ValueError ,TypeError ):
            form .instance .fields =[]
        return super ().form_valid (form )

    def get_success_url (self ):
        """
        Redirects to the DataCaptureSet list after a successful creation.
        ---
        Redirige a la lista de conjuntos de captura tras una creación correcta.
        """
        django_messages .success (
        self .request ,
        f"Conjunto de captura '{self.object.name}' creado correctamente."
        )
        return "/panel/datacapturesets/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        Adds existing_fields_json for pre-population on edit (empty list on create).
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        Añade existing_fields_json para prerellenar en edición (lista vacía en creación).
        """
        import json 
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="datacapturesets"
        context ["action"]="Crear"
        context ["section_workers"]=[]
        context ["existing_fields_json"]=json .dumps ([])
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class DataCaptureSetUpdateView (AdminRoleRequiredMixin ,UpdateView ):
    """
    Allows an ADMIN to update an existing DataCaptureSet belonging to their company.
    Prevents editing DataCaptureSets from other companies.
    The 'fields' JSONField is updated by the JS dynamic row builder in the template.
    ---
    Permite a un ADMIN actualizar un DataCaptureSet existente de su empresa.
    Impide editar conjuntos de captura de otras empresas.
    El JSONField 'fields' se actualiza por el constructor de filas JS dinámico del template.
    """

    model =DataCaptureSet 
    form_class =DataCaptureSetForm 
    template_name ="panel/datacapturesets/form.html"

    def get_queryset (self ):
        """
        Restricts the queryset to DataCaptureSet records of the authenticated user's company.
        ---
        Restringe el queryset a los registros DataCaptureSet de la empresa del usuario autenticado.
        """
        return DataCaptureSet .objects .filter (
        company =self .request .user .company_user .company 
        )

    def form_valid (self ,form ):
        """
        Deserialises the 'fields_json' hidden input into the model's JSONField before saving.
        ---
        Deserializa el campo oculto 'fields_json' en el JSONField del modelo antes de guardar.
        """
        import json 
        raw_json =self .request .POST .get ("fields_json","[]")
        try :
            form .instance .fields =json .loads (raw_json )
        except (ValueError ,TypeError ):
            form .instance .fields =[]
        return super ().form_valid (form )

    def get_success_url (self ):
        """
        Redirects to the DataCaptureSet list after a successful update.
        ---
        Redirige a la lista de conjuntos de captura tras una actualización correcta.
        """
        django_messages .success (
        self .request ,
        f"Conjunto de captura '{self.object.name}' actualizado correctamente."
        )
        return "/panel/datacapturesets/"

    def get_context_data (self ,**kwargs ):
        """
        Adds company, company_user, own_presence and action flag to template context.
        Serialises the existing fields JSONField for pre-population of the JS row builder.
        ---
        Añade company, company_user, own_presence y flag de acción al contexto de la plantilla.
        Serializa el JSONField fields existente para prerellenar el constructor de filas JS.
        """
        import json 
        context =super ().get_context_data (**kwargs )
        context ["company"]=self .request .user .company_user .company 
        context ["company_user"]=self .request .user .company_user 
        context ["own_presence"]=self ._get_own_presence ()
        context ["active_nav"]="datacapturesets"
        context ["action"]="Guardar"
        context ["section_workers"]=list (
        CompanyUser .objects .filter (
        company =self .request .user .company_user .company ,
        user__isnull =False ,
        contact_profile__section =self .object ,
        ).select_related ("user","contact").order_by ("user__username")
        )
        context ["existing_fields_json"]=json .dumps (self .object .fields or [])
        return context 

    def _get_own_presence (self ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        company_user =self .request .user .company_user 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()


class WorkOrderListView (SupervisorAccessMixin ,View ):
    """
    Lists WorkOrder records belonging to the authenticated user's company,
    split into four querysets for the tabbed UI introduced in Hito 8 / Bloque H:
      wo_queue    — PENDING or PROCESSING (polling tab).
      wo_error    — ERROR.
      wo_pending  — DONE, not yet reviewed (pending supervisor sign-off).
      wo_reviewed — DONE and reviewed.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).
    ---
    Lista los registros WorkOrder de la empresa del usuario autenticado,
    divididos en cuatro querysets para la UI de pestañas introducida en
    el Hito 8 / Bloque H:
      wo_queue    — PENDING o PROCESSING (pestaña con polling).
      wo_error    — ERROR.
      wo_pending  — DONE sin revisión (pendiente de validación del Supervisor).
      wo_reviewed — DONE y revisados.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    template_name ="panel/work_orders/list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .db .models import Q 
        from django .utils .timezone import now 
        from ivr_config .models import PresenceStatus 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Renders the tabbed work order list with four filtered querysets.
        The active tab defaults to "Pendiente revisión" when there are pending
        work orders; otherwise defaults to "En cola".
        Supports multi-level column sort via the sort_stack GET parameter
        (same mechanism as WorkOrderAdminHistoryView). Columns sortable from
        the PDF list: upload_date, pdf_display_name, uploaded_by_name.
        ---
        Renderiza la lista de partes con pestañas y cuatro querysets filtrados.
        La pestaña activa por defecto es "Pendiente revisión" si hay partes
        pendientes; en caso contrario, "En cola".
        Soporta ordenación multi-nivel via el parámetro GET sort_stack
        (mismo mecanismo que WorkOrderAdminHistoryView). Columnas ordenables
        desde la lista PDF: upload_date, pdf_display_name, uploaded_by_name.
        """
        company_user =request .user .company_user 
        company =company_user .company 









        _VALID_SORT_COLS =frozenset (
        ("upload_date","pdf_display_name","uploaded_by_name")
        )
        _VALID_SORT_DIRS =frozenset (("asc","desc"))



        _ORM_FIELD_MAP ={
        "upload_date":[("upload_date",)],
        "pdf_display_name":[("source_pdf__name",)],
        "uploaded_by_name":[
        ("uploaded_by__user__last_name",),
        ("uploaded_by__user__first_name",),
        ],
        }

        raw_stack =request .GET .get ("sort_stack","").strip ()

        sort_stack =[]
        if raw_stack :
            for token in raw_stack .split (","):
                token =token .strip ()
                if ":"not in token :
                    continue 
                col ,_ ,direction =token .partition (":")
                col =col .strip ()
                direction =direction .strip ()
                if col in _VALID_SORT_COLS and direction in _VALID_SORT_DIRS :
                    if not any (s [0 ]==col for s in sort_stack ):
                        sort_stack .append ((col ,direction ))
                        if len (sort_stack )>=3 :
                            break 

        if not sort_stack :
            sort_stack =[("upload_date","desc")]

        sort_primary_col =sort_stack [0 ][0 ]
        sort_primary_dir =sort_stack [0 ][1 ]
        sort_stack_str =",".join (f"{col}:{direction}"for col ,direction in sort_stack )

        def _build_order_by (stack ):
            """
            Builds an order_by tuple from the sort stack, applying the
            correct direction prefix ("-") for each column and its ORM fields.
            Secondary and tertiary columns act as tiebreakers in order.
            ---
            Construye una tupla order_by desde la pila de ordenación, aplicando
            el prefijo de dirección ("-") correcto a cada columna y sus campos ORM.
            Las columnas secundaria y terciaria actúan como desempate en orden.
            """
            order_fields =[]
            for col ,direction in stack :
                prefix ="-"if direction =="desc"else ""
                for field_tuple in _ORM_FIELD_MAP .get (col ,[(col ,)]):
                    for field in field_tuple :
                        order_fields .append (f"{prefix}{field}")
            return order_fields 

        orm_order =_build_order_by (sort_stack )






        wo_queue =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status__in =[
        WorkOrder .Status .PENDING ,
        WorkOrder .Status .PROCESSING ,
        ],
        )
        .order_by ("-upload_date")
        )
        wo_error =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .ERROR ,
        )
        .order_by ("-upload_date")
        )
        wo_pending =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .DONE ,
        reviewed =False ,
        )
        .select_related ("uploaded_by__user")
        .order_by (*orm_order )
        )
        wo_reviewed =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        )
        .select_related ("reviewed_by__user","uploaded_by__user")
        .order_by (*orm_order )
        )





        default_tab ="pending"if wo_pending .exists ()else "queue"

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        "wo_queue":wo_queue ,
        "wo_error":wo_error ,
        "wo_pending":wo_pending ,
        "wo_reviewed":wo_reviewed ,
        "default_tab":default_tab ,
        "sort_stack_str":sort_stack_str ,
        "sort_primary_col":sort_primary_col ,
        "sort_primary_dir":sort_primary_dir ,
        })

    def post (self ,request ):
        """
        Handles bulk actions submitted from the PDF pipeline list tabs.
        Supported bulk_op values:
          mark_reviewed   — sets reviewed=True on selected PDF_UPLOAD WorkOrders.
          unmark_reviewed — sets reviewed=False on selected PDF_UPLOAD WorkOrders.
          delete          — deletes selected PDF_UPLOAD WorkOrders (CASCADE).
        All operations are scoped to the authenticated user's company.

        ---

        Gestiona acciones en lote enviadas desde las pestañas de la lista PDF.
        Valores de bulk_op soportados:
          mark_reviewed   — establece reviewed=True en los WorkOrders seleccionados.
          unmark_reviewed — establece reviewed=False en los WorkOrders seleccionados.
          delete          — elimina los WorkOrders seleccionados (CASCADE).
        Todas las operaciones están acotadas a la empresa del usuario autenticado.
        """
        from django .utils .timezone import now as _now 
        from django .contrib import messages as django_messages 

        company_user =request .user .company_user 
        company =company_user .company 
        active_tab =request .POST .get ("active_tab","pending")
        bulk_op =request .POST .get ("bulk_op","")
        raw_pks =request .POST .getlist ("pks")



        try :
            pk_list =[int (p )for p in raw_pks if str (p ).strip ().isdigit ()]
        except (ValueError ,AttributeError ):
            pk_list =[]

        if not pk_list :
            django_messages .warning (request ,"No se ha seleccionado ningún parte.")
            return redirect (f"{request.path}?tab={active_tab}")

        qs =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        pk__in =pk_list ,
        )
        )

        if bulk_op =="mark_reviewed":

            updated =qs .filter (status =WorkOrder .Status .DONE ).update (
            reviewed =True ,
            reviewed_by =company_user ,
            reviewed_at =_now (),
            )
            django_messages .success (
            request ,
            f"{updated} parte{'s' if updated != 1 else ''} "
            f"marcado{'s' if updated != 1 else ''} como revisado{'s' if updated != 1 else ''}."
            )

        elif bulk_op =="unmark_reviewed":

            updated =qs .filter (status =WorkOrder .Status .DONE ,reviewed =True ).update (
            reviewed =False ,
            reviewed_by =None ,
            reviewed_at =None ,
            )
            django_messages .success (
            request ,
            f"Revisión desmarcada en {updated} parte{'s' if updated != 1 else ''}."
            )

        elif bulk_op =="delete":

            count =qs .count ()
            qs .delete ()
            django_messages .success (
            request ,
            f"{count} parte{'s' if count != 1 else ''} "
            f"eliminado{'s' if count != 1 else ''} correctamente."
            )

        else :
            django_messages .warning (
            request ,
            f"Operación en lote desconocida: '{bulk_op}'."
            )

        return redirect (f"{request.path}?tab={active_tab}")








class DigitalWorkOrderListView (SupervisorAccessMixin ,View ):
    """
    Lists WorkOrder records with source IN (DIGITAL, GENERATED) for the
    authenticated user's company, split into three querysets for the
    tabbed UI:
      wo_pending  — status=DONE, reviewed=False (pending supervisor sign-off).
      wo_reviewed — status=DONE, reviewed=True  (Excel download available).
      wo_error    — status=ERROR.

    Supports optional GET filters:
      operator_pk — restrict to a specific CompanyUser (WORKSHOP role).
      period_pk   — restrict to work orders whose entries fall within a
                    specific WorkPeriod date range.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).
    ---
    Lista los registros WorkOrder con source IN (DIGITAL, GENERATED) de la
    empresa del usuario autenticado, divididos en tres querysets para la
    UI de pestañas:
      wo_pending  — status=DONE, reviewed=False (pendiente de validación).
      wo_reviewed — status=DONE, reviewed=True  (descarga Excel disponible).
      wo_error    — status=ERROR.

    Soporta filtros GET opcionales:
      operator_pk — restringir a un CompanyUser concreto (rol WORKSHOP).
      period_pk   — restringir a partes cuyas entradas caen dentro del rango
                    de fechas de un WorkPeriod concreto.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    template_name ="panel/work_orders/digital_list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Builds the three querysets (pending / reviewed / error) scoped to
        DIGITAL and GENERATED sources and renders digital_list.html.
        Optional GET filters operator_pk and period_pk are applied when present.
        The default active tab is "pending" if wo_pending has results;
        otherwise "reviewed".

        ---

        Construye los tres querysets (pendiente / revisados / error) acotados
        a orígenes DIGITAL y GENERATED y renderiza digital_list.html.
        Los filtros GET opcionales operator_pk y period_pk se aplican cuando
        están presentes. La pestaña activa por defecto es "pending" si
        wo_pending tiene resultados; en caso contrario "reviewed".
        """
        from ivr_config .models import WorkPeriod 

        cu =request .user .company_user 
        company =cu .company 





        try :
            operator_pk =int (request .GET .get ("operator_pk",""))
        except (ValueError ,TypeError ):
            operator_pk =None 

        try :
            period_pk =int (request .GET .get ("period_pk",""))
        except (ValueError ,TypeError ):
            period_pk =None 



        period_start =None 
        period_end =None 
        if period_pk :
            try :
                wp =WorkPeriod .objects .get (pk =period_pk ,company_user__company =company )
                period_start =wp .start_date 
                period_end =wp .end_date 
            except WorkPeriod .DoesNotExist :
                period_pk =None 



        _digital_sources =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ]
        _base =WorkOrder .objects .filter (
        company =company ,
        source__in =_digital_sources ,
        )



        if operator_pk :
            _base =_base .filter (uploaded_by__pk =operator_pk )



        if period_start :
            _base =_base .filter (entries__work_date__gte =period_start ).distinct ()
        if period_end :
            _base =_base .filter (entries__work_date__lte =period_end ).distinct ()



        wo_pending =_base .filter (status =WorkOrder .Status .DONE ,reviewed =False ).order_by ("-upload_date")
        wo_reviewed =_base .filter (status =WorkOrder .Status .DONE ,reviewed =True ).select_related ("uploaded_by__user").order_by ("-upload_date")
        wo_error =_base .filter (status =WorkOrder .Status .ERROR ).order_by ("-upload_date")



        default_tab ="pending"if wo_pending .exists ()else "reviewed"



        operators =(
        CompanyUser .objects 
        .filter (company =company ,is_active =True ,role =CompanyUser .ROLE_WORKSHOP )
        .select_related ("user")
        .order_by ("user__last_name","user__first_name")
        )



        periods =(
        WorkPeriod .objects 
        .filter (company_user__company =company ,end_date__isnull =False )
        .order_by ("-end_date")
        .distinct ()
        )

        context ={
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"digital_list",
        "wo_pending":wo_pending ,
        "wo_reviewed":wo_reviewed ,
        "wo_error":wo_error ,
        "default_tab":default_tab ,
        "operators":operators ,
        "periods":periods ,
        "operator_pk":operator_pk ,
        "period_pk":period_pk ,
        }
        return render (request ,self .template_name ,context )


class WorkOrderUploadView (SupervisorAccessMixin ,View ):
    """
    Handles PDF upload for work order processing.
    On POST: validates the file, runs a duplicate detection check before
    creating the WorkOrder, enqueues the Celery processing task immediately
    via process_work_order_pdf.delay_on_commit() and redirects to the list.

    Duplicate detection (Hito 8, Paso 5, Bloque E):
      Before creating a new WorkOrder the view extracts the worker name and
      work period from the uploaded PDF filename and checks for an existing
      WorkOrder for the same company with the same worker name and an
      overlapping period. If a duplicate is found and the POST does not
      include confirm_overwrite=1, the upload form is re-rendered with the
      duplicate_wo context variable so the template can show a warning modal.
      If the user confirms overwrite, the duplicate WorkOrder (and all its
      cascade-deleted children) is deleted before the new one is created.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).
    ---
    Gestiona la carga de PDF para el procesamiento de partes de trabajo.
    En POST: valida el archivo, ejecuta una comprobación de duplicado antes
    de crear el WorkOrder, encola la tarea Celery inmediatamente mediante
    process_work_order_pdf.delay_on_commit() y redirige a la lista.

    Detección de duplicado (Hito 8, Paso 5, Bloque E):
      Antes de crear un nuevo WorkOrder la vista extrae el nombre del operario
      y el periodo de trabajo del nombre del PDF cargado y comprueba si existe
      un WorkOrder de la misma empresa con el mismo operario y periodo solapado.
      Si se detecta un duplicado y el POST no incluye confirm_overwrite=1, el
      formulario de carga se vuelve a renderizar con la variable de contexto
      duplicate_wo para que la plantilla muestre un modal de advertencia.
      Si el usuario confirma la sobrescritura, el WorkOrder duplicado (y todos
      sus hijos eliminados en cascada) se elimina antes de crear el nuevo.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    template_name ="panel/work_orders/upload.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .db .models import Q 
        from django .utils .timezone import now 
        from ivr_config .models import PresenceStatus 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Renders the PDF upload form.
        ---
        Renderiza el formulario de carga de PDF.
        """
        company_user =request .user .company_user 
        return render (request ,self .template_name ,{
        "company":company_user .company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        })

    def post (self ,request ):
        """
        Processes the uploaded PDF: validates the file, computes its SHA-256
        hash, runs a two-level duplicate detection check, optionally deletes
        the existing duplicate WorkOrder on confirmed overwrite, creates a new
        WorkOrder (storing the hash) and enqueues the Celery processing task.

        Duplicate detection — two levels (Hito 8 / Bloque I):
          LEVEL 1 — Exact hash match:
            SHA-256 of the incoming file is compared against source_pdf_hash
            for all WorkOrders of the same company. A hash match means the
            exact same file has been uploaded before, regardless of filename.
          LEVEL 2 — Worker + period overlap (only if no hash match):
            The incoming PDF filename is used to derive the worker name and,
            when possible, the work period. Existing WorkOrderEntry records for
            the same company + worker_name + overlapping work_date range are
            queried. A match here means a different file covering the same
            data has already been processed.

          duplicate_reason carries "exact" or "content" so the template can
          show a differentiated warning message in the modal.

          On confirmed overwrite the duplicate WorkOrder (and all its cascade-
          deleted children) is removed before the new one is created.
        ---
        Procesa el PDF cargado: valida el archivo, calcula su hash SHA-256,
        ejecuta la detección de duplicado en dos niveles, elimina opcionalmente
        el WorkOrder duplicado si el usuario confirma la sobrescritura, crea
        un nuevo WorkOrder (guardando el hash) y encola la tarea Celery.

        Detección de duplicado — dos niveles (Hito 8 / Bloque I):
          NIVEL 1 — Hash exacto:
            El SHA-256 del fichero entrante se compara contra source_pdf_hash
            de todos los WorkOrders de la misma empresa. Un match de hash
            significa que se ha subido exactamente el mismo fichero antes,
            con independencia del nombre.
          NIVEL 2 — Operario + solapamiento de periodo (solo si no hay hash match):
            El nombre del PDF se usa para derivar el nombre del operario y,
            cuando es posible, el periodo de trabajo. Se consultan los registros
            WorkOrderEntry de la misma empresa + worker_name + rango work_date
            solapado. Un match aquí significa que un fichero diferente pero con
            los mismos datos ya ha sido procesado.

          duplicate_reason transporta "exact" o "content" para que la plantilla
          muestre un mensaje diferenciado en el modal.

          Al confirmar la sobrescritura el WorkOrder duplicado (y todos sus hijos
          eliminados en cascada) se elimina antes de crear el nuevo.
        """
        import hashlib 

        from django .contrib import messages as django_messages 
        from work_order_processor .services import _worker_name_from_pdf_path 
        from work_order_processor .tasks import _extract_period_from_pdf_name 
        from work_order_processor .models import WorkOrderEntry 

        company_user =request .user .company_user 
        company =company_user .company 
        pdf_file =request .FILES .get ("source_pdf")





        if not pdf_file :
            django_messages .error (request ,"Debes seleccionar un archivo PDF.")
            return render (request ,self .template_name ,{
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            })

        if not pdf_file .name .lower ().endswith (".pdf"):
            django_messages .error (request ,"El archivo debe tener extensión .pdf.")
            return render (request ,self .template_name ,{
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            })











        pdf_file .seek (0 )
        incoming_hash =hashlib .sha256 (pdf_file .read ()).hexdigest ()
        pdf_file .seek (0 )













        incoming_name =pdf_file .name 
        incoming_worker =_worker_name_from_pdf_path (incoming_name )
        date_from ,date_to =_extract_period_from_pdf_name (incoming_name )

        duplicate_wo =None 
        duplicate_reason =None 



        hash_duplicate =(
        WorkOrder .objects 
        .filter (company =company ,source_pdf_hash =incoming_hash )
        .exclude (source_pdf_hash ="")
        .first ()
        )
        if hash_duplicate :
            duplicate_wo =hash_duplicate 
            duplicate_reason ="exact"



        if not duplicate_wo and incoming_worker :
            entry_qs =WorkOrderEntry .objects .filter (
            work_order__company =company ,
            worker_name =incoming_worker ,
            ).select_related ("work_order")

            if date_from and date_to :




                entry_qs =entry_qs .filter (
                work_date__gte =date_from ,
                work_date__lte =date_to ,
                )



            existing_entry =entry_qs .first ()
            if existing_entry :
                duplicate_wo =existing_entry .work_order 
                duplicate_reason ="content"







        if not duplicate_wo and incoming_worker and date_from and date_to :
            conflicting_entries =(
            WorkOrderEntry .objects 
            .filter (
            work_order__company =company ,
            worker_name =incoming_worker ,
            work_date__gte =date_from ,
            work_date__lte =date_to ,
            )
            .exclude (work_order__source_pdf_hash =incoming_hash )
            .exclude (work_order__source_pdf_hash ="")
            .select_related ("work_order")
            .order_by ("work_date")
            )
            if conflicting_entries .exists ():


                seen_dates =set ()
                duplicate_dates =[]
                for entry in conflicting_entries :
                    if entry .work_date :
                        date_key =entry .work_date .strftime ("%d/%m/%y")
                        if date_key not in seen_dates :
                            seen_dates .add (date_key )
                            duplicate_dates .append (date_key )

                first_entry =conflicting_entries .first ()
                duplicate_wo =first_entry .work_order 
                duplicate_reason ="duplicate_entries"





        if duplicate_wo and not request .POST .get ("confirm_overwrite"):
            ctx ={
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            "duplicate_wo":duplicate_wo ,
            "duplicate_reason":duplicate_reason ,
            "pdf_file_name":incoming_name ,
            }


            if duplicate_reason =="duplicate_entries":
                ctx ["duplicate_dates"]=duplicate_dates 
            return render (request ,self .template_name ,ctx )



        if duplicate_wo and request .POST .get ("confirm_overwrite"):
            dup_pk =duplicate_wo .pk 
            duplicate_wo .delete ()
            django_messages .warning (
            request ,
            f"El parte duplicado #{dup_pk} y todos sus datos han sido "
            f"eliminados. Procesando el nuevo PDF."
            )
























        from django .db import transaction 

        with transaction .atomic ():









            race_duplicate =(
            WorkOrder .objects 
            .filter (company =company ,source_pdf_hash =incoming_hash )
            .exclude (source_pdf_hash ="")
            .select_for_update ()
            .first ()
            )
            if race_duplicate :
                django_messages .info (
                request ,
                f"El fichero ya había sido registrado como Parte "
                f"#{race_duplicate.pk}. No se ha creado un duplicado."
                )
                return redirect ("panel:work_order_list")

            work_order =WorkOrder .objects .create (
            company =company ,
            uploaded_by =company_user ,
            source_pdf =pdf_file ,
            source_pdf_hash =incoming_hash ,
            source_pdf_name =incoming_name ,
            )

        process_work_order_pdf .delay_on_commit (work_order .pk )

        django_messages .success (
        request ,
        f"PDF cargado correctamente (Parte #{work_order.pk}). "
        f"El procesamiento ha sido encolado y comenzará en instantes."
        )
        return redirect ("panel:work_order_list")


class WorkOrderEditView (SupervisorAccessMixin ,View ):
    """
    Displays and processes inline edits for all WorkOrderEntryLine records
    belonging to a given WorkOrder. Lines are grouped by their parent
    WorkOrderEntry (page / date) for readability.

    GET  — Renders the editable table grouped by page.
    POST — Two actions dispatched via hidden input `action`:
      "save_line"      : Saves a single WorkOrderEntryLine identified by `line_pk`.
                         Recomputes delta_hours from hc/hf and re-resolves
                         machine_asset from the updated machine_norm.
      "regenerate"     : Re-enqueues the Excel generation task for this WorkOrder
                         (status reset to PENDING) and redirects to the appropriate list.

    Access rules:
      WORKSHOP  — can only access their own DIGITAL/GENERATED parts. Access to
                   PDF parts or parts belonging to other operators is denied.
                   After save/regenerate, redirects to panel:operator_history.
      SUPERVISOR/ADMIN — full access to all parts of the company.
                   After save/regenerate, redirects to panel:digital_work_order_list
                   for DIGITAL/GENERATED parts or panel:work_order_list for PDF parts.
    Access is restricted to the authenticated company (multicompany guard).

    ---

    Muestra y procesa ediciones inline para todos los registros WorkOrderEntryLine
    de un WorkOrder dado. Las líneas se agrupan por su WorkOrderEntry padre
    (página / fecha) para facilitar la lectura.

    GET  — Renderiza la tabla editable agrupada por página.
    POST — Dos acciones despachadas mediante el input oculto `action`:
      "save_line"      : Guarda un único WorkOrderEntryLine identificado por `line_pk`.
                         Recalcula delta_hours desde hc/hf y re-resuelve machine_asset
                         desde el machine_norm actualizado.
      "regenerate"     : Re-encola la tarea de generación de Excel para este WorkOrder
                         (estado reseteado a PENDING) y redirige a la lista apropiada.

    Reglas de acceso:
      WORKSHOP  — solo puede acceder a sus propios partes DIGITAL/GENERATED. El acceso
                   a partes PDF o de otros operarios queda denegado.
                   Tras guardar/regenerar redirige a panel:operator_history.
      SUPERVISOR/ADMIN — acceso completo a todos los partes de la empresa.
                   Tras guardar/regenerar redirige a panel:digital_work_order_list
                   para partes DIGITAL/GENERATED o a panel:work_order_list para PDF.
    El acceso está restringido a la empresa autenticada (guardia multiempresa).
    """

    template_name ="panel/work_orders/edit.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _get_work_order (self ,pk ,company ):
        """
        Retrieves a WorkOrder scoped to the given company.
        Raises WorkOrder.DoesNotExist if not found or belongs to another company.
        ---
        Recupera un WorkOrder acotado a la empresa dada.
        Lanza WorkOrder.DoesNotExist si no se encuentra o pertenece a otra empresa.
        """
        return WorkOrder .objects .get (pk =pk ,company =company )

    def _build_groups (self ,work_order ):
        """
        Returns a list of dicts grouping WorkOrderEntryLine records by their
        parent WorkOrderEntry, ordered by page number.
        Each dict contains:
          - entry  : WorkOrderEntry instance.
          - lines  : list of WorkOrderEntryLine instances for that entry.

        ---

        Devuelve una lista de dicts que agrupan los registros WorkOrderEntryLine
        por su WorkOrderEntry padre, ordenados por número de página.
        Cada dict contiene:
          - entry  : instancia de WorkOrderEntry.
          - lines  : lista de instancias WorkOrderEntryLine de esa entrada.
        """
        entries =(
        WorkOrderEntry .objects 
        .filter (work_order =work_order )
        .prefetch_related ("lines__machine_asset")
        .order_by ("page_number")
        )
        groups =[]
        for entry in entries :
            lines =list (entry .lines .order_by ("line_number"))








            day_total_raw =sum (
            (l .delta_hours for l in lines if l .delta_hours is not None ),
            0 ,
            )
            day_total =round (day_total_raw ,2 )if any (
            l .delta_hours is not None for l in lines 
            )else None 














            if day_total is None :
                day_css =""
            elif day_total <8 :
                day_css ="day-total-short"
            elif day_total <=12 :
                day_css ="day-total-normal"
            elif day_total <=16 :
                day_css ="day-total-warning"
            else :
                day_css ="day-total-danger"

            groups .append ({
            "entry":entry ,
            "lines":lines ,
            "day_total":day_total ,
            "day_css":day_css ,
            })
        return groups 

    def get (self ,request ,pk ):
        """
        Renders the inline edit table for the given WorkOrder.
        WORKSHOP role: access restricted to own DIGITAL/GENERATED parts only.
        SUPERVISOR/ADMIN: full access to all parts of the company.
        Passes is_digital to the template for conditional title rendering.
        ---
        Renderiza la tabla de edición inline para el WorkOrder dado.
        Rol WORKSHOP: acceso restringido a sus propios partes DIGITAL/GENERATED.
        SUPERVISOR/ADMIN: acceso completo a todos los partes de la empresa.
        Pasa is_digital al template para renderizado condicional del título.
        """
        from work_order_processor .models import WorkOrderEntry 
        from django .urls import reverse as _reverse_get 
        company_user =request .user .company_user 
        company =company_user .company 
        _is_workshop =company_user .role =="WORKSHOP"

        try :
            work_order =self ._get_work_order (pk ,company )
        except WorkOrder .DoesNotExist :
            django_messages .error (request ,"Parte de trabajo no encontrado.")
            if _is_workshop :
                return redirect (_reverse_get ("panel:operator_history"))
            return redirect (_reverse_get ("panel:work_order_list"))







        _is_digital =work_order .source in (
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        )
        if _is_workshop :
            if not _is_digital or work_order .uploaded_by !=company_user :
                django_messages .error (
                request ,
                "No tienes permiso para editar este parte."
                )
                return redirect (_reverse_get ("panel:operator_history"))

        groups =self ._build_groups (work_order )








        if _is_workshop :
            back_url =_reverse_get ("panel:operator_history")
        else :
            from_param =request .GET .get ("from","")
            if from_param =="digital":
                back_url =_reverse_get ("panel:digital_work_order_list")
            elif from_param =="taller":
                back_url =_reverse_get ("panel:work_order_admin_history")+"?tab=pending"
            elif _is_digital :
                back_url =_reverse_get ("panel:digital_work_order_list")
            else :
                back_url =_reverse_get ("panel:work_order_list")








        from work_order_processor .models import WorkdayGap 
        if _is_digital :
            workday_gaps =list (
            WorkdayGap .objects 
            .filter (work_order =work_order )
            .select_related ("absence_category")
            .order_by ("gap_start")
            )
        else :
            workday_gaps =[]

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        "work_order":work_order ,
        "groups":groups ,
        "back_url":back_url ,
        "workday_gaps":workday_gaps ,
        "is_digital":_is_digital ,
        })

    def post (self ,request ,pk ):
        """
        Dispatches POST actions: save_line or regenerate.
        WORKSHOP role: access restricted to own DIGITAL/GENERATED parts.
        Redirects after action respect the role:
          WORKSHOP         — panel:operator_history.
          SUPERVISOR/ADMIN — panel:digital_work_order_list (DIGITAL/GENERATED)
                              or panel:work_order_list (PDF).
        ---
        Despacha las acciones POST: save_line o regenerate.
        Rol WORKSHOP: acceso restringido a sus propios partes DIGITAL/GENERATED.
        Las redirecciones tras la acción respetan el rol:
          WORKSHOP         — panel:operator_history.
          SUPERVISOR/ADMIN — panel:digital_work_order_list (DIGITAL/GENERATED)
                              o panel:work_order_list (PDF).
        """
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        _compute_delta_hours ,
        _normalise_machine_code ,
        _resolve_machine_asset ,
        )
        from datetime import time as dt_time 
        from django .urls import reverse as _reverse_post 
        import json 

        company_user =request .user .company_user 
        company =company_user .company 
        _is_workshop_post =company_user .role =="WORKSHOP"

        try :
            work_order =self ._get_work_order (pk ,company )
        except WorkOrder .DoesNotExist :
            django_messages .error (request ,"Parte de trabajo no encontrado.")
            if _is_workshop_post :
                return redirect (_reverse_post ("panel:operator_history"))
            return redirect (_reverse_post ("panel:work_order_list"))





        _is_digital_post =work_order .source in (
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        )
        if _is_workshop_post :
            if not _is_digital_post or work_order .uploaded_by !=company_user :
                django_messages .error (
                request ,
                "No tienes permiso para editar este parte."
                )
                return redirect (_reverse_post ("panel:operator_history"))



        if _is_workshop_post :
            _list_url =_reverse_post ("panel:operator_history")
        elif _is_digital_post :
            _list_url =_reverse_post ("panel:digital_work_order_list")
        else :
            _list_url =_reverse_post ("panel:work_order_list")

        action =request .POST .get ("action","")





        if action =="regenerate":
            work_order .status =WorkOrder .Status .PENDING 
            work_order .excel_file =None 
            work_order .error_log =""
            work_order .save (update_fields =["status","excel_file","error_log"])
            from work_order_processor .services import generate_work_order_excel 
            from work_order_processor .tasks import process_work_order_pdf 
            generate_work_order_excel (work_order .pk )
            django_messages .success (
            request ,
            f"Excel regenerado correctamente para el Parte #{work_order.pk}."
            )
            return redirect (_list_url )





        if action =="save_line":
            line_pk =request .POST .get ("line_pk")
            try :
                line =WorkOrderEntryLine .objects .select_related (
                "entry__work_order"
                ).get (pk =line_pk ,entry__work_order =work_order )
            except WorkOrderEntryLine .DoesNotExist :
                django_messages .error (request ,"Línea no encontrada.")
                return redirect ("panel:work_order_edit",pk =pk )



            raw_norm =request .POST .get ("machine_norm","").strip ()
            norm =_normalise_machine_code (raw_norm )if raw_norm else raw_norm 
            asset =_resolve_machine_asset (norm ,company =company )if norm else None 



            def _parse_time_str (val ):
                """Parses HH:MM string into time, returns None on failure.
                --- Parsea cadena HH:MM a time, devuelve None si falla."""
                if not val :
                    return None 
                try :
                    parts =val .strip ().split (":")
                    return dt_time (int (parts [0 ]),int (parts [1 ]))
                except (ValueError ,IndexError ):
                    return None 

            hc =_parse_time_str (request .POST .get ("hc",""))
            hf =_parse_time_str (request .POST .get ("hf",""))
            delta =_compute_delta_hours (hc ,hf )



            flags_raw =request .POST .get ("flags","").strip ()
            flags =[f .strip ()for f in flags_raw .split (",")if f .strip ()]if flags_raw else []



            line .machine_norm =norm 
            line .machine_asset =asset 
            line .fault_description =request .POST .get ("fault_description","").strip ()
            line .repair_notes =request .POST .get ("repair_notes","").strip ()
            line .hc =hc 
            line .hf =hf 
            line .or_val =request .POST .get ("or_val","").strip ()
            line .delta_hours =delta 
            line .flags =flags 
            line .save (update_fields =[
            "machine_norm","machine_asset","fault_description",
            "repair_notes","hc","hf","or_val","delta_hours","flags",
            ])

            django_messages .success (
            request ,
            f"Bloque {line.line_number} de la página "
            f"{line.entry.page_number} guardado correctamente."
            )
            return redirect ("panel:work_order_edit",pk =pk )



        django_messages .warning (request ,"Acción no reconocida.")
        return redirect ("panel:work_order_edit",pk =pk )



class WorkOrderStatusFragmentView (SupervisorAccessMixin ,View ):
    """
    Returns the HTML status fragment for a single WorkOrder, used by HTMX polling.

    GET /panel/work-orders/<pk>/status/
        Renders only the _status_fragment.html partial for the WorkOrder identified
        by pk and scoped to the authenticated user's company. If the status is
        PENDING or PROCESSING the partial includes HTMX polling attributes so the
        browser continues polling every 4 seconds. When the status reaches DONE or
        ERROR the returned fragment contains no HTMX attributes and polling stops
        automatically.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Devuelve el fragmento HTML de estado de un WorkOrder, consumido por el polling HTMX.

    GET /panel/work-orders/<pk>/status/
        Renderiza únicamente el parcial _status_fragment.html para el WorkOrder
        identificado por pk, acotado a la empresa del usuario autenticado. Si el
        estado es PENDING o PROCESSING el parcial incluye atributos HTMX de polling
        para que el navegador siga consultando cada 4 segundos. Cuando el estado
        alcanza DONE o ERROR el fragmento devuelto no contiene atributos HTMX y el
        polling se detiene automáticamente.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def get (self ,request ,pk ):
        """
        Returns the rendered _status_fragment.html partial for the requested WorkOrder.
        Raises HTTP 404 if the WorkOrder does not exist or belongs to another company.
        ---
        Devuelve el parcial _status_fragment.html renderizado para el WorkOrder solicitado.
        Lanza HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.
        """
        from django .shortcuts import get_object_or_404 

        wo =get_object_or_404 (
        WorkOrder ,
        pk =pk ,
        company =request .user .company_user .company ,
        )
        return render (
        request ,
        "panel/work_orders/_status_fragment.html",
        {"wo":wo },
        )


class WorkOrderLineSaveView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that saves a single WorkOrderEntryLine and returns the updated
    <tr> row as an HTML fragment consumed by the inline editor.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/save/
         Expected POST fields (all optional — missing fields are treated as empty):
           machine_norm        : str  — normalised machine code.
           fault_description  : str  — fault description.
           repair_notes          : str  — repair description.
           hc                  : str  — start time  HH:MM.
           hf                  : str  — end time    HH:MM.
           or_val              : str  — repair order reference.
           flags               : str  — comma-separated flag list.
         Server recomputes delta_hours from hc/hf and re-resolves machine_asset
         from the updated machine_norm. Returns the rendered _line_row.html partial
         (a single <tr> element) with HTTP 200.
         Returns HTTP 404 if the WorkOrder or line do not exist or belong to another
         company. Returns HTTP 400 on an unexpected processing error.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Endpoint HTMX que guarda un único WorkOrderEntryLine y devuelve la fila <tr>
    actualizada como fragmento HTML consumido por el editor inline.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/save/
         Campos POST esperados (todos opcionales — los ausentes se tratan como vacíos):
           machine_norm        : str  — código de máquina normalizado.
           fault_description  : str  — descripción de la avería.
           repair_notes          : str  — descripción de la reparación.
           hc                  : str  — hora de comienzo HH:MM.
           hf                  : str  — hora de fin      HH:MM.
           or_val              : str  — referencia de orden de reparación.
           flags               : str  — lista de flags separada por comas.
         El servidor recalcula delta_hours desde hc/hf y re-resuelve machine_asset
         desde el machine_norm actualizado. Devuelve el parcial _line_row.html
         renderizado (un único elemento <tr>) con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder o la línea no existen o pertenecen a
         otra empresa. Devuelve HTTP 400 ante un error de procesamiento inesperado.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Saves the WorkOrderEntryLine identified by line_pk, recomputes derived
        fields and returns the updated <tr> row as an HTMX fragment.
        ---
        Guarda el WorkOrderEntryLine identificado por line_pk, recalcula los campos
        derivados y devuelve la fila <tr> actualizada como fragmento HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from work_order_processor .models import WorkOrderEntryLine 
        from work_order_processor .services import (
        _compute_delta_hours ,
        _normalise_machine_code ,
        _resolve_machine_asset ,
        )
        from datetime import time as dt_time 





        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )





        line =get_object_or_404 (
        WorkOrderEntryLine .objects .select_related (
        "entry__work_order",
        "machine_asset",
        ),
        pk =line_pk ,
        entry__work_order =wo ,
        )





        raw_norm =request .POST .get ("machine_norm","").strip ()
        norm =_normalise_machine_code (raw_norm )if raw_norm else raw_norm 
        asset =_resolve_machine_asset (norm ,company =company )if norm else None 





        def _parse_time_str (val ):
            """Parses HH:MM string into time object, returns None on failure.
            --- Parsea cadena HH:MM a objeto time, devuelve None si falla."""
            if not val :
                return None 
            try :
                parts =val .strip ().split (":")
                return dt_time (int (parts [0 ]),int (parts [1 ]))
            except (ValueError ,IndexError ):
                return None 

        hc =_parse_time_str (request .POST .get ("hc",""))
        hf =_parse_time_str (request .POST .get ("hf",""))
        delta =_compute_delta_hours (hc ,hf )





        flags_raw =request .POST .get ("flags","").strip ()
        flags =[f .strip ()for f in flags_raw .split (",")if f .strip ()]if flags_raw else []





        line .machine_norm =norm 
        line .machine_asset =asset 
        line .fault_description =request .POST .get ("fault_description","").strip ()
        line .repair_notes =request .POST .get ("repair_notes","").strip ()
        line .hc =hc 
        line .hf =hf 
        line .or_val =request .POST .get ("or_val","").strip ()
        line .delta_hours =delta 
        line .flags =flags 
        line .save (update_fields =[
        "machine_norm","machine_asset","fault_description",
        "repair_notes","hc","hf","or_val","delta_hours","flags",
        ])





        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":line ,
        "wo_pk":wo .pk ,
        "entry":line .entry ,
        },
        )




class WorkOrderEntrySaveDateView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that saves the work_date of a WorkOrderEntry and returns
    the updated group header fragment so the inline editor reflects the change
    without a full-page reload.

    POST /panel/work-orders/<wo_pk>/entries/<entry_pk>/save-date/
         Expected POST fields:
           work_date : str — date in YYYY-MM-DD format. Empty string clears
                             the date and keeps uncertain_date unchanged.
         On a valid non-empty date, uncertain_date is automatically set to
         False — the supervisor's explicit correction removes the uncertainty.
         Returns the rendered _entry_group_fragment.html partial with HTTP 200.
         Returns HTTP 400 on an invalid date format.
         Returns HTTP 404 if the WorkOrder or entry do not exist or belong to
         another company.

    ---

    Endpoint HTMX que guarda el work_date de un WorkOrderEntry y devuelve
    el fragmento de cabecera del grupo actualizado para que el editor inline
    refleje el cambio sin recarga completa de página.

    POST /panel/work-orders/<wo_pk>/entries/<entry_pk>/save-date/
         Campos POST esperados:
           work_date : str — fecha en formato YYYY-MM-DD. Cadena vacía limpia
                             la fecha y deja uncertain_date sin cambios.
         Con una fecha válida no vacía, uncertain_date se pone automáticamente
         a False — la corrección explícita del supervisor elimina la incertidumbre.
         Devuelve el parcial _entry_group_fragment.html renderizado con HTTP 200.
         Devuelve HTTP 400 ante un formato de fecha inválido.
         Devuelve HTTP 404 si el WorkOrder o entry no existen o son de otra empresa.
    """

    def post (self ,request ,wo_pk ,entry_pk ):
        """
        Persists the corrected work_date on the WorkOrderEntry and returns
        the group header fragment as an HTMX response.
        ---
        Persiste el work_date corregido en el WorkOrderEntry y devuelve el
        fragmento de cabecera del grupo como respuesta HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 
        from work_order_processor .models import WorkOrderEntry 
        from datetime import date as dt_date 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        entry =get_object_or_404 (WorkOrderEntry ,pk =entry_pk ,work_order =wo )

        raw_date =request .POST .get ("work_date","").strip ()

        if raw_date :


            try :
                parsed =dt_date .fromisoformat (raw_date )
            except ValueError :
                return HttpResponse (
                "Formato de fecha inválido. Use YYYY-MM-DD.",
                status =400 ,
                )
            entry .work_date =parsed 


            entry .uncertain_date =False 
        else :
            entry .work_date =None 



        entry .save (update_fields =["work_date","uncertain_date"])





        from decimal import Decimal 
        lines =list (
        entry .lines .select_related ("machine_asset").order_by ("line_number")
        )
        delta_sum =sum (
        (ln .delta_hours for ln in lines if ln .delta_hours is not None ),
        Decimal ("0"),
        )
        if delta_sum >0 :
            total_hours =float (delta_sum )
            if total_hours <8 :
                day_css ="day-total-short"
            elif total_hours <=12 :
                day_css ="day-total-normal"
            elif total_hours <=16 :
                day_css ="day-total-warning"
            else :
                day_css ="day-total-danger"
            day_total =round (total_hours ,2 )
        else :
            day_total =None 
            day_css =""

        group ={
        "entry":entry ,
        "lines":lines ,
        "day_total":day_total ,
        "day_css":day_css ,
        }

        return render (
        request ,
        "panel/work_orders/_entry_group_fragment.html",
        {
        "group":group ,
        "wo_pk":wo .pk ,
        },
        )


class MachineAssetAutocompleteView (SupervisorAccessMixin ,View ):
    """
    Lightweight JSON autocomplete endpoint for MachineAsset records.
    Returns up to 10 active assets matching the query string against
    code, brand_model and plate fields (case-insensitive icontains),
    scoped to the authenticated user's company.

    GET /panel/fleet/autocomplete/?q=<query>
        Returns:
          {"results": [{"code": "...", "label": "..."}, ...]}
        where label is "code — brand_model" or just "code" when brand_model
        is empty. Results are ordered by code ascending.

    Used by the machine_norm field autocomplete in _line_row.html.

    ---

    Endpoint JSON ligero de autocompletado para registros MachineAsset.
    Devuelve hasta 10 activos activos que coincidan con la cadena de consulta
    sobre los campos code, brand_model y plate (icontains sin distinción de
    mayúsculas), acotado a la empresa del usuario autenticado.

    GET /panel/fleet/autocomplete/?q=<query>
        Devuelve:
          {"results": [{"code": "...", "label": "..."}, ...]}
        donde label es "code — brand_model" o solo "code" cuando brand_model
        está vacío. Resultados ordenados por code ascendente.

    Usado por el autocompletado del campo machine_norm en _line_row.html.
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns a JSON list of matching MachineAsset records for the given query.
        ---
        Devuelve un JSON con los MachineAsset coincidentes para la consulta dada.
        """
        from django .http import JsonResponse 
        from fleet .models import MachineAsset 

        company =request .user .company_user .company 
        q =request .GET .get ("q","").strip ()

        qs =MachineAsset .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("code")

        if q :
            qs =qs .filter (
            Q (code__icontains =q )
            |Q (brand_model__icontains =q )
            |Q (plate__icontains =q )
            )

        results =[]
        for asset in qs [:10 ]:
            label =asset .code 
            if asset .brand_model :
                label =f"{asset.code} — {asset.brand_model}"
            results .append ({"code":asset .code ,"label":label })

        return JsonResponse ({"results":results })



class WorkOrderLineInsertView (SupervisorAccessMixin ,View ):
    """
    Creates a new empty WorkOrderEntryLine after a given line within the same
    WorkOrderEntry group and returns the new row as an HTMX fragment.

    POST /panel/work-orders/<wo_pk>/lines/insert/
         Expected POST fields:
           after_line_pk : int — pk of the line after which to insert.
           entry_pk      : int — pk of the WorkOrderEntry group.
         Creates a new WorkOrderEntryLine with line_number = after_line.line_number + 1,
         shifting the line_number of all subsequent lines in the same entry up by 1.
         Returns the rendered _line_row.html partial for the new line with HTTP 200.
         Returns HTTP 404 if the WorkOrder, entry or reference line do not belong
         to the authenticated company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Crea un nuevo WorkOrderEntryLine vacío tras una línea dada dentro del mismo
    grupo WorkOrderEntry y devuelve la nueva fila como fragmento HTMX.

    POST /panel/work-orders/<wo_pk>/lines/insert/
         Campos POST esperados:
           after_line_pk : int — pk de la línea tras la que insertar.
           entry_pk      : int — pk del grupo WorkOrderEntry.
         Crea un nuevo WorkOrderEntryLine con line_number = after_line.line_number + 1,
         desplazando el line_number de todas las líneas posteriores del mismo entry
         en +1. Devuelve el parcial _line_row.html renderizado para la nueva línea
         con HTTP 200. Devuelve HTTP 404 si el WorkOrder, entry o línea de referencia
         no pertenecen a la empresa autenticada.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ):
        """
        Inserts a new empty WorkOrderEntryLine after the specified reference line
        and returns the rendered _line_row.html fragment for the new row.
        ---
        Inserta un nuevo WorkOrderEntryLine vacío tras la línea de referencia
        especificada y devuelve el fragmento _line_row.html renderizado para la fila.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponseBadRequest 

        company =request .user .company_user .company 



        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )



        try :
            after_line_pk =int (request .POST .get ("after_line_pk",""))
            entry_pk =int (request .POST .get ("entry_pk",""))
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [INSERT] Parámetros after_line_pk / entry_pk inválidos.")



        after_line =get_object_or_404 (
        WorkOrderEntryLine ,
        pk =after_line_pk ,
        entry__work_order =wo ,
        entry__pk =entry_pk ,
        )
        entry =after_line .entry 







        from django .db import transaction 
        with transaction .atomic ():


            WorkOrderEntryLine .objects .filter (
            entry =entry ,
            line_number__gt =after_line .line_number ,
            ).update (line_number =django_models .F ("line_number")+1 )



            new_line =WorkOrderEntryLine .objects .create (
            entry =entry ,
            line_number =after_line .line_number +1 ,
            machine_norm ="",
            machine_raw ="",
            fault_description ="",
            repair_notes ="",
            hc =None ,
            hf =None ,
            or_val ="",
            delta_hours =None ,
            flags =[],
            machine_asset =None ,
            )

        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":new_line ,
        "wo_pk":wo .pk ,
        "entry":entry ,
        },
        )


class WorkOrderLineReorderView (SupervisorAccessMixin ,View ):
    """
    Accepts a new ordering for WorkOrderEntryLine records within a single
    WorkOrderEntry group and persists the updated line_number values.

    POST /panel/work-orders/<wo_pk>/lines/reorder/
         Expected POST fields:
           entry_pk        : int         — pk of the WorkOrderEntry group.
           line_pks[]      : list[int]   — ordered list of WorkOrderEntryLine pks
                                           in the desired new order.
         Updates line_number for each line to match the position in line_pks[].
         Returns HTTP 200 with JSON {"ok": true} on success.
         Returns HTTP 400 on invalid parameters.
         Returns HTTP 404 if the WorkOrder or entry do not belong to the company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Acepta un nuevo orden para los registros WorkOrderEntryLine dentro de un único
    grupo WorkOrderEntry y persiste los valores line_number actualizados.

    POST /panel/work-orders/<wo_pk>/lines/reorder/
         Campos POST esperados:
           entry_pk        : int         — pk del grupo WorkOrderEntry.
           line_pks[]      : list[int]   — lista ordenada de pks de WorkOrderEntryLine
                                           en el nuevo orden deseado.
         Actualiza line_number de cada línea para que coincida con su posición en
         line_pks[]. Devuelve HTTP 200 con JSON {"ok": true} en caso de éxito.
         Devuelve HTTP 400 ante parámetros inválidos. Devuelve HTTP 404 si el
         WorkOrder o entry no pertenecen a la empresa.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ):
        """
        Persists the new line_number ordering for the lines of a WorkOrderEntry.
        ---
        Persiste el nuevo orden de line_number para las líneas de un WorkOrderEntry.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import JsonResponse ,HttpResponseBadRequest 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )



        try :
            entry_pk =int (request .POST .get ("entry_pk",""))
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [REORDER] Parámetro entry_pk inválido.")

        entry =get_object_or_404 (WorkOrderEntry ,pk =entry_pk ,work_order =wo )



        try :
            line_pks =[int (pk )for pk in request .POST .getlist ("line_pks[]")]
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [REORDER] Parámetro line_pks[] inválido.")

        if not line_pks :
            return HttpResponseBadRequest ("# [REORDER] Lista line_pks[] vacía.")



        lines_map ={
        line .pk :line 
        for line in WorkOrderEntryLine .objects .filter (entry =entry )
        }



        bulk_update =[]
        for position ,pk in enumerate (line_pks ,start =1 ):
            line =lines_map .get (pk )
            if line is None :
                continue 
            line .line_number =position 
            bulk_update .append (line )

        WorkOrderEntryLine .objects .bulk_update (bulk_update ,["line_number"])

        return JsonResponse ({"ok":True })


class WorkOrderLineRestoreView (SupervisorAccessMixin ,View ):
    """
    Restores a single WorkOrderEntryLine to its original Gemini-extracted
    values by looking up the corresponding block in the parent WorkOrderEntry's
    raw_gemini_response using the line's line_number as the index.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/restore/
         Locates the bloque at raw_gemini_response["entradas"][line_number - 1],
         overwrites only the fields of the target line (machine_raw, machine_norm,
         machine_asset, fault_description, repair_notes, hc, hf, or_val,
         delta_hours, flags) and returns the rendered _line_row.html partial
         for that single row with HTTP 200.
         Returns HTTP 404 if the WorkOrder or line do not belong to the company.
         Returns HTTP 400 if no raw_gemini_response is stored or the block index
         is out of range.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Restaura un único WorkOrderEntryLine a sus valores originales extraídos por
    Gemini localizando el bloque correspondiente en el raw_gemini_response del
    WorkOrderEntry padre usando el line_number de la línea como índice.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/restore/
         Localiza el bloque en raw_gemini_response["entradas"][line_number - 1],
         sobreescribe únicamente los campos de la línea objetivo (machine_raw,
         machine_norm, machine_asset, fault_description, repair_notes, hc, hf,
         or_val, delta_hours, flags) y devuelve el parcial _line_row.html
         renderizado para esa única fila con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder o la línea no pertenecen a la empresa.
         Devuelve HTTP 400 si no hay raw_gemini_response almacenado o el índice
         de bloque está fuera de rango.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Restores the single WorkOrderEntryLine identified by line_pk from its
        corresponding block in raw_gemini_response and returns the updated row
        as an HTMX fragment.
        ---
        Restaura el único WorkOrderEntryLine identificado por line_pk desde su
        bloque correspondiente en raw_gemini_response y devuelve la fila
        actualizada como fragmento HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponseBadRequest 
        from work_order_processor .services import (
        _normalise_machine_code ,
        _resolve_machine_asset ,
        _compute_delta_hours ,
        _parse_time ,
        )

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )



        line =get_object_or_404 (
        WorkOrderEntryLine .objects .select_related ("entry","machine_asset"),
        pk =line_pk ,
        entry__work_order =wo ,
        )
        entry =line .entry 










        raw =entry .raw_gemini_response 

        if not raw or not isinstance (raw ,dict ):


            machine_norm =_normalise_machine_code (line .machine_raw or "")
            machine_asset =_resolve_machine_asset (machine_norm ,company =company )if machine_norm else None 
            delta =_compute_delta_hours (line .hc ,line .hf ,deduct_lunch =False )

            line .machine_norm =machine_norm 
            line .machine_asset =machine_asset 
            line .delta_hours =delta 
            line .save (update_fields =["machine_norm","machine_asset","delta_hours"])

            return render (
            request ,
            "panel/work_orders/_line_row.html",
            {
            "line":line ,
            "wo_pk":wo .pk ,
            "entry":entry ,
            },
            )



        entradas =raw .get ("entradas")or []
        block_index =line .line_number -1 

        if block_index <0 or block_index >=len (entradas ):
            return HttpResponseBadRequest (
            f"# [RESTORE] Índice de bloque {block_index} fuera de rango "
            f"(entradas disponibles: {len(entradas)})."
            )

        bloque =entradas [block_index ]



        machine_raw =(bloque .get ("machine_raw")or "").strip ()
        machine_norm =_normalise_machine_code (machine_raw )
        machine_asset =_resolve_machine_asset (machine_norm ,company =company )
        hc =_parse_time (bloque .get ("hc"))
        hf =_parse_time (bloque .get ("hf"))
        delta =_compute_delta_hours (hc ,hf )
        flags =bloque .get ("flags")or []
        if not isinstance (flags ,list ):
            flags =[]

        line .machine_raw =machine_raw 
        line .machine_norm =machine_norm 
        line .machine_asset =machine_asset 
        line .fault_description =(bloque .get ("fault_description")or "")
        line .repair_notes =(bloque .get ("repair_notes")or "")
        line .hc =hc 
        line .hf =hf 
        line .or_val =(bloque .get ("or_val")or "")
        line .delta_hours =delta 
        line .flags =flags 
        line .save (update_fields =[
        "machine_raw","machine_norm","machine_asset",
        "fault_description","repair_notes","hc","hf",
        "or_val","delta_hours","flags",
        ])

        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":line ,
        "wo_pk":wo .pk ,
        "entry":entry ,
        },
        )


class WorkOrderLineDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a single WorkOrderEntryLine identified by line_pk, scoped to the
    authenticated company. Returns an empty HTTP 200 response so that HTMX
    removes the <tr> row from the DOM via hx-swap="outerHTML" with an empty
    response body.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/delete/
         Returns HTTP 404 if the WorkOrder or line do not exist or belong to
         another company. Restricted to ADMIN role.

    ---

    Elimina un único WorkOrderEntryLine identificado por line_pk, acotado a la
    empresa autenticada. Devuelve una respuesta HTTP 200 vacía para que HTMX
    elimine la fila <tr> del DOM via hx-swap="outerHTML" con cuerpo vacío.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/delete/
         Devuelve HTTP 404 si el WorkOrder o la línea no existen o pertenecen a
         otra empresa. Restringido al rol ADMIN.
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Deletes the WorkOrderEntryLine identified by line_pk and returns an
        empty response for HTMX to remove the row from the DOM.
        ---
        Elimina el WorkOrderEntryLine identificado por line_pk y devuelve una
        respuesta vacía para que HTMX elimine la fila del DOM.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        line =get_object_or_404 (
        WorkOrderEntryLine ,
        pk =line_pk ,
        entry__work_order =wo ,
        )
        line .delete ()


        return HttpResponse ("")


class WorkOrderDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a WorkOrder and all its cascade-deleted children (WorkOrderEntry,
    WorkOrderEntryLine, source PDF file, Excel file) scoped to the authenticated
    company. Redirects to the work order list on success.

    POST /panel/work-orders/<pk>/delete/
         Returns HTTP 404 if the WorkOrder does not exist or belongs to another
         company. Restricted to ADMIN role.

    ---

    Elimina un WorkOrder y todos sus hijos eliminados en cascada (WorkOrderEntry,
    WorkOrderEntryLine, PDF original, archivo Excel) acotado a la empresa
    autenticada. Redirige a la lista de partes tras el éxito.

    POST /panel/work-orders/<pk>/delete/
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.
         Restringido al rol ADMIN.
    """

    def post (self ,request ,pk ):
        """
        Deletes the WorkOrder identified by pk, scoped to the authenticated company.
        ---
        Elimina el WorkOrder identificado por pk, acotado a la empresa autenticada.
        """
        from django .shortcuts import get_object_or_404 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )
        wo_pk =wo .pk 
        wo .delete ()
        django_messages .success (
        request ,
        f"Parte #{wo_pk} eliminado correctamente."
        )
        return redirect ("panel:work_order_list")


class WorkOrderMarkReviewedView (SupervisorAccessMixin ,View ):
    """
    Toggles the reviewed flag on a WorkOrder identified by pk, scoped to the
    authenticated company. Returns an HTML fragment (_review_badge_fragment.html)
    for HTMX to swap inline in list.html.

    POST /panel/work-orders/<pk>/review/
         If reviewed is currently False:
           Sets reviewed=True, reviewed_by=request.user.company_user,
           reviewed_at=now().
         If reviewed is currently True:
           Sets reviewed=False, reviewed_by=None, reviewed_at=None.
         Returns the rendered _review_badge_fragment.html partial with HTTP 200.
         Returns HTTP 404 if the WorkOrder does not exist or belongs to another
         company.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Alterna el flag reviewed de un WorkOrder identificado por pk, acotado a la
    empresa autenticada. Devuelve un fragmento HTML (_review_badge_fragment.html)
    para que HTMX lo intercambie inline en list.html.

    POST /panel/work-orders/<pk>/review/
         Si reviewed es actualmente False:
           Establece reviewed=True, reviewed_by=request.user.company_user,
           reviewed_at=now().
         Si reviewed es actualmente True:
           Establece reviewed=False, reviewed_by=None, reviewed_at=None.
         Devuelve el parcial _review_badge_fragment.html renderizado con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ,pk ):
        """
        Toggles reviewed on the WorkOrder identified by pk and returns the
        rendered review badge fragment for HTMX inline swap.
        ---
        Alterna reviewed en el WorkOrder identificado por pk y devuelve el
        fragmento del badge de revisión para el intercambio inline de HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .utils .timezone import now as tz_now 

        company =request .user .company_user .company 
        company_user =request .user .company_user 

        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )

        if wo .reviewed :


            wo .reviewed =False 
            wo .reviewed_by =None 
            wo .reviewed_at =None 
            wo .save (update_fields =["reviewed","reviewed_by","reviewed_at"])
        else :


            wo .reviewed =True 
            wo .reviewed_by =company_user 
            wo .reviewed_at =tz_now ()
            wo .save (update_fields =["reviewed","reviewed_by","reviewed_at"])

        return render (
        request ,
        "panel/work_orders/_review_badge_fragment.html",
        {"wo":wo },
        )


class WorkOrderExportView (SupervisorAccessMixin ,View ):
    """
    Generates and returns a multi-sheet Excel file concatenating the individual
    Excel reports of a selection of WorkOrder records identified by their pks.

    POST /panel/work-orders/export/
         Expected POST fields:
           pks         : list[int] (repeating field "pks") — primary keys of the
                         WorkOrder records to include. Only DONE records belonging
                         to the authenticated company are exported.
           export_mode : str — "single_sheet" or "multi_sheet" (Hito 8 / Bloque H).
                         single_sheet: one sheet with all entries ordered by
                           worker_name then date_key; an opaque header row marks
                           each new worker block.
                         multi_sheet: one sheet per worker with individual header.
         Returns HTTP 400 if no valid pks are provided.
         Returns HTTP 404 if the authenticated user has no CompanyUser profile.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Genera y devuelve un archivo Excel multi-hoja que concatena los informes
    individuales de una selección de registros WorkOrder identificados por sus pks.

    POST /panel/work-orders/export/
         Campos POST esperados:
           pks         : list[int] (campo repetido "pks") — claves primarias de
                         los WorkOrder a incluir. Solo se exportan DONE de la empresa.
           export_mode : str — "single_sheet" o "multi_sheet" (Hito 8 / Bloque H).
                         single_sheet: una sola hoja con todas las entradas ordenadas
                           por operario y fecha; una fila cabecera separadora por
                           cada nuevo operario.
                         multi_sheet: una hoja por operario con su membrete.
         Devuelve HTTP 400 si no se proporcionan pks válidos.
         Devuelve HTTP 404 si el usuario no tiene perfil CompanyUser.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ):
        """
        Builds the export Excel file from the selected WorkOrder pks.
        Supports two export modes controlled by the POST field export_mode:

          single_sheet — One flat sheet with all WorkOrderEntryLine records
            from all selected WorkOrders, grouped first by worker_name then
            by work_date. A dark-blue separator row bearing the worker name
            is inserted before the first block of each new worker. All data
            is read directly from the DB — no Excel regeneration is triggered.

          multi_sheet  — One sheet per distinct worker_name. Each sheet is
            built by copying the first sheet of the individual Excel stored
            in WorkOrder.excel_file (regenerated on-the-fly if missing).
            Sheet title is truncated to 31 characters (Excel limit).

        Both modes return an HttpResponse with Content-Disposition attachment.
        Returns HTTP 400 on invalid or missing pks or on unknown export_mode.

        ---

        Construye el Excel de exportación a partir de los pks de WorkOrder
        seleccionados. Soporta dos modos controlados por el campo POST
        export_mode:

          single_sheet — Una hoja plana con todos los WorkOrderEntryLine de
            todos los WorkOrders seleccionados, agrupados por worker_name y
            después por work_date. Una fila separadora azul oscuro con el
            nombre del operario se inserta antes del primer bloque de cada
            nuevo operario. Los datos se leen directamente de la BD — no se
            regenera ningún Excel individual.

          multi_sheet  — Una hoja por worker_name distinto. Cada hoja se
            construye copiando la primera hoja del Excel individual almacenado
            en WorkOrder.excel_file (regenerado al vuelo si falta).
            El título de hoja se trunca a 31 caracteres (límite de Excel).

        Ambos modos devuelven HttpResponse con Content-Disposition attachment.
        Devuelve HTTP 400 ante pks inválidos/ausentes o export_mode desconocido.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        generate_work_order_excel as _gen_excel ,
        _worker_name_from_pdf_path ,
        )

        company =request .user .company_user .company 
        export_mode =request .POST .get ("export_mode","single_sheet").strip ()





        if export_mode not in ("single_sheet","multi_sheet"):
            return HttpResponseBadRequest (
            f"# [EXPORT] Modo de exportación desconocido: {export_mode!r}."
            )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (pk )for pk in raw_pks if pk ]
        except (ValueError ,TypeError ):
            return HttpResponseBadRequest ("# [EXPORT] Parámetros pks inválidos.")

        if not pk_list :
            return HttpResponseBadRequest (
            "# [EXPORT] No se han seleccionado partes para exportar."
            )







        work_orders =list (
        WorkOrder .objects 
        .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        )
        .order_by ("pk")
        )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [EXPORT] Ninguno de los partes seleccionados está revisado y en estado DONE. "
            "Solo se pueden exportar partes marcados como revisados."
            )





        def _copy_sheet (src_sheet ,dest_wb ,title ):
            """
            Copies src_sheet (cells, styles, column widths, row heights) into
            a new sheet named title in dest_wb.
            ---
            Copia src_sheet (celdas, estilos, anchos de columna, alturas de
            fila) en una nueva hoja llamada title en dest_wb.
            """
            dest_sheet =dest_wb .create_sheet (title =title [:31 ])
            for row in src_sheet .iter_rows ():
                for cell in row :
                    dest_cell =dest_sheet .cell (
                    row =cell .row ,column =cell .column ,value =cell .value 
                    )
                    if cell .has_style :
                        dest_cell .font =cell .font .copy ()
                        dest_cell .fill =cell .fill .copy ()
                        dest_cell .alignment =cell .alignment .copy ()
                        dest_cell .border =cell .border .copy ()
                        dest_cell .number_format =cell .number_format 
            for col_letter ,col_dim in src_sheet .column_dimensions .items ():
                dest_sheet .column_dimensions [col_letter ].width =col_dim .width 
            for row_num ,row_dim in src_sheet .row_dimensions .items ():
                dest_sheet .row_dimensions [row_num ].height =row_dim .height 
            return dest_sheet 





        def _get_worker_name (wo ):
            """
            Returns the worker name from the PDF filename, or a fallback label.
            ---
            Devuelve el nombre del operario del nombre del PDF, o etiqueta de
            reserva.
            """
            if wo .source_pdf :
                return _worker_name_from_pdf_path (wo .source_pdf .name )
            return f"Operario #{wo.pk}"












        if export_mode =="single_sheet":





            wo_map ={wo .pk :wo for wo in work_orders }

            lines_qs =(
            WorkOrderEntryLine .objects 
            .filter (entry__work_order__in =work_orders )
            .select_related (
            "entry",
            "entry__work_order",
            "machine_asset",
            )
            .order_by (
            "entry__work_order__pk",
            "entry__work_date",
            "entry__page_number",
            "line_number",
            )
            )





            enriched =[]
            for line in lines_qs :
                entry =line .entry 
                wo =entry .work_order 
                worker_name =_get_worker_name (wo )
                date_key =entry .work_date or ""
                enriched .append ({
                "worker_name":worker_name ,
                "date_key":date_key ,
                "line":line ,
                "entry":entry ,
                })



            enriched .sort (key =lambda r :(
            r ["worker_name"],
            r ["date_key"]if r ["date_key"]else "",
            ))





            wb =openpyxl .Workbook ()
            ws =wb .active 
            ws .title ="EXPORTACION"



            _SEP_BG ="1F4E79"
            _SEP_FG ="FFFFFF"

            def _make_sep_fill ():
                return PatternFill (
                fill_type ="solid",
                start_color =_SEP_BG ,
                end_color =_SEP_BG ,
                )


            headers =[
            "OPERARIO","FECHA","BLOQUE","MÁQUINA (NORM)",
            "MÁQUINA (RAW)","ACTIVO RESUELTO","DESCRIPCIÓN AVERÍA",
            "REPARACIÓN","H.C.","H.F.","Δ HORAS","O.R.","FLAGS",
            ]
            NUM_COLS =len (headers )

            for col_idx ,hdr in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =hdr )
                cell .font =Font (bold =True ,color =_SEP_FG )
                cell .fill =_make_sep_fill ()
                cell .alignment =Alignment (horizontal ="center",vertical ="center")
            ws .row_dimensions [1 ].height =20 

            current_worker =None 
            data_row =2 

            for rec in enriched :
                worker_name =rec ["worker_name"]
                line =rec ["line"]
                entry =rec ["entry"]



                if worker_name !=current_worker :
                    ws .merge_cells (
                    start_row =data_row ,start_column =1 ,
                    end_row =data_row ,end_column =NUM_COLS ,
                    )
                    sep_cell =ws .cell (row =data_row ,column =1 ,
                    value =worker_name )
                    sep_cell .font =Font (bold =True ,color =_SEP_FG ,size =11 )
                    sep_cell .fill =_make_sep_fill ()
                    sep_cell .alignment =Alignment (horizontal ="left",
                    vertical ="center")
                    ws .row_dimensions [data_row ].height =22 
                    data_row +=1 
                    current_worker =worker_name 


                date_display =(
                entry .work_date .strftime ("%d/%m/%Y")
                if entry .work_date else ""
                )
                asset_code =(
                line .machine_asset .code if line .machine_asset else ""
                )
                hc_display =(
                line .hc .strftime ("%H:%M")if line .hc else ""
                )
                hf_display =(
                line .hf .strftime ("%H:%M")if line .hf else ""
                )
                delta_display =(
                str (line .delta_hours )if line .delta_hours is not None else ""
                )
                flags_display =", ".join (line .flags )if line .flags else ""

                row_values =[
                worker_name ,
                date_display ,
                line .line_number ,
                line .machine_norm ,
                line .machine_raw ,
                asset_code ,
                line .fault_description ,
                line .repair_notes ,
                hc_display ,
                hf_display ,
                delta_display ,
                line .or_val ,
                flags_display ,
                ]
                for col_idx ,val in enumerate (row_values ,start =1 ):
                    cell =ws .cell (row =data_row ,column =col_idx ,value =val )
                    cell .alignment =Alignment (vertical ="center",wrap_text =False )
                ws .row_dimensions [data_row ].height =16 
                data_row +=1 



            for col_idx ,hdr in enumerate (headers ,start =1 ):
                col_letter =openpyxl .utils .get_column_letter (col_idx )
                ws .column_dimensions [col_letter ].width =min (
                max (len (hdr )+2 ,12 ),60 
                )

            if ws .max_row <2 :
                return HttpResponseBadRequest (
                "# [EXPORT] No hay líneas de datos en los partes seleccionados."
                )










        else :


            from collections import OrderedDict 
            groups :dict [str ,list ]=OrderedDict ()
            for wo in work_orders :
                worker_name =_get_worker_name (wo )
                groups .setdefault (worker_name ,[]).append (wo )

            wb =openpyxl .Workbook ()
            wb .remove (wb .active )

            for worker_name ,wo_list in groups .items ():




                sheet_built =False 
                for wo in wo_list :
                    try :


                        if not wo .excel_file :
                            _gen_excel (wo .pk )
                            wo .refresh_from_db (fields =["excel_file"])
                        if not wo .excel_file :
                            continue 
                        with wo .excel_file .open ("rb")as f :
                            buf =io .BytesIO (f .read ())
                        src_wb =openpyxl .load_workbook (buf )
                        src_sheet =src_wb .worksheets [0 ]


                        sheet_title =(
                        (worker_name [:28 ]+f"#{wo.pk}")
                        if worker_name else f"Parte#{wo.pk}"
                        )
                        _copy_sheet (src_sheet ,wb ,sheet_title )
                        sheet_built =True 
                        break 
                    except Exception :
                        continue 

                if not sheet_built :


                    placeholder =wb .create_sheet (
                    title =(worker_name [:28 ]if worker_name else "Sin datos")[:31 ]
                    )
                    placeholder .cell (row =1 ,column =1 ,
                    value ="No se pudo generar el Excel para este operario.")

            if not wb .worksheets :
                return HttpResponseBadRequest (
                "# [EXPORT] No se pudo generar ninguna hoja para los partes seleccionados."
                )





        output =io .BytesIO ()
        wb .save (output )
        output .seek (0 )

        filename =f"EXPORTACION_{tz_now().strftime('%d-%m-%y')}.xlsx"
        response =HttpResponse (
        output .read (),
        content_type =(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


class WorkOrderDuplicateSearchView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that detects duplicate WorkOrders for the authenticated
    company by querying WorkOrderEntry records sharing identical
    (worker_name, work_date) tuples across distinct WorkOrders.
    Returns a rendered _duplicates_fragment.html partial with the grouped
    results, or an informational fragment when no duplicates are found.

    POST /panel/work-orders/duplicates/search/
         Returns HTTP 200 with the rendered fragment in all cases.
         Returns HTTP 404 if the authenticated user has no CompanyUser profile.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Endpoint HTMX que detecta WorkOrders duplicados para la empresa autenticada
    consultando registros WorkOrderEntry que comparten tuplas (worker_name,
    work_date) idénticas entre distintos WorkOrders. Devuelve el parcial
    _duplicates_fragment.html renderizado con los resultados agrupados, o un
    fragmento informativo cuando no se detectan duplicados.

    POST /panel/work-orders/duplicates/search/
         Devuelve HTTP 200 con el fragmento renderizado en todos los casos.
         Devuelve HTTP 404 si el usuario autenticado no tiene perfil CompanyUser.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ):
        """
        Executes the duplicate detection query scoped to the authenticated
        company and returns the rendered _duplicates_fragment.html partial.

        Detection logic mirrors detect_duplicate_entries management command:
          1. Query (company, worker_name, work_date) groups where more than one
             distinct WorkOrder contributes WorkOrderEntry records.
          2. For each group, fetch all implicated WorkOrders ordered by pk.
          3. Identify keeper (highest pk) and candidates for deletion (lower pks).

        ---

        Ejecuta la consulta de detección de duplicados acotada a la empresa
        autenticada y devuelve el parcial _duplicates_fragment.html renderizado.

        La lógica de detección es equivalente al comando de gestión
        detect_duplicate_entries:
          1. Consultar grupos (empresa, operario, fecha) donde más de un
             WorkOrder distinto aporta registros WorkOrderEntry.
          2. Por cada grupo, obtener todos los WorkOrders implicados ordenados
             por pk.
          3. Identificar el conservado (pk más alto) y los candidatos a
             eliminación (pks inferiores).
        """
        from django .db .models import Count 
        from work_order_processor .models import WorkOrderEntry 

        company =request .user .company_user .company 





        raw_groups =list (
        WorkOrderEntry .objects 
        .filter (work_order__company =company )
        .values ("worker_name","work_date")
        .annotate (wo_count =Count ("work_order_id",distinct =True ))
        .filter (wo_count__gt =1 )
        .order_by ("worker_name","work_date")
        )

        if not raw_groups :


            return render (
            request ,
            "panel/work_orders/_duplicates_fragment.html",
            {"duplicate_groups":[],"no_duplicates":True },
            )





        enriched_groups =[]

        for raw in raw_groups :
            worker_name =raw ["worker_name"]or ""
            work_date =raw ["work_date"]



            implicated =list (
            WorkOrder .objects 
            .filter (
            company =company ,
            entries__worker_name =worker_name ,
            entries__work_date =work_date ,
            )
            .select_related ("uploaded_by__user")
            .distinct ()
            .order_by ("pk")
            )

            if len (implicated )<2 :


                continue 

            enriched_groups .append ({
            "worker_name":worker_name ,
            "work_date":work_date ,
            "work_orders":implicated ,
            "keeper":implicated [-1 ],
            "to_delete":implicated [:-1 ],
            })

        return render (
        request ,
        "panel/work_orders/_duplicates_fragment.html",
        {
        "duplicate_groups":enriched_groups ,
        "no_duplicates":False ,
        "company_user":request .user .company_user ,
        },
        )


class WorkOrderDuplicateDeleteView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that deletes a single WorkOrder identified by pk, scoped
    to the authenticated company. Intended for use in the duplicates panel:
    the caller is responsible for ensuring the pk belongs to a duplicate
    group (not the keeper). Returns an empty HTTP 200 response so that HTMX
    removes the corresponding row from the DOM via hx-swap="outerHTML".

    POST /panel/work-orders/duplicates/<pk>/delete/
         Returns HTTP 404 if the WorkOrder does not exist or belongs to
         another company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Endpoint HTMX que elimina un único WorkOrder identificado por pk, acotado
    a la empresa autenticada. Diseñado para su uso en el panel de duplicados:
    el llamador es responsable de garantizar que el pk pertenece a un grupo
    duplicado (no al conservado). Devuelve una respuesta HTTP 200 vacía para
    que HTMX elimine la fila correspondiente del DOM via hx-swap="outerHTML".

    POST /panel/work-orders/duplicates/<pk>/delete/
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,pk ):
        """
        Deletes the WorkOrder identified by pk, cascade-removing all its
        children (WorkOrderEntry, WorkOrderEntryLine, source PDF, Excel file).
        Returns an empty response body for HTMX outerHTML swap.

        ---

        Elimina el WorkOrder identificado por pk, eliminando en cascada todos
        sus hijos (WorkOrderEntry, WorkOrderEntryLine, PDF original, Excel).
        Devuelve cuerpo vacío para el swap outerHTML de HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )
        wo .delete ()



        return HttpResponse ("")


class AnalyticsView (SupervisorAccessMixin ,View ):
    """
    Renders the analytics dashboard for the authenticated user's company.
    Serves the template shell only — all data is fetched client-side via
    the AnalyticsDataView JSON endpoint (/panel/analytics/data/).

    ---

    Renderiza el panel de analítica para la empresa del usuario autenticado.
    Sirve únicamente el shell del template — todos los datos se obtienen
    en el cliente mediante el endpoint JSON AnalyticsDataView
    (/panel/analytics/data/).
    """

    template_name ="panel/analytics.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Renders the analytics page. No chart data is computed here —
        the JS layer fetches it from /panel/analytics/data/.
        ---
        Renderiza la página de analítica. No se calculan datos de gráfico
        aquí — la capa JS los obtiene de /panel/analytics/data/.
        """
        company_user =request .user .company_user 
        company =company_user .company 

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"analytics",
        })


class AnalyticsDataView (SupervisorAccessMixin ,View ):
    """
    JSON endpoint that returns all WorkOrderEntryLine records for the
    authenticated user's company, enriched with machine asset metadata,
    work date and WorkOrder reference. The client-side chart builder
    consumes this payload to filter, aggregate and render Plotly charts
    without further server round-trips.

    Response schema:
    {
        "lines": [
            {
                "id":          int,
                "work_date":   "YYYY-MM-DD" | null,
                "work_order":  int,
                "pdf_name":    str,
                "code":      str,
                "brand_model": str,
                "delta_hours": float | null,
                "weekday":     int | null   // 0=Mon … 4=Fri
            },
            ...
        ],
        "work_orders": [
            {"id": int, "label": str},
            ...
        ],
        "assets": [
            {"code": str, "brand_model": str},
            ...
        ]
    }

    ---

    Endpoint JSON que devuelve todos los registros WorkOrderEntryLine de la
    empresa del usuario autenticado, enriquecidos con metadatos del activo,
    fecha de trabajo y referencia al WorkOrder. El constructor de gráficos
    client-side consume este payload para filtrar, agregar y renderizar
    gráficos Plotly sin más round-trips al servidor.

    Esquema de respuesta:
    {
        "lines": [
            {
                "id":           int,
                "work_date":    "YYYY-MM-DD" | null,
                "work_order":   int,
                "pdf_name":     str,
                "code":       str,
                "brand_model": str,
                "delta_hours":  float | null,
                "weekday":      int | null   // 0=Lun … 4=Vie
            },
            ...
        ],
        "work_orders": [
            {"id": int, "label": str},
            ...
        ],
        "assets": [
            {"code": str, "brand_model": str},
            ...
        ]
    }
    """



    _WEEKDAY_LABELS =["Lunes","Martes","Miércoles","Jueves","Viernes"]

    def get (self ,request ):
        """
        Queries and serialises all WorkOrderEntryLine records for the company.
        Returns HTTP 200 with a JSON payload or HTTP 403 if the user has no
        associated CompanyUser profile.

        ---

        Consulta y serializa todos los registros WorkOrderEntryLine de la empresa.
        Devuelve HTTP 200 con payload JSON o HTTP 403 si el usuario no tiene
        perfil CompanyUser asociado.
        """
        import json as _json 
        from django .http import JsonResponse 

        try :
            company =request .user .company_user .company 
        except AttributeError :
            return JsonResponse ({"error":"Sin perfil de empresa asociado."},status =403 )





        qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        machine_asset__isnull =False ,
        entry__work_date__isnull =False ,
        )
        .select_related (
        "machine_asset",
        "entry",
        "entry__work_order",
        )
        .order_by ("entry__work_date","machine_asset__code")
        )

        lines =[]
        for line in qs :
            work_date =line .entry .work_date 
            delta =float (line .delta_hours )if line .delta_hours is not None else None 
            pdf_name =line .entry .work_order .source_pdf .name .split ("/")[-1 ]


            import re as _re 
            pdf_label =_re .sub (r'_[A-Za-z0-9]{7}(\.[^.]+)$',r'',pdf_name )

            lines .append ({
            "id":line .pk ,
            "work_date":work_date .isoformat ()if work_date else None ,
            "work_order":line .entry .work_order_id ,
            "pdf_name":pdf_label ,
            "code":line .machine_asset .code ,
            "brand_model":line .machine_asset .brand_model ,
            "delta_hours":delta ,
            "weekday":work_date .weekday ()if work_date else None ,
            })







        wo_qs =(
        WorkOrder .objects 
        .filter (company =company ,status =WorkOrder .Status .DONE )
        .order_by ("id")
        )
        work_orders =[]
        for wo in wo_qs :
            raw =wo .source_pdf .name .split ("/")[-1 ]
            label =_re .sub (r'_[A-Za-z0-9]{7}(\.[^.]+)$',r'',raw )
            work_orders .append ({"id":wo .pk ,"label":label })



        seen_assets :dict [str ,str ]={}
        for line in lines :
            c =line ["code"]
            if c not in seen_assets :
                seen_assets [c ]=line ["brand_model"]
        assets =[
        {"code":c ,"brand_model":m }
        for c ,m in sorted (seen_assets .items ())
        ]

        return JsonResponse ({
        "lines":lines ,
        "work_orders":work_orders ,
        "assets":assets ,
        })






class AnalyticsLabView (AdminRoleRequiredMixin ,View ):
    """
    Renders the Unified Analytics Laboratory shell for the authenticated user's
    company. Passes selector data (operators, machines, fault categories, default
    date range) to the template. All chart and table data is fetched client-side
    via AnalyticsLabDataView (/panel/analytics/lab/data/).

    Only dimensions that have data available for the company are passed in
    context so the dimension selector never shows empty options.

    ---

    Renderiza el shell del Laboratorio de Análisis Unificado para la empresa
    del usuario autenticado. Pasa al template los datos de los selectores
    (operarios, máquinas, familias de avería, rango de fechas por defecto).
    Todos los datos de gráficos y tablas se obtienen en el cliente mediante
    AnalyticsLabDataView (/panel/analytics/lab/data/).

    Solo se pasan en contexto las dimensiones que tienen datos disponibles para
    la empresa, de modo que el selector de dimensión nunca muestra opciones vacías.
    """

    template_name ="panel/analytics_lab.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Builds context and renders the lab template. Determines which dimensions
        have data (count > 0) so only populated dimensions appear in the selector.
        Date defaults: date_from = first day of current month, date_to = today.
        ---
        Construye el contexto y renderiza el template del laboratorio. Determina
        qué dimensiones tienen datos (count > 0) para que solo aparezcan en el
        selector las dimensiones con contenido.
        Fechas por defecto: date_from = primer día del mes actual, date_to = hoy.
        """
        from django .utils .timezone import localdate 
        from budgets .models import Budget 

        company_user =request .user .company_user 
        company =company_user .company 





        today =localdate ()
        date_from_def =today .replace (day =1 )
        date_to_def =today 





        has_d1 =(
        WorkOrderEntry .objects 
        .filter (
        work_order__company =company ,
        worker_name__gt ="",
        )
        .values ("worker_name")
        .distinct ()
        .count ()
        )>0 

        has_d2 =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        machine_asset__isnull =False ,
        )
        .values ("machine_asset")
        .distinct ()
        .count ()
        )>0 

        has_d3 =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        fault_category__gt ="",
        )
        .values ("fault_category")
        .distinct ()
        .count ()
        )>0 

        has_d4 =(
        WorkOrderEntry .objects 
        .filter (
        work_order__company =company ,
        work_date__isnull =False ,
        )
        .exists ()
        )

        has_d5 =Budget .objects .filter (company =company ).count ()>0 





        operators =[]
        if has_d1 :
            operators =list (
            WorkOrderEntry .objects 
            .filter (work_order__company =company ,worker_name__gt ="")
            .values_list ("worker_name",flat =True )
            .distinct ()
            .order_by ("worker_name")
            )

        machines =[]
        if has_d2 :
            machines =list (
            MachineAsset .objects 
            .filter (company =company )
            .values ("pk","code","brand_model")
            .order_by ("code")
            )

        fault_categories =[]
        if has_d3 :






            _fault_cat_map ={
            "TYRES_RUNNING_GEAR":"Neumáticos y rodadura",
            "BRAKES_STEERING_SUSPENSION":"Frenos, dirección y suspensión",
            "HYDRAULIC":"Hidráulica",
            "BODYWORK_CHASSIS":"Carrocería y chasis",
            "ENGINE_TRANSMISSION":"Motor y transmisión",
            "LIFTING_STRUCTURE":"Estructura de elevación",
            "ELECTRICAL_ELECTRONIC":"Eléctrico y electrónico",
            "OTHER":"Otros",
            }
            _raw_keys =list (
            WorkOrderEntryLine .objects 
            .filter (
            entry__work_order__company =company ,
            fault_category__gt ="",
            )
            .values_list ("fault_category",flat =True )
            .distinct ()
            .order_by ("fault_category")
            )
            fault_categories =[
            {"key":k ,"label":_fault_cat_map .get (k ,k )}
            for k in _raw_keys 
            ]

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"analytics_lab",
        "date_from_default":date_from_def .isoformat (),
        "date_to_default":date_to_def .isoformat (),
        "has_d1":has_d1 ,
        "has_d2":has_d2 ,
        "has_d3":has_d3 ,
        "has_d4":has_d4 ,
        "has_d5":has_d5 ,
        "operators":operators ,
        "machines":machines ,
        "fault_categories":fault_categories ,
        })


class AnalyticsLabDataView (AdminRoleRequiredMixin ,View ):
    """
    JSON endpoint for the Unified Analytics Laboratory. Receives dimension,
    entity_pk, date_from, date_to, granularity and chart_type parameters and
    returns a structured JSON payload consumed by the ECharts frontend.

    GET /panel/analytics/lab/data/
        Parameters:
          dimension   (str)  — d1 | d2 | d3 | d4 | d5
          entity_pk   (int)  — pk of operator (worker_name), machine, fault
                               category; not required for d4/d5.
          date_from   (str)  — YYYY-MM-DD
          date_to     (str)  — YYYY-MM-DD
          granularity (str)  — day | week | month  (default: month)
          chart_type  (str)  — bar | line | scatter | pie | heatmap | treemap
                               (default depends on dimension)

    Response schema:
    {
        "ok":      bool,
        "chart": {
            "type":   str,
            "title":  str,
            "xAxis":  [str, ...],
            "series": [{"name": str, "data": [float|int, ...]}, ...]
        },
        "table": {
            "columns": [str, ...],
            "rows":    [[val, ...], ...]
        },
        "summary": {
            "total_hours":        float,
            "total_parts":        int,
            "avg_hours_per_part": float
        }
    }

    ---

    Endpoint JSON para el Laboratorio de Análisis Unificado. Recibe los parámetros
    dimension, entity_pk, date_from, date_to, granularity y chart_type y devuelve
    un payload JSON estructurado consumido por el frontend ECharts.

    GET /panel/analytics/lab/data/
        Parámetros:
          dimension   (str)  — d1 | d2 | d3 | d4 | d5
          entity_pk   (int)  — pk de operario (worker_name), máquina o familia;
                               no requerido para d4/d5.
          date_from   (str)  — YYYY-MM-DD
          date_to     (str)  — YYYY-MM-DD
          granularity (str)  — day | week | month  (por defecto: month)
          chart_type  (str)  — bar | line | scatter | pie | heatmap | treemap
                               (por defecto según dimensión)

    Esquema de respuesta:
    {
        "ok":      bool,
        "chart": {
            "type":   str,
            "title":  str,
            "xAxis":  [str, ...],
            "series": [{"name": str, "data": [float|int, ...]}, ...]
        },
        "table": {
            "columns": [str, ...],
            "rows":    [[val, ...], ...]
        },
        "summary": {
            "total_hours":        float,
            "total_parts":        int,
            "avg_hours_per_part": float
        }
    }
    """



    _DEFAULT_CHART ={
    "d1":"bar",
    "d2":"bar",
    "d3":"bar",
    "d4":"bar",
    "d5":"bar",
    }







    _FAULT_CAT_LABELS ={
    "TYRES_RUNNING_GEAR":"Neumáticos y rodadura",
    "BRAKES_STEERING_SUSPENSION":"Frenos, dirección y suspensión",
    "HYDRAULIC":"Hidráulica",
    "BODYWORK_CHASSIS":"Carrocería y chasis",
    "ENGINE_TRANSMISSION":"Motor y transmisión",
    "LIFTING_STRUCTURE":"Estructura de elevación",
    "ELECTRICAL_ELECTRONIC":"Eléctrico y electrónico",
    "OTHER":"Otros",
    }

    def _translate_fault_cat (self ,key ):
        """
        Returns the Spanish display label for a FaultCategory internal key.
        Falls back to the raw key if no translation is found, so unknown
        future values never break the UI.
        ---
        Devuelve la etiqueta en castellano para una clave interna de
        FaultCategory. Retrocede a la clave cruda si no encuentra traducción,
        para que valores futuros desconocidos no rompan la UI.
        """
        return self ._FAULT_CAT_LABELS .get (key ,key )





    @staticmethod 
    def _bucket (work_date ,granularity ):
        """
        Returns a string label representing the time bucket for work_date
        given the requested granularity (day / week / month).
        ---
        Devuelve una etiqueta de cadena representando el bucket temporal
        para work_date según la granularidad solicitada (day / week / month).
        """
        if granularity =="day":
            return work_date .strftime ("%Y-%m-%d")
        elif granularity =="week":


            return work_date .strftime ("%G-W%V")
        else :


            return work_date .strftime ("%Y-%m")





    @staticmethod 
    def _parse_date (value ,fallback ):
        """
        Parses a YYYY-MM-DD string into a date object. Returns fallback on
        any parsing error.
        ---
        Parsea una cadena YYYY-MM-DD en un objeto date. Devuelve fallback
        ante cualquier error de parseo.
        """
        from datetime import date as _date ,datetime as _datetime 
        try :
            return _datetime .strptime (value .strip (),"%Y-%m-%d").date ()
        except (ValueError ,AttributeError ):
            return fallback 






    def _handle_d1 (self ,company ,entity_pk ,date_from ,date_to ,
    granularity ,chart_type ):
        """
        D1 — Operator analysis.
        Metrics: worked hours per bucket, number of entries, parts per machine,
        dominant fault categories, time evolution, hours/part ratio, worked days.
        ---
        D1 — Análisis de operario.
        Métricas: horas trabajadas por bucket, número de entradas, partes por
        máquina, familias de avería más frecuentes, evolución temporal,
        ratio horas/parte, días trabajados.
        """
        from django .db .models import Sum ,Count 
        from decimal import Decimal 



        worker_name =None 
        if entity_pk :
            try :
                pk_int =int (entity_pk )




            except (ValueError ,TypeError ):
                pass 


            worker_name =str (entity_pk )



        qs =(
        WorkOrderEntry .objects 
        .filter (
        work_order__company =company ,
        work_date__gte =date_from ,
        work_date__lte =date_to ,
        work_date__isnull =False ,
        )
        )
        if worker_name :
            qs =qs .filter (worker_name =worker_name )





        entry_pks =list (qs .values_list ("pk",flat =True ))
        lines_qs =(
        WorkOrderEntryLine .objects 
        .filter (entry__pk__in =entry_pks ,delta_hours__isnull =False )
        )



        from collections import defaultdict 
        bucket_hours =defaultdict (float )
        bucket_parts =defaultdict (int )
        bucket_wo_sets =defaultdict (set )

        for line in lines_qs .select_related ("entry"):
            b =self ._bucket (line .entry .work_date ,granularity )
            bucket_hours [b ]+=float (line .delta_hours or 0 )
            bucket_wo_sets [b ].add (line .entry .work_order_id )

        for entry in qs :
            b =self ._bucket (entry .work_date ,granularity )
            bucket_parts [b ]+=1 



        all_buckets =sorted (
        set (bucket_hours .keys ())|set (bucket_parts .keys ())
        )

        hours_series =[round (bucket_hours .get (b ,0.0 ),2 )for b in all_buckets ]
        parts_series =[bucket_parts .get (b ,0 )for b in all_buckets ]





        if chart_type =="heatmap":


            from itertools import product 

            operators_in_range =list (
            WorkOrderEntry .objects 
            .filter (
            work_order__company =company ,
            work_date__gte =date_from ,
            work_date__lte =date_to ,
            worker_name__gt ="",
            )
            .values_list ("worker_name",flat =True )
            .distinct ()
            .order_by ("worker_name")
            )
            faults_in_range =[
            self ._translate_fault_cat (k )
            for k in sorted (
            WorkOrderEntryLine .objects 
            .filter (
            entry__work_order__company =company ,
            entry__work_date__gte =date_from ,
            entry__work_date__lte =date_to ,
            fault_category__gt ="",
            )
            .values_list ("fault_category",flat =True )
            .distinct ()
            )
            ]



            heat_data =defaultdict (int )
            for line in (
            WorkOrderEntryLine .objects 
            .filter (
            entry__work_order__company =company ,
            entry__work_date__gte =date_from ,
            entry__work_date__lte =date_to ,
            fault_category__gt ="",
            )
            .select_related ("entry")
            ):
                op =line .entry .worker_name 
                fc =self ._translate_fault_cat (line .fault_category )
                if op and fc :
                    heat_data [(op ,fc )]+=1 



            hm_points =[]
            for xi ,op in enumerate (operators_in_range ):
                for yi ,fc in enumerate (faults_in_range ):
                    hm_points .append ([xi ,yi ,heat_data .get ((op ,fc ),0 )])


            total_hours =sum (
            float (l .delta_hours or 0 )
            for l in WorkOrderEntryLine .objects .filter (
            entry__pk__in =entry_pks ,delta_hours__isnull =False 
            )
            )
            total_parts =qs .count ()

            table_rows =[
            [op ,fc ,heat_data .get ((op ,fc ),0 )]
            for (op ,fc )in sorted (heat_data .keys ())
            if heat_data [(op ,fc )]>0 
            ]

            return {
            "ok":True ,
            "chart":{
            "type":"heatmap",
            "title":f"Heatmap Operario × Familia ({date_from} / {date_to})",
            "xAxis":operators_in_range ,
            "yAxis":faults_in_range ,
            "series":[{"name":"Intervenciones","data":hm_points }],
            "visualMap":{
            "min":0 ,
            "max":max ((p [2 ]for p in hm_points ),default =1 ),
            },
            },
            "table":{
            "columns":["Operario","Familia de avería","Intervenciones"],
            "rows":table_rows ,
            },
            "summary":{
            "total_hours":round (total_hours ,2 ),
            "total_parts":total_parts ,
            "avg_hours_per_part":round (total_hours /total_parts ,2 )if total_parts else 0.0 ,
            },
            }





        title =(
        f"Operario: {worker_name or 'Todos'} — Horas trabajadas "
        f"({date_from} / {date_to})"
        )
        total_hours =sum (hours_series )
        total_parts =sum (parts_series )

        table_rows =[
        [b ,round (bucket_hours .get (b ,0.0 ),2 ),bucket_parts .get (b ,0 )]
        for b in all_buckets 
        ]

        return {
        "ok":True ,
        "chart":{
        "type":chart_type if chart_type in ("bar","line")else "bar",
        "title":title ,
        "xAxis":all_buckets ,
        "series":[
        {"name":"Horas trabajadas","data":hours_series },
        {"name":"Nº entradas","data":parts_series },
        ],
        },
        "table":{
        "columns":["Periodo","Horas trabajadas","Nº entradas"],
        "rows":table_rows ,
        },
        "summary":{
        "total_hours":round (total_hours ,2 ),
        "total_parts":total_parts ,
        "avg_hours_per_part":round (total_hours /total_parts ,2 )if total_parts else 0.0 ,
        },
        }

    def _handle_d2 (self ,company ,entity_pk ,date_from ,date_to ,
    granularity ,chart_type ):
        """
        D2 — Machine / Cost Centre analysis.
        Metrics: accumulated labour hours, number of interventions, dominant
        fault categories, most frequent operators, cost in hours per bucket,
        approximate MTBF, time evolution.
        ---
        D2 — Análisis de Máquina / Centro de Gasto.
        Métricas: horas de mano de obra acumuladas, número de intervenciones,
        familias de avería dominantes, operarios que más intervienen, coste en
        horas por bucket, MTBF aproximado, evolución temporal.
        """
        from django .db .models import Sum ,Count 
        from collections import defaultdict 



        machine =None 
        if entity_pk :
            try :
                machine =MachineAsset .objects .get (pk =int (entity_pk ),company =company )
            except (ValueError ,TypeError ,MachineAsset .DoesNotExist ):
                pass 



        lines_qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        entry__work_date__gte =date_from ,
        entry__work_date__lte =date_to ,
        entry__work_date__isnull =False ,
        machine_asset__isnull =False ,
        )
        .select_related ("entry","machine_asset")
        )
        if machine :
            lines_qs =lines_qs .filter (machine_asset =machine )





        if chart_type =="scatter":
            scatter_data =defaultdict (lambda :{"hours":0.0 ,"count":0 ,"code":""})
            for line in lines_qs :
                pk_m =line .machine_asset .pk 
                scatter_data [pk_m ]["hours"]+=float (line .delta_hours or 0 )
                scatter_data [pk_m ]["count"]+=1 
                scatter_data [pk_m ]["code"]=line .machine_asset .code 

            scatter_points =[
            [round (v ["hours"],2 ),v ["count"],v ["code"]]
            for v in scatter_data .values ()
            ]
            total_hours =sum (p [0 ]for p in scatter_points )
            total_parts =sum (p [1 ]for p in scatter_points )

            return {
            "ok":True ,
            "chart":{
            "type":"scatter",
            "title":f"Máquinas — Horas vs Intervenciones ({date_from} / {date_to})",
            "xAxis":[],
            "series":[{"name":"Horas vs Intervenciones","data":scatter_points }],
            },
            "table":{
            "columns":["Código máquina","Horas acumuladas","Intervenciones"],
            "rows":[
            [v ["code"],round (v ["hours"],2 ),v ["count"]]
            for v in sorted (scatter_data .values (),key =lambda x :x ["hours"],reverse =True )
            ],
            },
            "summary":{
            "total_hours":round (total_hours ,2 ),
            "total_parts":total_parts ,
            "avg_hours_per_part":round (total_hours /total_parts ,2 )if total_parts else 0.0 ,
            },
            }





        if chart_type =="pie":
            fault_counts =defaultdict (int )
            for line in lines_qs .filter (fault_category__gt =""):
                fault_counts [self ._translate_fault_cat (line .fault_category )]+=1 

            pie_data =[
            {"name":fc ,"value":cnt }
            for fc ,cnt in sorted (fault_counts .items (),key =lambda x :x [1 ],reverse =True )
            ]
            total_parts =sum (d ["value"]for d in pie_data )

            return {
            "ok":True ,
            "chart":{
            "type":"pie",
            "title":f"Familias de avería — {machine.code if machine else 'Todas las máquinas'} ({date_from} / {date_to})",
            "xAxis":[],
            "series":[{"name":"Intervenciones","data":pie_data }],
            },
            "table":{
            "columns":["Familia de avería","Intervenciones"],
            "rows":[[d ["name"],d ["value"]]for d in pie_data ],
            },
            "summary":{
            "total_hours":0.0 ,
            "total_parts":total_parts ,
            "avg_hours_per_part":0.0 ,
            },
            }





        bucket_hours =defaultdict (float )
        bucket_counts =defaultdict (int )

        for line in lines_qs :
            b =self ._bucket (line .entry .work_date ,granularity )
            bucket_hours [b ]+=float (line .delta_hours or 0 )
            bucket_counts [b ]+=1 

        all_buckets =sorted (set (bucket_hours .keys ())|set (bucket_counts .keys ()))
        hours_series =[round (bucket_hours .get (b ,0.0 ),2 )for b in all_buckets ]
        count_series =[bucket_counts .get (b ,0 )for b in all_buckets ]

        label =machine .code if machine else "Todas las máquinas"
        title =f"Máquina: {label} — Horas de mano de obra ({date_from} / {date_to})"

        total_hours =sum (hours_series )
        total_parts =sum (count_series )

        return {
        "ok":True ,
        "chart":{
        "type":chart_type if chart_type in ("bar","line")else "bar",
        "title":title ,
        "xAxis":all_buckets ,
        "series":[
        {"name":"Horas M.O.","data":hours_series },
        {"name":"Intervenciones","data":count_series },
        ],
        },
        "table":{
        "columns":["Periodo","Horas M.O.","Intervenciones"],
        "rows":[
        [b ,round (bucket_hours .get (b ,0.0 ),2 ),bucket_counts .get (b ,0 )]
        for b in all_buckets 
        ],
        },
        "summary":{
        "total_hours":round (total_hours ,2 ),
        "total_parts":total_parts ,
        "avg_hours_per_part":round (total_hours /total_parts ,2 )if total_parts else 0.0 ,
        },
        }

    def _handle_d3 (self ,company ,entity_pk ,date_from ,date_to ,
    granularity ,chart_type ):
        """
        D3 — Fault Category analysis.
        Metrics: frequency per bucket, most affected machines, operators that
        handle it most, average hours per intervention, time evolution,
        distribution by machine.
        ---
        D3 — Análisis de Familia de Avería.
        Métricas: frecuencia por bucket, máquinas más afectadas, operarios
        que más la atienden, horas medias por intervención, evolución temporal,
        distribución por máquina.
        """
        from collections import defaultdict 

        fault_cat =str (entity_pk )if entity_pk else None 



        lines_qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        entry__work_date__gte =date_from ,
        entry__work_date__lte =date_to ,
        entry__work_date__isnull =False ,
        fault_category__gt ="",
        )
        .select_related ("entry","entry__work_order","machine_asset")
        )
        if fault_cat :
            lines_qs =lines_qs .filter (fault_category =fault_cat )





        if chart_type in ("bar","stacked"):
            cat_bucket =defaultdict (lambda :defaultdict (int ))
            all_buckets_set =set ()

            for line in lines_qs :
                b =self ._bucket (line .entry .work_date ,granularity )
                fc =self ._translate_fault_cat (line .fault_category )
                cat_bucket [fc ][b ]+=1 
                all_buckets_set .add (b )

            all_buckets =sorted (all_buckets_set )
            all_cats =sorted (cat_bucket .keys ())
            series_list =[
            {
            "name":fc ,
            "data":[cat_bucket [fc ].get (b ,0 )for b in all_buckets ],
            "stack":"averia",
            }
            for fc in all_cats 
            ]
            total_parts =sum (
            sum (d ["data"])for d in series_list 
            )



            table_rows =[
            [fc ]+[cat_bucket [fc ].get (b ,0 )for b in all_buckets ]
            for fc in all_cats 
            ]
            table_cols =["Familia"]+all_buckets 

            return {
            "ok":True ,
            "chart":{
            "type":"bar",
            "title":f"Familias de avería — Distribución temporal ({date_from} / {date_to})",
            "xAxis":all_buckets ,
            "series":series_list ,
            },
            "table":{
            "columns":table_cols ,
            "rows":table_rows ,
            },
            "summary":{
            "total_hours":0.0 ,
            "total_parts":total_parts ,
            "avg_hours_per_part":0.0 ,
            },
            }





        if chart_type =="treemap":
            machine_fault =defaultdict (lambda :defaultdict (int ))
            for line in lines_qs :
                m_code =line .machine_asset .code if line .machine_asset else "Sin máquina"
                fc =self ._translate_fault_cat (line .fault_category )
                machine_fault [m_code ][fc ]+=1 

            treemap_data =[]
            for m_code ,faults in sorted (machine_fault .items ()):
                children =[
                {"name":fc ,"value":cnt }
                for fc ,cnt in sorted (faults .items (),key =lambda x :x [1 ],reverse =True )
                ]
                treemap_data .append ({
                "name":m_code ,
                "value":sum (c ["value"]for c in children ),
                "children":children ,
                })

            total_parts =sum (d ["value"]for d in treemap_data )

            table_rows =[
            [m_code ,fc ,machine_fault [m_code ][fc ]]
            for m_code in sorted (machine_fault .keys ())
            for fc in sorted (machine_fault [m_code ].keys ())
            if machine_fault [m_code ][fc ]>0 
            ]

            return {
            "ok":True ,
            "chart":{
            "type":"treemap",
            "title":f"Distribución por máquina — {self._translate_fault_cat(fault_cat) if fault_cat else 'Todas las familias'} ({date_from} / {date_to})",
            "xAxis":[],
            "series":[{"name":"Intervenciones","data":treemap_data }],
            },
            "table":{
            "columns":["Máquina","Familia de avería","Intervenciones"],
            "rows":table_rows ,
            },
            "summary":{
            "total_hours":0.0 ,
            "total_parts":total_parts ,
            "avg_hours_per_part":0.0 ,
            },
            }





        bucket_counts =defaultdict (int )
        for line in lines_qs :
            b =self ._bucket (line .entry .work_date ,granularity )
            bucket_counts [b ]+=1 

        all_buckets =sorted (bucket_counts .keys ())
        count_series =[bucket_counts .get (b ,0 )for b in all_buckets ]
        total_parts =sum (count_series )

        return {
        "ok":True ,
        "chart":{
        "type":"line",
        "title":f"Evolución temporal — {self._translate_fault_cat(fault_cat) if fault_cat else 'Todas las familias'} ({date_from} / {date_to})",
        "xAxis":all_buckets ,
        "series":[{"name":self ._translate_fault_cat (fault_cat )if fault_cat else "Todas las familias","data":count_series }],
        },
        "table":{
        "columns":["Periodo","Intervenciones"],
        "rows":[[b ,bucket_counts .get (b ,0 )]for b in all_buckets ],
        },
        "summary":{
        "total_hours":0.0 ,
        "total_parts":total_parts ,
        "avg_hours_per_part":0.0 ,
        },
        }

    def _handle_d4 (self ,company ,entity_pk ,date_from ,date_to ,
    granularity ,chart_type ):
        """
        D4 — Time Period cross-analysis.
        Metrics: total hours, processed entries, distribution by operator,
        distribution by fault family, top intervened machines, cross-period
        comparison.
        ---
        D4 — Análisis de Periodo Temporal cruzado.
        Métricas: horas totales, entradas procesadas, distribución por operario,
        distribución por familia, top máquinas intervenidas, comparativa entre
        periodos.
        """
        from collections import defaultdict 



        entries_qs =(
        WorkOrderEntry .objects 
        .filter (
        work_order__company =company ,
        work_date__gte =date_from ,
        work_date__lte =date_to ,
        work_date__isnull =False ,
        )
        )
        lines_qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        entry__work_date__gte =date_from ,
        entry__work_date__lte =date_to ,
        entry__work_date__isnull =False ,
        )
        .select_related ("entry","machine_asset")
        )





        op_bucket_hours =defaultdict (lambda :defaultdict (float ))
        all_buckets_set =set ()
        all_ops_set =set ()

        for line in lines_qs .filter (delta_hours__isnull =False ):
            b =self ._bucket (line .entry .work_date ,granularity )
            op =line .entry .worker_name or "Sin operario"
            op_bucket_hours [op ][b ]+=float (line .delta_hours or 0 )
            all_buckets_set .add (b )
            all_ops_set .add (op )

        all_buckets =sorted (all_buckets_set )
        all_ops =sorted (all_ops_set )

        series_list =[
        {
        "name":op ,
        "data":[round (op_bucket_hours [op ].get (b ,0.0 ),2 )for b in all_buckets ],
        }
        for op in all_ops 
        ]

        total_hours =sum (
        v 
        for op_data in op_bucket_hours .values ()
        for v in op_data .values ()
        )
        total_parts =entries_qs .count ()



        table_rows =[
        [op ]+[round (op_bucket_hours [op ].get (b ,0.0 ),2 )for b in all_buckets ]
        for op in all_ops 
        ]
        table_cols =["Operario"]+all_buckets 

        return {
        "ok":True ,
        "chart":{
        "type":chart_type if chart_type in ("bar","line")else "bar",
        "title":f"Resumen de periodo: {date_from} / {date_to}",
        "xAxis":all_buckets ,
        "series":series_list ,
        },
        "table":{
        "columns":table_cols ,
        "rows":table_rows ,
        },
        "summary":{
        "total_hours":round (total_hours ,2 ),
        "total_parts":total_parts ,
        "avg_hours_per_part":round (total_hours /total_parts ,2 )if total_parts else 0.0 ,
        },
        }

    def _handle_d5 (self ,company ,entity_pk ,date_from ,date_to ,
    granularity ,chart_type ):
        """
        D5 — Budget / Assistance analysis.
        Metrics: budgets per insurer, average amounts, service distribution,
        temporal evolution. Only available when Budget records exist for the
        company.
        ---
        D5 — Análisis de Presupuestos / Asistencia.
        Métricas: presupuestos por aseguradora, importes medios, distribución
        de servicios, evolución temporal. Solo disponible si existen registros
        Budget para la empresa.
        """
        from collections import defaultdict 
        from budgets .models import Budget 



        budgets_qs =(
        Budget .objects 
        .filter (
        company =company ,
        service_date__gte =date_from ,
        service_date__lte =date_to ,
        service_date__isnull =False ,
        )
        .select_related ("insurer")
        )





        if chart_type in ("bar",):
            insurer_data =defaultdict (lambda :{"amount":0.0 ,"count":0 })
            for budget in budgets_qs :
                label =budget .insurer .name if budget .insurer else "Sin aseguradora"
                insurer_data [label ]["amount"]+=float (budget .total_amount or 0 )
                insurer_data [label ]["count"]+=1 

            sorted_insurers =sorted (insurer_data .items (),key =lambda x :x [1 ]["amount"],reverse =True )
            x_labels =[k for k ,_ in sorted_insurers ]
            amount_series =[round (v ["amount"],2 )for _ ,v in sorted_insurers ]
            count_series =[v ["count"]for _ ,v in sorted_insurers ]

            total_amount =sum (amount_series )
            total_parts =sum (count_series )

            table_rows =[
            [label ,round (v ["amount"],2 ),v ["count"],
            round (v ["amount"]/v ["count"],2 )if v ["count"]else 0.0 ]
            for label ,v in sorted_insurers 
            ]

            return {
            "ok":True ,
            "chart":{
            "type":"bar",
            "title":f"Presupuestos por aseguradora ({date_from} / {date_to})",
            "xAxis":x_labels ,
            "series":[
            {"name":"Importe total (€)","data":amount_series },
            {"name":"Nº presupuestos","data":count_series },
            ],
            },
            "table":{
            "columns":["Aseguradora","Importe total (€)","Nº presupuestos","Importe medio (€)"],
            "rows":table_rows ,
            },
            "summary":{
            "total_hours":0.0 ,
            "total_parts":total_parts ,
            "avg_hours_per_part":round (total_amount /total_parts ,2 )if total_parts else 0.0 ,
            },
            }





        bucket_amounts =defaultdict (float )
        bucket_counts =defaultdict (int )

        for budget in budgets_qs :
            b =self ._bucket (budget .service_date ,granularity )
            bucket_amounts [b ]+=float (budget .total_amount or 0 )
            bucket_counts [b ]+=1 

        all_buckets =sorted (set (bucket_amounts .keys ())|set (bucket_counts .keys ()))
        total_parts =sum (bucket_counts .values ())
        total_amount =sum (bucket_amounts .values ())

        return {
        "ok":True ,
        "chart":{
        "type":"line",
        "title":f"Evolución temporal de presupuestos ({date_from} / {date_to})",
        "xAxis":all_buckets ,
        "series":[
        {"name":"Importe (€)","data":[round (bucket_amounts .get (b ,0.0 ),2 )for b in all_buckets ]},
        {"name":"Nº presupuestos","data":[bucket_counts .get (b ,0 )for b in all_buckets ]},
        ],
        },
        "table":{
        "columns":["Periodo","Importe (€)","Nº presupuestos"],
        "rows":[
        [b ,round (bucket_amounts .get (b ,0.0 ),2 ),bucket_counts .get (b ,0 )]
        for b in all_buckets 
        ],
        },
        "summary":{
        "total_hours":0.0 ,
        "total_parts":total_parts ,
        "avg_hours_per_part":round (total_amount /total_parts ,2 )if total_parts else 0.0 ,
        },
        }






    def get (self ,request ):
        """
        Dispatches to the appropriate dimension handler and returns the
        structured JSON payload. Returns HTTP 400 on missing required
        parameters and HTTP 403 on missing CompanyUser profile.
        ---
        Despacha al manejador de dimensión correspondiente y devuelve el
        payload JSON estructurado. Retorna HTTP 400 ante parámetros
        requeridos faltantes y HTTP 403 ante perfil CompanyUser ausente.
        """
        from django .http import JsonResponse 
        from django .utils .timezone import localdate 
        from datetime import timedelta 

        try :
            company =request .user .company_user .company 
        except AttributeError :
            return JsonResponse ({"ok":False ,"error":"Sin perfil de empresa asociado."},status =403 )





        dimension =request .GET .get ("dimension","").strip ().lower ()
        entity_pk =request .GET .get ("entity_pk","").strip ()or None 
        granularity =request .GET .get ("granularity","month").strip ().lower ()
        chart_type =request .GET .get ("chart_type","").strip ().lower ()



        if dimension not in ("d1","d2","d3","d4","d5"):
            return JsonResponse (
            {"ok":False ,"error":f"Dimensión no válida: '{dimension}'."},
            status =400 ,
            )



        if granularity not in ("day","week","month"):
            granularity ="month"



        if not chart_type :
            chart_type =self ._DEFAULT_CHART .get (dimension ,"bar")



        today =localdate ()
        fallback_from =today .replace (day =1 )
        fallback_to =today 

        raw_from =request .GET .get ("date_from","").strip ()
        raw_to =request .GET .get ("date_to","").strip ()
        date_from =self ._parse_date (raw_from ,fallback_from )
        date_to =self ._parse_date (raw_to ,fallback_to )

        if date_from >date_to :
            date_from ,date_to =date_to ,date_from 





        handlers ={
        "d1":self ._handle_d1 ,
        "d2":self ._handle_d2 ,
        "d3":self ._handle_d3 ,
        "d4":self ._handle_d4 ,
        "d5":self ._handle_d5 ,
        }
        try :
            result =handlers [dimension ](
            company ,entity_pk ,date_from ,date_to ,granularity ,chart_type 
            )
        except Exception as exc :
            logger .exception (
            "# [ANALYTICS LAB DATA] Error en dimension=%s entity_pk=%s: %s",
            dimension ,entity_pk ,exc ,
            )
            return JsonResponse (
            {"ok":False ,"error":f"Error interno al procesar la dimensión {dimension}."},
            status =500 ,
            )

        return JsonResponse (result )


class AnalyticsLabExportView (AdminRoleRequiredMixin ,View ):
    """
    Generates and streams an Excel file from the table data of the Analytics
    Laboratory. The POST body is expected to contain JSON-encoded 'columns'
    and 'rows' fields matching the table payload returned by AnalyticsLabDataView.

    POST /panel/analytics/lab/export/
         Body params (form-encoded):
           dimension  (str)       — used for filename.
           date_from  (str)       — used for filename.
           date_to    (str)       — used for filename.
           columns    (JSON str)  — list of column header strings.
           rows       (JSON str)  — list of row value lists.

    Returns HttpResponse with Content-Disposition attachment (.xlsx).
    Returns HTTP 400 on missing or malformed payload.
    ---
    Genera y devuelve en streaming un Excel desde los datos de tabla del
    Laboratorio de Análisis. El cuerpo del POST debe contener los campos
    JSON 'columns' y 'rows' que coinciden con el payload de tabla devuelto
    por AnalyticsLabDataView.

    POST /panel/analytics/lab/export/
         Parámetros del cuerpo (form-encoded):
           dimension  (str)       — usado para el nombre del archivo.
           date_from  (str)       — usado para el nombre del archivo.
           date_to    (str)       — usado para el nombre del archivo.
           columns    (JSON str)  — lista de cadenas de cabecera de columna.
           rows       (JSON str)  — lista de listas de valores de fila.

    Devuelve HttpResponse con Content-Disposition attachment (.xlsx).
    Devuelve HTTP 400 ante payload ausente o malformado.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Parses the JSON-encoded columns and rows from the POST body,
        builds an openpyxl workbook and streams it as an xlsx attachment.
        ---
        Parsea columns y rows codificados en JSON del cuerpo POST,
        construye un libro openpyxl y lo devuelve en streaming como
        adjunto xlsx.
        """
        import io 
        import json as _json 
        import openpyxl 
        from openpyxl .styles import Font ,PatternFill ,Alignment 
        from django .http import HttpResponse ,HttpResponseBadRequest 





        try :
            columns =_json .loads (request .POST .get ("columns","[]"))
            rows =_json .loads (request .POST .get ("rows","[]"))
            dimension =request .POST .get ("dimension","lab").strip ()
            date_from =request .POST .get ("date_from","").strip ()
            date_to =request .POST .get ("date_to","").strip ()
        except (_json .JSONDecodeError ,TypeError )as exc :
            logger .warning ("# [ANALYTICS LAB EXPORT] Payload inválido: %s",exc )
            return HttpResponseBadRequest ("# [ANALYTICS LAB EXPORT] Payload JSON inválido.")

        if not columns or not isinstance (rows ,list ):
            return HttpResponseBadRequest (
            "# [ANALYTICS LAB EXPORT] Se requieren 'columns' y 'rows' en el cuerpo."
            )





        wb =openpyxl .Workbook ()
        ws =wb .active 
        ws .title ="Laboratorio"



        header_font =Font (bold =True ,color ="FFFFFF")
        header_fill =PatternFill (fill_type ="solid",fgColor ="2C3E50")
        header_align =Alignment (horizontal ="center",vertical ="center",wrap_text =True )

        for col_idx ,col_header in enumerate (columns ,start =1 ):
            cell =ws .cell (row =1 ,column =col_idx ,value =str (col_header ))
            cell .font =header_font 
            cell .fill =header_fill 
            cell .alignment =header_align 



        for row_idx ,row in enumerate (rows ,start =2 ):
            for col_idx ,cell_value in enumerate (row ,start =1 ):


                if hasattr (cell_value ,"__float__"):
                    cell_value =float (cell_value )
                ws .cell (row =row_idx ,column =col_idx ,value =cell_value )



        for col_cells in ws .columns :
            max_length =max (
            (len (str (c .value ))if c .value is not None else 0 for c in col_cells ),
            default =8 ,
            )
            ws .column_dimensions [col_cells [0 ].column_letter ].width =min (max_length +4 ,60 )





        buf =io .BytesIO ()
        wb .save (buf )
        buf .seek (0 )

        filename =f"lab_{dimension}_{date_from}_{date_to}.xlsx"
        response =HttpResponse (
        buf .getvalue (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        logger .info (
        "# [ANALYTICS LAB EXPORT] Exportación dimension=%s (%d filas) por %s.",
        dimension ,len (rows ),request .user .username ,
        )
        return response 


class AnalyticsProfileListCreateView (SupervisorAccessMixin ,View ):
    """
    JSON endpoint for listing and creating/updating AnalyticsProfile records
    belonging to the authenticated CompanyUser.

    GET  /panel/analytics/profiles/
         Returns a JSON array of all profiles for the current user, ordered
         by name. Each item exposes: id, nombre, config.

    POST /panel/analytics/profiles/
         Body (JSON): {"nombre": str, "config": {...}}
         Creates a new profile or updates an existing one with the same name
         (upsert semantics — unique_together enforced at model level).
         Returns the saved profile as JSON with HTTP 200.
         Returns HTTP 400 on missing or invalid payload.
    ---
    Endpoint JSON para listar y crear/actualizar registros AnalyticsProfile
    del CompanyUser autenticado.

    GET  /panel/analytics/profiles/
         Devuelve un array JSON con todos los perfiles del usuario actual,
         ordenados por nombre. Cada elemento expone: id, nombre, config.

    POST /panel/analytics/profiles/
         Cuerpo (JSON): {"nombre": str, "config": {...}}
         Crea un perfil nuevo o actualiza uno existente con el mismo nombre
         (semántica upsert — unicidad aplicada a nivel de modelo).
         Devuelve el perfil guardado como JSON con HTTP 200.
         Devuelve HTTP 400 si el payload está ausente o es inválido.
    """

    def get (self ,request ):
        """
        Returns the list of AnalyticsProfile records for the current CompanyUser.
        ---
        Devuelve la lista de registros AnalyticsProfile del CompanyUser actual.
        """
        import json as _json 
        from django .http import JsonResponse 

        try :
            company_user =request .user .company_user 
        except AttributeError :
            return JsonResponse ({"error":"Sin perfil de empresa asociado."},status =403 )

        profiles =(
        AnalyticsProfile .objects 
        .filter (company_user =company_user )
        .order_by ("nombre")
        .values ("id","nombre","config")
        )
        return JsonResponse ({"profiles":list (profiles )})

    def post (self ,request ):
        """
        Creates or updates an AnalyticsProfile for the current CompanyUser.
        Uses update_or_create so that saving an existing name overwrites its
        config rather than raising a uniqueness error.
        ---
        Crea o actualiza un AnalyticsProfile para el CompanyUser actual.
        Usa update_or_create para que guardar un nombre existente sobreescriba
        su config en lugar de lanzar un error de unicidad.
        """
        import json as _json 
        from django .http import JsonResponse 

        try :
            company_user =request .user .company_user 
        except AttributeError :
            return JsonResponse ({"error":"Sin perfil de empresa asociado."},status =403 )


        try :
            payload =_json .loads (request .body )
        except (ValueError ,TypeError ):
            return JsonResponse ({"error":"Cuerpo JSON inválido."},status =400 )

        nombre =payload .get ("nombre","").strip ()
        config =payload .get ("config")

        if not nombre :
            return JsonResponse ({"error":"El campo 'nombre' es obligatorio."},status =400 )
        if not isinstance (config ,dict ):
            return JsonResponse ({"error":"El campo 'config' debe ser un objeto JSON."},status =400 )



        profile ,_ =AnalyticsProfile .objects .update_or_create (
        company_user =company_user ,
        nombre =nombre ,
        defaults ={"config":config },
        )

        return JsonResponse ({
        "id":profile .pk ,
        "nombre":profile .nombre ,
        "config":profile .config ,
        })


class WorkshopAssetAutocompleteView (WorkshopRequiredMixin ,View ):
    """
    JSON endpoint returning MachineAsset records for the authenticated
    CompanyUser's company. Supports incremental search via the optional
    'q' GET parameter (matches against codigo and brand_model).
    Used by the operator work-order entry form (Hito 7 / Paso 5).

    GET /panel/operator/assets/?q=<query>
        Returns a JSON array of {codigo, brand_model} objects, max 20 results.
        If 'q' is absent or blank, returns the first 20 active assets ordered
        by codigo.
    ---
    Endpoint JSON que devuelve registros MachineAsset de la empresa del
    CompanyUser autenticado. Admite búsqueda incremental mediante el parámetro
    GET opcional 'q' (busca en codigo y brand_model).
    Usado por el formulario de entrada de partes del operario (Hito 7 / Paso 5).

    GET /panel/operator/assets/?q=<query>
        Devuelve un array JSON de objetos {codigo, brand_model}, máx. 20 resultados.
        Si 'q' está ausente o vacío, devuelve los primeros 20 activos ordenados
        por codigo.
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns a filtered list of active MachineAsset records as JSON.
        ---
        Devuelve una lista filtrada de registros MachineAsset activos como JSON.
        """
        from django .http import JsonResponse 
        from fleet .models import MachineAsset 

        try :
            company =request .user .company_user .company 
        except AttributeError :
            return JsonResponse ({"error":"Sin perfil de empresa."},status =403 )

        q =request .GET .get ("q","").strip ()
        qs =MachineAsset .objects .filter (company =company ,is_active =True )

        if q :


            qs =qs .filter (
            django_models .Q (code__icontains =q )|
            django_models .Q (brand_model__icontains =q )
            )

        assets =list (
        qs .order_by ("code")
        .values ("code","brand_model")[:20 ]
        )
        return JsonResponse ({"assets":assets })


class WorkOrderEntryUploadView (WorkshopRequiredMixin ,View ):
    """
    Handles the operator Upload path (Via C) for new work orders.
    Accepts a photo or PDF file, rasterises it and sends it to Gemini Vision
    via extract_work_order_page_full() which returns both work blocks (front)
    and spare-part lines (back). The extracted data is stored in the session
    and the user is redirected to WorkOrderEntryConfirmView for full validation
    before any database write occurs.

    GET  /panel/operator/upload/
         Renders the upload form (upload_entry.html).
    POST /panel/operator/upload/
         Processes the uploaded file, stores extraction result in session,
         redirects to /panel/operator/confirm/.
    ---
    Gestiona la vía Upload (Vía C) del operario para partes nuevos.
    Acepta una foto o PDF, lo rasteriza y lo envía a Gemini Vision mediante
    extract_work_order_page_full() que devuelve tanto los bloques de trabajo
    (cara delantera) como las líneas de repuesto (cara trasera). Los datos
    extraídos se almacenan en la sesión y el usuario es redirigido a
    WorkOrderEntryConfirmView para validación completa antes de escribir en BD.

    GET  /panel/operator/upload/
         Renderiza el formulario de subida (upload_entry.html).
    POST /panel/operator/upload/
         Procesa el archivo subido, almacena el resultado de extracción en
         sesión, redirige a /panel/operator/confirm/.
    """

    template_name ="panel/operator/upload_entry.html"

    def _get_context (self ,request ,error =None ):
        """
        Builds base template context for the upload view.
        ---
        Construye el contexto base para la vista de subida.
        """
        cu =request .user .company_user 
        return {
        "company":cu .company ,
        "company_user":cu ,
        "active_nav":"operator_dashboard",
        "error":error ,
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the upload form.
        ---
        Renderiza el formulario de subida.
        """
        return render (request ,self .template_name ,self ._get_context (request ))

    def post (self ,request ,*args ,**kwargs ):
        """
        Processes the uploaded file through Gemini Vision (full prompt).
        Stores the structured extraction result in the session and redirects
        to the confirmation view. On Gemini failure (confidence=FAILED) the
        form is re-rendered with a descriptive error message.
        ---
        Procesa el archivo subido mediante Gemini Vision (prompt completo).
        Almacena el resultado de extracción estructurado en la sesión y redirige
        a la vista de confirmación. En caso de fallo de Gemini (confidence=FAILED)
        se re-renderiza el formulario con un mensaje de error descriptivo.
        """
        import io 
        from pdf2image import convert_from_bytes 
        from PIL import Image 
        from work_order_processor .services import extract_work_order_page_full 

        uploaded_file =request .FILES .get ("work_order_file")
        if not uploaded_file :
            return render (
            request ,self .template_name ,
            self ._get_context (request ,error ="Debes seleccionar un archivo.")
            )

        file_bytes =uploaded_file .read ()
        content_type =uploaded_file .content_type or ""

        try :


            if "pdf"in content_type or uploaded_file .name .lower ().endswith (".pdf"):


                pages =convert_from_bytes (file_bytes ,dpi =200 ,first_page =1 ,last_page =1 )
                if not pages :
                    raise ValueError ("# No se pudieron extraer páginas del PDF.")
                buf =io .BytesIO ()
                pages [0 ].save (buf ,format ="PNG")
                image_bytes =buf .getvalue ()
            else :


                img =Image .open (io .BytesIO (file_bytes )).convert ("RGB")
                buf =io .BytesIO ()
                img .save (buf ,format ="PNG")
                image_bytes =buf .getvalue ()

        except Exception as exc :
            logger .error (
            "# [Upload] Error al rasterizar archivo: %s",exc ,exc_info =True 
            )
            return render (
            request ,self .template_name ,
            self ._get_context (
            request ,
            error =(
            "No se pudo procesar el archivo. "
            "Asegúrate de que es una imagen o PDF válido."
            ),
            )
            )



        extraction =extract_work_order_page_full (image_bytes )

        if extraction .get ("extraction_confidence")=="FAILED":
            return render (
            request ,self .template_name ,
            self ._get_context (
            request ,
            error =(
            "Gemini no pudo extraer datos del archivo. "
            "Comprueba que la imagen sea legible y corresponda "
            "a un parte de trabajo."
            ),
            )
            )



        request .session ["operator_upload_extraction"]=extraction 
        request .session .modified =True 

        logger .info (
        "# [Upload] Extracción completada. Confianza: %s | "
        "Entradas: %d | Repuestos: %d. Redirigiendo a confirmación.",
        extraction .get ("extraction_confidence"),
        len (extraction .get ("entradas",[])),
        len (extraction .get ("repuestos",[])),
        )

        return redirect ("/panel/operator/confirm/")












def _resolve_operator_schedule (cu ,company ):
    """
    Resolves the effective WorkdaySchedule for a CompanyUser following
    the Gate 4 priority chain:
      1. CompanyUser.workday_schedule (individual assignment).
      2. First active Section of the operator's Contact that has a
         workday_schedule assigned.
      3. Company-level default WorkdaySchedule (is_default=True).
      4. None — no schedule available.
    Returns the resolved WorkdaySchedule instance or None.
    ---
    Resuelve el WorkdaySchedule efectivo para un CompanyUser siguiendo
    la cadena de prioridad de Gate 4:
      1. CompanyUser.workday_schedule (asignacion individual).
      2. Primera Section activa del Contact del operario que tenga
         workday_schedule asignado.
      3. WorkdaySchedule por defecto de empresa (is_default=True).
      4. None — no hay horario disponible.
    Devuelve la instancia WorkdaySchedule resuelta o None.
    """
    from ivr_config .models import WorkdaySchedule as _WDS_R 
    from ivr_config .models import Contact as _Contact_R 

    if cu .workday_schedule_id :
        return cu .workday_schedule 

    contact =(
    _Contact_R .objects 
    .filter (company_user =cu )
    .prefetch_related ("sections__workday_schedule")
    .first ()
    )
    if contact is not None :
        section_schedule =next (
        (
        sec .workday_schedule 
        for sec in contact .sections .filter (
        is_active =True ,workday_schedule__isnull =False 
        ).select_related ("workday_schedule").order_by ("name")
        ),
        None ,
        )
        if section_schedule is not None :
            return section_schedule 

    return _WDS_R .objects .filter (company =company ,is_default =True ).first ()


def _parse_entry_lines_from_post (POST ,company ):
    """
    Parses and resolves work-block entry lines submitted via POST.

    Resolution strategy for machine_asset (two-pass):
      Pass 1 — direct iexact on machine_raw: covers autocomplete selections
               where the field contains the exact asset.code string.
      Pass 2 — iexact on _normalise_machine_code(machine_raw): covers OCR
               and handwritten input where normalisation is required.

    Absence (PERSONAL asset) handling:
      When the selected machine_asset matches the PERSONAL asset code,
      the backend reads entrada_{i}_absence_category from POST, resolves
      the AbsenceCategory, and overrides fault_description with its label.
      If requires_note=True the operator must supply repair_notes (validated
      in Gate 1). The resolved AbsenceCategory instance is stored in the
      'absence_category' key of the dict so that save_blocks and close_order
      can create the synthetic WorkdayGap record.

    Returns a list of dicts ready to feed the integrity gate and the
    atomic persistence block.
    ---
    Parsea y resuelve las líneas de entrada de bloque de trabajo enviadas
    por POST.

    Estrategia de resolución para machine_asset (dos pasadas):
      Pasada 1 — iexact directo sobre machine_raw: cubre selecciones del
                 autocompletado donde el campo contiene el asset.code exacto.
      Pasada 2 — iexact sobre _normalise_machine_code(machine_raw): cubre
                 entrada OCR y manuscrita donde se requiere normalización.

    Gestión de ausencias (activo PERSONAL):
      Cuando el machine_asset seleccionado coincide con el código PERSONAL,
      el backend lee entrada_{i}_absence_category del POST, resuelve la
      AbsenceCategory y sobreescribe fault_description con su label.
      Si requires_note=True el operario debe proporcionar repair_notes
      (validado en Gate 1). La instancia AbsenceCategory resuelta se
      almacena en la clave 'absence_category' del dict para que save_blocks
      y close_order puedan crear el registro WorkdayGap sintético.

    Devuelve una lista de dicts lista para la barrera de integridad y el
    bloque de persistencia atómica.
    """
    import json as _json 
    from datetime import time as _dt_time 
    from fleet .models import MachineAsset 
    from work_order_processor .services import (
    _normalise_machine_code ,
    _compute_delta_hours ,
    )
    from work_order_processor .management .commands .seed_personal_asset import (
    PERSONAL_ASSET_CODE as _PERSONAL_CODE ,
    )

    num_entradas =int (POST .get ("num_entradas","1")or "1")
    entry_lines_data =[]

    for i in range (1 ,num_entradas +1 ):
        pfx =f"entrada_{i}_"
        machine_raw =POST .get (f"{pfx}machine_raw","").strip ()
        desc_averia =POST .get (f"{pfx}fault_description","").strip ()
        repair_notes =POST .get (f"{pfx}repair_notes","").strip ()
        hc_str =POST .get (f"{pfx}hc","").strip ()
        hf_str =POST .get (f"{pfx}hf","").strip ()
        or_val =POST .get (f"{pfx}or_val","").strip ()
        flags_raw =POST .get (f"{pfx}flags","[]")

        def _parse_t (s ):
            """
            Parses HH:MM string into time object, returns None on failure.
            ---
            Parsea cadena HH:MM a objeto time, devuelve None en fallo.
            """
            if not s :
                return None 
            try :
                parts =s .split (":")
                return _dt_time (int (parts [0 ]),int (parts [1 ]))
            except (ValueError ,IndexError ):
                return None 

        hc =_parse_t (hc_str )
        hf =_parse_t (hf_str )

        machine_norm =_normalise_machine_code (machine_raw )
        machine_asset =None 

        if machine_raw :


            try :
                machine_asset =MachineAsset .objects .get (
                code__iexact =machine_raw ,company =company 
                )
            except (MachineAsset .DoesNotExist ,MachineAsset .MultipleObjectsReturned ):
                machine_asset =MachineAsset .objects .filter (
                code__iexact =machine_raw ,company =company 
                ).first ()

        if machine_asset is None and machine_norm :


            try :
                machine_asset =MachineAsset .objects .get (
                code__iexact =machine_norm ,company =company 
                )
            except (MachineAsset .DoesNotExist ,MachineAsset .MultipleObjectsReturned ):
                machine_asset =MachineAsset .objects .filter (
                code__iexact =machine_norm ,company =company 
                ).first ()

        delta_hours =_compute_delta_hours (hc ,hf ,deduct_lunch =False )

        try :
            flags =_json .loads (flags_raw )if flags_raw else []
        except (ValueError ,TypeError ):
            flags =[]





        from decimal import Decimal ,InvalidOperation as _InvalidOp 

        def _parse_decimal (raw_val ):
            """
            Converts a POST string to Decimal or returns None on failure.
            ---
            Convierte una cadena POST a Decimal o devuelve None en caso de fallo.
            """
            v =(raw_val or "").strip ().replace (",",".")
            if not v :
                return None 
            try :
                return Decimal (v )
            except _InvalidOp :
                return None 

        odometer_reading =_parse_decimal (POST .get (f"entrada_{i}_odometer_reading",""))
        engine_hours_reading =_parse_decimal (POST .get (f"entrada_{i}_engine_hours_reading",""))
        crane_hours_reading =_parse_decimal (POST .get (f"entrada_{i}_crane_hours_reading",""))
























        absence_category_obj =None 
        _is_personal_block =(
        machine_asset is not None 
        and machine_asset .code .upper ()==_PERSONAL_CODE .upper ()
        )
        if _is_personal_block :
            _abs_cat_pk_raw =POST .get (f"{pfx}absence_category","").strip ()
            if _abs_cat_pk_raw :
                try :
                    from ivr_config .models import AbsenceCategory as _AbsCatParse 
                    absence_category_obj =_AbsCatParse .objects .get (
                    pk =int (_abs_cat_pk_raw ),
                    company =company ,
                    is_active =True ,
                    )





                    desc_averia =absence_category_obj .label 
                except (ValueError ,TypeError ):
                    logger .warning (
                    "# [_parse_entry_lines] absence_category pk=%r inválido "
                    "para entrada_%d — se omite.",
                    _abs_cat_pk_raw ,i ,
                    )
                except Exception as _abs_exc :
                    logger .warning (
                    "# [_parse_entry_lines] Error resolviendo AbsenceCategory "
                    "pk=%r para entrada_%d: %s",
                    _abs_cat_pk_raw ,i ,_abs_exc ,
                    )

        entry_lines_data .append ({
        "line_number":i ,
        "machine_raw":machine_raw ,
        "machine_norm":machine_norm or "",
        "machine_asset":machine_asset ,
        "fault_description":desc_averia ,
        "repair_notes":repair_notes ,
        "hc":hc ,
        "hf":hf ,
        "or_val":or_val ,
        "delta_hours":delta_hours ,
        "flags":flags ,
        "odometer_reading":odometer_reading ,
        "engine_hours_reading":engine_hours_reading ,
        "crane_hours_reading":crane_hours_reading ,
        "absence_category":absence_category_obj ,
        "is_personal":_is_personal_block ,
        })

    return entry_lines_data 


def _parse_spare_parts_from_post (POST ,company ,entry_lines_data =None ):
    """
    Parses and resolves spare-part lines submitted via POST.

    The select field for each spare part now delivers vehiculo_raw directly
    (the CdG value chosen by the operator from the unique list of machine_raw
    values in the current work blocks, or a free-text value when "Otro" is
    selected). Resolution against MachineAsset is attempted in two passes for
    every repuesto regardless of origin. If no match is found, vehicle_asset
    is None and cg_incident is set to True so the persistence layer can set
    WorkOrder.has_cg_incident = True.

    entry_lines_data is accepted for backwards compatibility but is no longer
    used to populate vehiculo_raw — the POST value is authoritative.

    Returns a list of dicts ready to feed the integrity gate and the
    atomic persistence block.
    ---
    Parsea y resuelve las líneas de repuesto enviadas por POST.

    El select de cada repuesto ahora entrega vehiculo_raw directamente
    (el valor CdG elegido por el operario de la lista única de machine_raw
    de los bloques de trabajo actuales, o un valor de texto libre cuando se
    selecciona "Otro"). La resolución contra MachineAsset se intenta en dos
    pasadas para cada repuesto independientemente del origen. Si no hay
    coincidencia, vehicle_asset es None y cg_incident se activa para que la
    capa de persistencia pueda establecer WorkOrder.has_cg_incident = True.

    entry_lines_data se acepta por compatibilidad hacia atrás pero ya no se
    usa para poblar vehiculo_raw — el valor del POST es autoritativo.

    Devuelve una lista de dicts lista para la barrera de integridad y el
    bloque de persistencia atómica.
    """
    from decimal import Decimal ,InvalidOperation 
    from fleet .models import MachineAsset as _MachineAsset 
    from work_order_processor .services import _normalise_machine_code as _norm_code 

    num_repuestos =int (POST .get ("num_repuestos","0")or "0")
    spare_parts_data =[]

    for r in range (1 ,num_repuestos +1 ):
        pfx =f"repuesto_{r}_"
        referencia =POST .get (f"{pfx}referencia","").strip ()
        material =POST .get (f"{pfx}material","").strip ()
        unidades_str =POST .get (f"{pfx}unidades","").strip ()
        origen =POST .get (f"{pfx}origen","WAREHOUSE").strip ()
        proveedor =POST .get (f"{pfx}proveedor","").strip ()






        _cdg_raw =POST .get (f"{pfx}vehiculo_raw","").strip ()
        if _cdg_raw =="__otro__":
            vehiculo_raw =POST .get (f"{pfx}cdg_free","").strip ()
        else :
            vehiculo_raw =_cdg_raw 

        quantity =None 
        if unidades_str :
            try :
                quantity =Decimal (unidades_str .replace (",","."))
            except InvalidOperation :
                quantity =None 

        if origen not in ("SUPPLIER","WAREHOUSE"):
            origen ="WAREHOUSE"









        veh_asset =None 
        cg_incident =False 

        if vehiculo_raw and company is not None :


            try :
                veh_asset =_MachineAsset .objects .get (
                code__iexact =vehiculo_raw ,company =company 
                )
            except (_MachineAsset .DoesNotExist ,_MachineAsset .MultipleObjectsReturned ):
                veh_asset =_MachineAsset .objects .filter (
                code__iexact =vehiculo_raw ,company =company 
                ).first ()

            if veh_asset is None :


                norm =_norm_code (vehiculo_raw )
                if norm :
                    try :
                        veh_asset =_MachineAsset .objects .get (
                        code__iexact =norm ,company =company 
                        )
                    except (_MachineAsset .DoesNotExist ,_MachineAsset .MultipleObjectsReturned ):
                        veh_asset =_MachineAsset .objects .filter (
                        code__iexact =norm ,company =company 
                        ).first ()

        if veh_asset is None and vehiculo_raw :


            cg_incident =True 

        spare_parts_data .append ({
        "line_number":r ,
        "referencia":referencia ,
        "vehiculo_raw":vehiculo_raw ,
        "vehicle_asset":veh_asset ,
        "material":material ,
        "quantity":quantity ,
        "source":origen ,
        "supplier":proveedor if origen =="SUPPLIER"else "",
        "flags":[],
        "cg_incident":cg_incident ,
        })

    return spare_parts_data 


class WorkOrderEntryConfirmView (WorkshopRequiredMixin ,View ):
    """
    Confirmation and persistence view for the operator Upload path (Via C).
    Reads the extraction result stored in the session by WorkOrderEntryUploadView,
    renders a fully editable confirmation form and, on POST, atomically persists
    the validated data to the database:
      WorkOrder (status=DONE, source_pdf blank) +
      WorkOrderEntry (page 1, confidence from Gemini) +
      N × WorkOrderEntryLine +
      M × SparePartLine per entry line.
    After persistence, generate_work_order_excel() is called synchronously.

    GET  /panel/operator/confirm/
         Renders the confirmation form pre-filled from session data.
         Redirects to the upload view if no session data is found.
    POST /panel/operator/confirm/
         Validates, persists and redirects to the work-order list on success.
    ---
    Vista de confirmación y persistencia para la vía Upload del operario (Vía C).
    Lee el resultado de extracción almacenado en sesión por WorkOrderEntryUploadView,
    renderiza un formulario de confirmación completamente editable y, en POST,
    persiste atómicamente los datos validados en la base de datos:
      WorkOrder (status=DONE, source_pdf en blanco) +
      WorkOrderEntry (página 1, confianza de Gemini) +
      N × WorkOrderEntryLine +
      M × SparePartLine por línea de entrada.
    Tras la persistencia, generate_work_order_excel() se llama de forma síncrona.

    GET  /panel/operator/confirm/
         Renderiza el formulario de confirmación pre-rellenado con datos de sesión.
         Redirige a la vista de subida si no hay datos de sesión.
    POST /panel/operator/confirm/
         Valida, persiste y redirige a la lista de partes en caso de éxito.
    """

    template_name ="panel/operator/confirm_entry.html"

    def _get_company_user (self ,request ):
        """
        Returns the CompanyUser for the authenticated request user.
        ---
        Devuelve el CompanyUser del usuario autenticado en la solicitud.
        """
        return request .user .company_user 

    def _get_context_base (self ,request ):
        """
        Returns the base template context with company and navigation data.
        ---
        Devuelve el contexto base con empresa y datos de navegación.
        """
        cu =self ._get_company_user (request )
        return {
        "company":cu .company ,
        "company_user":cu ,
        "active_nav":"operator_dashboard",
        }

    def _resolve_machine (self ,company ,raw_code ):
        """
        Attempts to resolve a raw machine code string to a MachineAsset
        belonging to the given company, applying the same D4 normalisation
        used by the historical pipeline (_normalise_machine_code).
        Returns the MachineAsset instance or None if no match is found.
        ---
        Intenta resolver un código de máquina bruto a un MachineAsset
        perteneciente a la empresa dada, aplicando la misma normalización D4
        usada por el pipeline histórico (_normalise_machine_code).
        Devuelve la instancia MachineAsset o None si no hay coincidencia.
        """
        from work_order_processor .services import _normalise_machine_code 
        from fleet .models import MachineAsset 

        if not raw_code :
            return None 
        norm =_normalise_machine_code (raw_code )
        if not norm :
            return None 
        try :
            return MachineAsset .objects .get (code__iexact =norm ,company =company )
        except MachineAsset .DoesNotExist :
            return None 
        except MachineAsset .MultipleObjectsReturned :
            return MachineAsset .objects .filter (
            code__iexact =norm ,company =company 
            ).first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the confirmation form using extraction data from the session.
        Redirects to the upload view if the session contains no extraction data.
        ---
        Renderiza el formulario de confirmación usando los datos de extracción
        de la sesión. Redirige a la vista de subida si la sesión no contiene
        datos de extracción.
        """
        from fleet .models import MachineAsset 

        extraction =request .session .get ("operator_upload_extraction")
        if not extraction :
            logger .warning (
            "# [Confirm] No hay datos de extracción en sesión. "
            "Redirigiendo a la vista de subida."
            )
            return redirect ("/panel/operator/upload/")

        cu =self ._get_company_user (request )
        company =cu .company 



        entradas_enriched =[]
        for idx ,entrada in enumerate (extraction .get ("entradas",[]),start =1 ):
            raw_code =entrada .get ("machine_raw")or ""
            machine_asset =self ._resolve_machine (company ,raw_code )
            entradas_enriched .append ({
            "idx":idx ,
            "machine_raw":raw_code ,
            "machine_asset":machine_asset ,
            "fault_description":entrada .get ("fault_description")or "",
            "repair_notes":entrada .get ("repair_notes")or "",
            "hc":entrada .get ("hc")or "",
            "hf":entrada .get ("hf")or "",
            "or_val":entrada .get ("or_val")or "",
            "flags":entrada .get ("flags")or [],
            })



        repuestos_enriched =[]
        for ridx ,rep in enumerate (extraction .get ("repuestos",[]),start =1 ):
            veh_raw =rep .get ("vehiculo_raw")or ""
            vehicle_asset =self ._resolve_machine (company ,veh_raw )
            repuestos_enriched .append ({
            "ridx":ridx ,
            "referencia":rep .get ("referencia")or "",
            "vehiculo_raw":veh_raw ,
            "vehicle_asset":vehicle_asset ,
            "material":rep .get ("material")or "",
            "unidades":rep .get ("unidades"),
            "origen":rep .get ("origen")or "WAREHOUSE",
            "proveedor":rep .get ("proveedor")or "",
            "flags":rep .get ("flags")or [],
            })



        assets =list (
        MachineAsset .objects .filter (company =company ,is_active =True )
        .order_by ("code")
        .values ("code","brand_model")
        )

        _min_date_get =_get_min_allowed_date (cu )
        context =self ._get_context_base (request )
        context .update ({
        "extraction":extraction ,
        "fecha":extraction .get ("fecha")or "",
        "uncertain_date":extraction .get ("uncertain_date",False ),
        "confidence":extraction .get ("extraction_confidence",""),
        "entradas_enriched":entradas_enriched ,
        "repuestos_enriched":repuestos_enriched ,
        "assets":assets ,
        "num_entradas":len (entradas_enriched ),
        "num_repuestos":len (repuestos_enriched ),
        "min_date":_min_date_get .isoformat ()if _min_date_get else "",
        })
        return render (request ,self .template_name ,context )

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the confirmed form data and atomically persists:
          WorkOrder → WorkOrderEntry → WorkOrderEntryLine(s) → SparePartLine(s).
        Generates the Excel report synchronously after persistence.
        Clears the session extraction data on success.
        On validation failure, re-renders the confirmation form with error context.
        ---
        Valida los datos del formulario confirmado y persiste atómicamente:
          WorkOrder → WorkOrderEntry → WorkOrderEntryLine(s) → SparePartLine(s).
        Genera el informe Excel de forma síncrona tras la persistencia.
        Elimina los datos de extracción de la sesión en caso de éxito.
        En caso de fallo de validación, re-renderiza el formulario con contexto
        de error.
        """
        import json as _json 
        from datetime import date ,time as dt_time 
        from decimal import Decimal ,InvalidOperation 
        from django .db import transaction 
        from django .utils import timezone 
        from fleet .models import MachineAsset 
        from work_order_processor .models import (
        WorkOrder ,WorkOrderEntry ,WorkOrderEntryLine ,SparePartLine ,
        )
        from work_order_processor .services import (
        generate_work_order_excel ,
        _normalise_machine_code ,
        _compute_delta_hours ,
        )

        cu =self ._get_company_user (request )
        company =cu .company 
        POST =request .POST 















        _gate0_fecha_str =POST .get ("fecha","").strip ()
        _gate0_work_date =None 
        if _gate0_fecha_str :
            from datetime import datetime as _dt0 
            for _fmt0 in ("%d/%m/%Y","%Y-%m-%d"):
                try :
                    _gate0_work_date =_dt0 .strptime (_gate0_fecha_str ,_fmt0 ).date ()
                    break 
                except ValueError :
                    continue 

        if _gate0_work_date is not None :



            _min_date_c =_get_min_allowed_date (cu )
            if _min_date_c is not None and _gate0_work_date <_min_date_c :
                from datetime import timedelta as _td_c 
                _last_rev_c =_min_date_c -_td_c (days =1 )
                context =self ._get_context_base (request )
                context .update ({
                "error":(
                f"No puedes introducir un parte con fecha "
                f"{_gate0_work_date.strftime('%d/%m/%Y')}. "
                f"El ultimo parte revisado es del "
                f"{_last_rev_c.strftime('%d/%m/%Y')} y ya ha sido auditado. "
                f"La fecha minima permitida es "
                f"{_min_date_c.strftime('%d/%m/%Y')}."
                ),
                "fecha":_gate0_fecha_str ,
                "uncertain_date":False ,
                "confidence":"",
                "entradas_enriched":[],
                "repuestos_enriched":[],
                "num_entradas":0 ,
                "num_repuestos":0 ,
                "min_date":_min_date_c .isoformat (),
                })
                return render (request ,self .template_name ,context )

            from django .urls import reverse as _rev0 
            _existing_entry0 =WorkOrderEntry .objects .filter (
            work_order__company =company ,
            work_order__uploaded_by =cu ,
            work_order__source__in =[
            WorkOrder .Source .DIGITAL ,
            WorkOrder .Source .GENERATED ,
            ],
            work_order__reviewed =False ,
            work_date =_gate0_work_date ,
            ).select_related ("work_order").first ()

            if _existing_entry0 is not None :


                _gate0_lines =_parse_entry_lines_from_post (POST ,company )
                _gate0_spare =_parse_spare_parts_from_post (
                POST ,company ,entry_lines_data =_gate0_lines 
                )
                request .session ["pending_merge_lines"]=_serialize_pending_lines (
                _gate0_lines ,_gate0_spare ,_gate0_work_date 
                )
                return redirect (
                _rev0 (
                "panel:operator_merge",
                kwargs ={"entry_pk":_existing_entry0 .pk },
                )
                )





        fecha_str =POST .get ("fecha","").strip ()
        work_date =None 
        if fecha_str :
            for fmt in ("%d/%m/%Y","%Y-%m-%d"):
                try :
                    from datetime import datetime 
                    work_date =datetime .strptime (fecha_str ,fmt ).date ()
                    break 
                except ValueError :
                    continue 










        entry_lines_data =_parse_entry_lines_from_post (POST ,company )
        spare_parts_data =_parse_spare_parts_from_post (
        POST ,company ,entry_lines_data =entry_lines_data 
        )

































        integrity_errors =[]


        if not work_date :
            integrity_errors .append (
            "La fecha del parte es obligatoria y debe tener formato DD/MM/AAAA."
            )


        if not entry_lines_data :
            integrity_errors .append (
            "El parte debe contener al menos un bloque de trabajo."
            )

        for ld in entry_lines_data :
            blk =f"Bloque {ld['line_number']}"
            if not ld ["machine_raw"]:
                integrity_errors .append (
                f"{blk}: el código de máquina es obligatorio."
                )
            elif ld ["machine_asset"]is None :
                integrity_errors .append (
                f"{blk}: el código '{ld['machine_raw']}' no se ha podido "
                f"identificar en el catálogo de flota. "
                f"Corrígelo antes de guardar."
                )
            if not ld ["hc"]:
                integrity_errors .append (
                f"{blk}: la hora de inicio (H.C.) es obligatoria."
                )
            if not ld ["hf"]:
                integrity_errors .append (
                f"{blk}: la hora de fin (H.F.) es obligatoria."
                )
            if ld ["hc"]and ld ["hf"]and ld ["delta_hours"]is not None :
                if ld ["delta_hours"]<=0 :
                    integrity_errors .append (
                    f"{blk}: la H.F. debe ser posterior a la H.C. "
                    f"(Δ horas calculado: {ld['delta_hours']})."
                    )
            if not ld ["fault_description"]:
                integrity_errors .append (
                f"{blk}: la descripción de la avería es obligatoria."
                )
            if not ld ["repair_notes"]:
                integrity_errors .append (
                f"{blk}: la descripción de la reparación realizada es obligatoria."
                )







        for ld in entry_lines_data :
            if ld ["machine_asset"]is not None :
                asset =ld ["machine_asset"]
                blk =f"Bloque {ld['line_number']}"
                if asset .has_odometer :
                    reading =ld .get ("odometer_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de km (odómetro) obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de km no puede ser cero para {asset.code} "
                        f"(ya tiene partes anteriores registrados)."
                        )
                if asset .has_engine_hours :
                    reading =ld .get ("engine_hours_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de horómetro motor obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de horómetro motor no puede ser cero "
                        f"para {asset.code} (ya tiene partes anteriores registrados)."
                        )
                if asset .has_crane_hours :
                    reading =ld .get ("crane_hours_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de horómetro grúa obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de horómetro grúa no puede ser cero "
                        f"para {asset.code} (ya tiene partes anteriores registrados)."
                        )


        for spd in spare_parts_data :
            rep =f"Repuesto {spd['line_number']}"
            if not spd ["material"]:
                integrity_errors .append (
                f"{rep}: la descripción del material es obligatoria."
                )
            if spd ["quantity"]is None or spd ["quantity"]<=0 :
                integrity_errors .append (
                f"{rep}: las unidades deben ser un número positivo."
                )

        if integrity_errors :




            entradas_enriched_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":ld ["flags"],
            }
            for ld in entry_lines_data 
            ]
            spare_enriched_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":spd ["flags"],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":" | ".join (integrity_errors ),
            "fecha":fecha_str ,
            "uncertain_date":False ,
            "confidence":POST .get ("confidence",""),
            "entradas_enriched":entradas_enriched_post ,
            "repuestos_enriched":spare_enriched_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            })
            return render (request ,self .template_name ,context )







        if not POST .get ("save_confirmed")and _form_action !="save_blocks":
            entradas_enriched_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":ld ["flags"],
            }
            for ld in entry_lines_data 
            ]
            spare_enriched_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":spd ["flags"],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":None ,
            "fecha":fecha_str ,
            "uncertain_date":False ,
            "confidence":POST .get ("confidence",""),
            "entradas_enriched":entradas_enriched_post ,
            "repuestos_enriched":spare_enriched_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            })
            return render (request ,self .template_name ,context )





        from work_order_processor .validators import (
        run_intra_part_validation ,
        parse_blocks_from_post ,
        validate_inter_overlap ,
        TimeBlock ,
        )

        num_entradas_post =int (POST .get ("num_entradas",len (entry_lines_data )))
        _blocks =parse_blocks_from_post (POST ,num_entradas_post ,entry_lines_data =entry_lines_data )
        _intra =run_intra_part_validation (_blocks )

        if not _intra .ok and _form_action !="save_blocks":
            entradas_enriched_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":ld ["flags"],
            }
            for ld in entry_lines_data 
            ]
            spare_enriched_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":spd ["flags"],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":" | ".join (e .message for e in _intra .errors ),
            "fecha":fecha_str ,
            "uncertain_date":False ,
            "confidence":POST .get ("confidence",""),
            "entradas_enriched":entradas_enriched_post ,
            "repuestos_enriched":spare_enriched_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            })
            return render (request ,self .template_name ,context )




        try :
            with transaction .atomic ():


                worker_name =(
                cu .user .get_full_name ()or cu .user .username 
                ).upper ()










                date_tag =(
                work_date .strftime ("%d-%m-%Y")if work_date else "SIN-FECHA"
                )
                synthetic_name =f"{worker_name}_{date_tag}.pdf"

                work_order =WorkOrder (
                company =company ,
                uploaded_by =cu ,
                source =WorkOrder .Source .DIGITAL ,
                status =WorkOrder .Status .DONE ,
                total_pages =1 ,
                processed_pages =1 ,
                reviewed =False ,
                )
                work_order .source_pdf .name =synthetic_name 
                work_order .save ()













                from work_order_processor .services import _coerce_confidence as _cc 
                _session_extraction =request .session .get (
                "operator_upload_extraction",{}
                )
                _gemini_confidence =_cc (
                _session_extraction .get ("extraction_confidence")
                )
                _gemini_uncertain_date =bool (
                _session_extraction .get ("uncertain_date",False )
                )
                entry =WorkOrderEntry .objects .create (
                work_order =work_order ,
                page_number =1 ,
                worker_name =worker_name ,
                work_date =work_date ,
                uncertain_date =_gemini_uncertain_date ,
                extraction_confidence =_gemini_confidence ,
                raw_gemini_response =None ,
                )


                created_lines ={}
                created_line_pks =[]

                for ld in entry_lines_data :
                    line =WorkOrderEntryLine .objects .create (
                    entry =entry ,
                    line_number =ld ["line_number"],
                    machine_asset =ld ["machine_asset"],
                    machine_raw =ld ["machine_raw"],
                    machine_norm =ld ["machine_norm"],
                    fault_description =ld ["fault_description"],
                    repair_notes =ld ["repair_notes"],
                    hc =ld ["hc"],
                    hf =ld ["hf"],
                    or_val =ld ["or_val"],
                    delta_hours =ld ["delta_hours"],
                    flags =ld ["flags"],
                    odometer_reading =ld .get ("odometer_reading"),
                    engine_hours_reading =ld .get ("engine_hours_reading"),
                    crane_hours_reading =ld .get ("crane_hours_reading"),
                    )
                    created_lines [ld ["line_number"]]=line 
                    created_line_pks .append (line .pk )



                for spd in spare_parts_data :




                    target_line =next (iter (created_lines .values ()),None )
                    if target_line is None :
                        continue 

                    SparePartLine .objects .create (
                    entry_line =target_line ,
                    line_number =spd ["line_number"],
                    reference =spd ["referencia"],
                    vehicle =spd ["vehicle_asset"],
                    material =spd ["material"],
                    quantity =spd ["quantity"],
                    source =spd ["source"],
                    supplier =spd ["supplier"],
                    flags =spd ["flags"],
                    )





            if any (spd .get ("cg_incident")for spd in spare_parts_data ):
                WorkOrder .objects .filter (pk =work_order .pk ).update (has_cg_incident =True )
                logger .warning (
                "# [Confirm] WorkOrder #%d marcado con has_cg_incident=True: "
                "al menos un repuesto tiene un CdG no resuelto en catálogo.",
                work_order .pk ,
                )

            logger .info (
            "# [Confirm] WorkOrder #%d creado correctamente. "
            "Entradas: %d | Repuestos: %d.",
            work_order .pk ,
            len (entry_lines_data ),
            len (spare_parts_data ),
            )





            import json as _json_mod 
            _zero_raw =POST .get ("zero_meters_confirmed","").strip ()
            if _zero_raw :
                try :
                    _zero_data =_json_mod .loads (_zero_raw )
                    for _bIdx_str ,_meter_list in _zero_data .items ():
                        try :
                            _bIdx =int (_bIdx_str )
                        except (ValueError ,TypeError ):
                            continue 
                        _line =created_lines .get (_bIdx )
                        if _line is None :
                            continue 
                        _asset =_line .machine_asset 
                        _line_fields =[]
                        _asset_fields =[]
                        for _m in _meter_list :
                            _name =_m .get ("name","")
                            if "odometer"in _name :
                                _line .odometer_reading =None 
                                _line_fields .append ("odometer_reading")
                                if _asset and _asset .has_odometer :
                                    _asset .has_odometer =False 
                                    _asset_fields .append ("has_odometer")
                            elif "engine_hours"in _name :
                                _line .engine_hours_reading =None 
                                _line_fields .append ("engine_hours_reading")
                                if _asset and _asset .has_engine_hours :
                                    _asset .has_engine_hours =False 
                                    _asset_fields .append ("has_engine_hours")
                            elif "crane_hours"in _name :
                                _line .crane_hours_reading =None 
                                _line_fields .append ("crane_hours_reading")
                                if _asset and _asset .has_crane_hours :
                                    _asset .has_crane_hours =False 
                                    _asset_fields .append ("has_crane_hours")
                        if _line_fields :
                            _line .save (update_fields =_line_fields )
                        if _asset and _asset_fields :
                            _asset .save (update_fields =list (set (_asset_fields )))
                            logger .info (
                            "# [Confirm] MachineAsset %s: flags desactivados: %s.",
                            _asset .code ,_asset_fields ,
                            )
                except (_json_mod .JSONDecodeError ,Exception )as _ze :
                    logger .warning (
                    "# [Confirm] Error procesando zero_meters_confirmed: %s",_ze 
                    )



            for ld in entry_lines_data :
                _asset =ld .get ("machine_asset")
                if _asset and _asset .first_repair :
                    _asset .first_repair =False 
                    _asset .save (update_fields =["first_repair"])
                    logger .info (
                    "# [Confirm] MachineAsset %s: first_repair=False.",
                    _asset .code ,
                    )

        except Exception as exc :
            logger .error (
            "# [Confirm] Error en persistencia atómica: %s",exc ,exc_info =True 
            )
            context =self ._get_context_base (request )
            context ["error"]=(
            f"Error al guardar el parte: {exc}. "
            "Por favor, inténtalo de nuevo o contacta con el administrador."
            )
            return render (request ,self .template_name ,context )










        from work_order_processor .models import WorkdayGap as _WDG_CO 
        for _co_ld in entry_lines_data :
            if not _co_ld .get ("is_personal"):
                continue 
            _co_abs_cat =_co_ld .get ("absence_category")
            _co_hc =_co_ld .get ("hc")
            _co_hf =_co_ld .get ("hf")
            if _co_hc is None or _co_hf is None :
                continue 
            _co_dur_min =max (
            0 ,
            (_co_hf .hour *60 +_co_hf .minute )
            -(_co_hc .hour *60 +_co_hc .minute ),
            )
            _WDG_CO .objects .create (
            work_order =work_order ,
            gap_type =_WDG_CO .GapType .GAP ,
            gap_start =_co_hc ,
            gap_end =_co_hf ,
            duration_minutes =_co_dur_min ,
            absence_category =_co_abs_cat ,
            note =_co_ld .get ("repair_notes",""),
            resolved =True ,
            )
            logger .info (
            "# [FormView/close_order] WorkdayGap sintético creado. "
            "work_order_pk=%r gap_start=%r gap_end=%r absence_cat=%r",
            work_order .pk ,_co_hc ,_co_hf ,
            _co_abs_cat .label if _co_abs_cat else None ,
            )










        for _lpk in created_line_pks :
            _line_obj =WorkOrderEntryLine .objects .filter (pk =_lpk ).first ()
            if _line_obj is None :
                continue 
            _cached =find_cached_classification (
            fault_description =_line_obj .fault_description ,
            repair_notes =_line_obj .repair_notes ,
            company =company ,
            )
            if _cached :
                WorkOrderEntryLine .objects .filter (pk =_lpk ).update (
                fault_category =_cached [0 ],
                fault_subcategory =_cached [1 ],
                )
                logger .info (
                "# [Confirm] Clasificación copiada desde caché para "
                "WorkOrderEntryLine pk=%d: category=%s subcategory=%s.",
                _lpk ,_cached [0 ],_cached [1 ],
                )
            else :
                classify_fault_line .apply_async (
                args =[_lpk ],
                queue ="work_orders",
                )
                logger .info (
                "# [Confirm] classify_fault_line encolada para "
                "WorkOrderEntryLine pk=%d.",
                _lpk ,
                )




        try :
            generate_work_order_excel (work_order .pk )
            logger .info (
            "# [Confirm] Excel generado correctamente para WorkOrder #%d.",
            work_order .pk ,
            )
        except Exception as exc :
            logger .warning (
            "# [Confirm] Excel no generado para WorkOrder #%d: %s.",
            work_order .pk ,exc ,
            )

















        if work_date is not None and _form_action !="save_blocks":
            from decimal import Decimal as _Dec_C2 
            from ivr_config .models import WorkerAbsence as _WA_C2 
            _total_hours_c2 =sum (
            (ld ["delta_hours"]for ld in _parsed_lines if ld .get ("delta_hours")is not None ),
            _Dec_C2 ("0"),
            )
            _has_absence_c2 =_WA_C2 .objects .filter (
            company_user =cu ,
            start_date__lte =work_date ,
            end_date__gte =work_date ,
            ).exists ()
            if _total_hours_c2 <_Dec_C2 ("8")and not _has_absence_c2 :
                _missing_c2 =_Dec_C2 ("8")-_total_hours_c2 
                context =self ._get_context_base (request )
                extraction =request .session .get ("operator_upload_extraction",{})
                context .update ({
                "error":(
                f"La jornada del parte suma {_total_hours_c2} h, "
                f"pero se requieren al menos 8 h. "
                f"Faltan {_missing_c2} h para completar la jornada. "
                f"Añade los bloques de trabajo que faltan o registra "
                f"una ausencia justificada para esta fecha."
                ),
                "extraction":extraction ,
                "fecha":POST .get ("fecha",""),
                "uncertain_date":False ,
                "confidence":"",
                "entradas_enriched":[],
                "repuestos_enriched":[],
                "num_entradas":0 ,
                "num_repuestos":0 ,
                "min_date":_get_min_allowed_date (cu ).isoformat ()if _get_min_allowed_date (cu )else "",
                })
                return render (request ,self .template_name ,context )















        if work_date is not None :
            from ivr_config .models import WorkdaySchedule as _WDS_C2 
            from django .urls import reverse as _rev_g4c2 
            _schedule_g4c2 =(
            cu .workday_schedule 
            if cu .workday_schedule_id 
            else (




            next (
            (
            _sec .workday_schedule 
            for _sec in (
            _contact_g4c2 .sections 
            .filter (is_active =True ,workday_schedule__isnull =False )
            .select_related ("workday_schedule")
            .order_by ("name")
            )
            ),
            None ,
            )
            if (_contact_g4c2 :=(
            Contact .objects .filter (company_user =cu )
            .prefetch_related ("sections__workday_schedule")
            .first ()
            ))is not None 
            else None 
            )or _WDS_C2 .objects .filter (
            company =company ,is_default =True 
            ).first ()
            )
            _gaps_g4c2 =_detect_workday_gaps (
            entry_lines_data ,_schedule_g4c2 ,work_date 
            )
            if _gaps_g4c2 :




                from django .db import transaction as _tx_g4c2 
                _worker_name_g4 =(
                cu .user .get_full_name ()or cu .user .username 
                ).upper ()
                _date_tag_g4 =work_date .strftime ("%d-%m-%Y")
                _synth_g4 =f"{_worker_name_g4}_{_date_tag_g4}_DRAFT.pdf"
                with _tx_g4c2 .atomic ():
                    _wo_draft =WorkOrder (
                    company =company ,
                    uploaded_by =cu ,
                    source =WorkOrder .Source .DIGITAL ,
                    status =WorkOrder .Status .PENDING_GAPS ,
                    total_pages =1 ,
                    processed_pages =1 ,
                    reviewed =False ,
                    )
                    _wo_draft .source_pdf .name =_synth_g4 
                    _wo_draft .save ()
                    _entry_draft =WorkOrderEntry .objects .create (
                    work_order =_wo_draft ,
                    page_number =1 ,
                    worker_name =_worker_name_g4 ,
                    work_date =work_date ,
                    uncertain_date =False ,
                    extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                    raw_gemini_response =None ,
                    )
                    _line_num_g4 =1 
                    for _ld in entry_lines_data :
                        _ln =WorkOrderEntryLine .objects .create (
                        entry =_entry_draft ,
                        line_number =_line_num_g4 ,
                        machine_asset =_ld .get ("machine_asset"),
                        machine_raw =_ld .get ("machine_raw",""),
                        machine_norm ="",
                        fault_description =_ld .get ("fault_description",""),
                        repair_notes =_ld .get ("repair_notes",""),
                        hc =_ld .get ("hc"),
                        hf =_ld .get ("hf"),
                        or_val =_ld .get ("or_val",""),
                        delta_hours =_ld .get ("delta_hours"),
                        flags =[],
                        odometer_reading =_ld .get ("odometer_reading"),
                        engine_hours_reading =_ld .get ("engine_hours_reading"),
                        crane_hours_reading =_ld .get ("crane_hours_reading"),
                        )
                        _line_num_g4 +=1 
                    from work_order_processor .models import WorkdayGap as _WDG_C2 
                    for _gap in _gaps_g4c2 :
                        _WDG_C2 .objects .create (
                        work_order =_wo_draft ,
                        gap_type =_gap ["gap_type"],
                        gap_start =_gap ["gap_start"],
                        gap_end =_gap ["gap_end"],
                        duration_minutes =_gap ["duration_minutes"],
                        )
                request .session ["pending_gaps_wo_pk"]=_wo_draft .pk 
                request .session .modified =True 
                logger .info (
                "# [ConfirmView/Gate4] PENDING_GAPS borrador pk=%d creado. "
                "%d gap(s) detectado(s). Redirigiendo a resolución.",
                _wo_draft .pk ,len (_gaps_g4c2 ),
                )
                return redirect (
                _rev_g4c2 (
                "panel:operator_gap_resolution",
                kwargs ={"wo_draft_pk":_wo_draft .pk },
                )
                )


        request .session .pop ("operator_upload_extraction",None )
        request .session .modified =True 





        _inter =validate_inter_overlap (
        company_user =cu ,
        work_date =work_date ,
        blocks =_blocks ,
        exclude_work_order_pk =work_order .pk ,
        )

        if _inter .has_overlap :


            WorkOrder .objects .filter (
            pk__in =[work_order .pk ]+_inter .conflicting_ids 
            ).update (has_overlap_incident =True )
            logger .warning (
            "# [Confirm] Solapamiento inter-parte detectado. "
            "WorkOrder #%d solapa con: %s.",
            work_order .pk ,
            _inter .conflicting_ids ,
            )
            django_messages .warning (
            request ,
            f"Parte #{work_order.pk} guardado con incidencia de solapamiento."
            )
            context =self ._get_context_base (request )
            context .update ({
            "overlap_incidents":True ,
            "new_work_order_pk":work_order .pk ,
            "conflicting_parts":[
            {"pk":pk ,"fecha":fecha }
            for pk ,fecha in zip (
            _inter .conflicting_ids ,
            _inter .conflicting_dates ,
            )
            ],
            "part_saved":True ,
            })
            return render (request ,self .template_name ,context )

        django_messages .success (
        request ,
        f"Parte de trabajo registrado correctamente (#{work_order.pk}). "
        f"El informe Excel está disponible en tu historial."
        )




        from django .urls import reverse as _reverse 
        _cu =request .user .company_user 
        if _cu .role in (CompanyUser .ROLE_ADMIN ,CompanyUser .ROLE_SUPERVISOR ):
            return redirect (_reverse ("panel:work_order_admin_history"))
        return redirect (_reverse ("panel:operator_history"))


class WorkOrderEntryFormView (WorkshopRequiredMixin ,View ):
    """
    Structured web form entry path for work orders (Via A).
    Allows WORKSHOP and ADMIN users to submit a daily work-order part
    directly via a multi-block web form, with no AI dependency and zero cost.

    GET  /panel/operator/form/
         Renders an empty form with one default work block and an empty
         spare-parts section. Additional blocks and spare-part rows can
         be added dynamically via JavaScript.
    POST /panel/operator/form/
         Applies the same integrity gate as WorkOrderEntryConfirmView and,
         on success, atomically persists:
           WorkOrder (status=DONE, source_pdf blank) +
           WorkOrderEntry (page 1, confidence=HIGH) +
           N x WorkOrderEntryLine +
           M x SparePartLine per entry line.
         Generates the Excel report synchronously after persistence.
    ---
    Via de entrada mediante formulario web estructurado para partes de
    trabajo (Via A). Permite a usuarios WORKSHOP y ADMIN enviar un parte
    diario directamente mediante un formulario web multi-bloque, sin
    dependencia de IA y con coste cero.

    GET  /panel/operator/form/
         Renderiza un formulario vacio con un bloque de trabajo por defecto
         y una seccion de repuestos vacia. Bloques y repuestos adicionales
         se anaden dinamicamente via JavaScript.
    POST /panel/operator/form/
         Aplica la misma barrera de integridad que WorkOrderEntryConfirmView
         y, en caso de exito, persiste atomicamente:
           WorkOrder (status=DONE, source_pdf en blanco) +
           WorkOrderEntry (pagina 1, confianza=HIGH) +
           N x WorkOrderEntryLine +
           M x SparePartLine por linea de entrada.
         Genera el informe Excel de forma sincrona tras la persistencia.
    """

    template_name ="panel/operator/form_entry.html"

    def _get_company_user (self ,request ):
        """
        Returns the CompanyUser for the authenticated request user.
        ---
        Devuelve el CompanyUser del usuario autenticado en la solicitud.
        """
        return request .user .company_user 

    def _get_context_base (self ,request ):
        """
        Returns the base template context with company and navigation data.
        Provides the list of active MachineAsset records for autocomplete and
        the list of open repair orders (BreakdownTicket with is_repair_order=True
        and status != RESOLVED) available to the authenticated operator.
        Repair orders without assigned_to are available to all operators.
        Repair orders assigned to this operator are included with priority.
        ---
        Devuelve el contexto base con empresa y datos de navegación.
        Proporciona la lista de MachineAsset activos para autocompletado y
        la lista de órdenes de reparación abiertas (BreakdownTicket con
        is_repair_order=True y status != RESOLVED) disponibles para el operario.
        Las OTs sin assigned_to están disponibles para cualquier operario.
        Las OTs asignadas a este operario se incluyen con prioridad.
        """
        from fleet .models import MachineAsset 
        from chat .models import BreakdownTicket 
        from django .db .models import Q as _Q 
        cu =self ._get_company_user (request )
        company =cu .company 
        assets =list (
        MachineAsset .objects .filter (company =company ,is_active =True )
        .order_by ("code")
        .values ("code","brand_model")
        )


        repair_orders =list (
        BreakdownTicket .objects 
        .filter (
        room__company =company ,
        is_repair_order =True ,
        )
        .exclude (status =BreakdownTicket .STATUS_RESOLVED )
        .filter (
        _Q (assigned_to__isnull =True )|_Q (assigned_to =cu )
        )
        .select_related ("machine","section")
        .order_by ("-urgency","created_at")
        )
        return {
        "company":company ,
        "company_user":cu ,
        "active_nav":"operator_dashboard",
        "assets":assets ,
        "repair_orders":repair_orders ,
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the work-order entry form.
        In edit mode (wo_pk present in URL kwargs): loads the existing
        unreviewed digital WorkOrder belonging to the operator, pre-fills
        all fields and passes edit_mode=True to the template so the POST
        handler knows to delete the original before creating the new one.
        In create mode: renders an empty form with one default block.
        Passes min_date to the template for client-side date enforcement.
        ---
        Renderiza el formulario de entrada de partes.
        En modo edicion (wo_pk en URL kwargs): carga el WorkOrder digital
        sin revisar del operario, prerellena todos los campos y pasa
        edit_mode=True al template para que el POST elimine el original
        antes de crear el nuevo. En modo creacion: formulario vacio.
        Pasa min_date para validacion de fecha en el lado cliente.
        """
        from work_order_processor .models import WorkOrder as _WO_E ,SparePartLine as _SPL_E 
        cu =self ._get_company_user (request )
        company =cu .company 
        min_date =_get_min_allowed_date (cu )
        wo_pk =kwargs .get ("wo_pk")

        if wo_pk is not None :


            try :
                wo_edit =_WO_E .objects .get (
                pk =wo_pk ,
                company =company ,
                uploaded_by =cu ,
                reviewed =False ,
                source__in =[
                _WO_E .Source .DIGITAL ,
                _WO_E .Source .GENERATED ,
                ],
                )
            except _WO_E .DoesNotExist :
                django_messages .error (
                request ,
                "El parte no existe, ya ha sido revisado o no te pertenece.",
                )
                return redirect ("/panel/operator/history/")



            entries =list (wo_edit .entries .prefetch_related ("lines").all ())
            first_entry =entries [0 ]if entries else None 
            fecha_str =(
            first_entry .work_date .strftime ("%Y-%m-%d")
            if first_entry and first_entry .work_date else ""
            )
            entradas_enriched =[]
            repuestos_enriched =[]
            ridx =1 
            for entry in entries :
                for line in entry .lines .order_by ("line_number"):
                    entradas_enriched .append ({
                    "idx":len (entradas_enriched )+1 ,
                    "machine_raw":line .machine_raw or "",
                    "machine_asset":line .machine_asset ,
                    "fault_description":line .fault_description or "",
                    "repair_notes":line .repair_notes or "",
                    "hc":line .hc .strftime ("%H:%M")if line .hc else "",
                    "hf":line .hf .strftime ("%H:%M")if line .hf else "",
                    "or_val":line .or_val or "",
                    "flags":line .flags or [],
                    "odometer_reading":float (line .odometer_reading )if line .odometer_reading is not None else "",
                    "engine_hours_reading":float (line .engine_hours_reading )if line .engine_hours_reading is not None else "",
                    "crane_hours_reading":float (line .crane_hours_reading )if line .crane_hours_reading is not None else "",
                    })
                    for spare in _SPL_E .objects .filter (entry_line =line ).order_by ("line_number"):
                        repuestos_enriched .append ({
                        "ridx":ridx ,
                        "referencia":spare .reference or "",
                        "vehiculo_raw":"",
                        "vehicle_asset":spare .vehicle ,
                        "material":spare .material or "",
                        "unidades":str (spare .quantity )if spare .quantity is not None else "",
                        "origen":spare .source or "WAREHOUSE",
                        "proveedor":spare .supplier or "",
                        "unit_price":str (spare .unit_price )if spare .unit_price is not None else "",
                        "flags":spare .flags or [],
                        })
                        ridx +=1 



            _schedule_edit =_resolve_operator_schedule (cu ,company )
            _lunch_start_edit =""
            _lunch_end_edit =""
            _first_hc_edit =""
            _show_lunch_edit =False 
            _end_time_morning_edit =""
            _end_time_afternoon_edit =""
            if _schedule_edit and not _schedule_edit .is_intensive :
                _show_lunch_edit =True 
                if _schedule_edit .end_time_morning :
                    _lunch_start_edit =_schedule_edit .end_time_morning .strftime ("%H:%M")
                    _end_time_morning_edit =_lunch_start_edit 
                if _schedule_edit .start_time_afternoon :
                    _lunch_end_edit =_schedule_edit .start_time_afternoon .strftime ("%H:%M")
            if _schedule_edit and _schedule_edit .end_time_afternoon :
                _end_time_afternoon_edit =_schedule_edit .end_time_afternoon .strftime ("%H:%M")
            elif _schedule_edit and _schedule_edit .is_intensive and _schedule_edit .end_time_morning :
                _end_time_afternoon_edit =_schedule_edit .end_time_morning .strftime ("%H:%M")
            if _schedule_edit and _schedule_edit .start_time_morning :
                _first_hc_edit =_schedule_edit .start_time_morning .strftime ("%H:%M")


            _first_hf_edit =_end_time_morning_edit if _end_time_morning_edit else _end_time_afternoon_edit 

            context =self ._get_context_base (request )
            context .update ({
            "edit_mode":True ,
            "edit_wo_pk":wo_pk ,
            "num_entradas":len (entradas_enriched )or 1 ,
            "num_repuestos":len (repuestos_enriched ),
            "fecha":fecha_str ,
            "entradas_enriched":entradas_enriched ,
            "repuestos_enriched":repuestos_enriched ,
            "min_date":min_date .isoformat ()if min_date else "",
            "lunch_break_start":_lunch_start_edit ,
            "lunch_break_end":_lunch_end_edit ,
            "first_block_hc":_first_hc_edit ,
            "first_block_hf":_first_hf_edit ,
            "show_lunch_break":_show_lunch_edit ,
            "end_time_morning":_end_time_morning_edit ,
            "end_time_afternoon":_end_time_afternoon_edit ,
            })
            return render (request ,self .template_name ,context )



        from datetime import date as _date_ip 
        from work_order_processor .models import WorkOrder as _WO_IP ,SparePartLine as _SPL_IP 
        _today_ip =_date_ip .today ()
        _in_progress_wo =_WO_IP .objects .filter (
        company =company ,
        uploaded_by =cu ,
        source =_WO_IP .Source .DIGITAL ,
        status =_WO_IP .Status .IN_PROGRESS ,
        ).order_by ("-upload_date").first ()

        if _in_progress_wo is not None :


            _ip_entries =list (_in_progress_wo .entries .prefetch_related ("lines").all ())
            _ip_first_entry =_ip_entries [0 ]if _ip_entries else None 
            _ip_fecha_str =(
            _ip_first_entry .work_date .strftime ("%Y-%m-%d")
            if _ip_first_entry and _ip_first_entry .work_date else ""
            )
            _ip_entradas =[]
            _ip_repuestos =[]
            _ip_ridx =1 
            for _ip_entry in _ip_entries :
                for _ip_line in _ip_entry .lines .order_by ("line_number"):
                    _ip_entradas .append ({
                    "idx":len (_ip_entradas )+1 ,
                    "machine_raw":_ip_line .machine_raw or "",
                    "machine_asset":_ip_line .machine_asset ,
                    "fault_description":_ip_line .fault_description or "",
                    "repair_notes":_ip_line .repair_notes or "",
                    "hc":_ip_line .hc .strftime ("%H:%M")if _ip_line .hc else "",
                    "hf":_ip_line .hf .strftime ("%H:%M")if _ip_line .hf else "",
                    "or_val":_ip_line .or_val or "",
                    "flags":_ip_line .flags or [],
                    "odometer_reading":float (_ip_line .odometer_reading )if _ip_line .odometer_reading is not None else "",
                    "engine_hours_reading":float (_ip_line .engine_hours_reading )if _ip_line .engine_hours_reading is not None else "",
                    "crane_hours_reading":float (_ip_line .crane_hours_reading )if _ip_line .crane_hours_reading is not None else "",
                    })
                    for _ip_spare in _SPL_IP .objects .filter (entry_line =_ip_line ).order_by ("line_number"):
                        _ip_repuestos .append ({
                        "ridx":_ip_ridx ,
                        "referencia":_ip_spare .reference or "",
                        "vehiculo_raw":"",
                        "vehicle_asset":_ip_spare .vehicle ,
                        "material":_ip_spare .material or "",
                        "unidades":str (_ip_spare .quantity )if _ip_spare .quantity is not None else "",
                        "origen":_ip_spare .source or "WAREHOUSE",
                        "proveedor":_ip_spare .supplier or "",
                        "unit_price":str (_ip_spare .unit_price )if _ip_spare .unit_price is not None else "",
                        "flags":_ip_spare .flags or [],
                        })
                        _ip_ridx +=1 


            _schedule_ip =_resolve_operator_schedule (cu ,company )
            _ip_lunch_start =""
            _ip_lunch_end =""
            _ip_first_hc =""
            _ip_show_lunch =False 
            _ip_end_time_morning =""
            _ip_end_time_afternoon =""
            if _schedule_ip and not _schedule_ip .is_intensive :
                _ip_show_lunch =True 
                if _schedule_ip .end_time_morning :
                    _ip_lunch_start =_schedule_ip .end_time_morning .strftime ("%H:%M")
                    _ip_end_time_morning =_ip_lunch_start 
                if _schedule_ip .start_time_afternoon :
                    _ip_lunch_end =_schedule_ip .start_time_afternoon .strftime ("%H:%M")
            if _schedule_ip and _schedule_ip .end_time_afternoon :
                _ip_end_time_afternoon =_schedule_ip .end_time_afternoon .strftime ("%H:%M")
            elif _schedule_ip and _schedule_ip .is_intensive and _schedule_ip .end_time_morning :
                _ip_end_time_afternoon =_schedule_ip .end_time_morning .strftime ("%H:%M")
            if _schedule_ip and _schedule_ip .start_time_morning :
                _ip_first_hc =_schedule_ip .start_time_morning .strftime ("%H:%M")
            _ip_first_hf =_ip_end_time_morning if _ip_end_time_morning else _ip_end_time_afternoon 
            _ip_no_lunch =_ip_first_entry .no_lunch_break if _ip_first_entry else False 
            logger .info (
            "# [I2-DIAG-GET] _ip_first_entry=%r lb_start_bd=%r lb_end_bd=%r "
            "no_lunch_bd=%r _ip_lunch_start=%r _ip_lunch_end=%r",
            _ip_first_entry ,
            _ip_first_entry .lunch_break_start if _ip_first_entry else None ,
            _ip_first_entry .lunch_break_end if _ip_first_entry else None ,
            _ip_first_entry .no_lunch_break if _ip_first_entry else None ,
            _ip_lunch_start ,_ip_lunch_end ,
            )
            context =self ._get_context_base (request )
            from ivr_config .models import AbsenceCategory as _AbsCat 
            from fleet .models import MachineAsset as _MA 
            from work_order_processor .management .commands .seed_personal_asset import PERSONAL_ASSET_CODE 
            _absence_cats =list (
            _AbsCat .objects .filter (company =company ,is_active =True )
            .order_by ("order","label")
            .values ("id","label","requires_note")
            )
            context .update ({
            "in_progress_mode":True ,
            "in_progress_wo_pk":_in_progress_wo .pk ,
            "num_entradas":len (_ip_entradas ),
            "num_repuestos":len (_ip_repuestos ),
            "fecha":_ip_fecha_str ,
            "entradas_enriched":_ip_entradas ,
            "repuestos_enriched":_ip_repuestos ,
            "min_date":min_date .isoformat ()if min_date else "",
            "lunch_break_start":_ip_first_entry .lunch_break_start .strftime ("%H:%M")if _ip_first_entry and _ip_first_entry .lunch_break_start else "",
            "lunch_break_end":_ip_first_entry .lunch_break_end .strftime ("%H:%M")if _ip_first_entry and _ip_first_entry .lunch_break_end else "",
            "first_block_hc":_ip_first_hc ,
            "first_block_hf":_ip_first_hf ,
            "no_lunch_break":_ip_no_lunch ,
            "show_lunch_break":_ip_show_lunch ,
            "end_time_morning":_ip_end_time_morning ,
            "end_time_afternoon":_ip_end_time_afternoon ,
            "absence_categories":_json_fix .dumps (_absence_cats ),
            "personal_asset_code":PERSONAL_ASSET_CODE ,
            })
            return render (request ,self .template_name ,context )








        _schedule_create =_resolve_operator_schedule (cu ,company )
        _lunch_start =""
        _lunch_end =""
        _first_hc =""
        _end_time_morning_create =""
        _end_time_afternoon_create =""
        if _schedule_create is not None :
            if not _schedule_create .is_intensive :


                if _schedule_create .end_time_morning :
                    _lunch_start =_schedule_create .end_time_morning .strftime ("%H:%M")
                if _schedule_create .start_time_afternoon :
                    _lunch_end =_schedule_create .start_time_afternoon .strftime ("%H:%M")


            if _schedule_create .start_time_morning :
                _first_hc =_schedule_create .start_time_morning .strftime ("%H:%M")
            if not _schedule_create .is_intensive and _schedule_create .end_time_morning :
                _end_time_morning_create =_schedule_create .end_time_morning .strftime ("%H:%M")
            if _schedule_create .end_time_afternoon :
                _end_time_afternoon_create =_schedule_create .end_time_afternoon .strftime ("%H:%M")
            elif _schedule_create .is_intensive and _schedule_create .end_time_morning :
                _end_time_afternoon_create =_schedule_create .end_time_morning .strftime ("%H:%M")

        import json as _json_fix 
        from ivr_config .models import AbsenceCategory as _AbsCat 
        from work_order_processor .management .commands .seed_personal_asset import PERSONAL_ASSET_CODE 
        _absence_cats =list (
        _AbsCat .objects .filter (company =company ,is_active =True )
        .order_by ("order","label")
        .values ("id","label","requires_note")
        )
        context =self ._get_context_base (request )
        context .update ({
        "num_entradas":1 ,
        "num_repuestos":0 ,
        "fecha":"",
        "min_date":min_date .isoformat ()if min_date else "",
        "lunch_break_start":_lunch_start ,
        "lunch_break_end":_lunch_end ,
        "first_block_hc":_first_hc ,
        "first_block_hf":_end_time_morning_create if _end_time_morning_create else _end_time_afternoon_create ,
        "no_lunch_break":False ,
        "show_lunch_break":bool (_lunch_start ),
        "end_time_morning":_end_time_morning_create ,
        "end_time_afternoon":_end_time_afternoon_create ,
        "absence_categories":_json_fix .dumps (_absence_cats ),
        "personal_asset_code":PERSONAL_ASSET_CODE ,
        })
        return render (request ,self .template_name ,context )

    def post (self ,request ,*args ,**kwargs ):
        """
        Parses, validates and atomically persists the submitted form data.
        Applies the same integrity gate used by WorkOrderEntryConfirmView.post().
        On validation failure re-renders the form with error context, preserving
        all data already entered by the operator.
        ---
        Parsea, valida y persiste atomicamente los datos del formulario enviado.
        Aplica la misma barrera de integridad que WorkOrderEntryConfirmView.post().
        En caso de fallo re-renderiza el formulario con contexto de error,
        preservando todos los datos introducidos por el operario.
        """
        import json as _json 
        from datetime import datetime ,time as dt_time 
        from decimal import Decimal ,InvalidOperation 
        from django .db import transaction 
        from fleet .models import MachineAsset 
        from work_order_processor .models import (
        WorkOrder ,WorkOrderEntry ,WorkOrderEntryLine ,SparePartLine ,
        )
        from work_order_processor .services import (
        generate_work_order_excel ,
        _normalise_machine_code ,
        _compute_delta_hours ,
        )

        cu =self ._get_company_user (request )
        company =cu .company 
        POST =request .POST 





        from datetime import time as _dt_time_lb 
        def _parse_time_field (raw ):
            """
            Parses a HH:MM string into a datetime.time or None.
            ---
            Parsea una cadena HH:MM a datetime.time o None.
            """
            if not raw :
                return None 
            try :
                parts =raw .strip ().split (":")
                return _dt_time_lb (int (parts [0 ]),int (parts [1 ]))
            except (ValueError ,IndexError ):
                return None 

        _lb_start_raw =POST .get ("lunch_break_start","").strip ()
        _lb_end_raw =POST .get ("lunch_break_end","").strip ()
        _lb_start =_parse_time_field (_lb_start_raw )
        _lb_end =_parse_time_field (_lb_end_raw )


        _no_lunch_break =POST .get ("no_lunch_break","")=="1"


        _form_action =POST .get ("form_action","close_order").strip ()


        _post_schedule =_resolve_operator_schedule (cu ,company )
        _post_lb_start =""
        _post_lb_end =""
        _post_show_lunch =False 
        if _post_schedule and not _post_schedule .is_intensive :
            _post_show_lunch =True 
            if _post_schedule .end_time_morning :
                _post_lb_start =_post_schedule .end_time_morning .strftime ("%H:%M")
            if _post_schedule .start_time_afternoon :
                _post_lb_end =_post_schedule .start_time_afternoon .strftime ("%H:%M")
        logger .info (
        "# [I2-DIAG] lunch_break_start_raw=%r lb_start=%r "
        "lunch_break_end_raw=%r lb_end=%r no_lunch_break=%r form_action=%r",
        _lb_start_raw ,_lb_start ,_lb_end_raw ,_lb_end ,
        _no_lunch_break ,POST .get ("form_action",""),
        )




        fecha_str =POST .get ("fecha","").strip ()
        work_date =None 
        if fecha_str :
            for fmt in ("%d/%m/%Y","%Y-%m-%d"):
                try :
                    work_date =datetime .strptime (fecha_str ,fmt ).date ()
                    break 
                except ValueError :
                    continue 







        logger .info (
        "# [I2-DIAG-WORKDATE] work_date=%r form_action=%r",
        work_date ,_form_action ,
        )
        if work_date is not None :
            _min_date =_get_min_allowed_date (cu )
            logger .info (
            "# [I2-DIAG-MINDATE] work_date=%r _min_date=%r form_action=%r",
            work_date ,_min_date ,_form_action ,
            )
            if _min_date is not None and work_date <_min_date :
                from datetime import timedelta as _td_fd 
                _last_rev =_min_date -_td_fd (days =1 )
                context =self ._get_context_base (request )
                context .update ({
                "error":(
                f"No puedes introducir un parte con fecha "
                f"{work_date.strftime('%d/%m/%Y')}. "
                f"El ultimo parte revisado es del "
                f"{_last_rev.strftime('%d/%m/%Y')} y ya ha sido auditado. "
                f"La fecha minima permitida es "
                f"{_min_date.strftime('%d/%m/%Y')}."
                ),
                "fecha":fecha_str ,
                "entradas_enriched":[],
                "repuestos_enriched":[],
                "num_entradas":1 ,
                "num_repuestos":0 ,
                "min_date":_min_date .isoformat (),
                "lunch_break_start":_post_lb_start ,
                "lunch_break_end":_post_lb_end ,
                "show_lunch_break":_post_show_lunch ,
                })
                return render (request ,self .template_name ,context )









        _edit_wo_pk_pre =POST .get ("edit_wo_pk","").strip ()
        if _edit_wo_pk_pre :
            try :
                _wo_orig_pre =WorkOrder .objects .get (
                pk =int (_edit_wo_pk_pre ),
                company =company ,
                uploaded_by =cu ,
                reviewed =False ,
                source__in =[
                WorkOrder .Source .DIGITAL ,
                WorkOrder .Source .GENERATED ,
                ],
                )
                _wo_orig_pre .delete ()
            except (WorkOrder .DoesNotExist ,ValueError ):


                pass 















        if work_date is not None :
            from django .urls import reverse as _rev0 
            from work_order_processor .models import WorkOrder as _WO0 ,WorkOrderEntry as _WOE0 
            _existing_entry0 =_WOE0 .objects .filter (
            work_order__company =company ,
            work_order__uploaded_by =cu ,
            work_order__source__in =[
            _WO0 .Source .DIGITAL ,
            _WO0 .Source .GENERATED ,
            ],
            work_order__reviewed =False ,
            work_date =work_date ,
            ).select_related ("work_order").first ()

            if _existing_entry0 is not None :









                _is_own_in_progress =(
                _existing_entry0 .work_order .status ==_WO0 .Status .IN_PROGRESS 
                and _existing_entry0 .work_order .uploaded_by_id ==cu .pk 
                )
                if not _is_own_in_progress :
                    logger .info (
                    "# [I2-DIAG-GATE0] duplicado detectado. form_action=%r "
                    "existing_entry0_pk=%r is_own_in_progress=%r",
                    _form_action ,_existing_entry0 .pk ,_is_own_in_progress ,
                    )


                    _gate0_lines =_parse_entry_lines_from_post (POST ,company )
                    _gate0_spare =_parse_spare_parts_from_post (
                    POST ,company ,entry_lines_data =_gate0_lines 
                    )
                    request .session ["pending_merge_lines"]=_serialize_pending_lines (
                    _gate0_lines ,_gate0_spare ,work_date 
                    )
                    return redirect (
                    _rev0 (
                    "panel:operator_merge",
                    kwargs ={"entry_pk":_existing_entry0 .pk },
                    )
                    )











        entry_lines_data =_parse_entry_lines_from_post (POST ,company )
        spare_parts_data =_parse_spare_parts_from_post (
        POST ,company ,entry_lines_data =entry_lines_data 
        )





        integrity_errors =[]

        if not work_date :
            integrity_errors .append (
            "La fecha del parte es obligatoria y debe tener formato DD/MM/AAAA."
            )

        if not entry_lines_data :
            integrity_errors .append ("El parte debe contener al menos un bloque de trabajo.")

        for ld in entry_lines_data :
            blk =f"Bloque {ld['line_number']}"
            if not ld ["machine_raw"]:
                integrity_errors .append (f"{blk}: el codigo de maquina es obligatorio.")
            elif ld ["machine_asset"]is None :
                integrity_errors .append (
                f"{blk}: el codigo '{ld['machine_raw']}' no se ha podido "
                f"identificar en el catalogo de flota. Corrigelo antes de guardar."
                )
            if not ld ["hc"]:
                integrity_errors .append (f"{blk}: la hora de inicio (H.C.) es obligatoria.")
            if not ld ["hf"]:
                integrity_errors .append (f"{blk}: la hora de fin (H.F.) es obligatoria.")
            if ld ["hc"]and ld ["hf"]and ld ["delta_hours"]is not None :
                if ld ["delta_hours"]<=0 :
                    integrity_errors .append (
                    f"{blk}: la H.F. debe ser posterior a la H.C. "
                    f"(Delta horas calculado: {ld['delta_hours']})."
                    )
            if not ld ["fault_description"]:
                integrity_errors .append (
                f"{blk}: la descripcion de la averia es obligatoria."
                )
            if not ld ["repair_notes"]:
                integrity_errors .append (
                f"{blk}: la descripcion de la reparacion realizada es obligatoria."
                )







        for ld in entry_lines_data :
            if ld ["machine_asset"]is not None :
                asset =ld ["machine_asset"]
                blk =f"Bloque {ld['line_number']}"
                if asset .has_odometer :
                    reading =ld .get ("odometer_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de km (odometro) obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de km no puede ser cero para {asset.code} "
                        f"(ya tiene partes anteriores registrados)."
                        )
                if asset .has_engine_hours :
                    reading =ld .get ("engine_hours_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de horometro motor obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de horometro motor no puede ser cero "
                        f"para {asset.code} (ya tiene partes anteriores registrados)."
                        )
                if asset .has_crane_hours :
                    reading =ld .get ("crane_hours_reading")
                    if reading is None :
                        integrity_errors .append (
                        f"{blk}: lectura de horometro grua obligatoria para {asset.code}."
                        )
                    elif reading ==0 and not asset .first_repair :
                        integrity_errors .append (
                        f"{blk}: la lectura de horometro grua no puede ser cero "
                        f"para {asset.code} (ya tiene partes anteriores registrados)."
                        )

        for spd in spare_parts_data :
            rep =f"Repuesto {spd['line_number']}"
            if not spd ["material"]:
                integrity_errors .append (f"{rep}: la descripcion del material es obligatoria.")
            if spd ["quantity"]is None or spd ["quantity"]<=0 :
                integrity_errors .append (f"{rep}: las unidades deben ser un numero positivo.")

        if integrity_errors :
            logger .info (
            "# [I2-DIAG-INTEGRITY] form_action=%r errors=%r",
            _form_action if "_form_action"in dir ()else POST .get ("form_action",""),
            integrity_errors ,
            )
            entradas_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":[],
            }
            for ld in entry_lines_data 
            ]
            repuestos_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":[],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":" | ".join (integrity_errors ),
            "fecha":fecha_str ,
            "entradas_enriched":entradas_post ,
            "repuestos_enriched":repuestos_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            "lunch_break_start":_post_lb_start ,
            "lunch_break_end":_post_lb_end ,
            "show_lunch_break":_post_show_lunch ,
            })
            return render (request ,self .template_name ,context )





        if not POST .get ("save_confirmed"):
            entradas_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":[],
            }
            for ld in entry_lines_data 
            ]
            repuestos_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":[],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":None ,
            "fecha":fecha_str ,
            "entradas_enriched":entradas_post ,
            "repuestos_enriched":repuestos_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            "lunch_break_start":_post_lb_start ,
            "lunch_break_end":_post_lb_end ,
            "show_lunch_break":_post_show_lunch ,
            })
            return render (request ,self .template_name ,context )





        from work_order_processor .validators import (
        run_intra_part_validation ,
        parse_blocks_from_post ,
        validate_inter_overlap ,
        TimeBlock ,
        )

        num_entradas_post =int (POST .get ("num_entradas",len (entry_lines_data )))
        _blocks =parse_blocks_from_post (POST ,num_entradas_post ,entry_lines_data =entry_lines_data )
        _intra =run_intra_part_validation (_blocks )

        if not _intra .ok :


            _error_msgs =[e .message for e in _intra .errors ]
            if _intra .warnings :
                _error_msgs +=[f"[AVISO] {w.message}"for w in _intra .warnings ]
            entradas_post =[
            {
            "idx":ld ["line_number"],
            "machine_raw":ld ["machine_raw"],
            "machine_asset":ld ["machine_asset"],
            "fault_description":ld ["fault_description"],
            "repair_notes":ld ["repair_notes"],
            "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
            "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
            "or_val":ld ["or_val"],
            "flags":[],
            }
            for ld in entry_lines_data 
            ]
            repuestos_post =[
            {
            "ridx":spd ["line_number"],
            "referencia":spd ["referencia"],
            "vehiculo_raw":spd ["vehiculo_raw"],
            "vehicle_asset":spd ["vehicle_asset"],
            "material":spd ["material"],
            "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
            "origen":spd ["source"],
            "proveedor":spd ["supplier"],
            "flags":[],
            }
            for spd in spare_parts_data 
            ]
            context =self ._get_context_base (request )
            context .update ({
            "error":" | ".join (_error_msgs ),
            "fecha":fecha_str ,
            "entradas_enriched":entradas_post ,
            "repuestos_enriched":repuestos_post ,
            "num_entradas":len (entry_lines_data ),
            "num_repuestos":len (spare_parts_data ),
            "lunch_break_start":_post_lb_start ,
            "lunch_break_end":_post_lb_end ,
            "show_lunch_break":_post_show_lunch ,
            })
            return render (request ,self .template_name ,context )



        _meter_warnings =[w .message for w in _intra .warnings ]













        if work_date is not None :
            from decimal import Decimal as _Dec_C 
            from ivr_config .models import WorkerAbsence as _WA_C 
            _total_hours_c =sum (
            (ld ["delta_hours"]for ld in entry_lines_data if ld ["delta_hours"]is not None ),
            _Dec_C ("0"),
            )
            _has_absence_c =_WA_C .objects .filter (
            company_user =cu ,
            start_date__lte =work_date ,
            end_date__gte =work_date ,
            ).exists ()
            if _total_hours_c <_Dec_C ("8")and not _has_absence_c :
                _missing_c =_Dec_C ("8")-_total_hours_c 
                entradas_post_c =[
                {
                "idx":ld ["line_number"],
                "machine_raw":ld ["machine_raw"],
                "machine_asset":ld ["machine_asset"],
                "fault_description":ld ["fault_description"],
                "repair_notes":ld ["repair_notes"],
                "hc":ld ["hc"].strftime ("%H:%M")if ld ["hc"]else "",
                "hf":ld ["hf"].strftime ("%H:%M")if ld ["hf"]else "",
                "or_val":ld ["or_val"],
                "flags":[],
                }
                for ld in entry_lines_data 
                ]
                repuestos_post_c =[
                {
                "ridx":spd ["line_number"],
                "referencia":spd ["referencia"],
                "vehiculo_raw":spd ["vehiculo_raw"],
                "vehicle_asset":spd ["vehicle_asset"],
                "material":spd ["material"],
                "unidades":str (spd ["quantity"])if spd ["quantity"]is not None else "",
                "origen":spd ["source"],
                "proveedor":spd ["supplier"],
                "flags":[],
                }
                for spd in spare_parts_data 
                ]
                context =self ._get_context_base (request )
                context .update ({
                "error":(
                f"La jornada del parte suma {_total_hours_c} h, "
                f"pero se requieren al menos 8 h. "
                f"Faltan {_missing_c} h para completar la jornada. "
                f"Añade los bloques de trabajo que faltan o registra "
                f"una ausencia justificada para esta fecha."
                ),
                "fecha":fecha_str ,
                "entradas_enriched":entradas_post_c ,
                "repuestos_enriched":repuestos_post_c ,
                "num_entradas":len (entry_lines_data ),
                "num_repuestos":len (spare_parts_data ),
                "min_date":_get_min_allowed_date (cu ).isoformat ()if _get_min_allowed_date (cu )else "",
                })
                return render (request ,self .template_name ,context )










        if work_date is not None and _form_action !="save_blocks":
            from ivr_config .models import WorkdaySchedule as _WDS_FA 
            from django .urls import reverse as _rev_g4fa 
            _schedule_g4fa =(
            cu .workday_schedule 
            if cu .workday_schedule_id 
            else (




            next (
            (
            _sec .workday_schedule 
            for _sec in (
            _contact_g4fa .sections 
            .filter (is_active =True ,workday_schedule__isnull =False )
            .select_related ("workday_schedule")
            .order_by ("name")
            )
            ),
            None ,
            )
            if (_contact_g4fa :=(
            Contact .objects .filter (company_user =cu )
            .prefetch_related ("sections__workday_schedule")
            .first ()
            ))is not None 
            else None 
            )or _WDS_FA .objects .filter (
            company =company ,is_default =True 
            ).first ()
            )




            _gaps_g4fa =[]if _no_lunch_break else _detect_workday_gaps (
            entry_lines_data ,_schedule_g4fa ,work_date 
            )
            if _gaps_g4fa :
                from django .db import transaction as _tx_g4fa 
                _worker_name_g4fa =(
                cu .user .get_full_name ()or cu .user .username 
                ).upper ()
                _date_tag_g4fa =work_date .strftime ("%d-%m-%Y")
                _synth_g4fa =f"{_worker_name_g4fa}_{_date_tag_g4fa}_DRAFT.pdf"
                with _tx_g4fa .atomic ():
                    _wo_draft_fa =WorkOrder (
                    company =company ,
                    uploaded_by =cu ,
                    source =WorkOrder .Source .DIGITAL ,
                    status =WorkOrder .Status .PENDING_GAPS ,
                    total_pages =1 ,
                    processed_pages =1 ,
                    reviewed =False ,
                    )
                    _wo_draft_fa .source_pdf .name =_synth_g4fa 
                    _wo_draft_fa .save ()
                    _entry_draft_fa =WorkOrderEntry .objects .create (
                    work_order =_wo_draft_fa ,
                    page_number =1 ,
                    worker_name =_worker_name_g4fa ,
                    work_date =work_date ,
                    uncertain_date =False ,
                    extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                    raw_gemini_response =None ,
                    )
                    _line_num_g4fa =1 
                    for _ld in entry_lines_data :
                        WorkOrderEntryLine .objects .create (
                        entry =_entry_draft_fa ,
                        line_number =_line_num_g4fa ,
                        machine_asset =_ld .get ("machine_asset"),
                        machine_raw =_ld .get ("machine_raw",""),
                        machine_norm ="",
                        fault_description =_ld .get ("fault_description",""),
                        repair_notes =_ld .get ("repair_notes",""),
                        hc =_ld .get ("hc"),
                        hf =_ld .get ("hf"),
                        or_val =_ld .get ("or_val",""),
                        delta_hours =_ld .get ("delta_hours"),
                        flags =[],
                        odometer_reading =_ld .get ("odometer_reading"),
                        engine_hours_reading =_ld .get ("engine_hours_reading"),
                        crane_hours_reading =_ld .get ("crane_hours_reading"),
                        )
                        _line_num_g4fa +=1 
                    from work_order_processor .models import WorkdayGap as _WDG_FA 
                    for _gap in _gaps_g4fa :
                        _WDG_FA .objects .create (
                        work_order =_wo_draft_fa ,
                        gap_type =_gap ["gap_type"],
                        gap_start =_gap ["gap_start"],
                        gap_end =_gap ["gap_end"],
                        duration_minutes =_gap ["duration_minutes"],
                        )
                request .session ["pending_gaps_wo_pk"]=_wo_draft_fa .pk 
                request .session .modified =True 
                logger .info (
                "# [FormView/Gate4] PENDING_GAPS borrador pk=%d creado. "
                "%d gap(s) detectado(s). Redirigiendo a resolución.",
                _wo_draft_fa .pk ,len (_gaps_g4fa ),
                )
                return redirect (
                _rev_g4fa (
                "panel:operator_gap_resolution",
                kwargs ={"wo_draft_pk":_wo_draft_fa .pk },
                )
                )


















        if _form_action =="save_blocks":
            logger .info (
            "# [I2-DIAG-PRE] entrando en save_blocks. work_date=%r "
            "num_entry_lines=%r _lb_start=%r _lb_end=%r",
            work_date ,len (entry_lines_data )if entry_lines_data else -1 ,
            _lb_start ,_lb_end ,
            )
            from django .db import transaction as _tx_ip 
            from work_order_processor .models import (
            WorkOrder as _WO_IP2 ,
            WorkOrderEntry as _WOE_IP2 ,
            WorkOrderEntryLine as _WOEL_IP2 ,
            SparePartLine as _SPL_IP2 ,
            )
            _ip_wo_pk_post =POST .get ("in_progress_wo_pk","").strip ()
            try :
                with _tx_ip .atomic ():
                    _worker_name_ip =(
                    cu .user .get_full_name ()or cu .user .username 
                    ).upper ()
                    _date_tag_ip =(
                    work_date .strftime ("%d-%m-%Y")if work_date else "SIN-FECHA"
                    )
                    _synth_ip =f"{_worker_name_ip}_{_date_tag_ip}.pdf"



                    if _ip_wo_pk_post :
                        try :
                            _ip_wo =_WO_IP2 .objects .get (
                            pk =int (_ip_wo_pk_post ),
                            company =company ,
                            uploaded_by =cu ,
                            status =_WO_IP2 .Status .IN_PROGRESS ,
                            )
                        except (_WO_IP2 .DoesNotExist ,ValueError ):
                            _ip_wo =None 
                    else :
                        _ip_wo =_WO_IP2 .objects .filter (
                        company =company ,
                        uploaded_by =cu ,
                        status =_WO_IP2 .Status .IN_PROGRESS ,
                        ).first ()

                    if _ip_wo is None :


                        _ip_wo =_WO_IP2 (
                        company =company ,
                        uploaded_by =cu ,
                        source =_WO_IP2 .Source .DIGITAL ,
                        status =_WO_IP2 .Status .IN_PROGRESS ,
                        total_pages =1 ,
                        processed_pages =1 ,
                        reviewed =False ,
                        )
                        _ip_wo .source_pdf .name =_synth_ip 
                        _ip_wo .save ()
                        _ip_entry =_WOE_IP2 .objects .create (
                        work_order =_ip_wo ,
                        page_number =1 ,
                        worker_name =_worker_name_ip ,
                        work_date =work_date ,
                        uncertain_date =False ,
                        extraction_confidence =_WOE_IP2 .Confidence .HIGH ,
                        raw_gemini_response =None ,
                        lunch_break_start =None if _no_lunch_break else _lb_start ,
                        lunch_break_end =None if _no_lunch_break else _lb_end ,
                        no_lunch_break =_no_lunch_break ,
                        )
                    else :




                        _ip_entry =_ip_wo .entries .first ()
                        if _ip_entry is None :
                            _ip_entry =_WOE_IP2 .objects .create (
                            work_order =_ip_wo ,
                            page_number =1 ,
                            worker_name =_worker_name_ip ,
                            work_date =work_date ,
                            uncertain_date =False ,
                            extraction_confidence =_WOE_IP2 .Confidence .HIGH ,
                            raw_gemini_response =None ,
                            lunch_break_start =None if _no_lunch_break else _lb_start ,
                            lunch_break_end =None if _no_lunch_break else _lb_end ,
                            no_lunch_break =_no_lunch_break ,
                            )
                        else :


                            _ip_entry .lunch_break_start =None if _no_lunch_break else _lb_start 
                            _ip_entry .lunch_break_end =None if _no_lunch_break else _lb_end 
                            _ip_entry .no_lunch_break =_no_lunch_break 
                            _ip_entry .save (update_fields =[
                            "lunch_break_start","lunch_break_end","no_lunch_break"
                            ])


                        _ip_entry .lines .all ().delete ()






























                    from decimal import Decimal as _Dec_ip 
                    from work_order_processor .services import _compute_delta_hours as _cdh_ip 


                    _ip_eff_lb_start =_lb_start 
                    _ip_eff_lb_end =_lb_end 
                    if not _no_lunch_break and (
                    _ip_eff_lb_start is None or _ip_eff_lb_end is None 
                    ):


                        if _post_schedule and not _post_schedule .is_intensive :
                            if _ip_eff_lb_start is None :
                                _ip_eff_lb_start =_post_schedule .end_time_morning 
                            if _ip_eff_lb_end is None :
                                _ip_eff_lb_end =_post_schedule .start_time_afternoon 
                    def _ip_to_min (t ):
                        """
                        Converts a time object to total minutes from midnight.
                        Returns 0 when t is None.
                        ---
                        Convierte un objeto time a minutos totales desde medianoche.
                        Devuelve 0 cuando t es None.
                        """
                        return t .hour *60 +t .minute if t is not None else 0 
                    def _ip_calc_overlap (hc_t ,hf_t ,lb_s ,lb_e ,no_lb ):
                        """
                        Computes the lunch overlap in minutes between a work block
                        [hc_t, hf_t] and the lunch window [lb_s, lb_e].
                        Returns 0 if no_lb is True or any value is None.
                        ---
                        Calcula el solapamiento de la pausa de comida en minutos entre
                        un bloque de trabajo [hc_t, hf_t] y la ventana [lb_s, lb_e].
                        Devuelve 0 si no_lb es True o algún valor es None.
                        """
                        if no_lb or hc_t is None or hf_t is None or lb_s is None or lb_e is None :
                            return 0 
                        return max (
                        0 ,
                        min (_ip_to_min (hf_t ),_ip_to_min (lb_e ))
                        -max (_ip_to_min (hc_t ),_ip_to_min (lb_s )),
                        )











                    _ip_pause_changed =(
                    _ip_entry .no_lunch_break !=_no_lunch_break 
                    or _ip_entry .lunch_break_start !=(
                    None if _no_lunch_break else _ip_eff_lb_start 
                    )
                    or _ip_entry .lunch_break_end !=(
                    None if _no_lunch_break else _ip_eff_lb_end 
                    )
                    )
                    if _ip_pause_changed :
                        logger .info (
                        "# [FormView/save_blocks] Pausa de comida modificada. "
                        "Recalculando líneas existentes. "
                        "entry_pk=%r old_lb_start=%r old_lb_end=%r old_no_lb=%r "
                        "new_lb_start=%r new_lb_end=%r new_no_lb=%r",
                        _ip_entry .pk ,
                        _ip_entry .lunch_break_start ,_ip_entry .lunch_break_end ,
                        _ip_entry .no_lunch_break ,
                        _ip_eff_lb_start ,_ip_eff_lb_end ,_no_lunch_break ,
                        )
                        for _ip_existing_line in _ip_entry .lines .all ():
                            _ip_ex_hc =_ip_existing_line .hc 
                            _ip_ex_hf =_ip_existing_line .hf 
                            if _ip_ex_hc is None or _ip_ex_hf is None :


                                continue 


                            _ip_gross =_cdh_ip (_ip_ex_hc ,_ip_ex_hf ,deduct_lunch =False )
                            if _ip_gross is None :
                                continue 


                            _ip_ex_overlap =_ip_calc_overlap (
                            _ip_ex_hc ,_ip_ex_hf ,
                            _ip_eff_lb_start ,_ip_eff_lb_end ,
                            _no_lunch_break ,
                            )
                            if _ip_ex_overlap >0 :
                                _ip_new_delta =_Dec_ip (str (_ip_gross ))-_Dec_ip (_ip_ex_overlap )/_Dec_ip ("60")
                                _ip_new_delta =max (_Dec_ip ("0"),_ip_new_delta )
                            else :
                                _ip_new_delta =_Dec_ip (str (_ip_gross ))
                            _ip_existing_line .delta_hours =_ip_new_delta 
                            _ip_existing_line .save (update_fields =["delta_hours"])
                            logger .info (
                            "# [FormView/save_blocks] Línea pk=%r recalculada. "
                            "hc=%r hf=%r gross=%r overlap_min=%r new_delta=%r",
                            _ip_existing_line .pk ,
                            _ip_ex_hc ,_ip_ex_hf ,_ip_gross ,
                            _ip_ex_overlap ,_ip_new_delta ,
                            )


                    _ip_created_lines ={}
                    for _ip_ld in entry_lines_data :
                        _ip_line_num =_ip_ld ["line_number"]
                        _ip_hc =_ip_ld .get ("hc")
                        _ip_hf =_ip_ld .get ("hf")
                        _ip_overlap_min =_ip_calc_overlap (
                        _ip_hc ,_ip_hf ,
                        _ip_eff_lb_start ,_ip_eff_lb_end ,
                        _no_lunch_break ,
                        )
                        _ip_delta_raw =_ip_ld ["delta_hours"]
                        if _ip_delta_raw is not None and _ip_overlap_min >0 :
                            _ip_delta_net =_Dec_ip (str (_ip_delta_raw ))-_Dec_ip (_ip_overlap_min )/_Dec_ip ("60")
                            _ip_delta_net =max (_Dec_ip ("0"),_ip_delta_net )
                        else :
                            _ip_delta_net =_ip_delta_raw 
                        _ip_line_obj =_WOEL_IP2 .objects .create (
                        entry =_ip_entry ,
                        line_number =_ip_line_num ,
                        machine_asset =_ip_ld .get ("machine_asset"),
                        machine_raw =_ip_ld .get ("machine_raw",""),
                        machine_norm =_ip_ld .get ("machine_norm",""),
                        fault_description =_ip_ld .get ("fault_description",""),
                        repair_notes =_ip_ld .get ("repair_notes",""),
                        hc =_ip_ld .get ("hc"),
                        hf =_ip_ld .get ("hf"),
                        or_val =_ip_ld .get ("or_val",""),
                        delta_hours =_ip_delta_net ,
                        flags =[],
                        odometer_reading =_ip_ld .get ("odometer_reading"),
                        engine_hours_reading =_ip_ld .get ("engine_hours_reading"),
                        crane_hours_reading =_ip_ld .get ("crane_hours_reading"),
                        )
                        _ip_created_lines [_ip_line_num ]=_ip_line_obj 



                    for _ip_spd in spare_parts_data :
                        _ip_target =next (iter (_ip_created_lines .values ()),None )
                        if _ip_target is None :
                            continue 
                        _SPL_IP2 .objects .create (
                        entry_line =_ip_target ,
                        line_number =_ip_spd ["line_number"],
                        reference =_ip_spd ["referencia"],
                        vehicle =_ip_spd ["vehicle_asset"],
                        material =_ip_spd ["material"],
                        quantity =_ip_spd ["quantity"],
                        source =_ip_spd ["source"],
                        supplier =_ip_spd ["supplier"],
                        flags =_ip_spd ["flags"],
                        )















                    from work_order_processor .models import WorkdayGap as _WDG_IP 
                    _WDG_IP .objects .filter (
                    work_order =_ip_wo ,
                    resolved =True ,
                    ).filter (
                    gap_type =_WDG_IP .GapType .GAP ,
                    ).delete ()
                    for _ip_ld_p in entry_lines_data :
                        if not _ip_ld_p .get ("is_personal"):
                            continue 
                        _ip_abs_cat =_ip_ld_p .get ("absence_category")
                        _ip_p_hc =_ip_ld_p .get ("hc")
                        _ip_p_hf =_ip_ld_p .get ("hf")
                        if _ip_p_hc is None or _ip_p_hf is None :
                            continue 
                        _ip_dur_min =max (
                        0 ,
                        (_ip_p_hf .hour *60 +_ip_p_hf .minute )
                        -(_ip_p_hc .hour *60 +_ip_p_hc .minute ),
                        )
                        _WDG_IP .objects .create (
                        work_order =_ip_wo ,
                        gap_type =_WDG_IP .GapType .GAP ,
                        gap_start =_ip_p_hc ,
                        gap_end =_ip_p_hf ,
                        duration_minutes =_ip_dur_min ,
                        absence_category =_ip_abs_cat ,
                        note =_ip_ld_p .get ("repair_notes",""),
                        resolved =True ,
                        )
                        logger .info (
                        "# [FormView/save_blocks] WorkdayGap sintético creado. "
                        "entry_pk=%r gap_start=%r gap_end=%r absence_cat=%r",
                        _ip_entry .pk ,_ip_p_hc ,_ip_p_hf ,
                        _ip_abs_cat .label if _ip_abs_cat else None ,
                        )

                logger .info (
                "# [FormView/save_blocks] WorkOrder IN_PROGRESS #%d actualizado. "
                "Bloques guardados: %d. entry_pk=%r lb_start=%r lb_end=%r no_lunch=%r",
                _ip_wo .pk ,len (entry_lines_data ),
                _ip_entry .pk ,
                _ip_entry .lunch_break_start ,
                _ip_entry .lunch_break_end ,
                _ip_entry .no_lunch_break ,
                )
            except Exception as _exc_ip :
                logger .error (
                "# [FormView/save_blocks] Error en persistencia: %s",
                _exc_ip ,exc_info =True ,
                )
                context =self ._get_context_base (request )
                context ["error"]=(
                f"Error al guardar los bloques: {_exc_ip}. "
                "Por favor, inténtalo de nuevo."
                )
                return render (request ,self .template_name ,context )

            django_messages .success (
            request ,
            f"{len(entry_lines_data)} bloque(s) guardado(s). "
            "Puedes añadir más bloques o cerrar el parte cuando termines."
            )
            return redirect ("/panel/operator/form/")










        _ip_wo_pk_close =POST .get ("in_progress_wo_pk","").strip ()
        if not _ip_wo_pk_close :
            from work_order_processor .models import WorkOrder as _WO_CL 
            _ip_wo_close =_WO_CL .objects .filter (
            company =company ,
            uploaded_by =cu ,
            status =_WO_CL .Status .IN_PROGRESS ,
            ).first ()
            if _ip_wo_close :
                _ip_wo_pk_close =str (_ip_wo_close .pk )





        edit_wo_pk =POST .get ("edit_wo_pk","").strip ()or _ip_wo_pk_close 
        if edit_wo_pk :
            try :
                _wo_orig =WorkOrder .objects .get (
                pk =int (edit_wo_pk ),
                company =company ,
                uploaded_by =cu ,
                reviewed =False ,
                source__in =[
                WorkOrder .Source .DIGITAL ,
                WorkOrder .Source .GENERATED ,
                ],
                )
                _wo_orig .delete ()
            except (WorkOrder .DoesNotExist ,ValueError ):


                pass 

        try :
            with transaction .atomic ():
                worker_name =(
                cu .user .get_full_name ()or cu .user .username 
                ).upper ()

                work_order =WorkOrder (
                company =company ,
                uploaded_by =cu ,
                source =WorkOrder .Source .DIGITAL ,
                status =WorkOrder .Status .DONE ,
                total_pages =1 ,
                processed_pages =1 ,
                reviewed =False ,
                )









                date_tag =(
                work_date .strftime ("%d-%m-%Y")if work_date else "SIN-FECHA"
                )
                synthetic_name =f"{worker_name}_{date_tag}.pdf"

                work_order .source_pdf .name =synthetic_name 
                work_order .save ()

                entry =WorkOrderEntry .objects .create (
                work_order =work_order ,
                page_number =1 ,
                worker_name =worker_name ,
                work_date =work_date ,
                uncertain_date =False ,
                extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                raw_gemini_response =None ,
                lunch_break_start =None if _no_lunch_break else _lb_start ,
                lunch_break_end =None if _no_lunch_break else _lb_end ,
                no_lunch_break =_no_lunch_break ,
                )

                created_lines ={}
                created_line_pks =[]
















                from decimal import Decimal as _Dec_lb 


                _co_eff_lb_start =_lb_start 
                _co_eff_lb_end =_lb_end 
                if not _no_lunch_break and (
                _co_eff_lb_start is None or _co_eff_lb_end is None 
                ):


                    if _post_schedule and not _post_schedule .is_intensive :
                        if _co_eff_lb_start is None :
                            _co_eff_lb_start =_post_schedule .end_time_morning 
                        if _co_eff_lb_end is None :
                            _co_eff_lb_end =_post_schedule .start_time_afternoon 
                def _co_to_min (t ):
                    """
                    Converts a time object to total minutes from midnight.
                    Returns 0 when t is None.
                    ---
                    Convierte un objeto time a minutos totales desde medianoche.
                    Devuelve 0 cuando t es None.
                    """
                    return t .hour *60 +t .minute if t is not None else 0 
                def _co_calc_overlap (hc_t ,hf_t ,lb_s ,lb_e ,no_lb ):
                    """
                    Computes the lunch overlap in minutes between a work block
                    [hc_t, hf_t] and the lunch window [lb_s, lb_e].
                    Returns 0 if no_lb is True or any value is None.
                    ---
                    Calcula el solapamiento de la pausa de comida en minutos entre
                    un bloque de trabajo [hc_t, hf_t] y la ventana [lb_s, lb_e].
                    Devuelve 0 si no_lb es True o algún valor es None.
                    """
                    if no_lb or hc_t is None or hf_t is None or lb_s is None or lb_e is None :
                        return 0 
                    return max (
                    0 ,
                    min (_co_to_min (hf_t ),_co_to_min (lb_e ))
                    -max (_co_to_min (hc_t ),_co_to_min (lb_s )),
                    )
                for ld in entry_lines_data :
                    _line_num =ld ["line_number"]
                    _co_hc =ld .get ("hc")
                    _co_hf =ld .get ("hf")
                    _overlap_min =_co_calc_overlap (
                    _co_hc ,_co_hf ,
                    _co_eff_lb_start ,_co_eff_lb_end ,
                    _no_lunch_break ,
                    )
                    _delta_raw =ld ["delta_hours"]
                    if _delta_raw is not None and _overlap_min >0 :
                        _delta_net =_Dec_lb (str (_delta_raw ))-_Dec_lb (_overlap_min )/_Dec_lb ("60")
                        _delta_net =max (_Dec_lb ("0"),_delta_net )
                    else :
                        _delta_net =_delta_raw 

                    line =WorkOrderEntryLine .objects .create (
                    entry =entry ,
                    line_number =ld ["line_number"],
                    machine_asset =ld ["machine_asset"],
                    machine_raw =ld ["machine_raw"],
                    machine_norm =ld ["machine_norm"],
                    fault_description =ld ["fault_description"],
                    repair_notes =ld ["repair_notes"],
                    hc =ld ["hc"],
                    hf =ld ["hf"],
                    or_val =ld ["or_val"],
                    delta_hours =_delta_net ,
                    flags =ld ["flags"],
                    odometer_reading =ld .get ("odometer_reading"),
                    engine_hours_reading =ld .get ("engine_hours_reading"),
                    crane_hours_reading =ld .get ("crane_hours_reading"),
                    )
                    created_lines [ld ["line_number"]]=line 
                    created_line_pks .append (line .pk )

                for spd in spare_parts_data :




                    target_line =next (iter (created_lines .values ()),None )
                    if target_line is None :
                        continue 
                    SparePartLine .objects .create (
                    entry_line =target_line ,
                    line_number =spd ["line_number"],
                    reference =spd ["referencia"],
                    vehicle =spd ["vehicle_asset"],
                    material =spd ["material"],
                    quantity =spd ["quantity"],
                    source =spd ["source"],
                    supplier =spd ["supplier"],
                    flags =spd ["flags"],
                    )





            if any (spd .get ("cg_incident")for spd in spare_parts_data ):
                WorkOrder .objects .filter (pk =work_order .pk ).update (has_cg_incident =True )
                logger .warning (
                "# [FormView] WorkOrder #%d marcado con has_cg_incident=True: "
                "al menos un repuesto tiene un CdG no resuelto en catálogo.",
                work_order .pk ,
                )

            logger .info (
            "# [FormView] WorkOrder #%d creado correctamente (Via A). "
            "Bloques: %d | Repuestos: %d.",
            work_order .pk ,
            len (entry_lines_data ),
            len (spare_parts_data ),
            )





            import json as _json_mod 
            _zero_raw =POST .get ("zero_meters_confirmed","").strip ()
            if _zero_raw :
                try :
                    _zero_data =_json_mod .loads (_zero_raw )
                    for _bIdx_str ,_meter_list in _zero_data .items ():
                        try :
                            _bIdx =int (_bIdx_str )
                        except (ValueError ,TypeError ):
                            continue 
                        _line =created_lines .get (_bIdx )
                        if _line is None :
                            continue 
                        _asset =_line .machine_asset 
                        _line_fields =[]
                        _asset_fields =[]
                        for _m in _meter_list :
                            _name =_m .get ("name","")
                            if "odometer"in _name :
                                _line .odometer_reading =None 
                                _line_fields .append ("odometer_reading")
                                if _asset and _asset .has_odometer :
                                    _asset .has_odometer =False 
                                    _asset_fields .append ("has_odometer")
                            elif "engine_hours"in _name :
                                _line .engine_hours_reading =None 
                                _line_fields .append ("engine_hours_reading")
                                if _asset and _asset .has_engine_hours :
                                    _asset .has_engine_hours =False 
                                    _asset_fields .append ("has_engine_hours")
                            elif "crane_hours"in _name :
                                _line .crane_hours_reading =None 
                                _line_fields .append ("crane_hours_reading")
                                if _asset and _asset .has_crane_hours :
                                    _asset .has_crane_hours =False 
                                    _asset_fields .append ("has_crane_hours")
                        if _line_fields :
                            _line .save (update_fields =_line_fields )
                        if _asset and _asset_fields :
                            _asset .save (update_fields =list (set (_asset_fields )))
                            logger .info (
                            "# [FormView] MachineAsset %s: flags desactivados: %s.",
                            _asset .code ,_asset_fields ,
                            )
                except (_json_mod .JSONDecodeError ,Exception )as _ze :
                    logger .warning (
                    "# [FormView] Error procesando zero_meters_confirmed: %s",_ze 
                    )



            for ld in entry_lines_data :
                _asset =ld .get ("machine_asset")
                if _asset and _asset .first_repair :
                    _asset .first_repair =False 
                    _asset .save (update_fields =["first_repair"])
                    logger .info (
                    "# [FormView] MachineAsset %s: first_repair=False.",
                    _asset .code ,
                    )

        except Exception as exc :
            logger .error (
            "# [FormView] Error en persistencia atomica: %s",exc ,exc_info =True 
            )
            context =self ._get_context_base (request )
            context ["error"]=(
            f"Error al guardar el parte: {exc}. "
            "Por favor, intentalo de nuevo o contacta con el administrador."
            )
            return render (request ,self .template_name ,context )










        for _lpk in created_line_pks :
            _line_obj =WorkOrderEntryLine .objects .filter (pk =_lpk ).first ()
            if _line_obj is None :
                continue 
            _cached =find_cached_classification (
            fault_description =_line_obj .fault_description ,
            repair_notes =_line_obj .repair_notes ,
            company =company ,
            )
            if _cached :
                WorkOrderEntryLine .objects .filter (pk =_lpk ).update (
                fault_category =_cached [0 ],
                fault_subcategory =_cached [1 ],
                )
                logger .info (
                "# [FormView] Clasificación copiada desde caché para "
                "WorkOrderEntryLine pk=%d: category=%s subcategory=%s.",
                _lpk ,_cached [0 ],_cached [1 ],
                )
            else :
                classify_fault_line .apply_async (
                args =[_lpk ],
                queue ="work_orders",
                )
                logger .info (
                "# [FormView] classify_fault_line encolada para "
                "WorkOrderEntryLine pk=%d.",
                _lpk ,
                )




        try :
            generate_work_order_excel (work_order .pk )
            logger .info (
            "# [FormView] Excel generado correctamente para WorkOrder #%d.",
            work_order .pk ,
            )
        except Exception as exc :
            logger .warning (
            "# [FormView] Excel no generado para WorkOrder #%d: %s.",
            work_order .pk ,exc ,
            )





        _inter =validate_inter_overlap (
        company_user =cu ,
        work_date =work_date ,
        blocks =_blocks ,
        exclude_work_order_pk =work_order .pk ,
        )

        if _inter .has_overlap :


            WorkOrder .objects .filter (
            pk__in =[work_order .pk ]+_inter .conflicting_ids 
            ).update (has_overlap_incident =True )
            logger .warning (
            "# [FormView] Solapamiento inter-parte detectado. "
            "WorkOrder #%d solapa con: %s.",
            work_order .pk ,
            _inter .conflicting_ids ,
            )
            django_messages .warning (
            request ,
            f"Parte #{work_order.pk} guardado con incidencia de solapamiento."
            )
            context =self ._get_context_base (request )
            context .update ({
            "overlap_incidents":True ,
            "new_work_order_pk":work_order .pk ,
            "conflicting_parts":[
            {"pk":pk ,"fecha":fecha }
            for pk ,fecha in zip (
            _inter .conflicting_ids ,
            _inter .conflicting_dates ,
            )
            ],
            "part_saved":True ,
            })
            return render (request ,self .template_name ,context )

        if _meter_warnings :
            django_messages .warning (
            request ,
            "Parte guardado con avisos de contadores: "+" | ".join (_meter_warnings ),
            )

        django_messages .success (
        request ,
        f"Parte de trabajo registrado correctamente (#{work_order.pk}). "
        f"El informe Excel está disponible en la lista de partes."
        )






        from django .urls import reverse as _reverse_co 
        if cu .role =="WORKSHOP":
            return redirect (_reverse_co ("panel:operator_history"))
        return redirect (_reverse_co ("panel:digital_work_order_list"))


class WorkOrderEntryHistoryView (WorkshopRequiredMixin ,View ):
    """
    Four-tab personal history view for WORKSHOP role operators.
    ADMIN and SUPERVISOR are automatically redirected to WorkOrderAdminHistoryView.

    Tabs:
      1 — Periodo actual  : work orders within the operator's active WorkPeriod
                            (end_date=None or end_date >= today). Total period hours.
                            Read-only.
      2 — Histórico       : work orders from closed WorkPeriod records, grouped
                            by period descending. Read-only.
      3 — Horas extra     : calculation of overtime for the active period.
                            Formula: worked_hours - (working_days * 8).
                            working_days = Mon–Fri count from start_date to today.
                            Read-only viewer, no editing.
      4 — Ausencias       : WorkerAbsence records for the authenticated operator.
                            Read-only viewer (type, dates, notes).

    GET /panel/operator/history/

    ---

    Vista de historial personal de cuatro pestanas para operarios con rol WORKSHOP.
    ADMIN y SUPERVISOR son redirigidos automaticamente a WorkOrderAdminHistoryView.

    Pestanas:
      1 — Periodo actual  : partes del WorkPeriod activo del operario
                            (end_date=None o end_date >= hoy). Horas totales del periodo.
                            Solo lectura.
      2 — Historico       : partes de WorkPeriod cerrados agrupados por periodo
                            en orden descendente. Solo lectura.
      3 — Horas extra     : calculo de horas extra para el periodo activo.
                            Formula: horas_trabajadas - (dias_laborables * 8).
                            dias_laborables = conteo lun–vie desde start_date a hoy.
                            Solo visor, sin edicion.
      4 — Ausencias       : registros WorkerAbsence del operario autenticado.
                            Solo lectura (tipo, fechas, notas).

    GET /panel/operator/history/
    """

    template_name ="panel/operator/history.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _count_working_days (self ,start_date ,end_date ):
        """
        Counts Monday–Friday days (working days) in the closed interval
        [start_date, end_date]. Both dates are inclusive.
        Returns 0 if start_date > end_date.
        ---
        Cuenta los dias de lunes a viernes (dias laborables) en el intervalo
        cerrado [start_date, end_date]. Ambas fechas son inclusivas.
        Devuelve 0 si start_date > end_date.
        """
        from datetime import timedelta 
        count =0 
        current =start_date 
        while current <=end_date :
            if current .weekday ()<5 :
                count +=1 
            current +=timedelta (days =1 )
        return count 

    def _enrich_work_orders_for_period (self ,qs ):
        """
        Converts a WorkOrder queryset into a list of enriched dicts for period
        tab rendering. Each dict includes pk, fecha, num_bloques, horas_totales,
        reviewed. Accumulates total hours as Decimal.
        ---
        Convierte un queryset de WorkOrder en una lista de dicts enriquecidos
        para el renderizado de la pestana de periodo. Cada dict incluye pk, fecha,
        num_bloques, horas_totales, reviewed. Acumula horas totales como Decimal.
        """
        from decimal import Decimal 
        result =[]
        total_hours =Decimal ("0")
        for wo in qs :
            entries_list =list (wo .entries .all ())
            first_entry =entries_list [0 ]if entries_list else None 
            work_date =first_entry .work_date if first_entry else None 
            num_bloques =sum (entry .lines .count ()for entry in entries_list )
            horas_totales =sum (
            (line .delta_hours 
            for entry in entries_list 
            for line in entry .lines .all ()
            if line .delta_hours is not None ),
            Decimal ("0"),
            )
            total_hours +=horas_totales 
            result .append ({
            "pk":wo .pk ,
            "fecha":work_date ,
            "num_bloques":num_bloques ,
            "horas_totales":horas_totales ,
            "reviewed":wo .reviewed ,
            "has_overlap_incident":wo .has_overlap_incident ,
            "source":wo .source ,
            })
        return result ,total_hours 

    def get (self ,request ,*args ,**kwargs ):
        """
        Builds the four-tab WORKSHOP history and renders the template.
        ADMIN and SUPERVISOR are redirected immediately to the admin history view.
        ---
        Construye el historial de cuatro pestanas WORKSHOP y renderiza el template.
        ADMIN y SUPERVISOR son redirigidos inmediatamente a la vista de historial
        de administrador.
        """
        from decimal import Decimal 
        from datetime import date as dt_date 
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence ,WorkPeriod 

        cu =request .user .company_user 
        company =cu .company 
        role =cu .role 
        today =dt_date .today ()



        if role in (CompanyUser .ROLE_ADMIN ,CompanyUser .ROLE_SUPERVISOR ):
            return redirect (reverse ("panel:work_order_admin_history"))

        active_tab =request .GET .get ("tab","current_period")






        sort_col =request .GET .get ("sort","fecha")
        sort_dir =request .GET .get ("dir","asc")
        if sort_col not in ("fecha","num_bloques","horas_totales","reviewed"):
            sort_col ="fecha"
        if sort_dir not in ("asc","desc"):
            sort_dir ="asc"








        def _base_qs ():
            return (
            WorkOrder .objects 
            .filter (
            company =company ,
            uploaded_by =cu ,
            source__in =[
            WorkOrder .Source .DIGITAL ,
            WorkOrder .Source .GENERATED ,
            ],
            )
            .prefetch_related (
            Prefetch (
            "entries",
            queryset =WorkOrderEntry .objects .prefetch_related ("lines"),
            )
            )
            .order_by ("entries__work_date")
            )




        active_period =(
        WorkPeriod .objects 
        .filter (
        company_user =cu ,
        )
        .filter (
        Q (end_date__isnull =True )|Q (end_date__gte =today )
        )
        .order_by ("-start_date")
        .first ()
        )

        current_period_list =[]
        current_period_hours =Decimal ("0")

        if active_period :


            period_qs =_base_qs ().filter (
            entries__work_date__gte =active_period .start_date ,
            )
            if active_period .end_date :
                period_qs =period_qs .filter (
                entries__work_date__lte =active_period .end_date ,
                )
            period_qs =period_qs .distinct ()
        else :




            period_qs =_base_qs ().distinct ()

        current_period_list ,current_period_hours =(
        self ._enrich_work_orders_for_period (period_qs )
        )



        def _sort_key_history (item ):
            """
            Returns the sort key for a current_period_list dict entry.
            None values sort last in both directions.
            ---
            Devuelve la clave de ordenación para un dict de current_period_list.
            Los valores None se ordenan al final en ambas direcciones.
            """
            val =item .get (sort_col )
            if val is None :


                return (1 ,None )
            return (0 ,val )

        current_period_list =sorted (
        current_period_list ,
        key =_sort_key_history ,
        reverse =(sort_dir =="desc"),
        )




        closed_periods =(
        WorkPeriod .objects 
        .filter (company_user =cu ,end_date__isnull =False ,end_date__lt =today )
        .order_by ("-start_date")
        )

        period_groups =[]
        for period in closed_periods :
            pqs =_base_qs ().filter (
            entries__work_date__gte =period .start_date ,
            entries__work_date__lte =period .end_date ,
            ).distinct ()
            wo_list ,period_hours =self ._enrich_work_orders_for_period (pqs )
            period_label =(
            period .label 
            or f"{period.start_date:%d/%m/%Y} – {period.end_date:%d/%m/%Y}"
            )
            period_groups .append ({
            "label":period_label ,
            "total_hours":period_hours ,
            "work_orders":wo_list ,
            })








        overtime_hours =Decimal ("0")
        working_days_count =0 

        if active_period :
            period_start_for_calc =active_period .start_date 
            period_end_for_calc =(
            active_period .end_date 
            if active_period .end_date and active_period .end_date <today 
            else today 
            )
        else :







            from work_order_processor .models import WorkOrderEntry as _WOE_OT 
            earliest =(
            _WOE_OT .objects 
            .filter (work_order__company =cu .company ,work_order__uploaded_by =cu )
            .order_by ("work_date")
            .values_list ("work_date",flat =True )
            .first ()
            )
            if earliest is None :


                working_days_count =0 
                overtime_hours =Decimal ("0")


                context ={
                "company":company ,
                "company_user":cu ,
                "own_presence":self ._get_own_presence (cu ),
                "active_nav":"operator_history",
                "active_tab":active_tab ,

                "sort_col":sort_col ,
                "sort_dir":sort_dir ,
                "active_period":active_period ,
                "current_period_list":current_period_list ,
                "current_period_hours":current_period_hours ,
                "period_groups":period_groups ,
                "working_days_count":0 ,
                "overtime_hours":Decimal ("0"),
                "overtime_worked_hours":Decimal ("0"),
                "absences":WorkerAbsence .objects .filter (company_user =cu ).order_by ("-start_date"),
                }
                return render (request ,self .template_name ,context )
            period_start_for_calc =earliest 
            period_end_for_calc =today 

        working_days_count =self ._count_working_days (
        period_start_for_calc ,period_end_for_calc 
        )
        expected_hours =Decimal (working_days_count )*Decimal ("8")
        overtime_hours =current_period_hours -expected_hours 




        absences =(
        WorkerAbsence .objects 
        .filter (company_user =cu )
        .order_by ("-start_date")
        )

        context ={
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"operator_history",
        "active_tab":active_tab ,

        "sort_col":sort_col ,
        "sort_dir":sort_dir ,

        "active_period":active_period ,
        "current_period_list":current_period_list ,
        "current_period_hours":current_period_hours ,

        "period_groups":period_groups ,




        "working_days_count":working_days_count ,
        "overtime_hours":overtime_hours ,
        "overtime_worked_hours":current_period_hours ,

        "absences":absences ,
        }
        return render (request ,self .template_name ,context )


class WorkerAbsenceCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a WorkerAbsence record from the absence modal in admin_history.html.
    Receives the form POST from the modal's form action pointing to
    'panel:worker_absence_create'. On success or failure, always redirects
    back to the Absences tab of WorkOrderAdminHistoryView.

    POST /panel/worker-absences/create/
        Body params:
          company_user_pk (int)  — pk of the target CompanyUser (must belong to company).
          absence_type    (str)  — one of the WorkerAbsence.ABSENCE_* constants.
          start_date      (str)  — ISO date YYYY-MM-DD.
          end_date        (str)  — ISO date YYYY-MM-DD.
          notes           (str)  — optional free text.

    ---

    Crea un registro WorkerAbsence desde el modal de ausencias de admin_history.html.
    Recibe el POST del formulario del modal cuya accion apunta a
    'panel:worker_absence_create'. En caso de exito o fallo, siempre redirige
    de vuelta a la pestana Ausencias de WorkOrderAdminHistoryView.

    POST /panel/worker-absences/create/
        Parametros del cuerpo:
          company_user_pk (int)  — pk del CompanyUser objetivo (debe pertenecer a empresa).
          absence_type    (str)  — una de las constantes WorkerAbsence.ABSENCE_*.
          start_date      (str)  — fecha ISO YYYY-MM-DD.
          end_date        (str)  — fecha ISO YYYY-MM-DD.
          notes           (str)  — texto libre opcional.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the POST data, creates the WorkerAbsence and redirects.
        Performs server-side validation: company scope, absence_type whitelist,
        date ordering (start_date <= end_date).
        ---
        Valida los datos POST, crea el WorkerAbsence y redirige.
        Realiza validacion en el servidor: scope de empresa, lista blanca de
        absence_type, ordenacion de fechas (start_date <= end_date).
        """
        from datetime import datetime 
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            cu_pk =int (request .POST .get ("company_user_pk",""))
            target_cu =CompanyUser .objects .get (pk =cu_pk ,company =company )
        except (ValueError ,TypeError ,CompanyUser .DoesNotExist ):
            django_messages .error (
            request ,
            "Operario no encontrado o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )





        absence_type =request .POST .get ("absence_type","").strip ()
        valid_types ={k for k ,_ in WorkerAbsence .ABSENCE_CHOICES }
        if absence_type not in valid_types :
            django_messages .error (
            request ,
            f"Tipo de ausencia '{absence_type}' no válido.",
            )
            return redirect (ABSENCES_TAB_URL )





        def _parse_iso (value ):
            """Parses YYYY-MM-DD string, returns date or None. / Parsea cadena YYYY-MM-DD."""
            try :
                return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
            except (ValueError ,AttributeError ):
                return None 

        start_date =_parse_iso (request .POST .get ("start_date",""))
        end_date =_parse_iso (request .POST .get ("end_date",""))

        if not start_date or not end_date :
            django_messages .error (
            request ,
            "Las fechas de inicio y fin son obligatorias y deben tener formato YYYY-MM-DD.",
            )
            return redirect (ABSENCES_TAB_URL )

        if start_date >end_date :
            django_messages .error (
            request ,
            "La fecha de inicio no puede ser posterior a la fecha de fin.",
            )
            return redirect (ABSENCES_TAB_URL )

        notes =request .POST .get ("notes","").strip ()





        WorkerAbsence .objects .create (
        company_user =target_cu ,
        absence_type =absence_type ,
        start_date =start_date ,
        end_date =end_date ,
        registered_by =cu ,
        notes =notes ,
        )

        operator_name =(
        target_cu .user .get_full_name ()or target_cu .user .username 
        )
        django_messages .success (
        request ,
        f"Ausencia de {operator_name} registrada correctamente "
        f"({start_date:%d/%m/%Y} – {end_date:%d/%m/%Y}).",
        )
        return redirect (ABSENCES_TAB_URL )


class WorkerAbsenceUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing WorkerAbsence record from the absence modal in
    admin_history.html. Receives the form POST from the edit modal.
    On success or failure, always redirects back to the Absences tab of
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/update/
        Body params:
          absence_type (str) — one of the WorkerAbsence.ABSENCE_* constants.
          start_date   (str) — ISO date YYYY-MM-DD.
          end_date     (str) — ISO date YYYY-MM-DD.
          notes        (str) — optional free text.
    ---
    Actualiza un registro WorkerAbsence existente desde el modal de edición
    de admin_history.html. Recibe el POST del formulario del modal de edición.
    En caso de éxito o fallo, siempre redirige a la pestaña Ausencias de
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/update/
        Parámetros del cuerpo:
          absence_type (str) — una de las constantes WorkerAbsence.ABSENCE_*.
          start_date   (str) — fecha ISO YYYY-MM-DD.
          end_date     (str) — fecha ISO YYYY-MM-DD.
          notes        (str) — texto libre opcional.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates the POST data, updates the WorkerAbsence and redirects.
        Performs server-side validation: company scope, absence_type whitelist,
        date ordering (start_date <= end_date).
        ---
        Valida los datos POST, actualiza el WorkerAbsence y redirige.
        Realiza validacion en servidor: scope de empresa, lista blanca de
        absence_type, ordenacion de fechas (start_date <= end_date).
        """
        from datetime import datetime 
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            absence =WorkerAbsence .objects .get (
            pk =pk ,
            company_user__company =company ,
            )
        except WorkerAbsence .DoesNotExist :
            django_messages .error (
            request ,
            "Ausencia no encontrada o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )





        absence_type =request .POST .get ("absence_type","").strip ()
        valid_types ={k for k ,_ in WorkerAbsence .ABSENCE_CHOICES }
        if absence_type not in valid_types :
            django_messages .error (
            request ,
            f"Tipo de ausencia '{absence_type}' no válido.",
            )
            return redirect (ABSENCES_TAB_URL )





        def _parse_iso (value ):
            """Parses YYYY-MM-DD string, returns date or None. / Parsea cadena YYYY-MM-DD."""
            try :
                return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
            except (ValueError ,AttributeError ):
                return None 

        start_date =_parse_iso (request .POST .get ("start_date",""))
        end_date =_parse_iso (request .POST .get ("end_date",""))

        if not start_date or not end_date :
            django_messages .error (
            request ,
            "Las fechas de inicio y fin son obligatorias y deben tener formato YYYY-MM-DD.",
            )
            return redirect (ABSENCES_TAB_URL )

        if start_date >end_date :
            django_messages .error (
            request ,
            "La fecha de inicio no puede ser posterior a la fecha de fin.",
            )
            return redirect (ABSENCES_TAB_URL )

        notes =request .POST .get ("notes","").strip ()





        absence .absence_type =absence_type 
        absence .start_date =start_date 
        absence .end_date =end_date 
        absence .notes =notes 
        absence .save (update_fields =["absence_type","start_date","end_date","notes"])

        operator_name =(
        absence .company_user .user .get_full_name ()
        or absence .company_user .user .username 
        )
        django_messages .success (
        request ,
        f"Ausencia de {operator_name} actualizada correctamente "
        f"({start_date:%d/%m/%Y} – {end_date:%d/%m/%Y}).",
        )
        return redirect (ABSENCES_TAB_URL )


class WorkerAbsenceDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a WorkerAbsence record scoped to the authenticated user's company.
    On success or failure, always redirects back to the Absences tab of
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/delete/
    ---
    Elimina un registro WorkerAbsence acotado a la empresa del usuario
    autenticado. En caso de éxito o fallo, siempre redirige a la pestaña
    Ausencias de WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/delete/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Resolves the WorkerAbsence by pk scoped to the company, deletes it
        and redirects back to the Absences tab.
        ---
        Resuelve el WorkerAbsence por pk acotado a la empresa, lo elimina
        y redirige de vuelta a la pestaña Ausencias.
        """
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            absence =WorkerAbsence .objects .select_related (
            "company_user__user"
            ).get (
            pk =pk ,
            company_user__company =company ,
            )
        except WorkerAbsence .DoesNotExist :
            django_messages .error (
            request ,
            "Ausencia no encontrada o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )

        operator_name =(
        absence .company_user .user .get_full_name ()
        or absence .company_user .user .username 
        )
        absence .delete ()

        django_messages .success (
        request ,
        f"Ausencia de {operator_name} eliminada correctamente.",
        )
        return redirect (ABSENCES_TAB_URL )








def _serialize_pending_lines (parsed_lines ,parsed_repuestos ,parsed_date ):
    """
    Serialises the incoming form lines and spare parts into a JSON-safe
    list of dicts suitable for storage in the Django session.

    Each line dict contains:
        machine_raw, machine_asset_pk, fault_description, repair_notes,
        hc ("HH:MM"|null), hf ("HH:MM"|null), delta_hours (str|null),
        odometer_reading (str|null), engine_hours_reading (str|null),
        crane_hours_reading (str|null),
        repuestos: [{material, reference, quantity, source,
                     supplier, unit_price}]

    The date is stored as an ISO string so the merge view can
    reconstruct it.
    ---
    Serializa las lineas del formulario entrante y los repuestos en una
    lista de dicts JSON-safe apta para su almacenamiento en la sesion.

    La fecha se almacena como cadena ISO para que la vista de merge
    pueda reconstruirla.
    """
    serialized =[]
    for ld in parsed_lines :
        hc_val =ld ["hc"].strftime ("%H:%M")if ld ["hc"]else None 
        hf_val =ld ["hf"].strftime ("%H:%M")if ld ["hf"]else None 



        line_repuestos =[]
        for spd in parsed_repuestos :
            unit_price_val =str (spd ["unit_price"])if spd .get ("unit_price")is not None else None 
            line_repuestos .append ({
            "material":spd ["material"],
            "reference":spd ["referencia"],
            "quantity":str (spd ["quantity"])if spd ["quantity"]is not None else None ,
            "source":spd ["source"],
            "supplier":spd ["supplier"],
            "unit_price":unit_price_val ,
            })

        serialized .append ({
        "machine_raw":ld ["machine_raw"],
        "machine_asset_pk":ld ["machine_asset"].pk if ld ["machine_asset"]else None ,
        "fault_description":ld ["fault_description"],
        "repair_notes":ld ["repair_notes"],
        "hc":hc_val ,
        "hf":hf_val ,
        "delta_hours":str (ld ["delta_hours"])if ld ["delta_hours"]is not None else None ,
        "odometer_reading":str (ld ["odometer_reading"])if ld .get ("odometer_reading")is not None else None ,
        "engine_hours_reading":str (ld ["engine_hours_reading"])if ld .get ("engine_hours_reading")is not None else None ,
        "crane_hours_reading":str (ld ["crane_hours_reading"])if ld .get ("crane_hours_reading")is not None else None ,
        "repuestos":line_repuestos ,
        })

    return {
    "lines":serialized ,
    "work_date":parsed_date .isoformat ()if parsed_date else None ,
    }


def _get_min_allowed_date (cu ):









    from datetime import timedelta 
    from work_order_processor .models import WorkOrderEntry 
    last_reviewed =(
    WorkOrderEntry .objects 
    .filter (
    work_order__company =cu .company ,
    work_order__uploaded_by =cu ,
    work_order__reviewed =True ,
    )
    .order_by ("-work_date")
    .values_list ("work_date",flat =True )
    .first ()
    )
    if last_reviewed is None :
        return None 
    return last_reviewed +timedelta (days =1 )


def _detect_workday_gaps (entry_lines ,schedule ,work_date ):
    """
    Detects workday gaps and deviations in a set of parsed entry lines
    against a WorkdaySchedule reference timetable.

    Supports both intensive (single morning tract) and split-shift schedules
    (morning + afternoon tracts). The midday window in split-shift mode is
    classified as LUNCH_BREAK — not a blocking GAP — to allow the operator
    to justify it with a simplified lunch confirmation dialog.

    Checks performed (all require schedule to be non-None):
      LATE_START   — first block starts after start_time_morning + tolerance.
      EARLY_END    — last block ends before effective end_time - tolerance.
                     Intensive: end_time_morning. Split: end_time_afternoon.
      LUNCH_BREAK  — split-shift only: midday window between end_time_morning
                     and start_time_afternoon not covered by work blocks.
      GAP          — uncovered interval >= tolerance_minutes between consecutive
                     blocks, excluding the lunch window in split-shift mode.

    Lines with null hc or hf are silently ignored in all checks.

    Parameters:
        entry_lines (list[dict]) — parsed line dicts with keys 'hc' and 'hf'
                                   (time objects or HH:MM strings).
        schedule    (WorkdaySchedule | None) — reference timetable. If None,
                                   the function returns an empty list
                                   (Gate 4 disabled).
        work_date   (date)        — date of the work order (reserved for
                                   future use, e.g. holiday calendars).

    Returns:
        list[dict] — one dict per detected gap/deviation, each with keys:
            gap_type         (str)  — "GAP" | "LATE_START" | "EARLY_END"
                                      | "LUNCH_BREAK"
            gap_start        (time) — start of the uncovered interval.
            gap_end          (time) — end of the uncovered interval.
            duration_minutes (int)  — duration in minutes.

    ---

    Detecta lagunas y desviaciones de jornada en un conjunto de líneas de
    entrada parseadas contra un horario de referencia WorkdaySchedule.

    Soporta jornada intensiva (solo tramo de mañana) y turno partido
    (tramo de mañana + tarde). La ventana de mediodía en turno partido se
    clasifica como LUNCH_BREAK — no como GAP bloqueante — para permitir al
    operario justificarla con un diálogo simplificado de confirmación de comida.

    Comprobaciones realizadas (todas requieren schedule no nulo):
      LATE_START   — el primer bloque comienza después de start_time_morning
                     + tolerancia.
      EARLY_END    — el último bloque termina antes de la hora de salida
                     efectiva - tolerancia. Intensiva: end_time_morning.
                     Partida: end_time_afternoon.
      LUNCH_BREAK  — solo turno partido: ventana de mediodía entre
                     end_time_morning y start_time_afternoon no cubierta
                     por bloques de trabajo.
      GAP          — intervalo sin cubrir >= tolerance_minutes entre bloques
                     consecutivos, excluyendo la ventana de comida en turno
                     partido.

    Las líneas con hc o hf nulos se ignoran silenciosamente en todas las
    comprobaciones.

    Parámetros:
        entry_lines (list[dict]) — dicts de líneas parseadas con claves 'hc' y
                                   'hf' (objetos time o cadenas HH:MM).
        schedule    (WorkdaySchedule | None) — horario de referencia. Si es
                                   None, la función devuelve lista vacía
                                   (Gate 4 desactivado).
        work_date   (date)        — fecha del parte (reservado para uso futuro,
                                   p. ej. calendarios de festivos).

    Retorna:
        list[dict] — un dict por laguna/desviación detectada, con claves:
            gap_type         (str)  — "GAP" | "LATE_START" | "EARLY_END"
                                      | "LUNCH_BREAK"
            gap_start        (time) — inicio del intervalo sin cubrir.
            gap_end          (time) — fin del intervalo sin cubrir.
            duration_minutes (int)  — duración en minutos.
    """
    from datetime import datetime as _dt_gaps ,time as _time_gaps 



    if schedule is None :
        return []

    def _to_t (val ):
        """
        Coerces a time object or HH:MM string to a time instance.
        Returns None on failure.
        ---
        Convierte un objeto time o cadena HH:MM a una instancia time.
        Devuelve None en caso de fallo.
        """
        if val is None :
            return None 
        if hasattr (val ,"hour"):
            return val 
        try :
            return _dt_gaps .strptime (str (val ).strip (),"%H:%M").time ()
        except ValueError :
            return None 

    def _mins (t ):
        """
        Converts a time object to total minutes since midnight.
        ---
        Convierte un objeto time a minutos totales desde medianoche.
        """
        return t .hour *60 +t .minute 












    start_morning =getattr (schedule ,"start_time_morning",None )or getattr (schedule ,"start_time",None )
    end_morning =getattr (schedule ,"end_time_morning",None )or getattr (schedule ,"end_time",None )
    is_intensive =getattr (schedule ,"is_intensive",True )
    start_afternoon =None if is_intensive else getattr (schedule ,"start_time_afternoon",None )
    end_afternoon =None if is_intensive else getattr (schedule ,"end_time_afternoon",None )

    if start_morning is None or end_morning is None :


        return []



    effective_end =end_afternoon if (not is_intensive and end_afternoon )else end_morning 





    valid_blocks =[]
    for ld in entry_lines :
        hc =_to_t (ld .get ("hc"))
        hf =_to_t (ld .get ("hf"))
        if hc is not None and hf is not None :
            valid_blocks .append ((hc ,hf ))

    if not valid_blocks :
        return []

    valid_blocks .sort (key =lambda b :_mins (b [0 ]))

    gaps =[]
    tol =schedule .tolerance_minutes 





    first_hc =valid_blocks [0 ][0 ]
    if _mins (first_hc )>_mins (start_morning )+tol :
        gaps .append ({
        "gap_type":"LATE_START",
        "gap_start":start_morning ,
        "gap_end":first_hc ,
        "duration_minutes":_mins (first_hc )-_mins (start_morning ),
        })





    last_hf =valid_blocks [-1 ][1 ]
    if _mins (last_hf )<_mins (effective_end )-tol :
        gaps .append ({
        "gap_type":"EARLY_END",
        "gap_start":last_hf ,
        "gap_end":effective_end ,
        "duration_minutes":_mins (effective_end )-_mins (last_hf ),
        })












    if not is_intensive and start_afternoon and end_morning :
        lunch_start_m =_mins (end_morning )
        lunch_end_m =_mins (start_afternoon )


        lunch_covered =any (
        _mins (hf )>=lunch_start_m and _mins (hc )<=lunch_end_m 
        for hc ,hf in valid_blocks 
        )
        if not lunch_covered :
            gaps .append ({
            "gap_type":"LUNCH_BREAK",
            "gap_start":end_morning ,
            "gap_end":start_afternoon ,
            "duration_minutes":lunch_end_m -lunch_start_m ,
            })







    for i in range (len (valid_blocks )-1 ):
        hf_curr =valid_blocks [i ][1 ]
        hc_next =valid_blocks [i +1 ][0 ]
        gap_mins =_mins (hc_next )-_mins (hf_curr )

        if gap_mins <tol :
            continue 





        if not is_intensive and start_afternoon and end_morning :
            lunch_start_m =_mins (end_morning )
            lunch_end_m =_mins (start_afternoon )
            gap_start_m =_mins (hf_curr )
            gap_end_m =_mins (hc_next )
            if gap_start_m >=lunch_start_m and gap_end_m <=lunch_end_m :
                continue 

        gaps .append ({
        "gap_type":"GAP",
        "gap_start":hf_curr ,
        "gap_end":hc_next ,
        "duration_minutes":gap_mins ,
        })

    return gaps 


def _detect_overlaps (existing_lines ,new_lines ):
    """
    Detects time overlaps between existing WorkOrderEntryLine records
    and a list of new line dicts from the session pending_merge_lines.

    Overlap condition (open intervals): hc_e < hf_n AND hc_n < hf_e.
    Lines with null hc or hf are silently ignored.

    Returns a list of tuples:
        (idx_e, idx_n, hc_e_str, hf_e_str, hc_n_str, hf_n_str)
    where idx_e is 1-based into existing_lines and idx_n is 1-based
    into new_lines.
    ---
    Detecta solapamientos temporales entre WorkOrderEntryLine existentes
    y la lista de dicts de lineas nuevas del payload pending_merge_lines.

    Condicion (intervalos abiertos): hc_e < hf_n AND hc_n < hf_e.
    Las lineas con hc o hf nulos se ignoran.

    Devuelve lista de tuplas:
        (idx_e, idx_n, hc_e_str, hf_e_str, hc_n_str, hf_n_str)
    """
    from datetime import datetime as _dt_ov 

    def _t (val ):
        """
        Parses HH:MM string to time, returns None on failure.
        ---
        Parsea cadena HH:MM a time, devuelve None en fallo.
        """
        if val is None :
            return None 
        try :
            return _dt_ov .strptime (str (val ).strip (),"%H:%M").time ()
        except ValueError :
            return None 

    conflicts =[]
    for idx_e ,existing_line in enumerate (existing_lines ,start =1 ):
        hc_e =existing_line .hc 
        hf_e =existing_line .hf 
        if hc_e is None or hf_e is None :
            continue 
        for idx_n ,new_line in enumerate (new_lines ,start =1 ):
            hc_n =_t (new_line .get ("hc"))
            hf_n =_t (new_line .get ("hf"))
            if hc_n is None or hf_n is None :
                continue 


            if hc_e <hf_n and hc_n <hf_e :
                conflicts .append ((
                idx_e ,
                idx_n ,
                hc_e .strftime ("%H:%M"),
                hf_e .strftime ("%H:%M"),
                hc_n .strftime ("%H:%M"),
                hf_n .strftime ("%H:%M"),
                ))
    return conflicts 


class WorkOrderEntryMergeView (WorkshopRequiredMixin ,View ):
    """
    Merge view for resolving conflicts when an operator tries to submit
    a new work order on a date that already has an unreviewed digital or
    generated entry. Presents existing and incoming lines side by side
    so the operator can choose one of three actions:

      discard_new      — keep existing entry, discard incoming lines.
      discard_existing — delete existing WorkOrder (CASCADE) and create
                         a new one from the pending session lines.
      merge            — append incoming lines to the existing entry.
                         Only available when no time overlaps exist.

    The operator may edit hc/hf before choosing an action.
    Client-side JS recalculates overlaps in real time; the server
    revalidates on every merge POST.

    GET  /panel/operator/merge/<int:entry_pk>/
    POST /panel/operator/merge/<int:entry_pk>/
         merge_action: discard_new | discard_existing | merge

    Accessible to WORKSHOP and ADMIN roles (WorkshopRequiredMixin).
    ---
    Vista de merge para resolver conflictos cuando un operario envia
    un parte en una fecha con entrada digital sin revisar preexistente.

    GET  /panel/operator/merge/<int:entry_pk>/
    POST /panel/operator/merge/<int:entry_pk>/
         merge_action: discard_new | discard_existing | merge

    Accesible para roles WORKSHOP y ADMIN (WorkshopRequiredMixin).
    """

    template_name ="panel/operator/merge_entry.html"





    def _get_pending (self ,request ):
        """
        Returns the pending_merge_lines dict from the session, or None.
        ---
        Devuelve el dict pending_merge_lines de la sesion, o None.
        """
        return request .session .get ("pending_merge_lines")

    def _clear_pending (self ,request ):
        """
        Removes pending_merge_lines from the session.
        ---
        Elimina pending_merge_lines de la sesion.
        """
        request .session .pop ("pending_merge_lines",None )
        request .session .modified =True 

    def _resolve_entry (self ,entry_pk ,company ,cu ):
        """
        Retrieves the WorkOrderEntry by pk, scoped to operator company
        and user. Returns None if not found or inaccessible.
        ---
        Recupera el WorkOrderEntry por pk, acotado a empresa y usuario
        del operario. Devuelve None si no existe o no es accesible.
        """
        try :
            return WorkOrderEntry .objects .select_related ("work_order").get (
            pk =entry_pk ,
            work_order__company =company ,
            work_order__uploaded_by =cu ,
            work_order__source__in =[
            WorkOrder .Source .DIGITAL ,
            WorkOrder .Source .GENERATED ,
            ],
            work_order__reviewed =False ,
            )
        except WorkOrderEntry .DoesNotExist :
            return None 

    def _parse_time_str (self ,val ):
        """
        Parses HH:MM string to datetime.time, returns None on failure.
        ---
        Parsea cadena HH:MM a datetime.time, devuelve None en fallo.
        """
        from datetime import time as _time 
        if not val :
            return None 
        try :
            parts =str (val ).strip ().split (":")
            return _time (int (parts [0 ]),int (parts [1 ]))
        except (ValueError ,IndexError ):
            return None 

    def _to_decimal (self ,val ):
        """
        Converts str/int/float to Decimal, returns None on failure.
        ---
        Convierte str/int/float a Decimal, devuelve None en fallo.
        """
        from decimal import Decimal ,InvalidOperation 
        if val is None :
            return None 
        try :
            return Decimal (str (val ))
        except InvalidOperation :
            return None 

    def _parse_edited_hc_hf (self ,POST ,prefix ,count ):
        """
        Parses operator-edited hc/hf from POST for a set of lines
        identified by prefix and count (1-based).
        Returns a list of dicts: [{hc: HH:MM|None, hf: HH:MM|None}].
        ---
        Parsea hc/hf editados del POST para un conjunto de lineas
        identificadas por prefix y count (base 1).
        """
        result =[]
        for i in range (1 ,count +1 ):
            hc_raw =POST .get (f"{prefix}{i}_hc","").strip ()or None 
            hf_raw =POST .get (f"{prefix}{i}_hf","").strip ()or None 
            result .append ({"hc":hc_raw ,"hf":hf_raw })
        return result 

    def _build_context (self ,company ,cu ,existing_entry ,existing_lines ,
    new_lines ,work_date_iso ,conflicts ,merge_error =None ):
        """
        Builds the template context dict for the merge view.
        ---
        Construye el dict de contexto del template para la vista de merge.
        """
        return {
        "company":company ,
        "company_user":cu ,
        "active_nav":"operator_dashboard",
        "existing_entry":existing_entry ,
        "existing_lines":existing_lines ,
        "new_lines":new_lines ,
        "work_date":work_date_iso ,
        "conflicts":conflicts ,
        "has_conflicts":bool (conflicts ),
        "merge_error":merge_error ,
        }

    def _create_lines_from_session (self ,new_lines ,edited_new ,
    target_entry ,start_line_number ,company ):
        """
        Creates WorkOrderEntryLine and SparePartLine records from the
        session pending data inside an already-open atomic block.
        edited_new may provide operator-corrected hc/hf values.
        ---
        Crea WorkOrderEntryLine y SparePartLine desde los datos pendientes
        de la sesion dentro de un bloque atomico ya abierto.
        edited_new puede aportar hc/hf corregidos por el operario.
        """
        from fleet .models import MachineAsset 
        from work_order_processor .models import SparePartLine 
        from work_order_processor .services import (
        _normalise_machine_code ,
        _compute_delta_hours ,
        )

        for idx ,line_data in enumerate (new_lines ):
            edits =edited_new [idx ]if idx <len (edited_new )else {}
            hc_str =edits .get ("hc")or line_data .get ("hc")
            hf_str =edits .get ("hf")or line_data .get ("hf")
            hc_val =self ._parse_time_str (hc_str )
            hf_val =self ._parse_time_str (hf_str )
            delta =_compute_delta_hours (hc_val ,hf_val ,deduct_lunch =False )



            asset =None 
            _asset_pk =line_data .get ("machine_asset_pk")
            if _asset_pk is not None :
                try :
                    asset =MachineAsset .objects .get (
                    pk =_asset_pk ,company =company 
                    )
                except MachineAsset .DoesNotExist :
                    pass 

            machine_raw =line_data .get ("machine_raw","")
            machine_norm =_normalise_machine_code (machine_raw )

            new_line =WorkOrderEntryLine .objects .create (
            entry =target_entry ,
            line_number =start_line_number +idx ,
            machine_asset =asset ,
            machine_raw =machine_raw ,
            machine_norm =machine_norm or "",
            fault_description =line_data .get ("fault_description",""),
            repair_notes =line_data .get ("repair_notes",""),
            hc =hc_val ,
            hf =hf_val ,
            or_val ="",
            delta_hours =delta ,
            flags =[],
            odometer_reading =self ._to_decimal (line_data .get ("odometer_reading")),
            engine_hours_reading =self ._to_decimal (line_data .get ("engine_hours_reading")),
            crane_hours_reading =self ._to_decimal (line_data .get ("crane_hours_reading")),
            )



            for rep_idx ,rep in enumerate (line_data .get ("repuestos",[]),start =1 ):
                SparePartLine .objects .create (
                entry_line =new_line ,
                line_number =rep_idx ,
                reference =rep .get ("reference",""),
                vehicle =None ,
                material =rep .get ("material",""),
                quantity =self ._to_decimal (rep .get ("quantity")),
                source =rep .get ("source","WAREHOUSE"),
                supplier =rep .get ("supplier",""),
                flags =[],
                )





    def get (self ,request ,entry_pk ,*args ,**kwargs ):
        """
        Renders the merge resolution page. Detects initial overlaps.
        Redirects to operator history with an error if session data is
        missing or the entry is inaccessible.
        ---
        Renderiza la pagina de resolucion de merge. Detecta solapamientos
        iniciales. Redirige al historial si faltan datos o el entry no
        es accesible.
        """
        from django .urls import reverse 

        cu =request .user .company_user 
        company =cu .company 

        pending =self ._get_pending (request )
        if not pending :
            django_messages .error (
            request ,
            "No hay datos de parte pendiente en sesion. "
            "El formulario ha expirado o fue enviado ya.",
            )
            return redirect (reverse ("panel:operator_history"))

        existing_entry =self ._resolve_entry (entry_pk ,company ,cu )
        if existing_entry is None :
            self ._clear_pending (request )
            django_messages .error (
            request ,
            "El parte existente no se ha encontrado o no es accesible.",
            )
            return redirect (reverse ("panel:operator_history"))

        existing_lines =list (existing_entry .lines .order_by ("line_number"))
        new_lines =pending .get ("lines",[])
        work_date_iso =pending .get ("work_date")

        conflicts =_detect_overlaps (existing_lines ,new_lines )

        context =self ._build_context (
        company ,cu ,existing_entry ,existing_lines ,
        new_lines ,work_date_iso ,conflicts ,
        )
        return render (request ,self .template_name ,context )

    def post (self ,request ,entry_pk ,*args ,**kwargs ):
        """
        Processes the merge action chosen by the operator.

        discard_new      — Clears the session and redirects to history.
        discard_existing — Deletes the existing WorkOrder (CASCADE) and
                           creates a new one from the pending session data.
        merge            — Re-validates overlaps with edited hc/hf values.
                           Appends new lines to existing entry if no conflicts.
        ---
        Procesa la accion de merge elegida por el operario.

        discard_new      — Limpia sesion y redirige al historial.
        discard_existing — Elimina WorkOrder existente y crea nuevo desde sesion.
        merge            — Revalida solapamientos con hc/hf editados. Anade
                           lineas al entry existente si no hay conflictos.
        """
        from datetime import datetime as _dtp 
        from django .db import transaction 
        from django .urls import reverse 
        from work_order_processor .services import generate_work_order_excel 

        cu =request .user .company_user 
        company =cu .company 

        pending =self ._get_pending (request )
        if not pending :
            django_messages .error (
            request ,
            "No hay datos de parte pendiente en sesion. "
            "El formulario ha expirado o fue enviado ya.",
            )
            return redirect (reverse ("panel:operator_history"))

        existing_entry =self ._resolve_entry (entry_pk ,company ,cu )
        if existing_entry is None :
            self ._clear_pending (request )
            django_messages .error (
            request ,
            "El parte existente no se ha encontrado o no es accesible.",
            )
            return redirect (reverse ("panel:operator_history"))

        merge_action =request .POST .get ("merge_action","").strip ()
        existing_lines =list (existing_entry .lines .order_by ("line_number"))
        new_lines =pending .get ("lines",[])
        work_date_iso =pending .get ("work_date")




        if merge_action =="discard_new":
            self ._clear_pending (request )
            django_messages .success (
            request ,
            "Parte nuevo descartado. Se conserva el parte existente.",
            )
            return redirect (reverse ("panel:operator_history"))















        if merge_action in ("discard_existing","merge"):
            _work_date_g4mv =None 
            if work_date_iso :
                try :
                    _work_date_g4mv =_dtp .strptime (work_date_iso ,"%Y-%m-%d").date ()
                except ValueError :
                    pass 

            if _work_date_g4mv is not None :
                from ivr_config .models import WorkdaySchedule as _WDS_MV 
                from django .urls import reverse as _rev_g4mv 
                _schedule_g4mv =(
                cu .workday_schedule 
                if cu .workday_schedule_id 
                else (




                next (
                (
                _sec .workday_schedule 
                for _sec in (
                _contact_g4mv .sections 
                .filter (is_active =True ,workday_schedule__isnull =False )
                .select_related ("workday_schedule")
                .order_by ("name")
                )
                ),
                None ,
                )
                if (_contact_g4mv :=(
                Contact .objects .filter (company_user =cu )
                .prefetch_related ("sections__workday_schedule")
                .first ()
                ))is not None 
                else None 
                )or _WDS_MV .objects .filter (
                company =company ,is_default =True 
                ).first ()
                )
                _gaps_g4mv =_detect_workday_gaps (
                new_lines ,_schedule_g4mv ,_work_date_g4mv 
                )
                if _gaps_g4mv :
                    from django .db import transaction as _tx_g4mv 
                    _worker_name_g4mv =(
                    cu .user .get_full_name ()or cu .user .username 
                    ).upper ()
                    _date_tag_g4mv =_work_date_g4mv .strftime ("%d-%m-%Y")
                    _synth_g4mv =(
                    f"{_worker_name_g4mv}_{_date_tag_g4mv}_MERGE_DRAFT.pdf"
                    )
                    with _tx_g4mv .atomic ():
                        _wo_draft_mv =WorkOrder (
                        company =company ,
                        uploaded_by =cu ,
                        source =WorkOrder .Source .DIGITAL ,
                        status =WorkOrder .Status .PENDING_GAPS ,
                        total_pages =1 ,
                        processed_pages =1 ,
                        reviewed =False ,
                        )
                        _wo_draft_mv .source_pdf .name =_synth_g4mv 
                        _wo_draft_mv .save ()
                        _entry_draft_mv =WorkOrderEntry .objects .create (
                        work_order =_wo_draft_mv ,
                        page_number =1 ,
                        worker_name =_worker_name_g4mv ,
                        work_date =_work_date_g4mv ,
                        uncertain_date =False ,
                        extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                        raw_gemini_response =None ,
                        )
                        _line_num_g4mv =1 
                        for _nd in new_lines :
                            _ma_pk =_nd .get ("machine_asset_pk")
                            _ma_mv =None 
                            if _ma_pk :
                                try :
                                    from fleet .models import MachineAsset as _MA_MV 
                                    _ma_mv =_MA_MV .objects .filter (
                                    pk =_ma_pk ,company =company 
                                    ).first ()
                                except Exception :
                                    pass 
                            from datetime import datetime as _dtm_mv 
                            def _tp (v ):
                                if not v :
                                    return None 
                                try :
                                    return _dtm_mv .strptime (str (v ).strip (),"%H:%M").time ()
                                except ValueError :
                                    return None 
                            from decimal import Decimal as _Dec_mv 
                            WorkOrderEntryLine .objects .create (
                            entry =_entry_draft_mv ,
                            line_number =_line_num_g4mv ,
                            machine_asset =_ma_mv ,
                            machine_raw =_nd .get ("machine_raw",""),
                            machine_norm ="",
                            fault_description =_nd .get ("fault_description",""),
                            repair_notes =_nd .get ("repair_notes",""),
                            hc =_tp (_nd .get ("hc")),
                            hf =_tp (_nd .get ("hf")),
                            or_val ="",
                            delta_hours =(
                            _Dec_mv (_nd ["delta_hours"])
                            if _nd .get ("delta_hours")else None 
                            ),
                            flags =[],
                            odometer_reading =(
                            _Dec_mv (_nd ["odometer_reading"])
                            if _nd .get ("odometer_reading")else None 
                            ),
                            engine_hours_reading =(
                            _Dec_mv (_nd ["engine_hours_reading"])
                            if _nd .get ("engine_hours_reading")else None 
                            ),
                            crane_hours_reading =(
                            _Dec_mv (_nd ["crane_hours_reading"])
                            if _nd .get ("crane_hours_reading")else None 
                            ),
                            )
                            _line_num_g4mv +=1 
                        from work_order_processor .models import WorkdayGap as _WDG_MV 
                        for _gap in _gaps_g4mv :
                            _WDG_MV .objects .create (
                            work_order =_wo_draft_mv ,
                            gap_type =_gap ["gap_type"],
                            gap_start =_gap ["gap_start"],
                            gap_end =_gap ["gap_end"],
                            duration_minutes =_gap ["duration_minutes"],
                            )
                    request .session ["pending_gaps_wo_pk"]=_wo_draft_mv .pk 
                    request .session .modified =True 
                    logger .info (
                    "# [MergeView/Gate4] PENDING_GAPS borrador pk=%d creado. "
                    "%d gap(s) detectado(s). Redirigiendo a resolución.",
                    _wo_draft_mv .pk ,len (_gaps_g4mv ),
                    )
                    return redirect (
                    _rev_g4mv (
                    "panel:operator_gap_resolution",
                    kwargs ={"wo_draft_pk":_wo_draft_mv .pk },
                    )
                    )




        if merge_action =="discard_existing":
            work_date =None 
            if work_date_iso :
                try :
                    work_date =_dtp .strptime (work_date_iso ,"%Y-%m-%d").date ()
                except ValueError :
                    pass 

            try :
                with transaction .atomic ():


                    existing_entry .work_order .delete ()

                    worker_name =(
                    cu .user .get_full_name ()or cu .user .username 
                    ).upper ()
                    date_tag =(
                    work_date .strftime ("%d-%m-%Y")if work_date else "SIN-FECHA"
                    )
                    synthetic_name =f"{worker_name}_{date_tag}.pdf"

                    new_wo =WorkOrder (
                    company =company ,
                    uploaded_by =cu ,
                    source =WorkOrder .Source .DIGITAL ,
                    status =WorkOrder .Status .DONE ,
                    total_pages =1 ,
                    processed_pages =1 ,
                    reviewed =False ,
                    )
                    new_wo .source_pdf .name =synthetic_name 
                    new_wo .save ()

                    new_entry =WorkOrderEntry .objects .create (
                    work_order =new_wo ,
                    page_number =1 ,
                    worker_name =worker_name ,
                    work_date =work_date ,
                    uncertain_date =False ,
                    extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                    raw_gemini_response =None ,
                    )



                    self ._create_lines_from_session (
                    new_lines ,[],new_entry ,1 ,company 
                    )

            except Exception as exc :
                logger .error (
                "# [MergeView] Error en discard_existing: %s",exc ,exc_info =True 
                )
                django_messages .error (
                request ,
                f"Error al procesar el merge: {exc}. Intentalo de nuevo.",
                )
                return redirect (reverse ("panel:operator_history"))

            self ._clear_pending (request )





            for _line in new_entry .lines .all ():
                _cached =find_cached_classification (
                fault_description =_line .fault_description ,
                repair_notes =_line .repair_notes ,
                company =company ,
                )
                if _cached :
                    WorkOrderEntryLine .objects .filter (pk =_line .pk ).update (
                    fault_category =_cached [0 ],
                    fault_subcategory =_cached [1 ],
                    )
                    logger .info (
                    "# [MergeView/discard_existing] Clasificación copiada "
                    "desde caché para WorkOrderEntryLine pk=%d: "
                    "category=%s subcategory=%s.",
                    _line .pk ,_cached [0 ],_cached [1 ],
                    )
                else :
                    classify_fault_line .apply_async (
                    args =[_line .pk ],
                    queue ="work_orders",
                    )
                    logger .info (
                    "# [MergeView/discard_existing] classify_fault_line "
                    "encolada para WorkOrderEntryLine pk=%d.",
                    _line .pk ,
                    )

            try :
                generate_work_order_excel (new_wo .pk )
            except Exception as exc :
                logger .warning (
                "# [MergeView] Excel no generado para WorkOrder #%d: %s",
                new_wo .pk ,exc ,
                )
            django_messages .success (
            request ,
            "Parte existente sustituido correctamente. "
            "El nuevo parte ha sido registrado.",
            )
            return redirect (reverse ("panel:operator_history"))




        if merge_action =="merge":
            edited_existing =self ._parse_edited_hc_hf (
            request .POST ,"existing_line_",len (existing_lines )
            )
            edited_new =self ._parse_edited_hc_hf (
            request .POST ,"new_line_",len (new_lines )
            )



            class _LineCopy :
                pass 

            existing_copies =[]
            for line ,edits in zip (existing_lines ,edited_existing ):
                copy =_LineCopy ()
                copy .hc =self ._parse_time_str (edits ["hc"])if edits ["hc"]else line .hc 
                copy .hf =self ._parse_time_str (edits ["hf"])if edits ["hf"]else line .hf 
                existing_copies .append (copy )

            new_copies =[]
            for nd ,edits in zip (new_lines ,edited_new ):
                new_copies .append ({
                "hc":edits ["hc"]if edits ["hc"]else nd .get ("hc"),
                "hf":edits ["hf"]if edits ["hf"]else nd .get ("hf"),
                })

            conflicts =_detect_overlaps (existing_copies ,new_copies )

            if conflicts :
                context =self ._build_context (
                company ,cu ,existing_entry ,existing_lines ,
                new_lines ,work_date_iso ,conflicts ,
                merge_error =(
                "No es posible fusionar: existen solapamientos horarios. "
                "Edita los horarios para resolver los conflictos."
                ),
                )
                return render (request ,self .template_name ,context )



            try :
                with transaction .atomic ():
                    start_number =existing_entry .lines .count ()+1 
                    self ._create_lines_from_session (
                    new_lines ,edited_new ,existing_entry ,start_number ,company 
                    )
            except Exception as exc :
                logger .error (
                "# [MergeView] Error en merge: %s",exc ,exc_info =True 
                )
                django_messages .error (
                request ,
                f"Error al fusionar el parte: {exc}. Intentalo de nuevo.",
                )
                return redirect (reverse ("panel:operator_history"))

            self ._clear_pending (request )







            _total_lines =existing_entry .lines .count ()
            _new_start_nr =_total_lines -len (new_lines )+1 
            for _line in existing_entry .lines .filter (line_number__gte =_new_start_nr ):
                _cached =find_cached_classification (
                fault_description =_line .fault_description ,
                repair_notes =_line .repair_notes ,
                company =company ,
                )
                if _cached :
                    WorkOrderEntryLine .objects .filter (pk =_line .pk ).update (
                    fault_category =_cached [0 ],
                    fault_subcategory =_cached [1 ],
                    )
                    logger .info (
                    "# [MergeView/merge] Clasificación copiada desde caché "
                    "para WorkOrderEntryLine pk=%d: "
                    "category=%s subcategory=%s.",
                    _line .pk ,_cached [0 ],_cached [1 ],
                    )
                else :
                    classify_fault_line .apply_async (
                    args =[_line .pk ],
                    queue ="work_orders",
                    )
                    logger .info (
                    "# [MergeView/merge] classify_fault_line encolada "
                    "para WorkOrderEntryLine pk=%d.",
                    _line .pk ,
                    )

            try :
                generate_work_order_excel (existing_entry .work_order .pk )
            except Exception as exc :
                logger .warning (
                "# [MergeView] Excel no regenerado para WorkOrder #%d: %s",
                existing_entry .work_order .pk ,exc ,
                )
            django_messages .success (
            request ,
            f"Parte fusionado correctamente. Tareas anadidas al parte del "
            f"{work_date_iso or 'fecha desconocida'}.",
            )
            return redirect (reverse ("panel:operator_history"))



        django_messages .warning (
        request ,
        "Accion de merge no reconocida. No se ha realizado ningun cambio.",
        )
        return redirect (reverse ("panel:operator_history"))


class WorkdayGapResolutionView (WorkshopRequiredMixin ,View ):
    """
    Allows a WORKSHOP operator to justify each workday gap detected by Gate 4
    before the draft work order is promoted from PENDING_GAPS to DONE.

    GET  /panel/operator/gaps/<int:wo_draft_pk>/
         Retrieves the pending WorkdayGap records for the draft WorkOrder and
         renders the resolution form. Each unresolved gap is presented with an
         AbsenceCategory selector and an optional note field.
         Redirects to operator_history if the draft is not found or does not
         belong to the authenticated operator.

    POST /panel/operator/gaps/<int:wo_draft_pk>/
         Validates that every gap has an AbsenceCategory assigned, and that a
         note is provided when AbsenceCategory.requires_note=True.
         On success:
           - Persists AbsenceCategory + note on each WorkdayGap.
           - Marks all gaps as resolved=True.
           - Promotes the WorkOrder from PENDING_GAPS to DONE.
           - Clears the pending_gaps_wo_pk session key.
           - Enqueues classify_fault_line for each entry line.
           - Generates the Excel report synchronously.
           - Redirects to operator_history with a success message.
         The "Volver y editar" action (POST field back_to_form=1) discards
         the draft WorkOrder and redirects to operator_dashboard so the
         operator can resubmit with corrected times.

    ---

    Permite a un operario WORKSHOP justificar cada laguna de jornada detectada
    por Gate 4 antes de que el borrador de parte sea promovido de PENDING_GAPS
    a DONE.

    GET  /panel/operator/gaps/<int:wo_draft_pk>/
         Recupera los registros WorkdayGap pendientes del WorkOrder borrador y
         renderiza el formulario de resolución. Cada gap no resuelto se presenta
         con un selector de AbsenceCategory y un campo de nota opcional.
         Redirige a operator_history si el borrador no se encuentra o no
         pertenece al operario autenticado.

    POST /panel/operator/gaps/<int:wo_draft_pk>/
         Valida que cada gap tenga una AbsenceCategory asignada, y que se
         proporcione nota cuando AbsenceCategory.requires_note=True.
         En caso de éxito:
           - Persiste AbsenceCategory + nota en cada WorkdayGap.
           - Marca todos los gaps como resolved=True.
           - Promueve el WorkOrder de PENDING_GAPS a DONE.
           - Limpia la clave de sesión pending_gaps_wo_pk.
           - Encola classify_fault_line para cada línea del entry.
           - Genera el informe Excel de forma síncrona.
           - Redirige a operator_history con mensaje de éxito.
         La acción "Volver y editar" (campo POST back_to_form=1) descarta el
         borrador WorkOrder y redirige a operator_dashboard para que el operario
         pueda reenviar con horas corregidas.
    """

    template_name ="panel/operator/gap_resolution.html"

    def _get_draft (self ,wo_draft_pk ,company ,cu ):
        """
        Resolves the PENDING_GAPS WorkOrder draft by pk, scoped to the
        authenticated operator and company. Returns None if not found.
        ---
        Resuelve el borrador WorkOrder PENDING_GAPS por pk, acotado al operario
        autenticado y la empresa. Devuelve None si no se encuentra.
        """
        return WorkOrder .objects .filter (
        pk =wo_draft_pk ,
        company =company ,
        uploaded_by =cu ,
        status =WorkOrder .Status .PENDING_GAPS ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).first ()

    def get (self ,request ,wo_draft_pk ,*args ,**kwargs ):
        """
        Renders the gap resolution form for the given PENDING_GAPS draft.
        ---
        Renderiza el formulario de resolución de gaps para el borrador dado.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 
        from work_order_processor .models import WorkdayGap 

        cu =request .user .company_user 
        company =cu .company 

        draft =self ._get_draft (wo_draft_pk ,company ,cu )
        if draft is None :
            django_messages .error (
            request ,
            "El parte borrador no se ha encontrado o ya fue procesado.",
            )
            return redirect (reverse ("panel:operator_history"))

        gaps =WorkdayGap .objects .filter (
        work_order =draft ,resolved =False 
        ).order_by ("gap_start")

        absence_categories =AbsenceCategory .objects .filter (
        company =company ,is_active =True 
        ).order_by ("order","label")

















        first_entry =draft .entries .first ()
        raw_lines =(
        list (first_entry .lines .order_by ("hc"))
        if first_entry else []
        )



        gaps_list =list (gaps )
        gap_by_start ={g .gap_start :g for g in gaps_list }













        for line in raw_lines :
            matched =gap_by_start .get (line .hf )









            line .next_gap =(
            matched 
            if matched and matched .gap_type =="GAP"
            else None 
            )








        late_start_gap =next (
        (g for g in gaps_list if g .gap_type =="LATE_START"),None 
        )
        early_end_gap =next (
        (g for g in gaps_list if g .gap_type =="EARLY_END"),None 
        )

        context ={
        "company":company ,
        "company_user":cu ,
        "active_nav":"operator_dashboard",
        "draft":draft ,
        "gaps":gaps_list ,
        "absence_categories":absence_categories ,
        "work_date":first_entry .work_date if first_entry else None ,
        "entry_lines":raw_lines ,
        "late_start_gap":late_start_gap ,
        "early_end_gap":early_end_gap ,
        }
        return render (request ,self .template_name ,context )

    def post (self ,request ,wo_draft_pk ,*args ,**kwargs ):
        """
        Processes the gap resolution form. On "Volver y editar" discards
        the draft. Otherwise validates and persists all gap justifications,
        promotes the WorkOrder to DONE and enqueues post-processing tasks.
        ---
        Procesa el formulario de resolución de gaps. Con "Volver y editar"
        descarta el borrador. Si no, valida y persiste todas las justificaciones,
        promueve el WorkOrder a DONE y encola las tareas de post-procesamiento.
        """
        from django .urls import reverse 
        from django .db import transaction 
        from ivr_config .models import AbsenceCategory 
        from work_order_processor .models import WorkdayGap 
        from work_order_processor .services import (
        generate_work_order_excel ,
        find_cached_classification ,
        )

        cu =request .user .company_user 
        company =cu .company 

        draft =self ._get_draft (wo_draft_pk ,company ,cu )
        if draft is None :
            django_messages .error (
            request ,
            "El parte borrador no se ha encontrado o ya fue procesado.",
            )
            return redirect (reverse ("panel:operator_history"))















        if request .POST .get ("back_to_form"):
            from django .urls import reverse as _rev_btf 
            from work_order_processor .models import WorkOrder as _WO_BTF 




            _WO_BTF .objects .filter (pk =draft .pk ).update (
            status =_WO_BTF .Status .DONE ,
            )
            request .session .pop ("pending_gaps_wo_pk",None )
            request .session .modified =True 
            django_messages .info (
            request ,
            "Puedes corregir las horas del parte y enviarlo de nuevo.",
            )
            return redirect (
            _rev_btf ("panel:operator_form_edit",kwargs ={"wo_pk":draft .pk })
            )














        gaps =list (
        WorkdayGap .objects .filter (
        work_order =draft ,resolved =False 
        ).order_by ("gap_start")
        )

        validation_errors =[]
        resolutions =[]

        for gap in gaps :
            label_gap =f"Laguna {gap.gap_start:%H:%M}–{gap.gap_end:%H:%M}"

            if gap .gap_type ==WorkdayGap .GapType .LUNCH_BREAK :




                lunch_had_raw =request .POST .get (f"gap_{gap.pk}_lunch_had","").strip ()
                note_val =request .POST .get (f"gap_{gap.pk}_note","").strip ()
                lunch_time_raw =request .POST .get (f"gap_{gap.pk}_lunch_time","").strip ()

                if lunch_had_raw not in ("yes","no"):
                    validation_errors .append (
                    f"{label_gap}: debes indicar si has parado a comer."
                    )
                    continue 

                lunch_had =lunch_had_raw =="yes"

                if not lunch_had and not note_val :
                    validation_errors .append (
                    f"{label_gap}: si no has parado a comer, debes indicar el motivo."
                    )
                    continue 



                lunch_time =None 
                if lunch_had and lunch_time_raw :
                    from datetime import datetime as _dt_lt 
                    try :
                        lunch_time =_dt_lt .strptime (lunch_time_raw ,"%H:%M").time ()
                    except ValueError :
                        lunch_time =None 

                resolutions .append ((gap ,{
                "type":"lunch",
                "lunch_had":lunch_had ,
                "lunch_time":lunch_time ,
                "note":note_val ,
                }))

            else :




                field_cat =f"gap_{gap.pk}_category"
                field_note =f"gap_{gap.pk}_note"

                cat_pk_raw =request .POST .get (field_cat ,"").strip ()
                note_val =request .POST .get (field_note ,"").strip ()

                if not cat_pk_raw :
                    validation_errors .append (
                    f"{label_gap}: debes seleccionar un motivo de ausencia."
                    )
                    continue 

                try :
                    absence_cat =AbsenceCategory .objects .get (
                    pk =int (cat_pk_raw ),company =company ,is_active =True 
                    )
                except (AbsenceCategory .DoesNotExist ,ValueError ,TypeError ):
                    validation_errors .append (
                    f"{label_gap}: la categoría seleccionada no es válida."
                    )
                    continue 

                if absence_cat .requires_note and not note_val :
                    validation_errors .append (
                    f"{label_gap}: la categoría '{absence_cat.label}' "
                    f"requiere una nota explicativa."
                    )
                    continue 

                resolutions .append ((gap ,{
                "type":"standard",
                "absence_cat":absence_cat ,
                "note":note_val ,
                }))

        if validation_errors :


            from ivr_config .models import AbsenceCategory as _AC 
            absence_categories =_AC .objects .filter (
            company =company ,is_active =True 
            ).order_by ("order","label")
            first_entry =draft .entries .first ()
            context ={
            "company":company ,
            "company_user":cu ,
            "active_nav":"operator_dashboard",
            "draft":draft ,
            "gaps":gaps ,
            "absence_categories":absence_categories ,
            "work_date":first_entry .work_date if first_entry else None ,
            "entry_lines":list (first_entry .lines .order_by ("hc"))if first_entry else [],
            "errors":validation_errors ,
            }
            return render (request ,self .template_name ,context )



















        from work_order_processor .management .commands .seed_personal_asset import (
        PERSONAL_ASSET_CODE as _PERSONAL_CODE_GR ,
        )
        from fleet .models import MachineAsset as _MA_GR 
        from work_order_processor .services import _compute_delta_hours as _cdh_gr 
        try :
            _personal_asset_gr =_MA_GR .objects .get (
            code__iexact =_PERSONAL_CODE_GR ,company =company 
            )
        except _MA_GR .DoesNotExist :
            _personal_asset_gr =None 
            logger .warning (
            "# [GapResolutionView] Activo PERSONAL no encontrado para empresa pk=%r. "
            "Las líneas PERSONAL no se crearán.",
            company .pk ,
            )
        try :
            with transaction .atomic ():
                first_entry_gr =draft .entries .first ()


                _next_line_num =1 
                if first_entry_gr :
                    existing_max =(
                    first_entry_gr .lines 
                    .order_by ("-line_number")
                    .values_list ("line_number",flat =True )
                    .first ()
                    )
                    _next_line_num =(existing_max or 0 )+1 
                for gap ,res in resolutions :
                    if res ["type"]=="lunch":
                        gap .lunch_had =res ["lunch_had"]
                        gap .lunch_time =res ["lunch_time"]
                        gap .note =res ["note"]
                        gap .resolved =True 
                        gap .save (update_fields =["lunch_had","lunch_time","note","resolved"])
                    else :
                        gap .absence_category =res ["absence_cat"]
                        gap .note =res ["note"]
                        gap .resolved =True 
                        gap .save (update_fields =["absence_category","note","resolved"])









                        if first_entry_gr is not None and _personal_asset_gr is not None :
                            _gr_delta =_cdh_gr (
                            gap .gap_start ,gap .gap_end ,deduct_lunch =False 
                            )
                            WorkOrderEntryLine .objects .create (
                            entry =first_entry_gr ,
                            line_number =_next_line_num ,
                            machine_asset =_personal_asset_gr ,
                            machine_raw =_personal_asset_gr .code ,
                            machine_norm =_personal_asset_gr .code ,
                            fault_description =res ["absence_cat"].label ,
                            repair_notes =res ["note"],
                            hc =gap .gap_start ,
                            hf =gap .gap_end ,
                            or_val ="",
                            delta_hours =_gr_delta ,
                            flags =[],
                            )
                            logger .info (
                            "# [GapResolutionView] WorkOrderEntryLine PERSONAL creado. "
                            "entry_pk=%r line_number=%r hc=%r hf=%r absence_cat=%r",
                            first_entry_gr .pk ,_next_line_num ,
                            gap .gap_start ,gap .gap_end ,
                            res ["absence_cat"].label ,
                            )
                            _next_line_num +=1 

                draft .status =WorkOrder .Status .DONE 
                draft .save (update_fields =["status"])

        except Exception as exc :
            logger .error (
            "# [GapResolutionView] Error al persistir resoluciones: %s",
            exc ,exc_info =True ,
            )
            django_messages .error (
            request ,
            f"Error al guardar la justificación: {exc}. "
            "Por favor, inténtalo de nuevo.",
            )
            return redirect (
            reverse (
            "panel:operator_gap_resolution",
            kwargs ={"wo_draft_pk":wo_draft_pk },
            )
            )



        request .session .pop ("pending_gaps_wo_pk",None )
        request .session .modified =True 





        first_entry =draft .entries .first ()
        if first_entry :
            for _line in first_entry .lines .all ():
                _cached =find_cached_classification (
                fault_description =_line .fault_description ,
                repair_notes =_line .repair_notes ,
                company =company ,
                )
                if _cached :
                    WorkOrderEntryLine .objects .filter (pk =_line .pk ).update (
                    fault_category =_cached [0 ],
                    fault_subcategory =_cached [1 ],
                    )
                    logger .info (
                    "# [GapResolutionView] Clasificación copiada desde caché "
                    "para WorkOrderEntryLine pk=%d: category=%s subcategory=%s.",
                    _line .pk ,_cached [0 ],_cached [1 ],
                    )
                else :
                    classify_fault_line .apply_async (
                    args =[_line .pk ],
                    queue ="work_orders",
                    )
                    logger .info (
                    "# [GapResolutionView] classify_fault_line encolada "
                    "para WorkOrderEntryLine pk=%d.",
                    _line .pk ,
                    )

        try :
            generate_work_order_excel (draft .pk )
            logger .info (
            "# [GapResolutionView] Excel generado para WorkOrder #%d.",
            draft .pk ,
            )
        except Exception as exc :
            logger .warning (
            "# [GapResolutionView] Excel no generado para WorkOrder #%d: %s.",
            draft .pk ,exc ,
            )

        django_messages .success (
        request ,
        f"Parte #{draft.pk} registrado correctamente. "
        f"Todas las lagunas de jornada han sido justificadas.",
        )
        return redirect (reverse ("panel:operator_history"))


class WorkdayScheduleView (SupervisorAccessMixin ,View ):
    """
    Creates or updates the WorkdaySchedule records for the authenticated
    user's company. Supports multiple named schedules per company.

    GET  /panel/workday-schedule/
         Renders the schedule management page with the list of existing
         schedules and a creation form.

    POST /panel/workday-schedule/   (action=create)
         Creates a new WorkdaySchedule for the company.

    POST /panel/workday-schedule/   (action=update, schedule_pk=<pk>)
         Updates an existing WorkdaySchedule.

    POST /panel/workday-schedule/   (action=delete, schedule_pk=<pk>)
         Deletes a WorkdaySchedule. Clears workday_schedule FK on all
         CompanyUsers assigned to it before deletion.

    ---

    Crea o actualiza los registros WorkdaySchedule de la empresa del usuario
    autenticado. Soporta múltiples horarios con nombre por empresa.

    GET  /panel/workday-schedule/
         Renderiza la página de gestión de horarios con la lista de horarios
         existentes y un formulario de creación.

    POST /panel/workday-schedule/   (action=create)
         Crea un nuevo WorkdaySchedule para la empresa.

    POST /panel/workday-schedule/   (action=update, schedule_pk=<pk>)
         Actualiza un WorkdaySchedule existente.

    POST /panel/workday-schedule/   (action=delete, schedule_pk=<pk>)
         Elimina un WorkdaySchedule. Limpia la FK workday_schedule en todos
         los CompanyUsers asignados a él antes de la eliminación.
    """

    template_name ="panel/workday/schedule_form.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _parse_time (self ,value ):
        """
        Parses an HH:MM time string and returns a time object, or None.
        ---
        Parsea una cadena HH:MM y devuelve un objeto time, o None.
        """
        from datetime import datetime 
        if not value :
            return None 
        try :
            return datetime .strptime (value .strip (),"%H:%M").time ()
        except ValueError :
            return None 

    def _build_context (self ,request ,errors =None ,post_data =None ):
        """
        Builds the template context with the list of existing schedules,
        the creation form data and validation errors.
        ---
        Construye el contexto con la lista de horarios existentes, los datos
        del formulario de creación y los errores de validación.
        """
        from ivr_config .models import WorkdaySchedule 
        cu =request .user .company_user 
        company =cu .company 
        return {
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"workday_schedule",
        "schedules":WorkdaySchedule .objects .filter (
        company =company 
        ).order_by ("label"),
        "errors":errors or [],
        "post_data":post_data or {},
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the workday schedule management page.
        ---
        Renderiza la página de gestión de horarios de jornada.
        """
        return render (request ,self .template_name ,self ._build_context (request ))

    def post (self ,request ,*args ,**kwargs ):
        """
        Dispatches create, update and delete actions for WorkdaySchedule.
        ---
        Despacha las acciones create, update y delete para WorkdaySchedule.
        """
        from django .urls import reverse 
        from ivr_config .models import WorkdaySchedule 

        cu =request .user .company_user 
        company =cu .company 
        action =request .POST .get ("action","").strip ()

        REDIRECT_URL =reverse ("panel:workday_schedule")





        if action =="delete":
            try :
                pk =int (request .POST .get ("schedule_pk",""))
                sched =WorkdaySchedule .objects .get (pk =pk ,company =company )


                CompanyUser .objects .filter (
                company =company ,workday_schedule =sched 
                ).update (workday_schedule =None )
                label =sched .label 
                sched .delete ()
                django_messages .success (
                request ,
                f"Horario '{label}' eliminado correctamente.",
                )
            except (WorkdaySchedule .DoesNotExist ,ValueError ,TypeError ):
                django_messages .error (
                request ,
                "Horario no encontrado o no pertenece a esta empresa.",
                )
            return redirect (REDIRECT_URL )





        from ivr_config .models import WorkdaySchedule as _WDS_choices 

        label_val =request .POST .get ("label","").strip ()
        season_val =request .POST .get ("season",_WDS_choices .Season .WINTER ).strip ()
        is_intensive_val =bool (request .POST .get ("is_intensive"))
        start_morning_val =request .POST .get ("start_time_morning","").strip ()
        end_morning_val =request .POST .get ("end_time_morning","").strip ()
        start_afternoon_val =request .POST .get ("start_time_afternoon","").strip ()
        end_afternoon_val =request .POST .get ("end_time_afternoon","").strip ()
        tol_val =request .POST .get ("tolerance_minutes","15").strip ()
        is_default_val =bool (request .POST .get ("is_default"))

        errors =[]
        if not label_val :
            errors .append ("El nombre del horario es obligatorio.")

        if season_val not in dict (_WDS_choices .Season .choices ):
            errors .append ("La temporada seleccionada no es válida.")
            season_val =_WDS_choices .Season .WINTER 

        start_morning =self ._parse_time (start_morning_val )
        if start_morning is None :
            errors .append ("La hora de entrada de mañana debe tener formato HH:MM.")

        end_morning =self ._parse_time (end_morning_val )
        if end_morning is None :
            errors .append ("La hora de salida de mañana debe tener formato HH:MM.")

        if start_morning and end_morning and end_morning <=start_morning :
            errors .append ("La hora de salida de mañana debe ser posterior a la de entrada.")



        start_afternoon =None 
        end_afternoon =None 
        if not is_intensive_val :
            start_afternoon =self ._parse_time (start_afternoon_val )
            end_afternoon =self ._parse_time (end_afternoon_val )
            if start_afternoon is None :
                errors .append ("La hora de entrada de tarde es obligatoria para jornada partida.")
            if end_afternoon is None :
                errors .append ("La hora de salida de tarde es obligatoria para jornada partida.")
            if start_afternoon and end_morning and start_afternoon <=end_morning :
                errors .append (
                "La hora de entrada de tarde debe ser posterior a la de salida de mañana."
                )
            if start_afternoon and end_afternoon and end_afternoon <=start_afternoon :
                errors .append ("La hora de salida de tarde debe ser posterior a la de entrada.")

        try :
            tolerance =int (tol_val )
            if tolerance <0 :
                raise ValueError 
        except (ValueError ,TypeError ):
            errors .append ("La tolerancia debe ser un número entero positivo.")
            tolerance =15 

        if errors :
            return render (
            request ,
            self .template_name ,
            self ._build_context (request ,errors =errors ,post_data =request .POST ),
            )





        if action =="update":
            try :
                pk =int (request .POST .get ("schedule_pk",""))
                sched =WorkdaySchedule .objects .get (pk =pk ,company =company )
                sched .label =label_val 
                sched .season =season_val 
                sched .is_intensive =is_intensive_val 
                sched .start_time_morning =start_morning 
                sched .end_time_morning =end_morning 
                sched .start_time_afternoon =start_afternoon 
                sched .end_time_afternoon =end_afternoon 
                sched .tolerance_minutes =tolerance 
                sched .is_default =is_default_val 
                sched .save ()
                django_messages .success (
                request ,
                f"Horario '{label_val}' actualizado correctamente.",
                )
            except (WorkdaySchedule .DoesNotExist ,ValueError ,TypeError ):
                django_messages .error (
                request ,
                "Horario no encontrado o no pertenece a esta empresa.",
                )
            return redirect (REDIRECT_URL )





        WorkdaySchedule .objects .create (
        company =company ,
        label =label_val ,
        season =season_val ,
        is_intensive =is_intensive_val ,
        start_time_morning =start_morning ,
        end_time_morning =end_morning ,
        start_time_afternoon =start_afternoon ,
        end_time_afternoon =end_afternoon ,
        tolerance_minutes =tolerance ,
        is_default =is_default_val ,
        )
        django_messages .success (
        request ,
        f"Horario '{label_val}' creado correctamente.",
        )
        return redirect (REDIRECT_URL )


class AbsenceCategoryListView (SupervisorAccessMixin ,View ):
    """
    Lists all AbsenceCategory records for the authenticated user's company.
    Renders the full absence category management page on GET.

    GET /panel/absence-categories/
    ---
    Lista todos los registros AbsenceCategory de la empresa del usuario
    autenticado. Renderiza la página completa de gestión en GET.

    GET /panel/absence-categories/
    """

    template_name ="panel/workday/absence_category_list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the absence category list page.
        ---
        Renderiza la página de listado de categorías de ausencia.
        """
        from ivr_config .models import AbsenceCategory 
        cu =request .user .company_user 
        company =cu .company 
        categories =AbsenceCategory .objects .filter (
        company =company 
        ).order_by ("order","label")
        context ={
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"absence_categories",
        "categories":categories ,
        }
        return render (request ,self .template_name ,context )


class AbsenceCategoryCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a new AbsenceCategory for the authenticated user's company.
    On success or failure, redirects to absence_category_list.

    POST /panel/absence-categories/create/
    ---
    Crea una nueva AbsenceCategory para la empresa del usuario autenticado.
    En caso de éxito o fallo, redirige a absence_category_list.

    POST /panel/absence-categories/create/
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates POST data and creates the AbsenceCategory.
        ---
        Valida los datos POST y crea la AbsenceCategory.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        label_val =request .POST .get ("label","").strip ()
        code_val =request .POST .get ("code","").strip ().upper ()
        requires_note =bool (request .POST .get ("requires_note"))
        is_justified =bool (request .POST .get ("is_justified"))
        order_val =request .POST .get ("order","0").strip ()

        if not label_val or not code_val :
            django_messages .error (
            request ,
            "El nombre y el código son obligatorios.",
            )
            return redirect (LIST_URL )

        try :
            order =int (order_val )
        except (ValueError ,TypeError ):
            order =0 

        if AbsenceCategory .objects .filter (company =company ,code =code_val ).exists ():
            django_messages .error (
            request ,
            f"Ya existe una categoría con el código '{code_val}' en esta empresa.",
            )
            return redirect (LIST_URL )

        AbsenceCategory .objects .create (
        company =company ,
        code =code_val ,
        label =label_val ,
        requires_note =requires_note ,
        is_justified =is_justified ,
        order =order ,
        is_active =True ,
        )
        django_messages .success (
        request ,
        f"Categoría '{label_val}' creada correctamente.",
        )
        return redirect (LIST_URL )


class AbsenceCategoryUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing AbsenceCategory belonging to the authenticated
    user's company. On success or failure, redirects to absence_category_list.

    POST /panel/absence-categories/<pk>/update/
    ---
    Actualiza una AbsenceCategory existente de la empresa del usuario
    autenticado. En caso de éxito o fallo, redirige a absence_category_list.

    POST /panel/absence-categories/<pk>/update/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates POST data and updates the AbsenceCategory.
        ---
        Valida los datos POST y actualiza la AbsenceCategory.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        try :
            category =AbsenceCategory .objects .get (pk =pk ,company =company )
        except AbsenceCategory .DoesNotExist :
            django_messages .error (
            request ,
            "Categoría no encontrada o no pertenece a esta empresa.",
            )
            return redirect (LIST_URL )

        label_val =request .POST .get ("label","").strip ()
        requires_note =bool (request .POST .get ("requires_note"))
        is_justified =bool (request .POST .get ("is_justified"))
        order_val =request .POST .get ("order","0").strip ()

        if not label_val :
            django_messages .error (request ,"El nombre es obligatorio.")
            return redirect (LIST_URL )

        try :
            order =int (order_val )
        except (ValueError ,TypeError ):
            order =category .order 

        category .label =label_val 
        category .requires_note =requires_note 
        category .is_justified =is_justified 
        category .order =order 
        category .save (update_fields =[
        "label","requires_note","is_justified","order"
        ])

        django_messages .success (
        request ,
        f"Categoría '{label_val}' actualizada correctamente.",
        )
        return redirect (LIST_URL )


class AbsenceCategoryToggleView (SupervisorAccessMixin ,View ):
    """
    Toggles the is_active flag of an AbsenceCategory. Deactivating a
    category hides it from the operator's gap resolution selector without
    deleting historical references in WorkdayGap records.

    POST /panel/absence-categories/<pk>/toggle/
    ---
    Alterna el flag is_active de una AbsenceCategory. Desactivar una
    categoría la oculta del selector del operario sin eliminar las
    referencias históricas en registros WorkdayGap.

    POST /panel/absence-categories/<pk>/toggle/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Toggles is_active and redirects to absence_category_list.
        ---
        Alterna is_active y redirige a absence_category_list.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        try :
            category =AbsenceCategory .objects .get (pk =pk ,company =company )
        except AbsenceCategory .DoesNotExist :
            django_messages .error (
            request ,
            "Categoría no encontrada o no pertenece a esta empresa.",
            )
            return redirect (LIST_URL )

        category .is_active =not category .is_active 
        category .save (update_fields =["is_active"])

        state ="activada"if category .is_active else "desactivada"
        django_messages .success (
        request ,
        f"Categoría '{category.label}' {state} correctamente.",
        )
        return redirect (LIST_URL )


class WorkOrderAdminHistoryView (SupervisorAccessMixin ,View ):

    """
    Four-tab management history view for ADMIN and SUPERVISOR roles.
    Provides a comprehensive interface for reviewing, exporting and managing
    all operators' work orders, absences and work periods within the company.

    Tabs:
      1 — Pendientes  : unreviewed work orders for all company operators.
                        Filters: operator, date, machine.
                        Actions: mark as reviewed, link to edit view.
      2 — Revisados   : reviewed work orders. Excel export available HERE ONLY
                        (single work order or date-range multi-export).
                        Filters: operator, date range, machine.
      3 — Histórico   : all work orders with cross-filters (operator, date
                        range, machine, review status). Read-only.
      4 — Ausencias   : WorkerAbsence records per operator with type/date
                        filters. Actions: create, edit, delete absence.
                        Action 'Generar partes del periodo': creates synthetic
                        WorkOrder records (one per working day Mon–Fri) for the
                        absence range, tagged with generated_by=current user.

    GET /panel/work-orders/history/
        Optional GET params:
          tab         (str)  — active tab: pending|reviewed|history|absences.
                               Default: pending.
          operator_pk (int)  — filter by CompanyUser pk (scoped to company).
          date_from   (str)  — ISO date YYYY-MM-DD start of range.
          date_to     (str)  — ISO date YYYY-MM-DD end of range.
          machine     (str)  — MachineAsset.code icontains filter.

    ---

    Vista de historial de gestion de cuatro pestanas para roles ADMIN y SUPERVISOR.
    Proporciona una interfaz completa para revisar, exportar y gestionar todos los
    partes de trabajo, ausencias y periodos de trabajo de los operarios de la empresa.

    Pestanas:
      1 — Pendientes  : partes sin revisar de todos los operarios de la empresa.
                        Filtros: operario, fecha, maquina.
                        Acciones: marcar como revisado, enlace a vista de edicion.
      2 — Revisados   : partes revisados. Exportacion Excel disponible SOLO AQUI
                        (parte individual o multi-exportacion por rango de fechas).
                        Filtros: operario, rango de fechas, maquina.
      3 — Historico   : todos los partes con filtros cruzados (operario, rango de
                        fechas, maquina, estado de revision). Solo lectura.
      4 — Ausencias   : registros WorkerAbsence por operario con filtros de tipo y
                        fecha. Acciones: alta, edicion, baja de ausencia.
                        Accion 'Generar partes del periodo': crea registros WorkOrder
                        sinteticos (uno por dia laborable lun–vie) para el rango de
                        la ausencia, etiquetados con generated_by=usuario actual.

    GET /panel/work-orders/history/
        Parametros GET opcionales:
          tab         (str)  — pestana activa: pending|reviewed|history|absences.
                               Por defecto: pending.
          operator_pk (int)  — filtrar por pk de CompanyUser (acotado a empresa).
          date_from   (str)  — fecha ISO YYYY-MM-DD inicio del rango.
          date_to     (str)  — fecha ISO YYYY-MM-DD fin del rango.
          machine     (str)  — filtro icontains sobre MachineAsset.code.
    """

    template_name ="panel/work_orders/admin_history.html"



    _MESES_ES ={
    1 :"Enero",2 :"Febrero",3 :"Marzo",4 :"Abril",
    5 :"Mayo",6 :"Junio",7 :"Julio",8 :"Agosto",
    9 :"Septiembre",10 :"Octubre",11 :"Noviembre",12 :"Diciembre",
    }

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _parse_date (self ,value ):
        """
        Parses an ISO date string (YYYY-MM-DD) and returns a date object,
        or None if the value is absent or malformed.
        ---
        Parsea una cadena de fecha ISO (YYYY-MM-DD) y devuelve un objeto date,
        o None si el valor esta ausente o malformado.
        """
        from datetime import datetime 
        if not value :
            return None 
        try :
            return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
        except ValueError :
            return None 

    def _build_base_queryset (self ,company ):
        """
        Returns the base WorkOrder queryset scoped to the company, restricted
        to DIGITAL and GENERATED sources (operator-entered parts only).
        PDF_UPLOAD parts belong exclusively to WorkOrderListView.
        All required related data is prefetched for the admin history tabs.
        ---
        Devuelve el queryset base de WorkOrder acotado a la empresa, restringido
        a los origenes DIGITAL y GENERATED (partes introducidos por operarios).
        Los partes PDF_UPLOAD pertenecen exclusivamente a WorkOrderListView.
        Todos los datos relacionados se prefetchean para las pestanas del historial.
        """
        return (
        WorkOrder .objects 
        .filter (
        company =company ,
        source__in =[
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        ],
        )
        .select_related ("uploaded_by__user","reviewed_by__user","generated_by__user")
        .prefetch_related (
        Prefetch (
        "entries",
        queryset =WorkOrderEntry .objects .prefetch_related ("lines"),
        )
        )
        .order_by ("-id")
        )

    def _apply_filters (self ,qs ,operator_pk ,date_from ,date_to ,machine ,company ,
    fault_category ="",q =""):
        """
        Applies the optional GET filters to a WorkOrder queryset.
        operator_pk filters by uploaded_by; date_from/date_to filter by the
        first entry's work_date; machine filters by machine_asset__code.
        fault_category filters by WorkOrderEntryLine.fault_category (exact).
        q performs a free-text OR search over fault_description and repair_notes.
        ---
        Aplica los filtros GET opcionales a un queryset de WorkOrder.
        operator_pk filtra por uploaded_by; date_from/date_to filtran por
        work_date del primer entry; machine filtra por machine_asset__code.
        fault_category filtra por WorkOrderEntryLine.fault_category (exacto).
        q realiza una búsqueda libre OR sobre fault_description y repair_notes.
        """
        if operator_pk :
            try :
                cu_pk =int (operator_pk )
                qs =qs .filter (uploaded_by__pk =cu_pk ,uploaded_by__company =company )
            except (ValueError ,TypeError ):
                pass 
        if date_from :
            qs =qs .filter (entries__work_date__gte =date_from )
        if date_to :
            qs =qs .filter (entries__work_date__lte =date_to )
        if machine :
            qs =qs .filter (
            entries__lines__machine_asset__code__icontains =machine 
            )
        if fault_category :


            qs =qs .filter (
            entries__lines__fault_category =fault_category 
            ).distinct ()
        if q :


            from django .db .models import Q as _Q 
            qs =qs .filter (
            _Q (entries__lines__fault_description__icontains =q )
            |_Q (entries__lines__repair_notes__icontains =q )
            )
        return qs .distinct ()

    def _enrich_work_orders (self ,qs ,active_fault_category =""):
        """
        Converts a WorkOrder queryset into a list of enriched dicts suitable
        for template rendering. Each dict includes pk, fecha, operator name,
        num_bloques, horas_totales and reviewed flag.
        When active_fault_category is provided (a FaultCategory internal value),
        the fault_category badge of every enriched dict is forced to the label
        of that category rather than being calculated from the dominant category
        across all lines. This ensures the badge is always coherent with the
        active filter — a work order returned by the filter is guaranteed to have
        at least one line matching the category, so forcing the label is correct.
        ---
        Convierte un queryset de WorkOrder en una lista de dicts enriquecidos
        adecuados para renderizado en template. Cada dict incluye pk, fecha,
        nombre del operario, num_bloques, horas_totales y flag reviewed.
        Cuando active_fault_category contiene un valor interno de FaultCategory,
        el badge fault_category de cada dict se fuerza al label de esa categoría
        en lugar de calcularse como la dominante sobre todas las líneas. Esto
        garantiza coherencia visual entre el filtro activo y el badge mostrado —
        un parte devuelto por el filtro tiene garantizada al menos una línea con
        esa categoría, por lo que forzar el label es semánticamente correcto.
        """
        from decimal import Decimal 
        result =[]
        for wo in qs :
            entries_list =list (wo .entries .all ())
            first_entry =entries_list [0 ]if entries_list else None 
            work_date =first_entry .work_date if first_entry else None 
            num_bloques =sum (entry .lines .count ()for entry in entries_list )
            horas_totales =sum (
            (line .delta_hours 
            for entry in entries_list 
            for line in entry .lines .all ()
            if line .delta_hours is not None ),
            Decimal ("0"),
            )
            operator_name =(
            wo .uploaded_by .user .get_full_name ()or wo .uploaded_by .user .username 
            if wo .uploaded_by else "Desconocido"
            )
            result .append ({
            "pk":wo .pk ,
            "fecha":work_date ,
            "operator_name":operator_name ,
            "operator_pk":wo .uploaded_by .pk if wo .uploaded_by else None ,
            "num_bloques":num_bloques ,
            "horas_totales":horas_totales ,
            "reviewed":wo .reviewed ,
            "reviewed_by":(
            wo .reviewed_by .user .get_full_name ()or wo .reviewed_by .user .username 
            if wo .reviewed_by else None 
            ),
            "reviewed_at":wo .reviewed_at ,
            "generated_by":(
            wo .generated_by .user .get_full_name ()or wo .generated_by .user .username 
            if wo .generated_by else None 
            ),
            "excel_url":wo .excel_file .url if wo .excel_file else None ,
            "fault_category":self ._dominant_fault_category (
            wo ,active_fault_category =active_fault_category 
            ),
            })
        return result 

    def _dominant_fault_category (self ,wo ,active_fault_category =""):
        """
        Returns the display label of the most frequent fault_category among all
        WorkOrderEntryLine records of this WorkOrder. Returns an empty string
        when no lines exist or all lines have an empty category.

        When active_fault_category is supplied (a FaultCategory internal value
        such as "ENGINE_TRANSMISSION"), the method short-circuits: it resolves
        the label for that value from FaultCategory.choices and returns it
        immediately, bypassing the frequency calculation. This is correct because
        any WorkOrder present in a fault_category-filtered queryset is guaranteed
        to contain at least one line with that exact category value — there is no
        semantic loss in forcing the label. The short-circuit also avoids
        redundant iteration over prefetched lines on every row render.
        ---
        Devuelve el label del fault_category más frecuente entre todas las
        WorkOrderEntryLine del WorkOrder. Devuelve cadena vacía cuando no hay
        líneas o todas tienen la categoría vacía.

        Cuando active_fault_category contiene un valor interno de FaultCategory
        (ej. "ENGINE_TRANSMISSION"), el método hace cortocircuito: resuelve el
        label de ese valor desde FaultCategory.choices y lo devuelve directamente,
        sin ejecutar el cálculo de frecuencia. Esto es semánticamente correcto
        porque cualquier WorkOrder presente en un queryset filtrado por
        fault_category tiene garantizada al menos una línea con ese valor exacto.
        El cortocircuito también evita iteración redundante sobre las líneas
        prefetcheadas en cada fila del render.
        """
        from collections import Counter 
        from work_order_processor .models import FaultCategory 
        label_map ={fc [0 ]:str (fc [1 ])for fc in FaultCategory .choices }







        if active_fault_category :
            label =label_map .get (active_fault_category )
            if label :
                return label 



        categories =[
        line .fault_category 
        for entry in wo .entries .all ()
        for line in entry .lines .all ()
        if line .fault_category 
        ]
        if not categories :
            return ""
        dominant =Counter (categories ).most_common (1 )[0 ][0 ]
        return label_map .get (dominant ,dominant )

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the four-tab admin history page. Resolves GET filters,
        builds each tab's queryset independently and passes all data to
        the template context.
        ---
        Renderiza la pagina de historial de administrador de cuatro pestanas.
        Resuelve los filtros GET, construye el queryset de cada pestana
        de forma independiente y pasa todos los datos al contexto del template.
        """
        cu =request .user .company_user 
        company =cu .company 




        active_tab =request .GET .get ("tab","pending")
        operator_pk =request .GET .get ("operator_pk","").strip ()
        date_from =self ._parse_date (request .GET .get ("date_from",""))
        date_to =self ._parse_date (request .GET .get ("date_to",""))
        machine =request .GET .get ("machine","").strip ()
        fault_category =request .GET .get ("fault_category","").strip ()
        q =request .GET .get ("q","").strip ()









        _VALID_SORT_COLS =frozenset (
        ("fecha","operator_name","horas_totales","reviewed","fault_category")
        )
        _VALID_SORT_DIRS =frozenset (("asc","desc"))

        raw_stack =request .GET .get ("sort_stack","").strip ()



        sort_stack =[]
        if raw_stack :
            for token in raw_stack .split (","):
                token =token .strip ()
                if ":"not in token :
                    continue 
                col ,_ ,direction =token .partition (":")
                col =col .strip ()
                direction =direction .strip ()
                if col in _VALID_SORT_COLS and direction in _VALID_SORT_DIRS :


                    if not any (s [0 ]==col for s in sort_stack ):
                        sort_stack .append ((col ,direction ))
                        if len (sort_stack )>=3 :
                            break 



        if not sort_stack :
            sort_stack =[("fecha","asc")]



        sort_primary_col =sort_stack [0 ][0 ]
        sort_primary_dir =sort_stack [0 ][1 ]



        sort_stack_str =",".join (f"{col}:{direction}"for col ,direction in sort_stack )





        operators =(
        CompanyUser .objects 
        .filter (
        company =company ,
        is_active =True ,
        role =CompanyUser .ROLE_WORKSHOP ,
        )
        .select_related ("user")
        .order_by ("user__last_name","user__first_name")
        )





        qs_pending =self ._build_base_queryset (company ).filter (reviewed =False )
        qs_pending =self ._apply_filters (
        qs_pending ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,
        )
        pending_list =self ._enrich_work_orders (qs_pending ,active_fault_category =fault_category )





        qs_reviewed =self ._build_base_queryset (company ).filter (reviewed =True )
        qs_reviewed =self ._apply_filters (
        qs_reviewed ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,
        )
        reviewed_list =self ._enrich_work_orders (qs_reviewed ,active_fault_category =fault_category )





        qs_history =self ._build_base_queryset (company ).filter (reviewed =True )
        qs_history =self ._apply_filters (
        qs_history ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,
        )
        history_list =self ._enrich_work_orders (qs_history ,active_fault_category =fault_category )


















        def _sort_key_for_col (item ,col ):
            """
            Returns a comparable sort key for a single column in an enriched
            work-order dict. None values and empty strings always sort last,
            regardless of the sort direction applied by the caller.
            ---
            Devuelve una clave de ordenación comparable para una sola columna
            en un dict enriquecido de parte. None y cadena vacía se ordenan
            siempre al final, independientemente de la dirección aplicada.
            """
            val =item .get (col )
            if val is None or val =="":


                return (1 ,"")
            return (0 ,val )

        def _apply_sort_stack (lst ,stack ):
            """
            Applies a multi-level stable sort to lst according to stack.
            stack is a list of (col, dir) pairs ordered from most to least
            significant. Sorting is applied from least significant to most
            significant to leverage Python's stable sort.
            ---
            Aplica una ordenación estable multi-nivel a lst según stack.
            stack es una lista de pares (col, dir) de más a menos significativo.
            La ordenación se aplica de menos significativo a más significativo
            para aprovechar la estabilidad del sort de Python.
            """
            result =list (lst )
            for col ,direction in reversed (stack ):
                reverse_flag =(direction =="desc")
                result =sorted (
                result ,
                key =lambda item ,_col =col :_sort_key_for_col (item ,_col ),
                reverse =reverse_flag ,
                )
            return result 

        pending_list =_apply_sort_stack (pending_list ,sort_stack )
        reviewed_list =_apply_sort_stack (reviewed_list ,sort_stack )
        history_list =_apply_sort_stack (history_list ,sort_stack )





        from ivr_config .models import WorkerAbsence 
        absence_qs =(
        WorkerAbsence .objects 
        .filter (company_user__company =company )
        .select_related (
        "company_user__user",
        "registered_by__user",
        )
        .order_by ("-start_date")
        )


        if operator_pk :
            try :
                absence_qs =absence_qs .filter (company_user__pk =int (operator_pk ))
            except (ValueError ,TypeError ):
                pass 
        absences_list =list (absence_qs )





        from ivr_config .models import WorkPeriod 
        from datetime import timedelta as _td 
        last_closed =(
        WorkPeriod .objects 
        .filter (
        company_user__company =company ,
        end_date__isnull =False ,
        )
        .order_by ("-end_date")
        .values_list ("end_date","start_date")
        .first ()
        )
        if last_closed :
            _last_end =last_closed [0 ]
            _last_start =last_closed [1 ]
            _period_len =_last_end -_last_start 
            suggested_period_start =(_last_end +_td (days =1 )).strftime ("%Y-%m-%d")
            suggested_period_end =(_last_end +_td (days =1 )+_period_len ).strftime ("%Y-%m-%d")
        else :
            suggested_period_start =""
            suggested_period_end =""

        context ={
        "company":cu .company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"work_order_admin_history",
        "active_tab":active_tab ,
        "operators":operators ,
        "operator_pk":operator_pk ,
        "date_from":request .GET .get ("date_from",""),
        "date_to":request .GET .get ("date_to",""),
        "machine":machine ,
        "fault_category":fault_category ,
        "q":q ,


        "fault_category_choices":[
        (fc [0 ],fc [1 ])
        for fc in WorkOrderEntryLine .fault_category .field .choices 
        if fc [0 ]
        ],



        "sort_stack_str":sort_stack_str ,
        "sort_primary_col":sort_primary_col ,
        "sort_primary_dir":sort_primary_dir ,
        "pending_list":pending_list ,
        "reviewed_list":reviewed_list ,
        "history_list":history_list ,
        "absences_list":absences_list ,
        "suggested_period_start":suggested_period_start ,
        "suggested_period_end":suggested_period_end ,
        }
        return render (request ,self .template_name ,context )

    def post (self ,request ,*args ,**kwargs ):
        """
        Dispatches POST actions for the admin history view.

        Supported actions:
          generate_absence_parts — creates synthetic WorkOrders from a WorkerAbsence.
          delete_work_order      — deletes a single WorkOrder by pk.
          bulk_action            — applies mark_reviewed, unmark_reviewed or delete
                                   to a set of PKs.

        ---

        Despacha las acciones POST de la vista de historial de administrador.

        Acciones soportadas:
          generate_absence_parts — crea WorkOrders sinteticos desde un WorkerAbsence.
          delete_work_order      — elimina un WorkOrder individual por pk.
          bulk_action            — aplica mark_reviewed, unmark_reviewed o delete
                                   a un conjunto de PKs.
        """
        from datetime import timedelta ,date as dt_date 
        from django .db import transaction 
        from ivr_config .models import WorkerAbsence 
        from work_order_processor .models import WorkOrder ,WorkOrderEntry ,WorkOrderEntryLine 
        from django .urls import reverse 

        cu =request .user .company_user 
        company =cu .company 
        action =request .POST .get ("action","").strip ()





        if action =="delete_work_order":
            active_tab =request .POST .get ("active_tab","pending")
            try :
                wo_pk =int (request .POST .get ("work_order_pk",""))
                work_order =WorkOrder .objects .get (
                pk =wo_pk ,
                company =company ,
                source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
                )
                work_order .delete ()
                django_messages .success (request ,f"Parte #{wo_pk} eliminado correctamente.")
            except (ValueError ,TypeError ,WorkOrder .DoesNotExist ):
                django_messages .error (
                request ,
                "Parte no encontrado o no pertenece a esta empresa.",
                )
            return redirect (
            reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
            )





        if action =="bulk_action":
            active_tab =request .POST .get ("active_tab","pending")
            bulk_op =request .POST .get ("bulk_op","").strip ()
            raw_pks =request .POST .getlist ("selected_pks")

            try :
                pk_list =[int (p )for p in raw_pks if p .strip ().isdigit ()]
            except (ValueError ,TypeError ):
                pk_list =[]

            if not pk_list :
                django_messages .warning (request ,"No se ha seleccionado ningún parte.")
                return redirect (
                reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
                )



            qs =WorkOrder .objects .filter (
            pk__in =pk_list ,
            company =company ,
            source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
            )

            if bulk_op =="mark_reviewed":
                from django .utils .timezone import now as tz_now 
                updated =qs .filter (reviewed =False ).update (
                reviewed =True ,
                reviewed_by =cu ,
                reviewed_at =tz_now (),
                )
                django_messages .success (
                request ,
                f"{updated} parte(s) marcado(s) como revisado(s).",
                )
            elif bulk_op =="unmark_reviewed":


                updated =qs .filter (reviewed =True ).update (
                reviewed =False ,
                reviewed_by =None ,
                reviewed_at =None ,
                )
                django_messages .success (
                request ,
                f"{updated} parte(s) desmarcado(s) como revisado(s).",
                )
            elif bulk_op =="delete":
                count =qs .count ()
                qs .delete ()
                django_messages .success (
                request ,
                f"{count} parte(s) eliminado(s) correctamente.",
                )
            else :
                django_messages .error (request ,"Operación en bloque no reconocida.")

            return redirect (
            reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
            )

        if action !="generate_absence_parts":
            django_messages .error (request ,"Acción no reconocida.")
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        try :
            absence_pk =int (request .POST .get ("absence_pk",""))
            absence =WorkerAbsence .objects .select_related ("company_user__user").get (
            pk =absence_pk ,
            company_user__company =company ,
            )
        except (ValueError ,TypeError ,WorkerAbsence .DoesNotExist ):
            django_messages .error (
            request ,
            "Registro de ausencia no encontrado o no pertenece a esta empresa.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        current_day =absence .start_date 
        working_days =[]
        while current_day <=absence .end_date :
            if current_day .weekday ()<5 :
                working_days .append (current_day )
            current_day +=timedelta (days =1 )

        if not working_days :
            django_messages .warning (
            request ,
            f"El rango de la ausencia ({absence.start_date} – {absence.end_date}) "
            f"no contiene días laborables (lunes a viernes). No se han generado partes.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")

        created_count =0 
        skipped_count =0 

        try :
            with transaction .atomic ():
                for work_day in working_days :


                    already_exists =WorkOrder .objects .filter (
                    company =company ,
                    uploaded_by =absence .company_user ,
                    entries__work_date =work_day ,
                    ).exists ()
                    if already_exists :
                        skipped_count +=1 
                        continue 



                    synthetic_name =(
                    f"AUSENCIA_{absence.get_absence_type_display().upper()}_"
                    f"{work_day.strftime('%Y%m%d')}_"
                    f"{absence.company_user.user.username.upper()}.pdf"
                    )

                    work_order =WorkOrder .objects .create (
                    company =company ,
                    uploaded_by =absence .company_user ,
                    generated_by =cu ,
                    source =WorkOrder .Source .GENERATED ,
                    status =WorkOrder .Status .DONE ,
                    )


                    work_order .source_pdf .name =synthetic_name 
                    work_order .save ()

                    entry =WorkOrderEntry .objects .create (
                    work_order =work_order ,
                    page_number =1 ,
                    worker_name =(
                    absence .company_user .user .get_full_name ()
                    or absence .company_user .user .username 
                    ).upper (),
                    work_date =work_day ,
                    uncertain_date =False ,
                    extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                    raw_gemini_response =None ,
                    )

                    WorkOrderEntryLine .objects .create (
                    entry =entry ,
                    line_number =1 ,
                    machine_asset =None ,
                    machine_raw ="",
                    machine_norm ="",
                    fault_description =absence .get_absence_type_display (),
                    repair_notes ="",
                    hc =None ,
                    hf =None ,
                    or_val ="",
                    delta_hours =8 ,
                    flags =[],
                    )
                    created_count +=1 

        except Exception as exc :
            logger .error (
            "# [AdminHistory] Error generando partes de ausencia pk=%d: %s",
            absence .pk ,exc ,exc_info =True ,
            )
            django_messages .error (
            request ,
            f"Error al generar los partes: {exc}. "
            "Por favor, inténtalo de nuevo o contacta con el administrador.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        msg_parts =[]
        if created_count :
            msg_parts .append (f"{created_count} parte(s) generado(s) correctamente.")
        if skipped_count :
            msg_parts .append (
            f"{skipped_count} día(s) omitido(s) por existir ya un parte para esa fecha."
            )
        django_messages .success (request ," ".join (msg_parts ))
        return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")


class WorkPeriodListView (SupervisorAccessMixin ,View ):
    """
    Lists all WorkPeriod records for the authenticated user's company,
    grouped by operator and ordered descending by start_date.
    Renders the full work_period_list.html page on GET.

    GET /panel/work-periods/
    ---
    Lista todos los registros WorkPeriod de la empresa del usuario autenticado,
    agrupados por operario y ordenados descendentemente por start_date.
    Renderiza la página completa work_period_list.html en GET.

    GET /panel/work-periods/
    """

    template_name ="panel/work_orders/work_period_list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Builds the context with all WorkPeriod records grouped by operator
        and renders the work_period_list template.
        ---
        Construye el contexto con todos los registros WorkPeriod agrupados
        por operario y renderiza el template work_period_list.
        """
        from ivr_config .models import WorkPeriod 

        cu =request .user .company_user 
        company =cu .company 

        operators =(
        CompanyUser .objects 
        .filter (company =company ,is_active =True ,role =CompanyUser .ROLE_WORKSHOP )
        .select_related ("user")
        .order_by ("user__last_name","user__first_name")
        )

        operator_groups =[]
        for operator in operators :
            periods =(
            WorkPeriod .objects 
            .filter (company_user =operator )
            .select_related ("created_by__user")
            .order_by ("-start_date")
            )
            operator_groups .append ({
            "operator":operator ,
            "periods":list (periods ),
            "has_open":any (p .end_date is None for p in periods ),
            })










        from datetime import date as _dt_date ,timedelta as _td 
        last_closed =(
        WorkPeriod .objects 
        .filter (
        company_user__company =company ,
        end_date__isnull =False ,
        )
        .order_by ("-end_date")
        .first ()
        )

        if last_closed and last_closed .end_date :
            _duration_days =(last_closed .end_date -last_closed .start_date ).days +1 
            _suggested_start =last_closed .end_date +_td (days =1 )
            _suggested_end =_suggested_start +_td (days =_duration_days -1 )
        else :


            _today =_dt_date .today ()
            if _today .day >=21 :
                _suggested_start =_today .replace (day =21 )
                _first_of_next =(_today .replace (day =1 )+_td (days =32 )).replace (day =1 )
                _suggested_end =_first_of_next .replace (day =20 )
            else :
                _first_of_this =_today .replace (day =1 )
                _prev_month_end =_first_of_this -_td (days =1 )
                _suggested_start =_prev_month_end .replace (day =21 )
                _suggested_end =_today .replace (day =20 )



        open_periods_exist =WorkPeriod .objects .filter (
        company_user__company =company ,
        end_date__isnull =True ,
        ).exists ()

        context ={
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"work_period_list",
        "operator_groups":operator_groups ,
        "operators":operators ,
        "suggested_start":_suggested_start .strftime ("%Y-%m-%d"),
        "suggested_end":_suggested_end .strftime ("%Y-%m-%d"),
        "has_open_periods":open_periods_exist ,
        }
        return render (request ,self .template_name ,context )


class WorkPeriodCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a new global WorkPeriod for ALL active WORKSHOP operators
    belonging to the authenticated supervisor's company.
    The work period is company-wide: the same start_date and label are
    applied to every active WORKSHOP operator in a single operation.
    Operators that already have an open period are skipped individually
    and reported back to the user via a warning message.
    On success or failure, redirects to work_period_list.

    POST /panel/work-periods/create/
    ---
    Crea un nuevo WorkPeriod global para TODOS los operarios WORKSHOP
    activos de la empresa del supervisor autenticado.
    El periodo de trabajo es de ámbito empresarial: la misma start_date
    y etiqueta se aplican a todos los operarios WORKSHOP activos.
    Los operarios que ya tienen un periodo abierto se omiten individualmente
    y se notifican al usuario mediante un mensaje de aviso.
    En éxito o fallo, redirige a work_period_list.

    POST /panel/work-periods/create/
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates POST data, creates a WorkPeriod for every active WORKSHOP
        operator in the company and redirects to work_period_list.
        Operators with an already-open period are skipped individually and
        reported back via a warning django message.
        ---
        Valida los datos POST, crea un WorkPeriod para cada operario WORKSHOP
        activo de la empresa y redirige a work_period_list.
        Los operarios con un periodo ya abierto se omiten individualmente y
        se notifican al usuario mediante un mensaje de aviso de Django.
        """
        from datetime import datetime 
        from django .urls import reverse 
        from ivr_config .models import WorkPeriod 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:work_period_list")



        raw_start =request .POST .get ("start_date","").strip ()
        try :
            start_date =datetime .strptime (raw_start ,"%Y-%m-%d").date ()
        except (ValueError ,AttributeError ):
            django_messages .error (
            request ,
            "La fecha de inicio es obligatoria y debe tener formato YYYY-MM-DD.",
            )
            return redirect (LIST_URL )

        label =request .POST .get ("label","").strip ()



        raw_end =request .POST .get ("end_date","").strip ()
        end_date_parsed =None 
        if raw_end :
            try :
                end_date_parsed =datetime .strptime (raw_end ,"%Y-%m-%d").date ()
                if end_date_parsed <start_date :
                    django_messages .error (
                    request ,
                    "La fecha de fin no puede ser anterior a la fecha de inicio.",
                    )
                    return redirect (LIST_URL )
            except (ValueError ,AttributeError ):
                end_date_parsed =None 



        workshop_operators =list (
        CompanyUser .objects 
        .filter (company =company ,is_active =True ,role =CompanyUser .ROLE_WORKSHOP )
        .select_related ("user")
        .order_by ("user__last_name","user__first_name")
        )

        if not workshop_operators :
            django_messages .error (
            request ,
            "No hay operarios de taller activos en la empresa. No se ha creado ningún periodo.",
            )
            return redirect (LIST_URL )



        created_count =0 
        skipped_names =[]

        for operator in workshop_operators :
            if WorkPeriod .objects .filter (
            company_user =operator ,end_date__isnull =True 
            ).exists ():
                skipped_names .append (
                operator .user .get_full_name ()or operator .user .username 
                )
                continue 

            WorkPeriod .objects .create (
            company_user =operator ,
            start_date =start_date ,
            end_date =end_date_parsed ,
            label =label ,
            created_by =cu ,
            )
            created_count +=1 



        if created_count >0 :
            django_messages .success (
            request ,
            f"Periodo de trabajo creado para {created_count} operario"
            f"{'s' if created_count != 1 else ''} "
            f"(inicio: {start_date:%d/%m/%Y}).",
            )
        else :
            django_messages .warning (
            request ,
            "No se ha creado ningún periodo: todos los operarios tienen ya un periodo abierto.",
            )

        if skipped_names :
            skipped_list =", ".join (skipped_names )
            django_messages .warning (
            request ,
            f"Operarios omitidos por tener periodo abierto: {skipped_list}.",
            )

        return redirect (LIST_URL )


class WorkPeriodCloseView (SupervisorAccessMixin ,View ):
    """
    Globally closes ALL open WorkPeriod records for the authenticated user's
    company in a single operation. No pk is received — the periods to close
    are derived from company scope.

    On successful close:
      - All open WorkPeriod records for the company are closed in bulk
        (end_date set to the submitted value).
      - All DIGITAL and GENERATED WorkOrder records whose entries fall within
        the closed period are marked reviewed=True in bulk.
      - One generate_period_excel Celery task is enqueued per reviewed WorkOrder.

    On success or failure, redirects to work_period_list.

    POST /panel/work-periods/close/
    ---
    Cierra GLOBALMENTE todos los WorkPeriod abiertos de la empresa del usuario
    autenticado en una única operación. No recibe pk — los periodos a cerrar se
    derivan del scope de la empresa.

    Al cerrar correctamente:
      - Todos los WorkPeriod abiertos de la empresa se cierran en bloque
        (end_date asignado al valor enviado).
      - Todos los WorkOrder DIGITAL y GENERATED cuyas entradas caen dentro del
        periodo cerrado se marcan reviewed=True en bloque.
      - Se encola una tarea Celery generate_period_excel por cada WorkOrder revisado.

    En éxito o fallo, redirige a work_period_list.

    POST /panel/work-periods/close/
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates end_date, closes all open WorkPeriod records for the company
        in bulk, marks the period's digital/generated WorkOrders as reviewed and
        enqueues Excel generation per WorkOrder. Redirects to work_period_list.

        Steps:
          1. Verify at least one open WorkPeriod exists for the company.
          2. Parse end_date from POST (YYYY-MM-DD format required).
          3. Derive period_start as the minimum start_date across all open periods.
          4. Validate end_date >= period_start.
          5. Capture pks of WorkOrders to review BEFORE the bulk update.
          6. Mark those WorkOrders reviewed=True in bulk.
          7. Close all open WorkPeriod records in bulk (end_date = end_date).
          8. Enqueue generate_period_excel for each reviewed WorkOrder pk.
          9. Redirect with descriptive success message.

        ---

        Valida end_date, cierra en bloque todos los WorkPeriod abiertos de la
        empresa, marca como revisados los WorkOrder digitales/generados del periodo
        y encola la generación de Excel por WorkOrder. Redirige a work_period_list.

        Pasos:
          1. Verificar que existe al menos un WorkPeriod abierto para la empresa.
          2. Parsear end_date del POST (formato YYYY-MM-DD obligatorio).
          3. Derivar period_start como la start_date mínima de todos los periodos abiertos.
          4. Validar end_date >= period_start.
          5. Capturar los pks de WorkOrder a revisar ANTES del bulk update.
          6. Marcar esos WorkOrder reviewed=True en bloque.
          7. Cerrar todos los WorkPeriod abiertos en bloque (end_date = end_date).
          8. Encolar generate_period_excel por cada pk de WorkOrder revisado.
          9. Redirigir con mensaje de éxito descriptivo.
        """
        from datetime import datetime 
        from django .db .models import Min 
        from django .urls import reverse 
        from django .utils .timezone import now as tz_now 
        from ivr_config .models import WorkPeriod 
        from work_order_processor .tasks import generate_period_excel 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:work_period_list")





        open_periods =WorkPeriod .objects .filter (
        company_user__company =company ,
        end_date__isnull =True ,
        )
        if not open_periods .exists ():
            django_messages .error (
            request ,
            "No hay ningún periodo activo abierto en esta empresa.",
            )
            return redirect (LIST_URL )





        raw_end =request .POST .get ("end_date","").strip ()
        try :
            end_date =datetime .strptime (raw_end ,"%Y-%m-%d").date ()
        except (ValueError ,AttributeError ):
            django_messages .error (
            request ,
            "La fecha de fin es obligatoria y debe tener formato YYYY-MM-DD.",
            )
            return redirect (LIST_URL )





        agg =open_periods .aggregate (min_start =Min ("start_date"))
        period_start =agg ["min_start"]





        if period_start and end_date <period_start :
            django_messages .error (
            request ,
            f"La fecha de fin ({end_date:%d/%m/%Y}) no puede ser anterior "
            f"a la fecha de inicio del periodo ({period_start:%d/%m/%Y}).",
            )
            return redirect (LIST_URL )









        work_orders_qs =WorkOrder .objects .filter (
        company =company ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        reviewed =False ,
        ).filter (
        entries__work_date__gte =period_start ,
        entries__work_date__lte =end_date ,
        ).distinct ()

        pks =list (work_orders_qs .values_list ("pk",flat =True ))
        reviewed_count =len (pks )





        if pks :
            WorkOrder .objects .filter (pk__in =pks ).update (
            reviewed =True ,
            reviewed_at =tz_now (),
            )





        closed_count =open_periods .count ()
        open_periods .update (end_date =end_date )





        for pk_val in pks :
            generate_period_excel .apply_async (
            args =[pk_val ],
            queue ="work_orders",
            )





        django_messages .success (
        request ,
        f"{closed_count} periodo(s) cerrado(s). "
        f"{reviewed_count} parte(s) marcado(s) como revisados. "
        f"{len(pks)} Excel(es) encolado(s).",
        )
        return redirect (LIST_URL )


class WorkOrderAdminExportView (SupervisorAccessMixin ,View ):
    """
    Excel export endpoint scoped exclusively to digital and generated work
    orders (WorkOrder.Source.DIGITAL and WorkOrder.Source.GENERATED).
    Designed for use from WorkOrderAdminHistoryView tab Revisados.

    Supports two export modes via the POST field export_mode:

      single_sheet — One flat sheet with all WorkOrderEntryLine records from
        the selected WorkOrders, grouped by operator name then work date.
        A dark-blue separator row bearing the operator name is inserted before
        each new operator block. Data is read directly from the DB.

      multi_sheet  — One sheet per distinct operator (worker_name derived from
        uploaded_by CompanyUser). Each sheet is built from the individual Excel
        stored in WorkOrder.excel_file (regenerated on-the-fly if missing).
        Sheet title truncated to 31 characters (Excel limit).

    Optional POST filter operator_pk restricts the export to a single operator.

    POST /panel/work-orders/admin-export/
         Body params:
           pks         (list[int]) — WorkOrder primary keys to export.
           export_mode (str)       — single_sheet | multi_sheet.
           operator_pk (int)       — optional operator filter.

    Returns HttpResponse with Content-Disposition attachment (xlsx).
    Returns HTTP 400 on invalid pks or unknown export_mode.
    ---
    Endpoint de exportación Excel exclusivo para partes digitales y generados
    (WorkOrder.Source.DIGITAL y WorkOrder.Source.GENERATED).
    Diseñado para su uso desde la pestaña Revisados de WorkOrderAdminHistoryView.

    Soporta dos modos de exportación via el campo POST export_mode:

      single_sheet — Una hoja plana con todos los WorkOrderEntryLine de los
        WorkOrders seleccionados, agrupados por nombre de operario y fecha.
        Una fila separadora azul oscuro con el nombre del operario se inserta
        antes de cada nuevo bloque de operario. Los datos se leen de la BD.

      multi_sheet  — Una hoja por operario distinto (worker_name derivado del
        CompanyUser uploaded_by). Cada hoja se construye desde el Excel
        individual almacenado en WorkOrder.excel_file (regenerado si falta).
        Título de hoja truncado a 31 caracteres (límite de Excel).

    El filtro POST opcional operator_pk restringe la exportación a un operario.

    POST /panel/work-orders/admin-export/
         Parámetros del cuerpo:
           pks         (list[int]) — claves primarias de WorkOrder a exportar.
           export_mode (str)       — single_sheet | multi_sheet.
           operator_pk (int)       — filtro de operario opcional.

    Devuelve HttpResponse con Content-Disposition attachment (xlsx).
    Devuelve HTTP 400 ante pks inválidos o export_mode desconocido.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Builds the admin export Excel file from the selected WorkOrder pks,
        restricted to DIGITAL and GENERATED sources only.
        Supports single_sheet and multi_sheet export modes.
        ---
        Construye el Excel de exportación admin desde los pks de WorkOrder
        seleccionados, restringido a orígenes DIGITAL y GENERATED únicamente.
        Soporta los modos de exportación single_sheet y multi_sheet.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        generate_work_order_excel as _gen_excel ,
        )

        cu =request .user .company_user 
        company =cu .company 





        export_mode =request .POST .get ("export_mode","single_sheet").strip ()
        if export_mode not in ("single_sheet","multi_sheet","digital_full"):
            return HttpResponseBadRequest (
            f"# [ADMIN EXPORT] Modo de exportación desconocido: {export_mode!r}."
            )







        if export_mode =="digital_full":
            return self ._build_digital_full_excel (request ,company ,pk_list ,operator_filter )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (pk )for pk in raw_pks if pk ]
        except (ValueError ,TypeError ):
            return HttpResponseBadRequest ("# [ADMIN EXPORT] Parámetros pks inválidos.")

        if not pk_list :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] No se han seleccionado partes para exportar."
            )





        operator_pk_raw =request .POST .get ("operator_pk","").strip ()
        operator_filter =None 
        if operator_pk_raw :
            try :
                operator_filter =int (operator_pk_raw )
            except (ValueError ,TypeError ):
                operator_filter =None 





        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).order_by ("pk")

        if operator_filter :
            qs =qs .filter (uploaded_by__pk =operator_filter )

        work_orders =list (qs )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] Ninguno de los partes seleccionados es digital/generado, "
            "está revisado y en estado DONE."
            )





        def _get_operator_name (wo ):
            """
            Returns the full name of the operator who submitted the work order,
            or a fallback label if uploaded_by is not set.
            ---
            Devuelve el nombre completo del operario que envió el parte,
            o una etiqueta de reserva si uploaded_by no está establecido.
            """
            if wo .uploaded_by :
                return (
                wo .uploaded_by .user .get_full_name ()
                or wo .uploaded_by .user .username 
                )
            return f"Operario #{wo.pk}"





        def _copy_sheet (src_sheet ,dest_wb ,title ):
            """
            Copies src_sheet (cells, styles, column widths, row heights) into
            a new sheet named title in dest_wb.
            ---
            Copia src_sheet (celdas, estilos, anchos de columna, alturas de
            fila) en una nueva hoja llamada title en dest_wb.
            """
            dest_sheet =dest_wb .create_sheet (title =title [:31 ])
            for row in src_sheet .iter_rows ():
                for cell in row :
                    dest_cell =dest_sheet .cell (
                    row =cell .row ,column =cell .column ,value =cell .value 
                    )
                    if cell .has_style :
                        dest_cell .font =cell .font .copy ()
                        dest_cell .fill =cell .fill .copy ()
                        dest_cell .alignment =cell .alignment .copy ()
                        dest_cell .border =cell .border .copy ()
                        dest_cell .number_format =cell .number_format 
            for col_letter ,col_dim in src_sheet .column_dimensions .items ():
                dest_sheet .column_dimensions [col_letter ].width =col_dim .width 
            for row_num ,row_dim in src_sheet .row_dimensions .items ():
                dest_sheet .row_dimensions [row_num ].height =row_dim .height 
            return dest_sheet 












        if export_mode =="single_sheet":

            wo_map ={wo .pk :wo for wo in work_orders }

            rows =(
            WorkOrderEntryLine .objects 
            .filter (entry__work_order__pk__in =list (wo_map .keys ()))
            .select_related (
            "entry__work_order",
            "entry",
            "machine_asset",
            )
            .order_by (
            "entry__work_order__uploaded_by__user__last_name",
            "entry__work_order__uploaded_by__user__first_name",
            "entry__work_date",
            "entry__work_order__pk",
            "line_number",
            )
            )

            wb =openpyxl .Workbook ()
            ws =wb .active 
            ws .title ="Partes digitales"


            header_fill =PatternFill ("solid",fgColor ="1F3864")
            header_font =Font (bold =True ,color ="FFFFFF",size =10 )
            headers =[
            "Operario","Fecha","Máquina / CdG",
            "Descripción avería","Notas reparación",
            "H. inicio","H. fin","Δ Horas",
            ]
            for col_idx ,h in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =h )
                cell .fill =header_fill 
                cell .font =header_font 
                cell .alignment =Alignment (horizontal ="center",vertical ="center")


            sep_fill =PatternFill ("solid",fgColor ="2F5496")
            sep_font =Font (bold =True ,color ="FFFFFF",size =10 )

            current_row =2 
            current_operator =None 

            for line in rows :
                wo =line .entry .work_order 
                operator_name =_get_operator_name (wo )
                work_date =line .entry .work_date 



                if operator_name !=current_operator :
                    current_operator =operator_name 
                    sep_cell =ws .cell (
                    row =current_row ,column =1 ,value =operator_name 
                    )
                    sep_cell .fill =sep_fill 
                    sep_cell .font =sep_font 
                    sep_cell .alignment =Alignment (vertical ="center")
                    ws .merge_cells (
                    start_row =current_row ,start_column =1 ,
                    end_row =current_row ,end_column =len (headers ),
                    )
                    current_row +=1 

                ws .cell (row =current_row ,column =1 ,value =operator_name )
                ws .cell (row =current_row ,column =2 ,
                value =work_date .strftime ("%d/%m/%Y")if work_date else "")
                ws .cell (row =current_row ,column =3 ,
                value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
                ws .cell (row =current_row ,column =4 ,value =line .fault_description or "")
                ws .cell (row =current_row ,column =5 ,value =line .repair_notes or "")
                ws .cell (row =current_row ,column =6 ,
                value =line .hc .strftime ("%H:%M")if line .hc else "")
                ws .cell (row =current_row ,column =7 ,
                value =line .hf .strftime ("%H:%M")if line .hf else "")
                ws .cell (row =current_row ,column =8 ,
                value =float (line .delta_hours )if line .delta_hours is not None else "")
                current_row +=1 


            for col in ws .columns :
                max_len =max (
                (len (str (cell .value ))for cell in col if cell .value ),
                default =10 ,
                )
                ws .column_dimensions [col [0 ].column_letter ].width =min (max_len +4 ,60 )

            buf =io .BytesIO ()
            wb .save (buf )
            buf .seek (0 )

            filename =f"partes_digitales_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response =HttpResponse (
            buf .read (),
            content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response ["Content-Disposition"]=f'attachment; filename="{filename}"'
            return response 












        dest_wb =openpyxl .Workbook ()
        dest_wb .remove (dest_wb .active )



        from collections import defaultdict 
        operator_groups =defaultdict (list )
        for wo in work_orders :
            operator_groups [_get_operator_name (wo )].append (wo )

        for operator_name ,wo_list in sorted (operator_groups .items ()):
            sheet_title =operator_name [:31 ]
            sheet_added =False 

            for wo in wo_list :


                if not wo .excel_file or not wo .excel_file .name :
                    try :
                        _gen_excel (wo .pk )
                        wo .refresh_from_db (fields =["excel_file"])
                    except Exception :
                        logger .warning (
                        "# [AdminExport] No se pudo regenerar Excel para WorkOrder #%d.",
                        wo .pk ,
                        )
                        continue 

                try :
                    src_wb =openpyxl .load_workbook (wo .excel_file .path )
                    src_sheet =src_wb .worksheets [0 ]
                    if not sheet_added :
                        _copy_sheet (src_sheet ,dest_wb ,sheet_title )
                        sheet_added =True 
                    else :


                        dest_sheet =dest_wb [sheet_title ]
                        start_row =dest_sheet .max_row +1 
                        for row in src_sheet .iter_rows (min_row =2 ):
                            for cell in row :
                                dest_cell =dest_sheet .cell (
                                row =start_row +cell .row -2 ,
                                column =cell .column ,
                                value =cell .value ,
                                )
                                if cell .has_style :
                                    dest_cell .font =cell .font .copy ()
                                    dest_cell .fill =cell .fill .copy ()
                                    dest_cell .alignment =cell .alignment .copy ()
                except Exception as exc :
                    logger .warning (
                    "# [AdminExport] Error procesando Excel WorkOrder #%d: %s",
                    wo .pk ,exc ,
                    )
                    continue 

        if not dest_wb .worksheets :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] No se pudo generar ninguna hoja Excel. "
            "Verifica que los partes seleccionados tienen Excel generado."
            )

        buf =io .BytesIO ()
        dest_wb .save (buf )
        buf .seek (0 )

        filename =f"partes_digitales_multi_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response =HttpResponse (
        buf .read (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


    def _build_digital_full_excel (self ,request ,company ,pk_list ,operator_filter ):
        """
        Builds a three-sheet Excel file for digital and generated work orders.

        Sheet 1 — Tareas: one row per WorkOrderEntryLine with all digital fields.
        Sheet 2 — Repuestos: one row per SparePartLine with supplier and price data.
        Sheet 3 — Incidencias de jornada: one row per WorkdayGap with resolution detail.

        Returns HttpResponse with Content-Disposition attachment (xlsx).
        Returns HTTP 400 if no valid digital/generated work orders are found.
        ---
        Construye un Excel de tres hojas para partes digitales y generados.

        Hoja 1 — Tareas: una fila por WorkOrderEntryLine con todos los campos digitales.
        Hoja 2 — Repuestos: una fila por SparePartLine con datos de proveedor y precio.
        Hoja 3 — Incidencias de jornada: una fila por WorkdayGap con detalle de resolucion.

        Devuelve HttpResponse con Content-Disposition attachment (xlsx).
        Devuelve HTTP 400 si no se encuentran partes digitales/generados validos.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import (
        WorkOrderEntry ,WorkOrderEntryLine ,SparePartLine ,WorkdayGap ,
        )



        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).order_by ("pk")

        if operator_filter :
            qs =qs .filter (uploaded_by__pk =operator_filter )

        work_orders =list (qs )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [DIGITAL FULL EXPORT] Ninguno de los partes seleccionados es "
            "digital/generado, esta revisado y en estado DONE."
            )

        wo_pks =[wo .pk for wo in work_orders ]



        def _op_name (wo ):
            if wo .uploaded_by :
                return (
                wo .uploaded_by .user .get_full_name ()
                or wo .uploaded_by .user .username 
                )
            return f"Operario #{wo.pk}"



        def _fmt_time (t ):
            return t .strftime ("%H:%M")if t else ""



        def _fmt_date (d ):
            return d .strftime ("%d/%m/%Y")if d else ""



        wo_map ={}
        entries =(
        WorkOrderEntry .objects 
        .filter (work_order__in =wo_pks )
        .select_related ("work_order__uploaded_by__user")
        )
        for entry in entries :
            wo_map [entry .work_order_id ]=(entry .work_order ,entry )


        header_fill =PatternFill ("solid",fgColor ="1F4E79")
        header_font =Font (bold =True ,color ="FFFFFF",size =10 )

        def _apply_header (ws ,headers ):
            for col_idx ,h in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =h )
                cell .fill =header_fill 
                cell .font =header_font 
                cell .alignment =Alignment (horizontal ="center",vertical ="center")
            ws .row_dimensions [1 ].height =18 

        def _autofit (ws ):
            for col in ws .columns :
                max_len =max (
                (len (str (cell .value ))for cell in col if cell .value ),
                default =10 ,
                )
                ws .column_dimensions [col [0 ].column_letter ].width =min (max_len +4 ,60 )

        wb =openpyxl .Workbook ()




        ws_tasks =wb .active 
        ws_tasks .title ="Tareas"
        task_headers =[
        "Operario","Fecha","Maquina / CdG","O.R.",
        "Descripcion averia","Notas reparacion",
        "H. inicio","H. fin","Delta Horas",
        "Km (odometro)","Horometro motor","Horometro grua",
        "Categoria averia","Subcategoria averia",
        ]
        _apply_header (ws_tasks ,task_headers )

        lines =(
        WorkOrderEntryLine .objects 
        .filter (entry__work_order__in =wo_pks )
        .select_related (
        "entry__work_order__uploaded_by__user",
        "machine_asset",
        )
        .order_by (
        "entry__work_order__uploaded_by__user__last_name",
        "entry__work_date",
        "line_number",
        )
        )

        row =2 
        for line in lines :
            entry =line .entry 
            wo =entry .work_order 
            ws_tasks .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_tasks .cell (row =row ,column =2 ,value =_fmt_date (entry .work_date ))
            ws_tasks .cell (row =row ,column =3 ,
            value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
            ws_tasks .cell (row =row ,column =4 ,value =line .or_val or "")
            ws_tasks .cell (row =row ,column =5 ,value =line .fault_description or "")
            ws_tasks .cell (row =row ,column =6 ,value =line .repair_notes or "")
            ws_tasks .cell (row =row ,column =7 ,value =_fmt_time (line .hc ))
            ws_tasks .cell (row =row ,column =8 ,value =_fmt_time (line .hf ))
            ws_tasks .cell (row =row ,column =9 ,
            value =float (line .delta_hours )if line .delta_hours is not None else "")
            ws_tasks .cell (row =row ,column =10 ,
            value =float (line .odometer_reading )if line .odometer_reading is not None else "")
            ws_tasks .cell (row =row ,column =11 ,
            value =float (line .engine_hours_reading )if line .engine_hours_reading is not None else "")
            ws_tasks .cell (row =row ,column =12 ,
            value =float (line .crane_hours_reading )if line .crane_hours_reading is not None else "")
            ws_tasks .cell (row =row ,column =13 ,
            value =line .get_fault_category_display ()if line .fault_category else "")
            ws_tasks .cell (row =row ,column =14 ,
            value =line .get_fault_subcategory_display ()if line .fault_subcategory else "")
            row +=1 

        _autofit (ws_tasks )




        ws_parts =wb .create_sheet (title ="Repuestos")
        parts_headers =[
        "Operario","Fecha","Maquina / CdG",
        "Referencia","Material","Cantidad",
        "Procedencia","Proveedor","Precio unitario",
        ]
        _apply_header (ws_parts ,parts_headers )

        spare_lines =(
        SparePartLine .objects 
        .filter (entry_line__entry__work_order__in =wo_pks )
        .select_related (
        "entry_line__entry__work_order__uploaded_by__user",
        "entry_line__machine_asset",
        "vehicle",
        )
        .order_by (
        "entry_line__entry__work_order__uploaded_by__user__last_name",
        "entry_line__entry__work_date",
        "entry_line__line_number",
        "line_number",
        )
        )

        row =2 
        for spare in spare_lines :
            line =spare .entry_line 
            entry =line .entry 
            wo =entry .work_order 
            ws_parts .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_parts .cell (row =row ,column =2 ,value =_fmt_date (entry .work_date ))
            ws_parts .cell (row =row ,column =3 ,
            value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
            ws_parts .cell (row =row ,column =4 ,value =spare .reference or "")
            ws_parts .cell (row =row ,column =5 ,value =spare .material or "")
            ws_parts .cell (row =row ,column =6 ,
            value =float (spare .quantity )if spare .quantity is not None else "")
            ws_parts .cell (row =row ,column =7 ,value =spare .get_source_display ())
            ws_parts .cell (row =row ,column =8 ,value =spare .supplier or "")
            ws_parts .cell (row =row ,column =9 ,
            value =float (spare .unit_price )if spare .unit_price is not None else "")
            row +=1 

        _autofit (ws_parts )




        ws_gaps =wb .create_sheet (title ="Incidencias de jornada")
        gaps_headers =[
        "Operario","Fecha","Tipo","Inicio","Fin",
        "Duracion (min)","Categoria ausencia",
        "Ha comido","Hora comida","Nota","Resuelto",
        ]
        _apply_header (ws_gaps ,gaps_headers )

        gaps =(
        WorkdayGap .objects 
        .filter (work_order__in =wo_pks )
        .select_related (
        "work_order__uploaded_by__user",
        "work_order__entries",
        "absence_category",
        )
        .prefetch_related ("work_order__entries")
        .order_by (
        "work_order__uploaded_by__user__last_name",
        "gap_start",
        )
        )

        row =2 
        for gap in gaps :
            wo =gap .work_order 
            entry =wo .entries .first ()
            ws_gaps .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_gaps .cell (row =row ,column =2 ,
            value =_fmt_date (entry .work_date )if entry else "")
            ws_gaps .cell (row =row ,column =3 ,value =gap .get_gap_type_display ())
            ws_gaps .cell (row =row ,column =4 ,value =_fmt_time (gap .gap_start ))
            ws_gaps .cell (row =row ,column =5 ,value =_fmt_time (gap .gap_end ))
            ws_gaps .cell (row =row ,column =6 ,value =gap .duration_minutes )
            ws_gaps .cell (row =row ,column =7 ,
            value =gap .absence_category .label if gap .absence_category else "")


            if gap .gap_type ==WorkdayGap .GapType .LUNCH_BREAK :
                ws_gaps .cell (row =row ,column =8 ,
                value ="Si"if gap .lunch_had else ("No"if gap .lunch_had is False else ""))
                ws_gaps .cell (row =row ,column =9 ,value =_fmt_time (gap .lunch_time ))
            else :
                ws_gaps .cell (row =row ,column =8 ,value ="")
                ws_gaps .cell (row =row ,column =9 ,value ="")
            ws_gaps .cell (row =row ,column =10 ,value =gap .note or "")
            ws_gaps .cell (row =row ,column =11 ,value ="Si"if gap .resolved else "No")
            row +=1 

        _autofit (ws_gaps )

        buf =io .BytesIO ()
        wb .save (buf )
        buf .seek (0 )

        filename =f"partes_digitales_completo_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response =HttpResponse (
        buf .read (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


class AnalyticsProfileDeleteView (SupervisorAccessMixin ,View ):
    """
    JSON endpoint for deleting a single AnalyticsProfile by primary key.

    DELETE /panel/analytics/profiles/<pk>/
           Deletes the profile only if it belongs to the authenticated
           CompanyUser. Returns HTTP 200 on success, HTTP 404 if not found
           or owned by a different user.
    ---
    Endpoint JSON para eliminar un AnalyticsProfile concreto por clave primaria.

    DELETE /panel/analytics/profiles/<pk>/
           Elimina el perfil solo si pertenece al CompanyUser autenticado.
           Devuelve HTTP 200 en caso de éxito, HTTP 404 si no se encuentra
           o pertenece a otro usuario.
    """

    def delete (self ,request ,pk ):
        """
        Deletes the AnalyticsProfile identified by pk, scoped to the current user.
        ---
        Elimina el AnalyticsProfile identificado por pk, acotado al usuario actual.
        """
        from django .http import JsonResponse 

        try :
            company_user =request .user .company_user 
        except AttributeError :
            return JsonResponse ({"error":"Sin perfil de empresa asociado."},status =403 )

        try :
            profile =AnalyticsProfile .objects .get (pk =pk ,company_user =company_user )
        except AnalyticsProfile .DoesNotExist :
            return JsonResponse ({"error":"Perfil no encontrado."},status =404 )

        nombre =profile .nombre 
        profile .delete ()
        return JsonResponse ({"deleted":True ,"nombre":nombre })






class MachineAssetListView (SupervisorAccessMixin ,View ):
    """
    List view for MachineAsset records belonging to the authenticated user's
    company. Supports filtering by family and is_active status.
    Renders the full list page on GET. HTMX partial refresh is triggered by
    the filter controls in the template.

    GET /panel/fleet/

    ---

    Vista de listado de registros MachineAsset pertenecientes a la empresa
    del usuario autenticado. Soporta filtrado por family e is_active.
    Renderiza la página completa en GET. El refresco parcial HTMX se activa
    desde los controles de filtro del template.

    GET /panel/fleet/
    """

    template_name ="panel/fleet/list.html"
    template_name_partial ="panel/fleet/_table_fragment.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _build_queryset (self ,company ,request ):
        """
        Builds the filtered MachineAsset queryset from GET parameters.
        Supported filters: family (str), is_active ('1'=active, '0'=inactive, ''=all),
        search (str, case-insensitive icontains on code, brand_model and plate).
        ---
        Construye el queryset filtrado de MachineAsset desde los parámetros GET.
        Filtros soportados: family (str), is_active ('1'=activo, '0'=inactivo, ''=todos),
        search (str, búsqueda icontains sobre code, brand_model y plate).
        """
        from fleet .models import MachineAsset 

        qs =MachineAsset .objects .filter (company =company ).order_by (
        "company_code","family","code"
        )

        family_filter =request .GET .get ("family","").strip ()
        if family_filter :
            qs =qs .filter (family__iexact =family_filter )

        active_filter =request .GET .get ("is_active","")
        if active_filter =="1":
            qs =qs .filter (is_active =True )
        elif active_filter =="0":
            qs =qs .filter (is_active =False )



        search_filter =request .GET .get ("search","").strip ()
        if search_filter :
            qs =qs .filter (
            Q (code__icontains =search_filter )
            |Q (brand_model__icontains =search_filter )
            |Q (plate__icontains =search_filter )
            )

        return qs 

    def _get_families (self ,company ):
        """
        Returns a sorted list of distinct family values for the company's assets.
        Used to populate the family filter dropdown.
        ---
        Retorna una lista ordenada de valores de family distintos para los activos
        de la empresa. Se usa para poblar el desplegable de filtro de familia.
        """
        from fleet .models import MachineAsset 

        return (
        MachineAsset .objects 
        .filter (company =company )
        .exclude (family ="")
        .values_list ("family",flat =True )
        .distinct ()
        .order_by ("family")
        )

    def _build_fragment_context (self ,company ,company_user ,request ):
        """
        Builds the shared context dict used by the table fragment.
        Applies the active filters, annotates each asset with use_count
        (number of associated WorkOrderEntryLine records) and preserves
        the current filter state so HTMX mutations do not reset the UI.
        ---
        Construye el dict de contexto compartido para el fragmento de tabla.
        Aplica los filtros activos, anota cada activo con use_count
        (número de WorkOrderEntryLine asociadas) y preserva el estado de
        filtros para que las mutaciones HTMX no reinicien la UI.
        """
        from django .db .models import Count 

        qs =self ._build_queryset (company ,request ).annotate (
        use_count =Count ("work_order_lines",distinct =True )
        )
        families =self ._get_families (company )

        return {
        "assets":qs ,
        "families":families ,
        "company":company ,
        "company_user":company_user ,
        "filter_family":request .GET .get ("family",""),
        "filter_is_active":request .GET .get ("is_active",""),
        "filter_search":request .GET .get ("search",""),
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the fleet list page or a partial HTMX table fragment.
        Detects HTMX requests via the HX-Request header and returns only
        the table fragment for partial page updates.
        ---
        Renderiza la página de listado de flota o un fragmento parcial HTMX.
        Detecta peticiones HTMX via la cabecera HX-Request y devuelve solo
        el fragmento de tabla para actualizaciones parciales.
        """
        from panel .forms import MachineAssetForm 

        company_user =request .user .company_user 
        company =company_user .company 
        fragment_ctx =self ._build_fragment_context (company ,company_user ,request )
        form =MachineAssetForm ()

        ctx ={
        **fragment_ctx ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"fleet",
        "form":form ,
        }

        if request .headers .get ("HX-Request"):
            return render (request ,self .template_name_partial ,ctx )
        return render (request ,self .template_name ,ctx )


class MachineAssetCreateView (AdminRoleRequiredMixin ,View ):
    """
    Creates a new MachineAsset for the authenticated user's company.
    Accepts HTMX POST requests and returns an updated table fragment on success,
    or a form fragment with validation errors on failure.

    POST /panel/fleet/create/

    ---

    Crea un nuevo MachineAsset para la empresa del usuario autenticado.
    Acepta peticiones POST HTMX y devuelve un fragmento de tabla actualizado
    en caso de éxito, o un fragmento de formulario con errores en caso de fallo.

    POST /panel/fleet/create/
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the form and creates the MachineAsset. On success returns
        the updated table fragment. On failure returns the form with errors.
        ---
        Valida el formulario y crea el MachineAsset. En caso de éxito devuelve
        el fragmento de tabla actualizado. En caso de fallo devuelve el formulario
        con errores.
        """
        from fleet .models import MachineAsset 
        from panel .forms import MachineAssetForm 

        company_user =request .user .company_user 
        company =company_user .company 
        form =MachineAssetForm (request .POST )

        list_view =MachineAssetListView ()

        if form .is_valid ():
            asset =form .save (commit =False )
            asset .company =company 
            asset .code =asset .code .strip ().upper ()
            asset .save ()
            fragment_ctx =list_view ._build_fragment_context (company ,company_user ,request )
            return render (request ,"panel/fleet/_table_fragment.html",fragment_ctx )

        return render (request ,"panel/fleet/_form_fragment.html",{
        "form":form ,
        "company_user":company_user ,
        "company":company ,
        "form_action":"create",
        })


class MachineAssetUpdateView (AdminRoleRequiredMixin ,View ):
    """
    Updates an existing MachineAsset belonging to the authenticated user's company.
    Accepts HTMX POST requests and returns an updated table fragment on success,
    or a form fragment with validation errors on failure.

    POST /panel/fleet/<pk>/update/

    ---

    Actualiza un MachineAsset existente perteneciente a la empresa del usuario
    autenticado. Acepta peticiones POST HTMX y devuelve un fragmento de tabla
    actualizado en caso de éxito, o un formulario con errores en caso de fallo.

    POST /panel/fleet/<pk>/update/
    """

    def get (self ,request ,pk ,*args ,**kwargs ):
        """
        Returns the pre-filled edit form fragment for the given asset pk.
        Called via HTMX GET from the edit modal trigger in the table.
        Returns 404 if the asset does not belong to the company.
        ---
        Devuelve el fragmento de formulario de edición pre-relleno para el pk dado.
        Invocado via HTMX GET desde el disparador del modal de edición en la tabla.
        Devuelve 404 si el activo no pertenece a la empresa.
        """
        from fleet .models import MachineAsset 
        from panel .forms import MachineAssetForm 
        from django .http import Http404 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            asset =MachineAsset .objects .get (pk =pk ,company =company )
        except MachineAsset .DoesNotExist :
            raise Http404 

        form =MachineAssetForm (instance =asset )
        return render (request ,"panel/fleet/_form_fragment.html",{
        "form":form ,
        "asset":asset ,
        "company_user":company_user ,
        "company":company ,
        "form_action":"update",
        })

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates the form and updates the MachineAsset identified by pk.
        Returns 404 if the asset does not belong to the company.
        ---
        Valida el formulario y actualiza el MachineAsset identificado por pk.
        Devuelve 404 si el activo no pertenece a la empresa.
        """
        from fleet .models import MachineAsset 
        from panel .forms import MachineAssetForm 
        from django .http import Http404 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            asset =MachineAsset .objects .get (pk =pk ,company =company )
        except MachineAsset .DoesNotExist :
            raise Http404 

        form =MachineAssetForm (request .POST ,instance =asset )

        list_view =MachineAssetListView ()

        if form .is_valid ():
            updated =form .save (commit =False )
            updated .code =updated .code .strip ().upper ()
            updated .save ()
            fragment_ctx =list_view ._build_fragment_context (company ,company_user ,request )
            return render (request ,"panel/fleet/_table_fragment.html",fragment_ctx )

        return render (request ,"panel/fleet/_form_fragment.html",{
        "form":form ,
        "asset":asset ,
        "company_user":company_user ,
        "company":company ,
        "form_action":"update",
        })


class MachineAssetDeactivateView (AdminRoleRequiredMixin ,View ):
    """
    Sets is_active=False on a MachineAsset belonging to the authenticated user's
    company. Does not delete the record — preserves historical data integrity.
    Returns an updated table row fragment via HTMX.

    POST /panel/fleet/<pk>/deactivate/

    ---

    Establece is_active=False en un MachineAsset perteneciente a la empresa del
    usuario autenticado. No elimina el registro — preserva la integridad del
    histórico. Devuelve un fragmento de fila de tabla actualizado via HTMX.

    POST /panel/fleet/<pk>/deactivate/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Marks the asset as inactive and returns the updated table fragment.
        Returns 404 if the asset does not belong to the company.
        ---
        Marca el activo como inactivo y devuelve el fragmento de tabla actualizado.
        Devuelve 404 si el activo no pertenece a la empresa.
        """
        from fleet .models import MachineAsset 
        from django .http import Http404 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            asset =MachineAsset .objects .get (pk =pk ,company =company )
        except MachineAsset .DoesNotExist :
            raise Http404 

        asset .is_active =False 
        asset .save (update_fields =["is_active"])

        list_view =MachineAssetListView ()
        fragment_ctx =list_view ._build_fragment_context (company ,company_user ,request )
        return render (request ,"panel/fleet/_table_fragment.html",fragment_ctx )


class MachineAssetReactivateView (AdminRoleRequiredMixin ,View ):
    """
    Sets is_active=True on a MachineAsset belonging to the authenticated user's
    company. Counterpart to MachineAssetDeactivateView — allows reversing an
    accidental deactivation. Returns an updated table fragment via HTMX.

    POST /panel/fleet/<pk>/reactivate/

    ---

    Establece is_active=True en un MachineAsset perteneciente a la empresa del
    usuario autenticado. Contraparte de MachineAssetDeactivateView — permite
    revertir una baja accidental. Devuelve un fragmento de tabla actualizado
    via HTMX.

    POST /panel/fleet/<pk>/reactivate/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Marks the asset as active and returns the updated table fragment.
        Returns 404 if the asset does not belong to the company.
        ---
        Marca el activo como activo y devuelve el fragmento de tabla actualizado.
        Devuelve 404 si el activo no pertenece a la empresa.
        """
        from fleet .models import MachineAsset 
        from django .http import Http404 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            asset =MachineAsset .objects .get (pk =pk ,company =company )
        except MachineAsset .DoesNotExist :
            raise Http404 

        asset .is_active =True 
        asset .save (update_fields =["is_active"])

        list_view =MachineAssetListView ()
        fragment_ctx =list_view ._build_fragment_context (company ,company_user ,request )
        return render (request ,"panel/fleet/_table_fragment.html",fragment_ctx )


class MachineAssetDeleteView (AdminRoleRequiredMixin ,View ):
    """
    Permanently deletes a MachineAsset only if it has no associated
    WorkOrderEntryLine records (referential integrity guard).
    Only available to ADMIN role.
    Returns an updated table fragment on success or a JSON error on failure.

    POST /panel/fleet/<pk>/delete/

    ---

    Elimina permanentemente un MachineAsset solo si no tiene registros
    WorkOrderEntryLine asociados (guardia de integridad referencial).
    Solo disponible para el rol ADMIN.
    Devuelve un fragmento de tabla actualizado en caso de éxito o un error
    JSON en caso de fallo.

    POST /panel/fleet/<pk>/delete/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Deletes the asset if it has no linked work-order lines.
        Returns HTTP 409 with a JSON error message if linked lines exist.
        Returns 404 if the asset does not belong to the company.
        ---
        Elimina el activo si no tiene líneas de parte asociadas.
        Devuelve HTTP 409 con un mensaje de error JSON si existen líneas vinculadas.
        Devuelve 404 si el activo no pertenece a la empresa.
        """
        from fleet .models import MachineAsset 
        from django .http import Http404 ,JsonResponse 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            asset =MachineAsset .objects .get (pk =pk ,company =company )
        except MachineAsset .DoesNotExist :
            raise Http404 



        if asset .work_order_lines .exists ():
            return JsonResponse (
            {
            "error":(
            f"No se puede eliminar '{asset.code}': tiene partes de trabajo "
            f"asociados. Use 'Dar de baja' para desactivarlo."
            )
            },
            status =409 ,
            )

        asset .delete ()

        list_view =MachineAssetListView ()
        fragment_ctx =list_view ._build_fragment_context (company ,company_user ,request )
        return render (request ,"panel/fleet/_table_fragment.html",fragment_ctx )


class WorkshopAssetDetailView (WorkshopRequiredMixin ,View ):
    """
    JSON endpoint that returns the meter-reading flags and current reference
    values for a single MachineAsset identified by its code, scoped to the
    authenticated user's company.

    Used by the three operator entry templates (form_entry, stt_entry,
    confirm_entry) to reveal/hide the odometer and hourmeter input fields
    dynamically when the operator selects a work block's Centre de Gasto.

    GET /panel/operator/assets/detail/?code=XX
        Returns HTTP 200 with JSON payload on success.
        Returns HTTP 404 if the asset does not exist for the company.
        Returns HTTP 400 if the `code` parameter is absent or blank.

    Response schema:
        {
          "has_odometer":     bool,
          "has_engine_hours": bool,
          "has_crane_hours":  bool,
          "mileage":          float | null,
          "hours":            float | null
        }

    ---

    Endpoint JSON que devuelve los flags de lecturas de contadores y los
    valores de referencia actuales para un MachineAsset identificado por su
    código, acotado a la empresa del usuario autenticado.

    Usado por los tres templates de entrada del operario (form_entry,
    stt_entry, confirm_entry) para revelar/ocultar dinámicamente los campos
    de odómetro y horómetro cuando el operario selecciona el Centro de Gasto
    de un bloque de trabajo.

    GET /panel/operator/assets/detail/?code=XX
        Devuelve HTTP 200 con payload JSON en caso de éxito.
        Devuelve HTTP 404 si el activo no existe para la empresa.
        Devuelve HTTP 400 si el parámetro `code` está ausente o en blanco.

    Esquema de respuesta:
        {
          "has_odometer":     bool,
          "has_engine_hours": bool,
          "has_crane_hours":  bool,
          "mileage":          float | null,
          "hours":            float | null
        }
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns meter-reading flags and reference values for the requested asset.
        Scoped to the authenticated user's company. Returns HTTP 400 on missing
        code and HTTP 404 if the asset does not exist for the company.
        ---
        Devuelve los flags de contadores y valores de referencia del activo
        solicitado. Acotado a la empresa del usuario autenticado. Devuelve
        HTTP 400 si falta el código y HTTP 404 si el activo no existe.
        """
        from django .http import JsonResponse 
        from fleet .models import MachineAsset 

        code =request .GET .get ("code","").strip ()
        if not code :
            return JsonResponse (
            {"error":"Parámetro 'code' obligatorio."},
            status =400 ,
            )

        company =request .user .company_user .company 

        try :
            asset =MachineAsset .objects .get (
            code__iexact =code ,
            company =company ,
            )
        except MachineAsset .DoesNotExist :
            return JsonResponse (
            {"error":f"Activo '{code}' no encontrado en catálogo."},
            status =404 ,
            )

        return JsonResponse ({
        "has_odometer":asset .has_odometer ,
        "has_engine_hours":asset .has_engine_hours ,
        "has_crane_hours":asset .has_crane_hours ,
        "first_repair":asset .first_repair ,
        "mileage":float (asset .mileage )if asset .mileage is not None else None ,
        "hours":float (asset .hours )if asset .hours is not None else None ,
        })


class WorkOrderMachineFilterView (SupervisorAccessMixin ,View ):
    """
    JSON endpoint returning distinct MachineAsset codes present in the
    WorkOrderEntryLine records of DIGITAL/GENERATED WorkOrders for the
    authenticated company, optionally filtered by operator and date range.
    Used by the admin history machine autocomplete (Bug B fix).

    GET /panel/work-orders/machines/
        Optional GET params:
          operator_pk (int)  — filter by uploaded_by CompanyUser pk.
          date_from   (str)  — ISO date YYYY-MM-DD start of range.
          date_to     (str)  — ISO date YYYY-MM-DD end of range.
        Returns: {"results": ["G12", "A44", ...]}

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Endpoint JSON que devuelve los códigos de MachineAsset distintos presentes
    en los WorkOrderEntryLine de los WorkOrders DIGITAL/GENERATED de la empresa
    autenticada, con filtro opcional por operario y rango de fechas.
    Usado por el autocompletado de máquina de admin_history (corrección Bug B).

    GET /panel/work-orders/machines/
        Parámetros GET opcionales:
          operator_pk (int)  — filtrar por pk de CompanyUser uploaded_by.
          date_from   (str)  — fecha ISO YYYY-MM-DD inicio del rango.
          date_to     (str)  — fecha ISO YYYY-MM-DD fin del rango.
        Devuelve: {"results": ["G12", "A44", ...]}

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns distinct MachineAsset codes present in the filtered WorkOrders.
        ---
        Devuelve los códigos de MachineAsset distintos en los WorkOrders filtrados.
        """
        from django .http import JsonResponse 
        from datetime import datetime as _dt_mf 
        from work_order_processor .models import WorkOrderEntryLine 

        company =request .user .company_user .company 
        operator_pk =request .GET .get ("operator_pk","").strip ()
        date_from_raw =request .GET .get ("date_from","").strip ()
        date_to_raw =request .GET .get ("date_to","").strip ()
        q_raw =request .GET .get ("q","").strip ()

        def _parse_iso (val ):
            """Parses YYYY-MM-DD string, returns date or None.
            --- Parsea cadena YYYY-MM-DD, devuelve date o None."""
            if not val :
                return None 
            try :
                return _dt_mf .strptime (val ,"%Y-%m-%d").date ()
            except ValueError :
                return None 

        date_from =_parse_iso (date_from_raw )
        date_to =_parse_iso (date_to_raw )



        qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        entry__work_order__source__in =[
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        ],
        machine_asset__isnull =False ,
        )
        )


        if operator_pk :
            try :
                qs =qs .filter (
                entry__work_order__uploaded_by__pk =int (operator_pk ),
                entry__work_order__uploaded_by__company =company ,
                )
            except (ValueError ,TypeError ):
                pass 


        if date_from :
            qs =qs .filter (entry__work_date__gte =date_from )
        if date_to :
            qs =qs .filter (entry__work_date__lte =date_to )


        if q_raw :
            qs =qs .filter (machine_asset__code__icontains =q_raw )

        codes =(
        qs 
        .values_list ("machine_asset__code",flat =True )
        .distinct ()
        .order_by ("machine_asset__code")
        )

        return JsonResponse ({"results":list (codes )})



class WorkOrderDescriptionAutocompleteView (WorkshopRequiredMixin ,View ):
    """
    Returns up to 8 unique values from WorkOrderEntryLine.fault_description
    or WorkOrderEntryLine.repair_notes that contain the query string (case-
    insensitive). Scoped to the authenticated user's company.

    Used by the description typeahead widget (_description_typeahead.html)
    in the three operator entry templates (form_entry, stt_entry,
    confirm_entry).

    GET /panel/operator/descriptions/?field=fault_description&q=XXX
    GET /panel/operator/descriptions/?field=repair_notes&q=XXX

    Response: {"results": ["value1", "value2", ...]}

    ---

    Devuelve hasta 8 valores únicos de WorkOrderEntryLine.fault_description
    o WorkOrderEntryLine.repair_notes que contengan la cadena de búsqueda
    (insensible a mayúsculas). Restringido a la empresa del usuario autenticado.

    Utilizado por el widget de typeahead de descripciones
    (_description_typeahead.html) en los tres templates de entrada del
    operario (form_entry, stt_entry, confirm_entry).

    GET /panel/operator/descriptions/?field=fault_description&q=XXX
    GET /panel/operator/descriptions/?field=repair_notes&q=XXX

    Respuesta: {"results": ["valor1", "valor2", ...]}
    """



    _ALLOWED_FIELDS ={"fault_description","repair_notes"}


    _MIN_QUERY_LEN =2 


    _MAX_RESULTS =8 

    def get (self ,request ,*args ,**kwargs ):
        """
        Validates the `field` and `q` parameters, queries the database and
        returns a JSON list of matching description values.

        Returns {"results": []} for invalid field, missing or too-short query.

        ---

        Valida los parámetros `field` y `q`, consulta la base de datos y
        devuelve una lista JSON de valores de descripción coincidentes.

        Devuelve {"results": []} para campo inválido, consulta ausente o
        demasiado corta.
        """
        from django .http import JsonResponse 
        from work_order_processor .models import WorkOrderEntryLine 

        field =request .GET .get ("field","").strip ()
        q =request .GET .get ("q","").strip ()



        if field not in self ._ALLOWED_FIELDS :
            return JsonResponse ({"results":[]})



        if len (q )<self ._MIN_QUERY_LEN :
            return JsonResponse ({"results":[]})

        company_user =request .user .company_user 
        company =company_user .company 



        lookup ={field +"__icontains":q }

        qs =(
        WorkOrderEntryLine .objects 
        .filter (entry__work_order__company =company ,**lookup )
        .exclude (**{field :""})
        .values_list (field ,flat =True )
        .distinct ()
        .order_by (field )
        [:self ._MAX_RESULTS ]
        )

        return JsonResponse ({"results":list (qs )})



class SectionDefaultRoleView (AdminRoleRequiredMixin ,View ):
    """
    AJAX endpoint that returns the default_role of a Section as JSON.
    Used by the user creation form to pre-fill the role selector when
    a section is chosen from the dropdown.
    ---
    Endpoint AJAX que devuelve el default_role de una Section como JSON.
    Usado por el formulario de creación de usuario para pre-rellenar el
    selector de rol cuando se elige una sección del desplegable.
    """

    def get (self ,request ,pk ,*args ,**kwargs ):
        """
        Returns JSON {"default_role": "<ROLE>"} for the requested section pk,
        restricted to the authenticated user's company.
        Returns HTTP 404 if the section does not exist or belongs to another company.
        ---
        Retorna JSON {"default_role": "<ROLE>"} para el pk de sección solicitado,
        restringido a la empresa del usuario autenticado.
        Retorna HTTP 404 si la sección no existe o pertenece a otra empresa.
        """
        from django .http import JsonResponse 
        try :
            section =Section .objects .get (
            pk =pk ,
            company =request .user .company_user .company ,
            )
        except Section .DoesNotExist :
            return JsonResponse ({"error":"Sección no encontrada."},status =404 )
        return JsonResponse ({"default_role":section .default_role })


class OwnProfileView (CompanyUserRequiredMixin ,View ):
    """
    Allows any authenticated CompanyUser to view and edit their own profile.
    Currently exposes only the alias field (chat IRC nick).
    Alias uniqueness within the company is enforced by OwnProfileForm.

    GET  — renders the profile form pre-populated with the current alias.
    POST — validates and saves the new alias, redirects back to the form.

    URL: GET/POST /panel/profile/
    ---
    Permite a cualquier CompanyUser autenticado ver y editar su propio perfil.
    Actualmente expone únicamente el campo alias (nick de chat IRC).
    La unicidad del alias dentro de la empresa la impone OwnProfileForm.

    GET  — renderiza el formulario de perfil prerellenado con el alias actual.
    POST — valida y guarda el nuevo alias, redirige de vuelta al formulario.

    URL: GET/POST /panel/profile/
    """

    template_name ="panel/profile/own_profile.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the own-profile form pre-populated with the current alias.
        ---
        Renderiza el formulario de perfil propio prerellenado con el alias actual.
        """
        from panel .forms import OwnProfileForm 
        company_user =request .user .company_user 
        form =OwnProfileForm (company_user =company_user )
        return render (request ,self .template_name ,{
        "form":form ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"own_profile",
        })






class BotManagementView (CompanyUserRequiredMixin ,View ):
    """
    Central management dashboard for the WhatsApp bot.
    Provides four functional blocks:
      1. Section onboarding — triggers the onboarding flow for all
         Contacts in a selected Section that have not yet completed it.
      2. Group broadcast — sends a free-form message to one or both
         WhatsApp workshop groups (Mechanical / Elevation) via the
         workshop ChatRoom SECTION rooms via Twilio WhatsApp.
      3. 1-to-1 broadcast — sends a free-form message individually
         to all active Contacts of the company via Twilio WhatsApp.
      4. Breakdown ticket viewer — lists active BreakdownTickets filtered
         by the authenticated user's workshop family. ADMIN and SUPERVISOR
         see all tickets with a family selector. WORKSHOPBOSS sees only
         their own family. WORKSHOP sees the family of the WORKSHOPBOSS
         assigned to their section.
    Access: ADMIN (all blocks), SUPERVISOR (viewer only),
    WORKSHOPBOSS and WORKSHOP (viewer only).
    ---
    Panel central de gestión del bot de WhatsApp.
    Proporciona cuatro bloques funcionales:
      1. Onboarding por sección — dispara el flujo de onboarding para todos
         los Contacts de una Section seleccionada que no lo hayan completado.
      2. Circular a grupos — envía un mensaje libre a uno o ambos grupos
         salas ChatRoom SECTION de taller via Twilio WhatsApp.
      3. Circular 1:1 — envía un mensaje libre individualmente a un conjunto
         de Contacts activos de la empresa via Twilio WhatsApp.
      4. Visor de averías — lista los BreakdownTickets activos filtrados
         por la familia de taller del usuario autenticado. ADMIN y SUPERVISOR
         ven todos los tickets con selector de familia. WORKSHOPBOSS ve solo
         su familia. WORKSHOP ve la familia del WORKSHOPBOSS asignado a su sección.
    Acceso: ADMIN (todos los bloques), SUPERVISOR (solo visor),
    WORKSHOPBOSS y WORKSHOP (solo visor).
    """

    template_name ="panel/bot/dashboard.html"





    ALLOWED_ROLES ={
    CompanyUser .ROLE_ADMIN ,
    CompanyUser .ROLE_SUPERVISOR ,
    CompanyUser .ROLE_WORKSHOPBOSS ,
    CompanyUser .ROLE_WORKSHOP ,
    }

    def _get_workshop_family_for_user (self ,company_user ):
        """
        Resolves the workshop family visible to the authenticated user.
        - ADMIN / SUPERVISOR: None (sees all, frontend applies selector).
        - WORKSHOPBOSS: their own workshop_family.
        - WORKSHOP: workshop_family of the WORKSHOPBOSS assigned to their
          section, resolved via Section → WORKSHOPBOSS chain.
        Returns None if no family can be resolved (tickets hidden).
        ---
        Resuelve la familia de taller visible para el usuario autenticado.
        - ADMIN / SUPERVISOR: None (ve todas, el frontend aplica el selector).
        - WORKSHOPBOSS: su propia workshop_family.
        - WORKSHOP: workshop_family del WORKSHOPBOSS asignado a su sección,
          resuelta via cadena Section → WORKSHOPBOSS.
        Devuelve None si no se puede resolver ninguna familia (tickets ocultos).
        """
        role =company_user .role 
        if role in (CompanyUser .ROLE_ADMIN ,CompanyUser .ROLE_SUPERVISOR ):
            return None 
        if role ==CompanyUser .ROLE_WORKSHOPBOSS :
            return company_user .workshop_family 
        if role ==CompanyUser .ROLE_WORKSHOP :


            from ivr_config .models import SectionContact 
            section_contact =(
            SectionContact .objects 
            .filter (
            contact__company_user =company_user ,
            section__company =company_user .company ,
            )
            .select_related ("section")
            .first ()
            )
            if not section_contact :
                return None 
            boss =(
            CompanyUser .objects 
            .filter (
            company =company_user .company ,
            role =CompanyUser .ROLE_WORKSHOPBOSS ,
            workshop_family__isnull =False ,
            )
            .first ()
            )
            return boss .workshop_family if boss else None 
        return None 

    def _get_breakdown_tickets (self ,company_user ,family_filter =None ):
        """
        Returns a queryset of active BreakdownTickets for the given company,
        filtered by workshop family when family_filter is provided.
        ADMIN and SUPERVISOR may pass an explicit family_filter from the
        GET parameter; other roles always use their resolved family.
        ---
        Devuelve un queryset de BreakdownTickets activos para la empresa dada,
        filtrado por familia de taller cuando se proporciona family_filter.
        ADMIN y SUPERVISOR pueden pasar un family_filter explícito desde el
        parámetro GET; otros roles siempre usan su familia resuelta.
        """
        from chat .models import BreakdownTicket 
        from ivr_config .models import WorkshopFamilyMapping 
        qs =(
        BreakdownTicket .objects 
        .filter (room__company =company_user .company )
        .exclude (status =BreakdownTicket .STATUS_RESOLVED )
        .select_related ("machine","contact","assigned_to")
        .order_by ("-created_at")
        )
        if family_filter :


            mapped_families =list (
            WorkshopFamilyMapping .objects 
            .filter (
            company =company_user .company ,
            workshop_family =family_filter ,
            )
            .values_list ("catalogue_family",flat =True )
            )
            if mapped_families :
                qs =qs .filter (machine__family__in =mapped_families )
            else :
                qs =qs .none ()
        return qs 

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the bot management dashboard with context data for all four
        functional blocks. Applies role-based visibility rules.
        ---
        Renderiza el panel de gestión del bot con datos de contexto para los
        cuatro bloques funcionales. Aplica las reglas de visibilidad por rol.
        """
        company_user =request .user .company_user 
        if company_user .role not in self .ALLOWED_ROLES :
            return redirect ("panel:dashboard")

        is_admin =company_user .role ==CompanyUser .ROLE_ADMIN 
        is_supervisor =company_user .role ==CompanyUser .ROLE_SUPERVISOR 
        can_manage =is_admin 



        resolved_family =self ._get_workshop_family_for_user (company_user )
        family_filter =resolved_family 
        if is_admin or is_supervisor :


            family_filter =request .GET .get ("family",None )or None 

        breakdown_tickets =self ._get_breakdown_tickets (company_user ,family_filter )





        from ivr_config .models import Section 
        sections =(
        Section .objects 
        .filter (company =company_user .company )
        .order_by ("name")
        if can_manage else []
        )


        broadcast_sections =(
        Section .objects 
        .filter (company =company_user .company ,is_broadcast_enabled =True )
        .order_by ("name")
        if can_manage else []
        )



        family_choices =(
        CompanyUser .WORKSHOP_FAMILY_CHOICES 
        if (is_admin or is_supervisor )else []
        )

        return render (request ,self .template_name ,{
        "active_nav":"bot_management",
        "can_manage":can_manage ,
        "is_admin":is_admin ,
        "is_supervisor":is_supervisor ,
        "sections":sections ,
        "broadcast_sections":broadcast_sections ,
        "breakdown_tickets":breakdown_tickets ,
        "family_choices":family_choices ,
        "family_filter":family_filter ,
        "company_user":company_user ,
        })

    def post (self ,request ,*args ,**kwargs ):
        """
        Handles the three bot management actions dispatched from the dashboard
        form via the hidden 'action' field:
          - onboarding: triggers the chat_onboarding quick-reply template for
            all contacts in a selected section that have not yet completed the
            alias onboarding flow.
          - group_broadcast: sends a free-form message 1:1 to all active
            CompanyUser members of one or both workshop ChatRoom SECTION rooms,
            resolved via WorkshopFamilyMapping. Persists a ChatMessage(OUTBOUND)
            in each affected room for panel history.
          - direct_broadcast: sends a free-form message 1:1 to all active
            Contacts of the company, optionally filtered by section.
        Only ADMIN role may reach this handler (can_manage guard).
        ---
        Gestiona las tres acciones de administración del bot despachadas desde
        el formulario del dashboard mediante el campo oculto 'action':
          - onboarding: lanza la plantilla quick-reply chat_onboarding para
            todos los contactos de una sección seleccionada que no hayan
            completado el flujo de alias onboarding.
          - group_broadcast: envía un mensaje libre 1:1 a todos los CompanyUser
            activos de una o ambas salas ChatRoom SECTION de taller, resueltas
            via WorkshopFamilyMapping. Persiste un ChatMessage(OUTBOUND) en
            cada sala afectada para el historial del panel.
          - direct_broadcast: envía un mensaje libre 1:1 a todos los Contacts
            activos de la empresa, opcionalmente filtrados por sección.
        Solo el rol ADMIN puede alcanzar este handler (guardia can_manage).
        """
        from whatsapp .services import WhatsAppChatService 
        from whatsapp .models import WhatsAppTemplate 
        from chat .models import ChatRoom ,ChatMessage 
        from ivr_config .models import Section ,Contact ,WorkshopFamilyMapping 

        company_user =request .user .company_user 
        is_admin =company_user .role ==CompanyUser .ROLE_ADMIN 
        if not is_admin :
            return redirect ("panel:bot_management")

        company =company_user .company 
        action =request .POST .get ("action","")



        bot_number =(
        PhoneNumber .objects 
        .filter (
        company =company ,
        capabilities__in =[
        PhoneNumber .CAPABILITY_WHATSAPP ,
        PhoneNumber .CAPABILITY_BOTH ,
        ],
        is_active =True ,
        )
        .values_list ("number",flat =True )
        .first ()
        )
        if not bot_number :
            django_messages .error (
            request ,
            "No se encontró ningún número WhatsApp activo para esta empresa.",
            )
            return redirect ("panel:bot_management")








        if action =="onboarding":
            section_id =request .POST .get ("section_id")
            if not section_id :
                django_messages .error (request ,"Debes seleccionar una sección.")
                return redirect ("panel:bot_management")
            try :
                section =Section .objects .get (pk =section_id ,company =company )
            except Section .DoesNotExist :
                django_messages .error (request ,"Sección no válida.")
                return redirect ("panel:bot_management")



            try :
                onboarding_template =WhatsAppTemplate .objects .get (
                name ="chat_onboarding",
                company =company ,
                )
                template_sid =onboarding_template .content_sid 
            except WhatsAppTemplate .DoesNotExist :
                django_messages .error (
                request ,
                "No se encontró el template chat_onboarding en la base de datos.",
                )
                return redirect ("panel:bot_management")



            pending_contacts =(
            Contact .objects 
            .filter (
            company =company ,
            section_assignments__section =section ,
            )
            .exclude (alias_onboarding_step =Contact .ALIAS_STEP_NONE )
            .exclude (
            company_user__isnull =False ,
            company_user__alias__isnull =False ,
            )
            .exclude (phone_number ="")
            .distinct ()
            )

            sent_count =0 
            for contact in pending_contacts :
                if not contact .phone_number :
                    continue 
                try :
                    WhatsAppChatService .send_quick_reply (
                    from_number =bot_number ,
                    to_number =contact .phone_number ,
                    content_sid =template_sid ,
                    content_variables ={
                    "1":contact .name ,
                    "2":company .name ,
                    },
                    )
                    sent_count +=1 
                except Exception as exc :
                    logger .error (
                    "# [BOT MGMT] Error enviando onboarding a %s: %s",
                    contact .phone_number ,exc ,
                    )

            django_messages .success (
            request ,
            f"Onboarding lanzado a {sent_count} contacto(s) de la sección '{section.name}'.",
            )
            return redirect ("panel:bot_management")

















        if action =="group_broadcast":
            selected_section_pks =request .POST .getlist ("section_pks")
            message_body =request .POST .get ("message","").strip ()
            if not selected_section_pks or not message_body :
                django_messages .error (
                request ,
                "Debes seleccionar al menos una sección y escribir un mensaje.",
                )
                return redirect ("panel:bot_management")



            from ivr_config .models import Section as _Section_BC 
            valid_sections =_Section_BC .objects .filter (
            pk__in =selected_section_pks ,
            company =company ,
            is_broadcast_enabled =True ,
            )
            if not valid_sections .exists ():
                django_messages .error (request ,"Ninguna sección seleccionada es válida para circulares.")
                return redirect ("panel:bot_management")

            from django .utils .timezone import now as _now_bc 
            from datetime import timedelta as _td_bc 
            import json as _json_bc 
            from whatsapp .models import WhatsAppSession as _WAS_BC 
            from whatsapp .models import WhatsAppTemplate as _WAT_BC 



            try :
                _renewal_template =_WAT_BC .objects .get (
                name ="chat_session_renewal",company =company 
                )
            except _WAT_BC .DoesNotExist :
                _renewal_template =None 
                logger .warning (
                "# [BOT MGMT] Template chat_session_renewal no encontrado "
                "para empresa pk=%r. Contactos fuera de ventana no recibirán renewal.",
                company .pk ,
                )

            _window_threshold =_now_bc ()-_td_bc (hours =24 )
            _created_at_iso =_now_bc ().isoformat ()
            total_sent =0 
            total_pending =0 
            for section in valid_sections :


                room =ChatRoom .objects .filter (
                company =company ,
                room_type =ChatRoom .ROOM_TYPE_SECTION ,
                section =section ,
                is_active =True ,
                ).first ()
                if room is None :
                    logger .warning (
                    "# [BOT MGMT] No hay ChatRoom SECTION activa para sección pk=%r (%s).",
                    section .pk ,section .name ,
                    )
                    continue 



                room_contacts =(
                Contact .objects 
                .filter (
                company =company ,
                sections =section ,
                company_user__isnull =False ,
                company_user__is_active =True ,
                )
                .exclude (phone_number ="")
                .distinct ()
                )

                room_sent =0 
                for contact in room_contacts :


                    _session =_WAS_BC .objects .filter (
                    company =company ,
                    phone_number =contact .phone_number ,
                    is_active =True ,
                    last_message_at__gte =_window_threshold ,
                    ).order_by ("-last_message_at").first ()

                    if _session is not None :


                        try :
                            WhatsAppChatService .send_reply (
                            from_number =bot_number ,
                            to_number =contact .phone_number ,
                            reply_text =message_body ,
                            )
                            room_sent +=1 
                            total_sent +=1 
                            logger .info (
                            "# [BOT MGMT] group_broadcast directo a %s (ventana activa).",
                            contact .phone_number ,
                            )
                        except Exception as exc :
                            logger .error (
                            "# [BOT MGMT] Error en group_broadcast directo a %s: %s",
                            contact .phone_number ,exc ,
                            )
                    else :




                        _out_session =_WAS_BC .objects .filter (
                        company =company ,
                        phone_number =contact .phone_number ,
                        ).order_by ("-session_start").first ()
                        if _out_session is None :
                            logger .warning (
                            "# [BOT MGMT] No hay sesión WhatsApp para %s — omitido.",
                            contact .phone_number ,
                            )
                            continue 


                        _pending =list (_out_session .pending_broadcast_messages or [])
                        _pending .append ({
                        "body":message_body ,
                        "created_at":_created_at_iso ,
                        })
                        _out_session .pending_broadcast_messages =_pending 
                        _out_session .save (update_fields =["pending_broadcast_messages"])


                        if _renewal_template :
                            try :
                                WhatsAppChatService .send_quick_reply (
                                from_number =bot_number ,
                                to_number =contact .phone_number ,
                                content_sid =_renewal_template .content_sid ,
                                content_variables ={
                                "1":contact .name or contact .phone_number ,
                                "2":company .name ,
                                "3":"/panel/",
                                },
                                )
                                total_pending +=1 
                                logger .info (
                                "# [BOT MGMT] chat_session_renewal enviado a %s —"
                                " circular encolada en pending_broadcast_messages.",
                                contact .phone_number ,
                                )
                            except Exception as exc :
                                logger .error (
                                "# [BOT MGMT] Error enviando renewal a %s: %s",
                                contact .phone_number ,exc ,
                                )



                if room_sent >0 :
                    ChatMessage .objects .create (
                    room =room ,
                    direction =ChatMessage .DIRECTION_OUTBOUND ,
                    body =message_body ,
                    whatsapp_sid ="",
                    )
                    logger .info (
                    "# [BOT MGMT] group_broadcast: %d directo(s) a sección '%s'.",
                    room_sent ,section .name ,
                    )

            django_messages .success (
            request ,
            f"Circular enviada: {total_sent} entregado(s) directamente, "
            f"{total_pending} renewal(s) enviado(s) con mensaje en cola.",
            )
            return redirect ("panel:bot_management")








        if action =="direct_broadcast":
            message_body =request .POST .get ("message","").strip ()
            section_id =request .POST .get ("section_id","").strip ()
            if not message_body :
                django_messages .error (request ,"Debes escribir un mensaje.")
                return redirect ("panel:bot_management")

            contacts_qs =Contact .objects .filter (
            company =company ,
            opt_out_broadcast =False ,
            ).exclude (phone_number ="")

            if section_id :
                contacts_qs =contacts_qs .filter (
                section_assignments__section_id =section_id ,
                section_assignments__section__company =company ,
                )

            sent_count =0 
            for contact in contacts_qs .distinct ():
                try :
                    WhatsAppChatService .send_reply (
                    from_number =bot_number ,
                    to_number =contact .phone_number ,
                    reply_text =message_body ,
                    )
                    sent_count +=1 
                except Exception as exc :
                    logger .error (
                    "# [BOT MGMT] Error en direct_broadcast a %s: %s",
                    contact .phone_number ,exc ,
                    )

            django_messages .success (
            request ,
            f"Circular 1:1 enviada a {sent_count} contacto(s).",
            )
            return redirect ("panel:bot_management")



        django_messages .warning (request ,"Acción no reconocida.")
        return redirect ("panel:bot_management")


class MachineAssetAnalyticsView (SupervisorAccessMixin ,View ):
    """
    Displays an activity report grouped by cost centre (MachineAsset).
    Aggregates total hours worked and total associated work-order entry lines
    per asset, with optional filters for date range, family and active status.
    Supports CSV export via the `export` GET parameter.
    Only accessible to ADMIN and SUPERVISOR roles (AdminRoleRequiredMixin).

    GET  /panel/fleet/analytics/
    GET  /panel/fleet/analytics/?export=csv

    ---

    Muestra un informe de actividad agrupado por centro de gasto (MachineAsset).
    Agrega el total de horas trabajadas y el total de líneas de parte asociadas
    por activo, con filtros opcionales por rango de fechas, familia y estado.
    Soporta exportación CSV mediante el parámetro GET `export`.
    Solo accesible para los roles ADMIN y SUPERVISOR (AdminRoleRequiredMixin).

    GET  /panel/fleet/analytics/
    GET  /panel/fleet/analytics/?export=csv
    """

    template_name ="panel/fleet/analytics.html"





    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _parse_date (self ,value ):
        """
        Parses an ISO date string (YYYY-MM-DD) into a date object.
        Returns None if the value is absent or malformed.
        ---
        Parsea una cadena de fecha ISO (YYYY-MM-DD) en un objeto date.
        Devuelve None si el valor está ausente o mal formado.
        """
        from datetime import date as _date 
        if not value :
            return None 
        try :
            parts =value .strip ().split ("-")
            return _date (int (parts [0 ]),int (parts [1 ]),int (parts [2 ]))
        except (ValueError ,IndexError ,AttributeError ):
            return None 

    def _build_report (self ,company ,request ):
        """
        Builds and returns the annotated analytics queryset and active
        filter values from GET parameters.

        Filters:
          date_from / date_to  — filter by WorkOrderEntryLine entry work_date.
          family               — exact match on MachineAsset.family.
          is_active            — '1' active only, '0' inactive only, '' all.

        Annotations per MachineAsset row:
          total_hours   — SUM of delta_hours from associated WorkOrderEntryLine.
          total_entries — COUNT of associated WorkOrderEntryLine (distinct).

        ---

        Construye y devuelve el queryset analítico anotado y los valores de
        filtro activos desde los parámetros GET.

        Filtros:
          date_from / date_to  — filtrar por work_date de la entrada de la línea.
          family               — coincidencia exacta con MachineAsset.family.
          is_active            — '1' solo activos, '0' solo inactivos, '' todos.

        Anotaciones por fila MachineAsset:
          total_hours   — SUMA de delta_hours de las WorkOrderEntryLine asociadas.
          total_entries — COUNT de WorkOrderEntryLine asociadas (distinct).
        """
        from fleet .models import MachineAsset 
        from django .db .models import Sum ,Count ,FloatField 
        from django .db .models .functions import Coalesce 



        date_from_raw =request .GET .get ("date_from","").strip ()
        date_to_raw =request .GET .get ("date_to","").strip ()
        family_filter =request .GET .get ("family","").strip ()
        active_filter =request .GET .get ("is_active","1")

        date_from =self ._parse_date (date_from_raw )
        date_to =self ._parse_date (date_to_raw )



        qs =MachineAsset .objects .filter (company =company )

        if family_filter :
            qs =qs .filter (family__iexact =family_filter )

        if active_filter =="1":
            qs =qs .filter (is_active =True )
        elif active_filter =="0":
            qs =qs .filter (is_active =False )






        line_filter_kwargs ={}
        if date_from :
            line_filter_kwargs ["work_order_lines__entry__work_date__gte"]=date_from 
        if date_to :
            line_filter_kwargs ["work_order_lines__entry__work_date__lte"]=date_to 






        if line_filter_kwargs :
            qs =qs .annotate (
            total_hours =Coalesce (
            Sum (
            "work_order_lines__delta_hours",
            filter =django_models .Q (**line_filter_kwargs ),
            ),
            0.0 ,
            output_field =FloatField (),
            ),
            total_entries =Coalesce (
            Count (
            "work_order_lines",
            filter =django_models .Q (**line_filter_kwargs ),
            distinct =True ,
            ),
            0 ,
            ),
            )
        else :
            qs =qs .annotate (
            total_hours =Coalesce (
            Sum ("work_order_lines__delta_hours"),
            0.0 ,
            output_field =FloatField (),
            ),
            total_entries =Coalesce (
            Count ("work_order_lines",distinct =True ),
            0 ,
            ),
            )

        qs =qs .order_by ("-total_hours","code")

        return qs ,{
        "date_from":date_from_raw ,
        "date_to":date_to_raw ,
        "filter_family":family_filter ,
        "filter_is_active":active_filter ,
        }

    def _get_families (self ,company ):
        """
        Returns a sorted list of distinct family values for the company.
        ---
        Retorna una lista ordenada de valores de family distintos para la empresa.
        """
        from fleet .models import MachineAsset 
        return (
        MachineAsset .objects 
        .filter (company =company )
        .exclude (family ="")
        .values_list ("family",flat =True )
        .distinct ()
        .order_by ("family")
        )





    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the analytics page or exports a CSV file when
        the `export=csv` GET parameter is present.
        ---
        Renderiza la página de analítica o exporta un CSV cuando
        el parámetro GET `export=csv` está presente.
        """
        import csv 
        from django .http import HttpResponse 

        company_user =request .user .company_user 
        company =company_user .company 

        qs ,filters =self ._build_report (company ,request )



        if request .GET .get ("export")=="csv":
            response =HttpResponse (content_type ="text/csv; charset=utf-8")
            response ["Content-Disposition"]=(
            'attachment; filename="centros_de_gasto_actividad.csv"'
            )


            response .write ("\ufeff")
            writer =csv .writer (response ,delimiter =";")
            writer .writerow ([
            "Código","Familia","Marca / Modelo","Matrícula",
            "Estado","Total horas","Total partes",
            ])
            for asset in qs :
                writer .writerow ([
                asset .code ,
                asset .family ,
                asset .brand_model ,
                asset .plate ,
                "Activo"if asset .is_active else "Inactivo",
                f"{asset.total_hours:.2f}"if asset .total_hours else "0.00",
                asset .total_entries ,
                ])
            return response 



        ctx ={
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"fleet_analytics",
        "assets":qs ,
        "families":self ._get_families (company ),
        "total_hours_sum":sum (
        a .total_hours for a in qs if a .total_hours 
        ),
        **filters ,
        }
        return render (request ,self .template_name ,ctx )








class CompanySettingsView (AdminRoleRequiredMixin ,View ):
    """
    Allows ADMIN users to edit the company-level text fields
    operation_bases and labor_calendar on the Company model.
    GET: renders the settings form pre-populated with the current values.
    POST: validates and saves both fields, then redirects back with a
    success message.
    ---
    Permite a los usuarios ADMIN editar los campos de texto de nivel empresa
    operation_bases y labor_calendar del modelo Company.
    GET: renderiza el formulario de configuración con los valores actuales.
    POST: valida y guarda ambos campos, luego redirige con mensaje de éxito.
    """

    template_name ="panel/company/settings.html"

    def get (self ,request ):
        """
        Render the company settings page with bases, schedules and night
        shift fields. The obsolete operation_bases and labor_calendar text
        fields have been replaced by live querysets from the database.
        ---
        Renderiza la página de configuración de empresa con bases, horarios
        y franja nocturna. Los campos de texto obsoletos operation_bases y
        labor_calendar han sido sustituidos por querysets en vivo desde BD.
        """
        import json as _json 
        from ivr_config .models import Company ,WorkdaySchedule 
        from budgets .models import Base 
        company_user =request .user .company_user 
        company =company_user .company 
        bases =Base .objects .filter (company =company ).order_by ("name")


        bases_data =[]
        for base in bases :
            holiday_count =0 
            if base .labor_calendar :
                try :
                    holiday_count =len (_json .loads (base .labor_calendar ))
                except (ValueError ,TypeError ):
                    holiday_count =0 
            bases_data .append ({
            "base":base ,
            "holiday_count":holiday_count ,
            })
        schedules =WorkdaySchedule .objects .filter (
        company =company 
        ).order_by ("label")
        ctx ={
        "company":company ,
        "company_user":company_user ,
        "own_presence":PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first (),
        "active_nav":"company_settings",
        "bases_data":bases_data ,
        "schedules":schedules ,
        }
        return render (request ,self .template_name ,ctx )

    def post (self ,request ):
        """
        Save night_start and night_end fields to the Company instance.
        The obsolete operation_bases and labor_calendar fields are no longer
        persisted from this form — they are managed via Base model and the
        sync_base_calendars command.
        ---
        Guarda los campos night_start y night_end en la instancia Company.
        Los campos obsoletos operation_bases y labor_calendar ya no se
        persisten desde este formulario — se gestionan via modelo Base y
        el comando sync_base_calendars.
        """
        import datetime as _dt 
        company_user =request .user .company_user 
        company =company_user .company 



        night_start_raw =request .POST .get ("night_start","").strip ()
        night_end_raw =request .POST .get ("night_end","").strip ()
        try :
            company .night_start =(
            _dt .time .fromisoformat (night_start_raw )
            if night_start_raw else _dt .time (22 ,0 )
            )
        except ValueError :
            company .night_start =_dt .time (22 ,0 )
        try :
            company .night_end =(
            _dt .time .fromisoformat (night_end_raw )
            if night_end_raw else _dt .time (6 ,0 )
            )
        except ValueError :
            company .night_end =_dt .time (6 ,0 )

        company .save (update_fields =[
        "night_start",
        "night_end",
        ])
        django_messages .success (request ,"Configuración de empresa guardada correctamente.")
        return redirect ("panel:company_settings")







class ExportTemplateListView (SupervisorAccessMixin ,View ):
    """
    Lists all ExportTemplate records belonging to the authenticated user.
    Returns a JSON response suitable for AJAX calls from the export modal,
    or renders a full HTML page for standalone template management.
    GET ?format=json — returns [{id, name, is_default, columns, sheet_format,
                                   operator_scope}] for the modal.
    GET             — renders panel/export_templates/list.html.
    ---
    Lista todos los registros ExportTemplate del usuario autenticado.
    Devuelve JSON para llamadas AJAX desde el modal de exportación,
    o renderiza una página HTML para la gestión autónoma de plantillas.
    GET ?format=json — lista [{id, name, is_default, ...}] para el modal.
    GET             — renderiza panel/export_templates/list.html.
    """

    template_name ="panel/export_templates/list.html"

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns the user's export templates as JSON or HTML.
        Auto-creates the default template if the user has none.
        ---
        Devuelve las plantillas del usuario como JSON o HTML.
        Crea la plantilla por defecto si el usuario no tiene ninguna.
        """
        from django .http import JsonResponse 
        from work_order_processor .models import ExportTemplate 

        cu =request .user .company_user 


        ExportTemplate .get_or_create_default (cu )
        templates =ExportTemplate .objects .filter (company_user =cu ).order_by ("-is_default","name")

        if request .GET .get ("format")=="json":
            data =[
            {
            "id":t .pk ,
            "name":t .name ,
            "is_default":t .is_default ,
            "columns":t .columns ,
            "sheet_format":t .sheet_format ,
            "operator_scope":t .operator_scope ,
            }
            for t in templates 
            ]
            return JsonResponse ({"templates":data })

        cu_obj =request .user .company_user 
        context ={
        "company":cu_obj .company ,
        "company_user":cu_obj ,
        "own_presence":PresenceStatus .objects .filter (
        company_user =cu_obj ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first (),
        "active_nav":"work_order_admin_history",
        "templates":templates ,


        "column_choices":[
        ("fecha","Fecha"),
        ("operario","Operario"),
        ("maquina","Máquina / CdG"),
        ("descripcion","Descripción avería"),
        ("notas","Notas reparación"),
        ("hc","H. inicio"),
        ("hf","H. fin"),
        ("delta_horas","Δ Horas"),
        ("estado","Estado"),
        ("familia","Familia avería"),
        ("origen","Origen"),
        ],
        }
        return render (request ,self .template_name ,context )


class ExportTemplateCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a new ExportTemplate for the authenticated user.
    Accepts POST with JSON body or form data.
    Returns JSON {id, name} on success or {error} on failure.
    ---
    Crea una nueva ExportTemplate para el usuario autenticado.
    Acepta POST con cuerpo JSON o datos de formulario.
    Devuelve JSON {id, name} en éxito o {error} en fallo.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates and creates the ExportTemplate.
        ---
        Valida y crea la ExportTemplate.
        """
        import json as _json 
        from django .http import JsonResponse 
        from work_order_processor .models import ExportTemplate 

        cu =request .user .company_user 

        try :
            body =_json .loads (request .body )
        except (ValueError ,TypeError ):
            body =request .POST 

        name =str (body .get ("name","")).strip ()
        columns =body .get ("columns",[])
        sheet_format =str (body .get ("sheet_format",ExportTemplate .SheetFormat .SINGLE_SHEET )).strip ()
        operator_scope =str (body .get ("operator_scope",ExportTemplate .OperatorScope .ALL )).strip ()
        is_default =bool (body .get ("is_default",False ))

        if not name :
            return JsonResponse ({"error":"El nombre es obligatorio."},status =400 )
        if not columns :
            return JsonResponse ({"error":"Selecciona al menos una columna."},status =400 )
        if sheet_format not in ExportTemplate .SheetFormat .values :
            return JsonResponse ({"error":"Formato de hoja no válido."},status =400 )
        if operator_scope not in ExportTemplate .OperatorScope .values :
            return JsonResponse ({"error":"Alcance de operarios no válido."},status =400 )
        if ExportTemplate .objects .filter (company_user =cu ,name =name ).exists ():
            return JsonResponse (
            {"error":f"Ya existe una plantilla con el nombre '{name}'."},status =400 
            )

        template =ExportTemplate .objects .create (
        company_user =cu ,
        name =name ,
        columns =list (columns ),
        sheet_format =sheet_format ,
        operator_scope =operator_scope ,
        is_default =is_default ,
        )
        logger .info (
        "# [EXPORT TEMPLATE] Plantilla pk=%s '%s' creada por %s.",
        template .pk ,template .name ,cu .user .username ,
        )
        return JsonResponse ({"id":template .pk ,"name":template .name },status =201 )


class ExportTemplateUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing ExportTemplate belonging to the authenticated user.
    Returns JSON {ok: true} on success or {error} on failure.
    ---
    Actualiza una ExportTemplate existente del usuario autenticado.
    Devuelve JSON {ok: true} en éxito o {error} en fallo.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates and applies the update.
        ---
        Valida y aplica la actualización.
        """
        import json as _json 
        from django .http import JsonResponse 
        from work_order_processor .models import ExportTemplate 

        cu =request .user .company_user 

        try :
            template =ExportTemplate .objects .get (pk =pk ,company_user =cu )
        except ExportTemplate .DoesNotExist :
            return JsonResponse ({"error":"Plantilla no encontrada."},status =404 )

        try :
            body =_json .loads (request .body )
        except (ValueError ,TypeError ):
            body =request .POST 

        name =str (body .get ("name",template .name )).strip ()
        columns =body .get ("columns",template .columns )
        sheet_format =str (body .get ("sheet_format",template .sheet_format )).strip ()
        operator_scope =str (body .get ("operator_scope",template .operator_scope )).strip ()
        is_default =bool (body .get ("is_default",template .is_default ))

        if not name :
            return JsonResponse ({"error":"El nombre es obligatorio."},status =400 )
        if not columns :
            return JsonResponse ({"error":"Selecciona al menos una columna."},status =400 )
        if sheet_format not in ExportTemplate .SheetFormat .values :
            return JsonResponse ({"error":"Formato de hoja no válido."},status =400 )
        if operator_scope not in ExportTemplate .OperatorScope .values :
            return JsonResponse ({"error":"Alcance de operarios no válido."},status =400 )
        if (
        name !=template .name 
        and ExportTemplate .objects .filter (company_user =cu ,name =name ).exists ()
        ):
            return JsonResponse (
            {"error":f"Ya existe una plantilla con el nombre '{name}'."},status =400 
            )

        template .name =name 
        template .columns =list (columns )
        template .sheet_format =sheet_format 
        template .operator_scope =operator_scope 
        template .is_default =is_default 
        template .save ()
        logger .info (
        "# [EXPORT TEMPLATE] Plantilla pk=%s '%s' actualizada por %s.",
        template .pk ,template .name ,cu .user .username ,
        )
        return JsonResponse ({"ok":True })


class ExportTemplateDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes an ExportTemplate belonging to the authenticated user.
    Returns JSON {ok: true} on success or {error} on failure.
    Cannot delete the last remaining template of a user.
    ---
    Elimina una ExportTemplate del usuario autenticado.
    Devuelve JSON {ok: true} en éxito o {error} en fallo.
    No permite eliminar la última plantilla del usuario.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Deletes the template and redirects to the list.
        ---
        Elimina la plantilla y redirige a la lista.
        """
        from django .http import JsonResponse 
        from work_order_processor .models import ExportTemplate 

        cu =request .user .company_user 

        try :
            template =ExportTemplate .objects .get (pk =pk ,company_user =cu )
        except ExportTemplate .DoesNotExist :
            return JsonResponse ({"error":"Plantilla no encontrada."},status =404 )

        if ExportTemplate .objects .filter (company_user =cu ).count ()<=1 :
            return JsonResponse (
            {"error":"No puedes eliminar tu única plantilla de exportación."},
            status =400 ,
            )

        name =template .name 
        template .delete ()
        logger .info (
        "# [EXPORT TEMPLATE] Plantilla '%s' eliminada por %s.",
        name ,cu .user .username ,
        )
        return JsonResponse ({"ok":True })


class WorkOrderAdminExportByTemplateView (SupervisorAccessMixin ,View ):
    """
    Generates and streams an Excel file from selected WorkOrder PKs
    using the configuration of the chosen ExportTemplate.

    POST /panel/work-orders/export-by-template/
         Body params:
           template_pk  (int)       — ExportTemplate pk (must belong to user).
           pks          (list[int]) — WorkOrder primary keys to export.
           operator_pks (list[int]) — optional operator filter when
                                       template.operator_scope == 'selection'.

    Returns HttpResponse with Content-Disposition attachment (xlsx).
    Returns HTTP 400 on invalid input or HTTP 404 on unknown template.
    ---
    Genera y devuelve en streaming un Excel desde los PKs de WorkOrder
    seleccionados usando la configuración de la ExportTemplate elegida.

    POST /panel/work-orders/export-by-template/
         Parámetros del cuerpo:
           template_pk  (int)       — pk de ExportTemplate (debe pertenecer al usuario).
           pks          (list[int]) — claves primarias de WorkOrder a exportar.
           operator_pks (list[int]) — filtro de operario opcional cuando
                                       template.operator_scope == 'selection'.

    Devuelve HttpResponse con Content-Disposition attachment (xlsx).
    Devuelve HTTP 400 en entrada inválida o HTTP 404 en plantilla desconocida.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Resolves the template and work orders, calls build_export_from_template
        and streams the resulting workbook as an xlsx attachment.
        ---
        Resuelve la plantilla y los partes, llama a build_export_from_template
        y devuelve el libro resultante como adjunto xlsx.
        """
        import io 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import ExportTemplate 
        from work_order_processor .services import build_export_from_template 

        cu =request .user .company_user 
        company =cu .company 





        try :
            template_pk =int (request .POST .get ("template_pk",""))
            template =ExportTemplate .objects .get (pk =template_pk ,company_user =cu )
        except (ValueError ,TypeError ,ExportTemplate .DoesNotExist ):
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] Plantilla no encontrada o no pertenece al usuario."
            )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (p )for p in raw_pks if str (p ).strip ().isdigit ()]
        except (ValueError ,TypeError ):
            pk_list =[]

        if not pk_list :
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] No se han seleccionado partes para exportar."
            )





        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).select_related (
        "uploaded_by__user",
        ).prefetch_related (
        Prefetch (
        "entries",
        queryset =WorkOrderEntry .objects .prefetch_related ("lines"),
        )
        ).order_by (
        "uploaded_by__user__last_name",
        "uploaded_by__user__first_name",
        "pk",
        )





        if template .operator_scope =="selection":
            raw_op_pks =request .POST .getlist ("operator_pks")
            try :
                op_pk_list =[int (p )for p in raw_op_pks if str (p ).strip ().isdigit ()]
            except (ValueError ,TypeError ):
                op_pk_list =[]
            if op_pk_list :
                qs =qs .filter (uploaded_by__pk__in =op_pk_list )

        if not qs .exists ():
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] Ninguno de los partes seleccionados es válido para exportar."
            )





        wb =build_export_from_template (template ,qs )
        buf =io .BytesIO ()
        wb .save (buf )
        buf .seek (0 )

        filename =(
        f"partes_digitales_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        response =HttpResponse (
        buf .getvalue (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        logger .info (
        "# [EXPORT BY TEMPLATE] Exportación '%s' (%d partes) por %s.",
        template .name ,qs .count (),cu .user .username ,
        )
        return response 
