

# /home/MiguelAeTxio/PROJECTS/EnterpriseBot/panel/views_workorders.py


from django.contrib import messages as django_messages
from django.views.generic import View, ListView
from django.shortcuts import redirect, render
from django.db.models import Q, Prefetch
from django.utils.timezone import now
from django.forms import modelformset_factory
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

from panel.mixins import SupervisorAccessMixin, AdminRoleRequiredMixin, WorkOrderFormAccessMixin, CompanyUserRequiredMixin
from ivr_config.models import (
    CompanyUser,
    PresenceStatus,
    Section,
)
from work_order_processor.models import (
    WorkOrder,
    WorkOrderEntry,
    WorkOrderEntryLine,
)
from work_order_processor.services import find_cached_classification
from work_order_processor.tasks import (
    classify_fault_line,
    generate_period_excel,
    process_work_order_pdf,
)
from fleet.models import MachineAsset
import logging
import plotly.graph_objects as go
import plotly.io as pio

logger = logging.getLogger(__name__)

class WorkOrderListView (SupervisorAccessMixin ,View ):
    """
    Lists WorkOrder records belonging to the authenticated user's company,
    split into four querysets for the tabbed UI introduced in Hito 8 / Bloque H:
      wo_queue    — PENDING or PROCESSING (polling tab).
      wo_error    — ERROR.
      wo_pending  — DONE, not yet reviewed (pending supervisor sign-off).
      wo_reviewed — DONE and reviewed.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).
    ---
    Lista los registros WorkOrder de la empresa del usuario autenticado,
    divididos en cuatro querysets para la UI de pestañas introducida en
    el Hito 8 / Bloque H:
      wo_queue    — PENDING o PROCESSING (pestaña con polling).
      wo_error    — ERROR.
      wo_pending  — DONE sin revisión (pendiente de validación del Supervisor).
      wo_reviewed — DONE y revisados.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    template_name ="panel/work_orders/list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .db .models import Q 
        from django .utils .timezone import now 
        from ivr_config .models import PresenceStatus 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Renders the tabbed work order list with four filtered querysets.
        The active tab defaults to "Pendiente revisión" when there are pending
        work orders; otherwise defaults to "En cola".
        Supports multi-level column sort via the sort_stack GET parameter
        (same mechanism as WorkOrderAdminHistoryView). Columns sortable from
        the PDF list: upload_date, pdf_display_name, uploaded_by_name.
        ---
        Renderiza la lista de partes con pestañas y cuatro querysets filtrados.
        La pestaña activa por defecto es "Pendiente revisión" si hay partes
        pendientes; en caso contrario, "En cola".
        Soporta ordenación multi-nivel via el parámetro GET sort_stack
        (mismo mecanismo que WorkOrderAdminHistoryView). Columnas ordenables
        desde la lista PDF: upload_date, pdf_display_name, uploaded_by_name.
        """
        company_user =request .user .company_user 
        company =company_user .company 









        _VALID_SORT_COLS =frozenset (
        ("upload_date","pdf_display_name","uploaded_by_name")
        )
        _VALID_SORT_DIRS =frozenset (("asc","desc"))



        _ORM_FIELD_MAP ={
        "upload_date":[("upload_date",)],
        "pdf_display_name":[("source_pdf__name",)],
        "uploaded_by_name":[
        ("uploaded_by__user__last_name",),
        ("uploaded_by__user__first_name",),
        ],
        }

        raw_stack =request .GET .get ("sort_stack","").strip ()

        sort_stack =[]
        if raw_stack :
            for token in raw_stack .split (","):
                token =token .strip ()
                if ":"not in token :
                    continue 
                col ,_ ,direction =token .partition (":")
                col =col .strip ()
                direction =direction .strip ()
                if col in _VALID_SORT_COLS and direction in _VALID_SORT_DIRS :
                    if not any (s [0 ]==col for s in sort_stack ):
                        sort_stack .append ((col ,direction ))
                        if len (sort_stack )>=3 :
                            break 

        if not sort_stack :
            sort_stack =[("upload_date","desc")]

        sort_primary_col =sort_stack [0 ][0 ]
        sort_primary_dir =sort_stack [0 ][1 ]
        sort_stack_str =",".join (f"{col}:{direction}"for col ,direction in sort_stack )

        def _build_order_by (stack ):
            """
            Builds an order_by tuple from the sort stack, applying the
            correct direction prefix ("-") for each column and its ORM fields.
            Secondary and tertiary columns act as tiebreakers in order.
            ---
            Construye una tupla order_by desde la pila de ordenación, aplicando
            el prefijo de dirección ("-") correcto a cada columna y sus campos ORM.
            Las columnas secundaria y terciaria actúan como desempate en orden.
            """
            order_fields =[]
            for col ,direction in stack :
                prefix ="-"if direction =="desc"else ""
                for field_tuple in _ORM_FIELD_MAP .get (col ,[(col ,)]):
                    for field in field_tuple :
                        order_fields .append (f"{prefix}{field}")
            return order_fields 

        orm_order =_build_order_by (sort_stack )






        wo_queue =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status__in =[
        WorkOrder .Status .PENDING ,
        WorkOrder .Status .PROCESSING ,
        ],
        )
        .order_by ("-upload_date")
        )
        wo_error =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .ERROR ,
        )
        .order_by ("-upload_date")
        )
        wo_pending =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .DONE ,
        reviewed =False ,
        )
        .select_related ("uploaded_by__user")
        .order_by (*orm_order )
        )
        wo_reviewed =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        )
        .select_related ("reviewed_by__user","uploaded_by__user")
        .order_by (*orm_order )
        )





        default_tab ="pending"if wo_pending .exists ()else "queue"

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        "wo_queue":wo_queue ,
        "wo_error":wo_error ,
        "wo_pending":wo_pending ,
        "wo_reviewed":wo_reviewed ,
        "default_tab":default_tab ,
        "sort_stack_str":sort_stack_str ,
        "sort_primary_col":sort_primary_col ,
        "sort_primary_dir":sort_primary_dir ,
        })

    def post (self ,request ):
        """
        Handles bulk actions submitted from the PDF pipeline list tabs.
        Supported bulk_op values:
          mark_reviewed   — sets reviewed=True on selected PDF_UPLOAD WorkOrders.
          unmark_reviewed — sets reviewed=False on selected PDF_UPLOAD WorkOrders.
          delete          — deletes selected PDF_UPLOAD WorkOrders (CASCADE).
        All operations are scoped to the authenticated user's company.

        ---

        Gestiona acciones en lote enviadas desde las pestañas de la lista PDF.
        Valores de bulk_op soportados:
          mark_reviewed   — establece reviewed=True en los WorkOrders seleccionados.
          unmark_reviewed — establece reviewed=False en los WorkOrders seleccionados.
          delete          — elimina los WorkOrders seleccionados (CASCADE).
        Todas las operaciones están acotadas a la empresa del usuario autenticado.
        """
        from django .utils .timezone import now as _now 
        from django .contrib import messages as django_messages 

        company_user =request .user .company_user 
        company =company_user .company 
        active_tab =request .POST .get ("active_tab","pending")
        bulk_op =request .POST .get ("bulk_op","")
        raw_pks =request .POST .getlist ("pks")



        try :
            pk_list =[int (p )for p in raw_pks if str (p ).strip ().isdigit ()]
        except (ValueError ,AttributeError ):
            pk_list =[]

        if not pk_list :
            django_messages .warning (request ,"No se ha seleccionado ningún parte.")
            return redirect (f"{request.path}?tab={active_tab}")

        qs =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source =WorkOrder .Source .PDF_UPLOAD ,
        pk__in =pk_list ,
        )
        )

        if bulk_op =="mark_reviewed":

            updated =qs .filter (status =WorkOrder .Status .DONE ).update (
            reviewed =True ,
            reviewed_by =company_user ,
            reviewed_at =_now (),
            )
            django_messages .success (
            request ,
            f"{updated} parte{'s' if updated != 1 else ''} "
            f"marcado{'s' if updated != 1 else ''} como revisado{'s' if updated != 1 else ''}."
            )

        elif bulk_op =="unmark_reviewed":

            updated =qs .filter (status =WorkOrder .Status .DONE ,reviewed =True ).update (
            reviewed =False ,
            reviewed_by =None ,
            reviewed_at =None ,
            )
            django_messages .success (
            request ,
            f"Revisión desmarcada en {updated} parte{'s' if updated != 1 else ''}."
            )

        elif bulk_op =="delete":

            count =qs .count ()
            qs .delete ()
            django_messages .success (
            request ,
            f"{count} parte{'s' if count != 1 else ''} "
            f"eliminado{'s' if count != 1 else ''} correctamente."
            )

        else :
            django_messages .warning (
            request ,
            f"Operación en lote desconocida: '{bulk_op}'."
            )

        return redirect (f"{request.path}?tab={active_tab}")








class WorkOrderUploadView (SupervisorAccessMixin ,View ):
    """
    Handles PDF upload for work order processing.
    On POST: validates the file, runs a duplicate detection check before
    creating the WorkOrder, enqueues the Celery processing task immediately
    via process_work_order_pdf.delay_on_commit() and redirects to the list.

    Duplicate detection (Hito 8, Paso 5, Bloque E):
      Before creating a new WorkOrder the view extracts the worker name and
      work period from the uploaded PDF filename and checks for an existing
      WorkOrder for the same company with the same worker name and an
      overlapping period. If a duplicate is found and the POST does not
      include confirm_overwrite=1, the upload form is re-rendered with the
      duplicate_wo context variable so the template can show a warning modal.
      If the user confirms overwrite, the duplicate WorkOrder (and all its
      cascade-deleted children) is deleted before the new one is created.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).
    ---
    Gestiona la carga de PDF para el procesamiento de partes de trabajo.
    En POST: valida el archivo, ejecuta una comprobación de duplicado antes
    de crear el WorkOrder, encola la tarea Celery inmediatamente mediante
    process_work_order_pdf.delay_on_commit() y redirige a la lista.

    Detección de duplicado (Hito 8, Paso 5, Bloque E):
      Antes de crear un nuevo WorkOrder la vista extrae el nombre del operario
      y el periodo de trabajo del nombre del PDF cargado y comprueba si existe
      un WorkOrder de la misma empresa con el mismo operario y periodo solapado.
      Si se detecta un duplicado y el POST no incluye confirm_overwrite=1, el
      formulario de carga se vuelve a renderizar con la variable de contexto
      duplicate_wo para que la plantilla muestre un modal de advertencia.
      Si el usuario confirma la sobrescritura, el WorkOrder duplicado (y todos
      sus hijos eliminados en cascada) se elimina antes de crear el nuevo.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    template_name ="panel/work_orders/upload.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        from django .db .models import Q 
        from django .utils .timezone import now 
        from ivr_config .models import PresenceStatus 
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ):
        """
        Renders the PDF upload form.
        ---
        Renderiza el formulario de carga de PDF.
        """
        company_user =request .user .company_user 
        return render (request ,self .template_name ,{
        "company":company_user .company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        })

    def post (self ,request ):
        """
        Processes the uploaded PDF: validates the file, computes its SHA-256
        hash, runs a two-level duplicate detection check, optionally deletes
        the existing duplicate WorkOrder on confirmed overwrite, creates a new
        WorkOrder (storing the hash) and enqueues the Celery processing task.

        Duplicate detection — two levels (Hito 8 / Bloque I):
          LEVEL 1 — Exact hash match:
            SHA-256 of the incoming file is compared against source_pdf_hash
            for all WorkOrders of the same company. A hash match means the
            exact same file has been uploaded before, regardless of filename.
          LEVEL 2 — Worker + period overlap (only if no hash match):
            The incoming PDF filename is used to derive the worker name and,
            when possible, the work period. Existing WorkOrderEntry records for
            the same company + worker_name + overlapping work_date range are
            queried. A match here means a different file covering the same
            data has already been processed.

          duplicate_reason carries "exact" or "content" so the template can
          show a differentiated warning message in the modal.

          On confirmed overwrite the duplicate WorkOrder (and all its cascade-
          deleted children) is removed before the new one is created.
        ---
        Procesa el PDF cargado: valida el archivo, calcula su hash SHA-256,
        ejecuta la detección de duplicado en dos niveles, elimina opcionalmente
        el WorkOrder duplicado si el usuario confirma la sobrescritura, crea
        un nuevo WorkOrder (guardando el hash) y encola la tarea Celery.

        Detección de duplicado — dos niveles (Hito 8 / Bloque I):
          NIVEL 1 — Hash exacto:
            El SHA-256 del fichero entrante se compara contra source_pdf_hash
            de todos los WorkOrders de la misma empresa. Un match de hash
            significa que se ha subido exactamente el mismo fichero antes,
            con independencia del nombre.
          NIVEL 2 — Operario + solapamiento de periodo (solo si no hay hash match):
            El nombre del PDF se usa para derivar el nombre del operario y,
            cuando es posible, el periodo de trabajo. Se consultan los registros
            WorkOrderEntry de la misma empresa + worker_name + rango work_date
            solapado. Un match aquí significa que un fichero diferente pero con
            los mismos datos ya ha sido procesado.

          duplicate_reason transporta "exact" o "content" para que la plantilla
          muestre un mensaje diferenciado en el modal.

          Al confirmar la sobrescritura el WorkOrder duplicado (y todos sus hijos
          eliminados en cascada) se elimina antes de crear el nuevo.
        """
        import hashlib 

        from django .contrib import messages as django_messages 
        from work_order_processor .services import _worker_name_from_pdf_path 
        from work_order_processor .tasks import _extract_period_from_pdf_name 
        from work_order_processor .models import WorkOrderEntry 

        company_user =request .user .company_user 
        company =company_user .company 
        pdf_file =request .FILES .get ("source_pdf")





        if not pdf_file :
            django_messages .error (request ,"Debes seleccionar un archivo PDF.")
            return render (request ,self .template_name ,{
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            })

        if not pdf_file .name .lower ().endswith (".pdf"):
            django_messages .error (request ,"El archivo debe tener extensión .pdf.")
            return render (request ,self .template_name ,{
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            })











        pdf_file .seek (0 )
        incoming_hash =hashlib .sha256 (pdf_file .read ()).hexdigest ()
        pdf_file .seek (0 )













        incoming_name =pdf_file .name 
        incoming_worker =_worker_name_from_pdf_path (incoming_name )
        date_from ,date_to =_extract_period_from_pdf_name (incoming_name )

        duplicate_wo =None 
        duplicate_reason =None 



        hash_duplicate =(
        WorkOrder .objects 
        .filter (company =company ,source_pdf_hash =incoming_hash )
        .exclude (source_pdf_hash ="")
        .first ()
        )
        if hash_duplicate :
            duplicate_wo =hash_duplicate 
            duplicate_reason ="exact"



        if not duplicate_wo and incoming_worker :
            entry_qs =WorkOrderEntry .objects .filter (
            work_order__company =company ,
            worker_name =incoming_worker ,
            ).select_related ("work_order")

            if date_from and date_to :




                entry_qs =entry_qs .filter (
                work_date__gte =date_from ,
                work_date__lte =date_to ,
                )



            existing_entry =entry_qs .first ()
            if existing_entry :
                duplicate_wo =existing_entry .work_order 
                duplicate_reason ="content"







        if not duplicate_wo and incoming_worker and date_from and date_to :
            conflicting_entries =(
            WorkOrderEntry .objects 
            .filter (
            work_order__company =company ,
            worker_name =incoming_worker ,
            work_date__gte =date_from ,
            work_date__lte =date_to ,
            )
            .exclude (work_order__source_pdf_hash =incoming_hash )
            .exclude (work_order__source_pdf_hash ="")
            .select_related ("work_order")
            .order_by ("work_date")
            )
            if conflicting_entries .exists ():


                seen_dates =set ()
                duplicate_dates =[]
                for entry in conflicting_entries :
                    if entry .work_date :
                        date_key =entry .work_date .strftime ("%d/%m/%y")
                        if date_key not in seen_dates :
                            seen_dates .add (date_key )
                            duplicate_dates .append (date_key )

                first_entry =conflicting_entries .first ()
                duplicate_wo =first_entry .work_order 
                duplicate_reason ="duplicate_entries"





        if duplicate_wo and not request .POST .get ("confirm_overwrite"):
            ctx ={
            "company":company ,
            "company_user":company_user ,
            "own_presence":self ._get_own_presence (company_user ),
            "active_nav":"work_orders",
            "duplicate_wo":duplicate_wo ,
            "duplicate_reason":duplicate_reason ,
            "pdf_file_name":incoming_name ,
            }


            if duplicate_reason =="duplicate_entries":
                ctx ["duplicate_dates"]=duplicate_dates 
            return render (request ,self .template_name ,ctx )



        if duplicate_wo and request .POST .get ("confirm_overwrite"):
            dup_pk =duplicate_wo .pk 
            duplicate_wo .delete ()
            django_messages .warning (
            request ,
            f"El parte duplicado #{dup_pk} y todos sus datos han sido "
            f"eliminados. Procesando el nuevo PDF."
            )
























        from django .db import transaction 

        with transaction .atomic ():









            race_duplicate =(
            WorkOrder .objects 
            .filter (company =company ,source_pdf_hash =incoming_hash )
            .exclude (source_pdf_hash ="")
            .select_for_update ()
            .first ()
            )
            if race_duplicate :
                django_messages .info (
                request ,
                f"El fichero ya había sido registrado como Parte "
                f"#{race_duplicate.pk}. No se ha creado un duplicado."
                )
                return redirect ("panel:work_order_list")

            work_order =WorkOrder .objects .create (
            company =company ,
            uploaded_by =company_user ,
            source_pdf =pdf_file ,
            source_pdf_hash =incoming_hash ,
            source_pdf_name =incoming_name ,
            )

        process_work_order_pdf .delay_on_commit (work_order .pk )

        django_messages .success (
        request ,
        f"PDF cargado correctamente (Parte #{work_order.pk}). "
        f"El procesamiento ha sido encolado y comenzará en instantes."
        )
        return redirect ("panel:work_order_list")


class WorkOrderEditView (SupervisorAccessMixin ,View ):
    """
    Displays and processes inline edits for all WorkOrderEntryLine records
    belonging to a given WorkOrder. Lines are grouped by their parent
    WorkOrderEntry (page / date) for readability.

    GET  — Renders the editable table grouped by page.
    POST — Two actions dispatched via hidden input `action`:
      "save_line"      : Saves a single WorkOrderEntryLine identified by `line_pk`.
                         Recomputes delta_hours from hc/hf and re-resolves
                         machine_asset from the updated machine_norm.
      "regenerate"     : Re-enqueues the Excel generation task for this WorkOrder
                         (status reset to PENDING) and redirects to the appropriate list.

    ⚠️ EXCLUSIVE TO PDF_UPLOAD PARTS. This view exists only to collate the
    historical PDF-derived data against what Gemini extracted from the
    original document — it has nothing in common with the digital-part
    workflow beyond writing to the same database tables. DIGITAL/GENERATED
    WorkOrders are hard-blocked in both get() and post(): they redirect to
    operator_form_edit, the single view that may ever display or edit a
    digital part, regardless of role. Access to this view itself is
    restricted to SUPERVISOR/WORKSHOPBOSS/ADMIN by SupervisorAccessMixin —
    WORKSHOP never reaches get()/post() here.
    Access is restricted to the authenticated company (multicompany guard).

    ---

    Muestra y procesa ediciones inline para todos los registros WorkOrderEntryLine
    de un WorkOrder dado. Las líneas se agrupan por su WorkOrderEntry padre
    (página / fecha) para facilitar la lectura.

    GET  — Renderiza la tabla editable agrupada por página.
    POST — Dos acciones despachadas mediante el input oculto `action`:
      "save_line"      : Guarda un único WorkOrderEntryLine identificado por `line_pk`.
                         Recalcula delta_hours desde hc/hf y re-resuelve machine_asset
                         desde el machine_norm actualizado.
      "regenerate"     : Re-encola la tarea de generación de Excel para este WorkOrder
                         (estado reseteado a PENDING) y redirige a la lista apropiada.

    ⚠️ EXCLUSIVA DE PARTES PDF_UPLOAD. Esta vista existe únicamente para
    cotejar los datos históricos extraídos del PDF contra lo que leyó
    Gemini del documento original — no tiene nada en común con el flujo de
    partes digitales, salvo que ambos escriben sobre las mismas tablas de
    base de datos. Los WorkOrder DIGITAL/GENERATED quedan bloqueados sin
    excepción tanto en get() como en post(): se redirigen a
    operator_form_edit, la única vista que puede mostrar o editar un parte
    digital, sea cual sea el rol. El acceso a esta vista está restringido
    a SUPERVISOR/WORKSHOPBOSS/ADMIN por SupervisorAccessMixin — el rol
    WORKSHOP nunca llega a get()/post() aquí.
    El acceso está restringido a la empresa autenticada (guardia multiempresa).
    """

    template_name ="panel/work_orders/edit.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _get_work_order (self ,pk ,company ):
        """
        Retrieves a WorkOrder scoped to the given company.
        Raises WorkOrder.DoesNotExist if not found or belongs to another company.
        ---
        Recupera un WorkOrder acotado a la empresa dada.
        Lanza WorkOrder.DoesNotExist si no se encuentra o pertenece a otra empresa.
        """
        return WorkOrder .objects .get (pk =pk ,company =company )

    def _build_groups (self ,work_order ):
        """
        Returns a list of dicts grouping WorkOrderEntryLine records by their
        parent WorkOrderEntry, ordered by page number.
        Each dict contains:
          - entry  : WorkOrderEntry instance.
          - lines  : list of WorkOrderEntryLine instances for that entry.

        ---

        Devuelve una lista de dicts que agrupan los registros WorkOrderEntryLine
        por su WorkOrderEntry padre, ordenados por número de página.
        Cada dict contiene:
          - entry  : instancia de WorkOrderEntry.
          - lines  : lista de instancias WorkOrderEntryLine de esa entrada.
        """
        entries =(
        WorkOrderEntry .objects 
        .filter (work_order =work_order )
        .prefetch_related ("lines__machine_asset")
        .order_by ("page_number")
        )
        groups =[]
        for entry in entries :
            lines =list (entry .lines .order_by ("line_number"))








            day_total_raw =sum (
            (l .delta_hours for l in lines if l .delta_hours is not None ),
            0 ,
            )
            day_total =round (day_total_raw ,2 )if any (
            l .delta_hours is not None for l in lines 
            )else None 














            if day_total is None :
                day_css =""
            elif day_total <8 :
                day_css ="day-total-short"
            elif day_total <=12 :
                day_css ="day-total-normal"
            elif day_total <=16 :
                day_css ="day-total-warning"
            else :
                day_css ="day-total-danger"

            groups .append ({
            "entry":entry ,
            "lines":lines ,
            "day_total":day_total ,
            "day_css":day_css ,
            })
        return groups 

    def get (self ,request ,pk ):
        """
        Renders the inline edit table for the given WorkOrder.
        Hard-blocks DIGITAL/GENERATED WorkOrders — see class docstring.
        ---
        Renderiza la tabla de edición inline para el WorkOrder dado.
        Bloquea sin excepción los WorkOrder DIGITAL/GENERATED — ver
        docstring de la clase.
        """
        from work_order_processor .models import WorkOrderEntry 
        from django .urls import reverse as _reverse_get 
        company_user =request .user .company_user 
        company =company_user .company 

        try :
            work_order =self ._get_work_order (pk ,company )
        except WorkOrder .DoesNotExist :
            django_messages .error (request ,"Parte de trabajo no encontrado.")
            return redirect (_reverse_get ("panel:work_order_list"))

        # Guardia absoluta: un parte DIGITAL/GENERATED jamás se ve ni se
        # edita desde esta vista, sea cual sea el rol o el enlace de origen.
        # Única vía permitida: operator_form_edit.
        if work_order .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            django_messages .error (
            request ,
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )
            return redirect (_reverse_get ("panel:operator_form_edit",kwargs ={"wo_pk":pk }))

        # Guard: block editing if work order belongs to a locked period.
        # Guardia: bloquear edición si el parte pertenece a un periodo liquidado.
        _first_entry_g =work_order .entries .order_by ("work_date").first ()
        if _first_entry_g and _first_entry_g .work_date :
            from ivr_config .models import WorkPeriod as _WP_G
            _locked_g =_WP_G .objects .filter (
                company_user =work_order .uploaded_by ,
                is_closed =True ,
                start_date__lte =_first_entry_g .work_date ,
                end_date__gte =_first_entry_g .work_date ,
            ).exists ()
            if _locked_g :
                django_messages .error (
                    request ,
                    "Este parte pertenece a un periodo ya liquidado y "
                    "no puede editarse.",
                )
                _back_g =request .GET .get ("back","")
                if _back_g :
                    from urllib.parse import unquote as _uq_g
                    return redirect (_uq_g (_back_g ))
                return redirect (_reverse_get ("panel:work_order_list"))

        groups =self ._build_groups (work_order )

        # Prefer explicit back= param (carries active filters).
        # Si viene back= explícito (con filtros activos) usarlo directamente.
        _back_param =request .GET .get ("back","").strip ()
        if _back_param :
            from urllib.parse import unquote as _unquote
            back_url =_unquote (_back_param )
        else :
            from_param =request .GET .get ("from","")
            if from_param =="taller":
                back_url =_reverse_get ("panel:work_order_admin_history")+"?tab=pending"
            else :
                back_url =_reverse_get ("panel:work_order_list")

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":company_user ,
        "own_presence":self ._get_own_presence (company_user ),
        "active_nav":"work_orders",
        "work_order":work_order ,
        "groups":groups ,
        "back_url":back_url ,
        })

    def post (self ,request ,pk ):
        """
        Dispatches POST actions: save_line or regenerate.
        Hard-blocks DIGITAL/GENERATED WorkOrders — see class docstring.
        ---
        Despacha las acciones POST: save_line o regenerate.
        Bloquea sin excepción los WorkOrder DIGITAL/GENERATED — ver
        docstring de la clase.
        """
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        _compute_delta_hours ,
        _normalise_machine_code ,
        _resolve_machine_asset ,
        )
        from datetime import time as dt_time 
        from django .urls import reverse as _reverse_post 
        import json 

        company_user =request .user .company_user 
        company =company_user .company 

        try :
            work_order =self ._get_work_order (pk ,company )
        except WorkOrder .DoesNotExist :
            django_messages .error (request ,"Parte de trabajo no encontrado.")
            return redirect (_reverse_post ("panel:work_order_list"))

        # Guardia absoluta: un parte DIGITAL/GENERATED jamás se ve ni se
        # edita desde esta vista, sea cual sea el rol o el enlace de origen.
        # Única vía permitida: operator_form_edit.
        if work_order .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            django_messages .error (
            request ,
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )
            return redirect (_reverse_post ("panel:operator_form_edit",kwargs ={"wo_pk":pk }))

        # Use back_url hidden field if present (carries active filters).
        # Usar el campo hidden back_url si viene (preserva filtros activos).
        _back_hidden =request .POST .get ("back_url","").strip ()
        if _back_hidden :
            from urllib.parse import unquote as _unquote_post
            _list_url =_unquote_post (_back_hidden )
        else :
            _list_url =_reverse_post ("panel:work_order_list")

        action =request .POST .get ("action","")





        if action =="regenerate":
            work_order .status =WorkOrder .Status .PENDING 
            work_order .excel_file =None 
            work_order .error_log =""
            work_order .save (update_fields =["status","excel_file","error_log"])
            from work_order_processor .services import generate_work_order_excel 
            from work_order_processor .tasks import process_work_order_pdf 
            generate_work_order_excel (work_order .pk )
            django_messages .success (
            request ,
            f"Excel regenerado correctamente para el Parte #{work_order.pk}."
            )
            return redirect (_list_url )





        if action =="save_line":
            line_pk =request .POST .get ("line_pk")
            try :
                line =WorkOrderEntryLine .objects .select_related (
                "entry__work_order"
                ).get (pk =line_pk ,entry__work_order =work_order )
            except WorkOrderEntryLine .DoesNotExist :
                django_messages .error (request ,"Línea no encontrada.")
                return redirect ("panel:work_order_edit",pk =pk )



            raw_norm =request .POST .get ("machine_norm","").strip ()
            norm =_normalise_machine_code (raw_norm )if raw_norm else raw_norm 
            asset =_resolve_machine_asset (norm ,company =company )if norm else None 



            def _parse_time_str (val ):
                """Parses HH:MM string into time, returns None on failure.
                --- Parsea cadena HH:MM a time, devuelve None si falla."""
                if not val :
                    return None 
                try :
                    parts =val .strip ().split (":")
                    return dt_time (int (parts [0 ]),int (parts [1 ]))
                except (ValueError ,IndexError ):
                    return None 

            hc =_parse_time_str (request .POST .get ("hc",""))
            hf =_parse_time_str (request .POST .get ("hf",""))
            delta =_compute_delta_hours (hc ,hf )



            flags_raw =request .POST .get ("flags","").strip ()
            flags =[f .strip ()for f in flags_raw .split (",")if f .strip ()]if flags_raw else []



            line .machine_norm =norm 
            line .machine_asset =asset 
            line .fault_description =request .POST .get ("fault_description","").strip ()
            line .repair_notes =request .POST .get ("repair_notes","").strip ()
            line .hc =hc 
            line .hf =hf 
            line .or_val =request .POST .get ("or_val","").strip ()
            line .delta_hours =delta 
            line .flags =flags 
            line .save (update_fields =[
            "machine_norm","machine_asset","fault_description",
            "repair_notes","hc","hf","or_val","delta_hours","flags",
            ])

            django_messages .success (
            request ,
            f"Bloque {line.line_number} de la página "
            f"{line.entry.page_number} guardado correctamente."
            )
            return redirect ("panel:work_order_edit",pk =pk )



        django_messages .warning (request ,"Acción no reconocida.")
        return redirect ("panel:work_order_edit",pk =pk )



class WorkOrderStatusFragmentView (SupervisorAccessMixin ,View ):
    """
    Returns the HTML status fragment for a single WorkOrder, used by HTMX polling.

    GET /panel/work-orders/<pk>/status/
        Renders only the _status_fragment.html partial for the WorkOrder identified
        by pk and scoped to the authenticated user's company. If the status is
        PENDING or PROCESSING the partial includes HTMX polling attributes so the
        browser continues polling every 4 seconds. When the status reaches DONE or
        ERROR the returned fragment contains no HTMX attributes and polling stops
        automatically.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Devuelve el fragmento HTML de estado de un WorkOrder, consumido por el polling HTMX.

    GET /panel/work-orders/<pk>/status/
        Renderiza únicamente el parcial _status_fragment.html para el WorkOrder
        identificado por pk, acotado a la empresa del usuario autenticado. Si el
        estado es PENDING o PROCESSING el parcial incluye atributos HTMX de polling
        para que el navegador siga consultando cada 4 segundos. Cuando el estado
        alcanza DONE o ERROR el fragmento devuelto no contiene atributos HTMX y el
        polling se detiene automáticamente.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def get (self ,request ,pk ):
        """
        Returns the rendered _status_fragment.html partial for the requested WorkOrder.
        Raises HTTP 404 if the WorkOrder does not exist or belongs to another company.
        ---
        Devuelve el parcial _status_fragment.html renderizado para el WorkOrder solicitado.
        Lanza HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.
        """
        from django .shortcuts import get_object_or_404 

        wo =get_object_or_404 (
        WorkOrder ,
        pk =pk ,
        company =request .user .company_user .company ,
        )
        return render (
        request ,
        "panel/work_orders/_status_fragment.html",
        {"wo":wo },
        )


class WorkOrderLineSaveView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that saves a single WorkOrderEntryLine and returns the updated
    <tr> row as an HTML fragment consumed by the inline editor.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/save/
         Expected POST fields (all optional — missing fields are treated as empty):
           machine_norm        : str  — normalised machine code.
           fault_description  : str  — fault description.
           repair_notes          : str  — repair description.
           hc                  : str  — start time  HH:MM.
           hf                  : str  — end time    HH:MM.
           or_val              : str  — repair order reference.
           flags               : str  — comma-separated flag list.
         Server recomputes delta_hours from hc/hf and re-resolves machine_asset
         from the updated machine_norm. Returns the rendered _line_row.html partial
         (a single <tr> element) with HTTP 200.
         Returns HTTP 404 if the WorkOrder or line do not exist or belong to another
         company. Returns HTTP 400 on an unexpected processing error.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Endpoint HTMX que guarda un único WorkOrderEntryLine y devuelve la fila <tr>
    actualizada como fragmento HTML consumido por el editor inline.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/save/
         Campos POST esperados (todos opcionales — los ausentes se tratan como vacíos):
           machine_norm        : str  — código de máquina normalizado.
           fault_description  : str  — descripción de la avería.
           repair_notes          : str  — descripción de la reparación.
           hc                  : str  — hora de comienzo HH:MM.
           hf                  : str  — hora de fin      HH:MM.
           or_val              : str  — referencia de orden de reparación.
           flags               : str  — lista de flags separada por comas.
         El servidor recalcula delta_hours desde hc/hf y re-resuelve machine_asset
         desde el machine_norm actualizado. Devuelve el parcial _line_row.html
         renderizado (un único elemento <tr>) con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder o la línea no existen o pertenecen a
         otra empresa. Devuelve HTTP 400 ante un error de procesamiento inesperado.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Saves the WorkOrderEntryLine identified by line_pk, recomputes derived
        fields and returns the updated <tr> row as an HTMX fragment.
        ---
        Guarda el WorkOrderEntryLine identificado por line_pk, recalcula los campos
        derivados y devuelve la fila <tr> actualizada como fragmento HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from work_order_processor .models import WorkOrderEntryLine 
        from work_order_processor .services import (
        _compute_delta_hours ,
        _normalise_machine_code ,
        _resolve_machine_asset ,
        )
        from datetime import time as dt_time 





        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )





        line =get_object_or_404 (
        WorkOrderEntryLine .objects .select_related (
        "entry__work_order",
        "machine_asset",
        ),
        pk =line_pk ,
        entry__work_order =wo ,
        )





        raw_norm =request .POST .get ("machine_norm","").strip ()
        norm =_normalise_machine_code (raw_norm )if raw_norm else raw_norm 
        asset =_resolve_machine_asset (norm ,company =company )if norm else None 





        def _parse_time_str (val ):
            """Parses HH:MM string into time object, returns None on failure.
            --- Parsea cadena HH:MM a objeto time, devuelve None si falla."""
            if not val :
                return None 
            try :
                parts =val .strip ().split (":")
                return dt_time (int (parts [0 ]),int (parts [1 ]))
            except (ValueError ,IndexError ):
                return None 

        hc =_parse_time_str (request .POST .get ("hc",""))
        hf =_parse_time_str (request .POST .get ("hf",""))
        delta =_compute_delta_hours (hc ,hf )





        flags_raw =request .POST .get ("flags","").strip ()
        flags =[f .strip ()for f in flags_raw .split (",")if f .strip ()]if flags_raw else []





        line .machine_norm =norm 
        line .machine_asset =asset 
        line .fault_description =request .POST .get ("fault_description","").strip ()
        line .repair_notes =request .POST .get ("repair_notes","").strip ()
        line .hc =hc 
        line .hf =hf 
        line .or_val =request .POST .get ("or_val","").strip ()
        line .delta_hours =delta 
        line .flags =flags 
        line .save (update_fields =[
        "machine_norm","machine_asset","fault_description",
        "repair_notes","hc","hf","or_val","delta_hours","flags",
        ])





        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":line ,
        "wo_pk":wo .pk ,
        "entry":line .entry ,
        },
        )




class WorkOrderEntrySaveDateView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that saves the work_date of a WorkOrderEntry and returns
    the updated group header fragment so the inline editor reflects the change
    without a full-page reload.

    POST /panel/work-orders/<wo_pk>/entries/<entry_pk>/save-date/
         Expected POST fields:
           work_date : str — date in YYYY-MM-DD format. Empty string clears
                             the date and keeps uncertain_date unchanged.
         On a valid non-empty date, uncertain_date is automatically set to
         False — the supervisor's explicit correction removes the uncertainty.
         Returns the rendered _entry_group_fragment.html partial with HTTP 200.
         Returns HTTP 400 on an invalid date format.
         Returns HTTP 404 if the WorkOrder or entry do not exist or belong to
         another company.

    ---

    Endpoint HTMX que guarda el work_date de un WorkOrderEntry y devuelve
    el fragmento de cabecera del grupo actualizado para que el editor inline
    refleje el cambio sin recarga completa de página.

    POST /panel/work-orders/<wo_pk>/entries/<entry_pk>/save-date/
         Campos POST esperados:
           work_date : str — fecha en formato YYYY-MM-DD. Cadena vacía limpia
                             la fecha y deja uncertain_date sin cambios.
         Con una fecha válida no vacía, uncertain_date se pone automáticamente
         a False — la corrección explícita del supervisor elimina la incertidumbre.
         Devuelve el parcial _entry_group_fragment.html renderizado con HTTP 200.
         Devuelve HTTP 400 ante un formato de fecha inválido.
         Devuelve HTTP 404 si el WorkOrder o entry no existen o son de otra empresa.
    """

    def post (self ,request ,wo_pk ,entry_pk ):
        """
        Persists the corrected work_date on the WorkOrderEntry and returns
        the group header fragment as an HTMX response.
        ---
        Persiste el work_date corregido en el WorkOrderEntry y devuelve el
        fragmento de cabecera del grupo como respuesta HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 
        from work_order_processor .models import WorkOrderEntry 
        from datetime import date as dt_date 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )
        entry =get_object_or_404 (WorkOrderEntry ,pk =entry_pk ,work_order =wo )

        raw_date =request .POST .get ("work_date","").strip ()

        if raw_date :


            try :
                parsed =dt_date .fromisoformat (raw_date )
            except ValueError :
                return HttpResponse (
                "Formato de fecha inválido. Use YYYY-MM-DD.",
                status =400 ,
                )
            entry .work_date =parsed 


            entry .uncertain_date =False 
        else :
            entry .work_date =None 



        entry .save (update_fields =["work_date","uncertain_date"])





        from decimal import Decimal 
        lines =list (
        entry .lines .select_related ("machine_asset").order_by ("line_number")
        )
        delta_sum =sum (
        (ln .delta_hours for ln in lines if ln .delta_hours is not None ),
        Decimal ("0"),
        )
        if delta_sum >0 :
            total_hours =float (delta_sum )
            if total_hours <8 :
                day_css ="day-total-short"
            elif total_hours <=12 :
                day_css ="day-total-normal"
            elif total_hours <=16 :
                day_css ="day-total-warning"
            else :
                day_css ="day-total-danger"
            day_total =round (total_hours ,2 )
        else :
            day_total =None 
            day_css =""

        group ={
        "entry":entry ,
        "lines":lines ,
        "day_total":day_total ,
        "day_css":day_css ,
        }

        return render (
        request ,
        "panel/work_orders/_entry_group_fragment.html",
        {
        "group":group ,
        "wo_pk":wo .pk ,
        },
        )


class MachineAssetAutocompleteView (SupervisorAccessMixin ,View ):
    """
    Lightweight JSON autocomplete endpoint for MachineAsset records.
    Returns up to 10 active assets matching the query string against
    code, brand_model and plate fields (case-insensitive icontains),
    scoped to the authenticated user's company.

    GET /panel/fleet/autocomplete/?q=<query>
        Returns:
          {"results": [{"code": "...", "label": "..."}, ...]}
        where label is "code — brand_model" or just "code" when brand_model
        is empty. Results are ordered by code ascending.

    Used by the machine_norm field autocomplete in _line_row.html.

    ---

    Endpoint JSON ligero de autocompletado para registros MachineAsset.
    Devuelve hasta 10 activos activos que coincidan con la cadena de consulta
    sobre los campos code, brand_model y plate (icontains sin distinción de
    mayúsculas), acotado a la empresa del usuario autenticado.

    GET /panel/fleet/autocomplete/?q=<query>
        Devuelve:
          {"results": [{"code": "...", "label": "..."}, ...]}
        donde label es "code — brand_model" o solo "code" cuando brand_model
        está vacío. Resultados ordenados por code ascendente.

    Usado por el autocompletado del campo machine_norm en _line_row.html.
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns a JSON list of matching MachineAsset records for the given query.
        ---
        Devuelve un JSON con los MachineAsset coincidentes para la consulta dada.
        """
        from django .http import JsonResponse 
        from fleet .models import MachineAsset 

        company =request .user .company_user .company 
        q =request .GET .get ("q","").strip ()

        qs =MachineAsset .objects .filter (
        company =company ,
        is_active =True ,
        ).order_by ("code")

        if q :
            qs =qs .filter (
            Q (code__icontains =q )
            |Q (brand_model__icontains =q )
            |Q (plate__icontains =q )
            )

        results =[]
        for asset in qs [:10 ]:
            label =asset .code 
            if asset .brand_model :
                label =f"{asset.code} — {asset.brand_model}"
            results .append ({"code":asset .code ,"label":label })

        return JsonResponse ({"results":results })



class WorkOrderLineInsertView (SupervisorAccessMixin ,View ):
    """
    Creates a new empty WorkOrderEntryLine after a given line within the same
    WorkOrderEntry group and returns the new row as an HTMX fragment.

    POST /panel/work-orders/<wo_pk>/lines/insert/
         Expected POST fields:
           after_line_pk : int — pk of the line after which to insert.
           entry_pk      : int — pk of the WorkOrderEntry group.
         Creates a new WorkOrderEntryLine with line_number = after_line.line_number + 1,
         shifting the line_number of all subsequent lines in the same entry up by 1.
         Returns the rendered _line_row.html partial for the new line with HTTP 200.
         Returns HTTP 404 if the WorkOrder, entry or reference line do not belong
         to the authenticated company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Crea un nuevo WorkOrderEntryLine vacío tras una línea dada dentro del mismo
    grupo WorkOrderEntry y devuelve la nueva fila como fragmento HTMX.

    POST /panel/work-orders/<wo_pk>/lines/insert/
         Campos POST esperados:
           after_line_pk : int — pk de la línea tras la que insertar.
           entry_pk      : int — pk del grupo WorkOrderEntry.
         Crea un nuevo WorkOrderEntryLine con line_number = after_line.line_number + 1,
         desplazando el line_number de todas las líneas posteriores del mismo entry
         en +1. Devuelve el parcial _line_row.html renderizado para la nueva línea
         con HTTP 200. Devuelve HTTP 404 si el WorkOrder, entry o línea de referencia
         no pertenecen a la empresa autenticada.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ):
        """
        Inserts a new empty WorkOrderEntryLine after the specified reference line
        and returns the rendered _line_row.html fragment for the new row.
        ---
        Inserta un nuevo WorkOrderEntryLine vacío tras la línea de referencia
        especificada y devuelve el fragmento _line_row.html renderizado para la fila.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponseBadRequest 

        company =request .user .company_user .company 



        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )



        try :
            after_line_pk =int (request .POST .get ("after_line_pk",""))
            entry_pk =int (request .POST .get ("entry_pk",""))
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [INSERT] Parámetros after_line_pk / entry_pk inválidos.")



        after_line =get_object_or_404 (
        WorkOrderEntryLine ,
        pk =after_line_pk ,
        entry__work_order =wo ,
        entry__pk =entry_pk ,
        )
        entry =after_line .entry 







        from django .db import transaction 
        with transaction .atomic ():

            # Shift line_number of subsequent lines DESCENDING (highest
            # first), one row at a time -- a bulk .update(F()+1) races
            # against the (entry_id, line_number) unique constraint under
            # MySQL, which does not guarantee row processing order for a
            # mass UPDATE. Processing highest-to-lowest guarantees each
            # target slot is already vacated before the next row claims
            # it. Diagnosed via error.log (2026-07-06):
            # IntegrityError Duplicate entry '1299-3' for key
            # ..._entry_id_line_number_..._uniq.
            # ---
            # Desplazar line_number de las lineas posteriores en orden
            # DESCENDENTE (la mayor primero), fila a fila -- un
            # .update(F()+1) en bloque compite contra la constraint unica
            # (entry_id, line_number) bajo MySQL, que no garantiza el
            # orden de procesamiento de filas en un UPDATE masivo.
            # Procesar de mayor a menor garantiza que cada hueco destino
            # ya esta libre antes de que la siguiente fila lo reclame.
            # Diagnosticado via error.log (2026-07-06): IntegrityError
            # Duplicate entry '1299-3' para la clave
            # ..._entry_id_line_number_..._uniq.
            lines_to_shift =list (
            WorkOrderEntryLine .objects .filter (
            entry =entry ,
            line_number__gt =after_line .line_number ,
            ).order_by ("-line_number")
            )
            for line_to_shift in lines_to_shift :
                line_to_shift .line_number =line_to_shift .line_number +1 
                line_to_shift .save (update_fields =["line_number"])



            new_line =WorkOrderEntryLine .objects .create (
            entry =entry ,
            line_number =after_line .line_number +1 ,
            machine_norm ="",
            machine_raw ="",
            fault_description ="",
            repair_notes ="",
            hc =None ,
            hf =None ,
            or_val ="",
            delta_hours =None ,
            flags =[],
            machine_asset =None ,
            )

        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":new_line ,
        "wo_pk":wo .pk ,
        "entry":entry ,
        },
        )


class WorkOrderEntryAddView (SupervisorAccessMixin ,View ):
    """
    Creates a new empty WorkOrderEntry (day) for a given WorkOrder and
    returns the rendered _entry_group_fragment.html for the new group.

    POST /panel/work-orders/<wo_pk>/entries/add/
         Expected POST fields:
           work_date : str — date in YYYY-MM-DD format.
         Creates a new WorkOrderEntry with the given work_date and one
         empty WorkOrderEntryLine (line_number=1). Returns the rendered
         _entry_group_fragment.html partial for the new group.
         Returns HTTP 400 if work_date is missing or invalid.
         Returns HTTP 404 if the WorkOrder does not belong to the company.
    ---
    Crea un WorkOrderEntry nuevo (dia) para un WorkOrder dado y devuelve
    el fragmento _entry_group_fragment.html renderizado para el nuevo grupo.

    POST /panel/work-orders/<wo_pk>/entries/add/
         Campos POST esperados:
           work_date : str — fecha en formato YYYY-MM-DD.
         Crea un WorkOrderEntry con la fecha dada y un WorkOrderEntryLine
         vacio (line_number=1). Devuelve el parcial _entry_group_fragment.html
         renderizado para el nuevo grupo.
         Devuelve HTTP 400 si work_date falta o es invalida.
         Devuelve HTTP 404 si el WorkOrder no pertenece a la empresa.
    """

    def post (self ,request ,wo_pk ):
        """
        Creates a new WorkOrderEntry + one empty WorkOrderEntryLine
        and returns the _entry_group_fragment.html partial.
        ---
        Crea un WorkOrderEntry nuevo + una WorkOrderEntryLine vacia
        y devuelve el parcial _entry_group_fragment.html.
        """
        from datetime import datetime as _dt
        from django.shortcuts import get_object_or_404
        from django.http import HttpResponseBadRequest
        from django.db import transaction

        company = request.user.company_user.company
        wo = get_object_or_404(WorkOrder, pk=wo_pk, company=company)
        if wo.source in (WorkOrder.Source.DIGITAL, WorkOrder.Source.GENERATED):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden(
                "Los partes digitales se editan única y exclusivamente desde "
                "su formulario de creación y edición."
            )

        raw_date = request.POST.get("work_date", "").strip()
        if not raw_date:
            return HttpResponseBadRequest(
                "# [ENTRY ADD] work_date es obligatorio."
            )
        try:
            work_date = _dt.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponseBadRequest(
                "# [ENTRY ADD] Formato de fecha invalido. Se esperaba YYYY-MM-DD."
            )

        worker_name = (
            wo.entries.first().worker_name
            if wo.entries.exists()
            else ""
        )

        with transaction.atomic():
            entry = WorkOrderEntry.objects.create(
                work_order=wo,
                page_number=wo.entries.count() + 1,
                worker_name=worker_name,
                work_date=work_date,
                uncertain_date=False,
                extraction_confidence=WorkOrderEntry.Confidence.HIGH,
                raw_gemini_response=None,
            )
            new_line = WorkOrderEntryLine.objects.create(
                entry=entry,
                line_number=1,
                machine_norm="",
                machine_raw="",
                fault_description="",
                repair_notes="",
                hc=None,
                hf=None,
                or_val="",
                delta_hours=None,
                flags=[],
                machine_asset=None,
            )

        logger.info(
            "# [ENTRY ADD] WorkOrderEntry pk=%s (fecha=%s) creado en WO pk=%s.",
            entry.pk, work_date, wo.pk,
        )

        lines = list(entry.lines.order_by("line_number"))
        group = {
            "entry":          entry,
            "lines":          lines,
            "day_total_hours": None,
            "day_total_class": "day-total-short",
        }
        return render(
            request,
            "panel/work_orders/_entry_group_fragment.html",
            {"group": group, "wo_pk": wo.pk},
        )


class WorkOrderLineReorderView (SupervisorAccessMixin ,View ):
    """
    Accepts a new ordering for WorkOrderEntryLine records within a single
    WorkOrderEntry group and persists the updated line_number values.

    POST /panel/work-orders/<wo_pk>/lines/reorder/
         Expected POST fields:
           entry_pk        : int         — pk of the WorkOrderEntry group.
           line_pks[]      : list[int]   — ordered list of WorkOrderEntryLine pks
                                           in the desired new order.
         Updates line_number for each line to match the position in line_pks[].
         Returns HTTP 200 with JSON {"ok": true} on success.
         Returns HTTP 400 on invalid parameters.
         Returns HTTP 404 if the WorkOrder or entry do not belong to the company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Acepta un nuevo orden para los registros WorkOrderEntryLine dentro de un único
    grupo WorkOrderEntry y persiste los valores line_number actualizados.

    POST /panel/work-orders/<wo_pk>/lines/reorder/
         Campos POST esperados:
           entry_pk        : int         — pk del grupo WorkOrderEntry.
           line_pks[]      : list[int]   — lista ordenada de pks de WorkOrderEntryLine
                                           en el nuevo orden deseado.
         Actualiza line_number de cada línea para que coincida con su posición en
         line_pks[]. Devuelve HTTP 200 con JSON {"ok": true} en caso de éxito.
         Devuelve HTTP 400 ante parámetros inválidos. Devuelve HTTP 404 si el
         WorkOrder o entry no pertenecen a la empresa.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ):
        """
        Persists the new line_number ordering for the lines of a WorkOrderEntry.
        ---
        Persiste el nuevo orden de line_number para las líneas de un WorkOrderEntry.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import JsonResponse ,HttpResponseBadRequest 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )



        try :
            entry_pk =int (request .POST .get ("entry_pk",""))
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [REORDER] Parámetro entry_pk inválido.")

        entry =get_object_or_404 (WorkOrderEntry ,pk =entry_pk ,work_order =wo )



        try :
            line_pks =[int (pk )for pk in request .POST .getlist ("line_pks[]")]
        except (TypeError ,ValueError ):
            return HttpResponseBadRequest ("# [REORDER] Parámetro line_pks[] inválido.")

        if not line_pks :
            return HttpResponseBadRequest ("# [REORDER] Lista line_pks[] vacía.")



        lines_map ={
        line .pk :line 
        for line in WorkOrderEntryLine .objects .filter (entry =entry )
        }



        bulk_update =[]
        for position ,pk in enumerate (line_pks ,start =1 ):
            line =lines_map .get (pk )
            if line is None :
                continue 
            line .line_number =position 
            bulk_update .append (line )

        WorkOrderEntryLine .objects .bulk_update (bulk_update ,["line_number"])

        return JsonResponse ({"ok":True })


class WorkOrderLineRestoreView (SupervisorAccessMixin ,View ):
    """
    Restores a single WorkOrderEntryLine to its original Gemini-extracted
    values by looking up the corresponding block in the parent WorkOrderEntry's
    raw_gemini_response using the line's line_number as the index.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/restore/
         Locates the bloque at raw_gemini_response["entradas"][line_number - 1],
         overwrites only the fields of the target line (machine_raw, machine_norm,
         machine_asset, fault_description, repair_notes, hc, hf, or_val,
         delta_hours, flags) and returns the rendered _line_row.html partial
         for that single row with HTTP 200.
         Returns HTTP 404 if the WorkOrder or line do not belong to the company.
         Returns HTTP 400 if no raw_gemini_response is stored or the block index
         is out of range.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Restaura un único WorkOrderEntryLine a sus valores originales extraídos por
    Gemini localizando el bloque correspondiente en el raw_gemini_response del
    WorkOrderEntry padre usando el line_number de la línea como índice.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/restore/
         Localiza el bloque en raw_gemini_response["entradas"][line_number - 1],
         sobreescribe únicamente los campos de la línea objetivo (machine_raw,
         machine_norm, machine_asset, fault_description, repair_notes, hc, hf,
         or_val, delta_hours, flags) y devuelve el parcial _line_row.html
         renderizado para esa única fila con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder o la línea no pertenecen a la empresa.
         Devuelve HTTP 400 si no hay raw_gemini_response almacenado o el índice
         de bloque está fuera de rango.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Restores the single WorkOrderEntryLine identified by line_pk from its
        corresponding block in raw_gemini_response and returns the updated row
        as an HTMX fragment.
        ---
        Restaura el único WorkOrderEntryLine identificado por line_pk desde su
        bloque correspondiente en raw_gemini_response y devuelve la fila
        actualizada como fragmento HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponseBadRequest 
        from work_order_processor .services import (
        _normalise_machine_code ,
        _resolve_machine_asset ,
        _compute_delta_hours ,
        _parse_time ,
        )

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )



        line =get_object_or_404 (
        WorkOrderEntryLine .objects .select_related ("entry","machine_asset"),
        pk =line_pk ,
        entry__work_order =wo ,
        )
        entry =line .entry 










        raw =entry .raw_gemini_response 

        if not raw or not isinstance (raw ,dict ):


            machine_norm =_normalise_machine_code (line .machine_raw or "")
            machine_asset =_resolve_machine_asset (machine_norm ,company =company )if machine_norm else None 
            delta =_compute_delta_hours (line .hc ,line .hf ,deduct_lunch =False )

            line .machine_norm =machine_norm 
            line .machine_asset =machine_asset 
            line .delta_hours =delta 
            line .save (update_fields =["machine_norm","machine_asset","delta_hours"])

            return render (
            request ,
            "panel/work_orders/_line_row.html",
            {
            "line":line ,
            "wo_pk":wo .pk ,
            "entry":entry ,
            },
            )



        entradas =raw .get ("entradas")or []
        block_index =line .line_number -1 

        if block_index <0 or block_index >=len (entradas ):
            return HttpResponseBadRequest (
            f"# [RESTORE] Índice de bloque {block_index} fuera de rango "
            f"(entradas disponibles: {len(entradas)})."
            )

        bloque =entradas [block_index ]



        machine_raw =(bloque .get ("machine_raw")or "").strip ()
        machine_norm =_normalise_machine_code (machine_raw )
        machine_asset =_resolve_machine_asset (machine_norm ,company =company )
        hc =_parse_time (bloque .get ("hc"))
        hf =_parse_time (bloque .get ("hf"))
        delta =_compute_delta_hours (hc ,hf )
        flags =bloque .get ("flags")or []
        if not isinstance (flags ,list ):
            flags =[]

        line .machine_raw =machine_raw 
        line .machine_norm =machine_norm 
        line .machine_asset =machine_asset 
        line .fault_description =(bloque .get ("fault_description")or "")
        line .repair_notes =(bloque .get ("repair_notes")or "")
        line .hc =hc 
        line .hf =hf 
        line .or_val =(bloque .get ("or_val")or "")
        line .delta_hours =delta 
        line .flags =flags 
        line .save (update_fields =[
        "machine_raw","machine_norm","machine_asset",
        "fault_description","repair_notes","hc","hf",
        "or_val","delta_hours","flags",
        ])

        return render (
        request ,
        "panel/work_orders/_line_row.html",
        {
        "line":line ,
        "wo_pk":wo .pk ,
        "entry":entry ,
        },
        )


class WorkOrderLineDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a single WorkOrderEntryLine identified by line_pk, scoped to the
    authenticated company. Returns an empty HTTP 200 response so that HTMX
    removes the <tr> row from the DOM via hx-swap="outerHTML" with an empty
    response body.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/delete/
         Returns HTTP 404 if the WorkOrder or line do not exist or belong to
         another company. Restricted to ADMIN role.

    ---

    Elimina un único WorkOrderEntryLine identificado por line_pk, acotado a la
    empresa autenticada. Devuelve una respuesta HTTP 200 vacía para que HTMX
    elimine la fila <tr> del DOM via hx-swap="outerHTML" con cuerpo vacío.

    POST /panel/work-orders/<wo_pk>/lines/<line_pk>/delete/
         Devuelve HTTP 404 si el WorkOrder o la línea no existen o pertenecen a
         otra empresa. Restringido al rol ADMIN.
    """

    def post (self ,request ,wo_pk ,line_pk ):
        """
        Deletes the WorkOrderEntryLine identified by line_pk and returns an
        empty response for HTMX to remove the row from the DOM.
        ---
        Elimina el WorkOrderEntryLine identificado por line_pk y devuelve una
        respuesta vacía para que HTMX elimine la fila del DOM.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =wo_pk ,company =company )
        if wo .source in (WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Los partes digitales se editan única y exclusivamente desde "
            "su formulario de creación y edición."
            )
        line =get_object_or_404 (
        WorkOrderEntryLine ,
        pk =line_pk ,
        entry__work_order =wo ,
        )
        entry =line .entry 
        line .delete ()

        # If the entry has no remaining lines, delete it too.
        # Si el entry ha quedado sin líneas, eliminarlo también.
        if not entry .lines .exists ():
            entry .delete ()

        return HttpResponse ("")


class WorkOrderDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a WorkOrder and all its cascade-deleted children (WorkOrderEntry,
    WorkOrderEntryLine, source PDF file, Excel file) scoped to the authenticated
    company. Redirects to the work order list on success.

    POST /panel/work-orders/<pk>/delete/
         Returns HTTP 404 if the WorkOrder does not exist or belongs to another
         company. Restricted to ADMIN role.

    ---

    Elimina un WorkOrder y todos sus hijos eliminados en cascada (WorkOrderEntry,
    WorkOrderEntryLine, PDF original, archivo Excel) acotado a la empresa
    autenticada. Redirige a la lista de partes tras el éxito.

    POST /panel/work-orders/<pk>/delete/
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.
         Restringido al rol ADMIN.
    """

    def post (self ,request ,pk ):
        """
        Deletes the WorkOrder identified by pk, scoped to the authenticated company.
        ---
        Elimina el WorkOrder identificado por pk, acotado a la empresa autenticada.
        """
        from django .shortcuts import get_object_or_404 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )
        wo_pk =wo .pk 
        wo .delete ()
        django_messages .success (
        request ,
        f"Parte #{wo_pk} eliminado correctamente."
        )
        return redirect ("panel:work_order_list")


class WorkOrderMarkReviewedView (SupervisorAccessMixin ,View ):
    """
    Toggles the reviewed flag on a WorkOrder identified by pk, scoped to the
    authenticated company. Returns an HTML fragment (_review_badge_fragment.html)
    for HTMX to swap inline in list.html.

    POST /panel/work-orders/<pk>/review/
         If reviewed is currently False:
           Sets reviewed=True, reviewed_by=request.user.company_user,
           reviewed_at=now().
         If reviewed is currently True:
           Sets reviewed=False, reviewed_by=None, reviewed_at=None.
         Returns the rendered _review_badge_fragment.html partial with HTTP 200.
         Returns HTTP 404 if the WorkOrder does not exist or belongs to another
         company.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Alterna el flag reviewed de un WorkOrder identificado por pk, acotado a la
    empresa autenticada. Devuelve un fragmento HTML (_review_badge_fragment.html)
    para que HTMX lo intercambie inline en list.html.

    POST /panel/work-orders/<pk>/review/
         Si reviewed es actualmente False:
           Establece reviewed=True, reviewed_by=request.user.company_user,
           reviewed_at=now().
         Si reviewed es actualmente True:
           Establece reviewed=False, reviewed_by=None, reviewed_at=None.
         Devuelve el parcial _review_badge_fragment.html renderizado con HTTP 200.
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ,pk ):
        """
        Toggles reviewed on the WorkOrder identified by pk and returns the
        rendered review badge fragment for HTMX inline swap.
        ---
        Alterna reviewed en el WorkOrder identificado por pk y devuelve el
        fragmento del badge de revisión para el intercambio inline de HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .utils .timezone import now as tz_now 

        company =request .user .company_user .company 
        company_user =request .user .company_user 

        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )

        if wo .reviewed :


            wo .reviewed =False 
            wo .reviewed_by =None 
            wo .reviewed_at =None 
            wo .save (update_fields =["reviewed","reviewed_by","reviewed_at"])
        else :


            wo .reviewed =True 
            wo .reviewed_by =company_user 
            wo .reviewed_at =tz_now ()
            wo .save (update_fields =["reviewed","reviewed_by","reviewed_at"])

        return render (
        request ,
        "panel/work_orders/_review_badge_fragment.html",
        {"wo":wo },
        )


class WorkOrderExportView (SupervisorAccessMixin ,View ):
    """
    Generates and returns a multi-sheet Excel file concatenating the individual
    Excel reports of a selection of WorkOrder records identified by their pks.

    POST /panel/work-orders/export/
         Expected POST fields:
           pks         : list[int] (repeating field "pks") — primary keys of the
                         WorkOrder records to include. Only DONE records belonging
                         to the authenticated company are exported.
           export_mode : str — "single_sheet" or "multi_sheet" (Hito 8 / Bloque H).
                         single_sheet: one sheet with all entries ordered by
                           worker_name then date_key; an opaque header row marks
                           each new worker block.
                         multi_sheet: one sheet per worker with individual header.
         Returns HTTP 400 if no valid pks are provided.
         Returns HTTP 404 if the authenticated user has no CompanyUser profile.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Genera y devuelve un archivo Excel multi-hoja que concatena los informes
    individuales de una selección de registros WorkOrder identificados por sus pks.

    POST /panel/work-orders/export/
         Campos POST esperados:
           pks         : list[int] (campo repetido "pks") — claves primarias de
                         los WorkOrder a incluir. Solo se exportan DONE de la empresa.
           export_mode : str — "single_sheet" o "multi_sheet" (Hito 8 / Bloque H).
                         single_sheet: una sola hoja con todas las entradas ordenadas
                           por operario y fecha; una fila cabecera separadora por
                           cada nuevo operario.
                         multi_sheet: una hoja por operario con su membrete.
         Devuelve HTTP 400 si no se proporcionan pks válidos.
         Devuelve HTTP 404 si el usuario no tiene perfil CompanyUser.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ):
        """
        Builds the export Excel file from the selected WorkOrder pks.
        Supports two export modes controlled by the POST field export_mode:

          single_sheet — One flat sheet with all WorkOrderEntryLine records
            from all selected WorkOrders, grouped first by worker_name then
            by work_date. A dark-blue separator row bearing the worker name
            is inserted before the first block of each new worker. All data
            is read directly from the DB — no Excel regeneration is triggered.

          multi_sheet  — One sheet per distinct worker_name. Each sheet is
            built by copying the first sheet of the individual Excel stored
            in WorkOrder.excel_file (regenerated on-the-fly if missing).
            Sheet title is truncated to 31 characters (Excel limit).

        Both modes return an HttpResponse with Content-Disposition attachment.
        Returns HTTP 400 on invalid or missing pks or on unknown export_mode.

        ---

        Construye el Excel de exportación a partir de los pks de WorkOrder
        seleccionados. Soporta dos modos controlados por el campo POST
        export_mode:

          single_sheet — Una hoja plana con todos los WorkOrderEntryLine de
            todos los WorkOrders seleccionados, agrupados por worker_name y
            después por work_date. Una fila separadora azul oscuro con el
            nombre del operario se inserta antes del primer bloque de cada
            nuevo operario. Los datos se leen directamente de la BD — no se
            regenera ningún Excel individual.

          multi_sheet  — Una hoja por worker_name distinto. Cada hoja se
            construye copiando la primera hoja del Excel individual almacenado
            en WorkOrder.excel_file (regenerado al vuelo si falta).
            El título de hoja se trunca a 31 caracteres (límite de Excel).

        Ambos modos devuelven HttpResponse con Content-Disposition attachment.
        Devuelve HTTP 400 ante pks inválidos/ausentes o export_mode desconocido.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        generate_work_order_excel as _gen_excel ,
        _worker_name_from_pdf_path ,
        )

        company =request .user .company_user .company 
        export_mode =request .POST .get ("export_mode","single_sheet").strip ()





        if export_mode not in ("single_sheet","multi_sheet"):
            return HttpResponseBadRequest (
            f"# [EXPORT] Modo de exportación desconocido: {export_mode!r}."
            )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (pk )for pk in raw_pks if pk ]
        except (ValueError ,TypeError ):
            return HttpResponseBadRequest ("# [EXPORT] Parámetros pks inválidos.")

        if not pk_list :
            return HttpResponseBadRequest (
            "# [EXPORT] No se han seleccionado partes para exportar."
            )







        work_orders =list (
        WorkOrder .objects 
        .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        )
        .order_by ("pk")
        )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [EXPORT] Ninguno de los partes seleccionados está revisado y en estado DONE. "
            "Solo se pueden exportar partes marcados como revisados."
            )





        def _copy_sheet (src_sheet ,dest_wb ,title ):
            """
            Copies src_sheet (cells, styles, column widths, row heights) into
            a new sheet named title in dest_wb.
            ---
            Copia src_sheet (celdas, estilos, anchos de columna, alturas de
            fila) en una nueva hoja llamada title en dest_wb.
            """
            dest_sheet =dest_wb .create_sheet (title =title [:31 ])
            for row in src_sheet .iter_rows ():
                for cell in row :
                    dest_cell =dest_sheet .cell (
                    row =cell .row ,column =cell .column ,value =cell .value 
                    )
                    if cell .has_style :
                        dest_cell .font =cell .font .copy ()
                        dest_cell .fill =cell .fill .copy ()
                        dest_cell .alignment =cell .alignment .copy ()
                        dest_cell .border =cell .border .copy ()
                        dest_cell .number_format =cell .number_format 
            for col_letter ,col_dim in src_sheet .column_dimensions .items ():
                dest_sheet .column_dimensions [col_letter ].width =col_dim .width 
            for row_num ,row_dim in src_sheet .row_dimensions .items ():
                dest_sheet .row_dimensions [row_num ].height =row_dim .height 
            return dest_sheet 





        def _get_worker_name (wo ):
            """
            Returns the worker name from the PDF filename, or a fallback label.
            ---
            Devuelve el nombre del operario del nombre del PDF, o etiqueta de
            reserva.
            """
            if wo .source_pdf :
                return _worker_name_from_pdf_path (wo .source_pdf .name )
            return f"Operario #{wo.pk}"












        if export_mode =="single_sheet":





            wo_map ={wo .pk :wo for wo in work_orders }

            lines_qs =(
            WorkOrderEntryLine .objects 
            .filter (entry__work_order__in =work_orders )
            .select_related (
            "entry",
            "entry__work_order",
            "machine_asset",
            )
            .order_by (
            "entry__work_order__pk",
            "entry__work_date",
            "entry__page_number",
            "line_number",
            )
            )





            enriched =[]
            for line in lines_qs :
                entry =line .entry 
                wo =entry .work_order 
                worker_name =_get_worker_name (wo )
                date_key =entry .work_date .isoformat ()if entry .work_date else ""
                enriched .append ({
                "worker_name":worker_name ,
                "date_key":date_key ,
                "line":line ,
                "entry":entry ,
                })



            enriched .sort (key =lambda r :(
            r ["worker_name"],
            r ["date_key"]if r ["date_key"]else "",
            ))





            wb =openpyxl .Workbook ()
            ws =wb .active 
            ws .title ="EXPORTACION"



            _SEP_BG ="1F4E79"
            _SEP_FG ="FFFFFF"

            def _make_sep_fill ():
                return PatternFill (
                fill_type ="solid",
                start_color =_SEP_BG ,
                end_color =_SEP_BG ,
                )


            headers =[
            "OPERARIO","FECHA","BLOQUE","MÁQUINA (NORM)",
            "MÁQUINA (RAW)","ACTIVO RESUELTO","DESCRIPCIÓN AVERÍA",
            "REPARACIÓN","H.C.","H.F.","Δ HORAS","O.R.","FLAGS",
            ]
            NUM_COLS =len (headers )

            for col_idx ,hdr in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =hdr )
                cell .font =Font (bold =True ,color =_SEP_FG )
                cell .fill =_make_sep_fill ()
                cell .alignment =Alignment (horizontal ="center",vertical ="center")
            ws .row_dimensions [1 ].height =20 

            current_worker =None 
            data_row =2 

            for rec in enriched :
                worker_name =rec ["worker_name"]
                line =rec ["line"]
                entry =rec ["entry"]



                if worker_name !=current_worker :
                    ws .merge_cells (
                    start_row =data_row ,start_column =1 ,
                    end_row =data_row ,end_column =NUM_COLS ,
                    )
                    sep_cell =ws .cell (row =data_row ,column =1 ,
                    value =worker_name )
                    sep_cell .font =Font (bold =True ,color =_SEP_FG ,size =11 )
                    sep_cell .fill =_make_sep_fill ()
                    sep_cell .alignment =Alignment (horizontal ="left",
                    vertical ="center")
                    ws .row_dimensions [data_row ].height =22 
                    data_row +=1 
                    current_worker =worker_name 


                date_display =(
                entry .work_date .strftime ("%d/%m/%Y")
                if entry .work_date else ""
                )
                asset_code =(
                line .machine_asset .code if line .machine_asset else ""
                )
                hc_display =(
                line .hc .strftime ("%H:%M")if line .hc else ""
                )
                hf_display =(
                line .hf .strftime ("%H:%M")if line .hf else ""
                )
                delta_display =(
                str (line .delta_hours )if line .delta_hours is not None else ""
                )
                flags_display =", ".join (line .flags )if line .flags else ""

                row_values =[
                worker_name ,
                date_display ,
                line .line_number ,
                line .machine_norm ,
                line .machine_raw ,
                asset_code ,
                line .fault_description ,
                line .repair_notes ,
                hc_display ,
                hf_display ,
                delta_display ,
                line .or_val ,
                flags_display ,
                ]
                for col_idx ,val in enumerate (row_values ,start =1 ):
                    cell =ws .cell (row =data_row ,column =col_idx ,value =val )
                    cell .alignment =Alignment (vertical ="center",wrap_text =False )
                ws .row_dimensions [data_row ].height =16 
                data_row +=1 



            for col_idx ,hdr in enumerate (headers ,start =1 ):
                col_letter =openpyxl .utils .get_column_letter (col_idx )
                ws .column_dimensions [col_letter ].width =min (
                max (len (hdr )+2 ,12 ),60 
                )

            if ws .max_row <2 :
                return HttpResponseBadRequest (
                "# [EXPORT] No hay líneas de datos en los partes seleccionados."
                )










        else :


            from collections import OrderedDict 
            groups :dict [str ,list ]=OrderedDict ()
            for wo in work_orders :
                worker_name =_get_worker_name (wo )
                groups .setdefault (worker_name ,[]).append (wo )

            wb =openpyxl .Workbook ()
            wb .remove (wb .active )

            for worker_name ,wo_list in groups .items ():




                sheet_built =False 
                for wo in wo_list :
                    try :


                        if not wo .excel_file :
                            _gen_excel (wo .pk )
                            wo .refresh_from_db (fields =["excel_file"])
                        if not wo .excel_file :
                            continue 
                        with wo .excel_file .open ("rb")as f :
                            buf =io .BytesIO (f .read ())
                        src_wb =openpyxl .load_workbook (buf )
                        src_sheet =src_wb .worksheets [0 ]


                        sheet_title =(
                        (worker_name [:28 ]+f"#{wo.pk}")
                        if worker_name else f"Parte#{wo.pk}"
                        )
                        _copy_sheet (src_sheet ,wb ,sheet_title )
                        sheet_built =True 
                        break 
                    except Exception :
                        continue 

                if not sheet_built :


                    placeholder =wb .create_sheet (
                    title =(worker_name [:28 ]if worker_name else "Sin datos")[:31 ]
                    )
                    placeholder .cell (row =1 ,column =1 ,
                    value ="No se pudo generar el Excel para este operario.")

            if not wb .worksheets :
                return HttpResponseBadRequest (
                "# [EXPORT] No se pudo generar ninguna hoja para los partes seleccionados."
                )





        output =io .BytesIO ()
        wb .save (output )
        output .seek (0 )

        filename =f"EXPORTACION_{tz_now().strftime('%d-%m-%y')}.xlsx"
        response =HttpResponse (
        output .read (),
        content_type =(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


class WorkOrderDuplicateSearchView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that detects duplicate WorkOrders for the authenticated
    company by querying WorkOrderEntry records sharing identical
    (worker_name, work_date) tuples across distinct WorkOrders.
    Returns a rendered _duplicates_fragment.html partial with the grouped
    results, or an informational fragment when no duplicates are found.

    POST /panel/work-orders/duplicates/search/
         Returns HTTP 200 with the rendered fragment in all cases.
         Returns HTTP 404 if the authenticated user has no CompanyUser profile.

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Endpoint HTMX que detecta WorkOrders duplicados para la empresa autenticada
    consultando registros WorkOrderEntry que comparten tuplas (worker_name,
    work_date) idénticas entre distintos WorkOrders. Devuelve el parcial
    _duplicates_fragment.html renderizado con los resultados agrupados, o un
    fragmento informativo cuando no se detectan duplicados.

    POST /panel/work-orders/duplicates/search/
         Devuelve HTTP 200 con el fragmento renderizado en todos los casos.
         Devuelve HTTP 404 si el usuario autenticado no tiene perfil CompanyUser.

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def post (self ,request ):
        """
        Executes the duplicate detection query scoped to the authenticated
        company and returns the rendered _duplicates_fragment.html partial.

        Detection logic mirrors detect_duplicate_entries management command:
          1. Query (company, worker_name, work_date) groups where more than one
             distinct WorkOrder contributes WorkOrderEntry records.
          2. For each group, fetch all implicated WorkOrders ordered by pk.
          3. Identify keeper (highest pk) and candidates for deletion (lower pks).

        ---

        Ejecuta la consulta de detección de duplicados acotada a la empresa
        autenticada y devuelve el parcial _duplicates_fragment.html renderizado.

        La lógica de detección es equivalente al comando de gestión
        detect_duplicate_entries:
          1. Consultar grupos (empresa, operario, fecha) donde más de un
             WorkOrder distinto aporta registros WorkOrderEntry.
          2. Por cada grupo, obtener todos los WorkOrders implicados ordenados
             por pk.
          3. Identificar el conservado (pk más alto) y los candidatos a
             eliminación (pks inferiores).
        """
        from django .db .models import Count 
        from work_order_processor .models import WorkOrderEntry 

        company =request .user .company_user .company 





        raw_groups =list (
        WorkOrderEntry .objects 
        .filter (work_order__company =company )
        .values ("worker_name","work_date")
        .annotate (wo_count =Count ("work_order_id",distinct =True ))
        .filter (wo_count__gt =1 )
        .order_by ("worker_name","work_date")
        )

        if not raw_groups :


            return render (
            request ,
            "panel/work_orders/_duplicates_fragment.html",
            {"duplicate_groups":[],"no_duplicates":True },
            )





        enriched_groups =[]

        for raw in raw_groups :
            worker_name =raw ["worker_name"]or ""
            work_date =raw ["work_date"]



            implicated =list (
            WorkOrder .objects 
            .filter (
            company =company ,
            entries__worker_name =worker_name ,
            entries__work_date =work_date ,
            )
            .select_related ("uploaded_by__user")
            .distinct ()
            .order_by ("pk")
            )

            if len (implicated )<2 :


                continue 

            enriched_groups .append ({
            "worker_name":worker_name ,
            "work_date":work_date ,
            "work_orders":implicated ,
            "keeper":implicated [-1 ],
            "to_delete":implicated [:-1 ],
            })

        return render (
        request ,
        "panel/work_orders/_duplicates_fragment.html",
        {
        "duplicate_groups":enriched_groups ,
        "no_duplicates":False ,
        "company_user":request .user .company_user ,
        },
        )


class WorkOrderDuplicateDeleteView (SupervisorAccessMixin ,View ):
    """
    HTMX endpoint that deletes a single WorkOrder identified by pk, scoped
    to the authenticated company. Intended for use in the duplicates panel:
    the caller is responsible for ensuring the pk belongs to a duplicate
    group (not the keeper). Returns an empty HTTP 200 response so that HTMX
    removes the corresponding row from the DOM via hx-swap="outerHTML".

    POST /panel/work-orders/duplicates/<pk>/delete/
         Returns HTTP 404 if the WorkOrder does not exist or belongs to
         another company.

    Restricted to the ADMIN role (AdminRoleRequiredMixin).

    ---

    Endpoint HTMX que elimina un único WorkOrder identificado por pk, acotado
    a la empresa autenticada. Diseñado para su uso en el panel de duplicados:
    el llamador es responsable de garantizar que el pk pertenece a un grupo
    duplicado (no al conservado). Devuelve una respuesta HTTP 200 vacía para
    que HTMX elimine la fila correspondiente del DOM via hx-swap="outerHTML".

    POST /panel/work-orders/duplicates/<pk>/delete/
         Devuelve HTTP 404 si el WorkOrder no existe o pertenece a otra empresa.

    Restringido al rol ADMIN (AdminRoleRequiredMixin).
    """

    def post (self ,request ,pk ):
        """
        Deletes the WorkOrder identified by pk, cascade-removing all its
        children (WorkOrderEntry, WorkOrderEntryLine, source PDF, Excel file).
        Returns an empty response body for HTMX outerHTML swap.

        ---

        Elimina el WorkOrder identificado por pk, eliminando en cascada todos
        sus hijos (WorkOrderEntry, WorkOrderEntryLine, PDF original, Excel).
        Devuelve cuerpo vacío para el swap outerHTML de HTMX.
        """
        from django .shortcuts import get_object_or_404 
        from django .http import HttpResponse 

        company =request .user .company_user .company 
        wo =get_object_or_404 (WorkOrder ,pk =pk ,company =company )
        wo .delete ()



        return HttpResponse ("")





class WorkerAbsenceCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a WorkerAbsence record from the absence modal in admin_history.html.
    Receives the form POST from the modal's form action pointing to
    'panel:worker_absence_create'. On success or failure, always redirects
    back to the Absences tab of WorkOrderAdminHistoryView.

    POST /panel/worker-absences/create/
        Body params:
          company_user_pk (int)  — pk of the target CompanyUser (must belong to company).
          absence_type    (str)  — one of the WorkerAbsence.ABSENCE_* constants.
          start_date      (str)  — ISO date YYYY-MM-DD.
          end_date        (str)  — ISO date YYYY-MM-DD.
          notes           (str)  — optional free text.

    ---

    Crea un registro WorkerAbsence desde el modal de ausencias de admin_history.html.
    Recibe el POST del formulario del modal cuya accion apunta a
    'panel:worker_absence_create'. En caso de exito o fallo, siempre redirige
    de vuelta a la pestana Ausencias de WorkOrderAdminHistoryView.

    POST /panel/worker-absences/create/
        Parametros del cuerpo:
          company_user_pk (int)  — pk del CompanyUser objetivo (debe pertenecer a empresa).
          absence_type    (str)  — una de las constantes WorkerAbsence.ABSENCE_*.
          start_date      (str)  — fecha ISO YYYY-MM-DD.
          end_date        (str)  — fecha ISO YYYY-MM-DD.
          notes           (str)  — texto libre opcional.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates the POST data, creates the WorkerAbsence and redirects.
        Performs server-side validation: company scope, absence_type whitelist,
        date ordering (start_date <= end_date).
        ---
        Valida los datos POST, crea el WorkerAbsence y redirige.
        Realiza validacion en el servidor: scope de empresa, lista blanca de
        absence_type, ordenacion de fechas (start_date <= end_date).
        """
        from datetime import datetime 
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            cu_pk =int (request .POST .get ("company_user_pk",""))
            target_cu =CompanyUser .objects .get (pk =cu_pk ,company =company )
        except (ValueError ,TypeError ,CompanyUser .DoesNotExist ):
            django_messages .error (
            request ,
            "Operario no encontrado o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )





        absence_type =request .POST .get ("absence_type","").strip ()
        valid_types ={k for k ,_ in WorkerAbsence .ABSENCE_CHOICES }
        if absence_type not in valid_types :
            django_messages .error (
            request ,
            f"Tipo de ausencia '{absence_type}' no válido.",
            )
            return redirect (ABSENCES_TAB_URL )





        def _parse_iso (value ):
            """Parses YYYY-MM-DD string, returns date or None. / Parsea cadena YYYY-MM-DD."""
            try :
                return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
            except (ValueError ,AttributeError ):
                return None 

        start_date =_parse_iso (request .POST .get ("start_date",""))
        end_date =_parse_iso (request .POST .get ("end_date",""))

        if not start_date or not end_date :
            django_messages .error (
            request ,
            "Las fechas de inicio y fin son obligatorias y deben tener formato YYYY-MM-DD.",
            )
            return redirect (ABSENCES_TAB_URL )

        if start_date >end_date :
            django_messages .error (
            request ,
            "La fecha de inicio no puede ser posterior a la fecha de fin.",
            )
            return redirect (ABSENCES_TAB_URL )

        notes =request .POST .get ("notes","").strip ()





        WorkerAbsence .objects .create (
        company_user =target_cu ,
        absence_type =absence_type ,
        start_date =start_date ,
        end_date =end_date ,
        registered_by =cu ,
        notes =notes ,
        )

        operator_name =(
        target_cu .user .get_full_name ()or target_cu .user .username 
        )
        django_messages .success (
        request ,
        f"Ausencia de {operator_name} registrada correctamente "
        f"({start_date:%d/%m/%Y} – {end_date:%d/%m/%Y}).",
        )
        return redirect (ABSENCES_TAB_URL )


class WorkerAbsenceUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing WorkerAbsence record from the absence modal in
    admin_history.html. Receives the form POST from the edit modal.
    On success or failure, always redirects back to the Absences tab of
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/update/
        Body params:
          absence_type (str) — one of the WorkerAbsence.ABSENCE_* constants.
          start_date   (str) — ISO date YYYY-MM-DD.
          end_date     (str) — ISO date YYYY-MM-DD.
          notes        (str) — optional free text.
    ---
    Actualiza un registro WorkerAbsence existente desde el modal de edición
    de admin_history.html. Recibe el POST del formulario del modal de edición.
    En caso de éxito o fallo, siempre redirige a la pestaña Ausencias de
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/update/
        Parámetros del cuerpo:
          absence_type (str) — una de las constantes WorkerAbsence.ABSENCE_*.
          start_date   (str) — fecha ISO YYYY-MM-DD.
          end_date     (str) — fecha ISO YYYY-MM-DD.
          notes        (str) — texto libre opcional.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates the POST data, updates the WorkerAbsence and redirects.
        Performs server-side validation: company scope, absence_type whitelist,
        date ordering (start_date <= end_date).
        ---
        Valida los datos POST, actualiza el WorkerAbsence y redirige.
        Realiza validacion en servidor: scope de empresa, lista blanca de
        absence_type, ordenacion de fechas (start_date <= end_date).
        """
        from datetime import datetime 
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            absence =WorkerAbsence .objects .get (
            pk =pk ,
            company_user__company =company ,
            )
        except WorkerAbsence .DoesNotExist :
            django_messages .error (
            request ,
            "Ausencia no encontrada o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )





        absence_type =request .POST .get ("absence_type","").strip ()
        valid_types ={k for k ,_ in WorkerAbsence .ABSENCE_CHOICES }
        if absence_type not in valid_types :
            django_messages .error (
            request ,
            f"Tipo de ausencia '{absence_type}' no válido.",
            )
            return redirect (ABSENCES_TAB_URL )





        def _parse_iso (value ):
            """Parses YYYY-MM-DD string, returns date or None. / Parsea cadena YYYY-MM-DD."""
            try :
                return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
            except (ValueError ,AttributeError ):
                return None 

        start_date =_parse_iso (request .POST .get ("start_date",""))
        end_date =_parse_iso (request .POST .get ("end_date",""))

        if not start_date or not end_date :
            django_messages .error (
            request ,
            "Las fechas de inicio y fin son obligatorias y deben tener formato YYYY-MM-DD.",
            )
            return redirect (ABSENCES_TAB_URL )

        if start_date >end_date :
            django_messages .error (
            request ,
            "La fecha de inicio no puede ser posterior a la fecha de fin.",
            )
            return redirect (ABSENCES_TAB_URL )

        notes =request .POST .get ("notes","").strip ()





        absence .absence_type =absence_type 
        absence .start_date =start_date 
        absence .end_date =end_date 
        absence .notes =notes 
        absence .save (update_fields =["absence_type","start_date","end_date","notes"])

        operator_name =(
        absence .company_user .user .get_full_name ()
        or absence .company_user .user .username 
        )
        django_messages .success (
        request ,
        f"Ausencia de {operator_name} actualizada correctamente "
        f"({start_date:%d/%m/%Y} – {end_date:%d/%m/%Y}).",
        )
        return redirect (ABSENCES_TAB_URL )


class WorkerAbsenceDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes a WorkerAbsence record scoped to the authenticated user's company.
    On success or failure, always redirects back to the Absences tab of
    WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/delete/
    ---
    Elimina un registro WorkerAbsence acotado a la empresa del usuario
    autenticado. En caso de éxito o fallo, siempre redirige a la pestaña
    Ausencias de WorkOrderAdminHistoryView.

    POST /panel/worker-absences/<pk>/delete/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Resolves the WorkerAbsence by pk scoped to the company, deletes it
        and redirects back to the Absences tab.
        ---
        Resuelve el WorkerAbsence por pk acotado a la empresa, lo elimina
        y redirige de vuelta a la pestaña Ausencias.
        """
        from django .urls import reverse 
        from ivr_config .models import WorkerAbsence 

        cu =request .user .company_user 
        company =cu .company 
        ABSENCES_TAB_URL =reverse ("panel:work_order_admin_history")+"?tab=absences"





        try :
            absence =WorkerAbsence .objects .select_related (
            "company_user__user"
            ).get (
            pk =pk ,
            company_user__company =company ,
            )
        except WorkerAbsence .DoesNotExist :
            django_messages .error (
            request ,
            "Ausencia no encontrada o no pertenece a esta empresa.",
            )
            return redirect (ABSENCES_TAB_URL )

        operator_name =(
        absence .company_user .user .get_full_name ()
        or absence .company_user .user .username 
        )
        absence .delete ()

        django_messages .success (
        request ,
        f"Ausencia de {operator_name} eliminada correctamente.",
        )
        return redirect (ABSENCES_TAB_URL )








class WorkdayScheduleView (SupervisorAccessMixin ,View ):
    """
    Creates or updates the WorkdaySchedule records for the authenticated
    user's company. Supports multiple named schedules per company.

    GET  /panel/workday-schedule/
         Renders the schedule management page with the list of existing
         schedules and a creation form.

    POST /panel/workday-schedule/   (action=create)
         Creates a new WorkdaySchedule for the company.

    POST /panel/workday-schedule/   (action=update, schedule_pk=<pk>)
         Updates an existing WorkdaySchedule.

    POST /panel/workday-schedule/   (action=delete, schedule_pk=<pk>)
         Deletes a WorkdaySchedule. Clears workday_schedule FK on all
         CompanyUsers assigned to it before deletion.

    ---

    Crea o actualiza los registros WorkdaySchedule de la empresa del usuario
    autenticado. Soporta múltiples horarios con nombre por empresa.

    GET  /panel/workday-schedule/
         Renderiza la página de gestión de horarios con la lista de horarios
         existentes y un formulario de creación.

    POST /panel/workday-schedule/   (action=create)
         Crea un nuevo WorkdaySchedule para la empresa.

    POST /panel/workday-schedule/   (action=update, schedule_pk=<pk>)
         Actualiza un WorkdaySchedule existente.

    POST /panel/workday-schedule/   (action=delete, schedule_pk=<pk>)
         Elimina un WorkdaySchedule. Limpia la FK workday_schedule en todos
         los CompanyUsers asignados a él antes de la eliminación.
    """

    template_name ="panel/workday/schedule_form.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _parse_time (self ,value ):
        """
        Parses an HH:MM time string and returns a time object, or None.
        ---
        Parsea una cadena HH:MM y devuelve un objeto time, o None.
        """
        from datetime import datetime 
        if not value :
            return None 
        try :
            return datetime .strptime (value .strip (),"%H:%M").time ()
        except ValueError :
            return None 

    def _build_context (self ,request ,errors =None ,post_data =None ):
        """
        Builds the template context with the list of existing schedules,
        the creation form data and validation errors.
        ---
        Construye el contexto con la lista de horarios existentes, los datos
        del formulario de creación y los errores de validación.
        """
        from ivr_config .models import WorkdaySchedule 
        cu =request .user .company_user 
        company =cu .company 
        return {
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"workday_schedule",
        "schedules":WorkdaySchedule .objects .filter (
        company =company 
        ).order_by ("label"),
        "errors":errors or [],
        "post_data":post_data or {},
        }

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the workday schedule management page.
        ---
        Renderiza la página de gestión de horarios de jornada.
        """
        return render (request ,self .template_name ,self ._build_context (request ))

    def post (self ,request ,*args ,**kwargs ):
        """
        Dispatches create, update and delete actions for WorkdaySchedule.
        ---
        Despacha las acciones create, update y delete para WorkdaySchedule.
        """
        from django .urls import reverse 
        from ivr_config .models import WorkdaySchedule 

        cu =request .user .company_user 
        company =cu .company 
        action =request .POST .get ("action","").strip ()

        REDIRECT_URL =reverse ("panel:workday_schedule")





        if action =="delete":
            try :
                pk =int (request .POST .get ("schedule_pk",""))
                sched =WorkdaySchedule .objects .get (pk =pk ,company =company )


                CompanyUser .objects .filter (
                company =company ,workday_schedule =sched 
                ).update (workday_schedule =None )
                label =sched .label 
                sched .delete ()
                django_messages .success (
                request ,
                f"Horario '{label}' eliminado correctamente.",
                )
            except (WorkdaySchedule .DoesNotExist ,ValueError ,TypeError ):
                django_messages .error (
                request ,
                "Horario no encontrado o no pertenece a esta empresa.",
                )
            return redirect (REDIRECT_URL )





        from ivr_config .models import WorkdaySchedule as _WDS_choices 

        label_val =request .POST .get ("label","").strip ()
        season_val =request .POST .get ("season",_WDS_choices .Season .WINTER ).strip ()
        is_intensive_val =bool (request .POST .get ("is_intensive"))
        start_morning_val =request .POST .get ("start_time_morning","").strip ()
        end_morning_val =request .POST .get ("end_time_morning","").strip ()
        start_afternoon_val =request .POST .get ("start_time_afternoon","").strip ()
        end_afternoon_val =request .POST .get ("end_time_afternoon","").strip ()
        tol_val =request .POST .get ("tolerance_minutes","15").strip ()
        is_default_val =bool (request .POST .get ("is_default"))

        errors =[]
        if not label_val :
            errors .append ("El nombre del horario es obligatorio.")

        if season_val not in dict (_WDS_choices .Season .choices ):
            errors .append ("La temporada seleccionada no es válida.")
            season_val =_WDS_choices .Season .WINTER 

        start_morning =self ._parse_time (start_morning_val )
        if start_morning is None :
            errors .append ("La hora de entrada de mañana debe tener formato HH:MM.")

        end_morning =self ._parse_time (end_morning_val )
        if end_morning is None :
            errors .append ("La hora de salida de mañana debe tener formato HH:MM.")

        if start_morning and end_morning and end_morning <=start_morning :
            errors .append ("La hora de salida de mañana debe ser posterior a la de entrada.")



        start_afternoon =None 
        end_afternoon =None 
        if not is_intensive_val :
            start_afternoon =self ._parse_time (start_afternoon_val )
            end_afternoon =self ._parse_time (end_afternoon_val )
            if start_afternoon is None :
                errors .append ("La hora de entrada de tarde es obligatoria para jornada partida.")
            if end_afternoon is None :
                errors .append ("La hora de salida de tarde es obligatoria para jornada partida.")
            if start_afternoon and end_morning and start_afternoon <=end_morning :
                errors .append (
                "La hora de entrada de tarde debe ser posterior a la de salida de mañana."
                )
            if start_afternoon and end_afternoon and end_afternoon <=start_afternoon :
                errors .append ("La hora de salida de tarde debe ser posterior a la de entrada.")

        try :
            tolerance =int (tol_val )
            if tolerance <0 :
                raise ValueError 
        except (ValueError ,TypeError ):
            errors .append ("La tolerancia debe ser un número entero positivo.")
            tolerance =15 

        if errors :
            return render (
            request ,
            self .template_name ,
            self ._build_context (request ,errors =errors ,post_data =request .POST ),
            )





        if action =="update":
            try :
                pk =int (request .POST .get ("schedule_pk",""))
                sched =WorkdaySchedule .objects .get (pk =pk ,company =company )
                sched .label =label_val 
                sched .season =season_val 
                sched .is_intensive =is_intensive_val 
                sched .start_time_morning =start_morning 
                sched .end_time_morning =end_morning 
                sched .start_time_afternoon =start_afternoon 
                sched .end_time_afternoon =end_afternoon 
                sched .tolerance_minutes =tolerance 
                sched .is_default =is_default_val 
                sched .save ()
                django_messages .success (
                request ,
                f"Horario '{label_val}' actualizado correctamente.",
                )
            except (WorkdaySchedule .DoesNotExist ,ValueError ,TypeError ):
                django_messages .error (
                request ,
                "Horario no encontrado o no pertenece a esta empresa.",
                )
            return redirect (REDIRECT_URL )





        WorkdaySchedule .objects .create (
        company =company ,
        label =label_val ,
        season =season_val ,
        is_intensive =is_intensive_val ,
        start_time_morning =start_morning ,
        end_time_morning =end_morning ,
        start_time_afternoon =start_afternoon ,
        end_time_afternoon =end_afternoon ,
        tolerance_minutes =tolerance ,
        is_default =is_default_val ,
        )
        django_messages .success (
        request ,
        f"Horario '{label_val}' creado correctamente.",
        )
        return redirect (REDIRECT_URL )


class AbsenceCategoryListView (SupervisorAccessMixin ,View ):
    """
    Lists all AbsenceCategory records for the authenticated user's company.
    Renders the full absence category management page on GET.

    GET /panel/absence-categories/
    ---
    Lista todos los registros AbsenceCategory de la empresa del usuario
    autenticado. Renderiza la página completa de gestión en GET.

    GET /panel/absence-categories/
    """

    template_name ="panel/workday/absence_category_list.html"

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the absence category list page.
        ---
        Renderiza la página de listado de categorías de ausencia.
        """
        from ivr_config .models import AbsenceCategory 
        cu =request .user .company_user 
        company =cu .company 
        categories =AbsenceCategory .objects .filter (
        company =company 
        ).order_by ("order","label")
        context ={
        "company":company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"absence_categories",
        "categories":categories ,
        }
        return render (request ,self .template_name ,context )


class AbsenceCategoryCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a new AbsenceCategory for the authenticated user's company.
    On success or failure, redirects to absence_category_list.

    POST /panel/absence-categories/create/
    ---
    Crea una nueva AbsenceCategory para la empresa del usuario autenticado.
    En caso de éxito o fallo, redirige a absence_category_list.

    POST /panel/absence-categories/create/
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Validates POST data and creates the AbsenceCategory.
        ---
        Valida los datos POST y crea la AbsenceCategory.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        label_val =request .POST .get ("label","").strip ()
        code_val =request .POST .get ("code","").strip ().upper ()
        requires_note =bool (request .POST .get ("requires_note"))
        is_justified =bool (request .POST .get ("is_justified"))
        order_val =request .POST .get ("order","0").strip ()

        if not label_val or not code_val :
            django_messages .error (
            request ,
            "El nombre y el código son obligatorios.",
            )
            return redirect (LIST_URL )

        try :
            order =int (order_val )
        except (ValueError ,TypeError ):
            order =0 

        if AbsenceCategory .objects .filter (company =company ,code =code_val ).exists ():
            django_messages .error (
            request ,
            f"Ya existe una categoría con el código '{code_val}' en esta empresa.",
            )
            return redirect (LIST_URL )

        AbsenceCategory .objects .create (
        company =company ,
        code =code_val ,
        label =label_val ,
        requires_note =requires_note ,
        is_justified =is_justified ,
        order =order ,
        is_active =True ,
        )
        django_messages .success (
        request ,
        f"Categoría '{label_val}' creada correctamente.",
        )
        return redirect (LIST_URL )


class AbsenceCategoryUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing AbsenceCategory belonging to the authenticated
    user's company. On success or failure, redirects to absence_category_list.

    POST /panel/absence-categories/<pk>/update/
    ---
    Actualiza una AbsenceCategory existente de la empresa del usuario
    autenticado. En caso de éxito o fallo, redirige a absence_category_list.

    POST /panel/absence-categories/<pk>/update/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Validates POST data and updates the AbsenceCategory.
        ---
        Valida los datos POST y actualiza la AbsenceCategory.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        try :
            category =AbsenceCategory .objects .get (pk =pk ,company =company )
        except AbsenceCategory .DoesNotExist :
            django_messages .error (
            request ,
            "Categoría no encontrada o no pertenece a esta empresa.",
            )
            return redirect (LIST_URL )

        label_val =request .POST .get ("label","").strip ()
        requires_note =bool (request .POST .get ("requires_note"))
        is_justified =bool (request .POST .get ("is_justified"))
        order_val =request .POST .get ("order","0").strip ()

        if not label_val :
            django_messages .error (request ,"El nombre es obligatorio.")
            return redirect (LIST_URL )

        try :
            order =int (order_val )
        except (ValueError ,TypeError ):
            order =category .order 

        category .label =label_val 
        category .requires_note =requires_note 
        category .is_justified =is_justified 
        category .order =order 
        category .save (update_fields =[
        "label","requires_note","is_justified","order"
        ])

        django_messages .success (
        request ,
        f"Categoría '{label_val}' actualizada correctamente.",
        )
        return redirect (LIST_URL )


class AbsenceCategoryToggleView (SupervisorAccessMixin ,View ):
    """
    Toggles the is_active flag of an AbsenceCategory. Deactivating a
    category hides it from the operator's gap resolution selector without
    deleting historical references in WorkdayGap records.

    POST /panel/absence-categories/<pk>/toggle/
    ---
    Alterna el flag is_active de una AbsenceCategory. Desactivar una
    categoría la oculta del selector del operario sin eliminar las
    referencias históricas en registros WorkdayGap.

    POST /panel/absence-categories/<pk>/toggle/
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Toggles is_active and redirects to absence_category_list.
        ---
        Alterna is_active y redirige a absence_category_list.
        """
        from django .urls import reverse 
        from ivr_config .models import AbsenceCategory 

        cu =request .user .company_user 
        company =cu .company 
        LIST_URL =reverse ("panel:absence_category_list")

        try :
            category =AbsenceCategory .objects .get (pk =pk ,company =company )
        except AbsenceCategory .DoesNotExist :
            django_messages .error (
            request ,
            "Categoría no encontrada o no pertenece a esta empresa.",
            )
            return redirect (LIST_URL )

        category .is_active =not category .is_active 
        category .save (update_fields =["is_active"])

        state ="activada"if category .is_active else "desactivada"
        django_messages .success (
        request ,
        f"Categoría '{category.label}' {state} correctamente.",
        )
        return redirect (LIST_URL )


class WorkOrderDetailView (CompanyUserRequiredMixin ,View ):
    """
    Read-only detail view of a single WorkOrder -- a column listing of
    the part's full content (entries, lines, spare parts), NOT the
    edit form. Added at Miguel Ángel's explicit request (2026-07-15):
    the "Revisados" tab only had "Borrar" as an action -- needed a way
    to actually look at a reviewed part's content without editing it.

    Access: ADMIN/SUPERVISOR/WORKSHOPBOSS only -- same gate as the
    "Ver" button in _wo_table.html (is_elevated), enforced here too so
    a WORKSHOP operator can't reach another operator's part by
    guessing the URL even though they never see the button.
    ---
    Vista de detalle de solo lectura de un WorkOrder -- un listado por
    columnas del contenido completo del parte (entries, líneas,
    repuestos), NO el formulario de edición. Añadida a petición
    explícita de Miguel Ángel (2026-07-15): la pestaña "Revisados" solo
    tenía "Borrar" como acción -- hacía falta una forma de consultar el
    contenido de un parte revisado sin editarlo.

    Acceso: solo ADMIN/SUPERVISOR/WORKSHOPBOSS -- mismo gate que el
    botón "Ver" en _wo_table.html (is_elevated), impuesto también aquí
    para que un operario WORKSHOP no pueda llegar al parte de otro
    adivinando la URL, aunque nunca vea el botón.
    """
    template_name = "panel/work_orders/detail.html"

    def get (self ,request ,pk ,*args ,**kwargs ):
        from django .shortcuts import get_object_or_404 as _get_404 
        from django .http import HttpResponseForbidden 

        cu =request .user .company_user 
        company =cu .company 

        if cu .role not in (
        CompanyUser .ROLE_ADMIN ,
        CompanyUser .ROLE_SUPERVISOR ,
        CompanyUser .ROLE_WORKSHOPBOSS ,
        ):
            return HttpResponseForbidden (
            "Acción no disponible para tu rol."
            )

        work_order =_get_404 (
        WorkOrder .objects .select_related (
        "uploaded_by__user","reviewed_by__user","generated_by__user",
        ),
        pk =pk ,company =company ,
        )
        entries =(
        work_order .entries 
        .order_by ("page_number")
        .prefetch_related (
        "lines__machine_asset",
        "lines__spare_parts",
        )
        )

        return render (request ,self .template_name ,{
        "company":company ,
        "company_user":cu ,
        "active_nav":"work_order_admin_history",
        "work_order":work_order ,
        "entries":entries ,
        })


class WorkOrderAdminHistoryView (WorkOrderFormAccessMixin ,View ):

    """
    SINGLE unified work-order history view for ALL roles with access to
    digital parts (ADMIN, SUPERVISOR, WORKSHOPBOSS, WORKSHOP).

    Replaces, as of 2026-07-10 (H17 S012), three separate views that were
    near-duplicates of this one: DigitalWorkOrderListView ("Partes
    Digitales"), history.WorkOrderHistoryListView ("Mis partes") and
    WorkOrderEntryHistoryView ("Historial" for WORKSHOP/WORKSHOPBOSS).
    All three were deleted outright (view + template + URL + nav entry) at
    Miguel Ángel's explicit request -- no dead code, no "marked as removed"
    leftovers.

    Role scope:
      ADMIN/SUPERVISOR/WORKSHOPBOSS — see all company operators. Operator
                        filter dropdown, "Marcar revisado" and "Eliminar"
                        actions available. Ausencias/Períodos tabs visible
                        (management tabs).
      WORKSHOP        — see only their own work orders (uploaded_by=self).
                        No operator filter, no "Marcar revisado", no
                        "Eliminar" — only "Editar" (routes to
                        operator_form_edit, the sole digital-part editor).
                        Ausencias/Períodos tabs hidden (management-only,
                        not part of what the old "Mis partes" offered).
                        Ausencias tab, when visible via direct interaction,
                        would be scoped to their own WorkerAbsence records
                        read-only — see is_elevated in context.
                        Pendientes tab shows an "horas extra" summary for
                        their own active WorkPeriod (folded in from the
                        deleted operator "Horas extra" tab).

    Tabs:
      1 — Pendientes  : unreviewed work orders (own, for WORKSHOP; all
                        company operators otherwise).
                        Filters: operator (elevated only), date, machine,
                        status.
                        Actions: mark as reviewed (elevated only), link to
                        edit view.
      2 — Revisados   : reviewed work orders. Excel export available HERE
                        ONLY, elevated roles only (single work order or
                        date-range multi-export).
                        Filters: operator (elevated only), date range,
                        machine, status.
      3 — Histórico   : all work orders with cross-filters (operator,
                        date range, machine, review status, status). Read-only.
      4 — Ausencias   : elevated roles only — WorkerAbsence records per
                        operator with type/date filters. Actions: create,
                        edit, delete absence. Action 'Generar partes del
                        periodo': creates synthetic WorkOrder records (one
                        per working day Mon–Fri) for the absence range,
                        tagged with generated_by=current user.
      5 — Períodos    : elevated roles only — WorkPeriod management.

    GET /panel/work-orders/history/
        Optional GET params:
          tab         (str)  — active tab: pending|reviewed|history|absences|periods.
                               Default: pending.
          operator_pk (int)  — filter by CompanyUser pk (scoped to company,
                               ignored for WORKSHOP — always own).
          date_from   (str)  — ISO date YYYY-MM-DD start of range.
          date_to     (str)  — ISO date YYYY-MM-DD end of range.
          machine     (str)  — MachineAsset.code icontains filter.
          status      (str)  — WorkOrder.Status code filter (folded in from
                               the deleted "Mis partes" Estado filter —
                               recovers the old "Error" tab of Partes
                               Digitales without a dedicated tab).

    ---

    ÚNICA vista unificada de historial de partes para TODOS los roles con
    acceso a partes digitales (ADMIN, SUPERVISOR, WORKSHOPBOSS, WORKSHOP).

    Sustituye, desde el 2026-07-10 (H17 S012), a tres vistas independientes
    que eran casi-duplicados de esta: DigitalWorkOrderListView ("Partes
    Digitales"), history.WorkOrderHistoryListView ("Mis partes") y
    WorkOrderEntryHistoryView ("Historial" para WORKSHOP/WORKSHOPBOSS).
    Las tres se eliminaron por completo (vista + plantilla + URL + entrada
    de menú) a petición expresa de Miguel Ángel — sin código muerto, sin
    restos "marcados como eliminados".

    Alcance por rol:
      ADMIN/SUPERVISOR/WORKSHOPBOSS — ven todos los operarios de la empresa.
                        Selector de operario, acciones "Marcar revisado" y
                        "Eliminar" disponibles. Pestañas Ausencias/Períodos
                        visibles (gestión).
      WORKSHOP        — ven solo sus propios partes (uploaded_by=self). Sin
                        selector de operario, sin "Marcar revisado", sin
                        "Eliminar" — solo "Editar" (enruta a
                        operator_form_edit, único editor de partes digitales).
                        Pestañas Ausencias/Períodos ocultas (son gestión, no
                        algo que ofreciera la antigua "Mis partes").
                        La pestaña Pendientes muestra un resumen de "horas
                        extra" de su periodo activo (traspasado desde la
                        pestaña "Horas extra" del operador, eliminada).

    Pestañas:
      1 — Pendientes  : partes sin revisar (propios para WORKSHOP; de toda
                        la empresa en el resto).
                        Filtros: operario (solo elevados), fecha, máquina,
                        estado.
                        Acciones: marcar como revisado (solo elevados),
                        enlace a vista de edición.
      2 — Revisados   : partes revisados. Exportación Excel disponible SOLO
                        AQUÍ, solo roles elevados (parte individual o
                        multi-exportación por rango de fechas).
                        Filtros: operario (solo elevados), rango de fechas,
                        máquina, estado.
      3 — Histórico   : todos los partes con filtros cruzados (operario,
                        rango de fechas, máquina, estado de revisión,
                        estado). Solo lectura.
      4 — Ausencias   : solo roles elevados — registros WorkerAbsence por
                        operario con filtros de tipo y fecha. Acciones:
                        alta, edición, baja de ausencia. Acción 'Generar
                        partes del periodo': crea registros WorkOrder
                        sintéticos (uno por día laborable lun–vie) para el
                        rango de la ausencia, etiquetados con
                        generated_by=usuario actual.
      5 — Períodos    : solo roles elevados — gestión de WorkPeriod.

    GET /panel/work-orders/history/
        Parámetros GET opcionales:
          tab         (str)  — pestaña activa: pending|reviewed|history|absences|periods.
                               Por defecto: pending.
          operator_pk (int)  — filtrar por pk de CompanyUser (acotado a
                               empresa, ignorado para WORKSHOP — siempre
                               el propio).
          date_from   (str)  — fecha ISO YYYY-MM-DD inicio del rango.
          date_to     (str)  — fecha ISO YYYY-MM-DD fin del rango.
          machine     (str)  — filtro icontains sobre MachineAsset.code.
          status      (str)  — filtro por código de WorkOrder.Status
                               (traspasado desde el filtro Estado de la
                               antigua "Mis partes" — recupera la vieja
                               pestaña "Error" de Partes Digitales sin
                               necesidad de una pestaña dedicada).
    """

    template_name ="panel/work_orders/admin_history.html"



    _MESES_ES ={
    1 :"Enero",2 :"Febrero",3 :"Marzo",4 :"Abril",
    5 :"Mayo",6 :"Junio",7 :"Julio",8 :"Agosto",
    9 :"Septiembre",10 :"Octubre",11 :"Noviembre",12 :"Diciembre",
    }

    def _get_own_presence (self ,company_user ):
        """
        Returns the current active PresenceStatus for the authenticated user.
        ---
        Retorna el PresenceStatus activo actual del usuario autenticado.
        """
        return PresenceStatus .objects .filter (
        company_user =company_user ,
        starts_at__lte =now (),
        ).filter (
        Q (ends_at__isnull =True )|Q (ends_at__gt =now ())
        ).order_by ("-starts_at").first ()

    def _parse_date (self ,value ):
        """
        Parses an ISO date string (YYYY-MM-DD) and returns a date object,
        or None if the value is absent or malformed.
        ---
        Parsea una cadena de fecha ISO (YYYY-MM-DD) y devuelve un objeto date,
        o None si el valor esta ausente o malformado.
        """
        from datetime import datetime 
        if not value :
            return None 
        try :
            return datetime .strptime (value .strip (),"%Y-%m-%d").date ()
        except ValueError :
            return None 

    def _build_base_queryset (self ,company ,owner =None ):
        """
        Returns the base WorkOrder queryset scoped to the company, restricted
        to DIGITAL and GENERATED sources (operator-entered parts only).
        PDF_UPLOAD parts belong exclusively to WorkOrderListView.
        When owner is provided (WORKSHOP role — always their own records,
        never the operator_pk filter), the queryset is further scoped to
        uploaded_by=owner.
        All required related data is prefetched for the admin history tabs.
        ---
        Devuelve el queryset base de WorkOrder acotado a la empresa, restringido
        a los origenes DIGITAL y GENERATED (partes introducidos por operarios).
        Los partes PDF_UPLOAD pertenecen exclusivamente a WorkOrderListView.
        Cuando se proporciona owner (rol WORKSHOP — siempre sus propios
        registros, nunca el filtro operator_pk), el queryset se acota además
        a uploaded_by=owner.
        Todos los datos relacionados se prefetchean para las pestanas del historial.
        """
        qs =(
        WorkOrder .objects 
        .filter (
        company =company ,
        source__in =[
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        ],
        )
        .select_related ("uploaded_by__user","reviewed_by__user","generated_by__user")
        .prefetch_related (
        Prefetch (
        "entries",
        queryset =WorkOrderEntry .objects .prefetch_related ("lines"),
        )
        )
        .order_by ("-id")
        )
        if owner is not None :
            qs =qs .filter (uploaded_by =owner )
        return qs 

    def _apply_filters (self ,qs ,operator_pk ,date_from ,date_to ,machine ,company ,
    fault_category ="",q ="",status ="",is_elevated =True ):
        """
        Applies the optional GET filters to a WorkOrder queryset.
        operator_pk filters by uploaded_by (ignored when is_elevated is False
        — WORKSHOP is already scoped to their own records by
        _build_base_queryset, and must never be able to widen the query to
        another operator via a crafted GET param); date_from/date_to filter
        by the first entry's work_date; machine filters by machine_asset__code.
        fault_category filters by WorkOrderEntryLine.fault_category (exact).
        q performs a free-text OR search over fault_description and repair_notes.
        status filters by WorkOrder.status (exact) — folded in from the
        deleted "Mis partes" Estado filter.
        ---
        Aplica los filtros GET opcionales a un queryset de WorkOrder.
        operator_pk filtra por uploaded_by (ignorado cuando is_elevated es
        False — WORKSHOP ya está acotado a sus propios registros por
        _build_base_queryset, y nunca debe poder ampliar la consulta a otro
        operario mediante un parámetro GET manipulado); date_from/date_to
        filtran por work_date del primer entry; machine filtra por
        machine_asset__code.
        fault_category filtra por WorkOrderEntryLine.fault_category (exacto).
        q realiza una búsqueda libre OR sobre fault_description y repair_notes.
        status filtra por WorkOrder.status (exacto) — traspasado desde el
        filtro Estado de la antigua "Mis partes".
        """
        if operator_pk and is_elevated :
            try :
                cu_pk =int (operator_pk )
                qs =qs .filter (uploaded_by__pk =cu_pk ,uploaded_by__company =company )
            except (ValueError ,TypeError ):
                pass 
        if date_from :
            qs =qs .filter (entries__work_date__gte =date_from )
        if date_to :
            qs =qs .filter (entries__work_date__lte =date_to )
        if machine :
            qs =qs .filter (
            entries__lines__machine_asset__code__icontains =machine 
            )
        if status :
            qs =qs .filter (status =status )
        if fault_category :


            qs =qs .filter (
            entries__lines__fault_category =fault_category 
            ).distinct ()
        if q :


            from django .db .models import Q as _Q 
            qs =qs .filter (
            _Q (entries__lines__fault_description__icontains =q )
            |_Q (entries__lines__repair_notes__icontains =q )
            )
        return qs .distinct ()

    def _enrich_work_orders (self ,qs ,active_fault_category =""):
        """
        Converts a WorkOrder queryset into a list of enriched dicts suitable
        for template rendering. Each dict includes pk, fecha, operator name,
        num_bloques, horas_totales, horas_extra, dieta (True if any entry of
        the work order has has_diet=True -- 2026-07-08, gap señalado por
        Miguel Ángel: la vista no exponía la dieta) and reviewed flag.
        When active_fault_category is provided (a FaultCategory internal value),
        the fault_category badge of every enriched dict is forced to the label
        of that category rather than being calculated from the dominant category
        across all lines. This ensures the badge is always coherent with the
        active filter — a work order returned by the filter is guaranteed to have
        at least one line matching the category, so forcing the label is correct.
        ---
        Convierte un queryset de WorkOrder en una lista de dicts enriquecidos
        adecuados para renderizado en template. Cada dict incluye pk, fecha,
        nombre del operario, num_bloques, horas_totales, horas_extra, dieta
        (True si algún entry del parte tiene has_diet=True -- 2026-07-08, gap
        señalado por Miguel Ángel: la vista no exponía la dieta) y flag
        reviewed. Cuando active_fault_category contiene un valor interno de
        FaultCategory, el badge fault_category de cada dict se fuerza al
        label de esa categoría en lugar de calcularse como la dominante sobre
        todas las líneas. Esto garantiza coherencia visual entre el filtro
        activo y el badge mostrado — un parte devuelto por el filtro tiene
        garantizada al menos una línea con esa categoría, por lo que forzar
        el label es semánticamente correcto.
        """
        from decimal import Decimal 
        result =[]
        for wo in qs :
            entries_list =list (wo .entries .all ())
            first_entry =entries_list [0 ]if entries_list else None 
            work_date =first_entry .work_date if first_entry else None 
            num_bloques =sum (entry .lines .count ()for entry in entries_list )
            horas_totales =sum (
            (line .delta_hours 
            for entry in entries_list 
            for line in entry .lines .all ()
            if line .delta_hours is not None ),
            Decimal ("0"),
            )
            operator_name =(
            wo .uploaded_by .user .get_full_name ()or wo .uploaded_by .user .username 
            if wo .uploaded_by else "Desconocido"
            )
            result .append ({
            "pk":wo .pk ,
            "fecha":work_date ,
            "operator_name":operator_name ,
            "operator_pk":wo .uploaded_by .pk if wo .uploaded_by else None ,
            "num_bloques":num_bloques ,
            "horas_totales":horas_totales ,
            "horas_extra":max (Decimal ("0"),horas_totales -Decimal ("8")),
            "dieta":any (entry .has_diet for entry in entries_list ),
            "reviewed":wo .reviewed ,
            "reviewed_by":(
            wo .reviewed_by .user .get_full_name ()or wo .reviewed_by .user .username 
            if wo .reviewed_by else None 
            ),
            "reviewed_at":wo .reviewed_at ,
            "generated_by":(
            wo .generated_by .user .get_full_name ()or wo .generated_by .user .username 
            if wo .generated_by else None 
            ),
            "excel_url":wo .excel_file .url if wo .excel_file else None ,
            "fault_category":self ._dominant_fault_category (
            wo ,active_fault_category =active_fault_category 
            ),
            # Source value used by admin_history.html to dispatch the correct
            # edit URL: operator_form_edit for DIGITAL/GENERATED, work_order_edit
            # for PDF parts.
            # Valor source usado por admin_history.html para despachar la URL
            # de edición correcta: operator_form_edit para DIGITAL/GENERATED,
            # work_order_edit para partes PDF.
            "source":wo .source ,
            # Pipeline status (PENDING/PROCESSING/DONE/ERROR/...), traspasado
            # desde la antigua "Mis partes" para el filtro/badge Estado y
            # para recuperar la vieja pestaña "Error" de Partes Digitales.
            "status":wo .status ,
            "status_display":wo .get_status_display (),
            # Clase Bootstrap del badge de Estado, resuelta aquí y no en el
            # template -- la plantilla solo debe renderizar, no decidir.
            # Bootstrap badge class for the Estado badge, resolved here and
            # not in the template -- templates only render, never decide.
            "status_badge_class":(
            "bg-danger"if wo .status =="ERROR"
            else ""if wo .status =="DONE"
            else "bg-secondary"
            ),
            })
        return result 

    def _dominant_fault_category (self ,wo ,active_fault_category =""):
        """
        Returns the display label of the most frequent fault_category among all
        WorkOrderEntryLine records of this WorkOrder. Returns an empty string
        when no lines exist or all lines have an empty category.

        When active_fault_category is supplied (a FaultCategory internal value
        such as "ENGINE_TRANSMISSION"), the method short-circuits: it resolves
        the label for that value from FaultCategory.choices and returns it
        immediately, bypassing the frequency calculation. This is correct because
        any WorkOrder present in a fault_category-filtered queryset is guaranteed
        to contain at least one line with that exact category value — there is no
        semantic loss in forcing the label. The short-circuit also avoids
        redundant iteration over prefetched lines on every row render.
        ---
        Devuelve el label del fault_category más frecuente entre todas las
        WorkOrderEntryLine del WorkOrder. Devuelve cadena vacía cuando no hay
        líneas o todas tienen la categoría vacía.

        Cuando active_fault_category contiene un valor interno de FaultCategory
        (ej. "ENGINE_TRANSMISSION"), el método hace cortocircuito: resuelve el
        label de ese valor desde FaultCategory.choices y lo devuelve directamente,
        sin ejecutar el cálculo de frecuencia. Esto es semánticamente correcto
        porque cualquier WorkOrder presente en un queryset filtrado por
        fault_category tiene garantizada al menos una línea con ese valor exacto.
        El cortocircuito también evita iteración redundante sobre las líneas
        prefetcheadas en cada fila del render.
        """
        from collections import Counter 
        from work_order_processor .models import FaultCategory 
        label_map ={fc [0 ]:str (fc [1 ])for fc in FaultCategory .choices }







        if active_fault_category :
            label =label_map .get (active_fault_category )
            if label :
                return label 



        categories =[
        line .fault_category 
        for entry in wo .entries .all ()
        for line in entry .lines .all ()
        if line .fault_category 
        ]
        if not categories :
            return ""
        dominant =Counter (categories ).most_common (1 )[0 ][0 ]
        return label_map .get (dominant ,dominant )

    def get (self ,request ,*args ,**kwargs ):
        """
        Renders the four-tab admin history page. Resolves GET filters,
        builds each tab's queryset independently and passes all data to
        the template context.
        ---
        Renderiza la pagina de historial de administrador de cuatro pestanas.
        Resuelve los filtros GET, construye el queryset de cada pestana
        de forma independiente y pasa todos los datos al contexto del template.
        """
        cu =request .user .company_user 
        company =cu .company 

        # Alcance por rol -- pieza central de la vista unificada (H17 S012):
        # ADMIN/SUPERVISOR/WORKSHOPBOSS ven todos los operarios; WORKSHOP
        # solo los suyos, sin excepcion, sin importar que parametros GET
        # lleguen (operator_pk se ignora para WORKSHOP en _apply_filters).
        # ---
        # Role scope -- central piece of the unified view (H17 S012):
        # ADMIN/SUPERVISOR/WORKSHOPBOSS see all operators; WORKSHOP sees
        # only their own, no exceptions, regardless of incoming GET params
        # (operator_pk is ignored for WORKSHOP inside _apply_filters).
        is_elevated =cu .role in (
        CompanyUser .ROLE_ADMIN ,
        CompanyUser .ROLE_SUPERVISOR ,
        CompanyUser .ROLE_WORKSHOPBOSS ,
        )
        _owner =None if is_elevated else cu 



        active_tab =request .GET .get ("tab","pending")
        operator_pk =request .GET .get ("operator_pk","").strip ()
        date_from =self ._parse_date (request .GET .get ("date_from",""))
        date_to =self ._parse_date (request .GET .get ("date_to",""))
        machine =request .GET .get ("machine","").strip ()
        fault_category =request .GET .get ("fault_category","").strip ()
        q =request .GET .get ("q","").strip ()
        status =request .GET .get ("status","").strip ()









        _VALID_SORT_COLS =frozenset (
        ("fecha","operator_name","horas_totales","horas_extra","reviewed","fault_category")
        )
        _VALID_SORT_DIRS =frozenset (("asc","desc"))

        raw_stack =request .GET .get ("sort_stack","").strip ()



        sort_stack =[]
        if raw_stack :
            for token in raw_stack .split (","):
                token =token .strip ()
                if ":"not in token :
                    continue 
                col ,_ ,direction =token .partition (":")
                col =col .strip ()
                direction =direction .strip ()
                if col in _VALID_SORT_COLS and direction in _VALID_SORT_DIRS :


                    if not any (s [0 ]==col for s in sort_stack ):
                        sort_stack .append ((col ,direction ))
                        if len (sort_stack )>=3 :
                            break 



        if not sort_stack :
            sort_stack =[("fecha","asc")]



        sort_primary_col =sort_stack [0 ][0 ]
        sort_primary_dir =sort_stack [0 ][1 ]



        sort_stack_str =",".join (f"{col}:{direction}"for col ,direction in sort_stack )





        operators =(
        CompanyUser .objects 
        .filter (
        company =company ,
        is_active =True ,
        role =CompanyUser .ROLE_WORKSHOP ,
        )
        .select_related ("user")
        .order_by ("user__last_name","user__first_name")
        )if is_elevated else CompanyUser .objects .none ()





        qs_pending =self ._build_base_queryset (company ,owner =_owner ).filter (reviewed =False )
        qs_pending =self ._apply_filters (
        qs_pending ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,status =status ,is_elevated =is_elevated ,
        )
        pending_list =self ._enrich_work_orders (qs_pending ,active_fault_category =fault_category )





        qs_reviewed =self ._build_base_queryset (company ,owner =_owner ).filter (reviewed =True )
        qs_reviewed =self ._apply_filters (
        qs_reviewed ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,status =status ,is_elevated =is_elevated ,
        )
        reviewed_list =self ._enrich_work_orders (qs_reviewed ,active_fault_category =fault_category )

        # Excluir de Revisados los partes cuya fecha cae dentro de un
        # WorkPeriod liquidado (is_closed=True) de su propio operario --
        # gap señalado por Miguel Ángel 2026-07-08: un parte liquidado ya
        # no debe seguir apareciendo en Revisados, solo en Histórico (que
        # sí muestra todos los periodos sin filtrar). Mismo patrón de rango
        # de fechas que WorkPeriodLockView usa para detectar partes dentro
        # de un periodo (ver arriba, comprobación de partes sin revisar).
        # Excludes from Revisados any part whose date falls within a
        # liquidated WorkPeriod (is_closed=True) of its own operator --
        # gap flagged by Miguel Ángel 2026-07-08: a liquidated part should
        # no longer show in Revisados, only in Histórico (which shows all
        # periods unfiltered). Same date-range pattern WorkPeriodLockView
        # already uses to detect parts within a period (see above, the
        # unreviewed-parts check).
        from ivr_config .models import WorkPeriod as _WP_closed 
        _closed_ranges_by_operator ={}
        for _op_pk ,_start ,_end in (
        _WP_closed .objects 
        .filter (company_user__company =company ,is_closed =True ,end_date__isnull =False )
        .values_list ("company_user_id","start_date","end_date")
        ):
            _closed_ranges_by_operator .setdefault (_op_pk ,[]).append ((_start ,_end ))

        def _falls_in_closed_period (wo_dict ):
            _ranges =_closed_ranges_by_operator .get (wo_dict ["operator_pk"])
            if not _ranges or wo_dict ["fecha"]is None :
                return False 
            return any (_s <=wo_dict ["fecha"]<=_e for _s ,_e in _ranges )

        reviewed_list =[wo for wo in reviewed_list if not _falls_in_closed_period (wo )]





        qs_history =self ._build_base_queryset (company ,owner =_owner ).filter (reviewed =True )
        qs_history =self ._apply_filters (
        qs_history ,operator_pk ,date_from ,date_to ,machine ,company ,
        fault_category =fault_category ,q =q ,status =status ,is_elevated =is_elevated ,
        )
        history_list =self ._enrich_work_orders (qs_history ,active_fault_category =fault_category )


















        def _sort_key_for_col (item ,col ):
            """
            Returns a comparable sort key for a single column in an enriched
            work-order dict. None values and empty strings always sort last,
            regardless of the sort direction applied by the caller.
            ---
            Devuelve una clave de ordenación comparable para una sola columna
            en un dict enriquecido de parte. None y cadena vacía se ordenan
            siempre al final, independientemente de la dirección aplicada.
            """
            val =item .get (col )
            if val is None or val =="":


                return (1 ,"")
            return (0 ,val )

        def _apply_sort_stack (lst ,stack ):
            """
            Applies a multi-level stable sort to lst according to stack.
            stack is a list of (col, dir) pairs ordered from most to least
            significant. Sorting is applied from least significant to most
            significant to leverage Python's stable sort.
            ---
            Aplica una ordenación estable multi-nivel a lst según stack.
            stack es una lista de pares (col, dir) de más a menos significativo.
            La ordenación se aplica de menos significativo a más significativo
            para aprovechar la estabilidad del sort de Python.
            """
            result =list (lst )
            for col ,direction in reversed (stack ):
                reverse_flag =(direction =="desc")
                result =sorted (
                result ,
                key =lambda item ,_col =col :_sort_key_for_col (item ,_col ),
                reverse =reverse_flag ,
                )
            return result 

        pending_list =_apply_sort_stack (pending_list ,sort_stack )
        reviewed_list =_apply_sort_stack (reviewed_list ,sort_stack )
        history_list =_apply_sort_stack (history_list ,sort_stack )





        from ivr_config .models import WorkerAbsence 
        absence_qs =(
        WorkerAbsence .objects 
        .filter (company_user__company =company )
        .select_related (
        "company_user__user",
        "registered_by__user",
        )
        .order_by ("-start_date")
        )

        if is_elevated :
            if operator_pk :
                try :
                    absence_qs =absence_qs .filter (company_user__pk =int (operator_pk ))
                except (ValueError ,TypeError ):
                    pass 
        else :
            # WORKSHOP: siempre las propias, sin excepcion -- pestaña de
            # solo lectura, traspasada desde la antigua operator_history.
            # WORKSHOP: always their own, no exceptions -- read-only tab,
            # folded in from the deleted operator_history.
            absence_qs =absence_qs .filter (company_user =cu )
        absences_list =list (absence_qs )





        if is_elevated :
            from ivr_config .models import WorkPeriod 
            from datetime import timedelta as _td 
            last_closed =(
            WorkPeriod .objects 
            .filter (
            company_user__company =company ,
            end_date__isnull =False ,
            )
            .order_by ("-end_date")
            .values_list ("end_date","start_date")
            .first ()
            )
            if last_closed :
                _last_end =last_closed [0 ]
                _last_start =last_closed [1 ]
                _period_len =_last_end -_last_start 
                suggested_period_start =(_last_end +_td (days =1 )).strftime ("%Y-%m-%d")
                suggested_period_end =(_last_end +_td (days =1 )+_period_len ).strftime ("%Y-%m-%d")
            else :
                suggested_period_start =""
                suggested_period_end =""

            # Build period_groups for the Períodos tab (WorkPeriodGroup-based).
            # Construir period_groups para la pestaña Períodos (basado en WorkPeriodGroup).
            from ivr_config.models import WorkPeriodGroup as _WPG_H
            _period_groups = list(
                _WPG_H.objects
                .filter(company=company)
                .select_related("created_by__user")
                .order_by("-start_date")
            )
            _total_operators = CompanyUser .objects .filter (
                company=company, is_active=True, role=CompanyUser.ROLE_WORKSHOP
            ).count()
            period_groups = []
            for _g in _period_groups:
                _op_count = _g.operator_periods.count()
                period_groups.append({
                    "group":           _g,
                    "operator_count":  _op_count,
                    "total_operators": _total_operators,
                })

            # Legacy fallback: keep period_operator_groups for backward compat.
            # Fallback legacy: mantener period_operator_groups por compatibilidad.
            period_operator_groups = []
        else :
            # WORKSHOP: la pestaña Períodos es gestión pura, no es parte de
            # lo que ofrecía ninguna de las vistas sustituidas -- se oculta
            # en el template, así que no hace falta calcular nada de esto.
            # WORKSHOP: the Períodos tab is pure management, not part of
            # what any of the replaced views offered -- hidden in the
            # template, so none of this needs computing.
            suggested_period_start =""
            suggested_period_end =""
            period_groups =[]
            period_operator_groups =[]

        # Resumen de horas extra del periodo activo -- traspasado desde la
        # antigua pestaña "Horas extra" de operator_history (eliminada).
        # Solo se calcula, y solo tiene sentido, para WORKSHOP.
        # ---
        # Active-period overtime summary -- folded in from the deleted
        # operator_history "Horas extra" tab. Only computed, and only
        # meaningful, for WORKSHOP.
        overtime_hours =None 
        if not is_elevated :
            from datetime import date as _date_OT 
            from ivr_config .models import WorkPeriod as _WP_OT 
            from decimal import Decimal as _Dec_OT 
            _active_period =(
            _WP_OT .objects 
            .filter (company_user =cu )
            .filter (Q (end_date__isnull =True )|Q (end_date__gte =_date_OT .today ()))
            .order_by ("-start_date")
            .first ()
            )
            if _active_period :
                overtime_hours =sum (
                (wo ["horas_extra"]for wo in pending_list +reviewed_list 
                if wo ["operator_pk"]==cu .pk 
                and wo ["fecha"]is not None 
                and _active_period .start_date <=wo ["fecha"]
                and (_active_period .end_date is None or wo ["fecha"]<=_active_period .end_date )
                ),
                _Dec_OT ("0"),
                )


        from work_order_processor.models import ExportTemplate as _ET
        _templates_propias = _ET.objects.filter(company_user=cu, is_global=False).order_by("-is_default", "name")
        _templates_globales = _ET.objects.filter(company=cu.company, is_global=True).order_by("name")

        # Compute aggregate totals for the Revisados tab.
        # Calcular totales agregados para la pestaña Revisados.
        from decimal import Decimal as _Dec
        reviewed_totals = {
            "horas_totales": sum(
                (wo["horas_totales"] for wo in reviewed_list
                 if wo["horas_totales"] is not None),
                _Dec("0"),
            ),
            "horas_extra": sum(
                (wo["horas_extra"] for wo in reviewed_list
                 if wo["horas_extra"] is not None),
                _Dec("0"),
            ),
            "dietas": sum(1 for wo in reviewed_list if wo["dieta"]),
        }

        context ={
        "company":cu .company ,
        "company_user":cu ,
        "own_presence":self ._get_own_presence (cu ),
        "active_nav":"work_order_admin_history",
        "active_tab":active_tab ,
        # Alcance por rol -- controla en el template el selector de operario,
        # los botones Marcar revisado/Eliminar y las pestañas Ausencias/Períodos.
        # Role scope -- controls the operator selector, the Marcar
        # revisado/Eliminar buttons and the Ausencias/Períodos tabs in the
        # template.
        "is_elevated":is_elevated ,
        "operators":operators ,
        "operator_pk":operator_pk ,
        "date_from":request .GET .get ("date_from",""),
        "date_to":request .GET .get ("date_to",""),
        "machine":machine ,
        "fault_category":fault_category ,
        "q":q ,
        "status":status ,
        "status_choices":WorkOrder .Status .choices ,
        # Resumen de horas extra del periodo activo -- solo WORKSHOP.
        # Active-period overtime summary -- WORKSHOP only.
        "overtime_hours":overtime_hours ,


        "fault_category_choices":[
        (fc [0 ],fc [1 ])
        for fc in WorkOrderEntryLine .fault_category .field .choices 
        if fc [0 ]
        ],



        "sort_stack_str":sort_stack_str ,
        "sort_primary_col":sort_primary_col ,
        "sort_primary_dir":sort_primary_dir ,
        "pending_list":pending_list ,
        "reviewed_list":reviewed_list ,
        "reviewed_totals":reviewed_totals ,
        "history_list":history_list ,
        "absences_list":absences_list ,
        "suggested_period_start":suggested_period_start ,
        "suggested_period_end":suggested_period_end ,
        "templates_propias":_templates_propias ,
        "templates_globales":_templates_globales ,
        "is_admin":cu .role =="ADMIN",
        "period_operator_groups":period_operator_groups ,
        "period_groups":period_groups ,
        "column_choices":[
        ("fecha","Fecha"),
        ("operario","Operario"),
        ("maquina","Máquina / CdG"),
        ("descripcion","Descripción avería"),
        ("notas","Notas reparación"),
        ("hc","H. inicio"),
        ("hf","H. fin"),
        ("delta_horas","Δ Horas"),
        ("estado","Estado"),
        ("familia","Familia avería"),
        ("origen","Origen"),
        ("horas_extra","H. Extra"),
        ("dietas","Dietas"),
        ],
        }
        return render (request ,self .template_name ,context )

    def post (self ,request ,*args ,**kwargs ):
        """
        Dispatches POST actions for the admin history view.

        Supported actions:
          generate_absence_parts — creates synthetic WorkOrders from a WorkerAbsence.
          delete_work_order      — deletes a single WorkOrder by pk.
          bulk_action            — applies mark_reviewed, unmark_reviewed or delete
                                   to a set of PKs.

        ---

        Despacha las acciones POST de la vista de historial de administrador.

        Acciones soportadas:
          generate_absence_parts — crea WorkOrders sinteticos desde un WorkerAbsence.
          delete_work_order      — elimina un WorkOrder individual por pk.
          bulk_action            — aplica mark_reviewed, unmark_reviewed o delete
                                   a un conjunto de PKs.
        """
        from datetime import timedelta ,date as dt_date 
        from django .db import transaction 
        from ivr_config .models import WorkerAbsence 
        from work_order_processor .models import WorkOrder ,WorkOrderEntry ,WorkOrderEntryLine 
        from django .urls import reverse 

        cu =request .user .company_user 
        company =cu .company 

        # Guardia de rol: todas las acciones de este post() son de gestion
        # (eliminar, marcar revisado en bloque, generar partes de ausencia)
        # -- WORKSHOP solo puede editar sus propios partes via
        # operator_form_edit, nunca actuar aqui. El mixin de la clase ahora
        # admite WORKSHOP (para el GET de solo lectura de su propio
        # historial), asi que este guard es imprescindible, no redundante.
        # ---
        # Role guard: every action in this post() is a management action
        # (delete, bulk mark-as-reviewed, generate absence parts) --
        # WORKSHOP may only edit their own parts via operator_form_edit,
        # never act here. The class mixin now admits WORKSHOP (for the
        # read-only GET of their own history), so this guard is essential,
        # not redundant.
        if cu .role not in (
        CompanyUser .ROLE_ADMIN ,
        CompanyUser .ROLE_SUPERVISOR ,
        CompanyUser .ROLE_WORKSHOPBOSS ,
        ):
            from django .http import HttpResponseForbidden 
            return HttpResponseForbidden (
            "Acción no disponible para tu rol."
            )

        action =request .POST .get ("action","").strip ()





        if action =="delete_work_order":
            active_tab =request .POST .get ("active_tab","pending")
            try :
                wo_pk =int (request .POST .get ("work_order_pk",""))
                work_order =WorkOrder .objects .get (
                pk =wo_pk ,
                company =company ,
                source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
                )
                work_order .delete ()
                django_messages .success (request ,f"Parte #{wo_pk} eliminado correctamente.")
            except (ValueError ,TypeError ,WorkOrder .DoesNotExist ):
                django_messages .error (
                request ,
                "Parte no encontrado o no pertenece a esta empresa.",
                )
            return redirect (
            reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
            )





        if action =="bulk_action":
            active_tab =request .POST .get ("active_tab","pending")
            bulk_op =request .POST .get ("bulk_op","").strip ()
            raw_pks =request .POST .getlist ("selected_pks")

            try :
                pk_list =[int (p )for p in raw_pks if p .strip ().isdigit ()]
            except (ValueError ,TypeError ):
                pk_list =[]

            if not pk_list :
                django_messages .warning (request ,"No se ha seleccionado ningún parte.")
                return redirect (
                reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
                )



            qs =WorkOrder .objects .filter (
            pk__in =pk_list ,
            company =company ,
            source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
            )

            if bulk_op =="mark_reviewed":
                from django .utils .timezone import now as tz_now 
                updated =qs .filter (reviewed =False ).update (
                reviewed =True ,
                reviewed_by =cu ,
                reviewed_at =tz_now (),
                )
                django_messages .success (
                request ,
                f"{updated} parte(s) marcado(s) como revisado(s).",
                )
            elif bulk_op =="unmark_reviewed":


                updated =qs .filter (reviewed =True ).update (
                reviewed =False ,
                reviewed_by =None ,
                reviewed_at =None ,
                )
                django_messages .success (
                request ,
                f"{updated} parte(s) desmarcado(s) como revisado(s).",
                )
            elif bulk_op =="delete":
                count =qs .count ()
                qs .delete ()
                django_messages .success (
                request ,
                f"{count} parte(s) eliminado(s) correctamente.",
                )
            else :
                django_messages .error (request ,"Operación en bloque no reconocida.")

            return redirect (
            reverse ("panel:work_order_admin_history")+f"?tab={active_tab}"
            )

        if action !="generate_absence_parts":
            django_messages .error (request ,"Acción no reconocida.")
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        try :
            absence_pk =int (request .POST .get ("absence_pk",""))
            absence =WorkerAbsence .objects .select_related ("company_user__user").get (
            pk =absence_pk ,
            company_user__company =company ,
            )
        except (ValueError ,TypeError ,WorkerAbsence .DoesNotExist ):
            django_messages .error (
            request ,
            "Registro de ausencia no encontrado o no pertenece a esta empresa.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        current_day =absence .start_date 
        working_days =[]
        while current_day <=absence .end_date :
            if current_day .weekday ()<5 :
                working_days .append (current_day )
            current_day +=timedelta (days =1 )

        if not working_days :
            django_messages .warning (
            request ,
            f"El rango de la ausencia ({absence.start_date} – {absence.end_date}) "
            f"no contiene días laborables (lunes a viernes). No se han generado partes.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")

        created_count =0 
        skipped_count =0 

        try :
            with transaction .atomic ():
                for work_day in working_days :


                    already_exists =WorkOrder .objects .filter (
                    company =company ,
                    uploaded_by =absence .company_user ,
                    entries__work_date =work_day ,
                    ).exists ()
                    if already_exists :
                        skipped_count +=1 
                        continue 



                    synthetic_name =(
                    f"AUSENCIA_{absence.get_absence_type_display().upper()}_"
                    f"{work_day.strftime('%Y%m%d')}_"
                    f"{absence.company_user.user.username.upper()}.pdf"
                    )

                    work_order =WorkOrder .objects .create (
                    company =company ,
                    uploaded_by =absence .company_user ,
                    generated_by =cu ,
                    source =WorkOrder .Source .GENERATED ,
                    status =WorkOrder .Status .DONE ,
                    )


                    work_order .source_pdf .name =synthetic_name 
                    work_order .save ()

                    entry =WorkOrderEntry .objects .create (
                    work_order =work_order ,
                    page_number =1 ,
                    worker_name =(
                    absence .company_user .user .get_full_name ()
                    or absence .company_user .user .username 
                    ).upper (),
                    work_date =work_day ,
                    uncertain_date =False ,
                    extraction_confidence =WorkOrderEntry .Confidence .HIGH ,
                    raw_gemini_response =None ,
                    )

                    WorkOrderEntryLine .objects .create (
                    entry =entry ,
                    line_number =1 ,
                    machine_asset =None ,
                    machine_raw ="",
                    machine_norm ="",
                    fault_description =absence .get_absence_type_display (),
                    repair_notes ="",
                    hc =None ,
                    hf =None ,
                    or_val ="",
                    delta_hours =8 ,
                    flags =[],
                    )
                    created_count +=1 

        except Exception as exc :
            logger .error (
            "# [AdminHistory] Error generando partes de ausencia pk=%d: %s",
            absence .pk ,exc ,exc_info =True ,
            )
            django_messages .error (
            request ,
            "Error al generar los partes. "
            "Por favor, inténtalo de nuevo o contacta con el administrador.",
            )
            return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")



        msg_parts =[]
        if created_count :
            msg_parts .append (f"{created_count} parte(s) generado(s) correctamente.")
        if skipped_count :
            msg_parts .append (
            f"{skipped_count} día(s) omitido(s) por existir ya un parte para esa fecha."
            )
        django_messages .success (request ," ".join (msg_parts ))
        return redirect (reverse ("panel:work_order_admin_history")+"?tab=absences")
class WorkPeriodLockView (AdminRoleRequiredMixin ,View ):
    """
    Toggles the is_closed flag of a single WorkPeriod identified by pk.
    ADMIN only.

    POST /panel/work-periods/<pk>/lock/

    Behaviour:
      - If is_closed=False → set is_closed=True  (lock / liquidate).
      - If is_closed=True  → set is_closed=False (unlock / reopen).

    Returns a redirect to the referring URL (HTTP_REFERER) or to
    panel:work_period_list as fallback.
    ---
    Alterna el flag is_closed de un WorkPeriod individual identificado por pk.
    Solo ADMIN.

    POST /panel/work-periods/<pk>/lock/

    Comportamiento:
      - Si is_closed=False → pone is_closed=True  (liquidar).
      - Si is_closed=True  → pone is_closed=False (reabrir).

    Devuelve una redirección a la URL referente (HTTP_REFERER) o a
    panel:work_period_list como fallback.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Toggle is_closed on the WorkPeriod and redirect back.
        ---
        Alterna is_closed en el WorkPeriod y redirige al origen.
        """
        from django .shortcuts import get_object_or_404
        from django .urls import reverse
        from ivr_config .models import WorkPeriod

        cu =request .user .company_user 
        wp =get_object_or_404 (
            WorkPeriod ,
            pk =pk ,
            company_user__company =cu .company ,
        )
        wp .is_closed =not wp .is_closed 
        wp .save (update_fields =["is_closed"])

        if wp .is_closed :
            # Check for unreviewed work orders within this period.
            # Comprobar si hay partes sin revisar dentro de este periodo.
            unreviewed_count = 0
            if wp .end_date :
                unreviewed_count = WorkOrder .objects .filter (
                    company =wp .company_user .company ,
                    uploaded_by =wp .company_user ,
                    source__in =[
                        WorkOrder .Source .DIGITAL ,
                        WorkOrder .Source .GENERATED ,
                    ],
                    reviewed =False ,
                    entries__work_date__gte =wp .start_date ,
                    entries__work_date__lte =wp .end_date ,
                ).distinct ().count ()
            if unreviewed_count :
                django_messages .warning (
                    request ,
                    f"Periodo {wp.start_date:%d/%m/%Y}"
                    f"{'–' + wp.end_date.strftime('%d/%m/%Y') if wp.end_date else ''}"
                    f" liquidado, pero hay {unreviewed_count} parte(s) sin revisar "
                    f"de {wp.company_user.user.get_full_name() or wp.company_user.user.username}.",
                )
            else :
                django_messages .success (
                    request ,
                    f"Periodo {wp.start_date:%d/%m/%Y}"
                    f"{'–' + wp.end_date.strftime('%d/%m/%Y') if wp.end_date else ''}"
                    f" liquidado. Los partes dentro del periodo ya no pueden editarse.",
                )
        else :
            django_messages .success (
                request ,
                f"Periodo {wp.start_date:%d/%m/%Y}"
                f"{'–' + wp.end_date.strftime('%d/%m/%Y') if wp.end_date else ''}"
                f" reabierto. Los partes dentro del periodo pueden editarse de nuevo.",
            )

        referer =request .META .get ("HTTP_REFERER","")
        return redirect (referer or reverse ("panel:work_period_list"))


class WorkOrderAdminExportView (SupervisorAccessMixin ,View ):
    """
    Excel export endpoint scoped exclusively to digital and generated work
    orders (WorkOrder.Source.DIGITAL and WorkOrder.Source.GENERATED).
    Designed for use from WorkOrderAdminHistoryView tab Revisados.

    Supports two export modes via the POST field export_mode:

      single_sheet — One flat sheet with all WorkOrderEntryLine records from
        the selected WorkOrders, grouped by operator name then work date.
        A dark-blue separator row bearing the operator name is inserted before
        each new operator block. Data is read directly from the DB.

      multi_sheet  — One sheet per distinct operator (worker_name derived from
        uploaded_by CompanyUser). Each sheet is built from the individual Excel
        stored in WorkOrder.excel_file (regenerated on-the-fly if missing).
        Sheet title truncated to 31 characters (Excel limit).

    Optional POST filter operator_pk restricts the export to a single operator.

    POST /panel/work-orders/admin-export/
         Body params:
           pks         (list[int]) — WorkOrder primary keys to export.
           export_mode (str)       — single_sheet | multi_sheet.
           operator_pk (int)       — optional operator filter.

    Returns HttpResponse with Content-Disposition attachment (xlsx).
    Returns HTTP 400 on invalid pks or unknown export_mode.
    ---
    Endpoint de exportación Excel exclusivo para partes digitales y generados
    (WorkOrder.Source.DIGITAL y WorkOrder.Source.GENERATED).
    Diseñado para su uso desde la pestaña Revisados de WorkOrderAdminHistoryView.

    Soporta dos modos de exportación via el campo POST export_mode:

      single_sheet — Una hoja plana con todos los WorkOrderEntryLine de los
        WorkOrders seleccionados, agrupados por nombre de operario y fecha.
        Una fila separadora azul oscuro con el nombre del operario se inserta
        antes de cada nuevo bloque de operario. Los datos se leen de la BD.

      multi_sheet  — Una hoja por operario distinto (worker_name derivado del
        CompanyUser uploaded_by). Cada hoja se construye desde el Excel
        individual almacenado en WorkOrder.excel_file (regenerado si falta).
        Título de hoja truncado a 31 caracteres (límite de Excel).

    El filtro POST opcional operator_pk restringe la exportación a un operario.

    POST /panel/work-orders/admin-export/
         Parámetros del cuerpo:
           pks         (list[int]) — claves primarias de WorkOrder a exportar.
           export_mode (str)       — single_sheet | multi_sheet.
           operator_pk (int)       — filtro de operario opcional.

    Devuelve HttpResponse con Content-Disposition attachment (xlsx).
    Devuelve HTTP 400 ante pks inválidos o export_mode desconocido.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Builds the admin export Excel file from the selected WorkOrder pks,
        restricted to DIGITAL and GENERATED sources only.
        Supports single_sheet and multi_sheet export modes.
        ---
        Construye el Excel de exportación admin desde los pks de WorkOrder
        seleccionados, restringido a orígenes DIGITAL y GENERATED únicamente.
        Soporta los modos de exportación single_sheet y multi_sheet.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import WorkOrderEntry ,WorkOrderEntryLine 
        from work_order_processor .services import (
        generate_work_order_excel as _gen_excel ,
        )

        cu =request .user .company_user 
        company =cu .company 





        export_mode =request .POST .get ("export_mode","single_sheet").strip ()
        if export_mode not in ("single_sheet","multi_sheet","digital_full"):
            return HttpResponseBadRequest (
            f"# [ADMIN EXPORT] Modo de exportación desconocido: {export_mode!r}."
            )







        if export_mode =="digital_full":
            return self ._build_digital_full_excel (request ,company ,pk_list ,operator_filter )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (pk )for pk in raw_pks if pk ]
        except (ValueError ,TypeError ):
            return HttpResponseBadRequest ("# [ADMIN EXPORT] Parámetros pks inválidos.")

        if not pk_list :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] No se han seleccionado partes para exportar."
            )





        operator_pk_raw =request .POST .get ("operator_pk","").strip ()
        operator_filter =None 
        if operator_pk_raw :
            try :
                operator_filter =int (operator_pk_raw )
            except (ValueError ,TypeError ):
                operator_filter =None 





        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).order_by ("pk")

        if operator_filter :
            qs =qs .filter (uploaded_by__pk =operator_filter )

        work_orders =list (qs )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] Ninguno de los partes seleccionados es digital/generado, "
            "está revisado y en estado DONE."
            )





        def _get_operator_name (wo ):
            """
            Returns the full name of the operator who submitted the work order,
            or a fallback label if uploaded_by is not set.
            ---
            Devuelve el nombre completo del operario que envió el parte,
            o una etiqueta de reserva si uploaded_by no está establecido.
            """
            if wo .uploaded_by :
                return (
                wo .uploaded_by .user .get_full_name ()
                or wo .uploaded_by .user .username 
                )
            return f"Operario #{wo.pk}"





        def _copy_sheet (src_sheet ,dest_wb ,title ):
            """
            Copies src_sheet (cells, styles, column widths, row heights) into
            a new sheet named title in dest_wb.
            ---
            Copia src_sheet (celdas, estilos, anchos de columna, alturas de
            fila) en una nueva hoja llamada title en dest_wb.
            """
            dest_sheet =dest_wb .create_sheet (title =title [:31 ])
            for row in src_sheet .iter_rows ():
                for cell in row :
                    dest_cell =dest_sheet .cell (
                    row =cell .row ,column =cell .column ,value =cell .value 
                    )
                    if cell .has_style :
                        dest_cell .font =cell .font .copy ()
                        dest_cell .fill =cell .fill .copy ()
                        dest_cell .alignment =cell .alignment .copy ()
                        dest_cell .border =cell .border .copy ()
                        dest_cell .number_format =cell .number_format 
            for col_letter ,col_dim in src_sheet .column_dimensions .items ():
                dest_sheet .column_dimensions [col_letter ].width =col_dim .width 
            for row_num ,row_dim in src_sheet .row_dimensions .items ():
                dest_sheet .row_dimensions [row_num ].height =row_dim .height 
            return dest_sheet 












        if export_mode =="single_sheet":

            wo_map ={wo .pk :wo for wo in work_orders }

            rows =(
            WorkOrderEntryLine .objects 
            .filter (entry__work_order__pk__in =list (wo_map .keys ()))
            .select_related (
            "entry__work_order",
            "entry",
            "machine_asset",
            )
            .order_by (
            "entry__work_order__uploaded_by__user__last_name",
            "entry__work_order__uploaded_by__user__first_name",
            "entry__work_date",
            "entry__work_order__pk",
            "line_number",
            )
            )

            wb =openpyxl .Workbook ()
            ws =wb .active 
            ws .title ="Partes digitales"


            header_fill =PatternFill ("solid",fgColor ="1F3864")
            header_font =Font (bold =True ,color ="FFFFFF",size =10 )
            headers =[
            "Operario","Fecha","Máquina / CdG",
            "Descripción avería","Notas reparación",
            "H. inicio","H. fin","Δ Horas",
            ]
            for col_idx ,h in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =h )
                cell .fill =header_fill 
                cell .font =header_font 
                cell .alignment =Alignment (horizontal ="center",vertical ="center")


            sep_fill =PatternFill ("solid",fgColor ="2F5496")
            sep_font =Font (bold =True ,color ="FFFFFF",size =10 )

            current_row =2 
            current_operator =None 

            for line in rows :
                wo =line .entry .work_order 
                operator_name =_get_operator_name (wo )
                work_date =line .entry .work_date 



                if operator_name !=current_operator :
                    current_operator =operator_name 
                    sep_cell =ws .cell (
                    row =current_row ,column =1 ,value =operator_name 
                    )
                    sep_cell .fill =sep_fill 
                    sep_cell .font =sep_font 
                    sep_cell .alignment =Alignment (vertical ="center")
                    ws .merge_cells (
                    start_row =current_row ,start_column =1 ,
                    end_row =current_row ,end_column =len (headers ),
                    )
                    current_row +=1 

                ws .cell (row =current_row ,column =1 ,value =operator_name )
                ws .cell (row =current_row ,column =2 ,
                value =work_date .strftime ("%d/%m/%Y")if work_date else "")
                ws .cell (row =current_row ,column =3 ,
                value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
                ws .cell (row =current_row ,column =4 ,value =line .fault_description or "")
                ws .cell (row =current_row ,column =5 ,value =line .repair_notes or "")
                ws .cell (row =current_row ,column =6 ,
                value =line .hc .strftime ("%H:%M")if line .hc else "")
                ws .cell (row =current_row ,column =7 ,
                value =line .hf .strftime ("%H:%M")if line .hf else "")
                ws .cell (row =current_row ,column =8 ,
                value =float (line .delta_hours )if line .delta_hours is not None else "")
                current_row +=1 


            for col in ws .columns :
                max_len =max (
                (len (str (cell .value ))for cell in col if cell .value ),
                default =10 ,
                )
                ws .column_dimensions [col [0 ].column_letter ].width =min (max_len +4 ,60 )

            buf =io .BytesIO ()
            wb .save (buf )
            buf .seek (0 )

            filename =f"partes_digitales_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response =HttpResponse (
            buf .read (),
            content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response ["Content-Disposition"]=f'attachment; filename="{filename}"'
            return response 












        dest_wb =openpyxl .Workbook ()
        dest_wb .remove (dest_wb .active )



        from collections import defaultdict 
        operator_groups =defaultdict (list )
        for wo in work_orders :
            operator_groups [_get_operator_name (wo )].append (wo )

        for operator_name ,wo_list in sorted (operator_groups .items ()):
            sheet_title =operator_name [:31 ]
            sheet_added =False 

            for wo in wo_list :


                if not wo .excel_file or not wo .excel_file .name :
                    try :
                        _gen_excel (wo .pk )
                        wo .refresh_from_db (fields =["excel_file"])
                    except Exception :
                        logger .warning (
                        "# [AdminExport] No se pudo regenerar Excel para WorkOrder #%d.",
                        wo .pk ,
                        )
                        continue 

                try :
                    src_wb =openpyxl .load_workbook (wo .excel_file .path )
                    src_sheet =src_wb .worksheets [0 ]
                    if not sheet_added :
                        _copy_sheet (src_sheet ,dest_wb ,sheet_title )
                        sheet_added =True 
                    else :


                        dest_sheet =dest_wb [sheet_title ]
                        start_row =dest_sheet .max_row +1 
                        for row in src_sheet .iter_rows (min_row =2 ):
                            for cell in row :
                                dest_cell =dest_sheet .cell (
                                row =start_row +cell .row -2 ,
                                column =cell .column ,
                                value =cell .value ,
                                )
                                if cell .has_style :
                                    dest_cell .font =cell .font .copy ()
                                    dest_cell .fill =cell .fill .copy ()
                                    dest_cell .alignment =cell .alignment .copy ()
                except Exception as exc :
                    logger .warning (
                    "# [AdminExport] Error procesando Excel WorkOrder #%d: %s",
                    wo .pk ,exc ,
                    )
                    continue 

        if not dest_wb .worksheets :
            return HttpResponseBadRequest (
            "# [ADMIN EXPORT] No se pudo generar ninguna hoja Excel. "
            "Verifica que los partes seleccionados tienen Excel generado."
            )

        buf =io .BytesIO ()
        dest_wb .save (buf )
        buf .seek (0 )

        filename =f"partes_digitales_multi_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response =HttpResponse (
        buf .read (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


    def _build_digital_full_excel (self ,request ,company ,pk_list ,operator_filter ):
        """
        Builds a three-sheet Excel file for digital and generated work orders.

        Sheet 1 — Tareas: one row per WorkOrderEntryLine with all digital fields.
        Sheet 2 — Repuestos: one row per SparePartLine with supplier and price data.
        Sheet 3 — Incidencias de jornada: one row per WorkdayGap with resolution detail.

        Returns HttpResponse with Content-Disposition attachment (xlsx).
        Returns HTTP 400 if no valid digital/generated work orders are found.
        ---
        Construye un Excel de tres hojas para partes digitales y generados.

        Hoja 1 — Tareas: una fila por WorkOrderEntryLine con todos los campos digitales.
        Hoja 2 — Repuestos: una fila por SparePartLine con datos de proveedor y precio.
        Hoja 3 — Incidencias de jornada: una fila por WorkdayGap con detalle de resolucion.

        Devuelve HttpResponse con Content-Disposition attachment (xlsx).
        Devuelve HTTP 400 si no se encuentran partes digitales/generados validos.
        """
        import io 
        import openpyxl 
        from openpyxl .styles import Alignment ,Font ,PatternFill 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import (
        WorkOrderEntry ,WorkOrderEntryLine ,SparePartLine ,WorkdayGap ,
        )



        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        status =WorkOrder .Status .DONE ,
        reviewed =True ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).order_by ("pk")

        if operator_filter :
            qs =qs .filter (uploaded_by__pk =operator_filter )

        work_orders =list (qs )

        if not work_orders :
            return HttpResponseBadRequest (
            "# [DIGITAL FULL EXPORT] Ninguno de los partes seleccionados es "
            "digital/generado, esta revisado y en estado DONE."
            )

        wo_pks =[wo .pk for wo in work_orders ]



        def _op_name (wo ):
            if wo .uploaded_by :
                return (
                wo .uploaded_by .user .get_full_name ()
                or wo .uploaded_by .user .username 
                )
            return f"Operario #{wo.pk}"



        def _fmt_time (t ):
            return t .strftime ("%H:%M")if t else ""



        def _fmt_date (d ):
            return d .strftime ("%d/%m/%Y")if d else ""



        wo_map ={}
        entries =(
        WorkOrderEntry .objects 
        .filter (work_order__in =wo_pks )
        .select_related ("work_order__uploaded_by__user")
        )
        for entry in entries :
            wo_map [entry .work_order_id ]=(entry .work_order ,entry )


        header_fill =PatternFill ("solid",fgColor ="1F4E79")
        header_font =Font (bold =True ,color ="FFFFFF",size =10 )

        def _apply_header (ws ,headers ):
            for col_idx ,h in enumerate (headers ,start =1 ):
                cell =ws .cell (row =1 ,column =col_idx ,value =h )
                cell .fill =header_fill 
                cell .font =header_font 
                cell .alignment =Alignment (horizontal ="center",vertical ="center")
            ws .row_dimensions [1 ].height =18 

        def _autofit (ws ):
            for col in ws .columns :
                max_len =max (
                (len (str (cell .value ))for cell in col if cell .value ),
                default =10 ,
                )
                ws .column_dimensions [col [0 ].column_letter ].width =min (max_len +4 ,60 )

        wb =openpyxl .Workbook ()




        ws_tasks =wb .active 
        ws_tasks .title ="Tareas"
        task_headers =[
        "Operario","Fecha","Maquina / CdG","O.R.",
        "Descripcion averia","Notas reparacion",
        "H. inicio","H. fin","Delta Horas",
        "Km (odometro)","Horometro motor","Horometro grua",
        "Categoria averia","Subcategoria averia",
        ]
        _apply_header (ws_tasks ,task_headers )

        lines =(
        WorkOrderEntryLine .objects 
        .filter (entry__work_order__in =wo_pks )
        .select_related (
        "entry__work_order__uploaded_by__user",
        "machine_asset",
        )
        .order_by (
        "entry__work_order__uploaded_by__user__last_name",
        "entry__work_date",
        "line_number",
        )
        )

        row =2 
        for line in lines :
            entry =line .entry 
            wo =entry .work_order 
            ws_tasks .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_tasks .cell (row =row ,column =2 ,value =_fmt_date (entry .work_date ))
            ws_tasks .cell (row =row ,column =3 ,
            value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
            ws_tasks .cell (row =row ,column =4 ,value =line .or_val or "")
            ws_tasks .cell (row =row ,column =5 ,value =line .fault_description or "")
            ws_tasks .cell (row =row ,column =6 ,value =line .repair_notes or "")
            ws_tasks .cell (row =row ,column =7 ,value =_fmt_time (line .hc ))
            ws_tasks .cell (row =row ,column =8 ,value =_fmt_time (line .hf ))
            ws_tasks .cell (row =row ,column =9 ,
            value =float (line .delta_hours )if line .delta_hours is not None else "")
            ws_tasks .cell (row =row ,column =10 ,
            value =float (line .odometer_reading )if line .odometer_reading is not None else "")
            ws_tasks .cell (row =row ,column =11 ,
            value =float (line .engine_hours_reading )if line .engine_hours_reading is not None else "")
            ws_tasks .cell (row =row ,column =12 ,
            value =float (line .crane_hours_reading )if line .crane_hours_reading is not None else "")
            ws_tasks .cell (row =row ,column =13 ,
            value =line .get_fault_category_display ()if line .fault_category else "")
            ws_tasks .cell (row =row ,column =14 ,
            value =line .get_fault_subcategory_display ()if line .fault_subcategory else "")
            row +=1 

        _autofit (ws_tasks )




        ws_parts =wb .create_sheet (title ="Repuestos")
        parts_headers =[
        "Operario","Fecha","Maquina / CdG",
        "Referencia","Material","Cantidad",
        "Procedencia","Proveedor","Precio unitario",
        ]
        _apply_header (ws_parts ,parts_headers )

        spare_lines =(
        SparePartLine .objects 
        .filter (entry_line__entry__work_order__in =wo_pks )
        .select_related (
        "entry_line__entry__work_order__uploaded_by__user",
        "entry_line__machine_asset",
        "vehicle",
        )
        .order_by (
        "entry_line__entry__work_order__uploaded_by__user__last_name",
        "entry_line__entry__work_date",
        "entry_line__line_number",
        "line_number",
        )
        )

        row =2 
        for spare in spare_lines :
            line =spare .entry_line 
            entry =line .entry 
            wo =entry .work_order 
            ws_parts .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_parts .cell (row =row ,column =2 ,value =_fmt_date (entry .work_date ))
            ws_parts .cell (row =row ,column =3 ,
            value =line .machine_asset .code if line .machine_asset else line .machine_raw or "")
            ws_parts .cell (row =row ,column =4 ,value =spare .reference or "")
            ws_parts .cell (row =row ,column =5 ,value =spare .material or "")
            ws_parts .cell (row =row ,column =6 ,
            value =float (spare .quantity )if spare .quantity is not None else "")
            ws_parts .cell (row =row ,column =7 ,value =spare .get_source_display ())
            ws_parts .cell (row =row ,column =8 ,value =spare .supplier or "")
            ws_parts .cell (row =row ,column =9 ,
            value =float (spare .unit_price )if spare .unit_price is not None else "")
            row +=1 

        _autofit (ws_parts )




        ws_gaps =wb .create_sheet (title ="Incidencias de jornada")
        gaps_headers =[
        "Operario","Fecha","Tipo","Inicio","Fin",
        "Duracion (min)","Categoria ausencia",
        "Ha comido","Hora comida","Nota","Resuelto",
        ]
        _apply_header (ws_gaps ,gaps_headers )

        gaps =(
        WorkdayGap .objects 
        .filter (work_order__in =wo_pks )
        .select_related (
        "work_order__uploaded_by__user",
        "work_order__entries",
        "absence_category",
        )
        .prefetch_related ("work_order__entries")
        .order_by (
        "work_order__uploaded_by__user__last_name",
        "gap_start",
        )
        )

        row =2 
        for gap in gaps :
            wo =gap .work_order 
            entry =wo .entries .first ()
            ws_gaps .cell (row =row ,column =1 ,value =_op_name (wo ))
            ws_gaps .cell (row =row ,column =2 ,
            value =_fmt_date (entry .work_date )if entry else "")
            ws_gaps .cell (row =row ,column =3 ,value =gap .get_gap_type_display ())
            ws_gaps .cell (row =row ,column =4 ,value =_fmt_time (gap .gap_start ))
            ws_gaps .cell (row =row ,column =5 ,value =_fmt_time (gap .gap_end ))
            ws_gaps .cell (row =row ,column =6 ,value =gap .duration_minutes )
            ws_gaps .cell (row =row ,column =7 ,
            value =gap .absence_category .label if gap .absence_category else "")


            if gap .gap_type ==WorkdayGap .GapType .LUNCH_BREAK :
                ws_gaps .cell (row =row ,column =8 ,
                value ="Si"if gap .lunch_had else ("No"if gap .lunch_had is False else ""))
                ws_gaps .cell (row =row ,column =9 ,value =_fmt_time (gap .lunch_time ))
            else :
                ws_gaps .cell (row =row ,column =8 ,value ="")
                ws_gaps .cell (row =row ,column =9 ,value ="")
            ws_gaps .cell (row =row ,column =10 ,value =gap .note or "")
            ws_gaps .cell (row =row ,column =11 ,value ="Si"if gap .resolved else "No")
            row +=1 

        _autofit (ws_gaps )

        buf =io .BytesIO ()
        wb .save (buf )
        buf .seek (0 )

        filename =f"partes_digitales_completo_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response =HttpResponse (
        buf .read (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        return response 


class WorkOrderMachineFilterView (SupervisorAccessMixin ,View ):
    """
    JSON endpoint returning distinct MachineAsset codes present in the
    WorkOrderEntryLine records of DIGITAL/GENERATED WorkOrders for the
    authenticated company, optionally filtered by operator and date range.
    Used by the admin history machine autocomplete (Bug B fix).

    GET /panel/work-orders/machines/
        Optional GET params:
          operator_pk (int)  — filter by uploaded_by CompanyUser pk.
          date_from   (str)  — ISO date YYYY-MM-DD start of range.
          date_to     (str)  — ISO date YYYY-MM-DD end of range.
        Returns: {"results": ["G12", "A44", ...]}

    Accessible to SUPERVISOR and ADMIN roles (SupervisorAccessMixin).

    ---

    Endpoint JSON que devuelve los códigos de MachineAsset distintos presentes
    en los WorkOrderEntryLine de los WorkOrders DIGITAL/GENERATED de la empresa
    autenticada, con filtro opcional por operario y rango de fechas.
    Usado por el autocompletado de máquina de admin_history (corrección Bug B).

    GET /panel/work-orders/machines/
        Parámetros GET opcionales:
          operator_pk (int)  — filtrar por pk de CompanyUser uploaded_by.
          date_from   (str)  — fecha ISO YYYY-MM-DD inicio del rango.
          date_to     (str)  — fecha ISO YYYY-MM-DD fin del rango.
        Devuelve: {"results": ["G12", "A44", ...]}

    Accesible para los roles SUPERVISOR y ADMIN (SupervisorAccessMixin).
    """

    def get (self ,request ,*args ,**kwargs ):
        """
        Returns distinct MachineAsset codes present in the filtered WorkOrders.
        ---
        Devuelve los códigos de MachineAsset distintos en los WorkOrders filtrados.
        """
        from django .http import JsonResponse 
        from datetime import datetime as _dt_mf 
        from work_order_processor .models import WorkOrderEntryLine 

        company =request .user .company_user .company 
        operator_pk =request .GET .get ("operator_pk","").strip ()
        date_from_raw =request .GET .get ("date_from","").strip ()
        date_to_raw =request .GET .get ("date_to","").strip ()
        q_raw =request .GET .get ("q","").strip ()

        def _parse_iso (val ):
            """Parses YYYY-MM-DD string, returns date or None.
            --- Parsea cadena YYYY-MM-DD, devuelve date o None."""
            if not val :
                return None 
            try :
                return _dt_mf .strptime (val ,"%Y-%m-%d").date ()
            except ValueError :
                return None 

        date_from =_parse_iso (date_from_raw )
        date_to =_parse_iso (date_to_raw )



        qs =(
        WorkOrderEntryLine .objects 
        .filter (
        entry__work_order__company =company ,
        entry__work_order__source__in =[
        WorkOrder .Source .DIGITAL ,
        WorkOrder .Source .GENERATED ,
        ],
        machine_asset__isnull =False ,
        )
        )


        if operator_pk :
            try :
                qs =qs .filter (
                entry__work_order__uploaded_by__pk =int (operator_pk ),
                entry__work_order__uploaded_by__company =company ,
                )
            except (ValueError ,TypeError ):
                pass 


        if date_from :
            qs =qs .filter (entry__work_date__gte =date_from )
        if date_to :
            qs =qs .filter (entry__work_date__lte =date_to )


        if q_raw :
            qs =qs .filter (machine_asset__code__icontains =q_raw )

        codes =(
        qs 
        .values_list ("machine_asset__code",flat =True )
        .distinct ()
        .order_by ("machine_asset__code")
        )

        return JsonResponse ({"results":list (codes )})





class ExportTemplateListView (SupervisorAccessMixin ,View ):
    """
    Lists all ExportTemplate records belonging to the authenticated user.
    Returns a JSON response suitable for AJAX calls from the export modal,
    or renders a full HTML page for standalone template management.
    GET ?format=json — returns [{id, name, is_default, columns, sheet_format,
                                   operator_scope}] for the modal.
    GET             — renders panel/export_templates/list.html.
    ---
    Lista todos los registros ExportTemplate del usuario autenticado.
    Devuelve JSON para llamadas AJAX desde el modal de exportación,
    o renderiza una página HTML para la gestión autónoma de plantillas.
    GET ?format=json — lista [{id, name, is_default, ...}] para el modal.
    GET             — renderiza panel/export_templates/list.html.
    """

    template_name ="panel/export_templates/list.html"

    def get(self, request, *args, **kwargs):
        """
        Returns the user's export templates (personal + global) as JSON or HTML.
        Auto-creates the default personal template if the user has none.
        JSON response includes is_global flag so the frontend can distinguish both types.
        ---
        Devuelve las plantillas del usuario (personales + globales) como JSON o HTML.
        Crea la plantilla personal por defecto si el usuario no tiene ninguna.
        La respuesta JSON incluye is_global para que el frontend distinga ambos tipos.
        """
        from django.http import JsonResponse
        from work_order_processor.models import ExportTemplate

        cu = request.user.company_user
        company = cu.company

        ExportTemplate.get_or_create_default(cu)

        templates_propias = ExportTemplate.objects.filter(
            company_user=cu,
            is_global=False,
        ).order_by("-is_default", "name")

        templates_globales = ExportTemplate.objects.filter(
            company=company,
            is_global=True,
        ).order_by("name")

        if request.GET.get("format") == "json":
            def _serialize(t, is_global=False):
                return {
                    "id": t.pk,
                    "name": t.name,
                    "is_default": t.is_default,
                    "is_global": is_global,
                    "columns": t.columns,
                    "sheet_format": t.sheet_format,
                    "operator_scope": t.operator_scope,
                }
            data = (
                [_serialize(t, is_global=False) for t in templates_propias]
                + [_serialize(t, is_global=True) for t in templates_globales]
            )
            return JsonResponse({"templates": data})

        _is_admin = cu.role == "ADMIN"
        context = {
            "company": company,
            "company_user": cu,
            "own_presence": PresenceStatus.objects.filter(
                company_user=cu,
                starts_at__lte=now(),
            ).filter(
                Q(ends_at__isnull=True) | Q(ends_at__gt=now())
            ).order_by("-starts_at").first(),
            "active_nav": "work_order_admin_history",
            "templates_propias": templates_propias,
            "templates_globales": templates_globales,
            "is_admin": _is_admin,
            "column_choices": [
                ("fecha", "Fecha"),
                ("operario", "Operario"),
                ("maquina", "Máquina / CdG"),
                ("descripcion", "Descripción avería"),
                ("notas", "Notas reparación"),
                ("hc", "H. inicio"),
                ("hf", "H. fin"),
                ("delta_horas", "Δ Horas"),
                ("estado", "Estado"),
                ("familia", "Familia avería"),
                ("origen", "Origen"),
                ("horas_extra", "H. Extra"),
                ("dietas", "Dietas"),
            ],
        }
        return render(request, self.template_name, context)


class ExportTemplateCreateView (SupervisorAccessMixin ,View ):
    """
    Creates a new ExportTemplate for the authenticated user.
    Accepts POST with JSON body or form data.
    Returns JSON {id, name} on success or {error} on failure.
    ---
    Crea una nueva ExportTemplate para el usuario autenticado.
    Acepta POST con cuerpo JSON o datos de formulario.
    Devuelve JSON {id, name} en éxito o {error} en fallo.
    """

    def post(self, request, *args, **kwargs):
        """
        Validates and creates the ExportTemplate.
        If the user is ADMIN and body includes is_global=true, creates a global
        template linked to the company (company_user=None).
        Otherwise creates a personal template for the authenticated user.
        ---
        Valida y crea la ExportTemplate.
        Si el usuario es ADMIN y el body incluye is_global=true, crea una plantilla
        global ligada a la empresa (company_user=None).
        En caso contrario, crea una plantilla personal del usuario autenticado.
        """
        import json as _json
        from django.http import JsonResponse
        from work_order_processor.models import ExportTemplate

        cu = request.user.company_user
        company = cu.company

        try:
            body = _json.loads(request.body)
        except (ValueError, TypeError):
            body = request.POST

        name = str(body.get("name", "")).strip()
        columns = body.get("columns", [])
        sheet_format = str(body.get("sheet_format", ExportTemplate.SheetFormat.SINGLE_SHEET)).strip()
        operator_scope = str(body.get("operator_scope", ExportTemplate.OperatorScope.ALL)).strip()
        is_default = bool(body.get("is_default", False))
        want_global = bool(body.get("is_global", False)) and cu.role == "ADMIN"

        if not name:
            return JsonResponse({"error": "El nombre es obligatorio."}, status=400)
        if not columns:
            return JsonResponse({"error": "Selecciona al menos una columna."}, status=400)
        if sheet_format not in ExportTemplate.SheetFormat.values:
            return JsonResponse({"error": "Formato de hoja no válido."}, status=400)
        if operator_scope not in ExportTemplate.OperatorScope.values:
            return JsonResponse({"error": "Alcance de operarios no válido."}, status=400)

        if want_global:
            if ExportTemplate.objects.filter(company=company, is_global=True, name=name).exists():
                return JsonResponse(
                    {"error": f"Ya existe una plantilla global con el nombre '{name}'."},
                    status=400,
                )
            template = ExportTemplate.objects.create(
                company_user=None,
                company=company,
                is_global=True,
                name=name,
                columns=list(columns),
                sheet_format=sheet_format,
                operator_scope=operator_scope,
                is_default=False,
            )
        else:
            if ExportTemplate.objects.filter(company_user=cu, is_global=False, name=name).exists():
                return JsonResponse(
                    {"error": f"Ya existe una plantilla con el nombre '{name}'."},
                    status=400,
                )
            template = ExportTemplate.objects.create(
                company_user=cu,
                company=None,
                is_global=False,
                name=name,
                columns=list(columns),
                sheet_format=sheet_format,
                operator_scope=operator_scope,
                is_default=is_default,
            )
        logger.info(
            "# [EXPORT TEMPLATE] Plantilla pk=%s '%s' creada por %s (global=%s).",
            template.pk, template.name, cu.user.username, want_global,
        )
        return JsonResponse({"id": template.pk, "name": template.name, "is_global": want_global}, status=201)


class ExportTemplateUpdateView (SupervisorAccessMixin ,View ):
    """
    Updates an existing ExportTemplate belonging to the authenticated user.
    Returns JSON {ok: true} on success or {error} on failure.
    ---
    Actualiza una ExportTemplate existente del usuario autenticado.
    Devuelve JSON {ok: true} en éxito o {error} en fallo.
    """

    def post(self, request, pk, *args, **kwargs):
        """
        Validates and applies the update.
        - Personal template: updated directly (must belong to the user).
        - Global template + ADMIN: updated directly.
        - Global template + non-ADMIN (SUPERVISOR): a personal copy is created
          with name + " (copia)" instead of editing the original.
        ---
        Valida y aplica la actualización.
        - Plantilla personal: se actualiza directamente (debe pertenecer al usuario).
        - Plantilla global + ADMIN: se actualiza directamente.
        - Plantilla global + no-ADMIN (SUPERVISOR): se crea una copia personal con
          nombre + " (copia)" en lugar de editar la original.
        """
        import json as _json
        from django.http import JsonResponse
        from work_order_processor.models import ExportTemplate

        cu = request.user.company_user
        company = cu.company

        # Resolve the template — either personal or global of the same company
        try:
            template = ExportTemplate.objects.get(pk=pk, company_user=cu, is_global=False)
            is_own_personal = True
        except ExportTemplate.DoesNotExist:
            is_own_personal = False
            try:
                template = ExportTemplate.objects.get(pk=pk, company=company, is_global=True)
            except ExportTemplate.DoesNotExist:
                return JsonResponse({"error": "Plantilla no encontrada."}, status=404)

        try:
            body = _json.loads(request.body)
        except (ValueError, TypeError):
            body = request.POST

        name = str(body.get("name", template.name)).strip()
        columns = body.get("columns", template.columns)
        sheet_format = str(body.get("sheet_format", template.sheet_format)).strip()
        operator_scope = str(body.get("operator_scope", template.operator_scope)).strip()
        is_default = bool(body.get("is_default", template.is_default))

        if not name:
            return JsonResponse({"error": "El nombre es obligatorio."}, status=400)
        if not columns:
            return JsonResponse({"error": "Selecciona al menos una columna."}, status=400)
        if sheet_format not in ExportTemplate.SheetFormat.values:
            return JsonResponse({"error": "Formato de hoja no válido."}, status=400)
        if operator_scope not in ExportTemplate.OperatorScope.values:
            return JsonResponse({"error": "Alcance de operarios no válido."}, status=400)

        # Global template + non-ADMIN → create personal copy, do not touch original
        if template.is_global and cu.role != "ADMIN":
            copy_name = f"{name} (copia)"
            # If a personal copy with that name already exists, just return it
            existing_copy = ExportTemplate.objects.filter(
                company_user=cu, is_global=False, name=copy_name
            ).first()
            if existing_copy:
                return JsonResponse({"ok": True, "copied": True, "id": existing_copy.pk, "name": existing_copy.name})
            copy = ExportTemplate.objects.create(
                company_user=cu,
                company=None,
                is_global=False,
                name=copy_name,
                columns=list(columns),
                sheet_format=sheet_format,
                operator_scope=operator_scope,
                is_default=False,
            )
            logger.info(
                "# [EXPORT TEMPLATE] Copia personal pk=%s '%s' creada por %s desde global pk=%s.",
                copy.pk, copy.name, cu.user.username, template.pk,
            )
            return JsonResponse({"ok": True, "copied": True, "id": copy.pk, "name": copy.name})

        # Personal template or global + ADMIN → update directly
        if is_own_personal and name != template.name:
            if ExportTemplate.objects.filter(company_user=cu, is_global=False, name=name).exists():
                return JsonResponse(
                    {"error": f"Ya existe una plantilla con el nombre '{name}'."},
                    status=400,
                )
        if template.is_global and name != template.name:
            if ExportTemplate.objects.filter(company=company, is_global=True, name=name).exists():
                return JsonResponse(
                    {"error": f"Ya existe una plantilla global con el nombre '{name}'."},
                    status=400,
                )

        template.name = name
        template.columns = list(columns)
        template.sheet_format = sheet_format
        template.operator_scope = operator_scope
        if not template.is_global:
            template.is_default = is_default
        template.save()
        logger.info(
            "# [EXPORT TEMPLATE] Plantilla pk=%s '%s' actualizada por %s.",
            template.pk, template.name, cu.user.username,
        )
        return JsonResponse({"ok": True})


class ExportTemplateDeleteView (SupervisorAccessMixin ,View ):
    """
    Deletes an ExportTemplate belonging to the authenticated user.
    Returns JSON {ok: true} on success or {error} on failure.
    Cannot delete the last remaining template of a user.
    ---
    Elimina una ExportTemplate del usuario autenticado.
    Devuelve JSON {ok: true} en éxito o {error} en fallo.
    No permite eliminar la última plantilla del usuario.
    """

    def post (self ,request ,pk ,*args ,**kwargs ):
        """
        Deletes the template and redirects to the list.
        ---
        Elimina la plantilla y redirige a la lista.
        """
        from django .http import JsonResponse 
        from work_order_processor .models import ExportTemplate 

        cu =request .user .company_user 

        try :
            template =ExportTemplate .objects .get (pk =pk ,company_user =cu )
        except ExportTemplate .DoesNotExist :
            return JsonResponse ({"error":"Plantilla no encontrada."},status =404 )

        if ExportTemplate .objects .filter (company_user =cu ).count ()<=1 :
            return JsonResponse (
            {"error":"No puedes eliminar tu única plantilla de exportación."},
            status =400 ,
            )

        name =template .name 
        template .delete ()
        logger .info (
        "# [EXPORT TEMPLATE] Plantilla '%s' eliminada por %s.",
        name ,cu .user .username ,
        )
        return JsonResponse ({"ok":True })


class WorkOrderAdminExportByTemplateView (SupervisorAccessMixin ,View ):
    """
    Generates and streams an Excel file from selected WorkOrder PKs
    using the configuration of the chosen ExportTemplate.

    POST /panel/work-orders/export-by-template/
         Body params:
           template_pk  (int)       — ExportTemplate pk (must belong to user).
           pks          (list[int]) — WorkOrder primary keys to export.
           operator_pks (list[int]) — optional operator filter when
                                       template.operator_scope == 'selection'.

    Returns HttpResponse with Content-Disposition attachment (xlsx).
    Returns HTTP 400 on invalid input or HTTP 404 on unknown template.
    ---
    Genera y devuelve en streaming un Excel desde los PKs de WorkOrder
    seleccionados usando la configuración de la ExportTemplate elegida.

    POST /panel/work-orders/export-by-template/
         Parámetros del cuerpo:
           template_pk  (int)       — pk de ExportTemplate (debe pertenecer al usuario).
           pks          (list[int]) — claves primarias de WorkOrder a exportar.
           operator_pks (list[int]) — filtro de operario opcional cuando
                                       template.operator_scope == 'selection'.

    Devuelve HttpResponse con Content-Disposition attachment (xlsx).
    Devuelve HTTP 400 en entrada inválida o HTTP 404 en plantilla desconocida.
    """

    def post (self ,request ,*args ,**kwargs ):
        """
        Resolves the template and work orders, calls build_export_from_template
        and streams the resulting workbook as an xlsx attachment.
        ---
        Resuelve la plantilla y los partes, llama a build_export_from_template
        y devuelve el libro resultante como adjunto xlsx.
        """
        import io 
        from django .http import HttpResponse ,HttpResponseBadRequest 
        from django .utils .timezone import now as tz_now 
        from work_order_processor .models import ExportTemplate 
        from work_order_processor .services import build_export_from_template 

        cu =request .user .company_user 
        company =cu .company 





        try :
            template_pk =int (request .POST .get ("template_pk",""))
            # Resolve template: personal (company_user=cu) OR global (company, is_global=True)
            # Resolver plantilla: personal (company_user=cu) O global (company, is_global=True)
            try :
                template =ExportTemplate .objects .get (pk =template_pk ,company_user =cu ,is_global =False )
            except ExportTemplate .DoesNotExist :
                template =ExportTemplate .objects .get (pk =template_pk ,company =company ,is_global =True )
        except (ValueError ,TypeError ,ExportTemplate .DoesNotExist ):
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] Plantilla no encontrada o no pertenece al usuario."
            )





        raw_pks =request .POST .getlist ("pks")
        try :
            pk_list =[int (p )for p in raw_pks if str (p ).strip ().isdigit ()]
        except (ValueError ,TypeError ):
            pk_list =[]

        if not pk_list :
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] No se han seleccionado partes para exportar."
            )





        # Note: prefetch_related is intentionally omitted here.
        # build_export_from_template builds its own WorkOrderEntryLine
        # queryset internally. Passing a queryset with prefetch annotations
        # as a subquery causes Django to emit duplicate rows.
        #
        # Nota: prefetch_related se omite intencionadamente aquí.
        # build_export_from_template construye su propio queryset de
        # WorkOrderEntryLine internamente. Pasar un queryset con anotaciones
        # prefetch como subquery hace que Django emita filas duplicadas.
        qs =WorkOrder .objects .filter (
        pk__in =pk_list ,
        company =company ,
        source__in =[WorkOrder .Source .DIGITAL ,WorkOrder .Source .GENERATED ],
        ).order_by (
        "uploaded_by__user__last_name",
        "uploaded_by__user__first_name",
        "pk",
        )





        if template .operator_scope =="selection":
            raw_op_pks =request .POST .getlist ("operator_pks")
            try :
                op_pk_list =[int (p )for p in raw_op_pks if str (p ).strip ().isdigit ()]
            except (ValueError ,TypeError ):
                op_pk_list =[]
            if op_pk_list :
                qs =qs .filter (uploaded_by__pk__in =op_pk_list )

        if not qs .exists ():
            return HttpResponseBadRequest (
            "# [EXPORT BY TEMPLATE] Ninguno de los partes seleccionados es válido para exportar."
            )





        wb =build_export_from_template (template ,qs )
        buf =io .BytesIO ()
        wb .save (buf )
        buf .seek (0 )

        filename =(
        f"partes_digitales_{tz_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        response =HttpResponse (
        buf .getvalue (),
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response ["Content-Disposition"]=f'attachment; filename="{filename}"'
        logger .info (
        "# [EXPORT BY TEMPLATE] Exportación '%s' (%d partes) por %s.",
        template .name ,qs .count (),cu .user .username ,
        )
        return response
# ===========================================================================
# WORK PERIOD GROUP — Vistas del nuevo modelo WorkPeriodGroup (Opción B)
# Arquitectura: el periodo es la entidad primaria de la UI. Los operarios
# son subordinados al grupo.
# ===========================================================================
class WorkPeriodGroupDetailView(SupervisorAccessMixin, View):
    """
    Shows detail of a single WorkPeriodGroup: header data, list of assigned
    operators with their individual WorkPeriod status, and list of active
    WORKSHOP operators NOT yet assigned to this group (available to add).

    GET /panel/work-period-groups/<pk>/
    ---
    Muestra el detalle de un WorkPeriodGroup: datos de cabecera, lista de
    operarios asignados con su estado WorkPeriod individual, y lista de
    operarios WORKSHOP activos aún NO asignados al grupo (disponibles
    para añadir).

    GET /panel/work-period-groups/<pk>/
    """

    template_name = "panel/work_orders/work_period_detail.html"

    def get(self, request, pk, *args, **kwargs):
        """
        Resolves the group, builds assigned and available operator lists,
        and renders the detail template.
        ---
        Resuelve el grupo, construye las listas de operarios asignados y
        disponibles, y renderiza el template de detalle.
        """
        from django.shortcuts import get_object_or_404
        from ivr_config.models import WorkPeriodGroup

        cu = request.user.company_user
        company = cu.company

        group = get_object_or_404(
            WorkPeriodGroup, pk=pk, company=company
        )

        # Operators already assigned to this group.
        # Operarios ya asignados a este grupo.
        assigned_periods = list(
            group.operator_periods
            .select_related("company_user__user")
            .order_by(
                "company_user__user__last_name",
                "company_user__user__first_name",
            )
        )
        assigned_cu_pks = {wp.company_user_id for wp in assigned_periods}

        # Active WORKSHOP operators NOT yet assigned to this group.
        # Operarios WORKSHOP activos aún NO asignados a este grupo.
        available_operators = list(
            CompanyUser.objects
            .filter(company=company, is_active=True, role=CompanyUser.ROLE_WORKSHOP)
            .exclude(pk__in=assigned_cu_pks)
            .select_related("user")
            .order_by("user__last_name", "user__first_name")
        )

        context = {
            "company":             company,
            "company_user":        cu,
            "active_nav":          "work_period_list",
            "group":               group,
            "assigned_periods":    assigned_periods,
            "available_operators": available_operators,
            "is_admin":            cu.role == "ADMIN",
        }
        return render(request, self.template_name, context)


class WorkPeriodGroupCreateView(SupervisorAccessMixin, View):
    """
    Creates a new WorkPeriodGroup and, optionally, individual WorkPeriod
    records for selected (or all) active WORKSHOP operators.

    POST /panel/work-period-groups/create/
         label       (str, required)
         start_date  (YYYY-MM-DD, required)
         end_date    (YYYY-MM-DD, optional)
         operator_pks (int list, optional — if absent, targets all WORKSHOP)
    ---
    Crea un nuevo WorkPeriodGroup y, opcionalmente, registros WorkPeriod
    individuales para los operarios WORKSHOP seleccionados (o todos).

    POST /panel/work-period-groups/create/
    """

    def post(self, request, *args, **kwargs):
        """
        Validates POST data, creates the group and individual WorkPeriod
        records, and redirects to the new group's detail page.
        ---
        Valida los datos POST, crea el grupo y los registros WorkPeriod
        individuales, y redirige al detalle del nuevo grupo.
        """
        from datetime import datetime
        from django.urls import reverse
        from ivr_config.models import WorkPeriodGroup, WorkPeriod

        cu = request.user.company_user
        company = cu.company
        LIST_URL = (
            reverse("panel:work_order_admin_history") + "?tab=periods"
        )

        # -- Parse label (required). --
        label = request.POST.get("label", "").strip()
        if not label:
            django_messages.error(
                request,
                "La etiqueta del periodo es obligatoria.",
            )
            return redirect(LIST_URL)

        # -- Parse start_date (required). --
        raw_start = request.POST.get("start_date", "").strip()
        try:
            start_date = datetime.strptime(raw_start, "%Y-%m-%d").date()
        except (ValueError, AttributeError):
            django_messages.error(
                request,
                "La fecha de inicio es obligatoria (formato YYYY-MM-DD).",
            )
            return redirect(LIST_URL)

        # -- Parse optional end_date. --
        raw_end = request.POST.get("end_date", "").strip()
        end_date = None
        if raw_end:
            try:
                end_date = datetime.strptime(raw_end, "%Y-%m-%d").date()
                if end_date < start_date:
                    django_messages.error(
                        request,
                        "La fecha de fin no puede ser anterior a la de inicio.",
                    )
                    return redirect(LIST_URL)
            except (ValueError, AttributeError):
                end_date = None

        # -- Create the group. --
        group = WorkPeriodGroup.objects.create(
            company=company,
            label=label,
            start_date=start_date,
            end_date=end_date,
            created_by=cu,
        )

        # -- Resolve target operators. --
        raw_op_pks = request.POST.getlist("operator_pks")
        selected_pks = []
        for raw_pk in raw_op_pks:
            try:
                selected_pks.append(int(raw_pk))
            except (ValueError, TypeError):
                pass

        base_qs = (
            CompanyUser.objects
            .filter(company=company, is_active=True, role=CompanyUser.ROLE_WORKSHOP)
            .select_related("user")
            .order_by("user__last_name", "user__first_name")
        )
        operators = list(
            base_qs.filter(pk__in=selected_pks) if selected_pks else base_qs
        )

        # -- Create individual WorkPeriod per operator. --
        # Skip operators that already have an open period (is_closed=False).
        # Omitir operarios con periodo abierto (is_closed=False).
        created_count = 0
        skipped_names = []
        for operator in operators:
            if WorkPeriod.objects.filter(
                company_user=operator, is_closed=False
            ).exists():
                skipped_names.append(
                    operator.user.get_full_name() or operator.user.username
                )
                continue
            WorkPeriod.objects.create(
                company_user=operator,
                group=group,
                start_date=start_date,
                end_date=end_date,
                label=label,
                created_by=cu,
            )
            created_count += 1

        if created_count > 0:
            django_messages.success(
                request,
                f"Periodo '{label}' creado con {created_count} operario"
                f"{'s' if created_count != 1 else ''}.",
            )
        else:
            django_messages.warning(
                request,
                f"Grupo '{label}' creado, pero ningún operario fue asignado "
                f"(todos tenían ya un periodo abierto).",
            )
        if skipped_names:
            django_messages.warning(
                request,
                "Omitidos por tener periodo abierto: "
                + ", ".join(skipped_names) + ".",
            )

        logger.info(
            "# [WorkPeriodGroup] Grupo '%s' creado por %s. "
            "%d operario(s) asignado(s).",
            label, cu.user.username, created_count,
        )
        return redirect(
            reverse("panel:work_period_group_detail", kwargs={"pk": group.pk})
        )


class WorkPeriodGroupAddOperatorView(SupervisorAccessMixin, View):
    """
    Adds one or more WORKSHOP operators to an existing WorkPeriodGroup by
    creating individual WorkPeriod records linked to that group.
    Operators already assigned or with an open period are skipped.

    POST /panel/work-period-groups/<pk>/add-operator/
         operator_pks (int list, required — one or more)
    ---
    Añade uno o más operarios WORKSHOP a un WorkPeriodGroup existente
    creando registros WorkPeriod individuales vinculados al grupo.
    Los operarios ya asignados o con periodo abierto se omiten.

    POST /panel/work-period-groups/<pk>/add-operator/
    """

    def post(self, request, pk, *args, **kwargs):
        """
        Resolves the group, validates operators and creates WorkPeriod
        records. Redirects to the group detail page.
        ---
        Resuelve el grupo, valida los operarios y crea los registros
        WorkPeriod. Redirige al detalle del grupo.
        """
        from django.shortcuts import get_object_or_404
        from django.urls import reverse
        from ivr_config.models import WorkPeriodGroup, WorkPeriod

        cu = request.user.company_user
        company = cu.company

        group = get_object_or_404(
            WorkPeriodGroup, pk=pk, company=company
        )
        DETAIL_URL = reverse(
            "panel:work_period_group_detail", kwargs={"pk": group.pk}
        )

        if group.is_closed:
            django_messages.error(
                request,
                f"El periodo '{group.label}' está liquidado. "
                f"Reabre el periodo antes de añadir operarios.",
            )
            return redirect(DETAIL_URL)

        # -- Resolve operator_pks. --
        raw_op_pks = request.POST.getlist("operator_pks")
        selected_pks = []
        for raw_pk in raw_op_pks:
            try:
                selected_pks.append(int(raw_pk))
            except (ValueError, TypeError):
                pass

        if not selected_pks:
            django_messages.error(
                request, "Selecciona al menos un operario."
            )
            return redirect(DETAIL_URL)

        operators = list(
            CompanyUser.objects
            .filter(
                pk__in=selected_pks,
                company=company,
                is_active=True,
                role=CompanyUser.ROLE_WORKSHOP,
            )
            .select_related("user")
        )

        # Already assigned to this group.
        # Ya asignados a este grupo.
        already_in_group = set(
            group.operator_periods.values_list("company_user_id", flat=True)
        )

        created_count = 0
        skipped_names = []
        for operator in operators:
            op_name = operator.user.get_full_name() or operator.user.username
            if operator.pk in already_in_group:
                skipped_names.append(f"{op_name} (ya en este periodo)")
                continue
            if WorkPeriod.objects.filter(
                company_user=operator, is_closed=False
            ).exclude(group=group).exists():
                skipped_names.append(f"{op_name} (tiene otro periodo abierto)")
                continue
            WorkPeriod.objects.create(
                company_user=operator,
                group=group,
                start_date=group.start_date,
                end_date=group.end_date,
                label=group.label,
                created_by=cu,
            )
            created_count += 1

        if created_count > 0:
            django_messages.success(
                request,
                f"{created_count} operario"
                f"{'s' if created_count != 1 else ''} "
                f"añadido{'s' if created_count != 1 else ''} al periodo "
                f"'{group.label}'.",
            )
        if skipped_names:
            django_messages.warning(
                request,
                "Omitidos: " + ", ".join(skipped_names) + ".",
            )

        logger.info(
            "# [WorkPeriodGroup] %d operario(s) añadido(s) al grupo pk=%d "
            "('%s') por %s.",
            created_count, group.pk, group.label, cu.user.username,
        )
        return redirect(DETAIL_URL)


class WorkPeriodGroupCloseView(SupervisorAccessMixin, View):
    """
    Closes a WorkPeriodGroup and all its subordinate WorkPeriod records.
    Also marks digital/generated WorkOrders within the period as reviewed
    and enqueues Excel generation per reviewed WorkOrder.

    POST /panel/work-period-groups/<pk>/close/
         end_date    (YYYY-MM-DD, required)
         force_close (1, optional — bypass unreviewed check)
    ---
    Cierra un WorkPeriodGroup y todos sus WorkPeriod subordinados.
    También marca como revisados los WorkOrder digitales/generados dentro
    del periodo y encola la generación de Excel por WorkOrder revisado.

    POST /panel/work-period-groups/<pk>/close/
    """

    def post(self, request, pk, *args, **kwargs):
        """
        Validates end_date, closes the group and individual periods in bulk,
        marks work orders reviewed and enqueues Excel tasks.
        ---
        Valida end_date, cierra el grupo y los periodos individuales en
        bloque, marca los partes revisados y encola las tareas Excel.
        """
        from datetime import datetime
        from django.shortcuts import get_object_or_404
        from django.urls import reverse
        from django.utils.timezone import now as tz_now
        from ivr_config.models import WorkPeriodGroup
        from work_order_processor.tasks import generate_period_excel

        cu = request.user.company_user
        company = cu.company

        group = get_object_or_404(
            WorkPeriodGroup, pk=pk, company=company
        )
        DETAIL_URL = reverse(
            "panel:work_period_group_detail", kwargs={"pk": group.pk}
        )

        if group.is_closed:
            django_messages.error(
                request,
                f"El periodo '{group.label}' ya está liquidado.",
            )
            return redirect(DETAIL_URL)

        # -- Parse end_date. --
        raw_end = request.POST.get("end_date", "").strip()
        try:
            end_date = datetime.strptime(raw_end, "%Y-%m-%d").date()
        except (ValueError, AttributeError):
            django_messages.error(
                request,
                "La fecha de fin es obligatoria (formato YYYY-MM-DD).",
            )
            return redirect(DETAIL_URL)

        if end_date < group.start_date:
            django_messages.error(
                request,
                f"La fecha de fin ({end_date:%d/%m/%Y}) no puede ser "
                f"anterior a la de inicio ({group.start_date:%d/%m/%Y}).",
            )
            return redirect(DETAIL_URL)

        # -- Check for unreviewed work orders (unless force_close). --
        force = request.POST.get("force_close", "") == "1"
        unreviewed_qs = WorkOrder.objects.filter(
            company=company,
            source__in=[WorkOrder.Source.DIGITAL, WorkOrder.Source.GENERATED],
            reviewed=False,
            uploaded_by__work_periods__group=group,
        ).distinct()
        unreviewed_count = unreviewed_qs.count()

        if unreviewed_count and not force:
            django_messages.warning(
                request,
                f"Hay {unreviewed_count} parte(s) sin revisar en este "
                f"periodo. Marca 'Liquidar igualmente' para continuar.",
            )
            return redirect(DETAIL_URL)

        # -- Mark work orders reviewed. --
        work_orders_qs = WorkOrder.objects.filter(
            company=company,
            source__in=[WorkOrder.Source.DIGITAL, WorkOrder.Source.GENERATED],
            reviewed=False,
            entries__work_date__gte=group.start_date,
            entries__work_date__lte=end_date,
            uploaded_by__work_periods__group=group,
        ).distinct()
        pks = list(work_orders_qs.values_list("pk", flat=True))
        if pks:
            WorkOrder.objects.filter(pk__in=pks).update(
                reviewed=True,
                reviewed_at=tz_now(),
            )

        # -- Close individual WorkPeriod records and the group. --
        group.operator_periods.filter(is_closed=False).update(
            end_date=end_date, is_closed=True
        )
        group.end_date = end_date
        group.is_closed = True
        group.save(update_fields=["end_date", "is_closed", "updated_at"])

        # -- Enqueue Excel generation. --
        for pk_val in pks:
            generate_period_excel.apply_async(
                args=[pk_val], queue="work_orders"
            )

        django_messages.success(
            request,
            f"Periodo '{group.label}' liquidado. "
            f"{len(pks)} parte(s) marcado(s) como revisados.",
        )
        logger.info(
            "# [WorkPeriodGroup] Grupo pk=%d '%s' cerrado por %s. "
            "%d partes revisados.",
            group.pk, group.label, cu.user.username, len(pks),
        )
        return redirect(DETAIL_URL)


class WorkPeriodGroupLockView(AdminRoleRequiredMixin, View):
    """
    Toggles the is_closed flag of a WorkPeriodGroup (ADMIN only).
    Also toggles the is_closed flag of all subordinate WorkPeriod records.

    POST /panel/work-period-groups/<pk>/lock/
    ---
    Alterna el flag is_closed de un WorkPeriodGroup (solo ADMIN).
    También alterna is_closed en todos los WorkPeriod subordinados.

    POST /panel/work-period-groups/<pk>/lock/
    """

    def post(self, request, pk, *args, **kwargs):
        """
        Toggles is_closed on the group and all its operator periods.
        ---
        Alterna is_closed en el grupo y todos sus periodos de operario.
        """
        from django.shortcuts import get_object_or_404
        from django.urls import reverse
        from ivr_config.models import WorkPeriodGroup

        cu = request.user.company_user
        group = get_object_or_404(
            WorkPeriodGroup, pk=pk, company=cu.company
        )

        new_state = not group.is_closed
        group.is_closed = new_state
        group.save(update_fields=["is_closed", "updated_at"])

        # Propagate to all subordinate periods.
        # Propagar a todos los periodos subordinados.
        group.operator_periods.all().update(is_closed=new_state)

        action = "liquidado" if new_state else "reabierto"
        django_messages.success(
            request,
            f"Periodo '{group.label}' {action}. "
            f"{group.operator_periods.count()} operario(s) actualizado(s).",
        )
        referer = request.META.get("HTTP_REFERER", "")
        return redirect(
            referer
            or reverse("panel:work_period_group_detail", kwargs={"pk": group.pk})
        )

