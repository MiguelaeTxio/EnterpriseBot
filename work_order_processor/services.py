# /home/MiguelAeTxio/PROJECTS/EnterpriseBot/work_order_processor/services.py

"""
Business logic services for the work_order_processor application.
Contains three primary services:
  - extract_work_order_page(): sends a rasterized PDF page to Gemini Vision
    and returns a structured multi-block dict (up to 4 work blocks per page).
  - generate_work_order_excel(): builds an Excel report conforming to the
    partes-trabajo skill v1.2 specification: 17 columns, configuration area
    (rows 1-3), colour-coded REVISION HORARIO, LEYENDA sheet and MANIFIESTO
    DE INCIDENCIAS block at the foot of the data sheet.
  - classify_fault(): sends fault_description + repair_notes of a single
    WorkOrderEntryLine to Gemini Flash (text-only, Vertex AI) and returns a
    (FaultCategory, FaultSubcategory) tuple. Called exclusively by the Celery
    task classify_fault_line (Hito 7 / S023).
  - find_cached_classification(): looks up an existing classified
    WorkOrderEntryLine within the same company whose fault_description and
    repair_notes match exactly. Returns a (category, subcategory) tuple if
    found, or None if no match exists. Used as a pre-enqueue gate to avoid
    unnecessary Gemini calls for repeated fault descriptions (Hito 7 / S023).

---

Servicios de lógica de negocio para la aplicación work_order_processor.
Contiene cuatro servicios principales:
  - extract_work_order_page(): envía una página del PDF rasterizada a Gemini
    Vision y devuelve un dict multi-bloque estructurado (hasta 4 bloques de
    trabajo por página).
  - generate_work_order_excel(): construye un informe Excel conforme a la
    especificación de la skill partes-trabajo v1.2: 17 columnas, área de
    configuración (filas 1-3), REVISIÓN HORARIO con código de colores, hoja
    LEYENDA y bloque MANIFIESTO DE INCIDENCIAS al pie de la hoja de datos.
  - classify_fault(): envía fault_description + repair_notes de una sola
    WorkOrderEntryLine a Gemini Flash (solo texto, Vertex AI) y devuelve una
    tupla (FaultCategory, FaultSubcategory). Llamado exclusivamente por la
    tarea Celery classify_fault_line (Hito 7 / S023).
  - find_cached_classification(): busca una WorkOrderEntryLine ya clasificada
    dentro de la misma empresa cuya fault_description y repair_notes coincidan
    exactamente. Devuelve una tupla (categoría, subcategoría) si encuentra
    coincidencia, o None si no existe. Se usa como gate previo al encolado
    para evitar llamadas innecesarias a Gemini en averías repetidas
    (Hito 7 / S023).
"""

import io
import json
import logging
import os
import re
import time as _time
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Any


def _parse_period_from_pdf_name(pdf_name: str) -> tuple[date | None, date | None]:
    """
    Minimal period parser for Excel title generation.
    Canonical format: NAME DD-MM-YY AL DD-MM-YY[.pdf]
    Accepts spaces or underscores as token separators.
    Returns (None, None) if the filename does not match.

    ---

    Parser de periodo mínimo para la generación del título Excel.
    Formato canónico: NOMBRE DD-MM-AA AL DD-MM-AA[.pdf]
    Acepta espacios o guiones bajos como separadores de tokens.
    Devuelve (None, None) si el nombre no coincide.
    """
    stem    = os.path.splitext(os.path.basename(pdf_name))[0]
    pattern = r'(\d{2})-(\d{2})-(\d{2})[_\s]+[Aa][Ll][_\s]+(\d{2})-(\d{2})-(\d{2})'
    m       = re.search(pattern, stem)
    if not m:
        return None, None
    try:
        start = date(2000 + int(m.group(3)), int(m.group(2)), int(m.group(1)))
        end   = date(2000 + int(m.group(6)), int(m.group(5)), int(m.group(4)))
        return start, end
    except ValueError:
        return None, None

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from django.utils import timezone
from google import genai
from google.genai.types import GenerateContentConfig, HttpOptions, Part, ThinkingConfig

from ai_services.gemini_client import get_gemini_client as _get_gemini_client
from fleet.models import MachineAsset
from .models import WorkOrder, WorkOrderEntry, WorkOrderEntryLine, FaultCategory, FaultSubcategory

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Gemini client initialisation — Hito S001/H10: migrated to the shared
# ai_services.gemini_client helper (DRY principle, see
# doc-master-enterprisebot section 4.1.1). The local _get_gemini_client()
# wrapper is gone; the name is kept as an alias via the import above so
# every call site below (3 occurrences) needs no further change.
#
# _GEMINI_MODEL stays pinned to gemini-2.5-flash here intentionally —
# this is the documented technical debt (doc-master-enterprisebot
# 4.1.1) pending migration to gemini-3.5-flash before 2026-09-16. New
# code (e.g. spare_parts.services) uses ai_services.gemini_client's
# DEFAULT_MODEL (gemini-3.5-flash) instead.
#
# Inicialización del cliente Gemini — Hito S001/H10: migrado al helper
# compartido ai_services.gemini_client (principio DRY, ver
# doc-master-enterprisebot sección 4.1.1). El wrapper local
# _get_gemini_client() desaparece; el nombre se conserva como alias vía
# el import de arriba para que cada punto de llamada de abajo (3
# ocurrencias) no requiera más cambios.
#
# _GEMINI_MODEL permanece fijado a gemini-2.5-flash aquí
# intencionadamente — es la deuda técnica documentada
# (doc-master-enterprisebot 4.1.1) pendiente de migrar a
# gemini-3.5-flash antes del 2026-09-16. El código nuevo (p.ej.
# spare_parts.services) usa el DEFAULT_MODEL de
# ai_services.gemini_client (gemini-3.5-flash) en su lugar.
# ---------------------------------------------------------------------------
_GEMINI_MODEL = "gemini-2.5-flash"

# Per-request timeout in milliseconds passed via GenerateContentConfig.http_options.
# With vertexai=True the client-level HttpOptions.timeout is not reliably forwarded
# to httpx on a per-request basis; the only guaranteed mechanism is to pass
# HttpOptions directly inside GenerateContentConfig on every generate_content call.
# 360 000 ms = 6 minutes — sufficient margin for complex handwritten pages.
#
# Timeout por petición en milisegundos, pasado mediante GenerateContentConfig.http_options.
# Con vertexai=True el HttpOptions.timeout a nivel de cliente no se propaga de forma
# fiable a httpx por petición; el único mecanismo garantizado es pasar HttpOptions
# directamente dentro de GenerateContentConfig en cada llamada a generate_content.
# 360 000 ms = 6 minutos — margen suficiente para páginas manuscritas complejas.
_GEMINI_TIMEOUT_MS = 60_000

# GenerateContentConfig reutilizable con timeout por petición garantizado.
# Reusable GenerateContentConfig with guaranteed per-request timeout.
_GEMINI_REQUEST_CONFIG = GenerateContentConfig(
    http_options=HttpOptions(timeout=_GEMINI_TIMEOUT_MS),
)


# ---------------------------------------------------------------------------
# Extraction prompt — multi-block / Prompt de extracción — multi-bloque
# ---------------------------------------------------------------------------
# Conforms to the partes-trabajo skill v1.2 JSON output specification.
# The operario field is intentionally left null — it is derived from the
# PDF filename in tasks.py, not from the handwritten text (D6 / annex V06).
#
# Conforme a la especificación JSON de salida de la skill partes-trabajo v1.2.
# El campo operario se deja intencionadamente null — se deriva del nombre del
# fichero PDF en tasks.py, no del texto manuscrito (D6 / anexo V06).

_EXTRACTION_PROMPT = """\
Eres un asistente especializado en la extracción de datos de partes de trabajo
manuscritos de una empresa de grúas y maquinaria industrial (grúas móviles,
plataformas elevadoras, autocargantes, cabezas tractoras, semirremolques,
carretillas elevadoras). Las averías y repair_noteses pertenecen siempre a este
tipo de vehículos y maquinaria pesada industrial.

Analiza la imagen del parte de trabajo manuscrito. Cada página es un parte
DIARIO con una única fecha en la cabecera y hasta 4 bloques de trabajo.

REGLAS OBLIGATORIAS:
1. Responde ÚNICAMENTE con un objeto JSON válido. Sin texto adicional, sin
   bloques de código markdown, sin explicaciones.
2. Si un campo no es legible o no aparece, usa null.
3. Fechas en formato "DD/MM/YYYY". Horas en formato "HH:MM".
4. Horarios SIEMPRE redondeados a fracciones de media hora (:00 o :30).
   Ejemplos: 09:20 → "09:30", 07:10 → "07:00", 17:45 → "18:00".
5. El campo "uncertain_date" es true solo si la fecha es genuinamente ilegible
   tras intentar deducirla por contexto. No lo uses si puedes leerla.
6. En "machine_raw" copia exactamente lo que lees tras aplicar las reglas de
   caligrafía (ver sección CALIGRAFÍA). Si aparece cualquier variante del
   alias Larios en contexto de máquina, normalízalo a "Larios". El tipo de
   vehículo que precede al alias va separado por un espacio (ej. "Furgon
   Larios", "Mercedes Larios"). NUNCA concatenar tipo y alias sin espacio.
   "Salida polígonos" NO es una máquina: deja machine_raw null en ese caso.
7. Interpreta abreviaturas técnicas en contexto de vehículos pesados:
   "hid." → hidráulico, "trans." → transmisión, "dir." → dirección,
   "mto." → mantenimiento, "ace." → aceite, "fil." → filtro.
8. Si un campo es de lectura incierta (no imposible, sino dudosa), inclúyelo
   con el valor más probable y añade el nombre del campo a "flags".
   Campos que pueden ir en flags: "FECHA", "H.C.", "H.F.", "DESCRIPCION",
   "MAQUINA".
9. Procesa SOLO los bloques que tengan al menos un campo relleno. Ignora
   bloques completamente vacíos.
10. "extraction_confidence" evalúa la calidad global de la página:
    "HIGH" = todos los campos principales legibles,
    "MEDIUM" = alguna duda menor,
    "LOW" = campos importantes ilegibles,
    "FAILED" = imagen ilegible o no es un parte de trabajo.

CALIGRAFÍA RÁPIDA — REGLAS DE INFERENCIA:
Estos operarios escriben con prisa. Aplicar SIEMPRE estas reglas antes de
marcar un campo como dudoso o ilegible.

A) CONFUSIÓN LETRA/NÚMERO EN CÓDIGOS DE MAQUINARIA:
   El formato de código es LETRA(S)-NÚMERO(S) o LETRA(S)NÚMERO(S).
   Si un carácter no encaja en su posición esperada, sustituir por el
   candidato morfológicamente más cercano:
   - "6" en posición de letra inicial → leer como "G" (formas similares).
   - "0" en posición de letra → leer como "O" y viceversa.
   - "1" en posición de letra → leer como "I" o "L".
   - "S" en posición numérica → leer como "5". El 5 escrito rápido
     degenera la curva inferior a un trazo vertical con visera horizontal
     arriba, resultando casi una "S" o una "L invertida".
   - "B" en posición numérica → leer como "8".
   - Punto "." como separador en código → tratar como guion "-".
   Ejemplos: "6-8" → "G-8", "A-S4" → "A-54", "B.42" → "B-42".

B) ALIAS DE EMPRESA "LARIOS":
   Cualquier variante del alias Larios escrita con prisa debe normalizarse
   a "Larios": Loriol, Lorios, Laros, Larios, Larjos, Laris → "Larios".
   El tipo de vehículo que precede al alias va separado por un espacio:
   "Furgon Larios", "Mercedes Larios", "Camion Larios".
   NUNCA concatenar tipo y alias sin espacio.

C) TACHONES Y CORRECCIONES:
   Si un valor está tachado y hay otro valor escrito cerca (encima, debajo,
   al lado), usar el valor escrito fuera del tachón. Ignorar completamente
   el valor tachado.

D) DESCRIPCIONES DE AVERÍA Y REPARACIÓN:
   Siempre tienen estructura VERBO + OBJETO. Los verbos típicos son:
   montar, desmontar, revisar, cambiar, reparar, sustituir, ajustar,
   soldar, comprobar, limpiar, engrasar, rellenar, purgar, apretar,
   soltar, instalar, retirar, aproximar, meter, sacar.
   Si lees un sustantivo aislado sin verbo claro, busca el trazo anterior
   o posterior que pueda ser el inicio de uno de estos verbos y completa
   la frase. Si aún hay ambigüedad, incluye el fragmento más probable y
   añade "DESCRIPCION" a flags.

E) CURVAS DEGENERADAS A TRAZOS RECTOS:
   Con la prisa, las curvas de letras y números se simplifican. Leer por
   la estructura global de la palabra o código, no trazo a trazo.
   Un "5" puede parecer una "S", un "2" puede parecer un "Z", un "3"
   puede parecer un "E" o un "F" truncado.

F) CÓDIGO DE MÁQUINA EN CAMPO KM:
   Los operarios a veces anotan el código de máquina en el campo KM:
   dejando el campo MAQUINA: en blanco. Si MAQUINA: está vacío pero
   KM: contiene un valor con formato de código de máquina (letras
   seguidas de guion y número, ej. G-8, A-54, B-42), usar ese valor
   como machine_raw. El valor numérico de KM queda en ese caso sin dato.

Formato de respuesta (claves exactas):
{
  "fecha": "<DD/MM/YYYY o null>",
  "uncertain_date": <true | false>,
  "extraction_confidence": "<HIGH | MEDIUM | LOW | FAILED>",
  "entradas": [
    {
      "machine_raw": "<código o alias normalizado tal como se lee, o null>",
      "fault_description": "<descripción de la avería o tarea, o null>",
      "repair_notes": "<descripción de la reparación realizada, o null>",
      "hc": "<HH:MM o null>",
      "hf": "<HH:MM o null>",
      "or_val": "<referencia O.R. o null>",
      "flags": ["CAMPO1", "CAMPO2"],
      "fault_category": "<CODIGO_CATEGORIA o null>",
      "fault_subcategory": "<CODIGO_SUBCATEGORIA o null>"
    }
  ]
}

TAXONOMÍA DE AVERÍAS — usa estos códigos exactos en fault_category y fault_subcategory:

Categorías (fault_category):
  ENGINE_TRANSMISSION        — Motor, transmisión, PTO, refrigeración, combustible
  HYDRAULIC                  — Bomba hidráulica, cilindros, válvulas, aceite, central
  ELECTRICAL_ELECTRONIC      — Cableado, sensores, mandos, iluminación, batería
  BRAKES_STEERING_SUSPENSION — Frenos, dirección, suspensión
  TYRES_RUNNING_GEAR         — Neumáticos, ejes, cadenas y rodadura oruga
  LIFTING_STRUCTURE          — Pluma, gancho/poleas, cable, rotación, estabilizadores,
                               mástil/horquillas, plataforma, quinta rueda, chasis semirremolque
  BODYWORK_CHASSIS           — Carrocería, chasis estructural
  OTHER                      — Cualquier avería que no encaje en los grupos anteriores

Subcategorías (fault_subcategory):
  ET_ENGINE | ET_TRANSMISSION | ET_PTO | ET_COOLING | ET_FUEL
  HY_PUMP | HY_CYLINDERS | HY_VALVES | HY_OIL | HY_CENTRAL
  EE_WIRING | EE_SENSORS | EE_CONTROLS | EE_LIGHTS | EE_BATTERY
  BSS_BRAKES | BSS_STEERING | BSS_SUSPENSION
  TRG_TYRES | TRG_AXLES | TRG_TRACKS
  LS_BOOM | LS_HOOK_PULLEYS | LS_CABLE | LS_ROTATION | LS_STABILIZERS |
  LS_MAST | LS_PLATFORM | LS_FIFTH_WHEEL | LS_CHASSIS_TRAILER
  BC_BODYWORK | BC_CHASSIS
  OT_OTHER

Regla: la subcategoría debe pertenecer a la categoría elegida (mismos prefijos).
Si la información es insuficiente o no encaja en ningún grupo, usa OTHER / OT_OTHER.
Si el campo fault_description y repair_notes son ambos nulos o ilegibles, usa null
en ambos campos de clasificación.
"""


# ---------------------------------------------------------------------------
# Public service: page extraction
# Servicio público: extracción de página
# ---------------------------------------------------------------------------

def extract_work_order_page(image_bytes: bytes) -> dict[str, Any]:
    """
    Sends a rasterized PDF page (PNG bytes) to Gemini Vision and returns
    a structured multi-block dict conforming to the partes-trabajo skill
    v1.2 JSON specification.

    Applies tolerant JSON parsing: strips markdown fences and locates the
    outermost JSON object within the response text. On total failure returns
    a safe fallback dict with an empty entradas list and confidence FAILED.

    ---

    Envía una página del PDF rasterizada (bytes PNG) a Gemini Vision y
    devuelve un dict multi-bloque estructurado conforme a la especificación
    JSON de la skill partes-trabajo v1.2.

    Aplica parseo JSON tolerante: elimina bloques markdown y localiza el
    objeto JSON más externo dentro del texto de respuesta. En caso de fallo
    total devuelve un dict de reserva seguro con lista entradas vacía y
    confianza FAILED.
    """
    client = _get_gemini_client()

    try:
        logger.info("# Gemini Vision: iniciando extracción de página.")

        # Retry loop for 429 RESOURCE_EXHAUSTED errors.
        # On quota exhaustion, wait 60 seconds and retry up to 3 times
        # before propagating the error as a FAILED page.
        # Bucle de reintento para errores 429 RESOURCE_EXHAUSTED.
        # En caso de cuota agotada, esperar 60 segundos y reintentar
        # hasta 3 veces antes de propagar el error como página FAILED.
        _MAX_RETRIES_429 = 3
        _response = None
        for _attempt in range(_MAX_RETRIES_429 + 1):
            try:
                _response = client.models.generate_content(
                    model=_GEMINI_MODEL,
                    contents=[
                        Part.from_bytes(data=image_bytes, mime_type="image/png"),
                        _EXTRACTION_PROMPT_FULL,
                    ],
                    config=_GEMINI_REQUEST_CONFIG,
                )
                break  # Success — exit retry loop.
            except Exception as _e429:
                _is_429 = (
                    hasattr(_e429, "status_code") and _e429.status_code == 429
                ) or "429" in str(_e429) or "RESOURCE_EXHAUSTED" in str(_e429)
                if _is_429 and _attempt < _MAX_RETRIES_429:
                    logger.warning(
                        "# Gemini Vision: 429 RESOURCE_EXHAUSTED — "
                        "reintento %d/%d en 60 segundos.",
                        _attempt + 1, _MAX_RETRIES_429,
                    )
                    _time.sleep(60)
                else:
                    raise

        response = _response
        raw_text = response.text.strip()
        logger.info("# Gemini Vision: respuesta recibida. Parseando JSON.")

        # Strip markdown fences / Eliminar bloques markdown.
        cleaned = re.sub(r"```(?:json)?|```", "", raw_text).strip()

        # Locate outermost JSON object / Localizar objeto JSON más externo.
        match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if not match:
            raise ValueError(
                f"No se encontró ningún objeto JSON en la respuesta de Gemini: "
                f"{raw_text[:200]}"
            )

        extracted: dict[str, Any] = json.loads(match.group())

        # Ensure entradas is always a list / Garantizar que entradas es siempre lista.
        if not isinstance(extracted.get("entradas"), list):
            extracted["entradas"] = []

        logger.info(
            "# Gemini Vision: extracción completada. Confianza: %s | Entradas: %d",
            extracted.get("extraction_confidence", "DESCONOCIDA"),
            len(extracted.get("entradas", [])),
        )

        # Post-processing: fix Larios alias concatenated without space.
        # The prompt instructs Gemini to separate them, but caligraphic OCR
        # sometimes produces "MERCEDESLARIOS", "FURGONLARIOS", etc.
        # We apply a deterministic Python fix: if machine_raw matches the
        # pattern VEHICLETYPE+LARIOS_VARIANT (no space), insert the space
        # and normalise the alias to "Larios".
        #
        # Post-procesado: corregir alias Larios concatenado sin espacio.
        # El prompt instruye a Gemini a separarlos, pero el OCR caligráfico
        # a veces produce "MERCEDESLARIOS", "FURGONLARIOS", etc.
        # Aplicamos un fix Python determinista: si machine_raw coincide con
        # el patrón TIPOVEHICULO+VARIANTE_LARIOS (sin espacio), se inserta
        # el espacio y se normaliza el alias a "Larios".
        _LARIOS_RE = re.compile(
            r'(?i)^([A-Za-z]+)(L[ao]r[aio][a-z]*)$'
        )
        for _entrada in (extracted.get("entradas") or []):
            _raw = (_entrada.get("machine_raw") or "").strip()
            if not _raw:
                continue
            _m = _LARIOS_RE.match(_raw)
            if _m:
                # Separate prefix and normalise alias to "Larios".
                # Separar prefijo y normalizar alias a "Larios".
                _prefix = _m.group(1).capitalize()
                _entrada["machine_raw"] = f"{_prefix} Larios"

        return extracted

    except Exception as exc:
        logger.error(
            "# Error crítico en extract_work_order_page: %s", exc, exc_info=True
        )
        return {
            "fecha":                None,
            "uncertain_date":       False,
            "extraction_confidence": "FAILED",
            "entradas":             [],
        }

# ---------------------------------------------------------------------------
# Extraction prompt — full (front + back of physical form)
# Prompt de extracción — completo (cara delantera + trasera del formulario)
# ---------------------------------------------------------------------------
# Used exclusively by extract_work_order_page_full() for new work orders
# submitted via the operator Upload view (WorkOrderEntryUploadView).
# The historical pipeline (tasks.py / Celery) continues using
# _EXTRACTION_PROMPT and extract_work_order_page() without modification.
#
# Usado exclusivamente por extract_work_order_page_full() para partes nuevos
# enviados via la vista Upload del operario (WorkOrderEntryUploadView).
# El pipeline histórico (tasks.py / Celery) sigue usando _EXTRACTION_PROMPT
# y extract_work_order_page() sin ninguna modificación.

_EXTRACTION_PROMPT_FULL = """Eres un asistente especializado en la extracción de datos de partes de trabajo
manuscritos de una empresa de grúas y maquinaria industrial (grúas móviles,
plataformas elevadoras, autocargantes, cabezas tractoras, semirremolques,
carretillas elevadoras). El parte tiene DOS caras:

CARA DELANTERA — Partes de trabajo diarios:
  Una única fecha en la cabecera y hasta 4 bloques de trabajo con campos:
  MAQUINA, DESCRIPCION AVERIA, REPARACION, H.C., H.F., O.R.

CARA TRASERA — Repuestos utilizados:
  Una tabla con filas de materiales consumidos. Columnas:
  REFERENCIA (ref. del albarán del proveedor), VEHICULO (código de máquina),
  MATERIAL (descripción del repuesto), UNIDADES (cantidad numérica), y
  PROCEDENCIA (proveedor externo o almacén interno).

REGLAS OBLIGATORIAS (aplican a ambas caras):
1. Responde ÚNICAMENTE con un objeto JSON válido. Sin texto adicional, sin
   bloques de código markdown, sin explicaciones.
2. Si un campo no es legible o no aparece, usa null.
3. Fechas en formato "DD/MM/YYYY". Horas en formato "HH:MM".
4. Horarios SIEMPRE redondeados a fracciones de media hora (:00 o :30).
   Ejemplos: 09:20 → "09:30", 07:10 → "07:00", 17:45 → "18:00".
5. El campo "uncertain_date" es true solo si la fecha es genuinamente ilegible
   tras intentar deducirla por contexto. No lo uses si puedes leerla.
6. En "machine_raw" copia exactamente lo que lees tras aplicar las reglas de
   caligrafía. Si aparece cualquier variante del alias Larios en contexto de
   máquina, normalízalo a "Larios". El tipo de vehículo que precede al alias
   va separado por un espacio (ej. "Furgon Larios"). NUNCA concatenar sin espacio.
   "Salida polígonos" NO es una máquina: deja machine_raw null en ese caso.
7. Interpreta abreviaturas técnicas en contexto de vehículos pesados:
   "hid." → hidráulico, "trans." → transmisión, "dir." → dirección,
   "mto." → mantenimiento, "ace." → aceite, "fil." → filtro.
8. Si un campo es de lectura incierta (no imposible, sino dudosa), inclúyelo
   con el valor más probable y añade el nombre del campo a "flags".
   Campos que pueden ir en flags de entradas: "FECHA", "H.C.", "H.F.",
   "DESCRIPCION", "MAQUINA".
   Campos que pueden ir en flags de repuestos: "REFERENCIA", "MATERIAL",
   "UNIDADES", "VEHICULO", "PROCEDENCIA".
9. Procesa SOLO los bloques de trabajo que tengan al menos un campo relleno.
   Procesa SOLO las filas de repuestos que tengan al menos un campo relleno.
   Ignora bloques y filas completamente vacíos.
10. "extraction_confidence" evalúa la calidad global de la página:
    "HIGH" = todos los campos principales legibles,
    "MEDIUM" = alguna duda menor,
    "LOW" = campos importantes ilegibles,
    "FAILED" = imagen ilegible o no es un parte de trabajo.

REGLAS ESPECÍFICAS — CARA TRASERA (REPUESTOS):
R1. "referencia" es el código o número de referencia del albarán del proveedor.
    Puede ser alfanumérico. Si no aparece, usa null.
R2. "vehiculo_raw" es el código o descripción de la máquina a la que se imputa
    el repuesto. Aplica las mismas reglas de caligrafía que para machine_raw.
R3. "material" es la descripción textual del repuesto o material. Puede inferirse
    de la referencia si el campo está vacío pero la referencia es descriptiva.
R4. "unidades" es un número (entero o decimal). Usa null si no es legible.
R5. "origen" debe ser exactamente "SUPPLIER" si el material proviene de un
    proveedor externo, o "WAREHOUSE" si proviene del almacén interno.
    Indicadores de proveedor: nombre de empresa, número de albarán externo.
    Indicadores de almacén: "almacén", "stock", "almacen", sin nombre de empresa.
    Si no se puede determinar con certeza, usa "WAREHOUSE" por defecto y añade
    "PROCEDENCIA" a flags.
R6. "proveedor" es el nombre del proveedor externo. Solo se rellena cuando
    origen = "SUPPLIER". En caso contrario usa null.

CALIGRAFÍA RÁPIDA — REGLAS DE INFERENCIA (aplican a ambas caras):
A) CONFUSIÓN LETRA/NÚMERO EN CÓDIGOS:
   - "6" en posición de letra inicial → leer como "G".
   - "0" en posición de letra → leer como "O" y viceversa.
   - "1" en posición de letra → leer como "I" o "L".
   - "S" en posición numérica → leer como "5".
   - "B" en posición numérica → leer como "8".
   - Punto "." como separador en código → tratar como guion "-".
B) ALIAS DE EMPRESA "LARIOS": cualquier variante normalizar a "Larios".
C) TACHONES: usar el valor escrito fuera del tachón, ignorar el tachado.
D) DESCRIPCIONES: siempre estructura VERBO + OBJETO.
E) CURVAS DEGENERADAS: leer por estructura global, no trazo a trazo.

F) MÚLTIPLES MÁQUINAS EN UN ÚNICO CAMPO:
   Si el operario ha escrito más de un código de máquina en el campo
   MAQUINA: separados por coma, punto y coma, la conjunción "y", barra
   oblicua o espacio (ej. "Z59, Z73, Z62", "T-11 y Z-107 Z-123",
   "F-22 y V04"), genera una entrada separada por cada código de máquina
   identificado, repitiendo en cada una la misma fault_description,
   repair_notes, hc, hf y or_val. Añade "MAQUINA" a flags en todas las
   entradas resultantes para indicar que proceden de un campo múltiple.

G) DESBORDAMIENTO DE LÍNEA — DOS TAREAS EN UN MISMO BLOQUE FÍSICO:
   El formulario tiene 4 filas de trabajo por página. Cuando un operario
   necesita más de 4 tareas, reutiliza filas existentes escribiendo una
   segunda tarea encima, debajo o al lado de la primera, usando el campo
   O.R. para anotar el horario de la segunda tarea.
   Si detectas en un bloque físico que el campo O.R. contiene un valor
   con formato HH:MM y los campos H.C./H.F. ya están ocupados con otro
   horario, genera DOS entradas separadas:
     - Entrada 1: hc y hf del horario principal, or_val vacío, misma máquina
       y descripción de la primera tarea si se puede distinguir.
     - Entrada 2: hc = hf de la entrada 1, hf = valor HH:MM del campo O.R.,
       or_val vacío, misma máquina, descripción de la segunda tarea si se
       puede distinguir (o la misma si no hay distinción visual clara).
   Añade "H.C." y "H.F." a flags en ambas entradas para indicar desdoblamiento.

H) CONFUSIÓN MORFOLÓGICA EN CÓDIGOS DE MÁQUINA — TODOS LOS CARACTERES:
   Los caracteres de los códigos de máquina pueden confundirse en cualquier
   posición (prefijo alfabético o bloque numérico) por escritura rápida.
   Pares de confusión confirmados en este catálogo:
     T ↔ 7  (la T con barra baja parece 7, el 7 sin gancho parece T)
     T ↔ 4  (la T con vertical corta parece 4)
     Z ↔ 2  (morfológicamente idénticos a velocidad)
     O ↔ 0  (indistinguibles en escritura rápida)
     L ↔ 1, I ↔ 1  (trazo único)
     S ↔ 5  (curva aplanada)
     G ↔ 6  (forma circular con trazo horizontal)
     B ↔ 8  (dos bucles apilados)
   Cuando un código leído no te resulte coherente con el catálogo de
   maquinaria de una empresa de grúas y plataformas elevadoras, aplica
   estas sustituciones para intentar leer el código correcto. Transcribe
   en machine_raw el código más probable e incluye "MAQUINA" en flags si
   hay ambigüedad residual.
   EXCEPCIÓN: los vehículos de la familia TURISMOS usan matrícula española
   (4 dígitos seguidos de 3 letras, ej. "2030JVK") — transcribir tal cual
   sin aplicar sustituciones.

I) TEXTO DE ETIQUETA DEL FORMULARIO CAPTURADO COMO CONTENIDO:
   Si fault_description o repair_notes contienen exclusivamente palabras
   que son etiquetas impresas del propio formulario ("Descripción",
   "Avería", "Reparación", "Tarea") sin ningún contenido real añadido,
   deja el campo vacío (null). Si la palabra de etiqueta aparece como
   parte de una frase con contenido real (ej. "Buscar avería eléctrica",
   "Reparar avería mangera cortada"), transcribe la frase completa — es
   contenido válido del operario, no una etiqueta del formulario.

Formato de respuesta (claves exactas):
{
  "fecha": "<DD/MM/YYYY o null>",
  "uncertain_date": <true | false>,
  "extraction_confidence": "<HIGH | MEDIUM | LOW | FAILED>",
  "entradas": [
    {
      "machine_raw": "<código o alias normalizado, o null>",
      "fault_description": "<descripción de la avería o tarea, o null>",
      "repair_notes": "<descripción de la reparación realizada, o null>",
      "hc": "<HH:MM o null>",
      "hf": "<HH:MM o null>",
      "or_val": "<referencia O.R. o null>",
      "flags": ["CAMPO1", "CAMPO2"],
      "fault_category": "<CODIGO_CATEGORIA o null>",
      "fault_subcategory": "<CODIGO_SUBCATEGORIA o null>"
    }
  ],
  "repuestos": [
    {
      "referencia": "<referencia albarán o null>",
      "vehiculo_raw": "<código máquina o null>",
      "material": "<descripción material o null>",
      "unidades": <número o null>,
      "origen": "<SUPPLIER | WAREHOUSE>",
      "proveedor": "<nombre proveedor o null>",
      "flags": ["CAMPO1"]
    }
  ]
}

TAXONOMÍA DE AVERÍAS — usa estos códigos exactos en fault_category y fault_subcategory:

Categorías (fault_category):
  ENGINE_TRANSMISSION        — Motor, transmisión, PTO, refrigeración, combustible
  HYDRAULIC                  — Bomba hidráulica, cilindros, válvulas, aceite, central
  ELECTRICAL_ELECTRONIC      — Cableado, sensores, mandos, iluminación, batería
  BRAKES_STEERING_SUSPENSION — Frenos, dirección, suspensión
  TYRES_RUNNING_GEAR         — Neumáticos, ejes, cadenas y rodadura oruga
  LIFTING_STRUCTURE          — Pluma, gancho/poleas, cable, rotación, estabilizadores,
                               mástil/horquillas, plataforma, quinta rueda, chasis semirremolque
  BODYWORK_CHASSIS           — Carrocería, chasis estructural
  OTHER                      — Cualquier avería que no encaje en los grupos anteriores

Subcategorías (fault_subcategory):
  ET_ENGINE | ET_TRANSMISSION | ET_PTO | ET_COOLING | ET_FUEL
  HY_PUMP | HY_CYLINDERS | HY_VALVES | HY_OIL | HY_CENTRAL
  EE_WIRING | EE_SENSORS | EE_CONTROLS | EE_LIGHTS | EE_BATTERY
  BSS_BRAKES | BSS_STEERING | BSS_SUSPENSION
  TRG_TYRES | TRG_AXLES | TRG_TRACKS
  LS_BOOM | LS_HOOK_PULLEYS | LS_CABLE | LS_ROTATION | LS_STABILIZERS |
  LS_MAST | LS_PLATFORM | LS_FIFTH_WHEEL | LS_CHASSIS_TRAILER
  BC_BODYWORK | BC_CHASSIS
  OT_OTHER

Regla: la subcategoría debe pertenecer a la categoría elegida (mismos prefijos).
Si la información es insuficiente o no encaja en ningún grupo, usa OTHER / OT_OTHER.
Si el campo fault_description y repair_notes son ambos nulos o ilegibles, usa null
en ambos campos de clasificación.
"""


# ---------------------------------------------------------------------------
# Public service: full page extraction (front + back)
# Servicio público: extracción completa de página (delantera + trasera)
# ---------------------------------------------------------------------------

def extract_work_order_page_full(image_bytes: bytes) -> dict:
    """
    Sends a rasterized work-order page (PNG bytes) to Gemini Vision using
    the full prompt (_EXTRACTION_PROMPT_FULL) which covers both the front
    (work blocks) and the back (spare parts table) of the physical form.

    Returns a structured dict with keys: fecha, uncertain_date,
    extraction_confidence, entradas (list of work blocks) and repuestos
    (list of spare part lines). On total failure returns a safe fallback
    dict with empty lists and confidence FAILED.

    This function is used exclusively by the operator Upload view
    (WorkOrderEntryUploadView) for new work orders submitted under the
    active system. The historical pipeline continues using
    extract_work_order_page() without modification.

    ---

    Envía una página de parte de trabajo rasterizada (bytes PNG) a Gemini
    Vision usando el prompt completo (_EXTRACTION_PROMPT_FULL) que cubre
    tanto la cara delantera (bloques de trabajo) como la trasera (tabla de
    repuestos) del formulario físico.

    Devuelve un dict estructurado con claves: fecha, uncertain_date,
    extraction_confidence, entradas (lista de bloques de trabajo) y repuestos
    (lista de líneas de repuesto). En caso de fallo total devuelve un dict
    de reserva seguro con listas vacías y confianza FAILED.

    Esta función es usada exclusivamente por la vista Upload del operario
    (WorkOrderEntryUploadView) para partes nuevos enviados bajo el sistema
    activo. El pipeline histórico sigue usando extract_work_order_page()
    sin modificación.
    """
    client = _get_gemini_client()

    try:
        logger.info(
            "# Gemini Vision (full): iniciando extracción completa de página "
            "(delantera + trasera)."
        )

        _MAX_RETRIES_429 = 3
        _response = None
        for _attempt in range(_MAX_RETRIES_429 + 1):
            try:
                _response = client.models.generate_content(
                    model=_GEMINI_MODEL,
                    contents=[
                        Part.from_bytes(data=image_bytes, mime_type="image/png"),
                        _EXTRACTION_PROMPT_FULL,
                    ],
                    config=_GEMINI_REQUEST_CONFIG,
                )
                break
            except Exception as _e429:
                _is_429 = (
                    hasattr(_e429, "status_code") and _e429.status_code == 429
                ) or "429" in str(_e429) or "RESOURCE_EXHAUSTED" in str(_e429)
                if _is_429 and _attempt < _MAX_RETRIES_429:
                    logger.warning(
                        "# Gemini Vision (full): 429 RESOURCE_EXHAUSTED — "
                        "reintento %d/%d en 60 segundos.",
                        _attempt + 1, _MAX_RETRIES_429,
                    )
                    _time.sleep(60)
                else:
                    raise

        response = _response
        raw_text = response.text.strip()
        logger.info(
            "# Gemini Vision (full): respuesta recibida. Parseando JSON."
        )

        # Strip markdown fences / Eliminar bloques markdown.
        cleaned = re.sub(r"```(?:json)?|```", "", raw_text).strip()

        # Locate outermost JSON object / Localizar objeto JSON más externo.
        match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if not match:
            raise ValueError(
                f"No se encontró ningún objeto JSON en la respuesta de Gemini: "
                f"{raw_text[:200]}"
            )

        extracted: dict = json.loads(match.group())

        # Ensure entradas and repuestos are always lists.
        # Garantizar que entradas y repuestos son siempre listas.
        if not isinstance(extracted.get("entradas"), list):
            extracted["entradas"] = []
        if not isinstance(extracted.get("repuestos"), list):
            extracted["repuestos"] = []

        logger.info(
            "# Gemini Vision (full): extracción completada. "
            "Confianza: %s | Entradas: %d | Repuestos: %d",
            extracted.get("extraction_confidence", "DESCONOCIDA"),
            len(extracted.get("entradas", [])),
            len(extracted.get("repuestos", [])),
        )

        # Post-processing: fix Larios alias concatenated without space.
        # Post-procesado: corregir alias Larios concatenado sin espacio.
        _LARIOS_RE = re.compile(r'(?i)^([A-Za-z]+)(L[ao]r[aio][a-z]*)$')
        for _entrada in (extracted.get("entradas") or []):
            _raw = (_entrada.get("machine_raw") or "").strip()
            if _raw:
                _m = _LARIOS_RE.match(_raw)
                if _m:
                    _prefix = _m.group(1).capitalize()
                    _entrada["machine_raw"] = f"{_prefix} Larios"

        # Apply same Larios fix to vehiculo_raw in repuestos.
        # Aplicar el mismo fix Larios a vehiculo_raw en repuestos.
        for _repuesto in (extracted.get("repuestos") or []):
            _vraw = (_repuesto.get("vehiculo_raw") or "").strip()
            if _vraw:
                _m = _LARIOS_RE.match(_vraw)
                if _m:
                    _prefix = _m.group(1).capitalize()
                    _repuesto["vehiculo_raw"] = f"{_prefix} Larios"

        return extracted

    except Exception as exc:
        logger.error(
            "# Error crítico en extract_work_order_page_full: %s",
            exc,
            exc_info=True,
        )
        return {
            "fecha":                 None,
            "uncertain_date":        False,
            "extraction_confidence": "FAILED",
            "entradas":              [],
            "repuestos":             [],
        }


# ---------------------------------------------------------------------------
# Helper functions / Funciones auxiliares
# ---------------------------------------------------------------------------

def _parse_date(value: str | None) -> date | None:
    """
    Parses a date string in DD/MM/YYYY format into a Python date.
    Returns None if value is None, empty or unparseable.

    ---

    Parsea una cadena de fecha en formato DD/MM/YYYY a un objeto date.
    Devuelve None si el valor es None, vacío o no parseable.
    """
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    logger.warning("# _parse_date: valor no parseable ignorado: %r", value)
    return None


def _parse_time(value: str | None) -> time | None:
    """
    Parses a time string in HH:MM format into a Python time.
    Returns None if value is None, empty or unparseable.

    ---

    Parsea una cadena de hora en formato HH:MM a un objeto time de Python.
    Devuelve None si el valor es None, vacío o no parseable.
    """
    if not value:
        return None
    try:
        parts = value.strip().split(":")
        return time(int(parts[0]), int(parts[1]))
    except (ValueError, TypeError, IndexError):
        logger.warning("# _parse_time: valor no parseable ignorado: %r", value)
        return None


def _coerce_confidence(value: str | None) -> str:
    """
    Validates the confidence value returned by Gemini against the allowed
    choices. Falls back to 'LOW' if the value is unrecognised.

    ---

    Valida el valor de confianza devuelto por Gemini contra las opciones
    permitidas. Devuelve 'LOW' si el valor no es reconocido.
    """
    allowed = {c.value for c in WorkOrderEntry.Confidence}
    if value and value.upper() in allowed:
        return value.upper()
    return WorkOrderEntry.Confidence.LOW


# Symmetric morphological confusion map for machine-code characters.
# Applied character-by-character at ANY position of the code (prefix or numeric
# block) by _resolve_machine_asset when the exact lookup fails.
# Each entry maps ONE character to its list of visually similar alternatives.
# Direction is fully bidirectional: digit↔letter and letter↔digit alike.
# Validated against the full MachineAsset catalogue (313 records, Hito 8 / S008).
#
# Mapa de confusión morfológica simétrico para caracteres de código de máquina.
# Se aplica carácter a carácter en CUALQUIER posición del código (prefijo o
# bloque numérico) por _resolve_machine_asset cuando la búsqueda exacta falla.
# Cada entrada mapea UN carácter a su lista de alternativas visualmente similares.
# La dirección es totalmente bidireccional: dígito↔letra y letra↔dígito.
_OCR_CONFUSION_MAP: dict[str, list[str]] = {
    # T written quickly degenerates to 7 (crossbar too low) or 4 (short vertical).
    # La T escrita rápido degenera a 7 (barra demasiado baja) o a 4 (vertical corta).
    "T": ["7", "4"],
    "7": ["T"],
    "4": ["T"],
    # Z is morphologically identical to 2 at speed.
    # La Z es morfológicamente idéntica al 2 a velocidad de escritura.
    "Z": ["2"],
    "2": ["Z"],
    # O and 0 are indistinguishable in fast handwriting.
    # La O y el 0 son indistinguibles en escritura rápida.
    "O": ["0"],
    "0": ["O"],
    # L/I and 1 share a single-stroke form.
    # L/I y 1 comparten trazo único.
    "L": ["1"],
    "I": ["1"],
    "1": ["L", "I"],
    # S degenerates to 5 when the curve flattens.
    # La S degenera a 5 cuando la curva se aplana.
    "S": ["5"],
    "5": ["S"],
    # G and 6 share a circular form with a horizontal stroke.
    # G y 6 comparten forma circular con trazo horizontal.
    "G": ["6"],
    "6": ["G"],
    # B and 8 share two stacked loops.
    # B y 8 comparten dos bucles apilados.
    "B": ["8"],
    "8": ["B"],
}

# Legacy map kept for backward compatibility with _normalise_machine_code.
# Only the numeric-block substitutions that are safe to apply unconditionally
# (never ambiguous in the numeric position) are preserved here.
# These are a strict subset of _OCR_CONFUSION_MAP.
#
# Mapa heredado mantenido por compatibilidad con _normalise_machine_code.
# Solo las sustituciones del bloque numérico que son seguras aplicar sin
# condición (nunca ambiguas en posición numérica) se conservan aquí.
# Son un subconjunto estricto de _OCR_CONFUSION_MAP.
_OCR_DIGIT_MAP: dict[str, str] = {
    "O": "0",
    "L": "1",
    "T": "7",
    "S": "5",
    "Z": "2",
    "G": "6",
}

# Inverse OCR map: digit misread → most likely letter candidate.
# Used in _normalise_machine_code step 6 for purely-numeric codes where the
# leading digit is actually a letter confounded by handwriting (e.g. 2→Z, 6→G).
# Only digits with a single unambiguous letter counterpart are included —
# digits with multiple plausible letter readings (e.g. 1→L/I) are excluded
# to avoid false positives.
#
# Mapa OCR inverso: dígito mal leído → letra candidata más probable.
# Usado en _normalise_machine_code paso 6 para códigos puramente numéricos
# donde el dígito inicial es en realidad una letra confundida por caligrafía
# (ej. 2→Z, 6→G). Solo se incluyen dígitos con un único candidato de letra
# inequívoco — los dígitos con múltiples lecturas plausibles (ej. 1→L/I)
# se excluyen para evitar falsos positivos.
_OCR_DIGIT_TO_LETTER_MAP: dict[str, str] = {
    "2": "Z",   # 2 ↔ Z  — trazo diagonal en escritura rápida
    "6": "G",   # 6 ↔ G  — forma circular con trazo horizontal
    "5": "S",   # 5 ↔ S  — curva degenerada a trazo vertical con visera
    "8": "B",   # 8 ↔ B  — dos bucles apilados
}


def _normalise_machine_code(raw: str | None) -> str:
    """
    Normalises a raw machine code string according to partes-trabajo skill
    directive D4, extended in Hito 8 (Paso 4, D2) with OCR character
    substitutions applied only to the numeric block, and further extended
    to handle purely-numeric codes where the leading digit is actually a
    letter misread by OCR (e.g. "294" where "2" is a misread "Z" → "Z94").

    Processing steps:
      1. Strip and uppercase.
      2. Remove internal spaces and equals signs.
      3. Multi-word alias: return uppercased with space preserved.
      4. Insert a hyphen between the leading letter(s) and the numeric block
         if not already present.
      5. Apply _OCR_DIGIT_MAP substitutions to the numeric block (right of
         hyphen) when the block contains exclusively digits.
      6. Purely-numeric codes (no letter prefix found after step 4): attempt
         to reinterpret the first digit as a letter via _OCR_DIGIT_TO_LETTER_MAP.
         If a mapping exists, return the candidate LETTER-REST form so the
         resolver can try it against the catalogue. If no mapping exists,
         return the numeric string as-is and let the resolver handle it.

    ---

    Normaliza un código de máquina bruto según la directriz D4 de la skill
    partes-trabajo, ampliada en el Hito 8 (Paso 4, D2) con sustituciones OCR
    sobre el bloque numérico, y extendida para manejar códigos puramente
    numéricos donde el dígito inicial es en realidad una letra mal leída por
    OCR (ej. "294" donde "2" es una "Z" mal leída → "Z94").

    Pasos de procesamiento:
      1. Strip y mayúsculas.
      2. Eliminar espacios internos y signos igual.
      3. Alias de múltiples palabras: devolver en mayúsculas conservando espacio.
      4. Insertar guion entre letras iniciales y bloque numérico si no está.
      5. Aplicar sustituciones _OCR_DIGIT_MAP al bloque numérico (derecha del
         guion) cuando el bloque contiene exclusivamente dígitos.
      6. Códigos puramente numéricos (sin prefijo de letras tras el paso 4):
         intentar reinterpretar el primer dígito como letra via
         _OCR_DIGIT_TO_LETTER_MAP. Si existe mapeo, devolver la forma candidata
         LETRA-RESTO para que el resolver la pruebe contra el catálogo. Si no
         existe mapeo, devolver la cadena numérica tal cual y dejar que el
         resolver la gestione.
    """
    if not raw:
        return ""
    # Multi-word alias: preserve space, return uppercase.
    # Alias de múltiples palabras: conservar espacio, devolver en mayúsculas.
    if " " in raw.strip():
        return raw.strip().upper()

    code = raw.strip().upper().replace(" ", "").replace("=", "")

    # Insert hyphen between leading letters and digits if absent.
    # Insertar guion entre letras iniciales y dígitos si no está presente.
    code = re.sub(r"^([A-Z]+)(\d)", r"\1-\2", code)

    # Apply OCR substitutions to the numeric block only (right of hyphen),
    # and only when the numeric block contains exclusively digits — never
    # when it contains letters (e.g. compound codes like "ZOP-Z124").
    # Applying the map to letter characters in the numeric block would
    # corrupt valid codes (Z→2 turning "Z124" into "2124").
    #
    # Aplicar sustituciones OCR solo al bloque numérico (derecha del guion),
    # y únicamente cuando el bloque numérico contiene exclusivamente dígitos —
    # nunca cuando contiene letras (ej. códigos compuestos como "ZOP-Z124").
    # Aplicar el mapa a caracteres de letra en el bloque numérico corrompería
    # códigos válidos (Z→2 convirtiendo "Z124" en "2124").
    if "-" in code:
        prefix, _, numeric = code.partition("-")
        if numeric.isdigit():
            numeric = "".join(_OCR_DIGIT_MAP.get(ch, ch) for ch in numeric)
        code = f"{prefix}-{numeric}"
        return code

    # Step 6 — Purely-numeric code: no letter prefix was found after hyphen
    # insertion. This occurs when OCR misread a leading letter as a digit
    # (e.g. "Z" → "2", "G" → "6", "S" → "5"). Attempt to reinterpret the
    # first character as a letter using the inverse OCR map.
    #
    # Paso 6 — Código puramente numérico: no se encontró prefijo de letras
    # tras la inserción del guion. Ocurre cuando el OCR confundió una letra
    # inicial con un dígito (ej. "Z" → "2", "G" → "6", "S" → "5"). Se intenta
    # reinterpretar el primer carácter como letra usando el mapa OCR inverso.
    if code.isdigit() and len(code) >= 2:
        first_digit = code[0]
        letter_candidate = _OCR_DIGIT_TO_LETTER_MAP.get(first_digit)
        if letter_candidate:
            rest = code[1:]
            return f"{letter_candidate}-{rest}"

    return code


# Regex to detect any Larios alias variant as the second word of machine_norm.
# All such combinations map to the catalogue code FURGLAR regardless of the
# vehicle type prefix (Mercedes, Furgon, Camion, etc.).
# Covers: "MERCEDES LARIOS", "FURGON LARIOS", "MERCEDES LORIOL", etc.
#
# Regex para detectar cualquier variante del alias Larios como segunda palabra
# de machine_norm. Todas estas combinaciones se mapean a FURGLAR independientemente
# del tipo de vehículo. Cubre: "MERCEDES LARIOS", "FURGON LARIOS", etc.
_LARIOS_TWO_WORD_RE = re.compile(
    r'(?i)^\S+\s+L[ao]r[aio][a-z]*$'
)


def _resolve_machine_asset(
    machine_norm: str,
    company=None,
) -> MachineAsset | None:
    """
    Attempts to resolve a normalised machine code to a MachineAsset record.
    Tries exact match first, then zero-padded variants (G-8 → G-08, G-08 → G-8).
    Returns None if no match is found.

    When 'company' is provided, all queries are scoped to that company.
    This is mandatory in multicompany contexts (panel views) to prevent
    cross-company asset resolution. Internal pipeline calls (tasks.py,
    services.py) may omit 'company' for backward compatibility.

    ---

    Intenta resolver un código de máquina normalizado a un registro MachineAsset.
    Prueba primero coincidencia exacta, luego variantes con ceros (G-8 → G-08,
    G-08 → G-8). Devuelve None si no se encuentra coincidencia.

    Cuando se proporciona 'company', todas las consultas se acotan a esa empresa.
    Esto es obligatorio en contextos multiempresa (vistas del panel) para evitar
    la resolución de activos entre empresas. Las llamadas internas del pipeline
    (tasks.py, services.py) pueden omitir 'company' por compatibilidad.
    """
    if not machine_norm:
        return None

    # Build base queryset — scoped to company when provided.
    # Construir queryset base — acotado a empresa cuando se proporciona.
    qs = MachineAsset.objects
    if company is not None:
        qs = qs.filter(company=company)

    # Larios alias resolution: any two-word value where the second word is a
    # Larios variant maps to FURGLAR regardless of the vehicle type prefix.
    # Resolución alias Larios: cualquier valor de dos palabras donde la segunda
    # es una variante de Larios se mapea a FURGLAR independientemente del tipo.
    if _LARIOS_TWO_WORD_RE.match(machine_norm):
        try:
            return qs.get(code="FURGLAR")
        except (MachineAsset.DoesNotExist, MachineAsset.MultipleObjectsReturned):
            pass

    # Exact match / Coincidencia exacta.
    try:
        return qs.get(code=machine_norm)
    except MachineAsset.DoesNotExist:
        pass
    except MachineAsset.MultipleObjectsReturned:
        logger.warning(
            "# _resolve_machine_asset: múltiples coincidencias para '%s'.", machine_norm
        )
        return None

    # Build candidate set: hyphen variants + no-hyphen variants.
    # The catalogue uses codes WITHOUT hyphens (e.g. "A54"), while the
    # normaliser inserts a hyphen (e.g. "A-54"). We must try both forms
    # plus zero-padding variants of each to ensure maximum resolution.
    #
    # Construir conjunto de candidatos: variantes con guion + sin guion.
    # El catálogo usa códigos SIN guion (ej. "A54"), mientras que el
    # normalizador inserta guion (ej. "A-54"). Hay que probar ambas formas
    # más variantes con relleno de ceros para maximizar la resolución.
    m = re.match(r"^([A-Z]+)-?(\d+)$", machine_norm)
    if m:
        letters, digits = m.group(1), m.group(2)
        candidates: set[str] = set()

        # Base forms: with and without hyphen.
        # Formas base: con y sin guion.
        for sep in ("-", ""):
            base = f"{letters}{sep}{digits}"
            candidates.add(base)
            # Zero-padding variants / Variantes con relleno de ceros.
            if len(digits) == 1:
                candidates.add(f"{letters}{sep}0{digits}")
                candidates.add(f"{letters}{sep}00{digits}")
            elif len(digits) == 2 and digits.startswith("0"):
                candidates.add(f"{letters}{sep}{digits[1:]}")
            elif len(digits) == 3 and digits.startswith("0"):
                candidates.add(f"{letters}{sep}{digits[1:]}")
                candidates.add(f"{letters}{sep}{digits[2:]}")

        # Remove the already-tried exact match to avoid redundant queries.
        # Eliminar la coincidencia exacta ya intentada para evitar consultas redundantes.
        candidates.discard(machine_norm)

        for candidate in sorted(candidates):
            try:
                return MachineAsset.objects.get(code=candidate)
            except (MachineAsset.DoesNotExist, MachineAsset.MultipleObjectsReturned):
                continue

    # ------------------------------------------------------------------
    # Morphological substitution candidates — Nivel 1 and Nivel 2.
    # Candidatos por sustitución morfológica — Nivel 1 y Nivel 2.
    #
    # Guard: Spanish vehicle licence plates (4 digits + 3 letters, e.g.
    # "2030JVK") are stored verbatim in the catalogue. Never apply
    # morphological substitutions to them — their format is unambiguous.
    #
    # Guarda: las matrículas españolas (4 dígitos + 3 letras, ej.
    # "2030JVK") se almacenan tal cual en el catálogo. Nunca aplicar
    # sustituciones morfológicas — su formato es inequívoco.
    # ------------------------------------------------------------------
    _PLATE_RE = re.compile(r"^\d{4}[A-Z]{3}$")
    if _PLATE_RE.match(machine_norm):
        return None

    def _substitution_candidates(code: str, max_subs: int) -> set[str]:
        """
        Generates all variants of `code` obtained by substituting up to
        `max_subs` characters using _OCR_CONFUSION_MAP. Each substitution
        replaces exactly one character at one position with one alternative.
        Returns a flat set of candidate strings (original excluded).

        ---

        Genera todas las variantes de `code` obtenidas sustituyendo hasta
        `max_subs` caracteres usando _OCR_CONFUSION_MAP. Cada sustitución
        reemplaza exactamente un carácter en una posición por una alternativa.
        Devuelve un conjunto plano de cadenas candidatas (el original excluido).
        """
        results: set[str] = set()

        def _apply_single(base: str) -> set[str]:
            """One-substitution pass over base. / Pasada de una sustitución sobre base."""
            variants: set[str] = set()
            for idx, ch in enumerate(base):
                for alt in _OCR_CONFUSION_MAP.get(ch, []):
                    variants.add(base[:idx] + alt + base[idx + 1:])
            return variants

        level1 = _apply_single(code)
        results.update(level1)

        if max_subs >= 2:
            for lvl1_code in level1:
                results.update(_apply_single(lvl1_code))

        results.discard(code)
        return results

    # Normalise the raw code to strip hyphens before generating substitution
    # candidates, then re-probe with and without hyphen insertion.
    # Normalizar el código crudo eliminando guiones antes de generar candidatos
    # de sustitución, y luego re-probar con y sin inserción de guion.
    stripped = machine_norm.replace("-", "")

    # Nivel 1 — single substitution / sustitución única.
    # Nivel 2 — two simultaneous substitutions / dos sustituciones simultáneas.
    for max_subs in (1, 2):
        sub_candidates: set[str] = _substitution_candidates(stripped, max_subs)
        # For each substitution candidate, probe both the raw form and the
        # hyphenated form (letter-prefix + hyphen + digits).
        # Para cada candidato de sustitución, probar la forma cruda y la
        # forma con guion (prefijo de letras + guion + dígitos).
        probe_set: set[str] = set()
        for cand in sub_candidates:
            probe_set.add(cand)
            # Insert hyphen between leading letters and digits if absent.
            # Insertar guion entre letras iniciales y dígitos si no está.
            hyphenated = re.sub(r"^([A-Z]+)(\d)", r"-", cand)
            probe_set.add(hyphenated)

        for candidate in sorted(probe_set):
            qs_probe = MachineAsset.objects
            if company is not None:
                qs_probe = qs_probe.filter(company=company)
            try:
                asset = qs_probe.get(code=candidate)
                logger.info(
                    "# _resolve_machine_asset: '%s' resuelto como '%s' "
                    "mediante sustitución morfológica (nivel %d).",
                    machine_norm, candidate, max_subs,
                )
                return asset
            except (MachineAsset.DoesNotExist, MachineAsset.MultipleObjectsReturned):
                continue

    return None


def _compute_delta_hours(
    hc: time | None,
    hf: time | None,
    deduct_lunch: bool = True,
) -> Decimal | None:
    """
    Computes the hours for a work block.

    When deduct_lunch=True (default, PDF pipeline): subtracts the fixed
    lunch break window (13:30–15:00, 90 min) if the block covers it, as
    defined in the partes-trabajo skill.

    When deduct_lunch=False (digital/form entry): returns the gross
    duration without any deduction. Gate 4 already handles the midday
    window as a LUNCH_BREAK gap, so no implicit deduction is needed.

    Returns a Decimal rounded to 2 decimal places, or None if either time
    is missing or hf <= hc.

    ---

    Calcula las horas de un bloque de trabajo.

    Con deduct_lunch=True (por defecto, pipeline PDF): descuenta la pausa
    de comida fija (13:30–15:00, 90 min) si el bloque la cubre, según la
    definición de la skill partes-trabajo.

    Con deduct_lunch=False (entrada digital/formulario): devuelve la
    duración bruta sin ningún descuento. Gate 4 ya gestiona la ventana de
    mediodía como laguna LUNCH_BREAK, por lo que no se necesita descuento
    implícito.

    Devuelve un Decimal redondeado a 2 decimales, o None si alguna hora
    falta o hf <= hc.
    """
    if not hc or not hf:
        return None

    # Convert to minutes since midnight / Convertir a minutos desde medianoche.
    hc_min = hc.hour * 60 + hc.minute
    hf_min = hf.hour * 60 + hf.minute

    if hf_min <= hc_min:
        return None

    total_min = hf_min - hc_min

    if deduct_lunch:
        # Lunch break deduction: 13:30–15:00 = 90 minutes (PDF pipeline only).
        # Descuento pausa comida: 13:30–15:00 = 90 min (solo pipeline PDF).
        lunch_start   = 13 * 60 + 30   # 810
        lunch_end     = 15 * 60        # 900
        overlap_start = max(hc_min, lunch_start)
        overlap_end   = min(hf_min, lunch_end)
        deduction     = max(0, overlap_end - overlap_start)
        total_min     = total_min - deduction

    net_h = Decimal(total_min) / Decimal(60)
    return net_h.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _worker_name_from_pdf_path(pdf_name: str) -> str:
    """
    Derives the worker's full name from the PDF filename.
    Expected format: NOMBRE_APELLIDO1_APELLIDO2_DD-MM_AL_DD-MM.pdf
    or any variant where the name occupies the leading underscore-separated
    tokens before a token matching DD-MM or AL.

    Returns the name in uppercase with spaces, e.g. "ALEJANDRO GARCIA LUQUE".
    Falls back to the raw filename stem if the format is unrecognised.

    ---

    Deriva el nombre completo del operario del nombre del fichero PDF.
    Formato esperado: NOMBRE_APELLIDO1_APELLIDO2_DD-MM_AL_DD-MM.pdf
    o cualquier variante donde el nombre ocupa los tokens iniciales separados
    por guion bajo antes de un token con formato DD-MM o AL.

    Devuelve el nombre en mayúsculas con espacios, ej: "ALEJANDRO GARCIA LUQUE".
    Devuelve el stem del nombre de fichero bruto si el formato no se reconoce.
    """
    import os
    stem = os.path.splitext(os.path.basename(pdf_name))[0]
    # Normalise separators: replace underscores with spaces for uniform splitting.
    # Normalizar separadores: reemplazar guiones bajos por espacios para split uniforme.
    tokens = stem.replace("_", " ").split(" ")

    # Collect tokens until the first one that starts with a digit.
    # That digit marks the start of the date range — everything before it
    # is part of the worker name.
    # Recoger tokens hasta el primero que empiece por dígito.
    # Ese dígito marca el inicio del rango de fechas — todo lo anterior
    # forma parte del nombre del operario.
    name_tokens: list[str] = []
    for tok in tokens:
        if tok and tok[0].isdigit():
            break
        name_tokens.append(tok.upper())

    return " ".join(name_tokens).strip() if name_tokens else stem.upper()


# ---------------------------------------------------------------------------
# Excel generation constants / Constantes de generación Excel
# ---------------------------------------------------------------------------

# Colour palette / Paleta de colores
_CLR_HEADER_BG    = "1F4E79"   # Azul oscuro — cabecera de columnas
_CLR_HEADER_FG    = "FFFFFF"   # Blanco — texto cabecera
_CLR_CONFIG_LABEL = "D6E4F0"   # Azul muy claro — etiquetas área config
_CLR_CONFIG_INPUT = "FFFF99"   # Amarillo — celdas de entrada C2 y C3
_CLR_DATE_FIRST   = "BDD7EE"   # Azul claro — primera entrada del día (col FECHA)
_CLR_DATE_UNCERT  = "FFD700"   # Dorado — fecha incierta
_CLR_MACHINE_UNKN = "FFD700"   # Dorado — máquina no identificada
_CLR_MACHINE_EMPT = "D9D9D9"   # Gris — máquina vacía
_CLR_REV_GREEN    = "70AD47"   # Verde — jornada estándar (= 8h)
_CLR_REV_BLUE     = "9DC3E6"   # Azul claro — horas extra (8–12h)
_CLR_REV_YELLOW   = "FFFF00"   # Amarillo — por debajo mínimo (7–8h)
_CLR_REV_ORANGE   = "F4B942"   # Naranja — jornada excesiva (12–16h)
_CLR_REV_RED      = "FF0000"   # Rojo — fuera de rango (< 7h ó > 16h)
_CLR_MANIFEST_BG  = "1F4E79"   # Azul oscuro — cabecera manifiesto
_CLR_MANIFEST_FG  = "FFFFFF"   # Blanco — texto cabecera manifiesto
_CLR_DAY_SHADE    = "EBF3FB"   # Azul muy claro — sombreado alterno de día

# Column definitions: (header, width)
# Definición de columnas: (cabecera, ancho)
_COLS = [
    ("FECHA",             14),   # A  1
    ("CÓDIGO / VEH.",     14),   # B  2
    ("MARCA / MODELO",    28),   # C  3
    ("KM",                10),   # D  4
    ("HORAS VEH.",        10),   # E  5
    ("DESCRIPCIÓN AVERÍA",44),   # F  6  — ganamos ancho al eliminar OPERARIO
    ("REPARACIÓN",        44),   # G  7  — ganamos ancho al eliminar OPERARIO
    ("H.C.",               8),   # H  8
    ("H.F.",               8),   # I  9
    ("O.R.",              12),   # J 10
    ("Δ HORAS (neta)",    12),   # K 11
    ("HORAS NETAS DÍA",  14),   # L 12
    ("HORAS EXTRAS",     12),   # M 13
    ("SALARIO EXTRAS",   14),   # N 14
    ("REVISIÓN HORARIO", 16),   # O 15
    ("COSTE M.O.",        14),   # P 16  — delta_hours × C3 (coste hora ordinaria)
]

_DATA_ROW_START = 5   # Datos comienzan en fila 5 (filas 1-3 config, 4 cabecera)


def _make_fill(hex_color: str) -> PatternFill:
    """Returns a solid PatternFill for the given hex colour string.
    --- Devuelve un PatternFill sólido para el color hex indicado."""
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _make_border_thin() -> Border:
    """Returns a thin border on all four sides.
    --- Devuelve un borde fino en los cuatro lados."""
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _revision_color(horas_netas_dia: Decimal | None) -> str:
    """
    Returns the hex colour code for the REVISION HORARIO cell based on the
    daily net hours total, according to the partes-trabajo skill colour table.

    ---

    Devuelve el código de color hex para la celda REVISIÓN HORARIO basándose
    en el total de horas netas del día, según la tabla de colores de la skill.
    """
    if horas_netas_dia is None:
        return _CLR_MACHINE_EMPT  # Gris si no hay dato.
    h = float(horas_netas_dia)
    if h == 8.0:
        return _CLR_REV_GREEN
    if 8.0 < h <= 12.0:
        return _CLR_REV_BLUE
    if 7.0 <= h < 8.0:
        return _CLR_REV_YELLOW
    if 12.0 < h <= 16.0:
        return _CLR_REV_ORANGE
    return _CLR_REV_RED   # < 7h ó > 16h


def _revision_text(horas_netas_dia: Decimal | None) -> str:
    """
    Returns the display text for the REVISION HORARIO cell.
    --- Devuelve el texto de visualización para la celda REVISIÓN HORARIO.
    """
    if horas_netas_dia is None:
        return ""
    h = float(horas_netas_dia)
    if h == 8.0:
        return "Jornada estándar"
    if 8.0 < h <= 12.0:
        return "Horas extra"
    if 7.0 <= h < 8.0:
        return "Por debajo mínimo"
    if 12.0 < h <= 16.0:
        return "Jornada excesiva"
    return "Fuera de rango"


# ---------------------------------------------------------------------------
# Public service: Excel generation (skill partes-trabajo v1.2 + col Q)
# Servicio público: generación Excel (skill partes-trabajo v1.2 + col Q)
# ---------------------------------------------------------------------------

def generate_work_order_excel(work_order_id: int) -> None:
    """
    Builds an Excel report conforming to the partes-trabajo skill v1.2
    specification, extended with column Q (COSTE M.O.) for labour cost
    imputation per work block.

    Structure:
      Row 1 — Title: PARTES DE TRABAJO — [OPERARIO] — [PERIOD]
      Row 2 — PRECIO HORA EXTRA (C2, yellow input cell)
      Row 3 — COSTE HORA ORDINARIA (C3, yellow input cell)
      Row 4 — Column headers (dark blue)
      Row 5+ — Data rows (one per WorkOrderEntryLine)
      Footer — MANIFIESTO DE INCIDENCIAS block (2 rows gap + header + rows)
    Second sheet: LEYENDA

    Saves the file to WorkOrder.excel_file and sets status to DONE.
    On failure sets status to ERROR and re-raises.

    ---

    Construye un informe Excel conforme a la especificación de la skill
    partes-trabajo v1.2, extendido con la columna Q (COSTE M.O.) para
    imputación de coste de mano de obra por bloque de trabajo.

    Estructura:
      Fila 1 — Título: PARTES DE TRABAJO — [OPERARIO] — [PERIODO]
      Fila 2 — PRECIO HORA EXTRA (C2, celda amarilla de entrada)
      Fila 3 — COSTE HORA ORDINARIA (C3, celda amarilla de entrada)
      Fila 4 — Cabeceras de columna (azul oscuro)
      Fila 5+ — Filas de datos (una por WorkOrderEntryLine)
      Pie    — Bloque MANIFIESTO DE INCIDENCIAS (2 filas separación + cabecera + filas)
    Segunda hoja: LEYENDA
    """
    work_order = WorkOrder.objects.get(pk=work_order_id)

    try:
        logger.info(
            "# Excel: iniciando generación (skill v1.2) para WorkOrder #%d.",
            work_order_id,
        )

        # ------------------------------------------------------------------
        # Gather data / Recopilar datos
        # ------------------------------------------------------------------
        entries = (
            work_order.entries
            .prefetch_related("lines__machine_asset")
            .order_by("page_number")
        )

        worker_name = ""
        if work_order.source_pdf:
            worker_name = _worker_name_from_pdf_path(work_order.source_pdf.name)

        # Build a flat list of row data grouped by day for HORAS NETAS DÍA.
        # Construir una lista plana de datos de fila agrupados por día para
        # HORAS NETAS DÍA.
        #
        # Structure: list of dicts, one per WorkOrderEntryLine, enriched with:
        #   - date_key      : date object (for day grouping)
        #   - is_first_day  : True for the first line of a new day
        #   - is_last_day   : True for the last line of a day (set in post-pass)
        #   - day_net_hours : Decimal total net hours for the day (set in post-pass)
        #   - uncertain_date: bool from parent WorkOrderEntry

        flat_rows: list[dict] = []

        for entry in entries:
            lines = list(entry.lines.order_by("line_number"))
            if not lines:
                continue

            date_key = entry.work_date

            for line in lines:
                flat_rows.append({
                    "date_key":       date_key,
                    "uncertain_date": entry.uncertain_date,
                    "worker_name":    entry.worker_name or worker_name,
                    "line":           line,
                    "is_first_day":   False,
                    "is_last_day":    False,
                    "day_net_hours":  None,
                    "day_shade":      False,
                })

        # Post-pass: compute day groups, is_first_day, is_last_day,
        # day_net_hours.
        # Post-pasada: calcular grupos de día, is_first_day, is_last_day,
        # day_net_hours.
        if flat_rows:
            # Group indices by date_key.
            # Agrupar índices por date_key.
            from collections import defaultdict
            day_indices: dict = defaultdict(list)
            for idx, row in enumerate(flat_rows):
                day_indices[row["date_key"]].append(idx)

            for day_ordinal, (day_key, indices) in enumerate(day_indices.items()):
                flat_rows[indices[0]]["is_first_day"] = True
                flat_rows[indices[-1]]["is_last_day"]  = True

                # Alternating day shade: even ordinal days get shaded.
                # Sombreado alterno: los días con índice ordinal par se sombrean.
                shade = (day_ordinal % 2 == 1)

                # Sum delta_hours for the day.
                # Sumar delta_hours del día.
                day_total = Decimal("0.00")
                all_valid = True
                for idx in indices:
                    dh = flat_rows[idx]["line"].delta_hours
                    if dh is not None:
                        day_total += dh
                    else:
                        all_valid = False

                day_net = day_total if all_valid or day_total > 0 else None
                for idx in indices:
                    flat_rows[idx]["day_net_hours"] = day_net
                    flat_rows[idx]["day_shade"]     = shade

        # ------------------------------------------------------------------
        # Build workbook / Construir libro
        # ------------------------------------------------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Partes de Trabajo"

        num_data_cols = len(_COLS)

        # --- Row 1: Title / Fila 1: Título ---
        # Period is read from the PDF filename (canonical format DD-MM-YY AL DD-MM-YY)
        # to avoid OCR year misreads polluting the title.
        # El periodo se lee del nombre del PDF (formato DD-MM-AA AL DD-MM-AA)
        # para evitar que errores de OCR en el año contaminen el título.
        period = ""
        if work_order.source_pdf:
            _p_start, _p_end = _parse_period_from_pdf_name(work_order.source_pdf.name)
            if _p_start and _p_end:
                period = (
                    f"{_p_start.strftime('%d/%m/%Y')} — {_p_end.strftime('%d/%m/%Y')}"
                )
        if not period and flat_rows:
            # Fallback: derive from data if filename parse fails.
            # Fallback: derivar de los datos si el parseo del nombre falla.
            dates = [r["date_key"] for r in flat_rows if r["date_key"]]
            if dates:
                period = (
                    f"{min(dates).strftime('%d/%m/%Y')} — {max(dates).strftime('%d/%m/%Y')}"
                )

        title_val = f"PARTES DE TRABAJO — {worker_name} — {period}".strip(" —")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=num_data_cols)
        title_cell       = ws.cell(row=1, column=1, value=title_val)
        title_cell.font  = Font(bold=True, size=13, color=_CLR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # --- Row 2: Precio hora extra / Fila 2: Precio hora extra ---
        ws.cell(row=2, column=1, value="PRECIO HORA EXTRA (euros/h):").fill = (
            _make_fill(_CLR_CONFIG_LABEL)
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        c2_cell       = ws.cell(row=2, column=3)
        c2_cell.fill  = _make_fill(_CLR_CONFIG_INPUT)
        c2_cell.number_format = '#,##0.00 "€"'
        c2_cell.alignment = Alignment(horizontal="center")

        # --- Row 3: Coste hora ordinaria / Fila 3: Coste hora ordinaria ---
        ws.cell(row=3, column=1, value="COSTE HORA ORDINARIA (euros/h):").fill = (
            _make_fill(_CLR_CONFIG_LABEL)
        )
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
        c3_cell       = ws.cell(row=3, column=3)
        c3_cell.fill  = _make_fill(_CLR_CONFIG_INPUT)
        c3_cell.number_format = '#,##0.00 "€"'
        c3_cell.alignment = Alignment(horizontal="center")

        # --- Row 4: Column headers / Fila 4: Cabeceras de columna ---
        hdr_font = Font(bold=True, color=_CLR_HEADER_FG)
        hdr_fill = _make_fill(_CLR_HEADER_BG)
        hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (hdr, width) in enumerate(_COLS, start=1):
            cell           = ws.cell(row=4, column=col_idx, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_aln
            cell.border    = _make_border_thin()
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[4].height = 30

        # --- Data rows / Filas de datos ---
        incidences: list[dict] = []   # Accumulated for the manifest.

        for row_offset, row_data in enumerate(flat_rows):
            r          = _DATA_ROW_START + row_offset
            line: WorkOrderEntryLine = row_data["line"]
            date_key   = row_data["date_key"]
            is_first   = row_data["is_first_day"]
            is_last    = row_data["is_last_day"]
            day_net    = row_data["day_net_hours"]
            fecha_inc  = row_data["uncertain_date"]
            day_shade  = row_data.get("day_shade", False)

            # Pre-fill shaded rows: apply alternating day background to all
            # columns before individual cell values are written. Cells with
            # their own fill (date, machine status, revision) will override.
            # Pre-rellenar filas sombreadas: aplicar fondo alterno de día en
            # todas las columnas antes de escribir los valores individuales.
            # Las celdas con relleno propio (fecha, estado máquina, revisión)
            # sobreescriben este fondo.
            if day_shade:
                for _sc in range(1, num_data_cols + 1):
                    ws.cell(row=r, column=_sc).fill = _make_fill(_CLR_DAY_SHADE)

            # Col A — FECHA (first entry of day only)
            # Col A — FECHA (solo primera entrada del día)
            if is_first:
                date_str = (
                    date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA"
                )
                a_cell        = ws.cell(row=r, column=1, value=date_str)
                a_cell.fill   = _make_fill(
                    _CLR_DATE_UNCERT if fecha_inc else _CLR_DATE_FIRST
                )
                a_cell.alignment = Alignment(horizontal="center", vertical="center")
                a_cell.border = _make_border_thin()

                # Incidencia de fecha — solo una vez por día (D1).
                # Date incidence — only once per day (D1).
                if fecha_inc and date_key:
                    incidences.append({
                        "fecha": date_str,
                        "tramo": f"{line.hc.strftime('%H:%M') if line.hc else '?'}–"
                                 f"{line.hf.strftime('%H:%M') if line.hf else '?'}",
                        "campo": "FECHA",
                        "descripcion": "Fecha incierta — verificar manuscrito original.",
                    })
            else:
                ws.cell(row=r, column=1).border = _make_border_thin()

            # Col B — CÓDIGO / VEH.
            code_val  = line.machine_norm or line.machine_raw or ""
            c_cell      = ws.cell(row=r, column=2, value=code_val)
            c_cell.border = _make_border_thin()
            if not code_val:
                c_cell.fill = _make_fill(_CLR_MACHINE_EMPT)
            elif not line.machine_asset:
                c_cell.fill = _make_fill(_CLR_MACHINE_UNKN)
                incidences.append({
                    "fecha": date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA",
                    "tramo": f"{line.hc.strftime('%H:%M') if line.hc else '?'}–"
                             f"{line.hf.strftime('%H:%M') if line.hf else '?'}",
                    "campo": "MAQUINA",
                    "descripcion": (
                        f"Código '{code_val}' no encontrado en catálogo "
                        "tras normalización D4."
                    ),
                })

            # Col C — MARCA / MODELO
            marca_val = (
                line.machine_asset.brand_model if line.machine_asset else ""
            )
            ws.cell(row=r, column=3,
                    value=marca_val).border = _make_border_thin()

            # Col D — KM (from MachineAsset catalogue snapshot)
            # Col D — KM (del snapshot del catálogo MachineAsset)
            kms_val = line.machine_asset.mileage if line.machine_asset else ""
            ws.cell(row=r, column=4,
                    value=kms_val).border = _make_border_thin()

            # Col E — HORAS VEH.
            horas_val = line.machine_asset.hours if line.machine_asset else ""
            ws.cell(row=r, column=5,
                    value=horas_val).border = _make_border_thin()

            # Col F — DESCRIPCIÓN AVERÍA
            g_cell = ws.cell(row=r, column=6,
                             value=line.fault_description or "")
            g_cell.alignment = Alignment(wrap_text=True, vertical="top")
            g_cell.border    = _make_border_thin()
            if "DESCRIPCION" in (line.flags or []):
                incidences.append({
                    "fecha": date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA",
                    "tramo": f"{line.hc.strftime('%H:%M') if line.hc else '?'}–"
                             f"{line.hf.strftime('%H:%M') if line.hf else '?'}",
                    "campo": "DESCRIPCIÓN",
                    "descripcion": "Descripción de difícil lectura — verificar manuscrito.",
                })

            # Col G — REPARACIÓN
            h_cell = ws.cell(row=r, column=7, value=line.repair_notes or "")
            h_cell.alignment = Alignment(wrap_text=True, vertical="top")
            h_cell.border    = _make_border_thin()

            # Col H — H.C.
            hc_str  = line.hc.strftime("%H:%M") if line.hc else ""
            i_cell  = ws.cell(row=r, column=8, value=hc_str)
            i_cell.alignment = Alignment(horizontal="center")
            i_cell.border    = _make_border_thin()
            if "H.C." in (line.flags or []):
                incidences.append({
                    "fecha": date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA",
                    "tramo": f"{hc_str or '?'}–{line.hf.strftime('%H:%M') if line.hf else '?'}",
                    "campo": "H.C.",
                    "descripcion": "H.C. de difícil lectura — verificar manuscrito.",
                })

            # Col I — H.F.
            hf_str  = line.hf.strftime("%H:%M") if line.hf else ""
            j_cell  = ws.cell(row=r, column=9, value=hf_str)
            j_cell.alignment = Alignment(horizontal="center")
            j_cell.border    = _make_border_thin()
            if "H.F." in (line.flags or []):
                incidences.append({
                    "fecha": date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA",
                    "tramo": f"{hc_str or '?'}–{hf_str or '?'}",
                    "campo": "H.F.",
                    "descripcion": "H.F. de difícil lectura — verificar manuscrito.",
                })

            # Horario inválido: HF <= HC / Invalid schedule: HF <= HC
            if line.hc and line.hf and line.hf <= line.hc:
                incidences.append({
                    "fecha": date_key.strftime("%d/%m/%Y") if date_key else "SIN FECHA",
                    "tramo": f"{hc_str}–{hf_str}",
                    "campo": "HORARIO",
                    "descripcion": "H.F. igual o anterior a H.C. — horario inválido.",
                })

            # Col J — O.R.
            ws.cell(row=r, column=10,
                    value=line.or_val or "").border = _make_border_thin()

            # Col K — Δ HORAS (neta)
            l_cell = ws.cell(
                row=r, column=11,
                value=float(line.delta_hours) if line.delta_hours is not None else "",
            )
            l_cell.number_format = '0.00" h"'
            l_cell.alignment     = Alignment(horizontal="center")
            l_cell.border        = _make_border_thin()

            # Col L — HORAS NETAS DÍA (last entry of day only)
            # Col L — HORAS NETAS DÍA (solo última entrada del día)
            if is_last and day_net is not None:
                m_cell = ws.cell(row=r, column=12, value=float(day_net))
                m_cell.number_format = '0.00" h"'
                m_cell.alignment     = Alignment(horizontal="center")
                m_cell.border        = _make_border_thin()
                m_cell.font          = Font(bold=True)

                # Incidencia D5: jornada < 8h / D5 incidence: shift < 8h
                if day_net < Decimal("8.00") and date_key:
                    incidences.append({
                        "fecha": date_key.strftime("%d/%m/%Y"),
                        "tramo": f"{hc_str or '?'}–{hf_str or '?'}",
                        "campo": "JORNADA",
                        "descripcion": (
                            f"Jornada diaria neta ({float(day_net):.2f} h) inferior "
                            "a 8 h — verificar con el operario."
                        ),
                    })
            else:
                ws.cell(row=r, column=12).border = _make_border_thin()

            # Col M — HORAS EXTRAS (formula, last entry only)
            # Col M — HORAS EXTRAS (fórmula, solo última entrada)
            if is_last and day_net is not None:
                n_cell = ws.cell(
                    row=r, column=13,
                    value=f'=IF(OR(L{r}="",L{r}<=0),"",L{r}-8)',
                )
                n_cell.number_format = '0.00" h"'
                n_cell.alignment     = Alignment(horizontal="center")
                n_cell.border        = _make_border_thin()
                # Red font if negative / Fuente roja si negativo.
            else:
                ws.cell(row=r, column=13).border = _make_border_thin()

            # Col N — SALARIO EXTRAS (formula, last entry only)
            # Only populated when HORAS EXTRAS > 0 to avoid negative salary values.
            # Col N — SALARIO EXTRAS (fórmula, solo última entrada)
            if is_last and day_net is not None:
                _or_formula = f'=IF(OR(M{r}="",M{r}<=0),"",M{r}*$C$2)'
                o_cell = ws.cell(
                    row=r, column=14,
                    value=_or_formula,
                )
                o_cell.number_format = '#,##0.00 "€"'
                o_cell.alignment     = Alignment(horizontal="center")
                o_cell.border        = _make_border_thin()
            else:
                ws.cell(row=r, column=14).border = _make_border_thin()

            # Col O — REVISIÓN HORARIO (last entry only)
            # Col O — REVISIÓN HORARIO (solo última entrada)
            if is_last:
                p_cell = ws.cell(
                    row=r, column=15,
                    value=_revision_text(day_net),
                )
                p_cell.fill      = _make_fill(_revision_color(day_net))
                p_cell.alignment = Alignment(horizontal="center", vertical="center")
                p_cell.border    = _make_border_thin()
                p_cell.font      = Font(bold=True)
            else:
                ws.cell(row=r, column=15).border = _make_border_thin()

            # Col P — COSTE M.O. (formula: delta_hours * $C$3)
            # Col P — COSTE M.O. (fórmula: delta_hours * $C$3)
            if line.delta_hours is not None:
                q_cell = ws.cell(
                    row=r, column=16,
                    value=f'=IFERROR(IF($C$3=0,"",K{r}*$C$3),"")',
                )
                q_cell.number_format = '#,##0.00 "€"'
                q_cell.alignment     = Alignment(horizontal="center")
                q_cell.border        = _make_border_thin()
            else:
                ws.cell(row=r, column=16).border = _make_border_thin()

        # --- Freeze panes / Fijar paneles ---
        ws.freeze_panes = f"A{_DATA_ROW_START}"

        # ------------------------------------------------------------------
        # TOTALS ROW / FILA DE TOTALES
        # One row below the last data row, spanning all 16 columns.
        # Una fila por debajo de la última fila de datos, 16 columnas.
        # ------------------------------------------------------------------
        last_data_row = _DATA_ROW_START + len(flat_rows) - 1
        totals_row    = last_data_row + 1

        tot_font = Font(bold=True, color=_CLR_HEADER_FG)
        tot_fill = _make_fill(_CLR_MANIFEST_BG)

        # Label spanning cols A–J / Etiqueta abarcando cols A–J
        ws.merge_cells(
            start_row=totals_row, start_column=1,
            end_row=totals_row,   end_column=10,
        )
        tot_label           = ws.cell(row=totals_row, column=1, value="TOTALES")
        tot_label.font      = tot_font
        tot_label.fill      = tot_fill
        tot_label.alignment = Alignment(horizontal="center", vertical="center")
        tot_label.border    = _make_border_thin()

        # Build SUMPRODUCT range strings / Construir cadenas de rango SUMPRODUCT
        _dr_start = _DATA_ROW_START
        _dr_end   = last_data_row

        def _sumif(col_letter):
            """Returns a SUMPRODUCT formula summing only numeric cells.
            Uses ISNUMBER to safely ignore empty-string cells returned by
            IF formulas, avoiding #VALUE! errors in the totals row.
            --- Devuelve una fórmula SUMPRODUCT que suma solo celdas numéricas.
            Usa ISNUMBER para ignorar celdas con cadena vacía devueltas por
            fórmulas IF, evitando errores #¡VALOR! en la fila de totales.
            """
            rng = f"{col_letter}{_dr_start}:{col_letter}{_dr_end}"
            return '=SUMPRODUCT(ISNUMBER(' + rng + ')*(' + rng + '+0))'

        # Col K — Total Δ HORAS (neta) / Total Δ HORAS (neta)
        tot_l               = ws.cell(row=totals_row, column=11, value=_sumif("K"))
        tot_l.number_format = '0.00" h"'
        tot_l.alignment     = Alignment(horizontal="center")
        tot_l.font          = tot_font
        tot_l.fill          = tot_fill
        tot_l.border        = _make_border_thin()

        # Col L — Total HORAS NETAS DÍA / Total HORAS NETAS DÍA
        tot_m               = ws.cell(row=totals_row, column=12, value=_sumif("L"))
        tot_m.number_format = '0.00" h"'
        tot_m.alignment     = Alignment(horizontal="center")
        tot_m.font          = tot_font
        tot_m.fill          = tot_fill
        tot_m.border        = _make_border_thin()

        # Col M — Total HORAS EXTRAS / Total HORAS EXTRAS
        tot_n               = ws.cell(row=totals_row, column=13, value=_sumif("M"))
        tot_n.number_format = '0.00" h"'
        tot_n.alignment     = Alignment(horizontal="center")
        tot_n.font          = tot_font
        tot_n.fill          = tot_fill
        tot_n.border        = _make_border_thin()

        # Col N — Total SALARIO EXTRAS / Total SALARIO EXTRAS
        tot_o               = ws.cell(row=totals_row, column=14, value=_sumif("N"))
        tot_o.number_format = '#,##0.00 "€"'
        tot_o.alignment     = Alignment(horizontal="center")
        tot_o.font          = tot_font
        tot_o.fill          = tot_fill
        tot_o.border        = _make_border_thin()

        # Col O — Leyenda dinámica precio hora extra.
        # Warns if B2=0 (price not entered), else shows applied price.
        # Avisa si B2=0 (precio no introducido), si no muestra el precio aplicado.
        _leg_warn   = ("ATENCIÓN: introduzca el precio de la hora extra "
                       "en C2 para calcular el salario total.")
        _leg_ok_str = '=IF($C$2=0,"' + _leg_warn + '","Precio hora extra aplicado: "&TEXT($C$2,"#,##0.00")&" EUR/h")'
        tot_p               = ws.cell(row=totals_row, column=15, value=_leg_ok_str)
        tot_p.alignment     = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
        tot_p.font          = Font(italic=True, bold=True, color=_CLR_HEADER_FG)
        tot_p.fill          = tot_fill
        tot_p.border        = _make_border_thin()

        # Col P — Total COSTE M.O. / Total COSTE M.O.
        tot_q               = ws.cell(row=totals_row, column=16, value=_sumif("P"))
        tot_q.number_format = '#,##0.00 "€"'
        tot_q.alignment     = Alignment(horizontal="center")
        tot_q.font          = tot_font
        tot_q.fill          = tot_fill
        tot_q.border        = _make_border_thin()

        ws.row_dimensions[totals_row].height = 30

        # ------------------------------------------------------------------
        # MANIFIESTO DE INCIDENCIAS / INCIDENCE MANIFEST
        # Extended to all 16 columns (A→P), matching report width.
        # Description column merged from E to P, wrap_text=False, fixed height.
        # Extendido a las 16 columnas (A→P), igual ancho que el reporte.
        # Columna descripción fusionada de E a P, wrap_text=False, altura fija.
        # ------------------------------------------------------------------
        manifest_start = totals_row + 3   # 2 blank rows gap + header

        num_inc = len(incidences)

        # Manifest title spanning all 17 cols / Título manifiesto en 17 cols
        ws.merge_cells(
            start_row=manifest_start, start_column=1,
            end_row=manifest_start,   end_column=num_data_cols,
        )
        mhdr = ws.cell(
            row=manifest_start, column=1,
            value=f"MANIFIESTO DE INCIDENCIAS  —  {num_inc} incidencia(s) detectada(s)",
        )
        mhdr.font      = Font(bold=True, color=_CLR_MANIFEST_FG, size=11)
        mhdr.fill      = _make_fill(_CLR_MANIFEST_BG)
        mhdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[manifest_start].height = 22

        # Column headers row / Fila de cabeceras del manifiesto
        manifest_col_row = manifest_start + 1
        for ci, ch in enumerate(["#", "FECHA", "TRAMO", "CAMPO"], start=1):
            cell           = ws.cell(row=manifest_col_row, column=ci, value=ch)
            cell.font      = Font(bold=True, color=_CLR_MANIFEST_FG)
            cell.fill      = _make_fill(_CLR_MANIFEST_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _make_border_thin()
        # Description header merged from col 5 to 17
        # Cabecera descripción fusionada de col 5 a 17
        ws.merge_cells(
            start_row=manifest_col_row, start_column=5,
            end_row=manifest_col_row,   end_column=num_data_cols,
        )
        desc_hdr           = ws.cell(row=manifest_col_row, column=5,
                                     value="DESCRIPCIÓN DE LA INCIDENCIA")
        desc_hdr.font      = Font(bold=True, color=_CLR_MANIFEST_FG)
        desc_hdr.fill      = _make_fill(_CLR_MANIFEST_BG)
        desc_hdr.alignment = Alignment(horizontal="center", vertical="center")
        desc_hdr.border    = _make_border_thin()
        ws.row_dimensions[manifest_col_row].height = 20

        if incidences:
            for inc_idx, inc in enumerate(incidences, start=1):
                inc_row = manifest_col_row + inc_idx
                ws.cell(row=inc_row, column=1,
                        value=inc_idx).border = _make_border_thin()
                ws.cell(row=inc_row, column=2,
                        value=inc["fecha"]).border = _make_border_thin()
                ws.cell(row=inc_row, column=3,
                        value=inc["tramo"]).border = _make_border_thin()
                ws.cell(row=inc_row, column=4,
                        value=inc["campo"]).border = _make_border_thin()
                # Description merged from col 5 to 17, single line, no wrap.
                # Descripción fusionada de col 5 a 17, línea única, sin wrap.
                ws.merge_cells(
                    start_row=inc_row, start_column=5,
                    end_row=inc_row,   end_column=num_data_cols,
                )
                desc_cell           = ws.cell(row=inc_row, column=5,
                                              value=inc["descripcion"])
                desc_cell.alignment = Alignment(horizontal="left",
                                                vertical="center",
                                                wrap_text=False)
                desc_cell.border    = _make_border_thin()
                ws.row_dimensions[inc_row].height = 16
        else:
            no_inc_row = manifest_col_row + 1
            ws.merge_cells(
                start_row=no_inc_row, start_column=1,
                end_row=no_inc_row,   end_column=num_data_cols,
            )
            no_inc           = ws.cell(row=no_inc_row, column=1,
                                       value="Sin incidencias registradas.")
            no_inc.font      = Font(italic=True)
            no_inc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[no_inc_row].height = 16

        # ------------------------------------------------------------------
        # LEYENDA sheet / Hoja LEYENDA
        # ------------------------------------------------------------------
        ws_ley = wb.create_sheet(title="LEYENDA")
        # Write value BEFORE merging — openpyxl generates invalid sheet2.xml
        # if merge_cells is called before the cell value is set.
        # Escribir el valor ANTES de fusionar — openpyxl genera sheet2.xml
        # inválido si merge_cells se llama antes de establecer el valor.
        ley_title           = ws_ley.cell(row=1, column=1,
                                          value="LEYENDA — REVISIÓN HORARIO")
        ley_title.font      = Font(bold=True, size=12, color=_CLR_HEADER_BG)
        ley_title.alignment = Alignment(horizontal="center", vertical="center")
        ws_ley.merge_cells("A1:C1")
        ws_ley.row_dimensions[1].height = 20

        leyenda_rows = [
            (_CLR_REV_GREEN,  "8 h",      "Jornada estándar"),
            (_CLR_REV_BLUE,   "8 h – 12 h", "Horas extraordinarias"),
            (_CLR_REV_YELLOW, "7 h – 8 h",  "Por debajo del mínimo"),
            (_CLR_REV_ORANGE, "12 h – 16 h","Jornada excesiva — revisar"),
            (_CLR_REV_RED,    "< 7 h ó > 16 h", "Fuera de rango — error probable"),
        ]
        for ley_idx, (color, rango, significado) in enumerate(leyenda_rows, start=2):
            color_cell           = ws_ley.cell(row=ley_idx, column=1, value="")
            color_cell.fill      = _make_fill(color)
            color_cell.border    = _make_border_thin()
            ws_ley.cell(row=ley_idx, column=2,
                        value=rango).border      = _make_border_thin()
            ws_ley.cell(row=ley_idx, column=3,
                        value=significado).border = _make_border_thin()

        ws_ley.column_dimensions["A"].width = 6
        ws_ley.column_dimensions["B"].width = 16
        ws_ley.column_dimensions["C"].width = 40

        # ------------------------------------------------------------------
        # REPUESTOS sheet / Hoja REPUESTOS
        # One row per SparePartLine linked to the WorkOrderEntryLine records
        # of this WorkOrder. Columns: FECHA, BLOQUE, REFERENCIA, VEHICULO,
        # MATERIAL, UNIDADES, ORIGEN, PROVEEDOR.
        # Only created when at least one SparePartLine exists.
        #
        # Hoja REPUESTOS: una fila por SparePartLine vinculada a las
        # WorkOrderEntryLine de este WorkOrder. Columnas: FECHA, BLOQUE,
        # REFERENCIA, VEHICULO, MATERIAL, UNIDADES, ORIGEN, PROVEEDOR.
        # Solo se crea cuando existe al menos un SparePartLine.
        # ------------------------------------------------------------------
        from work_order_processor.models import SparePartLine

        spare_parts = (
            SparePartLine.objects
            .filter(entry_line__entry__work_order=work_order)
            .select_related(
                "entry_line__entry",
                "entry_line__machine_asset",
                "vehicle",
            )
            .order_by(
                "entry_line__entry__page_number",
                "entry_line__line_number",
                "line_number",
            )
        )

        if spare_parts.exists():
            ws_rep = wb.create_sheet(title="Repuestos")

            # -- Header row / Fila de cabecera --
            _REP_COLS = [
                ("FECHA",      14),
                ("BLOQUE",     10),
                ("REFERENCIA", 18),
                ("VEHICULO",   14),
                ("MATERIAL",   40),
                ("UNIDADES",   10),
                ("ORIGEN",     12),
                ("PROVEEDOR",  24),
            ]
            rep_hdr_font = Font(bold=True, color=_CLR_HEADER_FG)
            rep_hdr_fill = _make_fill(_CLR_HEADER_BG)
            rep_hdr_aln  = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

            for ci, (hdr, width) in enumerate(_REP_COLS, start=1):
                cell           = ws_rep.cell(row=1, column=ci, value=hdr)
                cell.font      = rep_hdr_font
                cell.fill      = rep_hdr_fill
                cell.alignment = rep_hdr_aln
                cell.border    = _make_border_thin()
                ws_rep.column_dimensions[get_column_letter(ci)].width = width

            ws_rep.row_dimensions[1].height = 24

            # -- Data rows / Filas de datos --
            for rep_offset, spl in enumerate(spare_parts, start=2):
                entry      = spl.entry_line.entry
                date_str   = (
                    entry.work_date.strftime("%d/%m/%Y")
                    if entry.work_date else ""
                )
                bloque_val = f"Bloque {spl.entry_line.line_number}"

                # VEHICULO: prefer resolved MachineAsset code, fall back to
                # raw vehicle name stored on the SparePartLine.
                # VEHICULO: preferir código MachineAsset resuelto; si no,
                # usar el nombre de vehículo crudo almacenado en SparePartLine.
                if spl.vehicle:
                    vehiculo_val = spl.vehicle.code or spl.vehicle.brand_model or ""
                else:
                    # vehicle_raw is not stored on SparePartLine; use reference
                    # as fallback since it may carry the vehicle code in some cases.
                    # vehicle_raw no se almacena en SparePartLine; usar referencia
                    # como fallback ya que puede contener el código de vehículo.
                    vehiculo_val = ""

                origen_val   = spl.get_source_display()
                proveedor_val = spl.supplier if spl.source == SparePartLine.Source.SUPPLIER else ""

                row_vals = [
                    date_str,
                    bloque_val,
                    spl.reference  or "",
                    vehiculo_val,
                    spl.material   or "",
                    float(spl.quantity) if spl.quantity is not None else "",
                    origen_val,
                    proveedor_val,
                ]
                for ci, val in enumerate(row_vals, start=1):
                    cell           = ws_rep.cell(row=rep_offset, column=ci, value=val)
                    cell.border    = _make_border_thin()
                    cell.alignment = Alignment(vertical="center")
                    if ci == 6:
                        # UNIDADES — right-aligned, 2 decimal places.
                        # UNIDADES — alineación derecha, 2 decimales.
                        cell.alignment    = Alignment(horizontal="right",
                                                      vertical="center")
                        cell.number_format = "0.00"

                ws_rep.row_dimensions[rep_offset].height = 16

            ws_rep.freeze_panes = "A2"

            logger.info(
                "# Excel: hoja Repuestos generada con %d líneas para "
                "WorkOrder #%d.",
                spare_parts.count(),
                work_order_id,
            )

        # ------------------------------------------------------------------
        # Persist / Persistir
        # ------------------------------------------------------------------
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f"partes_{worker_name.replace(' ', '_').lower()}_"
            f"{work_order_id}_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        work_order.excel_file.save(filename, ContentFile(buffer.read()), save=False)
        work_order.status = WorkOrder.Status.DONE
        work_order.save(update_fields=["excel_file", "status"])

        logger.info(
            "# Excel: informe generado correctamente para WorkOrder #%d → %s.",
            work_order_id,
            filename,
        )

    except Exception as exc:
        logger.error(
            "# Error crítico en generate_work_order_excel para WorkOrder #%d: %s",
            work_order_id,
            exc,
            exc_info=True,
        )
        work_order.status    = WorkOrder.Status.ERROR
        work_order.error_log = f"Error en generación de Excel: {exc}"
        work_order.save(update_fields=["status", "error_log"])
        raise


# ---------------------------------------------------------------------------
# classify_fault — automatic fault classification via Gemini Flash (text-only)
# classify_fault — clasificación automática de avería vía Gemini Flash (solo texto)
# ---------------------------------------------------------------------------

# Response schema for classify_fault — guarantees both fields are always present
# in the JSON output. Values must be valid FaultCategory / FaultSubcategory codes.
#
# Esquema de respuesta para classify_fault — garantiza que ambos campos estén
# siempre presentes en el JSON de salida. Los valores deben ser códigos válidos
# de FaultCategory / FaultSubcategory.
_CLASSIFY_RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "fault_category":    {"type": "string"},
        "fault_subcategory": {"type": "string"},
    },
    "required": ["fault_category", "fault_subcategory"],
}

# Classification prompt — built once at module level to avoid repeated string
# construction. Embeds the full taxonomy so Gemini can pick the exact codes.
#
# Prompt de clasificación — construido una vez a nivel de módulo para evitar
# construcción repetida de cadenas. Embebe la taxonomía completa para que
# Gemini pueda elegir los códigos exactos.
_CLASSIFY_PROMPT = """Eres un sistema de clasificación automática de averías en maquinaria industrial
pesada (grúas, carretillas elevadoras, equipos de obra, camiones).

Se te proporciona la descripción de una avería y las notas de reparación de un
parte de trabajo de taller. Tu tarea es asignar el par de códigos más preciso
de la taxonomía siguiente.

TAXONOMÍA — Categorías principales (fault_category):
  ENGINE_TRANSMISSION        — Motor, transmisión, PTO, refrigeración, combustible
  HYDRAULIC                  — Bomba hidráulica, cilindros, válvulas, aceite, central
  ELECTRICAL_ELECTRONIC      — Cableado, sensores, mandos, iluminación, batería
  BRAKES_STEERING_SUSPENSION — Frenos, dirección, suspensión
  TYRES_RUNNING_GEAR         — Neumáticos, ejes, cadenas y rodadura oruga
  LIFTING_STRUCTURE          — Pluma, gancho/poleas, cable, rotación, estabilizadores,
                               mástil/horquillas, plataforma, quinta rueda, chasis semirremolque
  BODYWORK_CHASSIS           — Carrocería, chasis estructural
  OTHER                      — Cualquier avería que no encaje en los grupos anteriores

TAXONOMÍA — Subcategorías (fault_subcategory):
  ET_ENGINE | ET_TRANSMISSION | ET_PTO | ET_COOLING | ET_FUEL
  HY_PUMP | HY_CYLINDERS | HY_VALVES | HY_OIL | HY_CENTRAL
  EE_WIRING | EE_SENSORS | EE_CONTROLS | EE_LIGHTS | EE_BATTERY
  BSS_BRAKES | BSS_STEERING | BSS_SUSPENSION
  TRG_TYRES | TRG_AXLES | TRG_TRACKS
  LS_BOOM | LS_HOOK_PULLEYS | LS_CABLE | LS_ROTATION | LS_STABILIZERS |
  LS_MAST | LS_PLATFORM | LS_FIFTH_WHEEL | LS_CHASSIS_TRAILER
  BC_BODYWORK | BC_CHASSIS
  OT_OTHER

REGLAS OBLIGATORIAS:
1. Responde ÚNICAMENTE con un objeto JSON válido. Sin texto adicional, sin
   bloques de código markdown, sin explicaciones.
2. Los dos campos son obligatorios. Usa exactamente los códigos de la taxonomía.
3. La subcategoría debe pertenecer a la categoría elegida (mismos prefijos).
4. Si la información es insuficiente o no encaja, usa OTHER / OT_OTHER.

Formato de respuesta exacto:
{{
  "fault_category": "<CODIGO_CATEGORIA>",
  "fault_subcategory": "<CODIGO_SUBCATEGORIA>"
}}

Descripción de la avería: {fault_description}
Notas de reparación: {repair_notes}
"""

# Valid code sets — used for post-extraction validation.
# Conjuntos de códigos válidos — usados para validación post-extracción.
_VALID_CATEGORIES    = {c.value for c in FaultCategory}
_VALID_SUBCATEGORIES = {s.value for s in FaultSubcategory}


def classify_fault(
    fault_description: str,
    repair_notes: str,
) -> tuple[str, str]:
    """
    Sends fault_description and repair_notes to Gemini Flash (text-only,
    Vertex AI) and returns a (fault_category, fault_subcategory) tuple
    whose values are guaranteed to be valid FaultCategory / FaultSubcategory
    codes. On any error or invalid response, falls back to ("", "").

    Uses thinking_budget=0 and temperature=0.0 for deterministic, fast
    classification. max_output_tokens=64 is sufficient for the two-field JSON.
    A single attempt is made — retries are delegated to the Celery task layer
    (classify_fault_line). The 429-retry loop (3 attempts, 60 s wait) defined
    in the Celery task layer handles Vertex AI contention (Key Learning).

    ---

    Envía fault_description y repair_notes a Gemini Flash (solo texto,
    Vertex AI) y devuelve una tupla (fault_category, fault_subcategory)
    cuyos valores están garantizados como códigos válidos de FaultCategory /
    FaultSubcategory. Ante cualquier error o respuesta inválida, retorna ("", "").

    Usa thinking_budget=0 y temperature=0.0 para clasificación determinista
    y rápida. max_output_tokens=64 es suficiente para el JSON de dos campos.
    Se realiza un único intento — los reintentos se delegan a la capa de tarea
    Celery (classify_fault_line). El bucle de reintento por 429 (3 intentos,
    espera 60 s) definido en la capa Celery gestiona la contención de Vertex AI
    (Key Learning).
    """
    prompt = _CLASSIFY_PROMPT.format(
        fault_description=fault_description or "(sin descripción)",
        repair_notes=repair_notes or "(sin notas)",
    )

    try:
        client = _get_gemini_client()

        response = client.models.generate_content(
            model    = _GEMINI_MODEL,
            contents = [prompt],
            config   = GenerateContentConfig(
                http_options       = HttpOptions(timeout=30_000),
                response_mime_type = "application/json",
                response_schema    = _CLASSIFY_RESPONSE_SCHEMA,
                thinking_config    = ThinkingConfig(thinking_budget=0),
                temperature        = 0.0,
                max_output_tokens  = 64,
            ),
        )

        raw      = response.text.strip()
        parsed   = json.loads(raw)
        category = str(parsed.get("fault_category", "")).strip()
        subcat   = str(parsed.get("fault_subcategory", "")).strip()

        # Validate that both codes belong to the taxonomy.
        # Validar que ambos códigos pertenecen a la taxonomía.
        if category not in _VALID_CATEGORIES or subcat not in _VALID_SUBCATEGORIES:
            logger.warning(
                "# [classify_fault] Códigos fuera de taxonomía devueltos por Gemini: "
                "category=%r subcategory=%r. Se usará fallback vacío.",
                category,
                subcat,
            )
            return "", ""

        logger.info(
            "# [classify_fault] Clasificación completada: category=%s subcategory=%s.",
            category,
            subcat,
        )
        return category, subcat

    except Exception as exc:
        logger.error(
            "# [classify_fault] Error en clasificación Gemini: %s",
            exc,
            exc_info=True,
        )
        return "", ""


# ---------------------------------------------------------------------------
# find_cached_classification — exact-match lookup within the same company
# find_cached_classification — búsqueda por coincidencia exacta en la misma empresa
# ---------------------------------------------------------------------------

def find_cached_classification(
    fault_description: str,
    repair_notes: str,
    company,
) -> tuple[str, str] | None:
    """
    Looks up an already-classified WorkOrderEntryLine belonging to the same
    company whose fault_description and repair_notes match the provided values
    exactly (case-insensitive, leading/trailing whitespace stripped).

    Returns a (fault_category, fault_subcategory) tuple if a match with a
    non-empty fault_category is found, or None otherwise.

    This function is called as a pre-enqueue gate in the three WorkOrderEntryLine
    INSERT points (WorkOrderEntryConfirmView, WorkOrderEntryFormView,
    WorkOrderEntryMergeView). When a match is found the classification is copied
    directly, avoiding a Gemini Flash inference call. When no match is found,
    the caller enqueues classify_fault_line for asynchronous classification.

    Scope is deliberately limited to the same company: fault taxonomy varies
    significantly across companies (e.g. crane repairs vs. assembly work vs.
    administrative tasks), so cross-company matches would produce incorrect
    classifications.

    ---

    Busca una WorkOrderEntryLine ya clasificada de la misma empresa cuya
    fault_description y repair_notes coincidan exactamente con los valores
    proporcionados (sin distinción de mayúsculas, espacios iniciales/finales
    eliminados).

    Devuelve una tupla (fault_category, fault_subcategory) si existe una
    coincidencia con fault_category no vacío, o None en caso contrario.

    Esta función se invoca como gate previo al encolado en los tres puntos de
    INSERT de WorkOrderEntryLine (WorkOrderEntryConfirmView,
    WorkOrderEntryFormView, WorkOrderEntryMergeView). Cuando se encuentra
    coincidencia, la clasificación se copia directamente, evitando una llamada
    de inferencia a Gemini Flash. Cuando no hay coincidencia, el llamador
    encola classify_fault_line para clasificación asíncrona.

    El scope se limita deliberadamente a la misma empresa: la taxonomía de
    averías varía significativamente entre empresas (p.ej. reparaciones de
    grúa vs. trabajos de montaje vs. tareas administrativas), por lo que las
    coincidencias entre empresas producirían clasificaciones incorrectas.
    """
    # Normalise inputs — strip and lower for case-insensitive exact match.
    # Normalizar entradas — strip y lower para coincidencia exacta sin case.
    norm_description = (fault_description or "").strip().lower()
    norm_notes       = (repair_notes or "").strip().lower()

    try:
        match = (
            WorkOrderEntryLine.objects
            .filter(
                entry__work_order__company=company,
                fault_category__gt="",  # non-empty — already classified
            )
            .extra(
                where=[
                    "LOWER(TRIM(fault_description)) = %s",
                    "LOWER(TRIM(repair_notes)) = %s",
                ],
                params=[norm_description, norm_notes],
            )
            .values("fault_category", "fault_subcategory")
            .first()
        )

        if match:
            logger.info(
                "# [find_cached_classification] Coincidencia encontrada: "
                "category=%s subcategory=%s.",
                match["fault_category"],
                match["fault_subcategory"],
            )
            return match["fault_category"], match["fault_subcategory"]

        logger.debug(
            "# [find_cached_classification] Sin coincidencia para "
            "description=%r notes=%r. Se encolará classify_fault_line.",
            norm_description[:80],
            norm_notes[:80],
        )
        return None

    except Exception as exc:
        logger.error(
            "# [find_cached_classification] Error en lookup de caché: %s",
            exc,
            exc_info=True,
        )
        # On any DB error, fall through to Celery classification.
        # Ante cualquier error de BD, dejar pasar al encolado Celery.
        return None


# ---------------------------------------------------------------------------
# build_export_from_template — Excel generation from ExportTemplate
# build_export_from_template — generación de Excel desde ExportTemplate
# ---------------------------------------------------------------------------

def build_export_from_template(template, work_orders_qs):
    """
    Builds an openpyxl Workbook from a WorkOrder queryset according to the
    configuration stored in the given ExportTemplate instance.

    Supported template.sheet_format values:
      single_sheet — all rows on one sheet, grouped by operator then date.
                     A dark-blue separator row marks each new operator block.
      multi_sheet  — one sheet per distinct operator.

    The columns rendered in each row are controlled by template.columns,
    an ordered list of column keys. Valid keys and their sources:
      fecha       — WorkOrderEntry.work_date
      operario    — WorkOrder.uploaded_by (full name or username)
      maquina     — WorkOrderEntryLine.machine_asset.code or machine_raw
      descripcion — WorkOrderEntryLine.fault_description
      notas       — WorkOrderEntryLine.repair_notes
      hc          — WorkOrderEntryLine.hc
      hf          — WorkOrderEntryLine.hf
      delta_horas — WorkOrderEntryLine.delta_hours
      estado      — WorkOrder.reviewed (Revisado / Pendiente)
      familia     — WorkOrderEntryLine.fault_category
      origen      — WorkOrder.source
      horas_extra — horas sobre jornada estándar de 8h (max(0, sum(delta_hours) - 8) por entry)
      dietas      — 1 si el entry (día) tiene has_diet=True, 0 si no (por entry)

    Cada hoja incluye, antes de la cabecera de columnas, dos celdas de
    entrada para que Administración introduzca manualmente el precio de
    la hora ordinaria y de la hora extra (mismo patrón que
    generate_work_order_excel) -- 2026-07-07, solicitado por Miguel Ángel.
    No se generan fórmulas de salario por fila (fuera de alcance de esta
    petición); las celdas quedan como referencia de precio junto a las
    horas/dietas totales de cada hoja.

    Returns an openpyxl.Workbook instance ready for streaming.
    ---

    Construye un Workbook openpyxl desde un queryset de WorkOrder según
    la configuración almacenada en la instancia ExportTemplate dada.

    Valores de template.sheet_format soportados:
      single_sheet — todas las filas en una hoja, agrupadas por operario y fecha.
                     Una fila separadora azul oscuro marca cada nuevo operario.
      multi_sheet  — una hoja por operario distinto.

    Las columnas renderizadas se controlan por template.columns, una lista
    ordenada de claves. Claves válidas y sus orígenes: ver docstring en.

    Devuelve una instancia openpyxl.Workbook lista para streaming.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from decimal import Decimal
    from .models import WorkOrderEntry, WorkOrderEntryLine

    # ------------------------------------------------------------------
    # Column metadata: label and value extractor per key.
    # Metadatos de columna: etiqueta y extractor de valor por clave.
    # ------------------------------------------------------------------
    COLUMN_DEFS = {
        "fecha":       ("Fecha",            lambda wo, entry, line: entry.work_date),
        "operario":    ("Operario",         lambda wo, entry, line: (
                                                wo.uploaded_by.user.get_full_name()
                                                or wo.uploaded_by.user.username
                                            ) if wo.uploaded_by else ""),
        "maquina":     ("Máquina / CdG",    lambda wo, entry, line: (
                                                line.machine_asset.code
                                                if line.machine_asset
                                                else line.machine_raw or ""
                                            )),
        "descripcion": ("Descripción avería", lambda wo, entry, line: line.fault_description or ""),
        "notas":       ("Notas reparación", lambda wo, entry, line: line.repair_notes or ""),
        "hc":          ("H. inicio",        lambda wo, entry, line: (
                                                line.hc.strftime("%H:%M") if line.hc else ""
                                            )),
        "hf":          ("H. fin",           lambda wo, entry, line: (
                                                line.hf.strftime("%H:%M") if line.hf else ""
                                            )),
        "delta_horas": ("Δ Horas",          lambda wo, entry, line: (
                                                float(line.delta_hours)
                                                if line.delta_hours is not None else ""
                                            )),
        "estado":      ("Estado",           lambda wo, entry, line: (
                                                "Revisado" if wo.reviewed else "Pendiente"
                                            )),
        "familia":     ("Familia avería",   lambda wo, entry, line: line.fault_category or ""),
        "origen":      ("Origen",           lambda wo, entry, line: wo.source or ""),
        "horas_extra": ("H. Extra",          lambda wo, entry, line: (
                                                float(max(
                                                    Decimal("0"),
                                                    sum(
                                                        (l.delta_hours for l in entry.lines.all()
                                                         if l.delta_hours is not None),
                                                        Decimal("0"),
                                                    ) - Decimal("8")
                                                ))
                                            )),
        # Dietas — 1 si el entry (dia) tiene has_diet=True, 0 si no. Metrica
        # de entry, no de linea -- igual que delta_horas/horas_extra, se
        # acumula una sola vez por entry en la fila TOTAL (2026-07-07,
        # solicitado por Miguel Angel: "sumar el numero de dietas por
        # operario y periodo").
        "dietas":      ("Dietas",            lambda wo, entry, line: (
                                                1 if entry.has_diet else 0
                                            )),
    }

    # Resolve active column definitions in template order.
    # Resolver definiciones de columna activas en el orden de la plantilla.
    active_cols = [
        (key, COLUMN_DEFS[key])
        for key in template.columns
        if key in COLUMN_DEFS
    ]

    # ------------------------------------------------------------------
    # Numeric column keys — eligible for a TOTAL row at the end of each
    # data block. Only keys present in active_cols are summed.
    # Claves de columna numéricas — elegibles para la fila TOTAL al final
    # de cada bloque de datos. Solo se suman las que estén en active_cols.
    # ------------------------------------------------------------------
    NUMERIC_KEYS = {"delta_horas", "horas_extra", "dietas"}

    # ------------------------------------------------------------------
    # Styles / Estilos
    # ------------------------------------------------------------------
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sep_fill    = PatternFill("solid", fgColor="2F5496")
    sep_font    = Font(bold=True, color="FFFFFF", size=10)
    total_fill  = PatternFill("solid", fgColor="D6E4F0")
    total_font  = Font(bold=True, color="1F3864", size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------------
    # Helper: write a totals row on a given worksheet.
    # Writes "TOTAL" in column 1 and the accumulated sum for each
    # numeric column. Non-numeric columns are left blank.
    # Auxiliar: escribir la fila de totales en una hoja dada.
    # Escribe "TOTAL" en la columna 1 y la suma acumulada por columna
    # numérica. Las columnas no numéricas quedan en blanco.
    # ------------------------------------------------------------------
    def _write_totals_row(ws, row, col_defs, accumulators):
        """
        Writes a styled TOTAL row using the accumulated column sums.
        ---
        Escribe una fila TOTAL estilizada usando las sumas acumuladas.
        """
        for col_idx, (key, _) in enumerate(col_defs, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = center_align
            if col_idx == 1:
                cell.value = "TOTAL"
            elif key in NUMERIC_KEYS and key in accumulators:
                val = accumulators[key]
                cell.value = round(val, 2) if val else 0.0

    # ------------------------------------------------------------------
    # Build flat line list enriched with wo and entry references.
    # Construir lista plana de líneas enriquecidas con referencias wo y entry.
    # ------------------------------------------------------------------
    rows = (
        WorkOrderEntryLine.objects
        .filter(entry__work_order__in=work_orders_qs)
        .select_related(
            "entry__work_order__uploaded_by__user",
            "entry",
            "machine_asset",
        )
        .order_by(
            "entry__work_order__uploaded_by__user__last_name",
            "entry__work_order__uploaded_by__user__first_name",
            "entry__work_date",
            "entry__work_order__pk",
            "line_number",
        )
        # Guard against duplicate rows caused by prefetch_related annotations
        # on the incoming work_orders_qs when used as a subquery.
        # Protección contra filas duplicadas causadas por anotaciones
        # prefetch_related en work_orders_qs al usarse como subquery.
        .distinct()
    )

    # ------------------------------------------------------------------
    # Row layout — dos filas de celdas de precio antes de la cabecera de
    # columnas, mismo patron que generate_work_order_excel (2026-07-07,
    # solicitado por Miguel Angel).
    # Row layout — two price-input rows before the column header row,
    # same pattern as generate_work_order_excel.
    # ------------------------------------------------------------------
    _PRICE_ROW_EXTRA     = 1
    _PRICE_ROW_ORDINARY  = 2
    _HEADER_ROW          = 3
    _FIRST_DATA_ROW      = 4

    # ------------------------------------------------------------------
    # Helper: write the two price-input cells at the top of a sheet.
    # Column C holds the editable value; columns A-B carry the label.
    # Auxiliar: escribir las dos celdas de precio en la parte superior de
    # una hoja. La columna C contiene el valor editable; A-B llevan la
    # etiqueta.
    # ------------------------------------------------------------------
    def _write_price_cells(ws):
        """
        Writes 'PRECIO HORA EXTRA' (row 1) and 'COSTE HORA ORDINARIA'
        (row 2) label + input cells, mirroring generate_work_order_excel's
        C2/C3 mechanism.
        ---
        Escribe las celdas de etiqueta + entrada 'PRECIO HORA EXTRA'
        (fila 1) y 'COSTE HORA ORDINARIA' (fila 2), replicando el
        mecanismo C2/C3 de generate_work_order_excel.
        """
        ws.cell(row=_PRICE_ROW_EXTRA, column=1, value="PRECIO HORA EXTRA (euros/h):").fill = (
            _make_fill(_CLR_CONFIG_LABEL)
        )
        ws.merge_cells(start_row=_PRICE_ROW_EXTRA, start_column=1, end_row=_PRICE_ROW_EXTRA, end_column=2)
        price_extra_cell = ws.cell(row=_PRICE_ROW_EXTRA, column=3)
        price_extra_cell.fill = _make_fill(_CLR_CONFIG_INPUT)
        price_extra_cell.number_format = '#,##0.00 "€"'
        price_extra_cell.alignment = center_align

        ws.cell(row=_PRICE_ROW_ORDINARY, column=1, value="COSTE HORA ORDINARIA (euros/h):").fill = (
            _make_fill(_CLR_CONFIG_LABEL)
        )
        ws.merge_cells(start_row=_PRICE_ROW_ORDINARY, start_column=1, end_row=_PRICE_ROW_ORDINARY, end_column=2)
        price_ord_cell = ws.cell(row=_PRICE_ROW_ORDINARY, column=3)
        price_ord_cell.fill = _make_fill(_CLR_CONFIG_INPUT)
        price_ord_cell.number_format = '#,##0.00 "€"'
        price_ord_cell.alignment = center_align

    # ------------------------------------------------------------------
    # Helper: write a header row on a given worksheet.
    # Auxiliar: escribir la fila de cabecera en una hoja dada.
    # ------------------------------------------------------------------
    def _write_header(ws, col_defs):
        """
        Writes the header row (row _HEADER_ROW, below the two price cells)
        with the active column labels.
        ---
        Escribe la fila de cabecera (fila _HEADER_ROW, debajo de las dos
        celdas de precio) con las etiquetas de columna activas.
        """
        for col_idx, (key, (label, _extractor)) in enumerate(col_defs, start=1):
            cell           = ws.cell(row=_HEADER_ROW, column=col_idx, value=label)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align

    # ------------------------------------------------------------------
    # Helper: autofit column widths based on header and content.
    # Auxiliar: autoajustar ancho de columnas según cabecera y contenido.
    # ------------------------------------------------------------------
    def _autofit_columns(ws):
        """
        Sets each column width to the maximum of its header length and the
        longest cell value in that column, with a minimum of 10 and a cap of 60.
        ---
        Ajusta el ancho de cada columna al máximo entre la longitud de la
        cabecera y el valor más largo de la columna, con mínimo 10 y máximo 60.
        """
        from openpyxl.utils import get_column_letter
        for col_cells in ws.columns:
            max_len = 10
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # ------------------------------------------------------------------
    # SHEET FORMAT: single_sheet
    # FORMATO DE HOJA: single_sheet
    # ------------------------------------------------------------------
    if template.sheet_format == "single_sheet":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Partes digitales"
        _write_price_cells(ws)
        _write_header(ws, active_cols)

        current_row      = _FIRST_DATA_ROW
        current_operator = None
        # Accumulators for numeric columns (global, single sheet).
        # Acumuladores para columnas numéricas (global, hoja única).
        accumulators = {key: 0.0 for key, _ in active_cols if key in NUMERIC_KEYS}
        # Guard: NUMERIC_KEYS values (delta_horas, horas_extra) are entry-level
        # metrics stored redundantly on every line of the same entry. We must
        # only accumulate them once per entry, not once per line.
        # Guarda: los valores NUMERIC_KEYS (delta_horas, horas_extra) son
        # métricas de entry almacenadas de forma redundante en cada línea de
        # la misma entry. Solo deben acumularse una vez por entry, no por línea.
        seen_entry_pks: set = set()

        for line in rows:
            wo    = line.entry.work_order
            entry = line.entry
            op_name = (
                wo.uploaded_by.user.get_full_name() or wo.uploaded_by.user.username
                if wo.uploaded_by else ""
            )

            # Insert separator row when operator changes.
            # Insertar fila separadora cuando cambia el operario.
            if op_name != current_operator:
                current_operator = op_name
                sep_cell       = ws.cell(row=current_row, column=1, value=op_name)
                sep_cell.fill  = sep_fill
                sep_cell.font  = sep_font
                # Merge separator across all active columns.
                # Fusionar separador a lo largo de todas las columnas activas.
                if len(active_cols) > 1:
                    ws.merge_cells(
                        start_row=current_row, start_column=1,
                        end_row=current_row,   end_column=len(active_cols),
                    )
                current_row += 1

            is_first_line_of_entry = entry.pk not in seen_entry_pks
            seen_entry_pks.add(entry.pk)

            for col_idx, (key, (label, extractor)) in enumerate(active_cols, start=1):
                val = extractor(wo, entry, line)
                # NUMERIC_KEYS are entry-level metrics (delta_horas, horas_extra).
                # On subsequent lines of the same entry write blank so the
                # value appears only once per day (Opción B).
                # NUMERIC_KEYS son métricas de entry (delta_horas, horas_extra).
                # En líneas posteriores de la misma entry se escribe en blanco
                # para que el valor aparezca solo una vez por día (Opción B).
                if key in NUMERIC_KEYS and not is_first_line_of_entry:
                    val = ""
                ws.cell(row=current_row, column=col_idx, value=val)
                # Accumulate only on the first line of each entry.
                # Acumular solo en la primera línea de cada entry.
                if (
                    key in accumulators
                    and isinstance(val, (int, float))
                    and is_first_line_of_entry
                ):
                    accumulators[key] += val
            current_row += 1

        # Write global TOTAL row at the bottom of the sheet.
        # Escribir la fila TOTAL global al final de la hoja.
        if accumulators:
            _write_totals_row(ws, current_row, active_cols, accumulators)

        _autofit_columns(ws)
        return wb

    # ------------------------------------------------------------------
    # SHEET FORMAT: multi_sheet (one sheet per operator)
    # FORMATO DE HOJA: multi_sheet (una hoja por operario)
    # ------------------------------------------------------------------
    wb = openpyxl.Workbook()
    # Remove default empty sheet.
    # Eliminar la hoja vacía por defecto.
    wb.remove(wb.active)

    current_operator = None
    ws               = None
    current_row      = _FIRST_DATA_ROW
    # Per-sheet accumulators reset when a new sheet opens.
    # Acumuladores por hoja, reiniciados al abrir cada hoja nueva.
    sheet_accumulators = {}
    # Guard: NUMERIC_KEYS values are entry-level metrics — only accumulate
    # once per entry per sheet, not once per line.
    # Guarda: los valores NUMERIC_KEYS son métricas de entry — solo acumular
    # una vez por entry por hoja, no por línea.
    seen_entry_pks_sheet: set = set()

    for line in rows:
        wo    = line.entry.work_order
        entry = line.entry
        op_name = (
            wo.uploaded_by.user.get_full_name() or wo.uploaded_by.user.username
            if wo.uploaded_by else "Sin operario"
        )

        # Create a new sheet when operator changes.
        # Crear una nueva hoja cuando cambia el operario.
        if op_name != current_operator:
            # Write TOTAL row on the previous sheet before switching.
            # Escribir fila TOTAL en la hoja anterior antes de cambiar.
            if ws is not None and sheet_accumulators:
                _write_totals_row(ws, current_row, active_cols, sheet_accumulators)
            current_operator   = op_name
            ws                 = wb.create_sheet(title=op_name[:31])
            current_row        = _FIRST_DATA_ROW
            sheet_accumulators = {
                key: 0.0 for key, _ in active_cols if key in NUMERIC_KEYS
            }
            seen_entry_pks_sheet = set()
            _write_price_cells(ws)
            _write_header(ws, active_cols)

        is_first_line_of_entry = entry.pk not in seen_entry_pks_sheet
        seen_entry_pks_sheet.add(entry.pk)

        for col_idx, (key, (label, extractor)) in enumerate(active_cols, start=1):
            val = extractor(wo, entry, line)
            # NUMERIC_KEYS are entry-level metrics (delta_horas, horas_extra).
            # On subsequent lines of the same entry write blank (Opción B).
            # NUMERIC_KEYS son métricas de entry. En líneas posteriores
            # de la misma entry se escribe en blanco (Opción B).
            if key in NUMERIC_KEYS and not is_first_line_of_entry:
                val = ""
            ws.cell(row=current_row, column=col_idx, value=val)
            # Accumulate only on the first line of each entry.
            # Acumular solo en la primera línea de cada entry.
            if (
                key in sheet_accumulators
                and isinstance(val, (int, float))
                and is_first_line_of_entry
            ):
                sheet_accumulators[key] += val
        current_row += 1

    # Write TOTAL row on the last sheet.
    # Escribir fila TOTAL en la última hoja.
    if ws is not None and sheet_accumulators:
        _write_totals_row(ws, current_row, active_cols, sheet_accumulators)

    for sheet in wb.worksheets:
        _autofit_columns(sheet)
    return wb



